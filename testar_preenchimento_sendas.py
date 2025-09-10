#!/usr/bin/env python3
"""
Script de teste para preencher planilha de agendamento do Sendas
Usa o CNPJ 06.057.223/0233-84 conforme solicitado
"""

import os
import sys
from datetime import date, datetime
import logging

# Adicionar o caminho do projeto
sys.path.insert(0, os.path.abspath(os.path.dirname(__file__)))

from app import create_app, db
from app.portal.sendas.preencher_planilha import PreencherPlanilhaSendas

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def criar_planilha_exemplo():
    """
    Cria uma planilha de exemplo com a estrutura do Sendas
    para teste quando não houver planilha real baixada
    """
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    logger.info("📝 Criando planilha de exemplo...")
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Agendamento"
    
    # Cabeçalhos (linha 3)
    headers = [
        'Demanda',  # A
        'Razão Social - Fornecedor',  # B
        'Nome Fantasia - Fornecedor',  # C
        'Unidade de destino',  # D
        'UF Destino',  # E
        'Fluxo de operação',  # F
        'Código do pedido Cliente',  # G
        'Código Produto Cliente',  # H
        'Código Produto SKU Fornecedor',  # I
        'EAN',  # J
        'Setor',  # K
        'Número do pedido Trizy',  # L
        'Descrição do Item',  # M
        'Quantidade total',  # N
        'Saldo disponível',  # O
        'Unidade de medida',  # P
        'Quantidade entrega',  # Q
        'Data sugerida de entrega',  # R
        'ID de agendamento (opcional)',  # S
        'Reserva de Slot (opcional)',  # T
        'Característica da carga',  # U
        'Característica do veículo',  # V
        'Transportadora CNPJ (opcional)',  # W
        'Observação/ Fornecedor (opcional)'  # X
    ]
    
    # Adicionar cabeçalhos na linha 3
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    
    # Dados de exemplo (linha 4)
    exemplo = [
        '',  # A - Demanda (será preenchido)
        'NACOM GOYA IND COM ALIMENTOS LTDA',  # B
        'NACOM GOYA',  # C
        'SENDAS 923 CD MANAUS',  # D
        'AM',  # E
        'Recebimento',  # F
        '19447861-923',  # G
        '93734',  # H
        '-',  # I
        '-',  # J
        '-',  # K
        '2998070',  # L
        'AZEITONA PTA CAMPO BELO FAT 1,01KG',  # M
        '20',  # N
        '20',  # O
        'CX',  # P
        '',  # Q - Quantidade entrega (será preenchido)
        '',  # R - Data sugerida (será preenchido)
        '-',  # S
        '-',  # T
        '',  # U - Característica carga (será preenchido)
        '',  # V - Característica veículo (será preenchido)
        '-',  # W
        ''  # X - Observação (será preenchido)
    ]
    
    # Adicionar exemplo na linha 4
    for col, valor in enumerate(exemplo, 1):
        ws.cell(row=4, column=col, value=valor)
    
    # Salvar planilha
    caminho = os.path.join(
        os.path.dirname(__file__),
        'app/portal/sendas/downloads',
        f'planilha_exemplo_sendas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )
    
    # Criar diretório se não existir
    os.makedirs(os.path.dirname(caminho), exist_ok=True)
    
    wb.save(caminho)
    logger.info(f"  ✅ Planilha de exemplo criada: {caminho}")
    
    return caminho


def testar_preenchimento():
    """
    Testa o preenchimento da planilha com o CNPJ especificado
    """
    logger.info("=" * 80)
    logger.info("🧪 TESTE DE PREENCHIMENTO DE PLANILHA SENDAS")
    logger.info("=" * 80)
    
    # CNPJ para teste
    cnpj_teste = "06.057.223/0233-84"
    logger.info(f"📍 CNPJ de teste: {cnpj_teste}")
    
    # Data de agendamento (hoje + 3 dias)
    data_agendamento = date.today()
    for _ in range(3):
        data_agendamento = date(
            data_agendamento.year,
            data_agendamento.month,
            data_agendamento.day + 1
        )
    logger.info(f"📅 Data de agendamento: {data_agendamento.strftime('%d/%m/%Y')}")
    
    # Verificar se existe planilha baixada
    downloads_dir = os.path.join(
        os.path.dirname(__file__),
        'app/portal/sendas/downloads'
    )
    
    arquivo_origem = None
    
    # Procurar por planilha existente
    if os.path.exists(downloads_dir):
        for arquivo in os.listdir(downloads_dir):
            if arquivo.endswith('.xlsx') and 'sendas' in arquivo.lower():
                arquivo_origem = os.path.join(downloads_dir, arquivo)
                logger.info(f"📂 Usando planilha existente: {arquivo}")
                break
    
    # Se não encontrou, criar planilha de exemplo
    if not arquivo_origem:
        logger.warning("⚠️ Nenhuma planilha baixada encontrada")
        arquivo_origem = criar_planilha_exemplo()
    
    # Criar instância do preenchedor
    logger.info("\n🔧 Iniciando preenchimento...")
    app = create_app()
    
    with app.app_context():
        preenchedor = PreencherPlanilhaSendas()
        
        # Arquivo de destino
        arquivo_destino = os.path.join(
            downloads_dir,
            "agendamento_preenchido.xlsx"
        )
        
        try:
            # Preencher planilha
            arquivo_preenchido = preenchedor.preencher_planilha(
                arquivo_origem=arquivo_origem,
                cnpj=cnpj_teste,
                data_agendamento=data_agendamento,
                arquivo_destino=arquivo_destino
            )
            
            if arquivo_preenchido:
                logger.info("\n" + "=" * 80)
                logger.info("✅ TESTE CONCLUÍDO COM SUCESSO!")
                logger.info("=" * 80)
                logger.info(f"📄 Planilha preenchida salva em:")
                logger.info(f"   {arquivo_preenchido}")
                logger.info("\n💡 Verifique o arquivo para confirmar a integridade dos dados")
                
                # Verificar se o arquivo existe e tem conteúdo
                if os.path.exists(arquivo_preenchido):
                    tamanho = os.path.getsize(arquivo_preenchido)
                    logger.info(f"   Tamanho do arquivo: {tamanho:,} bytes")
                    
                    # Tentar ler o arquivo para validação básica
                    try:
                        from openpyxl import load_workbook
                        wb = load_workbook(arquivo_preenchido)
                        ws = wb.active
                        
                        # Contar linhas preenchidas
                        linhas_preenchidas = 0
                        for row in range(4, ws.max_row + 1):
                            if ws.cell(row=row, column=1).value:  # Coluna A - Demanda
                                linhas_preenchidas += 1
                        
                        logger.info(f"   Linhas de dados preenchidas: {linhas_preenchidas}")
                        
                        # Verificar campos principais
                        if linhas_preenchidas > 0:
                            logger.info("\n📋 Verificação de campos (primeira linha de dados):")
                            primeira_linha = 4
                            
                            campos_verificar = [
                                (1, "Demanda"),
                                (17, "Quantidade entrega"),
                                (18, "Data sugerida"),
                                (21, "Característica carga"),
                                (22, "Característica veículo"),
                                (24, "Observação")
                            ]
                            
                            for col, nome in campos_verificar:
                                valor = ws.cell(row=primeira_linha, column=col).value
                                if valor:
                                    logger.info(f"   ✓ {nome}: {valor}")
                                else:
                                    logger.warning(f"   ⚠️ {nome}: (vazio)")
                        
                        wb.close()
                        
                    except Exception as e:
                        logger.error(f"⚠️ Erro ao validar arquivo: {e}")
                
            else:
                logger.error("\n" + "=" * 80)
                logger.error("❌ TESTE FALHOU!")
                logger.error("=" * 80)
                logger.error("Nenhum dado encontrado para o CNPJ especificado")
                logger.info("\n💡 Dicas:")
                logger.info("1. Verifique se existe dados para o CNPJ 06.057.223/0233-84")
                logger.info("2. Verifique se há pedidos na CarteiraPrincipal ou Separacao")
                logger.info("3. Verifique se o CNPJ está correto no banco de dados")
                
        except Exception as e:
            logger.error("\n" + "=" * 80)
            logger.error("❌ ERRO NO TESTE!")
            logger.error("=" * 80)
            logger.error(f"Erro: {e}")
            
            import traceback
            logger.error("\nStack trace:")
            traceback.print_exc()
            
            logger.info("\n💡 Possíveis causas:")
            logger.info("1. Banco de dados não acessível")
            logger.info("2. Tabelas DE-PARA não populadas")
            logger.info("3. Estrutura da planilha diferente do esperado")


if __name__ == "__main__":
    testar_preenchimento()