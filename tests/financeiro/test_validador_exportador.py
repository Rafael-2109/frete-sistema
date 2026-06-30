"""Testes do exportador Excel do Validador de Titulos x Bancos."""

from openpyxl import load_workbook

from app.financeiro.services.validador_titulos.comparador import ResultadoComparacao
from app.financeiro.services.validador_titulos.exportador import gerar_excel
from app.financeiro.services.validador_titulos.service import ResultadoValidacao


def _resultado_exemplo():
    comp = ResultadoComparacao(
        duplicados=[{"nf_parc": "100-1", "bancos": ["AGIS", "SRM"], "qtd_bancos": 2,
                     "tem_recompra": True}],
        faturado_sem_boleto=[{"nf_parc": "200-1", "cliente": "ACME", "cnpj": "1",
                              "uf": "SP", "vencimento": None, "valor": 100.0,
                              "tipo_titulo": "AGIS", "empresa": 3}],
        boleto_sem_nota=[{"nf_parc": "300-1", "bancos": ["VORTX"]}],
        nao_identificados=[{"banco": "VORTX", "original": "147569.0", "valor": 50.0,
                            "vencimento": None}],
    )
    return ResultadoValidacao(
        resultado=comp,
        resumo={"qtd_boletos": 3, "qtd_duplicados": 1, "qtd_faturado_sem_boleto": 1,
                "qtd_boleto_sem_nota": 1, "qtd_nao_identificados": 1},
    )


def test_gera_excel_com_todas_as_abas():
    buffer = gerar_excel(_resultado_exemplo())
    wb = load_workbook(buffer)
    nomes = set(wb.sheetnames)
    assert {"RESUMO", "DUPLICADOS", "FATURADO SEM BOLETO",
            "BOLETO SEM NOTA", "NAO IDENTIFICADOS"} <= nomes


def test_aba_duplicados_tem_linha_de_dados():
    buffer = gerar_excel(_resultado_exemplo())
    wb = load_workbook(buffer)
    ws = wb["DUPLICADOS"]
    linhas = list(ws.iter_rows(values_only=True))
    assert linhas[0][0] == "NF-PARC"          # cabecalho
    assert linhas[1][0] == "100-1"            # dado
    assert "AGIS" in str(linhas[1][1])        # bancos
    assert linhas[1][3] in ("OK", "RECOMPRA OK")  # flag de recompra


def test_faturado_sem_boleto_tem_cliente():
    buffer = gerar_excel(_resultado_exemplo())
    wb = load_workbook(buffer)
    ws = wb["FATURADO SEM BOLETO"]
    linhas = list(ws.iter_rows(values_only=True))
    assert any("ACME" in str(c) for c in linhas[1])


def test_caractere_ilegal_nao_quebra_excel():
    """Valores com caractere de controle (NUL) nao podem derrubar a geracao."""
    comp = ResultadoComparacao(
        boleto_sem_nota=[{"nf_parc": "\x00\x00147768-2", "bancos": ["VORTX"]}],
        nao_identificados=[{"banco": "VORTX", "original": "lixo\x00aqui",
                            "valor": 10.0, "vencimento": None}],
    )
    rv = ResultadoValidacao(resultado=comp, resumo={})
    buffer = gerar_excel(rv)  # nao deve levantar
    wb = load_workbook(buffer)
    ws = wb["BOLETO SEM NOTA"]
    linhas = list(ws.iter_rows(values_only=True))
    # valor saiu limpo, sem o NUL
    assert linhas[1][0] == "147768-2"
