"""
Testes do orquestrador do Validador de Titulos x Bancos.

Junta parsers de banco + CP-NACOM + faturamento + comparador. O faturamento
e injetado nos testes (evita tocar o banco).
"""

import pytest
from openpyxl import Workbook

from app.financeiro.services.validador_titulos.service import processar_validacao


def _xlsx_banco_agis(caminho, *recebiveis):
    wb = Workbook()
    ws = wb.active
    ws.append(["Recebivel", "Tipo", "Modalidade", "Valor Face", "Valor Aberto"])
    for rec in recebiveis:
        ws.append([rec, "DUPLICATA", "DESCONTO", "100", "100"])
    wb.save(caminho)


def _xlsx_banco_srm(caminho, *docs):
    wb = Workbook()
    ws = wb.active
    ws.append(["Nro, Documento", "Codigo", "Data Vencimento", "Valor"])
    for d in docs:
        ws.append([d, "x", "16/06/2026", "100"])
    wb.save(caminho)


def _xlsx_cp(caminho, *titulos_parc):
    wb = Workbook()
    wb.active.title = "RESUMO"
    ws = wb.create_sheet("CP - NACOM")
    ws.append(["Vencimento", "EMISSAO", "F. PAGTO", "Fornecedor", "Titulo", "Parc", "VALOR"])
    for titulo, parc in titulos_parc:
        ws.append(["45231", "", "BOLETO", "ACME", titulo, parc, "100"])
    wb.save(caminho)


def test_pipeline_completo_duplicado_com_recompra(tmp_path):
    # Mesmo titulo 100-1 em AGIS e SRM => duplicado. Recompra lancada no CP => OK.
    agis = tmp_path / "agis.xlsx"; _xlsx_banco_agis(agis, "100/1")
    srm = tmp_path / "srm.xlsx"; _xlsx_banco_srm(srm, "100-001")
    cp = tmp_path / "cp.xlsx"; _xlsx_cp(cp, ("100", "1"))

    faturamento = [{"nf_parc": "100-1"}, {"nf_parc": "200-1", "cliente": "X"}]

    res = processar_validacao(
        caminhos_bancos={"AGIS": str(agis), "SRM": str(srm)},
        caminho_cp=str(cp),
        faturamento=faturamento,
    )

    assert res.resumo["qtd_duplicados"] == 1
    assert res.resultado.duplicados[0]["nf_parc"] == "100-1"
    assert res.resultado.duplicados[0]["tem_recompra"] is True
    # 200-1 faturado mas sem boleto
    assert [f["nf_parc"] for f in res.resultado.faturado_sem_boleto] == ["200-1"]
    assert res.resumo["qtd_recompras_cp"] == 1


def test_erro_em_um_banco_nao_derruba_os_outros(tmp_path):
    # AGIS valido; "GRAFENO" aponta para arquivo sem a coluna identificadora.
    agis = tmp_path / "agis.xlsx"; _xlsx_banco_agis(agis, "100/1")
    ruim = tmp_path / "ruim.xlsx"
    wb = Workbook(); wb.active.append(["coluna_errada"]); wb.save(ruim)
    cp = tmp_path / "cp.xlsx"; _xlsx_cp(cp)

    res = processar_validacao(
        caminhos_bancos={"AGIS": str(agis), "GRAFENO": str(ruim)},
        caminho_cp=str(cp),
        faturamento=[],
    )
    assert res.resumo["qtd_boletos"] == 1  # so o AGIS entrou
    assert "GRAFENO" in res.erros
