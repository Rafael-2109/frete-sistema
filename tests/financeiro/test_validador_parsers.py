"""
Testes dos parsers de base de banco do Validador de Titulos x Bancos.

Foca na logica pura `extrair_boletos(linhas, banco)`, que recebe a matriz ja
lida do arquivo e devolve a lista de boletos normalizados. O I/O (ler xlsx/
xlsb/csv) e testado a parte / validado contra os arquivos reais.
"""

import pytest

from app.financeiro.services.validador_titulos.parsers_bancos import (
    extrair_boletos,
    ler_arquivo,
    parsear_arquivo,
    BANCOS_SUPORTADOS,
)


# Cabecalhos reais (com acento) de GRAFENO/VORTX, precedidos das 4 linhas de resumo.
GRAFENO_HEADER = [
    "Usuario_Criador_da_Cobranca", "Conta_Origem", "Pagador", "Numero_Documento",
    "Seu_Número", "Regua_Cobranca", "Data_Criacao", "Data_Vencimento",
    "Valor_Cobranca", "Grupo_Cobranca",
]


def _grafeno_linhas(*data_rows):
    return [
        ["Periodo:", "-"],
        ["Cobrancas Abertas (pagas)", "0,00"],
        ["Cobrancas Abertas (em aberto)", "805.898,06"],
        ["Total:", "805.898,06"],
        GRAFENO_HEADER,
        *data_rows,
    ]


def _grafeno_row(seu_numero, pagador="CLIENTE X", venc="15/06/2026", valor="R$ 16.252,62"):
    return [
        "Allanda", "NACOM", pagador, "00.722/0001",
        seu_numero, "Apenas Registro", "25/03/2026", venc, valor, "Regra",
    ]


class TestExtrairBoletosGrafeno:

    def test_extrai_boleto_basico(self):
        linhas = _grafeno_linhas(_grafeno_row("146299/003"))
        boletos = extrair_boletos(linhas, "GRAFENO")
        assert len(boletos) == 1
        b = boletos[0]
        assert b["nf_parc"] == "146299-3"
        assert b["banco"] == "GRAFENO"
        assert b["original"] == "146299/003"

    def test_acha_cabecalho_apos_linhas_de_resumo(self):
        """O cabecalho real esta na 5a linha; as 4 de resumo nao podem virar boleto."""
        linhas = _grafeno_linhas(_grafeno_row("146299/003"), _grafeno_row("146367/003"))
        boletos = extrair_boletos(linhas, "GRAFENO")
        assert [b["nf_parc"] for b in boletos] == ["146299-3", "146367-3"]

    def test_extrai_valor_formato_brasileiro(self):
        linhas = _grafeno_linhas(_grafeno_row("146299/003", valor="R$ 16.252,62"))
        b = extrair_boletos(linhas, "GRAFENO")[0]
        assert b["valor"] == pytest.approx(16252.62)

    def test_extrai_pagador_e_vencimento(self):
        linhas = _grafeno_linhas(_grafeno_row("146299/003", pagador="ACME LTDA", venc="20/07/2026"))
        b = extrair_boletos(linhas, "GRAFENO")[0]
        assert b["pagador"] == "ACME LTDA"
        assert b["vencimento"] == "20/07/2026"

    def test_linha_vazia_e_ignorada(self):
        linhas = _grafeno_linhas(_grafeno_row("146299/003"), _grafeno_row(""))
        boletos = extrair_boletos(linhas, "GRAFENO")
        assert len(boletos) == 1

    def test_identificador_sujo_vira_nf_parc_none(self):
        """Identificador presente mas nao normalizavel (ex 147569.0) entra com nf_parc None."""
        linhas = _grafeno_linhas(_grafeno_row("147569.0"))
        b = extrair_boletos(linhas, "GRAFENO")[0]
        assert b["nf_parc"] is None
        assert b["original"] == "147569.0"


class TestExtrairBoletosSrm:

    def test_srm_documento_com_parcela(self):
        linhas = [
            ["Nro, Documento", "Codigo da Operacao", "Produto", "Data Vencimento",
             "Data Liquidacao", "Prazo", "Nome do Sacado", "Valor"],
            ["148466-001", "8898493", "GARKGCP", "16/06/2026", "", "0", "W J GIL", "4041.63"],
        ]
        b = extrair_boletos(linhas, "SRM")[0]
        assert b["nf_parc"] == "148466-1"
        assert b["valor"] == pytest.approx(4041.63)


class TestExtrairBoletosAgis:

    def test_agis_recebivel_barra(self):
        linhas = [
            ["Recebivel", "Tipo Recebivel", "Modalidade", "Valor Face", "Valor Aberto",
             "Status", "Observacao", "Desagio"],
            ["146826/5", "DUPLICATA", "DESCONTO", "1745.17", "1745.17", "ABERTO", "", "355.99"],
        ]
        b = extrair_boletos(linhas, "AGIS")[0]
        assert b["nf_parc"] == "146826-5"
        assert b["valor"] == pytest.approx(1745.17)


class TestBancoInvalido:

    def test_banco_desconhecido_levanta_erro(self):
        with pytest.raises(ValueError):
            extrair_boletos([["x"]], "BANCO_INEXISTENTE")

    def test_lista_de_bancos_suportados(self):
        assert set(BANCOS_SUPORTADOS) == {"SRM", "GRAFENO", "AGIS", "VORTX"}

    def test_cabecalho_nao_encontrado_levanta_erro(self):
        """Se a coluna identificadora nao existe, erro claro (arquivo errado)."""
        linhas = [["coluna_a", "coluna_b"], ["1", "2"]]
        with pytest.raises(ValueError):
            extrair_boletos(linhas, "GRAFENO")


class TestLerArquivo:
    """I/O: ler xlsx e csv para matriz de linhas."""

    def test_le_xlsx(self, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Recebivel", "Valor Aberto"])
        ws.append(["146826/5", "1745.17"])
        caminho = tmp_path / "agis.xlsx"
        wb.save(caminho)

        linhas = ler_arquivo(str(caminho))
        assert linhas[0][0] == "Recebivel"
        assert linhas[1][0] == "146826/5"

    def test_le_csv(self, tmp_path):
        caminho = tmp_path / "agis.csv"
        caminho.write_text("Recebivel;Valor Aberto\n146826/5;1745.17\n", encoding="utf-8")
        linhas = ler_arquivo(str(caminho))
        assert linhas[0][0] == "Recebivel"
        assert linhas[1][0] == "146826/5"

    def test_extensao_nao_suportada_levanta_erro(self, tmp_path):
        caminho = tmp_path / "x.pdf"
        caminho.write_text("nada", encoding="utf-8")
        with pytest.raises(ValueError):
            ler_arquivo(str(caminho))

    def test_le_aba_nomeada_xlsx(self, tmp_path):
        """Le uma aba especifica por nome (tolerante a espacos/hifens)."""
        from openpyxl import Workbook
        wb = Workbook()
        wb.active.title = "RESUMO"
        wb.active.append(["lixo"])
        ws2 = wb.create_sheet("CP - NACOM")
        ws2.append(["Titulo", "Parc"])
        ws2.append(["2023", "10"])
        caminho = tmp_path / "cp.xlsx"
        wb.save(caminho)

        linhas = ler_arquivo(str(caminho), aba="CP-NACOM")
        assert linhas[0][0] == "Titulo"
        assert linhas[1][0] == "2023"

    def test_aba_inexistente_levanta_erro(self, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        wb.active.title = "RESUMO"
        caminho = tmp_path / "cp.xlsx"
        wb.save(caminho)
        with pytest.raises(ValueError):
            ler_arquivo(str(caminho), aba="CP-NACOM")


class TestParsearArquivoIntegracao:
    """parsear_arquivo: le o arquivo e extrai boletos (ponta a ponta de um banco)."""

    def test_parsear_xlsx_agis(self, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Recebivel", "Tipo", "Modalidade", "Valor Face", "Valor Aberto"])
        ws.append(["146826/5", "DUPLICATA", "DESCONTO", "1745.17", "1745.17"])
        caminho = tmp_path / "agis.xlsx"
        wb.save(caminho)

        boletos = parsear_arquivo(str(caminho), "AGIS")
        assert len(boletos) == 1
        assert boletos[0]["nf_parc"] == "146826-5"
