"""Multi-abas (IMP-2026-06-11-002) — Excel com varias sheets na skill exportando-arquivos.

A skill so gerava 1 aba; relatorios com multiplas secoes (ex: vinculacao Teams com
Fantasmas Ativos / Bloqueados / Usuarios Sem Vinculo) forcavam workaround manual com
xlsxwriter via Bash. Este teste cobre `gerar_excel_multi_abas` de forma deterministica
(sem LLM) — preferencia do projeto: cobertura de skill = pytest.

Origem: dialogo de melhoria D8 (Agent SDK -> Claude Code), sugestao IMP-2026-06-11-002.
"""
import importlib.util
import os

import pytest

_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
_EXPORTAR = os.path.join(
    _REPO_ROOT, ".claude", "skills", "exportando-arquivos", "scripts", "exportar.py"
)


@pytest.fixture(scope="module")
def exportar():
    spec = importlib.util.spec_from_file_location("exportar_skill_multiabas", _EXPORTAR)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


@pytest.fixture(autouse=True)
def _isola_upload_dir(tmp_path, monkeypatch):
    # get_upload_folder() le AGENTE_FILES_ROOT; isolar para nao poluir /tmp real
    monkeypatch.setenv("AGENTE_FILES_ROOT", str(tmp_path))


def _load_sheets(filepath):
    import openpyxl

    wb = openpyxl.load_workbook(filepath)
    return wb.sheetnames, wb


class TestGerarExcelMultiAbas:
    def test_tres_abas_total_registros(self, exportar):
        abas = [
            {"titulo": "Aba A", "dados": [{"x": 1}, {"x": 2}]},
            {"titulo": "Aba B", "dados": [{"y": 3}]},
            {"titulo": "Aba C", "dados": [{"z": 4}, {"z": 5}, {"z": 6}]},
        ]
        filepath, filename, total = exportar.gerar_excel_multi_abas(abas, "rel")
        assert os.path.exists(filepath)
        assert filename.endswith("rel.xlsx")
        assert total == 6
        sheets, _ = _load_sheets(filepath)
        assert sheets == ["Aba A", "Aba B", "Aba C"]

    def test_titulo_truncado_31_chars(self, exportar):
        longo = "Resumo Geral Muito Longo Que Passa De Trinta E Um"
        abas = [{"titulo": longo, "dados": [{"a": 1}]}]
        filepath, _, _ = exportar.gerar_excel_multi_abas(abas, "trunc")
        sheets, _ = _load_sheets(filepath)
        assert len(sheets[0]) <= 31
        assert sheets[0] == longo[:31]

    def test_titulos_duplicados_recebem_sufixo(self, exportar):
        abas = [
            {"titulo": "Dup", "dados": [{"a": 1}]},
            {"titulo": "Dup", "dados": [{"a": 2}]},
            {"titulo": "Dup", "dados": [{"a": 3}]},
        ]
        filepath, _, _ = exportar.gerar_excel_multi_abas(abas, "dup")
        sheets, _ = _load_sheets(filepath)
        # Excel rejeita sheets com nome duplicado: todas devem ser unicas
        assert len(set(sheets)) == 3
        assert sheets[0] == "Dup"

    def test_aba_vazia_cria_sheet_sem_quebrar(self, exportar):
        abas = [
            {"titulo": "Com Dados", "dados": [{"a": 1}]},
            {"titulo": "Vazia", "dados": []},
        ]
        filepath, _, total = exportar.gerar_excel_multi_abas(abas, "vazia")
        sheets, _ = _load_sheets(filepath)
        assert "Vazia" in sheets
        assert total == 1

    def test_colunas_filtra_subconjunto(self, exportar):
        abas = [{
            "titulo": "Filtro",
            "dados": [{"nome": "A", "valor": 1, "interno": "x"}],
            "colunas": ["nome", "valor"],
        }]
        filepath, _, _ = exportar.gerar_excel_multi_abas(abas, "filtro")
        _, wb = _load_sheets(filepath)
        ws = wb["Filtro"]
        header = [c.value for c in ws[1]]
        assert header == ["nome", "valor"]
        assert "interno" not in header

    def test_abas_invalido_levanta_valueerror(self, exportar):
        with pytest.raises(ValueError):
            exportar.gerar_excel_multi_abas([], "x")
        with pytest.raises(ValueError):
            exportar.gerar_excel_multi_abas("nao-lista", "x")

    def test_aba_nao_dict_levanta_valueerror(self, exportar):
        with pytest.raises(ValueError):
            exportar.gerar_excel_multi_abas([["nao", "dict"]], "x")
