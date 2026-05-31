"""Testes do ContagemExportService (Excel base + relatório/plano)."""
import io
from decimal import Decimal

import openpyxl

from app.inventario.models import ContagemInventario, ContagemInventarioItem
from app.inventario.services.contagem_export_service import ContagemExportService
from app.utils.timezone import agora_utc_naive


def _contagem(db):
    c = ContagemInventario(
        codigo='CONT-EXP-1', empresa='FB', data_base=agora_utc_naive(),
        status='CONTABILIZADA', criado_em=agora_utc_naive(),
    )
    db.session.add(c)
    db.session.flush()
    db.session.add(ContagemInventarioItem(
        contagem_id=c.id, location_name='FB/Estoque', cod_produto='4000001',
        nome_produto='AZEITONA', lote='L1', company_id=1,
        qtd_esperada=Decimal('500'), reservado_esperado=Decimal('0'),
        contagem=Decimal('400'), ajuste=Decimal('-100'), classe='NORMAL'))
    db.session.add(ContagemInventarioItem(  # sem contagem -> não entra no relatório
        contagem_id=c.id, location_name='FB/Estoque', cod_produto='4000001',
        nome_produto='AZEITONA', lote='L2', company_id=1,
        qtd_esperada=Decimal('500'), reservado_esperado=Decimal('0'),
        contagem=None, ajuste=Decimal('0'), classe=None))
    db.session.flush()
    return c


def test_excel_base_gera_xlsx_valido(db):
    c = _contagem(db)
    data = ContagemExportService.excel_base(c)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    assert 'location_name' in header and 'AJUSTE' in header and 'CONTAGEM' in header
    assert ws.max_row == 3  # header + 2 itens da base


def test_excel_relatorio_so_contados(db):
    c = _contagem(db)
    data = ContagemExportService.excel_relatorio(c)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    assert 'ajuste' in header and 'classe' in header
    assert ws.max_row == 2  # header + só L1 (contagem != None)
