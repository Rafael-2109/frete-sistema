"""Fixtures compartilhadas para testes do módulo inventario."""
import io
import pytest
import openpyxl
from datetime import date
from app.inventario.models import CicloInventario


@pytest.fixture
def ciclo(db):
    c = CicloInventario(
        codigo='INV-TESTE-2026-05',
        data_snapshot=date(2026, 5, 16),
        descricao='Ciclo de teste',
        status='ATIVO',
        criado_por='pytest',
    )
    db.session.add(c)
    db.session.flush()
    return c


def _make_test_xlsx(rows_por_aba):
    """Cria xlsx em memória com abas FB/CD/LF + dados.

    rows_por_aba = {'FB': [(cod, lote, qtd), ...], ...}
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for aba, rows in rows_por_aba.items():
        ws = wb.create_sheet(aba)
        ws.append(['CODIGO', 'LOTE', 'QTD'])
        for r in rows:
            ws.append(list(r))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@pytest.fixture
def xlsx_valido():
    return _make_test_xlsx({
        'FB': [('4320147', '139/26', 100), ('208000041', '', 50)],
        'CD': [('4320147', '139/26', 200), ('103000037', '023/26', 75.5)],
        'LF': [('4320147', '139/26', 50)],
    })


@pytest.fixture
def xlsx_com_invalidos():
    """Inclui código que começa com letra (deve pular) e qtd negativa."""
    return _make_test_xlsx({
        'FB': [('CHAVE-X', '', 10), ('4320147', '139/26', 100)],
        'CD': [('208000041', '', -5)],
        'LF': [],
    })


def _make_xlsx_headers_reais():
    """xlsx com headers IDENTICOS aos da planilha fisica COMPILADO INV. 16.05.2026.

    FB:  CODIGO | DESCRIÇÃO | QUANTIDADE | VALIDADE | LOTE  | MEDIDA | LOCAL
    CD:  RUA/LADO | CODIGO | DESC | FINAL | LOTE | VALIDADE
    LF:  CODIGO | PRODUTO | QUANTIDADE/UN | LOTE | VALIDADE | LOCAL
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    fb = wb.create_sheet('FB')
    fb.append(['CODIGO', 'DESCRIÇÃO', 'QUANTIDADE', 'VALIDADE', 'LOTE ', 'MEDIDA', 'LOCAL'])
    fb.append(['4320147', 'AZEITONA VD', 100, None, '139/26', 'CX', 'A1'])
    fb.append(['208000041', 'PALMITO POUCH', 50.5, None, '', 'CX', 'B2'])

    cd = wb.create_sheet('CD')
    cd.append(['RUA/LADO', 'CODIGO', 'DESC', 'FINAL', 'LOTE', 'VALIDADE'])
    cd.append(['R01-A', '4320147', 'AZEITONA VD', 200, '139/26', None])
    cd.append(['R02-B', '103000037', 'INSUMO X', 75.5, '023/26', None])

    lf = wb.create_sheet('LF')
    lf.append(['CODIGO', 'PRODUTO', 'QUANTIDADE/UN', 'LOTE', 'VALIDADE', 'LOCAL'])
    lf.append(['4320147', 'AZEITONA VD', 50, '139/26', None, 'LF-01'])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@pytest.fixture
def xlsx_headers_reais():
    return _make_xlsx_headers_reais()
