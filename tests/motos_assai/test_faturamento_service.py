"""Testes do serviço de faturamento (geração Excel Q.P.A.)."""
import io
import uuid
from decimal import Decimal

import pytest
import openpyxl
from unittest.mock import patch

from app import db
from app.motos_assai.models import (
    AssaiPedidoVenda, AssaiPedidoVendaLoja, AssaiPedidoVendaItem,
    AssaiLoja, AssaiModelo,
    AssaiMoto, AssaiSeparacao,
    PEDIDO_STATUS_ABERTO, SEPARACAO_STATUS_FECHADA,
    EVENTO_ESTOQUE, EVENTO_MONTADA, EVENTO_DISPONIVEL,
)
from app.motos_assai.services import (
    # get_ou_criar_separacao foi renomeada para get_separacao_ativa
    # e perdeu o side-effect de criar implicitamente (Migration 17 corretivo).
    gerar_excel_qpa,
    registrar_chassi, finalizar_separacao,
    emitir_evento,
    criar_separacao_com_saldos,
)


def _uid():
    return uuid.uuid4().hex[:8].upper()


def _setup_separacao_fechada(app, admin):
    """Cria pedido + separação fechada com 1 chassi."""
    modelo_dot = AssaiModelo.query.filter_by(codigo='DOT').first()
    loja = AssaiLoja.query.first()

    uid = _uid()
    p = AssaiPedidoVenda(
        numero=f'TST-FAT-{uid}',
        # R4.2 (Big Bang Task 20): pedido fica ABERTO ate primeira NF.
        status=PEDIDO_STATUS_ABERTO,
        criado_por_id=admin.id,
    )
    db.session.add(p)
    db.session.flush()
    pvl = AssaiPedidoVendaLoja(pedido_id=p.id, loja_id=loja.id)
    db.session.add(pvl); db.session.flush()
    db.session.add(AssaiPedidoVendaItem(
        pedido_id=p.id, pedido_loja_id=pvl.id, loja_id=loja.id, modelo_id=modelo_dot.id,
        qtd_pedida=1, valor_unitario=Decimal('6900'), valor_total=Decimal('6900'),
    ))
    db.session.flush()

    chassi = f'TST_F_{_uid()}'
    m = AssaiMoto(chassi=chassi, modelo_id=modelo_dot.id, cor='PRETO')
    db.session.add(m)
    db.session.flush()
    emitir_evento(chassi, EVENTO_ESTOQUE, admin.id)
    emitir_evento(chassi, EVENTO_MONTADA, admin.id)
    emitir_evento(chassi, EVENTO_DISPONIVEL, admin.id)
    db.session.commit()

    # Cria sep EM_SEPARACAO explicitamente (registrar_chassi nao cria mais —
    # Migration 17 corretivo 2026-05-12).
    criar_separacao_com_saldos(
        pedido_id=p.id, loja_id=loja.id,
        alocacoes=[{'modelo_id': modelo_dot.id, 'qtd': 1}],
        operador_id=admin.id,
    )
    db.session.commit()

    registrar_chassi(p.id, loja.id, chassi, admin.id)
    sep = AssaiSeparacao.query.filter_by(pedido_id=p.id, loja_id=loja.id).first()
    finalizar_separacao(sep.id, admin.id)
    db.session.commit()

    return sep


def test_gerar_excel_estrutura_basica(app, admin_user):
    """gerar_excel_qpa deve retornar bytes com 2 abas: PEDIDO + BASE LOJAS."""
    with app.app_context():
        sep = _setup_separacao_fechada(app, admin_user)
        sep_id = sep.id

        with patch('app.motos_assai.services.faturamento_service.FileStorage') as mock_fs:
            mock_fs.return_value.save_file.return_value = f'motos_assai/solicitacoes/{sep_id}.xlsx'
            bytes_xlsx, s3_key = gerar_excel_qpa(sep_id, admin_user.id)

        assert isinstance(bytes_xlsx, bytes)
        assert len(bytes_xlsx) > 0
        assert s3_key

        wb = openpyxl.load_workbook(io.BytesIO(bytes_xlsx))
        assert 'PEDIDO' in wb.sheetnames, "Aba PEDIDO ausente"
        assert 'BASE LOJAS' in wb.sheetnames, "Aba BASE LOJAS ausente"

        ws_pedido = wb['PEDIDO']
        # Deve ter pelo menos 1 linha de dados de chassi
        valores = [ws_pedido.cell(row=r, column=2).value for r in range(1, ws_pedido.max_row + 1)]
        assert any(v and 'TST_F_' in str(v) for v in valores), "Chassi não encontrado no Excel"

        ws_lojas = wb['BASE LOJAS']
        assert ws_lojas.max_row > 1, "Aba BASE LOJAS deve ter lojas"

        db.session.rollback()


def test_gerar_excel_separacao_cancelada_falha(app, admin_user):
    """gerar_excel_qpa com separação CANCELADA deve levantar ValueError (H3).

    Mudanca (Plano Fase 2-3): EM_SEPARACAO agora e aceita para regeneracao
    apos substituicao cross-loja. Statuses validos: EM_SEPARACAO, FECHADA,
    CARREGADA, FATURADA. Apenas CANCELADA falha.
    """
    from app.motos_assai.models import SEPARACAO_STATUS_CANCELADA
    from app.motos_assai.services import cancelar_separacao

    with app.app_context():
        modelo_dot = AssaiModelo.query.filter_by(codigo='DOT').first()
        loja = AssaiLoja.query.first()
        uid = _uid()
        p = AssaiPedidoVenda(
            numero=f'TST-FAT2-{uid}',
            status=PEDIDO_STATUS_ABERTO,
            criado_por_id=admin_user.id,
        )
        db.session.add(p)
        db.session.flush()
        pvl2 = AssaiPedidoVendaLoja(pedido_id=p.id, loja_id=loja.id)
        db.session.add(pvl2); db.session.flush()
        db.session.add(AssaiPedidoVendaItem(
            pedido_id=p.id, pedido_loja_id=pvl2.id, loja_id=loja.id, modelo_id=modelo_dot.id,
            qtd_pedida=1, valor_unitario=Decimal('6900'), valor_total=Decimal('6900'),
        ))
        db.session.flush()

        # Cria sep EM_SEPARACAO + cancela
        criar_separacao_com_saldos(
            pedido_id=p.id, loja_id=loja.id,
            alocacoes=[{'modelo_id': modelo_dot.id, 'qtd': 1}],
            operador_id=admin_user.id,
        )
        db.session.commit()
        sep = AssaiSeparacao.query.filter_by(pedido_id=p.id, loja_id=loja.id).first()
        cancelar_separacao(sep.id, 'teste — sep cancelada', admin_user.id)
        db.session.commit()
        assert sep.status == SEPARACAO_STATUS_CANCELADA
        sep_id = sep.id

        # CANCELADA: service deve levantar ValueError
        import pytest as _pytest
        with _pytest.raises(ValueError, match='CANCELADA'):
            gerar_excel_qpa(sep_id, admin_user.id)

        db.session.rollback()
