"""Testes TDD para rotas Resultado por Frete (Tasks 6 + 7)."""
from io import BytesIO
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook


def _user(carvia=True, admin=True):
    u = MagicMock()
    u.is_authenticated = True
    u.sistema_carvia = carvia
    u.perfil = 'administrador' if admin else 'vendedor'
    u.email = 'test@bot'
    return u


def test_tela_resultado_frete_render(db, client):
    with patch('flask_login.utils._get_user', return_value=_user()):
        r = client.get('/carvia/resultado-frete')
    assert r.status_code == 200
    assert b'Resultado por Frete' in r.data


def test_tela_guard_sem_carvia(db, client):
    with patch('flask_login.utils._get_user', return_value=_user(carvia=False)):
        r = client.get('/carvia/resultado-frete')
    assert r.status_code in (301, 302)


def test_export_resultado_frete_duas_abas(db, client):
    with patch('flask_login.utils._get_user', return_value=_user()):
        r = client.get('/carvia/api/exportar/resultado-frete?data_inicio=2026-01-01&data_fim=2026-12-31')
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.data))
    assert 'Resumo' in wb.sheetnames
    assert 'Detalhe NF' in wb.sheetnames
