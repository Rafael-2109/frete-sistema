"""Testes das correcoes de export Excel CarVia:

1. DATAS como DATA nativa do Excel (nao texto) — para ordenarem cronologicamente
   (05/01/26 antes de 01/02/26), tanto no helper de duplo cabecalho quanto no
   gerador df-based `_gerar_excel`.
2. Coluna PESO CUBADO ao lado do peso da NF nos exports de NFs e Operacoes
   (volume x qtd x cubagem_minima==300, via MotoRecognitionService).
"""
from io import BytesIO
from datetime import date, datetime, timezone, timedelta
from decimal import Decimal
from unittest.mock import MagicMock, patch

import pandas as pd
from openpyxl import load_workbook


# --------------------------------------------------------------------------- #
#  Helpers
# --------------------------------------------------------------------------- #
def _user():
    u = MagicMock()
    u.is_authenticated = True
    u.sistema_carvia = True
    u.perfil = 'administrador'
    u.email = 'test@bot'
    return u


def _col_index(ws, grupo, label):
    """Indice 0-based da coluna (grupo na linha 1 merged, label na linha 2)."""
    grupos, atual = [], None
    for c in ws[1]:
        if c.value:
            atual = c.value
        grupos.append(atual)
    labels = [c.value for c in ws[2]]
    for i, (g, l) in enumerate(zip(grupos, labels)):
        if g == grupo and l == label:
            return i
    return None


def _linha_da_nf(ws, numero_nf):
    """Numero da linha (>=3) cujo campo NF/Numero == numero_nf. Outras NFs no
    banco de teste deslocam a ordenacao — nunca assumir linha 3."""
    idx = _col_index(ws, 'NF', 'Numero')
    assert idx is not None, 'coluna NF/Numero ausente'
    for row in range(3, ws.max_row + 1):
        if ws.cell(row=row, column=idx + 1).value == numero_nf:
            return row
    return None


def _criar_nf(db, numero, data_emissao):
    from app.carvia.models import CarviaNf
    nf = CarviaNf(
        numero_nf=numero,
        cnpj_emitente='11111111000111',
        nome_emitente='EMIT T',
        cnpj_destinatario='22222222000122',
        nome_destinatario='DEST T',
        uf_destinatario='SP',
        cidade_destinatario='SAO PAULO',
        data_emissao=data_emissao,
        valor_total=Decimal('500.00'),
        peso_bruto=Decimal('100.000'),
        status='ATIVA',
        tipo_fonte='MANUAL',
        criado_por='test',
    )
    db.session.add(nf)
    db.session.flush()
    return nf


def _criar_modelo(db, nome):
    """Modelo 1m x 1m x 1m (cm) => volume 1 m3; cubagem 300 => 300 kg/un."""
    from app.carvia.models import CarviaModeloMoto
    m = CarviaModeloMoto(
        nome=nome,
        comprimento=Decimal('100'),
        largura=Decimal('100'),
        altura=Decimal('100'),
        peso_medio=None,
        cubagem_minima=Decimal('300'),
        criado_por='test',
    )
    db.session.add(m)
    db.session.flush()
    return m


def _criar_item(db, nf_id, modelo_id, qtd):
    from app.carvia.models import CarviaNfItem
    it = CarviaNfItem(
        nf_id=nf_id,
        descricao='MOTO TESTE',
        quantidade=Decimal(str(qtd)),
        valor_unitario=Decimal('250.00'),
        valor_total_item=Decimal('500.00'),
        modelo_moto_id=modelo_id,
    )
    db.session.add(it)
    db.session.flush()
    return it


# --------------------------------------------------------------------------- #
#  1. Helper excel_export_helper (unit)
# --------------------------------------------------------------------------- #
def test_fmt_value_date_retorna_objeto_nao_string():
    from app.carvia.utils.excel_export_helper import _fmt_value
    out = _fmt_value(date(2026, 1, 5), 'date')
    assert isinstance(out, date)
    assert not isinstance(out, str)


def test_fmt_value_datetime_remove_tzinfo():
    from app.carvia.utils.excel_export_helper import _fmt_value
    aware = datetime(2026, 1, 5, 8, 30, tzinfo=timezone(timedelta(hours=-3)))
    out = _fmt_value(aware, 'datetime')
    assert isinstance(out, datetime)
    assert out.tzinfo is None  # Excel nao suporta timezone


def test_numberformat_date_e_datetime():
    from app.carvia.utils.excel_export_helper import _numberformat
    assert _numberformat('date') and 'YY' in _numberformat('date').upper()
    assert _numberformat('datetime') and 'HH' in _numberformat('datetime').upper()


# --------------------------------------------------------------------------- #
#  2. exportar_nfs — datas como data + peso cubado
# --------------------------------------------------------------------------- #
def test_export_nfs_data_emissao_e_data_nao_texto(db, client):
    _criar_nf(db, '88001', date(2026, 1, 5))
    db.session.commit()
    with patch('flask_login.utils._get_user', return_value=_user()):
        r = client.get('/carvia/api/exportar/nfs')
    assert r.status_code == 200
    ws = load_workbook(BytesIO(r.data)).active
    row = _linha_da_nf(ws, '88001')
    assert row is not None
    idx = _col_index(ws, 'NF', 'Data')
    assert idx is not None
    cell = ws.cell(row=row, column=idx + 1)
    assert isinstance(cell.value, (date, datetime)), f'esperado data, veio {type(cell.value)}'
    assert 'YY' in (cell.number_format or '').upper()


def test_export_nfs_inclui_peso_cubado(db, client):
    nf = _criar_nf(db, '88002', date(2026, 1, 10))
    m = _criar_modelo(db, 'MODELO_CUB_A')
    _criar_item(db, nf.id, m.id, qtd=2)  # 2 x 300 kg = 600 kg
    db.session.commit()
    with patch('flask_login.utils._get_user', return_value=_user()):
        r = client.get('/carvia/api/exportar/nfs')
    assert r.status_code == 200
    ws = load_workbook(BytesIO(r.data)).active
    row = _linha_da_nf(ws, '88002')
    assert row is not None
    idx = _col_index(ws, 'NF', 'Peso Cubado')
    assert idx is not None, 'coluna Peso Cubado ausente no grupo NF'
    cell = ws.cell(row=row, column=idx + 1)
    assert abs(float(cell.value) - 600.0) < 0.01, f'peso cubado errado: {cell.value}'


# --------------------------------------------------------------------------- #
#  3. exportar_operacoes — peso cubado
# --------------------------------------------------------------------------- #
def test_export_operacoes_inclui_peso_cubado(db, client):
    from app.carvia.models import CarviaOperacao, CarviaOperacaoNf
    nf = _criar_nf(db, '88003', date(2026, 2, 1))
    m = _criar_modelo(db, 'MODELO_CUB_B')
    _criar_item(db, nf.id, m.id, qtd=3)  # 3 x 300 = 900 kg
    op = CarviaOperacao(
        cte_numero='CTe-900',
        cte_valor=Decimal('1000.00'),
        cte_data_emissao=date(2026, 2, 2),
        cnpj_cliente='12345678000100',
        nome_cliente='Cli',
        uf_origem='SP', cidade_origem='SAO PAULO',
        uf_destino='RJ', cidade_destino='RIO DE JANEIRO',
        status='RASCUNHO', tipo_entrada='IMPORTADO', criado_por='test',
    )
    db.session.add(op)
    db.session.flush()
    db.session.add(CarviaOperacaoNf(operacao_id=op.id, nf_id=nf.id))
    db.session.commit()
    with patch('flask_login.utils._get_user', return_value=_user()):
        r = client.get('/carvia/api/exportar/operacoes')
    assert r.status_code == 200
    ws = load_workbook(BytesIO(r.data)).active
    row = _linha_da_nf(ws, '88003')
    assert row is not None
    idx = _col_index(ws, 'NF', 'Peso Cubado')
    assert idx is not None, 'coluna Peso Cubado ausente no grupo NF (operacoes)'
    cell = ws.cell(row=row, column=idx + 1)
    assert abs(float(cell.value) - 900.0) < 0.01, f'peso cubado errado: {cell.value}'


# --------------------------------------------------------------------------- #
#  4. _gerar_excel (df-based) — data como data
# --------------------------------------------------------------------------- #
def test_gerar_excel_escreve_date_como_data(app):
    from app.carvia.routes.exportacao_routes import _gerar_excel
    df = pd.DataFrame([
        {'Nome': 'A', 'Data': date(2026, 1, 5)},
        {'Nome': 'B', 'Data': date(2026, 2, 1)},
    ])
    with app.test_request_context():
        resp = _gerar_excel(df, 'Teste', 'teste')
        resp.direct_passthrough = False  # send_file usa passthrough; desligar p/ ler
        data = resp.get_data()
    ws = load_workbook(BytesIO(data)).active
    c1 = ws.cell(row=2, column=2)  # 1a linha de dados, coluna Data
    assert isinstance(c1.value, (date, datetime)), f'esperado data, veio {type(c1.value)}'
    assert 'YY' in (c1.number_format or '').upper()


def test_fmt_date_retorna_objeto_nao_string():
    from app.carvia.routes.exportacao_routes import _fmt_date, _fmt_datetime
    assert _fmt_date(date(2026, 1, 5)) == date(2026, 1, 5)
    assert _fmt_date(None) == ''
    out = _fmt_datetime(datetime(2026, 1, 5, 8, 30))
    assert isinstance(out, datetime)


# --------------------------------------------------------------------------- #
#  5. aplicar_formato_datas (helper compartilhado) — formata so' celulas data
# --------------------------------------------------------------------------- #
def test_aplicar_formato_datas_formata_data_e_ignora_texto():
    from openpyxl import Workbook
    from app.carvia.utils.excel_export_helper import aplicar_formato_datas
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value='Header')
    ws.cell(row=2, column=1, value=date(2026, 1, 5))      # data
    ws.cell(row=3, column=1, value=datetime(2026, 1, 5, 8, 30))  # datetime
    ws.cell(row=4, column=1, value='TOTAIS')              # texto -> nao mexe
    ws.cell(row=5, column=1, value=123.45)               # numero -> nao mexe
    aplicar_formato_datas(ws, min_row=2)
    assert 'YY' in ws.cell(row=2, column=1).number_format.upper()
    assert 'HH' in ws.cell(row=3, column=1).number_format.upper()
    assert 'YY' not in ws.cell(row=4, column=1).number_format.upper()
    assert 'YY' not in ws.cell(row=5, column=1).number_format.upper()
