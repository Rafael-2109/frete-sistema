# tests/carvia/test_export_nf_cidade.py
from io import BytesIO
from datetime import datetime
from decimal import Decimal
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook


def _user():
    u = MagicMock()
    u.is_authenticated = True
    u.sistema_carvia = True
    u.perfil = 'administrador'
    u.email = 'test@bot'
    return u


def _criar_nf_simples(db, numero, cidade, uf):
    from app.carvia.models import CarviaNf
    nf = CarviaNf(
        numero_nf=numero,
        cnpj_emitente='11111111000111',
        nome_emitente='EMIT T',
        cnpj_destinatario='22222222000122',
        nome_destinatario='DEST T',
        uf_destinatario=uf,
        cidade_destinatario=cidade,
        data_emissao=datetime(2026, 1, 10).date(),
        valor_total=Decimal('500.00'),
        peso_bruto=Decimal('100.000'),
        status='ATIVA',
        tipo_fonte='MANUAL',
        criado_por='test',
    )
    db.session.add(nf)
    db.session.flush()
    db.session.commit()  # savepoint commit — visivel ao request via join_transaction_mode
    return nf


def test_export_nfs_inclui_cidade_destino(db, client):
    _criar_nf_simples(db, '99001', 'RIBEIRAO PRETO', 'SP')
    with patch('flask_login.utils._get_user', return_value=_user()):
        r = client.get('/carvia/api/exportar/nfs')
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.data))
    ws = wb.active
    headers = [c.value for c in ws[2]]  # linha 2 = campos
    assert 'Cidade Dest' in headers
    valores = [c.value for row in ws.iter_rows(min_row=3) for c in row]
    assert 'RIBEIRAO PRETO' in valores
