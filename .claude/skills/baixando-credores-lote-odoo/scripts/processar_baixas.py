"""processar_baixas.py — skill `baixando-credores-lote-odoo` (passo 1a: PREVIEW).

Le a planilha de credores, localiza cada fatura de compra pelo `name` (coluna
FT REF = CMPMP), valida saldo/company/partner/journals e calcula o PLANO de pares
`account.payment` SICOOB (parcela) + DESAGIO (desagio) por vencimento.

ESTE SCRIPT NAO ESCREVE NO ODOO. Apenas READ (search_read). Produz:
  - relatorio JSON no stdout (resumo por status + detalhe por linha);
  - planilha anotada `<planilha>_PROCESSADO.xlsx` com o plano (copia; nunca
    destroi o original), ancorada pelo FT REF (CMPMP).

O WRITE real (criar/postar/reconciliar pagamento) e' o passo 1b e ainda NAO esta
implementado: `--confirmar` e' recusado de proposito (exit 2).

Escopo: FB(1)/LF(5). SC(3)/CD(4) -> BLOQUEADO_CROSS_COMPANY (Fase 2).

Exit: 4 preview OK (dry-run) · 1 erro de execucao (arquivo/conexao) · 2 uso/recusa.
"""
import argparse
import json
import os
import sys
from pathlib import Path

os.environ.setdefault('NACOM_QUIET_BOOT', '1')  # silencia prints de boot no import do app

_THIS = Path(__file__).resolve()
sys.path.insert(0, str(_THIS.parents[4]))  # .claude/skills/<skill>/scripts/<f> -> repo root

from app.odoo.estoque._cli_utils import criar_app_silencioso  # noqa: E402
from app.financeiro.services.baixa_credores_lote_service import (  # noqa: E402
    BaixaCredoresLoteService,
    parsear_planilha,
    _norm,
    _ALIASES,
    STATUS_OK,
)

import openpyxl  # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment  # noqa: E402

# Colunas de anotacao acrescentadas a planilha (apos as originais)
COLS_ANOTACAO = ['STATUS', 'MOTIVO', 'COMPANY', 'PARTNER (FATURA)', 'MOVE_ID',
                 'PAYABLE_LINE', 'RESIDUAL', 'TOTAL_PLANO', 'N_PARES', 'PLANO',
                 'AVISOS']

_FILL_OK = PatternFill('solid', fgColor='E2EFDA')
_FILL_BLOCK = PatternFill('solid', fgColor='FCE4D6')
_FILL_SKIP = PatternFill('solid', fgColor='FFF2CC')
_HDR_FILL = PatternFill('solid', fgColor='1F4E78')
_HDR_FONT = Font(bold=True, color='FFFFFF')


def _fmt_plano(res) -> str:
    return ' | '.join(
        f"{p.tipo} R$ {p.valor:.2f} j{p.journal_id} @{p.data.isoformat()}"
        for p in res.pares
    )


def _resultado_dict(res) -> dict:
    return {
        'linha': res.linha.idx,
        'credor': res.linha.credor,
        'ft_ref': res.linha.ft_ref,
        'empresa_planilha': res.linha.empresa,
        'status': res.status,
        'motivo': res.motivo,
        'company_id': res.company_id,
        'partner_id': res.partner_id,
        'partner_nome': res.partner_nome,
        'move_id': res.move_id,
        'payable_line_id': res.payable_line_id,
        'residual': res.residual,
        'total_plano': res.total,
        'n_pares': len(res.pares),
        'plano': _fmt_plano(res),
        'avisos': res.avisos,
    }


def _achar_col_ftref(ws) -> int:
    """Indice (1-based) da coluna FT REF na planilha, ancora para a anotacao."""
    for c in range(1, ws.max_column + 1):
        if _norm(ws.cell(row=1, column=c).value) in _ALIASES['ft_ref']:
            return c
    return 0


def gerar_planilha_anotada(planilha_in: str, saida: str, resultados) -> dict:
    """Copia a planilha original e anexa o plano por linha, ANCORADO pelo FT REF
    (nunca por indice posicional cego — guard F5). Retorna stats de fechamento."""
    wb = openpyxl.load_workbook(planilha_in)
    ws = wb.active
    base_cols = ws.max_column
    ftref_col = _achar_col_ftref(ws)

    # cabecalho das colunas de anotacao
    for i, h in enumerate(COLS_ANOTACAO):
        cell = ws.cell(row=1, column=base_cols + 1 + i, value=h)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal='center')

    por_idx = {r.linha.idx: r for r in resultados}
    anotadas = desalinhadas = 0
    for idx, res in por_idx.items():
        # ancoragem por FT REF: a celula da linha tem de bater com o ft_ref do resultado
        if ftref_col:
            valor_na_linha = _norm(ws.cell(row=idx, column=ftref_col).value)
            if valor_na_linha != _norm(res.linha.ft_ref):
                desalinhadas += 1
                res.avisos.append('DESALINHAMENTO: FT REF da linha nao bate com o plano')
        vals = [res.status, res.motivo, res.company_id, res.partner_nome, res.move_id,
                res.payable_line_id, res.residual, res.total, len(res.pares),
                _fmt_plano(res), '; '.join(res.avisos)]
        fill = _FILL_OK if res.status in STATUS_OK else (
            _FILL_SKIP if res.status in ('PULADO_SEM_DADOS', 'JA_PROCESSADO') else _FILL_BLOCK)
        for i, v in enumerate(vals):
            cell = ws.cell(row=idx, column=base_cols + 1 + i, value=v)
            cell.fill = fill
        anotadas += 1

    wb.save(saida)
    return {'linhas_anotadas': anotadas, 'desalinhamentos_ftref': desalinhadas}


def main() -> int:
    ap = argparse.ArgumentParser(description=(__doc__ or '').split('\n')[0])
    ap.add_argument('--planilha', required=True, help='caminho da planilha de credores (.xlsx)')
    ap.add_argument('--saida', default=None, help='saida anotada (default <planilha>_PROCESSADO.xlsx)')
    ap.add_argument('--sheet', default=None, help='nome da aba (default: ativa)')
    ap.add_argument('--credor', default=None, help='filtra por credor (substring, case-insensitive)')
    ap.add_argument('--tolerancia-saldo', type=float, default=0.01, help='tolerancia do guard de saldo')
    ap.add_argument('--user-id', type=int, default=None, help='id do usuario (auditoria; obrig. no 1b)')
    ap.add_argument('--quiet', action='store_true', help='suprime stdout do boot Flask')
    ap.add_argument('--confirmar', action='store_true',
                    help='[1b — NAO IMPLEMENTADO] efetivaria o WRITE no Odoo; recusado nesta versao')
    args = ap.parse_args()

    # Recusa explicita do WRITE (1b ainda nao existe) — evita falsa sensacao de pagamento.
    if args.confirmar:
        print(json.dumps({
            'status': 'RECUSADO',
            'erro': 'modo WRITE (--confirmar) e o passo 1b e ainda NAO esta implementado; '
                    'rode SEM --confirmar para o PREVIEW (READ-only).',
        }, ensure_ascii=False, indent=2))
        return 2

    planilha = Path(args.planilha)
    if not planilha.exists():
        print(json.dumps({'status': 'ERRO', 'erro': f'planilha nao encontrada: {planilha}'},
                         ensure_ascii=False, indent=2))
        return 1
    saida = args.saida or str(planilha.with_name(planilha.stem + '_PROCESSADO.xlsx'))

    try:
        linhas = parsear_planilha(str(planilha), sheet=args.sheet)
    except Exception as exc:  # noqa: BLE001
        print(json.dumps({'status': 'ERRO', 'erro': f'falha ao ler planilha: {exc}'},
                         ensure_ascii=False, indent=2))
        return 1

    if args.credor:
        alvo = args.credor.lower()
        linhas = [l for l in linhas if alvo in (l.credor or '').lower()]

    if not linhas:
        print(json.dumps({'status': 'VAZIO', 'planilha': str(planilha),
                          'erro': 'nenhuma linha de credor encontrada (apos filtro)'},
                         ensure_ascii=False, indent=2))
        return 1

    app = criar_app_silencioso(args.quiet)
    with app.app_context():
        try:
            svc = BaixaCredoresLoteService(tolerancia_saldo=args.tolerancia_saldo)
            resultados = svc.gerar_preview(linhas)
        except Exception as exc:  # noqa: BLE001
            print(json.dumps({'status': 'ERRO', 'erro': f'falha no preview (Odoo READ): {exc}'},
                             ensure_ascii=False, indent=2))
            return 1

    try:
        fechamento = gerar_planilha_anotada(str(planilha), saida, resultados)
    except Exception as exc:  # noqa: BLE001
        fechamento = {'erro_planilha_anotada': str(exc)}

    # resumo por status
    resumo = {}
    total_a_pagar = 0.0
    for r in resultados:
        resumo[r.status] = resumo.get(r.status, 0) + 1
        if r.status in STATUS_OK and r.total:
            total_a_pagar += r.total

    out = {
        'modo': 'PREVIEW (dry-run, READ-only — zero escrita no Odoo)',
        'planilha': str(planilha),
        'saida_anotada': saida,
        'total_linhas': len(resultados),
        'resumo_por_status': resumo,
        'total_a_pagar_dry_run_ok': round(total_a_pagar, 2),
        'fechamento_planilha': fechamento,
        'detalhe': [_resultado_dict(r) for r in resultados],
    }
    print(json.dumps(out, ensure_ascii=False, indent=2, default=str))
    return 4  # dry-run/preview OK


if __name__ == '__main__':
    sys.exit(main())
