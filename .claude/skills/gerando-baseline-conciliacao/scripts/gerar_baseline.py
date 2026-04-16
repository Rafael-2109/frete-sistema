"""
Gerador de baseline canonico de extratos pendentes (Marcus - user_id=18).

Formato travado em /memories/preferences.xml secao baseline_conciliacoes:
- 4 abas: Pendentes Mes x Journal, Pendentes, Conciliacoes Dia Anterior, Resumo
- Fonte: Odoo account.bank.statement.line is_reconciled=False
- Nomes reais via write_uid (nao SYNC_*)

Uso:
  python gerar_baseline.py [--output-dir /tmp] [--data-referencia 2026-04-17]

Saida:
  extratos_pendentes_mes_journal_<DDmmmYYYY>.xlsx

Ref: .claude/skills/gerando-baseline-conciliacao/SKILL.md
"""

import argparse
import os
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta

# Adicionar raiz do projeto ao path
_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..', '..'))
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)

JOURNALS_MONITORADOS = ['SICOOB', 'GRAFENO', 'BRADESCO', 'AGIS GARANTIDA', 'VORTX GRAFENO']
COMPANY_ID_FB = 1  # Nacom Goya / Conservas Campo Belo
TOP_N_PENDENTES = 500

MES_PT = {
    1: 'Jan', 2: 'Fev', 3: 'Mar', 4: 'Abr', 5: 'Mai', 6: 'Jun',
    7: 'Jul', 8: 'Ago', 9: 'Set', 10: 'Out', 11: 'Nov', 12: 'Dez',
}


def _fmt_mes_ano(data_str):
    """2026-04-16 -> '04/2026'"""
    if not data_str:
        return ''
    ymd = data_str.split(' ')[0]  # pode vir com hora
    y, m, _ = ymd.split('-')
    return f'{m}/{y}'


def _fmt_data_ref(d):
    """date(2026,4,17) -> '17Abr2026'"""
    return f'{d.day:02d}{MES_PT[d.month]}{d.year}'


def query_odoo_pendentes(odoo_conn):
    """Query 1: Agregacao Mes x Journal + Query 2: Top N detalhadas."""
    models = odoo_conn['models']
    uid = odoo_conn['uid']
    db = odoo_conn['db']
    password = odoo_conn['password']

    # Buscar IDs dos journals monitorados
    journals = models.execute_kw(db, uid, password,
        'account.journal', 'search_read',
        [[['name', 'in', JOURNALS_MONITORADOS],
          ['company_id', '=', COMPANY_ID_FB]]],
        {'fields': ['id', 'name']})
    journal_ids = [j['id'] for j in journals]
    journal_map = {j['id']: j['name'] for j in journals}

    if not journal_ids:
        print('[ERRO] Nenhum journal monitorado encontrado no Odoo')
        return {}, [], journal_map

    # Buscar todas as linhas pendentes em batches
    agg = defaultdict(lambda: {'linhas': 0, 'pgtos': 0, 'recebs': 0, 'vl_deb': 0.0, 'vl_cred': 0.0})
    todas_linhas = []

    offset = 0
    batch_size = 2000
    while True:
        linhas = models.execute_kw(db, uid, password,
            'account.bank.statement.line', 'search_read',
            [[['is_reconciled', '=', False],
              ['journal_id', 'in', journal_ids],
              ['company_id', '=', COMPANY_ID_FB]]],
            {'fields': ['journal_id', 'date', 'amount', 'payment_ref',
                        'partner_id', 'payment_id'],
             'limit': batch_size, 'offset': offset})
        if not linhas:
            break

        for linha in linhas:
            mes = _fmt_mes_ano(linha.get('date', ''))
            journal = journal_map.get(linha['journal_id'][0], 'UNKNOWN') if linha.get('journal_id') else 'UNKNOWN'
            key = (mes, journal)
            amount = linha.get('amount', 0.0)
            agg[key]['linhas'] += 1
            if amount < 0:
                agg[key]['pgtos'] += 1
                agg[key]['vl_deb'] += amount
            elif amount > 0:
                agg[key]['recebs'] += 1
                agg[key]['vl_cred'] += amount

            # Acumular para aba 2
            todas_linhas.append({
                'mes': mes,
                'journal': journal,
                'date': linha.get('date', '')[:10],
                'payment_ref': linha.get('payment_ref') or linha.get('name') or '',
                'partner': (linha.get('partner_id') or [None, ''])[1] if linha.get('partner_id') else '',
                'amount': amount,
                'payment_id': (linha.get('payment_id') or [None])[0] if linha.get('payment_id') else None,
            })

        offset += batch_size
        if len(linhas) < batch_size:
            break

    # Top N por valor absoluto
    todas_linhas.sort(key=lambda x: abs(x['amount']), reverse=True)
    pendentes_top = todas_linhas[:TOP_N_PENDENTES]

    return dict(agg), pendentes_top, journal_map


def query_conciliacoes_d1(odoo_conn, data_ref):
    """Query 3: Conciliacoes D-1 (UNIAO 3 fontes).

    Aba 3 - armadilha documentada: consultar apenas 1 fonte = contagem errada.
    """
    from app import db

    ontem = (data_ref - timedelta(days=1))
    inicio = f'{ontem.isoformat()} 00:00:00'
    fim = f'{ontem.isoformat()} 23:59:59'

    per_user = defaultdict(lambda: {'linhas': 0, 'pgtos': 0, 'recebs': 0, 'vl_deb': 0.0, 'vl_cred': 0.0})

    # Fonte 1: Odoo account.bank.statement.line
    try:
        models = odoo_conn['models']
        uid = odoo_conn['uid']
        odoo_db = odoo_conn['db']
        password = odoo_conn['password']

        journals = models.execute_kw(odoo_db, uid, password,
            'account.journal', 'search_read',
            [[['name', 'in', JOURNALS_MONITORADOS],
              ['company_id', '=', COMPANY_ID_FB]]],
            {'fields': ['id']})
        journal_ids = [j['id'] for j in journals]

        conciliacoes = models.execute_kw(odoo_db, uid, password,
            'account.bank.statement.line', 'search_read',
            [[['is_reconciled', '=', True],
              ['write_date', '>=', inicio],
              ['write_date', '<=', fim],
              ['journal_id', 'in', journal_ids],
              ['company_id', '=', COMPANY_ID_FB]]],
            {'fields': ['write_uid', 'amount']})

        user_ids = list({c['write_uid'][0] for c in conciliacoes if c.get('write_uid')})
        user_map = {}
        if user_ids:
            users = models.execute_kw(odoo_db, uid, password,
                'res.users', 'read',
                [user_ids, ['id', 'name']])
            user_map = {u['id']: u['name'] for u in users}

        for c in conciliacoes:
            if not c.get('write_uid'):
                continue
            nome = user_map.get(c['write_uid'][0], f"USER_{c['write_uid'][0]}")
            # Armadilha: pular SYNC_*
            if nome.upper().startswith('SYNC_'):
                continue
            amount = c.get('amount', 0.0)
            per_user[nome]['linhas'] += 1
            if amount < 0:
                per_user[nome]['pgtos'] += 1
                per_user[nome]['vl_deb'] += amount
            elif amount > 0:
                per_user[nome]['recebs'] += 1
                per_user[nome]['vl_cred'] += amount
    except Exception as e:
        print(f'[WARN] Fonte 1 (Odoo) falhou: {e}')

    # Fonte 2: Local lancamento_comprovante
    try:
        from sqlalchemy import text as sql_text
        result = db.session.execute(sql_text("""
            SELECT COALESCE(criado_por, 'SEM_USUARIO') AS usuario,
                   COALESCE(valor, 0) AS amount
            FROM lancamento_comprovante
            WHERE DATE(criado_em) = :ontem
        """), {'ontem': ontem})
        for row in result:
            nome = row[0]
            if not nome or nome.upper().startswith('SYNC_'):
                continue
            amount = float(row[1] or 0)
            per_user[nome]['linhas'] += 1
            if amount < 0:
                per_user[nome]['pgtos'] += 1
                per_user[nome]['vl_deb'] += amount
            elif amount > 0:
                per_user[nome]['recebs'] += 1
                per_user[nome]['vl_cred'] += amount
    except Exception as e:
        print(f'[WARN] Fonte 2 (lancamento_comprovante) falhou: {e}')

    # Fonte 3: Local carvia_conciliacoes
    try:
        from sqlalchemy import text as sql_text
        result = db.session.execute(sql_text("""
            SELECT COALESCE(criado_por, 'SEM_USUARIO') AS usuario,
                   COALESCE(valor_total, 0) AS amount
            FROM carvia_conciliacoes
            WHERE DATE(criado_em) = :ontem
        """), {'ontem': ontem})
        for row in result:
            nome = row[0]
            if not nome or nome.upper().startswith('SYNC_'):
                continue
            amount = float(row[1] or 0)
            per_user[nome]['linhas'] += 1
            if amount < 0:
                per_user[nome]['pgtos'] += 1
                per_user[nome]['vl_deb'] += amount
            elif amount > 0:
                per_user[nome]['recebs'] += 1
                per_user[nome]['vl_cred'] += amount
    except Exception as e:
        print(f'[WARN] Fonte 3 (carvia_conciliacoes) falhou: {e}')

    return dict(per_user)


def montar_excel(agg, pendentes, d1, _journal_map, data_ref, output_path):
    """Monta Excel com 4 abas canonicas."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Estilos
    bold = Font(bold=True)
    italic = Font(italic=True, size=10)
    fill_total = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    fill_verde = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fill_header = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    fmt_neg = '#,##0.00;[Red]-#,##0.00'

    # -------- Aba 1: Pendentes Mes x Journal --------
    ws1 = wb.active
    ws1.title = 'Pendentes Mes x Journal'
    headers1 = ['Mes', 'Journal', 'Linhas', 'PGTOS', 'Valor Debitos', 'RECEB.', 'Valor Creditos']
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = bold
        cell.fill = fill_header

    total = {'linhas': 0, 'pgtos': 0, 'vl_deb': 0.0, 'recebs': 0, 'vl_cred': 0.0}
    sorted_keys = sorted(agg.keys(), key=lambda k: (k[0], k[1]))
    for (mes, journal) in sorted_keys:
        d = agg[(mes, journal)]
        ws1.append([mes, journal, d['linhas'], d['pgtos'], d['vl_deb'], d['recebs'], d['vl_cred']])
        total['linhas'] += d['linhas']
        total['pgtos'] += d['pgtos']
        total['vl_deb'] += d['vl_deb']
        total['recebs'] += d['recebs']
        total['vl_cred'] += d['vl_cred']

    # Linha TOTAL
    ws1.append(['TOTAL', '', total['linhas'], total['pgtos'], total['vl_deb'],
                total['recebs'], total['vl_cred']])
    total_row = ws1.max_row
    for c in ws1[total_row]:
        c.font = bold
        c.fill = fill_total

    # Secao Evolucao Baseline
    ws1.append([])
    ws1.append(['Evolucao Baseline'])
    ws1.cell(row=ws1.max_row, column=1).font = bold
    ws1.append(['09/Abr/2026', 8684])
    ws1.append(['16/Abr/2026', 6985, 'delta=-1699'])
    delta = total['linhas'] - 6985
    data_ref_label = f'{data_ref.day:02d}/{MES_PT[data_ref.month]}/{data_ref.year}'
    ws1.append([data_ref_label, total['linhas'], f'delta={delta:+d}'])
    # Italico para secao
    for r in range(ws1.max_row - 2, ws1.max_row + 1):
        for c in ws1[r]:
            c.font = italic

    # Formatacao numeros
    for row in ws1.iter_rows(min_row=2, max_row=total_row):
        row[4].number_format = fmt_neg  # Valor Debitos
        row[6].number_format = '#,##0.00'  # Valor Creditos

    # Larguras de coluna
    for i, w in enumerate([10, 20, 10, 10, 18, 10, 18]):
        ws1.column_dimensions[get_column_letter(i + 1)].width = w

    # -------- Aba 2: Pendentes (detalhadas) --------
    ws2 = wb.create_sheet('Pendentes')
    headers2 = ['Mes', 'Journal', 'Data', 'Descricao', 'Partner', 'Valor', 'payment_id']
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = bold
        cell.fill = fill_header
    ws2.freeze_panes = 'A2'

    for p in pendentes:
        ws2.append([p['mes'], p['journal'], p['date'], p['payment_ref'],
                    p['partner'], p['amount'], p['payment_id']])

    for row in ws2.iter_rows(min_row=2):
        row[5].number_format = fmt_neg  # Valor preserva sinal

    for i, w in enumerate([10, 20, 12, 50, 30, 15, 12]):
        ws2.column_dimensions[get_column_letter(i + 1)].width = w

    # -------- Aba 3: Conciliacoes Dia Anterior --------
    ws3 = wb.create_sheet('Conciliacoes Dia Anterior')
    headers3 = ['Usuario', 'Linhas', 'Pgtos', 'Valor Debitos', 'Rec', 'Valor Creditos']
    ws3.append(headers3)
    for cell in ws3[1]:
        cell.font = bold
        cell.fill = fill_header

    if not d1:
        ws3.append(['Nenhuma conciliacao registrada', 0, 0, 0.0, 0, 0.0])
    else:
        total3 = {'linhas': 0, 'pgtos': 0, 'vl_deb': 0.0, 'recebs': 0, 'vl_cred': 0.0}
        for nome in sorted(d1.keys(), key=lambda n: d1[n]['linhas'], reverse=True):
            d = d1[nome]
            ws3.append([nome, d['linhas'], d['pgtos'], d['vl_deb'], d['recebs'], d['vl_cred']])
            total3['linhas'] += d['linhas']
            total3['pgtos'] += d['pgtos']
            total3['vl_deb'] += d['vl_deb']
            total3['recebs'] += d['recebs']
            total3['vl_cred'] += d['vl_cred']
        ws3.append(['TOTAL', total3['linhas'], total3['pgtos'], total3['vl_deb'],
                    total3['recebs'], total3['vl_cred']])
        tr = ws3.max_row
        for c in ws3[tr]:
            c.font = bold
            c.fill = fill_total

    for row in ws3.iter_rows(min_row=2):
        row[3].number_format = fmt_neg
        row[5].number_format = '#,##0.00'

    for i, w in enumerate([28, 10, 10, 18, 10, 18]):
        ws3.column_dimensions[get_column_letter(i + 1)].width = w

    # -------- Aba 4: Resumo (pivot) --------
    ws4 = wb.create_sheet('Resumo')
    ws4.append(['Rotulos de Linha', 'Soma de PGTOS', 'Soma de RECEB'])
    for cell in ws4[1]:
        cell.font = bold
        cell.fill = fill_header

    # Organizar pivot: mes -> journal -> {pgtos, recebs}
    pivot = defaultdict(lambda: defaultdict(lambda: {'pgtos': 0, 'recebs': 0}))
    for (mes, journal), d in agg.items():
        pivot[mes][journal]['pgtos'] += d['pgtos']
        pivot[mes][journal]['recebs'] += d['recebs']

    total_pgtos = 0
    total_recebs = 0
    for mes in sorted(pivot.keys()):
        # Subtotal do mes (verde claro)
        tot_mes_p = sum(j['pgtos'] for j in pivot[mes].values())
        tot_mes_r = sum(j['recebs'] for j in pivot[mes].values())
        ws4.append([mes, tot_mes_p, tot_mes_r])
        row_idx = ws4.max_row
        for c in ws4[row_idx]:
            c.fill = fill_verde
            c.font = bold
        total_pgtos += tot_mes_p
        total_recebs += tot_mes_r

        # Sub-itens (journals alfabetico)
        for journal in sorted(pivot[mes].keys()):
            j = pivot[mes][journal]
            ws4.append([f'  {journal}', j['pgtos'], j['recebs']])

    # TOTAL GERAL
    ws4.append(['Total Geral', total_pgtos, total_recebs])
    tr = ws4.max_row
    for c in ws4[tr]:
        c.font = bold
        c.fill = fill_total

    for i, w in enumerate([25, 18, 18]):
        ws4.column_dimensions[get_column_letter(i + 1)].width = w

    # Salvar
    wb.save(output_path)
    return {
        'aba1_linhas': len(agg),
        'aba2_linhas': len(pendentes),
        'aba3_usuarios': len(d1),
        'aba4_linhas': sum(len(journals) + 1 for journals in pivot.values()) + 1,
        'total_pendentes': total['linhas'],
    }


def main():
    parser = argparse.ArgumentParser(description='Gerar baseline canonico de extratos pendentes')
    parser.add_argument('--output-dir', default='/tmp', help='Diretorio de saida')
    parser.add_argument('--data-referencia', default=None, help='Data ref ISO (default hoje)')
    args = parser.parse_args()

    data_ref = datetime.fromisoformat(args.data_referencia).date() if args.data_referencia else date.today()
    output_file = os.path.join(args.output_dir, f'extratos_pendentes_mes_journal_{_fmt_data_ref(data_ref)}.xlsx')

    print(f'[{datetime.now().isoformat(timespec="seconds")}] Gerando baseline data_ref={data_ref.isoformat()} -> {output_file}')

    # Conectar Flask app + Odoo
    from app import create_app
    app = create_app()
    with app.app_context():
        from app.odoo.utils.connection import get_odoo_connection
        odoo_conn = get_odoo_connection()
        if not odoo_conn:
            print('[ERRO] Nao foi possivel conectar ao Odoo')
            sys.exit(1)

        print('[1/4] Consultando pendentes por Mes x Journal (Odoo)...')
        agg, pendentes, journal_map = query_odoo_pendentes(odoo_conn)
        print(f'       {len(agg)} combinacoes Mes x Journal, {len(pendentes)} linhas detalhadas (top N).')

        print('[2/4] Consultando conciliacoes D-1 (UNIAO 3 fontes)...')
        d1 = query_conciliacoes_d1(odoo_conn, data_ref)
        print(f'       {len(d1)} usuarios com conciliacoes em D-1.')

        print('[3/4] Calculando pivot para Resumo...')
        # Pivot e derivado dentro do montar_excel

        print('[4/4] Montando Excel...')
        result = montar_excel(agg, pendentes, d1, journal_map, data_ref, output_file)

    print()
    print(f'[OK] Baseline gerado: {output_file}')
    print(f'     Aba 1 (Pendentes Mes x Journal): {result["aba1_linhas"]} linhas')
    print(f'     Aba 2 (Pendentes): {result["aba2_linhas"]} linhas')
    print(f'     Aba 3 (Conciliacoes D-1): {result["aba3_usuarios"]} usuarios')
    print(f'     Aba 4 (Resumo): {result["aba4_linhas"]} linhas + Total Geral')
    print(f'     Total extratos pendentes: {result["total_pendentes"]}')


if __name__ == '__main__':
    main()
