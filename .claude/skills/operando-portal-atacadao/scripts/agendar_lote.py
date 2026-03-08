#!/usr/bin/env python3
"""
agendar_lote.py -- Agendamento em lote no portal Atacadao via upload de planilha.

Tres fases:
  FASE 1: Construir planilha a partir do CSV de saldo (ou usar planilha pronta)
  FASE 2: Upload, mapeamento de colunas e validacao no portal
  FASE 3: Salvar (somente sem --dry-run e sem inconsistencias)

Uso:
    # Modo 1: A partir do CSV de saldo + parametros
    python agendar_lote.py --saldo-csv /tmp/saldo_atacadao/saldo.csv \\
        --data 20/03/2026 --veiculo 7 --dry-run

    # Modo 2: Planilha ja preparada
    python agendar_lote.py --planilha /path/to/prepared.xlsx --dry-run

    # Confirmar (apos dry-run bem-sucedido)
    python agendar_lote.py --planilha /tmp/agendamento_atacadao/agendamento.xlsx --confirmar

Saida:
    JSON padronizado com resultado da validacao/agendamento.
"""
import argparse
import csv
import io
import logging
import math
import os
import re
import sys
from datetime import datetime
from pathlib import Path

# Adicionar raiz do projeto ao path para imports
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..", ".."))
sys.path.insert(0, PROJECT_ROOT)

from dotenv import load_dotenv
load_dotenv(os.path.join(PROJECT_ROOT, ".env"))

from atacadao_common import (
    gerar_saida,
    capturar_screenshot,
    verificar_sessao_sync,
    criar_sessao_download,
    ATACADAO_URLS,
    TIMEOUTS,
)

logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────
# Constantes
# ──────────────────────────────────────────────

OUTPUT_DIR = Path("/tmp/agendamento_atacadao")

URL_CARGAS_PLANILHA = ATACADAO_URLS["cargas_planilha"]

# CNPJ padrao do CD (Nacom Goya)
CNPJ_CD_PADRAO = "61724241000330"

SELETORES_LOTE = {
    # Pagina cargas-planilha
    "download_modelo": 'a.btn.btn-primary.pull-right, a[href*="download_modelo"]',
    "upload_browser": "#uploadForm > div > span > button.btn.btn-primary.inputfile-browser",
    "input_file": '#uploadForm input[type="file"]',
    "enviar": "#enviar",

    # Mapeamento de colunas (apos upload)
    # Botao real: [name=analisar] (jQuery handler faz PUT /analisa)
    "confirmar_mapeamento": "[name=analisar]",

    # Validacao
    "linha_valida": "tr td.bg-white",
    "linha_inconsistente": "tr td.bg-danger",
    "texto_critica": "i.text-danger",
    "texto_acerte": 'i:has-text("Acerte os itens")',
    "tabela_validacao": "table.table",

    # Salvar
    "salvar": "#salvar2",

    # Modal sucesso
    "modal_sucesso": '.modal-content:has-text("Registro criado com sucesso")',
    "link_listar_cargas": 'a.btn.btn-primary:has-text("Listar cargas")',
    "botao_ok": "#footerModalAlerta > div > div > button",
}

# Mapeamento Cod. planilha → nome veiculo (para validacao)
VEICULOS_PLANILHA = {
    "1": "Kombi/Van",
    "2": "Carreta-Graneleira",
    "3": "F4000-3/4 Bau",
    "4": "Toco-Bau",
    "5": "Truck-Bau",
    "6": "Truck-Sider",
    "7": "Carreta-Bau",
    "8": "Carreta-Sider",
    "9": "Carreta-Container",
    "10": "Bitrem-Graneleiro",
    "11": "Rodotrem-Bau",
    "12": "Truck-Graneleiro",
}


# ──────────────────────────────────────────────
# Validacao de entradas
# ──────────────────────────────────────────────

def _validar_data(data_str):
    """Valida formato de data(s) DD/MM/YYYY. Aceita multiplas separadas por virgula.

    Args:
        data_str: Data(s) como string, ex: "24/03/2026" ou "24/03/2026,25/03/2026,26/03/2026"

    Returns:
        list[str]: Lista de datas validadas

    Raises:
        ValueError: Se formato invalido
    """
    datas = [d.strip() for d in data_str.split(",") if d.strip()]
    if not datas:
        raise ValueError(f"Nenhuma data valida em: '{data_str}'")

    for d in datas:
        try:
            datetime.strptime(d, "%d/%m/%Y")
        except ValueError:
            raise ValueError(
                f"Data invalida: '{d}'. Formato esperado: DD/MM/YYYY"
            )

    return datas


def _validar_veiculo(cod_veiculo):
    """Valida codigo de veiculo da planilha.

    Args:
        cod_veiculo: Codigo do veiculo (string)

    Returns:
        str: Codigo validado

    Raises:
        ValueError: Se codigo invalido
    """
    cod = str(cod_veiculo)
    if cod not in VEICULOS_PLANILHA:
        codigos_validos = ", ".join(
            f"{k}={v}" for k, v in sorted(VEICULOS_PLANILHA.items(), key=lambda x: int(x[0]))
        )
        raise ValueError(
            f"Codigo de veiculo invalido: '{cod_veiculo}'. "
            f"Validos: {codigos_validos}"
        )
    return cod


def _limpar_cnpj(cnpj):
    """Limpa formatacao de CNPJ."""
    return re.sub(r"[.\-/]", "", (cnpj or "").strip())


def _formatar_cnpj(cnpj_limpo):
    """Formata CNPJ: 61724241000330 → 61.724.241/0003-30."""
    c = cnpj_limpo
    if len(c) == 14:
        return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:14]}"
    return c


# ──────────────────────────────────────────────
# FASE 1: Construir planilha a partir do CSV de saldo
# ──────────────────────────────────────────────

def _resolver_palletizacao(ean):
    """Busca palletizacao de um produto pelo EAN no cadastro local.

    Requer acesso ao banco de dados (create_app + app_context).

    Args:
        ean: Codigo EAN do produto

    Returns:
        tuple: (palletizacao, cod_produto, nome_produto)

    Raises:
        ValueError: Se produto nao encontrado ou sem palletizacao valida
    """
    _old_stdout = sys.stdout
    sys.stdout = io.StringIO()
    try:
        from app import create_app, db

        app = create_app()
        with app.app_context():
            resultado = db.session.execute(
                db.text("""
                    SELECT cod_produto, nome_produto, palletizacao
                    FROM cadastro_palletizacao
                    WHERE codigo_ean = :ean AND ativo = true
                    LIMIT 1
                """),
                {"ean": ean},
            ).fetchone()
    finally:
        sys.stdout = _old_stdout

    if not resultado:
        raise ValueError(f"Produto com EAN {ean} nao encontrado em cadastro_palletizacao")

    if not resultado.palletizacao or float(resultado.palletizacao) <= 0:
        raise ValueError(
            f"Produto {resultado.cod_produto} ({resultado.nome_produto}) "
            f"tem palletizacao={resultado.palletizacao}. Defina um valor valido."
        )

    return float(resultado.palletizacao), resultado.cod_produto, resultado.nome_produto


def _limpar_ean_saldo(valor):
    """Limpa formato EAN do portal: {="17898075642344"} → 17898075642344."""
    if not valor:
        return ""
    valor = valor.strip()
    match = re.match(r'\{="?([^"}\s]+)"?\}', valor)
    if match:
        return match.group(1).strip()
    match = re.match(r'="?([^"}\s]+)"?', valor)
    if match:
        return match.group(1).strip()
    return valor


def _limpar_ean_raw(valor):
    """Limpa formato EAN do CSV: {="17898075642344"} -> 17898075642344.

    Args:
        valor: String bruta do CSV

    Returns:
        str: EAN limpo (apenas digitos/texto)
    """
    if not valor:
        return ""
    valor = valor.strip()
    match = re.match(r'\{="?([^"}\s]+)"?\}', valor)
    if match:
        return match.group(1).strip()
    match = re.match(r'="?([^"}\s]+)"?', valor)
    if match:
        return match.group(1).strip()
    return valor


def _saldo_valido(valor):
    """Converte valor de saldo para float, retornando 0 em caso de erro.

    Args:
        valor: String do CSV (pode ter virgula como separador decimal)

    Returns:
        float: Valor numerico ou 0
    """
    try:
        return float((valor or "0").strip().replace(",", "."))
    except (ValueError, AttributeError):
        return 0


def construir_planilha(
    saldo_csv_path,
    data_agendamento,
    cod_veiculo,
    cnpj_cd=None,
    multiplicar=1,
    dividir_por=1,
    qtd_min_pallets=None,
    filtrar_ean=None,
    filtrar_filial=None,
    qtd_override=None,
):
    """Constroi planilha XLSX para upload a partir do CSV de saldo.

    Colunas da planilha de upload (9 colunas, A-I):
      A: Nr. da carga (int — sequencial por agendamento)
      B: CNPJ do CD (str — com formatacao XX.XXX.XXX/XXXX-XX)
      C: Filial (int — codigo loja)
      D: Nr. do pedido (int — pedido_cliente limpo)
      E: Codigo EAN (int com fmt='0' — preserva digitos)
      F: Quantidade a agendar (int)
      G: Data de (datetime com fmt='d-mmm')
      H: Data ate (datetime com fmt='d-mmm')
      I: Codigo veiculo (int)

    IMPORTANTE: O portal rejeita com HTTP 500 se tipos estiverem errados.
    EAN deve ser int (nao string), datas devem ser datetime (nao string DD/MM/YYYY).

    Args:
        saldo_csv_path: Caminho do CSV de saldo (output de consultar_saldo.py)
        data_agendamento: Data(s) para agendamento — string ou lista de strings DD/MM/YYYY
        cod_veiculo: Codigo do veiculo para planilha
        cnpj_cd: CNPJ do CD (default: 61.724.241/0003-30)
        multiplicar: Repetir cada agendamento N vezes (default: 1)
        dividir_por: Dividir quantidade por N (default: 1)
        qtd_min_pallets: Quantidade minima de pallets por agendamento (default: None)
        filtrar_ean: Filtrar por EAN especifico (default: None = todos)
        filtrar_filial: Filtrar por codigo de filial (default: None = todas)
        qtd_override: Quantidade fixa por item, override do saldo (default: None = usar saldo)

    Returns:
        tuple: (caminho_planilha, total_linhas, detalhes)
    """
    import openpyxl

    cnpj_cd = cnpj_cd or CNPJ_CD_PADRAO
    cnpj_cd_formatado = _formatar_cnpj(cnpj_cd)

    # Ler CSV de saldo
    csv_path = Path(saldo_csv_path)
    if not csv_path.exists():
        raise FileNotFoundError(f"CSV de saldo nao encontrado: {saldo_csv_path}")

    # Parsear CSV (mesmo mecanismo de consultar_saldo)
    registros = []
    for encoding in ["utf-8", "utf-8-sig", "iso-8859-1", "windows-1252"]:
        try:
            with open(csv_path, "r", encoding=encoding) as f:
                primeira_linha = f.readline()
                f.seek(0)
                delimitador = ";" if primeira_linha.count(";") > primeira_linha.count(",") else ","
                reader = csv.reader(f, delimiter=delimitador)
                headers = next(reader)  # pular header

                for row in reader:
                    if len(row) >= 11:
                        registros.append(row)
                break
        except (UnicodeDecodeError, csv.Error):
            continue

    if not registros:
        raise ValueError(f"CSV de saldo vazio ou ilegivel: {saldo_csv_path}")

    # GOTCHA: Remover pedidos fantasmas do Atacadao (saldo=1 por item).
    # O Atacadao gera pedidos com qtd=1 que nao representam saldo real.
    total_antes = len(registros)
    registros = [
        row for row in registros
        if _saldo_valido(row[7]) >= 2  # coluna H = saldo_disponivel
    ]
    if len(registros) < total_antes:
        logger.info(
            f"Filtro fantasma (saldo<2): {total_antes} -> {len(registros)} linhas "
            f"({total_antes - len(registros)} fantasmas removidos)"
        )

    if not registros:
        raise ValueError("Nenhum registro com saldo valido (>=2) apos remover fantasmas")

    # GOTCHA: Remover registros com EAN invalido (codigos internos do portal).
    # EANs validos do Atacadao comecam com "17". Prefixos 000, 037, 37, 57 sao internos.
    PREFIXOS_EAN_INVALIDOS = ("000", "037", "37", "57")
    total_antes_ean = len(registros)
    registros = [
        row for row in registros
        if not _limpar_ean_raw(row[5]).startswith(PREFIXOS_EAN_INVALIDOS)  # coluna F = EAN
    ]
    ean_invalidos = total_antes_ean - len(registros)
    if ean_invalidos:
        logger.info(
            f"Filtro EAN invalido: {total_antes_ean} -> {len(registros)} linhas "
            f"({ean_invalidos} com EAN interno removidos)"
        )

    if not registros:
        raise ValueError("Nenhum registro com EAN valido apos filtros")

    # Filtrar por EAN especifico (se solicitado)
    if filtrar_ean:
        total_antes_filtro = len(registros)
        registros = [
            row for row in registros
            if _limpar_ean_saldo(row[5] or "") == filtrar_ean
        ]
        logger.info(
            f"Filtro EAN={filtrar_ean}: {total_antes_filtro} -> {len(registros)} linhas"
        )
        if not registros:
            raise ValueError(
                f"Nenhum registro encontrado para EAN {filtrar_ean} no CSV de saldo"
            )

    # Filtrar por filial especifica (se solicitado)
    if filtrar_filial:
        total_antes_filial = len(registros)
        filial_str = str(filtrar_filial).strip()
        registros = [
            row for row in registros
            if (row[2] or "").strip() == filial_str
        ]
        logger.info(
            f"Filtro filial={filial_str}: {total_antes_filial} -> {len(registros)} linhas"
        )
        if not registros:
            raise ValueError(
                f"Nenhum registro encontrado para filial {filtrar_filial} no CSV de saldo"
            )

    logger.info(f"CSV lido: {len(registros)} linhas de saldo valido")

    # Normalizar datas para lista
    # Aceita: string unica, string com virgulas, ou lista
    if isinstance(data_agendamento, str):
        datas = [d.strip() for d in data_agendamento.split(",") if d.strip()]
    else:
        datas = list(data_agendamento)

    # Criar planilha usando template salvo (se existir) ou do zero
    # Template tem 18 colunas (A-R) e sheet "Planilha1"
    template_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "template_agendamento.xlsx"
    )
    if os.path.exists(template_path):
        logger.info(f"Usando template salvo: {template_path}")
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        # Template ja tem headers na linha 1
    else:
        logger.info("Template nao encontrado, criando planilha do zero")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Planilha1"
        # Headers EXATOS do template oficial (18 colunas)
        ws.append([
            "Nº da carga",
            "CNPJ do fornecedor",
            "Unidade de entrega",
            "Nº pedido",
            "Código Atacadão/EAN",
            "Qtde produto",
            "Data mínima",
            "Data desejada",
            "Código tipo de veículo",
            "CNPJ transportadora",
            "Placa do veículo",
            "Placa da carroceria",
            "CPF motorista",
            "Nome do motorista",
            "Celular do motorista",
            "RNTT Motorista",
            "CPF Ajudante 1",
            "CPF Ajudante 2",
        ])

    nr_carga = 1
    total_linhas = 0
    linhas_ignoradas = 0

    for data in datas:
        for row in registros:
            # Extrair campos do CSV de saldo (posicoes A-K)
            filial = (row[2] or "").strip()
            pedido = _limpar_ean_saldo(row[3] or "")
            ean = _limpar_ean_saldo(row[5] or "")
            saldo_str = (row[7] or "").strip().replace(",", ".")

            try:
                saldo_disponivel = float(saldo_str) if saldo_str else 0
            except ValueError:
                saldo_disponivel = 0

            if saldo_disponivel <= 0:
                linhas_ignoradas += 1
                continue

            # Calcular quantidade a agendar
            if qtd_override is not None and qtd_override > 0:
                qtd = qtd_override
            else:
                qtd = saldo_disponivel
                if dividir_por > 1:
                    qtd = math.ceil(saldo_disponivel / dividir_por)

            if qtd <= 0:
                linhas_ignoradas += 1
                continue

            # Multiplicar agendamentos (N vezes com nr_carga sequencial)
            # TIPOS DE CELULA CRITICOS (portal rejeita com 500 se errado):
            #   int: nr_carga, filial, pedido, qtd, veiculo
            #   int + fmt='0': EAN (preserva todos os digitos)
            #   datetime + fmt='d-mmm': datas
            #   str: CNPJ
            filial_int = int(filial) if filial.isdigit() else filial
            pedido_int = int(pedido) if pedido.isdigit() else pedido
            data_dt = datetime.strptime(data, "%d/%m/%Y")

            for _ in range(multiplicar):
                row_num = ws.max_row + 1

                ws.cell(row=row_num, column=1, value=nr_carga)          # A: int
                ws.cell(row=row_num, column=2, value=cnpj_cd_formatado)  # B: str
                ws.cell(row=row_num, column=3, value=filial_int)         # C: int
                ws.cell(row=row_num, column=4, value=pedido_int)         # D: int

                # E: EAN como int com formato "0" (todos os digitos)
                ean_cell = ws.cell(row=row_num, column=5, value=int(ean))
                ean_cell.number_format = '0'

                ws.cell(row=row_num, column=6, value=int(qtd))           # F: int

                # G/H: Datas como datetime com formato "d-mmm"
                dmin_cell = ws.cell(row=row_num, column=7, value=data_dt)
                dmin_cell.number_format = 'd-mmm'

                ddes_cell = ws.cell(row=row_num, column=8, value=data_dt)
                ddes_cell.number_format = 'd-mmm'

                ws.cell(row=row_num, column=9, value=int(cod_veiculo))   # I: int

                total_linhas += 1

            nr_carga += 1

    # Salvar planilha
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = OUTPUT_DIR / f"agendamento_{timestamp}.xlsx"
    wb.save(str(xlsx_path))

    detalhes = {
        "registros_saldo": len(registros),
        "linhas_geradas": total_linhas,
        "linhas_ignoradas_saldo_zero": linhas_ignoradas,
        "multiplicador": multiplicar,
        "divisor": dividir_por,
        "veiculo": f"{cod_veiculo} ({VEICULOS_PLANILHA.get(str(cod_veiculo), '?')})",
        "datas": datas,
        "cnpj_cd": cnpj_cd_formatado,
    }
    if filtrar_ean:
        detalhes["filtro_ean"] = filtrar_ean
    if filtrar_filial:
        detalhes["filtro_filial"] = str(filtrar_filial)
    if qtd_override is not None:
        detalhes["qtd_override"] = qtd_override

    logger.info(
        f"Planilha gerada: {xlsx_path} ({total_linhas} linhas, "
        f"{linhas_ignoradas} ignoradas por saldo=0)"
    )

    return str(xlsx_path), total_linhas, detalhes


# ──────────────────────────────────────────────
# FASE 2: Upload e validacao no portal
# ──────────────────────────────────────────────

def _fazer_upload(page, planilha_path):
    """Faz upload da planilha no portal /cargas-planilha.

    Fluxo real do portal (descoberto via JS inline):
    1. Cancela mapeamento anterior (DELETE /cargas-planilha/delete)
    2. Recarrega pagina para obter formulario de upload
    3. Seta arquivo no input#arquivo
    4. jQuery faz POST /cargas-planilha/create com FormData
    5. No sucesso, redireciona para /cargas-planilha?u=TIMESTAMP (mapeamento)

    Args:
        page: Playwright Page (sync)
        planilha_path: Caminho do arquivo XLSX

    Returns:
        bool: True se upload bem-sucedido e mapeamento visivel
    """
    try:
        # 1. Cancelar mapeamento anterior via AJAX (como o JS faz)
        cancel_result = page.evaluate("""() => {
            return new Promise((resolve) => {
                $.ajax({
                    url: window.location.origin + "/cargas-planilha/delete",
                    method: "DELETE",
                    success: function() { resolve("ok"); },
                    error: function(xhr) { resolve("error:" + xhr.status); }
                });
            });
        }""")
        logger.info(f"Cancelar anterior: {cancel_result}")

        # 2. Recarregar para obter formulario limpo
        page.goto(
            f"{page.evaluate('() => window.location.origin')}/cargas-planilha",
            wait_until="domcontentloaded",
            timeout=30000,
        )
        page.wait_for_timeout(2000)

        # 3. Localizar input file
        input_file = page.locator('#arquivo')
        if input_file.count() == 0:
            input_file = page.locator(SELETORES_LOTE["input_file"])
        if input_file.count() == 0:
            logger.error("Input de arquivo nao encontrado")
            return False

        # 4. Set file via Playwright
        logger.info(f"Enviando arquivo: {planilha_path}")
        input_file.first.set_input_files(planilha_path)
        page.wait_for_timeout(500)

        # 5. Fazer upload via AJAX direto (como o jQuery #enviar handler faz)
        upload_result = page.evaluate("""() => {
            return new Promise((resolve) => {
                var dataForm = new FormData(document.getElementById('uploadForm'));
                let uid = "7237&u=" + (new Date()).toISOString();
                $.ajax({
                    url: window.location.origin + "/cargas-planilha/create?uid=" + uid,
                    method: "POST",
                    data: dataForm,
                    cache: false,
                    contentType: false,
                    processData: false,
                    success: function(json) {
                        resolve({status: 200, ok: true});
                    },
                    error: function(xhr) {
                        resolve({
                            status: xhr.status,
                            ok: false,
                            body: xhr.responseText ? xhr.responseText.substring(0, 500) : ''
                        });
                    }
                });
            });
        }""")

        if not upload_result.get("ok"):
            logger.error(f"Upload falhou: HTTP {upload_result.get('status')} — {upload_result.get('body', '')}")
            return False

        logger.info(f"Upload OK (HTTP {upload_result['status']})")

        # 6. Recarregar para mapeamento (como o JS faz com redirect)
        page.goto(
            f"{page.evaluate('() => window.location.origin')}/cargas-planilha?u={int(datetime.now().timestamp() * 1000)}",
            wait_until="domcontentloaded",
            timeout=30000,
        )
        page.wait_for_timeout(3000)

        # 7. Verificar se mapeamento apareceu
        selects = page.locator(".select_coluna")
        if selects.count() > 0:
            logger.info(f"Mapeamento de colunas visivel ({selects.count()} selects)")
            return True

        # Fallback: verificar pelo botao analisar
        btn_analisar = page.locator(SELETORES_LOTE["confirmar_mapeamento"])
        if btn_analisar.count() > 0:
            logger.info("Mapeamento de colunas visivel (botao analisar encontrado)")
            return True

        logger.error("Mapeamento de colunas nao apareceu apos upload")
        return False

    except Exception as e:
        logger.error(f"Erro no upload: {e}")
        return False


def _preencher_mapeamento_manual(page):
    """Preenche dropdowns de mapeamento caso auto-fill nao funcione.

    O portal tem 9 selects na ordem:
      Nº da carga, CNPJ do fornecedor, Unidade de entrega, Nº pedido,
      Código Atacadão/EAN, Qtde produto, Data mínima, Data desejada,
      Código tipo de veículo

    Nossa planilha tem 9 colunas na mesma ordem.
    Cada select tem options com text = nome da coluna do planilha.
    Selecionamos coluna N para select N (1-indexed).
    """
    try:
        selects = page.locator("select")
        total_selects = selects.count()
        if total_selects == 0:
            logger.warning("Nenhum select de mapeamento encontrado")
            return False

        logger.info(f"Verificando {total_selects} selects de mapeamento...")

        # Nomes esperados das colunas (mesmo que headers do XLSX)
        colunas_esperadas = [
            "Nº da carga",
            "CNPJ do fornecedor",
            "Unidade de entrega",
            "Nº pedido",
            "Código Atacadão/EAN",
            "Qtde produto",
            "Data mínima",
            "Data desejada",
            "Código tipo de veículo",
        ]

        preenchidos = 0
        for i in range(min(total_selects, len(colunas_esperadas))):
            sel = selects.nth(i)
            valor_atual = sel.input_value()

            if valor_atual and valor_atual.strip():
                # Ja preenchido (auto-fill funcionou)
                preenchidos += 1
                continue

            # Tentar selecionar pelo label da coluna
            nome_coluna = colunas_esperadas[i]
            try:
                sel.select_option(label=nome_coluna, timeout=3000)
                preenchidos += 1
                logger.info(f"  Select {i + 1}: selecionado '{nome_coluna}'")
            except Exception:
                # Fallback: selecionar por indice (coluna i+1, pois 0=vazio)
                try:
                    options = sel.locator("option")
                    n_options = options.count()
                    # Opcao 0 geralmente e "Selecione...", coluna 1 = option 1
                    if n_options > i + 1:
                        val = options.nth(i + 1).get_attribute("value")
                        if val:
                            sel.select_option(value=val, timeout=3000)
                            preenchidos += 1
                            logger.info(f"  Select {i + 1}: selecionado por indice (value={val})")
                        else:
                            logger.warning(f"  Select {i + 1}: option sem value")
                    else:
                        logger.warning(f"  Select {i + 1}: apenas {n_options} options, preciso de {i + 2}")
                except Exception as e2:
                    logger.warning(f"  Select {i + 1}: fallback tambem falhou: {e2}")

        logger.info(f"Mapeamento: {preenchidos}/{len(colunas_esperadas)} colunas preenchidas")
        return preenchidos == len(colunas_esperadas)

    except Exception as e:
        logger.error(f"Erro ao preencher mapeamento manual: {e}")
        return False


def _confirmar_mapeamento(page):
    """Confirma o mapeamento de colunas e inicia validacao.

    Fluxo real do portal (descoberto via JS inline):
    1. Verifica se dropdowns estao preenchidos (auto-fill pelo template)
    2. Se nao, preenche manualmente (_preencher_mapeamento_manual)
    3. jQuery serializa #analiseForm e faz PUT /cargas-planilha/analisa
    4. No sucesso, redireciona para /cargas-planilha?u=TIMESTAMP
    5. Pagina recarregada mostra tabela de validacao

    O botao [name=analisar] dispara AJAX, nao form submit. O resultado
    aparece apos o redirect automatico do JS.

    Args:
        page: Playwright Page (sync)

    Returns:
        bool: True se mapeamento confirmado e validacao visivel
    """
    MAX_RETRIES = 2

    for attempt in range(MAX_RETRIES + 1):
        try:
            btn_analisar = page.locator(SELETORES_LOTE["confirmar_mapeamento"])
            if btn_analisar.count() == 0:
                logger.error("Botao [name=analisar] nao encontrado")
                return False

            # Verificar se dropdowns estao preenchidos (auto-fill)
            selects = page.locator(".select_coluna")
            total_selects = selects.count()
            preenchidos = 0
            for i in range(total_selects):
                val = selects.nth(i).input_value()
                if val and val.strip():
                    preenchidos += 1

            if preenchidos < total_selects:
                logger.info(
                    f"Auto-fill incompleto ({preenchidos}/{total_selects}). "
                    f"Preenchendo mapeamento manualmente..."
                )
                _preencher_mapeamento_manual(page)
                page.wait_for_timeout(1000)
            else:
                logger.info(f"Auto-fill OK ({preenchidos}/{total_selects} selects)")

            # Fazer PUT via AJAX direto (como o jQuery faz) para capturar resposta
            logger.info("Executando PUT /cargas-planilha/analisa via AJAX...")
            analisa_result = page.evaluate("""() => {
                return new Promise((resolve) => {
                    var dataForm = $("#analiseForm").serialize();
                    $.ajax({
                        url: window.location.origin + "/cargas-planilha/analisa",
                        method: "PUT",
                        data: dataForm,
                        success: function(json, textStatus, xhr) {
                            resolve({status: xhr.status, ok: true});
                        },
                        error: function(xhr) {
                            resolve({
                                status: xhr.status,
                                ok: false,
                                body: xhr.responseText ? xhr.responseText.substring(0, 500) : ''
                            });
                        }
                    });
                });
            }""")

            if analisa_result.get("ok"):
                logger.info(f"PUT /analisa retornou {analisa_result['status']} OK")

                # Recarregar pagina para ver resultado (como o JS faz com redirect)
                page.goto(
                    f"{page.evaluate('() => window.location.origin')}/cargas-planilha?u={int(datetime.now().timestamp() * 1000)}",
                    wait_until="domcontentloaded",
                    timeout=30000,
                )
                page.wait_for_timeout(3000)

                # Verificar se tabela de validacao apareceu
                tabela = page.locator(SELETORES_LOTE["tabela_validacao"])
                if tabela.count() > 0:
                    linhas = page.locator("table.table tbody tr")
                    if linhas.count() > 0:
                        logger.info(f"Validacao visivel: {linhas.count()} linhas")
                        return True

                # Verificar se botao salvar apareceu
                btn_salvar = page.locator(SELETORES_LOTE["salvar"])
                if btn_salvar.count() > 0:
                    logger.info("Botao salvar visivel — validacao concluida")
                    return True

                logger.warning("PUT OK mas tabela de validacao nao apareceu apos reload")
                return True  # PUT foi aceito, resultado pode nao ter tabela vazia

            else:
                status = analisa_result.get("status", 0)
                body = analisa_result.get("body", "")
                logger.error(f"PUT /analisa falhou: HTTP {status} — {body}")

                # Fechar modal de erro se existir
                page.evaluate("""() => {
                    var btn = document.querySelector('#modal-alert button, .modal button');
                    if (btn) btn.click();
                }""")
                page.wait_for_timeout(1000)

                if attempt < MAX_RETRIES:
                    logger.info(f"Tentativa {attempt + 1}/{MAX_RETRIES + 1} falhou. Retentando...")
                    page.wait_for_timeout(3000)
                    continue

                return False

        except Exception as e:
            logger.error(f"Erro ao confirmar mapeamento (tentativa {attempt + 1}): {e}")
            if attempt < MAX_RETRIES:
                page.wait_for_timeout(3000)
                continue
            return False

    return False


def _extrair_validacao(page):
    """Extrai resultado da validacao do portal.

    Colunas da tabela de validacao (10 colunas, indices 0-9):
      [0] Status + critica (ex: "Inconsistente Quantidade solicitada...")
      [1] Nr. carga
      [2] Filial
      [3] Data minima
      [4] Data desejada
      [5] CNPJ fornecedor + razao social
      [6] Pedido cliente / pedido portal
      [7] Codigo Atacadao / EAN
      [8] Quantidade
      [9] Veiculo

    Critica de saldo tem formato:
      "Quantidade solicitada do item {EAN} ,({qtd_solicitada}) é maior do que o saldo ({saldo})"
    Quando saldo > 0 mas < qtd, e parcial. Quando saldo = 0, nao tem nada disponivel.

    Args:
        page: Playwright Page (sync)

    Returns:
        dict: {validas, inconsistentes, detalhes_inconsistencias}
              Cada detalhe tem: linha, motivo, dados{carga, filial, ean, qtd, pedido,
              data_min, data_des, veiculo}, saldo_portal (int ou None),
              qtd_solicitada (int ou None), saldo_parcial (bool)
    """
    try:
        resultado = page.evaluate(r"""() => {
            const linhas = document.querySelectorAll('table.table tbody tr');
            let validas = 0;
            let inconsistentes = 0;
            const detalhes = [];

            linhas.forEach((tr, idx) => {
                const tds = tr.querySelectorAll('td');
                if (tds.length === 0) return;

                // Checar inconsistencia por bg-danger ou texto
                let isInconsistente = false;
                tds.forEach(td => {
                    if (td.classList.contains('bg-danger')) isInconsistente = true;
                });
                const col0 = tds[0].textContent.trim().replace(/\s+/g, ' ');
                if (col0.toLowerCase().includes('inconsistente')) isInconsistente = true;

                // Extrair dados da linha (todas as colunas)
                const dados = {
                    carga: tds.length > 1 ? tds[1].textContent.trim() : '',
                    filial: tds.length > 2 ? tds[2].textContent.trim() : '',
                    data_min: tds.length > 3 ? tds[3].textContent.trim() : '',
                    data_des: tds.length > 4 ? tds[4].textContent.trim() : '',
                    fornecedor: tds.length > 5 ? tds[5].textContent.trim() : '',
                    pedido: tds.length > 6 ? tds[6].textContent.trim() : '',
                    ean: tds.length > 7 ? tds[7].textContent.trim() : '',
                    qtd: tds.length > 8 ? tds[8].textContent.trim() : '',
                    veiculo: tds.length > 9 ? tds[9].textContent.trim() : '',
                };

                if (isInconsistente) {
                    inconsistentes++;

                    // Parsear saldo da critica:
                    // "Quantidade solicitada do item {EAN} ,({qtd}) é maior do que o saldo ({saldo})"
                    let saldo_portal = null;
                    let qtd_solicitada = null;
                    let saldo_parcial = false;

                    const matchSaldo = col0.match(/saldo\s*\((\d+)\)/i);
                    if (matchSaldo) saldo_portal = parseInt(matchSaldo[1]);

                    const matchQtd = col0.match(/,\s*\((\d+)\)\s*é\s*maior/i);
                    if (matchQtd) qtd_solicitada = parseInt(matchQtd[1]);

                    if (saldo_portal !== null && saldo_portal > 0) saldo_parcial = true;

                    detalhes.push({
                        linha: idx + 1,
                        motivo: col0,
                        dados: dados,
                        saldo_portal: saldo_portal,
                        qtd_solicitada: qtd_solicitada,
                        saldo_parcial: saldo_parcial
                    });
                } else {
                    validas++;
                }
            });

            return { validas, inconsistentes, detalhes };
        }""")

        return {
            "validas": resultado.get("validas", 0),
            "inconsistentes": resultado.get("inconsistentes", 0),
            "detalhes_inconsistencias": resultado.get("detalhes", []),
        }

    except Exception as e:
        logger.error(f"Erro ao extrair validacao: {e}")
        return {
            "validas": 0,
            "inconsistentes": 0,
            "detalhes_inconsistencias": [],
            "erro": str(e),
        }


# ──────────────────────────────────────────────
# FASE 3: Salvar
# ──────────────────────────────────────────────

def _salvar_agendamento(page):
    """Salva o agendamento no portal.

    Clica #salvar2, aguarda modal de sucesso, captura link de cargas.

    Args:
        page: Playwright Page (sync)

    Returns:
        dict: {status, link_cargas, controle} ou {status, erro}
    """
    try:
        btn_salvar = page.locator(SELETORES_LOTE["salvar"])
        if btn_salvar.count() == 0:
            return {"status": "erro", "erro": "Botao #salvar2 nao encontrado"}

        logger.info("Salvando agendamento...")
        btn_salvar.click()

        # Aguardar modal de sucesso (pode demorar)
        for _ in range(24):  # 2 min max
            # Verificar modal de sucesso
            modal = page.locator(SELETORES_LOTE["modal_sucesso"])
            if modal.count() > 0 and modal.first.is_visible():
                logger.info("Modal de sucesso detectado!")

                # Capturar link de listar cargas
                link_cargas = None
                controle = None
                try:
                    link_el = page.locator('a.btn.btn-primary[href*="cargas"]')
                    if link_el.count() > 0:
                        href = link_el.first.get_attribute("href")
                        if href:
                            link_cargas = href if href.startswith("http") else f"https://atacadao.hodiebooking.com.br{href}"
                            # Extrair controle UUID da URL
                            match = re.search(r'controle=([a-f0-9]+)', href)
                            if match:
                                controle = match.group(1)
                except Exception:
                    pass

                # Clicar OK para fechar modal
                try:
                    btn_ok = page.locator(SELETORES_LOTE["botao_ok"])
                    if btn_ok.count() > 0:
                        btn_ok.first.click()
                        page.wait_for_timeout(1000)
                except Exception:
                    pass

                return {
                    "status": "agendado",
                    "link_cargas": link_cargas,
                    "controle": controle,
                }

            # Verificar se apareceu erro
            alertas = page.locator(".alert-danger")
            if alertas.count() > 0:
                texto = alertas.first.text_content().strip()
                return {"status": "erro", "erro": f"Portal retornou erro: {texto}"}

            page.wait_for_timeout(5000)

        return {"status": "erro", "erro": "Timeout aguardando confirmacao (2 min)"}

    except Exception as e:
        logger.error(f"Erro ao salvar: {e}")
        return {"status": "erro", "erro": str(e)}


# ──────────────────────────────────────────────
# Orquestrador principal
# ──────────────────────────────────────────────

def agendar_lote(
    saldo_csv=None,
    planilha=None,
    data=None,
    veiculo=None,
    cnpj_cd=None,
    multiplicar=1,
    dividir_por=1,
    qtd_min_pallets=None,
    dry_run=False,
    confirmar=False,
    filtrar_ean=None,
    filtrar_filial=None,
    qtd_override=None,
):
    """Orquestra agendamento em lote no portal Atacadao.

    Args:
        saldo_csv: Caminho do CSV de saldo (modo 1)
        planilha: Caminho da planilha preparada (modo 2)
        data: Data(s) de agendamento DD/MM/YYYY — string ou lista (obrigatorio no modo 1)
        veiculo: Codigo do veiculo para planilha (obrigatorio no modo 1)
        cnpj_cd: CNPJ do CD (default: 61.724.241/0003-30)
        multiplicar: Repetir agendamentos N vezes (default: 1)
        dividir_por: Dividir quantidade por N (default: 1)
        qtd_min_pallets: Minimo de pallets por agendamento
        dry_run: Se True, validar sem salvar
        confirmar: Se True, salvar apos validacao
        filtrar_ean: Filtrar CSV por EAN especifico
        filtrar_filial: Filtrar CSV por codigo de filial
        qtd_override: Quantidade fixa por item (override do saldo)

    Returns:
        dict: Resultado JSON padronizado
    """
    planilha_path = None
    detalhes_construcao = None

    # ─── FASE 1: Obter/Construir planilha ───

    if planilha:
        # Modo 2: planilha pronta
        planilha_path = planilha
        if not Path(planilha_path).exists():
            return gerar_saida(False, erro=f"Planilha nao encontrada: {planilha_path}")
        logger.info(f"Usando planilha existente: {planilha_path}")

    elif saldo_csv:
        # Modo 1: construir a partir do CSV de saldo
        if not data:
            return gerar_saida(False, erro="--data e obrigatorio no modo --saldo-csv")
        if not veiculo:
            return gerar_saida(False, erro="--veiculo e obrigatorio no modo --saldo-csv")

        try:
            datas = _validar_data(data)
            veiculo = _validar_veiculo(veiculo)
        except ValueError as e:
            return gerar_saida(False, erro=str(e))

        try:
            planilha_path, total_linhas, detalhes_construcao = construir_planilha(
                saldo_csv_path=saldo_csv,
                data_agendamento=datas,
                cod_veiculo=veiculo,
                cnpj_cd=cnpj_cd,
                multiplicar=multiplicar,
                dividir_por=dividir_por,
                qtd_min_pallets=qtd_min_pallets,
                filtrar_ean=filtrar_ean,
                filtrar_filial=filtrar_filial,
                qtd_override=qtd_override,
            )
        except (FileNotFoundError, ValueError) as e:
            return gerar_saida(False, erro=str(e))

        if total_linhas == 0:
            return gerar_saida(
                False,
                erro="Nenhuma linha gerada (todo saldo = 0)",
                planilha_path=planilha_path,
                detalhes=detalhes_construcao,
            )

        logger.info(f"Planilha construida: {planilha_path} ({total_linhas} linhas)")

    else:
        return gerar_saida(
            False,
            erro="Informe --saldo-csv (com --data e --veiculo) ou --planilha",
        )

    # Se nao tem --dry-run nem --confirmar, apenas mostra preview da planilha
    if not dry_run and not confirmar:
        resultado = {
            "modo": "preview",
            "planilha_path": planilha_path,
        }
        if detalhes_construcao:
            resultado["detalhes"] = detalhes_construcao
        resultado["instrucao"] = (
            f"Planilha gerada em {planilha_path}. "
            "Use --dry-run para validar no portal ou --confirmar para agendar."
        )
        return gerar_saida(True, **resultado)

    # ─── FASE 2: Upload e validacao no portal ───

    pw = None
    browser = None
    try:
        pw, browser, _context, page = criar_sessao_download(headless=True)

        # Verificar sessao
        if not verificar_sessao_sync(page):
            return gerar_saida(
                False,
                erro="Sessao do Atacadao expirada",
                requer_login=True,
                planilha_path=planilha_path,
                instrucao="Faca re-login interativo: python -m app.portal.atacadao.login_interativo",
            )

        # Navegar para /cargas-planilha
        logger.info(f"Navegando para {URL_CARGAS_PLANILHA}...")
        page.goto(
            URL_CARGAS_PLANILHA,
            wait_until="domcontentloaded",
            timeout=TIMEOUTS["page_load"],
        )
        page.wait_for_timeout(2000)

        # Verificar se chegou na pagina correta
        url_atual = page.url.lower()
        if "login" in url_atual or "signin" in url_atual:
            return gerar_saida(
                False,
                erro="Redirecionado para login ao acessar /cargas-planilha",
                requer_login=True,
                planilha_path=planilha_path,
                instrucao="Faca re-login interativo: python -m app.portal.atacadao.login_interativo",
            )

        # Screenshot antes do upload
        screenshot_antes = capturar_screenshot(page, "cargas_planilha_antes")

        # Upload da planilha
        if not _fazer_upload(page, planilha_path):
            screenshot_erro = capturar_screenshot(page, "erro_upload")
            return gerar_saida(
                False,
                erro="Falha no upload da planilha",
                planilha_path=planilha_path,
                screenshot=screenshot_erro,
            )

        # Screenshot apos upload
        capturar_screenshot(page, "cargas_planilha_apos_upload")

        # Confirmar mapeamento de colunas
        if not _confirmar_mapeamento(page):
            screenshot_erro = capturar_screenshot(page, "erro_mapeamento")
            return gerar_saida(
                False,
                erro="Falha ao confirmar mapeamento de colunas",
                planilha_path=planilha_path,
                screenshot=screenshot_erro,
            )

        # Extrair resultado da validacao
        page.wait_for_timeout(3000)
        validacao = _extrair_validacao(page)

        # Screenshot da validacao
        screenshot_validacao = capturar_screenshot(page, "cargas_planilha_validacao")

        # ─── Decisao: dry-run ou confirmar ───

        if dry_run:
            resultado = {
                "modo": "dry-run",
                "planilha_path": planilha_path,
                "validacao": validacao,
                "screenshot": screenshot_validacao,
            }
            if detalhes_construcao:
                resultado["detalhes"] = detalhes_construcao
            if validacao.get("inconsistentes", 0) > 0:
                resultado["aviso"] = (
                    f"{validacao['inconsistentes']} linha(s) inconsistente(s). "
                    "Revise antes de confirmar."
                )
            else:
                resultado["instrucao"] = (
                    f"Validacao OK ({validacao.get('validas', 0)} linhas validas). "
                    f"Para confirmar: python agendar_lote.py --planilha {planilha_path} --confirmar"
                )
            return gerar_saida(True, **resultado)

        # Confirmar — verificar se ha inconsistencias
        if confirmar:
            if validacao.get("inconsistentes", 0) > 0:
                return gerar_saida(
                    False,
                    erro=(
                        f"{validacao['inconsistentes']} linha(s) inconsistente(s). "
                        "Nao e possivel salvar com inconsistencias."
                    ),
                    modo="confirmar_bloqueado",
                    planilha_path=planilha_path,
                    validacao=validacao,
                    screenshot=screenshot_validacao,
                )

            # ─── FASE 3: Salvar ───

            resultado_salvar = _salvar_agendamento(page)

            screenshot_final = capturar_screenshot(page, "cargas_planilha_final")

            resultado = {
                "modo": "confirmado",
                "planilha_path": planilha_path,
                "total_linhas": validacao.get("validas", 0),
                "validacao": validacao,
                "resultado": resultado_salvar,
                "screenshot": screenshot_final,
            }
            if detalhes_construcao:
                resultado["detalhes"] = detalhes_construcao

            sucesso = resultado_salvar.get("status") == "agendado"
            return gerar_saida(sucesso, **resultado)

    except RuntimeError as e:
        return gerar_saida(
            False,
            erro=str(e),
            planilha_path=planilha_path,
            requer_login="sessao" in str(e).lower() or "storage" in str(e).lower(),
            instrucao="Faca re-login interativo: python -m app.portal.atacadao.login_interativo",
        )

    except Exception as e:
        logger.error(f"Erro inesperado: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return gerar_saida(
            False,
            erro=f"Erro inesperado: {str(e)}",
            planilha_path=planilha_path,
        )

    finally:
        if browser:
            try:
                browser.close()
            except Exception:
                pass
        if pw:
            try:
                pw.stop()
            except Exception:
                pass


# ──────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Agendamento em lote no portal Atacadao via upload de planilha"
    )

    # Modo 1: A partir do CSV de saldo
    grupo_saldo = parser.add_argument_group("Modo 1: A partir do CSV de saldo")
    grupo_saldo.add_argument(
        "--saldo-csv",
        help="Caminho do CSV de saldo (output de consultar_saldo.py)",
    )
    grupo_saldo.add_argument(
        "--data",
        help="Data(s) de agendamento DD/MM/YYYY. Multiplas separadas por virgula: 24/03/2026,25/03/2026",
    )
    grupo_saldo.add_argument(
        "--veiculo",
        help="Codigo do veiculo para planilha (ex: 7=Carreta-Bau). Obrigatorio com --saldo-csv",
    )

    # Modo 2: Planilha pronta
    grupo_planilha = parser.add_argument_group("Modo 2: Planilha ja preparada")
    grupo_planilha.add_argument(
        "--planilha",
        help="Caminho da planilha XLSX preparada para upload",
    )

    # Filtros de produto/filial
    grupo_filtros = parser.add_argument_group("Filtros de produto e filial")
    grupo_filtros.add_argument(
        "--ean",
        default=None,
        help="Filtrar saldo por EAN especifico (ex: 17898075641163)",
    )
    grupo_filtros.add_argument(
        "--filial",
        default=None,
        help="Filtrar saldo por codigo de filial (ex: 183)",
    )

    # Quantidade
    grupo_qtd = parser.add_argument_group("Quantidade (override do saldo)")
    grupo_qtd.add_argument(
        "--qtd",
        type=int,
        default=None,
        help="Quantidade fixa por item em UNIDADES (override do saldo)",
    )
    grupo_qtd.add_argument(
        "--pallets",
        type=int,
        default=None,
        help="Quantidade em PALLETS (requer --ean para lookup de palletizacao no cadastro local)",
    )

    # Parametros de negocio
    grupo_negocio = parser.add_argument_group("Parametros de negocio")
    grupo_negocio.add_argument(
        "--cnpj-cd",
        default=None,
        help=f"CNPJ do CD para coluna B (default: {_formatar_cnpj(CNPJ_CD_PADRAO)})",
    )
    grupo_negocio.add_argument(
        "--multiplicar",
        type=int,
        default=1,
        help="Repetir cada agendamento N vezes (default: 1)",
    )
    grupo_negocio.add_argument(
        "--dividir-por",
        type=int,
        default=1,
        help="Dividir quantidade disponivel por N (default: 1)",
    )
    grupo_negocio.add_argument(
        "--qtd-min-pallets",
        type=int,
        default=None,
        help="Quantidade minima de pallets por agendamento",
    )

    # Modo de execucao
    grupo_exec = parser.add_argument_group("Modo de execucao")
    grupo_exec.add_argument(
        "--dry-run",
        action="store_true",
        help="Validar no portal sem salvar (RECOMENDADO na primeira execucao)",
    )
    grupo_exec.add_argument(
        "--confirmar",
        action="store_true",
        help="Salvar agendamento apos validacao (somente se sem inconsistencias)",
    )

    args = parser.parse_args()

    # Validacoes de argumentos
    if args.dry_run and args.confirmar:
        parser.error("--dry-run e --confirmar sao mutuamente exclusivos")

    if args.saldo_csv and args.planilha:
        parser.error("--saldo-csv e --planilha sao mutuamente exclusivos")

    if not args.saldo_csv and not args.planilha:
        parser.error("Informe --saldo-csv (com --data e --veiculo) ou --planilha")

    if args.qtd and args.pallets:
        parser.error("--qtd e --pallets sao mutuamente exclusivos")

    if args.pallets and not args.ean:
        parser.error("--pallets requer --ean para lookup de palletizacao no cadastro local")

    # Limpar CNPJ do CD
    cnpj_cd = None
    if args.cnpj_cd:
        cnpj_cd = _limpar_cnpj(args.cnpj_cd)
        if len(cnpj_cd) != 14:
            parser.error(f"CNPJ do CD invalido: {args.cnpj_cd}")

    # Resolver quantidade: --pallets converte para unidades via palletizacao do cadastro
    qtd_final = None
    if args.pallets:
        try:
            palletizacao, cod_produto, nome_produto = _resolver_palletizacao(args.ean)
            qtd_final = int(args.pallets * palletizacao)
            logger.info(
                f"Convertendo {args.pallets} pallets x {palletizacao} un/pallet = "
                f"{qtd_final} unidades ({cod_produto} - {nome_produto})"
            )
        except ValueError as e:
            parser.error(str(e))
    elif args.qtd:
        qtd_final = args.qtd

    resultado = agendar_lote(
        saldo_csv=args.saldo_csv,
        planilha=args.planilha,
        data=args.data,
        veiculo=args.veiculo,
        cnpj_cd=cnpj_cd,
        multiplicar=args.multiplicar,
        dividir_por=args.dividir_por,
        qtd_min_pallets=args.qtd_min_pallets,
        dry_run=args.dry_run,
        confirmar=args.confirmar,
        filtrar_ean=args.ean,
        filtrar_filial=args.filial,
        qtd_override=qtd_final,
    )

    if not resultado.get("sucesso"):
        sys.exit(1)


if __name__ == "__main__":
    main()
