#!/usr/bin/env python3
"""Gera a planilha de CONTROLE DE TITULOS A RECEBER (recebiveis) a partir do
sistema (tabela `contas_a_receber`, dados do Odoo enriquecidos) — para download
pelo Agente Web.

Reproduz deterministicamente o controle que era feito a mao (conciliacao Grafeno):
- Aba "Titulos": todos os titulos do escopo, com Situacao derivada
  (CONFIRMADO = pago / Vencido = venceu e nao pago / Em Aberto = a vencer).
- Aba "Vencidos": apenas os vencidos, AGRUPADOS por GESTOR de carteira
  (equipe_vendas do Odoo), com subtotal por grupo e total geral.

FONTE 100% determinística do banco — NAO recebe upload do usuario:
- titulo/parcela/cliente/vencimento/valor/parcela_paga -> contas_a_receber
- vendedor -> entregas_monitoradas (via entrega_monitorada_id)
- gestor   -> equipe_vendas do faturamento mais recente do mesmo CNPJ

LIMITE CONHECIDO (MVP): a marcacao CONFIRMADO usa `parcela_paga` do Odoo. Pagamentos
recebidos no extrato Grafeno mas ainda NAO baixados no Odoo aparecerao como Vencido/
Em Aberto (gap de conciliacao Grafeno, parado desde jan/2026). Esse delta e esperado.

Uso (CLI):
  python gerar_controle_recebiveis.py [--cliente "BORGES" --cliente 42385121]
        [--gestor "DENISE"] [--empresa 3] [--apenas-vencidos]
        [--desde 2025-06-01] [--session-id <sid>] [--nome controle_recebiveis]

Saida: JSON em stdout {sucesso, arquivo, url, resumo}. Arquivo em /tmp/agente_files/<sid>/.
"""
import argparse
import json
import os
import re
import uuid
from datetime import date, datetime, timedelta


# ---------------------------------------------------------------------------
# Diretorio de download (espelha exportando-arquivos / routes/_constants.py).
# NAO usar tempfile.gettempdir(): o subprocesso Bash do CLI usa TMPDIR=/tmp/claude-{uid},
# mas o gunicorn que SERVE o download le de /tmp/agente_files. (bug #787)
# ---------------------------------------------------------------------------
def _safe_sid(session_id: str) -> str:
    """Sanitiza session_id (anti path-traversal). fullmatch p/ rejeitar newline final."""
    sid = session_id or ''
    return sid if re.fullmatch(r'[A-Za-z0-9\-_]{1,64}', sid) else 'default'


def get_upload_folder(session_id: str) -> str:
    base_folder = os.path.join(os.environ.get('AGENTE_FILES_ROOT', '/tmp'), 'agente_files')
    folder = os.path.join(base_folder, _safe_sid(session_id))
    os.makedirs(folder, exist_ok=True)
    return folder


# ---------------------------------------------------------------------------
# Query — contas_a_receber + gestor(equipe_vendas) por CNPJ + vendedor.
# ---------------------------------------------------------------------------
SQL = """
WITH gestor_map AS (
    SELECT DISTINCT ON (regexp_replace(cnpj_cliente, '\\D', '', 'g'))
           regexp_replace(cnpj_cliente, '\\D', '', 'g') AS cnpj_num,
           equipe_vendas
    FROM faturamento_produto
    WHERE coalesce(equipe_vendas, '') <> ''
    ORDER BY regexp_replace(cnpj_cliente, '\\D', '', 'g'), data_fatura DESC NULLS LAST
)
SELECT c.titulo_nf, c.parcela, c.empresa, c.cnpj, c.raz_social, c.uf_cliente,
       c.emissao, c.vencimento, c.valor_original, c.valor_residual, c.parcela_paga,
       em.vendedor,
       gm.equipe_vendas AS gestor,
       CASE WHEN coalesce(c.parcela_paga, false) THEN 'CONFIRMADO'
            WHEN c.vencimento < CURRENT_DATE THEN 'Vencido'
            ELSE 'Em Aberto' END AS situacao
FROM contas_a_receber c
LEFT JOIN entregas_monitoradas em ON em.id = c.entrega_monitorada_id
LEFT JOIN gestor_map gm ON gm.cnpj_num = regexp_replace(c.cnpj, '\\D', '', 'g')
WHERE 1 = 1
  {f_desde}
  {f_empresa}
  {f_cliente}
  {f_gestor}
  {f_vencidos}
ORDER BY gm.equipe_vendas NULLS LAST, c.raz_social, c.vencimento, c.titulo_nf, c.parcela
"""


def buscar_titulos(filtros: dict) -> list:
    """Executa a query no banco da aplicacao (PROD no Agente Web)."""
    from sqlalchemy import text
    from app import create_app, db

    f_desde = "AND c.vencimento >= :desde" if filtros.get('desde') else ""
    f_empresa = "AND c.empresa = :empresa" if filtros.get('empresa') else ""
    f_vencidos = ("AND NOT coalesce(c.parcela_paga, false) AND c.vencimento < CURRENT_DATE"
                  if filtros.get('apenas_vencidos') else "")

    params = {}
    if filtros.get('desde'):
        params['desde'] = filtros['desde']
    if filtros.get('empresa'):
        params['empresa'] = filtros['empresa']

    # cliente: lista de termos (CNPJ digits OU nome) -> OR de ilike/cnpj
    f_cliente = ""
    cli = filtros.get('cliente') or []
    if cli:
        ors = []
        for i, termo in enumerate(cli):
            digits = re.sub(r'\D', '', termo)
            if len(digits) >= 6:  # parece CNPJ/raiz
                ors.append(f"regexp_replace(c.cnpj, '\\D', '', 'g') LIKE :cli{i}")
                params[f'cli{i}'] = digits + '%'
            else:
                ors.append(f"c.raz_social ILIKE :cli{i}")
                params[f'cli{i}'] = f'%{termo}%'
        f_cliente = "AND (" + " OR ".join(ors) + ")"

    f_gestor = ""
    if filtros.get('gestor'):
        f_gestor = "AND gm.equipe_vendas ILIKE :gestor"
        params['gestor'] = f"%{filtros['gestor']}%"

    sql = SQL.format(f_desde=f_desde, f_empresa=f_empresa, f_cliente=f_cliente,
                     f_gestor=f_gestor, f_vencidos=f_vencidos)

    app = create_app()
    with app.app_context():
        rows = db.session.execute(text(sql), params).mappings().all()
    return [dict(r) for r in rows]


# ---------------------------------------------------------------------------
# Gerador de Excel — PURO (recebe list[dict]), testavel sem banco.
# ---------------------------------------------------------------------------
def _to_date(v):
    if v is None or v == '':
        return None
    if isinstance(v, datetime):   # checar datetime ANTES de date (datetime e' subclasse de date)
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v)[:10]
    try:
        y, m, d = s.split('-')
        return date(int(y), int(m), int(d))
    except Exception:
        return None


def _grupo(t: dict) -> str:
    g = (t.get('gestor') or '').strip()
    if g:
        return g
    v = (t.get('vendedor') or '').strip()
    return v or '(sem gestor)'


def gerar_excel(titulos: list, filepath: str) -> dict:
    """Gera o .xlsx com abas Titulos + Vencidos(por gestor). Retorna resumo."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    VENC_FILL = PatternFill("solid", fgColor="FCE4E4")
    SUB_FILL = PatternFill("solid", fgColor="DDEBF7")
    BOLD = Font(bold=True)
    GRP_FONT = Font(bold=True, color="1F4E78")
    DATE_FMT = "dd/mm/yyyy"
    CURR_FMT = 'R$ #,##0.00'
    thin = Side(style="thin", color="D0D0D0")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center")

    cols = ["NF", "Parc.", "Cliente", "CNPJ", "UF", "Emissao", "Vencimento",
            "Vl Original", "Vl Residual", "Situacao", "Vendedor", "Gestor"]
    width = {"NF": 11, "Parc.": 7, "Cliente": 42, "CNPJ": 20, "UF": 6, "Emissao": 12,
             "Vencimento": 12, "Vl Original": 14, "Vl Residual": 14, "Situacao": 12,
             "Vendedor": 34, "Gestor": 22}

    def row_of(t):
        return {
            "NF": t.get("titulo_nf"), "Parc.": t.get("parcela") or "",
            "Cliente": t.get("raz_social") or "", "CNPJ": t.get("cnpj") or "",
            "UF": t.get("uf_cliente") or "", "Emissao": _to_date(t.get("emissao")),
            "Vencimento": _to_date(t.get("vencimento")),
            "Vl Original": float(t.get("valor_original") or 0),
            "Vl Residual": float(t.get("valor_residual") or 0),
            "Situacao": t.get("situacao") or "", "Vendedor": t.get("vendedor") or "",
            "Gestor": t.get("gestor") or "",
        }

    rows = [row_of(t) for t in titulos]
    # drop colunas totalmente vazias (UF, Gestor) — espelha o pedido da Martha
    drop = {c for c in ("UF", "Gestor") if all(not r[c] for r in rows)}
    use_cols = [c for c in cols if c not in drop]

    wb = Workbook()

    def write_header(ws, headers):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = BORDER
        for i, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = width.get(h, 14)
        ws.freeze_panes = "A2"

    def style_cell(cell, h, destaque=False):
        if h in ("Emissao", "Vencimento"):
            cell.number_format = DATE_FMT
        elif h in ("Vl Original", "Vl Residual"):
            cell.number_format = CURR_FMT
        cell.border = BORDER
        if destaque:
            cell.fill = VENC_FILL

    # ===== Aba Titulos =====
    ws = wb.active
    ws.title = "Titulos"
    write_header(ws, use_cols)
    for r in rows:
        rid = ws.max_row + 1
        venc = r["Situacao"] == "Vencido"
        for i, h in enumerate(use_cols, 1):
            cell = ws.cell(row=rid, column=i, value=r[h])
            style_cell(cell, h, destaque=venc)

    # ===== Aba Vencidos (por gestor) =====
    venc_cols = [c for c in use_cols if c != "Situacao"]
    wsv = wb.create_sheet("Vencidos")
    write_header(wsv, venc_cols)
    by_grp = {}
    for t, r in zip(titulos, rows):
        if r["Situacao"] != "Vencido":
            continue
        by_grp.setdefault(_grupo(t), []).append(r)

    g_orig = g_resid = 0.0
    for grp in sorted(by_grp):
        sub = by_grp[grp]
        hr = wsv.max_row + 1
        gc = wsv.cell(row=hr, column=1, value=f"Gestor: {grp}")
        gc.font = GRP_FONT
        wsv.merge_cells(start_row=hr, start_column=1, end_row=hr, end_column=len(venc_cols))
        s_orig = s_resid = 0.0
        for r in sub:
            rid = wsv.max_row + 1
            for i, h in enumerate(venc_cols, 1):
                style_cell(wsv.cell(row=rid, column=i, value=r[h]), h, destaque=True)
            s_orig += r["Vl Original"]; s_resid += r["Vl Residual"]
        # subtotal
        sr = wsv.max_row + 1
        wsv.cell(row=sr, column=1, value=f"Subtotal {grp}").font = BOLD
        for i, h in enumerate(venc_cols, 1):
            cell = wsv.cell(row=sr, column=i)
            cell.fill = SUB_FILL; cell.font = BOLD
            if h == "Vl Original":
                cell.value = s_orig; cell.number_format = CURR_FMT
            elif h == "Vl Residual":
                cell.value = s_resid; cell.number_format = CURR_FMT
        wsv.cell(row=wsv.max_row + 1, column=1, value="")  # separador
        g_orig += s_orig; g_resid += s_resid

    tr = wsv.max_row + 1
    wsv.cell(row=tr, column=1, value="TOTAL GERAL VENCIDOS")
    for i, h in enumerate(venc_cols, 1):
        cell = wsv.cell(row=tr, column=i)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        if h == "Vl Original":
            cell.value = g_orig; cell.number_format = CURR_FMT
        elif h == "Vl Residual":
            cell.value = g_resid; cell.number_format = CURR_FMT

    wb.save(filepath)

    from collections import Counter
    sit = Counter(r["Situacao"] for r in rows)
    return {
        "total": len(rows),
        "confirmado": sit.get("CONFIRMADO", 0),
        "vencido": sit.get("Vencido", 0),
        "em_aberto": sit.get("Em Aberto", 0),
        "gestores": sorted(by_grp.keys()),
        "valor_vencido": round(g_resid, 2),
    }


def main() -> int:
    ap = argparse.ArgumentParser(description="Gera planilha de controle de titulos a receber (recebiveis).")
    ap.add_argument("--cliente", action="append", default=[], help="CNPJ/raiz ou nome (repetivel)")
    ap.add_argument("--gestor", help="filtra por gestor/equipe (ILIKE), ex: DENISE")
    ap.add_argument("--empresa", type=int, help="1=FB, 2=SC, 3=CD")
    ap.add_argument("--apenas-vencidos", action="store_true", help="so titulos vencidos e nao pagos")
    ap.add_argument("--desde", help="vencimento >= YYYY-MM-DD (default: ultimos 12 meses)")
    ap.add_argument("--session-id", default="default", help="session do agente (p/ pasta de download)")
    ap.add_argument("--nome", default="controle_recebiveis", help="nome base do arquivo")
    a = ap.parse_args()

    desde = a.desde
    if not desde and not a.cliente and not a.gestor:
        # default seguro: evita varrer a tabela inteira
        desde = (date.today() - timedelta(days=365)).isoformat()

    filtros = {
        "cliente": a.cliente, "gestor": a.gestor, "empresa": a.empresa,
        "apenas_vencidos": a.apenas_vencidos, "desde": desde,
    }

    try:
        titulos = buscar_titulos(filtros)
    except Exception as e:  # noqa: BLE001
        print(json.dumps({"sucesso": False, "erro": f"falha na consulta: {e}"}, ensure_ascii=False))
        return 1

    if not titulos:
        print(json.dumps({"sucesso": False, "erro": "nenhum titulo encontrado para o filtro",
                          "filtros": filtros}, ensure_ascii=False))
        return 0

    file_id = str(uuid.uuid4())[:8]
    filename = f"{file_id}_{a.nome}.xlsx"
    sid = _safe_sid(a.session_id)
    filepath = os.path.join(get_upload_folder(a.session_id), filename)
    resumo = gerar_excel(titulos, filepath)

    if not os.path.exists(filepath) or os.path.getsize(filepath) <= 0:
        print(json.dumps({"sucesso": False, "erro": "arquivo gerado vazio/ausente"}, ensure_ascii=False))
        return 1

    print(json.dumps({
        "sucesso": True,
        "arquivo": filename,
        "url": f"/agente/api/files/{sid}/{filename}",
        "resumo": resumo,
        "nota": "Situacao CONFIRMADO vem de parcela_paga do Odoo; pagamentos so no extrato "
                "Grafeno ainda nao baixados aparecem como Vencido/Em Aberto.",
    }, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
