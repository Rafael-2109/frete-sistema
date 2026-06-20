#!/usr/bin/env python3
"""
Script: corrigindo_dados_assai.py

Skill WRITE — backfill / correcao manual de dados do modulo Motos Assai (B2B
Q.P.A.). Cobre o que os fluxos normais (upload PDF da NF, CCe, telas 1-a-1)
NAO alcancam: carga historica em lote a partir de planilha Excel, eventos de
estado com DATA RETROATIVA, cadastros (loja/modelo) e itens de pedido ABERTO.

Opera SOMENTE no lado ENTRADA/ESTOQUE/ESTADO/CADASTRO — 100% isolado da Nacom
(NAO cria espelho em `separacao`, NAO toca embarque/frete/financeiro). Reusa os
services do modulo (nunca SQL/UPDATE/DELETE cru) e respeita os guard-rails:

    - assai_moto_evento e APPEND-ONLY: correcao = NOVO evento (nunca UPDATE/DELETE).
    - FATURADA / SEPARADA / CARREGADA sao PROIBIDOS no backfill: nascem do fluxo
      oficial (separar / importar NF Q.P.A. com match BATEU). A skill bloqueia e
      orienta o caminho correto.
    - recebimento fisico e SOT de cor/modelo: a skill so define cor/modelo ao
      CRIAR a moto; nunca sobrescreve moto existente.

Modos (exatamente 1 por invocacao):
    --criar-moto          cria AssaiMoto + evento ESTOQUE (data retroativa opc.)
    --definir-estado      leva 1 chassi ao estado-alvo via cadeia de eventos
    --cadastrar-loja      cria AssaiLoja (--campos JSON)
    --corrigir-loja       atualiza AssaiLoja (--loja-id --campos JSON)
    --cadastrar-modelo    cria AssaiModelo (--campos JSON)
    --corrigir-modelo     atualiza AssaiModelo (--modelo-id --campos JSON)
    --item-pedido         add/edit/remove item de pedido ABERTO (--acao)
    --planilha-estado     carga em LOTE a partir de Excel (modo principal)

Args sempre: --user-id <id>. Sem --confirmar -> preview dry-run (exit 4).

Exit codes:
    0 - sucesso (efetivado)
    1 - validacao (dado invalido / estado-alvo proibido / arg faltando)
    2 - infra (DB / app boot / leitura de arquivo)
    3 - sem autorizacao (pode_acessar_motos_assai=False)
    4 - dry-run preview
    5 - conflito (UNIQUE / chassi ja em outro estado do fluxo de venda)
"""
import sys
import os
import json
import argparse
import contextlib
import io
from datetime import datetime, date
from decimal import Decimal

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../../../..')))

with contextlib.redirect_stdout(io.StringIO()):
    from app import create_app, db  # noqa: E402, F401


# ----- Serializacao / helpers de resposta -----------------------------------

def _json_default(obj):
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    if isinstance(obj, Decimal):
        return float(obj)
    return str(obj)


def _verificar_autorizacao(user_id):
    """Retorna (ok: bool, motivo: Optional[str])."""
    from app.auth.models import Usuario
    u = Usuario.query.get(user_id)
    if not u:
        return False, 'usuario_nao_encontrado'
    if not u.pode_acessar_motos_assai():
        return False, 'sem_permissao_motos_assai'
    return True, None


def _resp_dry_run(modo, **kwargs):
    out = {'dry_run': True, 'modo': modo, 'exit_code': 4}
    out.update(kwargs)
    return out


def _resp_validacao(erro, **kwargs):
    out = {'ok': False, 'erro': erro, 'tipo_erro': 'validacao', 'exit_code': 1}
    out.update(kwargs)
    return out


def _resp_conflito(erro, **kwargs):
    out = {'ok': False, 'erro': erro, 'tipo_erro': 'conflito', 'retry': True,
           'exit_code': 5}
    out.update(kwargs)
    return out


def _resp_autorizacao(motivo, user_id):
    return {'ok': False, 'erro': motivo, 'tipo_erro': 'autorizacao',
            'user_id': user_id, 'exit_code': 3}


def _resp_ok(modo, payload=None):
    out = {'ok': True, 'modo': modo, 'exit_code': 0}
    out.update(payload or {})
    return out


# ----- Datas (Brasil naive) -------------------------------------------------

def _parse_ocorrido_em(valor):
    """Converte string/datetime/date em datetime Brasil naive (sem tzinfo).

    Aceita: datetime/date (de openpyxl), 'YYYY-MM-DD[ HH:MM:SS]', 'DD/MM/YYYY'.
    Retorna None se valor vazio. Levanta ValueError em formato desconhecido.
    """
    if valor is None or valor == '':
        return None
    if isinstance(valor, datetime):
        return valor.replace(tzinfo=None)
    if isinstance(valor, date):
        return datetime(valor.year, valor.month, valor.day)
    s = str(valor).strip()
    if not s:
        return None
    formatos = ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S',
                '%d/%m/%Y', '%d/%m/%y')
    for fmt in formatos:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    raise ValueError(f'data invalida: {valor!r} (use YYYY-MM-DD ou DD/MM/YYYY)')


# ----- Maquina de estados de backfill (cadeia de eventos) -------------------
# Estados-alvo que o backfill PODE definir (lado entrada/estoque). FATURADA,
# SEPARADA, CARREGADA, CANCELADA, MOTO_FALTANDO sao do fluxo oficial e ficam
# de fora — sao bloqueados com orientacao.

def _eventos_consts():
    from app.motos_assai.models import (
        EVENTO_ESTOQUE, EVENTO_MONTADA, EVENTO_PENDENTE,
        EVENTO_PENDENCIA_RESOLVIDA, EVENTO_DISPONIVEL,
        EVENTO_REVERTIDA_PARA_MONTADA, EVENTO_DEMONSTRACAO,
    )
    return {
        'ESTOQUE': EVENTO_ESTOQUE, 'MONTADA': EVENTO_MONTADA,
        'PENDENTE': EVENTO_PENDENTE, 'PENDENCIA_RESOLVIDA': EVENTO_PENDENCIA_RESOLVIDA,
        'DISPONIVEL': EVENTO_DISPONIVEL, 'REVERTIDA': EVENTO_REVERTIDA_PARA_MONTADA,
        'DEMONSTRACAO': EVENTO_DEMONSTRACAO,
    }


# Estados de partida que NAO sao operaveis por este backfill (fluxo de venda).
ESTADOS_FLUXO_VENDA = {'SEPARADA', 'CARREGADA', 'FATURADA', 'CANCELADA',
                       'MOTO_FALTANDO'}

# Mapa de termos da planilha -> estado-alvo canonico.
STATUS_PLANILHA = {
    'ESTOQUE': 'ESTOQUE', 'EM ESTOQUE': 'ESTOQUE', 'ESTOQUE CD': 'ESTOQUE',
    'MONTADA': 'MONTADA', 'MONTADO': 'MONTADA', 'MONTAGEM': 'MONTADA',
    'PENDENTE': 'PENDENTE', 'PENDENCIA': 'PENDENTE', 'PENDÊNCIA': 'PENDENTE',
    'DISPONIVEL': 'DISPONIVEL', 'DISPONÍVEL': 'DISPONIVEL', 'PRONTA': 'DISPONIVEL',
    'DEMONSTRACAO': 'DEMONSTRACAO', 'DEMONSTRAÇÃO': 'DEMONSTRACAO', 'DEMO': 'DEMONSTRACAO',
}

ALVOS_PERMITIDOS = {'ESTOQUE', 'MONTADA', 'PENDENTE', 'DISPONIVEL', 'DEMONSTRACAO'}


def _normalizar_estado_partida(status):
    """Reduz estados transitorios ao efetivo para o BFS."""
    if status in ('PENDENCIA_RESOLVIDA', 'REVERTIDA_PARA_MONTADA'):
        return 'MONTADA'
    return status  # None / ESTOQUE / MONTADA / PENDENTE / DISPONIVEL / DEMONSTRACAO / fluxo-venda


def _arestas():
    """Grafo de transicoes de backfill. Cada aresta: (origem, [eventos], destino).

    Eventos sao nomes-curtos resolvidos em _eventos_consts(). Caminhos compostos
    (PENDENTE->MONTADA) emitem a sequencia correta (PENDENCIA_RESOLVIDA + MONTADA),
    preservando o desempate por id (ultimo evento = estado efetivo).
    """
    return [
        (None, ['ESTOQUE'], 'ESTOQUE'),
        ('ESTOQUE', ['MONTADA'], 'MONTADA'),
        ('ESTOQUE', ['PENDENTE'], 'PENDENTE'),
        ('MONTADA', ['DISPONIVEL'], 'DISPONIVEL'),
        ('MONTADA', ['PENDENTE'], 'PENDENTE'),
        ('DISPONIVEL', ['PENDENTE'], 'PENDENTE'),
        ('DISPONIVEL', ['REVERTIDA'], 'MONTADA'),
        ('PENDENTE', ['PENDENCIA_RESOLVIDA', 'MONTADA'], 'MONTADA'),
        ('ESTOQUE', ['DEMONSTRACAO'], 'DEMONSTRACAO'),
        ('MONTADA', ['DEMONSTRACAO'], 'DEMONSTRACAO'),
        ('PENDENTE', ['DEMONSTRACAO'], 'DEMONSTRACAO'),
        ('DISPONIVEL', ['DEMONSTRACAO'], 'DEMONSTRACAO'),
        ('DEMONSTRACAO', ['ESTOQUE'], 'ESTOQUE'),
    ]


def _planejar_cadeia(status_atual, alvo):
    """BFS do estado atual ao alvo. Retorna (eventos:list[str]|None, erro:str|None).

    eventos == [] -> ja esta no alvo (idempotente, skip).
    eventos == None -> impossivel/proibido (erro preenchido).
    """
    if alvo not in ALVOS_PERMITIDOS:
        return None, (
            f"estado-alvo '{alvo}' nao e operavel por backfill. Permitidos: "
            f"{sorted(ALVOS_PERMITIDOS)}. FATURADA nasce do import de NF Q.P.A. "
            f"(match BATEU); SEPARADA/CARREGADA do fluxo de separacao/carregamento "
            f"-> use registrando-evento-moto-assai ou suba a NF real."
        )

    partida = _normalizar_estado_partida(status_atual)
    if partida in ESTADOS_FLUXO_VENDA:
        return None, (
            f"chassi esta em '{status_atual}' (fluxo de venda). Backfill nao "
            f"opera a partir desse estado — reverta pelo fluxo oficial "
            f"(cancelar separacao/NF) antes."
        )
    if partida == alvo:
        return [], None

    # BFS sobre as arestas
    arestas = _arestas()
    adj = {}
    for origem, eventos, destino in arestas:
        adj.setdefault(origem, []).append((eventos, destino))

    from collections import deque
    fila = deque([(partida, [])])
    visitados = {partida}
    while fila:
        atual, caminho = fila.popleft()
        for eventos, destino in adj.get(atual, []):
            if destino in visitados:
                continue
            novo_caminho = caminho + eventos
            if destino == alvo:
                return novo_caminho, None
            visitados.add(destino)
            fila.append((destino, novo_caminho))

    return None, (
        f"nao ha caminho de '{status_atual or 'SEM EVENTOS'}' para '{alvo}' no "
        f"backfill. Verifique o estado atual do chassi."
    )


# Eventos da cadeia que exigem justificativa (motivo).
_EVENTOS_EXIGEM_MOTIVO = {'PENDENTE', 'REVERTIDA'}


# ----- Helpers de dominio ---------------------------------------------------

def _filtrar_campos_model(model_cls, dados):
    """Mantem so as chaves que sao colunas do model (exceto id/criado*).

    Evita passar lixo da planilha para o construtor. Retorna (limpos, ignorados).
    """
    colunas = {c.name for c in model_cls.__table__.columns}
    bloqueadas = {'id', 'criado_em', 'criada_em'}
    limpos, ignorados = {}, []
    for k, v in dados.items():
        if k in colunas and k not in bloqueadas:
            limpos[k] = v
        else:
            ignorados.append(k)
    return limpos, ignorados


def _resolver_modelo_id(modelo_texto, modelo_id):
    """Resolve modelo_id explicito OU texto via modelo_resolver. (id, erro)."""
    from app.motos_assai.models import AssaiModelo
    if modelo_id is not None:
        m = AssaiModelo.query.get(modelo_id)
        if not m:
            return None, f'modelo_id={modelo_id} nao encontrado'
        return m.id, None
    if modelo_texto:
        from app.motos_assai.services.modelo_resolver import resolver_modelo
        m = resolver_modelo(modelo_texto, origem='BACKFILL')
        if not m:
            return None, f"modelo '{modelo_texto}' nao resolvido (cadastre alias?)"
        return m.id, None
    return None, 'informe --modelo-id ou --modelo <texto>'


def _criar_moto(chassi, modelo_id, cor, motor, ano, ocorrido_em, user_id,
                fonte, confirmar):
    """Cria AssaiMoto (se nao existe) + emite ESTOQUE retroativo. Idempotente.

    Retorna dict-resposta. NAO commita (caller commita).
    """
    from app.motos_assai.models import AssaiMoto
    from app.motos_assai.services.moto_evento_service import (
        emitir_evento, status_efetivo,
    )
    from app.motos_assai.services.chassi_validator import validar_chassi

    chassi_norm = chassi.strip().upper()
    existente = AssaiMoto.query.filter_by(chassi=chassi_norm).first()
    status = status_efetivo(chassi_norm)

    # Idempotencia: moto ja existe E ja tem historico -> skip criacao/ESTOQUE
    if existente and status is not None:
        return _resp_ok('criar-moto', {
            'chassi': chassi_norm, 'skipped': True,
            'motivo_skip': f'moto ja existe com status_efetivo={status}',
            'moto_id': existente.id,
        })

    regex_check = validar_chassi(chassi_norm, modelo_id)

    if not confirmar:
        return _resp_dry_run('criar-moto', chassi=chassi_norm,
                             status_efetivo_atual=status,
                             acao_pretendida='criar AssaiMoto + evento ESTOQUE',
                             modelo_id=modelo_id, cor=cor,
                             ocorrido_em=ocorrido_em,
                             regex_check=regex_check)

    moto = existente
    if not moto:
        moto = AssaiMoto(chassi=chassi_norm, modelo_id=modelo_id,
                         cor=(cor or None), motor=(motor or None), ano=ano)
        db.session.add(moto)
        db.session.flush()

    ev = emitir_evento(
        chassi=chassi_norm, tipo=_eventos_consts()['ESTOQUE'],
        operador_id=user_id, ocorrido_em=ocorrido_em,
        observacao='Carga historica (backfill)',
        dados_extras={'origem': 'backfill:corrigindo-dados-assai',
                      'modo': 'criar-moto', 'fonte': fonte},
    )
    return _resp_ok('criar-moto', {
        'chassi': chassi_norm, 'moto_id': moto.id, 'evento_id': ev.id,
        'tipo': ev.tipo, 'regex_check': regex_check,
    })


def _emitir_cadeia(chassi_norm, cadeia, ocorrido_em, motivo, user_id, fonte,
                   modo, alvo):
    """Emite os eventos de uma cadeia JA planejada. NAO commita (caller commita).

    Usado tanto pelo modo pontual quanto pelo lote (planilha), garantindo que a
    execucao siga exatamente o plano calculado por _planejar_cadeia.
    """
    from app.motos_assai.services.moto_evento_service import emitir_evento
    consts = _eventos_consts()
    eventos_ids = []
    for nome in cadeia:
        ev = emitir_evento(
            chassi=chassi_norm, tipo=consts[nome], operador_id=user_id,
            ocorrido_em=ocorrido_em,
            observacao=(motivo if nome in _EVENTOS_EXIGEM_MOTIVO else
                        f'Backfill -> {alvo}'),
            dados_extras={'origem': 'backfill:corrigindo-dados-assai',
                          'modo': modo, 'fonte': fonte, 'alvo': alvo},
        )
        eventos_ids.append({'tipo': ev.tipo, 'evento_id': ev.id})
    return eventos_ids


def _aplicar_cadeia(chassi, alvo, ocorrido_em, motivo, user_id, fonte,
                    confirmar, modo='definir-estado'):
    """Leva o chassi ao estado-alvo via cadeia de eventos. NAO commita."""
    from app.motos_assai.models import AssaiMoto
    from app.motos_assai.services.moto_evento_service import status_efetivo

    chassi_norm = chassi.strip().upper()

    if not AssaiMoto.query.filter_by(chassi=chassi_norm).first():
        return _resp_validacao(
            f'chassi {chassi_norm} nao existe em assai_moto — crie a moto '
            f'primeiro (--criar-moto ou coluna MODELO na planilha).',
            chassi=chassi_norm,
        )

    status = status_efetivo(chassi_norm)
    cadeia, erro = _planejar_cadeia(status, alvo)
    if erro:
        # estado de fluxo de venda -> conflito (5); demais validacao (1)
        if _normalizar_estado_partida(status) in ESTADOS_FLUXO_VENDA:
            return _resp_conflito(erro, chassi=chassi_norm,
                                  status_efetivo_atual=status)
        return _resp_validacao(erro, chassi=chassi_norm,
                               status_efetivo_atual=status)

    if not cadeia:  # [] -> ja no alvo (idempotente). None nao ocorre (erro tratado acima)
        return _resp_ok(modo, {'chassi': chassi_norm, 'skipped': True,
                               'motivo_skip': f'ja esta em {alvo}',
                               'status_efetivo_atual': status})

    exige_motivo = any(e in _EVENTOS_EXIGEM_MOTIVO for e in cadeia)
    if exige_motivo and not (motivo and len(motivo.strip()) >= 3):
        return _resp_validacao(
            f'a cadeia ate {alvo} inclui PENDENTE/REVERTIDA e exige --motivo '
            f'(>=3 chars). Cadeia: {cadeia}', chassi=chassi_norm,
        )

    if not confirmar:
        return _resp_dry_run(modo, chassi=chassi_norm,
                             status_efetivo_atual=status, alvo=alvo,
                             cadeia_eventos=cadeia, ocorrido_em=ocorrido_em,
                             motivo=motivo)

    eventos_ids = _emitir_cadeia(chassi_norm, cadeia, ocorrido_em, motivo,
                                 user_id, fonte, modo, alvo)
    return _resp_ok(modo, {'chassi': chassi_norm, 'de': status, 'para': alvo,
                           'eventos': eventos_ids})


# ----- Comandos -------------------------------------------------------------

def _cmd_criar_moto(args):
    if not args.chassi:
        return _resp_validacao('--chassi obrigatorio para --criar-moto')
    modelo_id, erro = _resolver_modelo_id(args.modelo, args.modelo_id)
    if erro:
        return _resp_validacao(erro, chassi=args.chassi)
    try:
        ocorrido = _parse_ocorrido_em(args.ocorrido_em)
    except ValueError as e:
        return _resp_validacao(str(e))
    out = _criar_moto(args.chassi, modelo_id, args.cor, args.motor, args.ano,
                      ocorrido, args.user_id, 'pontual', args.confirmar)
    if args.confirmar and out.get('ok'):
        db.session.commit()
    return out


def _cmd_definir_estado(args):
    if not args.chassi:
        return _resp_validacao('--chassi obrigatorio para --definir-estado')
    if not args.estado:
        return _resp_validacao('--estado obrigatorio (ESTOQUE/MONTADA/PENDENTE/'
                               'DISPONIVEL/DEMONSTRACAO)')
    alvo = STATUS_PLANILHA.get(args.estado.strip().upper(),
                               args.estado.strip().upper())
    try:
        ocorrido = _parse_ocorrido_em(args.ocorrido_em)
    except ValueError as e:
        return _resp_validacao(str(e))
    out = _aplicar_cadeia(args.chassi, alvo, ocorrido, args.motivo,
                          args.user_id, 'pontual', args.confirmar)
    if args.confirmar and out.get('ok'):
        db.session.commit()
    return out


def _carregar_campos_json(args):
    if not args.campos:
        return None, _resp_validacao('--campos JSON obrigatorio (ex: '
                                     '\'{"numero":"14","nome":"Loja X"}\')')
    try:
        dados = json.loads(args.campos)
    except json.JSONDecodeError as e:
        return None, _resp_validacao(f'--campos nao e JSON valido: {e}')
    if not isinstance(dados, dict):
        return None, _resp_validacao('--campos deve ser um objeto JSON')
    return dados, None


def _cmd_cadastrar_loja(args):
    from app.motos_assai.models import AssaiLoja
    from app.motos_assai.services.loja_service import criar_loja, LojaJaExisteError
    dados, erro = _carregar_campos_json(args)
    if erro:
        return erro
    limpos, ignorados = _filtrar_campos_model(AssaiLoja, dados)
    if 'numero' not in limpos:
        return _resp_validacao("campo 'numero' obrigatorio para cadastrar loja",
                               ignorados=ignorados)
    if not args.confirmar:
        return _resp_dry_run('cadastrar-loja', campos=limpos, ignorados=ignorados,
                             acao_pretendida='criar AssaiLoja')
    try:
        loja = criar_loja(limpos, operador_id=args.user_id)
        return _resp_ok('cadastrar-loja', {'loja_id': loja.id, 'numero': loja.numero,
                                           'ignorados': ignorados})
    except LojaJaExisteError as e:
        return _resp_conflito(str(e))


def _cmd_corrigir_loja(args):
    from app.motos_assai.models import AssaiLoja
    from app.motos_assai.services.loja_service import atualizar_loja
    if args.loja_id is None:
        return _resp_validacao('--loja-id obrigatorio para --corrigir-loja')
    dados, erro = _carregar_campos_json(args)
    if erro:
        return erro
    limpos, ignorados = _filtrar_campos_model(AssaiLoja, dados)
    if not limpos:
        return _resp_validacao('nenhum campo valido em --campos', ignorados=ignorados)
    alvo = AssaiLoja.query.get(args.loja_id)
    if not alvo:
        return _resp_validacao(f'loja_id={args.loja_id} nao encontrada')
    if not args.confirmar:
        return _resp_dry_run('corrigir-loja', loja_id=args.loja_id,
                             campos=limpos, ignorados=ignorados,
                             valores_atuais={k: getattr(alvo, k, None) for k in limpos})
    loja = atualizar_loja(args.loja_id, limpos, operador_id=args.user_id)
    return _resp_ok('corrigir-loja', {'loja_id': loja.id, 'numero': loja.numero,
                                      'ignorados': ignorados})


def _cmd_cadastrar_modelo(args):
    from app.motos_assai.models import AssaiModelo
    from app.motos_assai.services.modelo_service import criar_modelo, ModeloJaExisteError
    dados, erro = _carregar_campos_json(args)
    if erro:
        return erro
    limpos, ignorados = _filtrar_campos_model(AssaiModelo, dados)
    if 'codigo' not in limpos:
        return _resp_validacao("campo 'codigo' obrigatorio para cadastrar modelo",
                               ignorados=ignorados)
    if not args.confirmar:
        return _resp_dry_run('cadastrar-modelo', campos=limpos, ignorados=ignorados)
    try:
        m = criar_modelo(limpos)
        return _resp_ok('cadastrar-modelo', {'modelo_id': m.id, 'codigo': m.codigo,
                                             'ignorados': ignorados})
    except ModeloJaExisteError as e:
        return _resp_conflito(str(e))


def _cmd_corrigir_modelo(args):
    from app.motos_assai.models import AssaiModelo
    from app.motos_assai.services.modelo_service import atualizar_modelo
    if args.modelo_id is None:
        return _resp_validacao('--modelo-id obrigatorio para --corrigir-modelo')
    dados, erro = _carregar_campos_json(args)
    if erro:
        return erro
    limpos, ignorados = _filtrar_campos_model(AssaiModelo, dados)
    if not limpos:
        return _resp_validacao('nenhum campo valido em --campos', ignorados=ignorados)
    alvo = AssaiModelo.query.get(args.modelo_id)
    if not alvo:
        return _resp_validacao(f'modelo_id={args.modelo_id} nao encontrado')
    if not args.confirmar:
        return _resp_dry_run('corrigir-modelo', modelo_id=args.modelo_id,
                             campos=limpos, ignorados=ignorados,
                             valores_atuais={k: getattr(alvo, k, None) for k in limpos})
    m = atualizar_modelo(args.modelo_id, limpos, operador_id=args.user_id)
    return _resp_ok('corrigir-modelo', {'modelo_id': m.id, 'codigo': m.codigo,
                                        'ignorados': ignorados})


def _cmd_item_pedido(args):
    from app.motos_assai.services.pedido_service import (
        adicionar_item_manual, editar_item_manual, remover_item_manual,
        PedidoVoeEdicaoError,
    )
    acao = (args.acao or '').strip().lower()
    if acao not in ('add', 'edit', 'remove'):
        return _resp_validacao('--acao deve ser add | edit | remove')

    if acao == 'add':
        faltam = [n for n, v in (('--pedido-id', args.pedido_id),
                                 ('--loja-id', args.loja_id),
                                 ('--modelo-id', args.modelo_id),
                                 ('--qtd', args.qtd),
                                 ('--valor', args.valor)) if v is None]
        if faltam:
            return _resp_validacao(f'args faltando para add: {", ".join(faltam)}')
        if not args.confirmar:
            return _resp_dry_run('item-pedido', acao='add', pedido_id=args.pedido_id,
                                 loja_id=args.loja_id, modelo_id=args.modelo_id,
                                 qtd=args.qtd, valor_unitario=args.valor)
        try:
            item = adicionar_item_manual(args.pedido_id, args.loja_id, args.modelo_id,
                                         args.qtd, args.valor, operador_id=args.user_id)
            return _resp_ok('item-pedido', {'acao': 'add', 'item_id': item.id,
                                            'qtd_pedida': item.qtd_pedida})
        except PedidoVoeEdicaoError as e:
            return _resp_validacao(str(e))

    if acao == 'edit':
        if args.item_id is None or args.qtd is None or args.valor is None:
            return _resp_validacao('edit exige --item-id, --qtd, --valor')
        if not args.confirmar:
            return _resp_dry_run('item-pedido', acao='edit', item_id=args.item_id,
                                 qtd=args.qtd, valor_unitario=args.valor)
        try:
            item = editar_item_manual(args.item_id, args.qtd, args.valor,
                                      operador_id=args.user_id)
            return _resp_ok('item-pedido', {'acao': 'edit', 'item_id': item.id,
                                            'qtd_pedida': item.qtd_pedida})
        except PedidoVoeEdicaoError as e:
            return _resp_validacao(str(e))

    # remove
    if args.item_id is None:
        return _resp_validacao('remove exige --item-id')
    if not args.confirmar:
        return _resp_dry_run('item-pedido', acao='remove', item_id=args.item_id)
    try:
        remover_item_manual(args.item_id, operador_id=args.user_id)
        return _resp_ok('item-pedido', {'acao': 'remove', 'item_id': args.item_id})
    except PedidoVoeEdicaoError as e:
        return _resp_validacao(str(e))


# ----- Planilha em lote -----------------------------------------------------

def _achar_header(ws, coluna_chassi):
    """Retorna (linha_idx, {nome_upper: col_idx}) procurando a celula chassi."""
    alvo = coluna_chassi.strip().upper()
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
        upper = [str(c).strip().upper() if c is not None else '' for c in row]
        if alvo in upper:
            mapa = {nome: i for i, nome in enumerate(upper) if nome}
            return r_idx, mapa
    return None, None


def _cmd_planilha_estado(args):
    if not args.excel:
        return _resp_validacao('--excel <caminho> obrigatorio para --planilha-estado')
    if not os.path.isfile(args.excel):
        return _resp_validacao(f'arquivo nao encontrado: {args.excel}')
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {'ok': False, 'erro': 'openpyxl nao disponivel', 'tipo_erro': 'infra',
                'exit_code': 2}

    col_chassi = args.coluna_chassi
    col_status = args.coluna_status
    col_modelo = args.coluna_modelo
    col_cor = args.coluna_cor
    col_data = args.coluna_data

    try:
        wb = load_workbook(args.excel, data_only=True, read_only=True)
    except Exception as e:
        return {'ok': False, 'erro': f'falha ao abrir Excel: {e}',
                'tipo_erro': 'infra', 'exit_code': 2}
    ws = wb[args.aba] if args.aba else wb.active

    header_row, mapa = _achar_header(ws, col_chassi)
    if not mapa or header_row is None:
        return _resp_validacao(
            f"coluna '{col_chassi}' nao encontrada no cabecalho (30 primeiras "
            f"linhas). Ajuste --coluna-chassi.")

    def _cel(row, nome):
        idx = mapa.get(nome.strip().upper()) if nome else None
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    resultados = []
    resumo = {'linhas': 0, 'criadas': 0, 'estado_aplicado': 0, 'puladas': 0,
              'erros': 0, 'sem_status': 0}
    fonte = os.path.basename(args.excel)

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        chassi = _cel(row, col_chassi)
        if chassi is None or str(chassi).strip() == '':
            continue
        chassi = str(chassi).strip().upper()
        resumo['linhas'] += 1
        linha_out = {'chassi': chassi}

        status_raw = _cel(row, col_status)
        status_txt = str(status_raw).strip().upper() if status_raw is not None else ''
        alvo = STATUS_PLANILHA.get(status_txt)
        if not alvo:
            resumo['sem_status'] += 1
            linha_out.update({'status': 'pulada', 'motivo': (
                f"STATUS '{status_txt or '(vazio)'}' nao mapeavel "
                f"(FATURADO/SEPARADO etc. -> use fluxo oficial)")})
            resultados.append(linha_out)
            continue

        try:
            ocorrido = _parse_ocorrido_em(_cel(row, col_data)) if col_data else None
        except ValueError as e:
            resumo['erros'] += 1
            linha_out.update({'status': 'erro', 'motivo': str(e)})
            resultados.append(linha_out)
            continue

        try:
            from app.motos_assai.models import AssaiMoto
            from app.motos_assai.services.moto_evento_service import status_efetivo

            existe = AssaiMoto.query.filter_by(chassi=chassi).first()
            status_atual = status_efetivo(chassi) if existe else None

            # 1) precisa criar a moto?
            vai_criar = False
            modelo_id = None
            if not existe:
                if not args.criar_faltantes:
                    resumo['erros'] += 1
                    linha_out.update({'status': 'erro', 'motivo':
                                      'moto inexistente (use --criar-faltantes)'})
                    resultados.append(linha_out)
                    continue
                modelo_id, erro_m = _resolver_modelo_id(_cel(row, col_modelo), None)
                if erro_m:
                    resumo['erros'] += 1
                    linha_out.update({'status': 'erro', 'motivo': erro_m})
                    resultados.append(linha_out)
                    continue
                vai_criar = True

            # 2) planejar a cadeia (estado-base = ESTOQUE quando vai criar, pois
            #    a criacao ja emite ESTOQUE; a cadeia parte dali ate o alvo)
            estado_base = 'ESTOQUE' if vai_criar else status_atual
            cadeia, erro_plan = _planejar_cadeia(estado_base, alvo)
            if erro_plan:
                resumo['erros'] += 1
                linha_out.update({'status': 'erro', 'alvo': alvo, 'motivo': erro_plan})
                resultados.append(linha_out)
                continue
            cadeia = cadeia or []  # erro_plan None garante list; normaliza p/ type-checker

            # 3) motivo obrigatorio se a cadeia inclui PENDENTE/REVERTIDA
            if any(e in _EVENTOS_EXIGEM_MOTIVO for e in cadeia) and not (
                    args.motivo and len(args.motivo.strip()) >= 3):
                resumo['erros'] += 1
                linha_out.update({'status': 'erro', 'alvo': alvo,
                                  'motivo': f'cadeia {cadeia} exige --motivo (>=3 chars)'})
                resultados.append(linha_out)
                continue

            # 4) nada a fazer (ja no alvo, sem criar)
            if not vai_criar and not cadeia:
                resumo['puladas'] += 1
                linha_out.update({'status': 'pulada', 'alvo': alvo,
                                  'motivo': f'ja esta em {alvo}'})
                resultados.append(linha_out)
                continue

            eventos_previstos = (['ESTOQUE'] if vai_criar else []) + cadeia

            # 5) dry-run: registra a previsao sem mutar
            if not args.confirmar:
                if vai_criar:
                    resumo['criadas'] += 1
                resumo['estado_aplicado'] += 1
                linha_out.update({'status': 'ok', 'alvo': alvo,
                                  'criaria_moto': vai_criar,
                                  'eventos_previstos': eventos_previstos})
                resultados.append(linha_out)
                continue

            # 6) confirmar: cria (emite ESTOQUE) + emite a cadeia restante.
            #    NAO commita aqui — commit unico no fim do lote (tudo-ou-nada).
            if vai_criar:
                r_criar = _criar_moto(chassi, modelo_id, _cel(row, col_cor), None,
                                      None, ocorrido, args.user_id, fonte, True)
                if not r_criar.get('ok'):
                    resumo['erros'] += 1
                    linha_out.update({'status': 'erro',
                                      'motivo': r_criar.get('erro', 'falha criar moto')})
                    resultados.append(linha_out)
                    continue
                resumo['criadas'] += 1
            if cadeia:
                _emitir_cadeia(chassi, cadeia, ocorrido, args.motivo,
                               args.user_id, fonte, 'planilha-estado', alvo)
            resumo['estado_aplicado'] += 1
            linha_out.update({'status': 'ok', 'alvo': alvo, 'criou_moto': vai_criar,
                              'eventos': eventos_previstos})
            resultados.append(linha_out)
        except Exception as e:  # noqa: BLE001 — isola a linha, segue o lote
            db.session.rollback()
            resumo['erros'] += 1
            linha_out.update({'status': 'erro', 'motivo': f'{type(e).__name__}: {e}'})
            resultados.append(linha_out)

    if args.confirmar:
        if resumo['erros'] == 0:
            db.session.commit()
        else:
            db.session.rollback()
            return {'ok': False, 'tipo_erro': 'validacao', 'exit_code': 1,
                    'modo': 'planilha-estado', 'fonte': fonte,
                    'erro': (f"{resumo['erros']} linha(s) com erro — NADA foi "
                             f"commitado (transacao revertida). Corrija e re-rode."),
                    'resumo': resumo, 'linhas': resultados}
        return _resp_ok('planilha-estado', {'fonte': fonte, 'resumo': resumo,
                                            'linhas': resultados})

    return _resp_dry_run('planilha-estado', fonte=fonte, resumo=resumo,
                         linhas=resultados,
                         aviso='dry-run: nada foi alterado. Use --confirmar.')


# ----- Faturamento / NF Q.P.A. (lado modulo, isolado da Nacom) --------------
# A NF Q.P.A. e o faturamento do MODULO (PJ Q.P.A.) — nao se mistura com o
# faturamento Nacom. Gravar a NF real (PDF) da' lastro ao evento FATURADA; o
# unico cruzamento e logistico (espelho da separacao -> embarque/frete), que
# so ocorre quando a separacao foi espelhada (fluxo oficial).

def _cmd_importar_nf(args):
    """Grava faturamento Q.P.A. a partir do PDF da NF (fluxo oficial, lastreado).

    Cria AssaiNfQpa + itens, roda o match (pode subir chassis a FATURADA e o
    pedido a FATURADO). Commita internamente (ponto de entrada oficial).
    """
    from app.motos_assai.services.parsers.nf_qpa_adapter import (
        importar_nf_qpa, NfQpaParseError, NfQpaJaImportadaError,
    )
    if not args.pdf:
        return _resp_validacao('--pdf <caminho> obrigatorio para --importar-nf')
    if not os.path.isfile(args.pdf):
        return _resp_validacao(f'arquivo nao encontrado: {args.pdf}')
    if not args.confirmar:
        return _resp_dry_run(
            'importar-nf', pdf=args.pdf,
            acao_pretendida='parsear PDF -> criar AssaiNfQpa + match '
                            '(pode subir chassis a FATURADA)')
    try:
        with open(args.pdf, 'rb') as fh:
            pdf_bytes = fh.read()
        nf = importar_nf_qpa(pdf_bytes, os.path.basename(args.pdf), args.user_id)
        return _resp_ok('importar-nf', {
            'nf_id': nf.id, 'chave_44': nf.chave_44, 'numero': nf.numero,
            'status_match': nf.status_match, 'loja_id': nf.loja_id})
    except NfQpaJaImportadaError as e:
        return _resp_conflito(str(e))
    except NfQpaParseError as e:
        return _resp_validacao(str(e))


def _cmd_corrigir_chassi_nf(args):
    """Troca chassi(s) numa NF Q.P.A. (alterar chassi em NF).

    Reusa aplicar_correcao_cce: substitui chassi antigo->novo em AssaiNfQpaItem,
    registra AssaiNfQpaItemVinculoHistorico, reverte FATURADA do antigo e re-roda
    o match. Caller commita.
    """
    from app.motos_assai.services.cancelamento_nf_service import (
        aplicar_correcao_cce, CancelamentoValidationError,
    )
    if args.nf_id is None:
        return _resp_validacao('--nf-id obrigatorio para --corrigir-chassi-nf')
    pares = []
    if args.pares_json:
        try:
            raw = json.loads(args.pares_json)
            pares = [(str(a).strip().upper(), str(b).strip().upper()) for a, b in raw]
        except Exception as e:  # noqa: BLE001
            return _resp_validacao(f'--pares-json invalido (use [["ANTIGO","NOVO"]]): {e}')
    elif args.de_chassi and args.para_chassi:
        pares = [(args.de_chassi.strip().upper(), args.para_chassi.strip().upper())]
    else:
        return _resp_validacao('informe --de-chassi e --para-chassi (ou --pares-json)')
    numero_cce = args.numero_cce or f'MANUAL-BACKFILL-u{args.user_id}'
    if not args.confirmar:
        return _resp_dry_run(
            'corrigir-chassi-nf', nf_id=args.nf_id, pares=pares, numero_cce=numero_cce,
            acao_pretendida='trocar chassi(s) na NF + reverter FATURADA do antigo + re-match')
    try:
        nf = aplicar_correcao_cce(args.nf_id, pares, numero_cce, args.user_id)
        db.session.commit()
        return _resp_ok('corrigir-chassi-nf', {
            'nf_id': nf.id, 'status_match': nf.status_match, 'pares': pares,
            'numero_cce': numero_cce})
    except CancelamentoValidationError as e:
        db.session.rollback()
        return _resp_validacao(str(e))


def _cmd_cancelar_nf(args):
    """Cancela NF Q.P.A. (reverte FATURADA dos chassis, limpa espelho/embarque)."""
    from app.motos_assai.services.cancelamento_nf_service import (
        cancelar_nf_qpa, CancelamentoValidationError,
    )
    if args.nf_id is None:
        return _resp_validacao('--nf-id obrigatorio para --cancelar-nf')
    if not (args.motivo and len(args.motivo.strip()) >= 3):
        return _resp_validacao('--motivo (>=3 chars) obrigatorio para --cancelar-nf')
    if not args.confirmar:
        return _resp_dry_run(
            'cancelar-nf', nf_id=args.nf_id, motivo=args.motivo,
            acao_pretendida='NF -> CANCELADA + reverter FATURADA dos chassis')
    try:
        nf = cancelar_nf_qpa(args.nf_id, args.motivo, args.user_id)
        db.session.commit()
        return _resp_ok('cancelar-nf', {'nf_id': nf.id, 'status_match': nf.status_match})
    except CancelamentoValidationError as e:
        db.session.rollback()
        return _resp_validacao(str(e))


def _cmd_vincular_nf(args):
    """Vincula NF NAO_RECONCILIADO a um pedido (match por CNPJ) + re-match.

    Util para 'baixar' faturamento de NF orfa cujo destinatario nao resolveu
    automaticamente (ex: 'SENDAS DISTRIBUIDORA S/A' sem LJ<n>). Caller commita.
    """
    from app.motos_assai.services.parsers.nf_qpa_adapter import (
        vincular_nf_manualmente, VincularNfError,
    )
    if args.nf_id is None or args.pedido_id is None:
        return _resp_validacao('--nf-id e --pedido-id obrigatorios para --vincular-nf')
    if not args.confirmar:
        return _resp_dry_run(
            'vincular-nf', nf_id=args.nf_id, pedido_id=args.pedido_id,
            acao_pretendida='vincular NF NAO_RECONCILIADO ao pedido (CNPJ) + re-match')
    try:
        res = vincular_nf_manualmente(args.nf_id, args.pedido_id, args.user_id)
        db.session.commit()
        return _resp_ok('vincular-nf', {'nf_id': args.nf_id, 'pedido_id': args.pedido_id,
                                        'resultado': res})
    except VincularNfError as e:
        db.session.rollback()
        return _resp_validacao(str(e))


def _cmd_registrar_nf_manual(args):
    """Grava faturamento Q.P.A. SEM PDF, a partir de um JSON com os dados da NF.

    Da' lastro fiscal ao faturamento (a propria NF), reusando o motor oficial
    (criar_nf_qpa_de_dados): cria AssaiNfQpa + itens, roda o match (pode subir
    FATURADA + espelhar a logistica Nacom). Financeiro Nacom NAO e' tocado.
    """
    from app.motos_assai.models import AssaiNfQpa, AssaiLoja
    if not args.nf_json:
        return _resp_validacao(
            '--nf-json obrigatorio. Estrutura: {"chave_44":"<44>","numero":"...",'
            '"loja_id":N ou "destinatario_cnpj":"...","valor_total":N,'
            '"data_emissao":"YYYY-MM-DD","itens":[{"chassi":"...","modelo":"...",'
            '"valor_unitario":N}]}')
    try:
        dados = json.loads(args.nf_json)
    except json.JSONDecodeError as e:
        return _resp_validacao(f'--nf-json invalido: {e}')
    if not isinstance(dados, dict):
        return _resp_validacao('--nf-json deve ser um objeto JSON')

    chave = str(dados.get('chave_44') or '').strip()
    if len(chave) != 44:
        return _resp_validacao(f'chave_44 invalida ({len(chave)} chars; precisa 44)')
    if AssaiNfQpa.query.filter_by(chave_44=chave).first():
        return _resp_conflito(f'NF {chave} ja importada (chave_44 UNIQUE)')
    chassis = [str(i.get('chassi') or '').strip().upper()
               for i in (dados.get('itens') or [])]
    chassis = [c for c in chassis if c]
    if not chassis:
        return _resp_validacao('NF sem itens/chassis — informe "itens":[{"chassi":...}]')

    # parse defensivo da data (deixa o erro claro no dry-run, nao so' no confirmar)
    if dados.get('data_emissao'):
        try:
            dados['data_emissao'] = _parse_ocorrido_em(dados['data_emissao'])
        except ValueError as e:
            return _resp_validacao(f'data_emissao: {e}')

    loja_preview = None
    if dados.get('loja_id'):
        loja_preview = AssaiLoja.query.get(dados['loja_id'])
        if not loja_preview:
            return _resp_validacao(f"loja_id={dados['loja_id']} nao encontrada")

    if not args.confirmar:
        return _resp_dry_run(
            'registrar-nf-manual', chave_44=chave, numero=dados.get('numero'),
            loja_id=(loja_preview.id if loja_preview else dados.get('loja_id')),
            destinatario_cnpj=dados.get('destinatario_cnpj'),
            n_itens=len(chassis), chassis=chassis,
            acao_pretendida='criar AssaiNfQpa + itens + match (pode subir chassis a '
                            'FATURADA e refletir na logistica Nacom)')
    try:
        from app.motos_assai.services.parsers.nf_qpa_adapter import (
            criar_nf_qpa_de_dados, NfQpaParseError, NfQpaJaImportadaError,
        )
        nf = criar_nf_qpa_de_dados(dados, args.user_id)
        return _resp_ok('registrar-nf-manual', {
            'nf_id': nf.id, 'chave_44': nf.chave_44, 'numero': nf.numero,
            'status_match': nf.status_match, 'loja_id': nf.loja_id,
            'nota': 'status_match NAO_RECONCILIADO = NF gravada (lastro) mas sem '
                    'separacao casada; DIVERGENTE/BATEU = casou.'})
    except NfQpaJaImportadaError as e:
        return _resp_conflito(str(e))
    except NfQpaParseError as e:
        return _resp_validacao(str(e))


def _cmd_registrar_devolucao_nfd(args):
    """Registra devolucao (NFd) de 1+ chassis de uma NF Q.P.A. ja FATURADA.

    Reusa devolucao_service.criar_devolucao: cada chassi volta a PENDENTE
    (conserto), AssaiNfQpaItem.devolvido=True e o saldo do modelo retorna ao
    pedido. A NF original NAO e cancelada (devolucao parcial e legitima).
    Idempotente por UNIQUE (nf_id, numero_nfd). Anexos (PDF/XML/imagem) NAO
    sao suportados via CLI — so pela tela. O service commita.
    """
    if args.nf_id is None:
        return _resp_validacao('--nf-id obrigatorio para --registrar-devolucao-nfd')
    if not (args.numero_nfd and args.numero_nfd.strip()):
        return _resp_validacao('--numero-nfd obrigatorio para --registrar-devolucao-nfd')
    if not (args.motivo and len(args.motivo.strip()) >= 3):
        return _resp_validacao('--motivo (>=3 chars) obrigatorio para --registrar-devolucao-nfd')
    if not args.data_devolucao:
        return _resp_validacao('--data-devolucao obrigatorio (YYYY-MM-DD ou DD/MM/YYYY)')
    try:
        dt = _parse_ocorrido_em(args.data_devolucao)
    except ValueError as e:
        return _resp_validacao(f'--data-devolucao: {e}')
    data_dev = dt.date() if dt else None
    if not data_dev:
        return _resp_validacao('--data-devolucao vazia/invalida')

    # chassis: --chassi (1) OU --chassis-json (lista)
    if args.chassis_json:
        try:
            raw = json.loads(args.chassis_json)
        except json.JSONDecodeError as e:
            return _resp_validacao(f'--chassis-json invalido (use ["X","Y"]): {e}')
        if not isinstance(raw, list):
            return _resp_validacao('--chassis-json deve ser uma lista JSON')
        chassis = [str(c).strip().upper() for c in raw if str(c).strip()]
    elif args.chassi:
        chassis = [args.chassi.strip().upper()]
    else:
        chassis = []
    if not chassis:
        return _resp_validacao('informe ao menos 1 chassi via --chassi ou --chassis-json')

    if not args.confirmar:
        return _resp_dry_run(
            'registrar-devolucao-nfd', nf_id=args.nf_id,
            numero_nfd=args.numero_nfd.strip(), data_devolucao=str(data_dev),
            motivo=args.motivo.strip(), n_chassis=len(chassis), chassis=chassis,
            acao_pretendida='criar AssaiDevolucaoNfd + por chassi: evento PENDENTE '
                            '(volta ao estoque) + AssaiNfQpaItem.devolvido=True + '
                            'saldo do modelo retorna; NF original NAO e cancelada')
    try:
        from app.motos_assai.services.devolucao_service import (
            criar_devolucao, DevolucaoValidationError,
        )
        dev = criar_devolucao(
            nf_id=args.nf_id, numero_nfd=args.numero_nfd.strip(),
            data_devolucao=data_dev, motivo=args.motivo.strip(),
            chassis=chassis, anexos=None, operador_id=args.user_id,
        )
        return _resp_ok('registrar-devolucao-nfd', {
            'devolucao_id': dev.id, 'numero_nfd': dev.numero_nfd,
            'nf_id': args.nf_id, 'n_chassis': len(chassis), 'chassis': chassis})
    except DevolucaoValidationError as e:
        db.session.rollback()
        return _resp_validacao(str(e))


# ----- Roteador -------------------------------------------------------------

COMANDOS = [
    ('criar_moto', _cmd_criar_moto),
    ('definir_estado', _cmd_definir_estado),
    ('cadastrar_loja', _cmd_cadastrar_loja),
    ('corrigir_loja', _cmd_corrigir_loja),
    ('cadastrar_modelo', _cmd_cadastrar_modelo),
    ('corrigir_modelo', _cmd_corrigir_modelo),
    ('item_pedido', _cmd_item_pedido),
    ('planilha_estado', _cmd_planilha_estado),
    ('importar_nf', _cmd_importar_nf),
    ('corrigir_chassi_nf', _cmd_corrigir_chassi_nf),
    ('cancelar_nf', _cmd_cancelar_nf),
    ('vincular_nf', _cmd_vincular_nf),
    ('registrar_nf_manual', _cmd_registrar_nf_manual),
    ('registrar_devolucao_nfd', _cmd_registrar_devolucao_nfd),
]


def _selecionar_comando(args):
    selecionados = [(n, fn) for n, fn in COMANDOS if getattr(args, n, False)]
    if len(selecionados) == 0:
        return None, 'Nenhum modo especificado. Use --help para ver as opcoes.'
    if len(selecionados) > 1:
        nomes = ', '.join(n for n, _ in selecionados)
        return None, f'Apenas um modo por vez. Selecionados: {nomes}'
    return selecionados[0], None


def main():
    p = argparse.ArgumentParser(
        prog='corrigindo_dados_assai',
        description='Skill WRITE de backfill / correcao manual do modulo Motos Assai',
    )
    # Modos (exclusivos)
    p.add_argument('--criar-moto', action='store_true', dest='criar_moto')
    p.add_argument('--definir-estado', action='store_true', dest='definir_estado')
    p.add_argument('--cadastrar-loja', action='store_true', dest='cadastrar_loja')
    p.add_argument('--corrigir-loja', action='store_true', dest='corrigir_loja')
    p.add_argument('--cadastrar-modelo', action='store_true', dest='cadastrar_modelo')
    p.add_argument('--corrigir-modelo', action='store_true', dest='corrigir_modelo')
    p.add_argument('--item-pedido', action='store_true', dest='item_pedido')
    p.add_argument('--planilha-estado', action='store_true', dest='planilha_estado')
    p.add_argument('--importar-nf', action='store_true', dest='importar_nf')
    p.add_argument('--corrigir-chassi-nf', action='store_true', dest='corrigir_chassi_nf')
    p.add_argument('--cancelar-nf', action='store_true', dest='cancelar_nf')
    p.add_argument('--vincular-nf', action='store_true', dest='vincular_nf')
    p.add_argument('--registrar-nf-manual', action='store_true', dest='registrar_nf_manual')
    p.add_argument('--registrar-devolucao-nfd', action='store_true',
                   dest='registrar_devolucao_nfd')

    # Args de dados
    p.add_argument('--chassi')
    p.add_argument('--modelo', help='Texto do modelo (resolvido via alias)')
    p.add_argument('--modelo-id', type=int, dest='modelo_id', default=None)
    p.add_argument('--cor', default=None)
    p.add_argument('--motor', default=None)
    p.add_argument('--ano', type=int, default=None)
    p.add_argument('--estado', help='Estado-alvo (ESTOQUE/MONTADA/PENDENTE/'
                                    'DISPONIVEL/DEMONSTRACAO)')
    p.add_argument('--ocorrido-em', dest='ocorrido_em', default=None,
                   help='Data retroativa do evento (YYYY-MM-DD ou DD/MM/YYYY)')
    p.add_argument('--motivo', default=None,
                   help='Justificativa (>=3 chars) p/ PENDENTE/REVERTIDA')
    p.add_argument('--campos', default=None, help='JSON de campos (loja/modelo)')
    p.add_argument('--loja-id', type=int, dest='loja_id', default=None)
    p.add_argument('--acao', default=None, help='item-pedido: add | edit | remove')
    p.add_argument('--pedido-id', type=int, dest='pedido_id', default=None)
    p.add_argument('--item-id', type=int, dest='item_id', default=None)
    p.add_argument('--qtd', type=int, default=None)
    p.add_argument('--valor', type=float, default=None)

    # NF Q.P.A. (faturamento / correcao de chassi em NF)
    p.add_argument('--pdf', default=None, help='Caminho do PDF da NF (importar-nf)')
    p.add_argument('--nf-id', type=int, dest='nf_id', default=None)
    p.add_argument('--de-chassi', dest='de_chassi', default=None,
                   help='Chassi antigo (corrigir-chassi-nf)')
    p.add_argument('--para-chassi', dest='para_chassi', default=None,
                   help='Chassi novo (corrigir-chassi-nf)')
    p.add_argument('--pares-json', dest='pares_json', default=None,
                   help='Multiplos pares: [["ANTIGO","NOVO"],...]')
    p.add_argument('--numero-cce', dest='numero_cce', default=None,
                   help='Identificador da correcao (corrigir-chassi-nf)')
    p.add_argument('--nf-json', dest='nf_json', default=None,
                   help='JSON da NF Q.P.A. p/ registro manual sem PDF '
                        '(chave_44, numero, loja_id/destinatario_cnpj, itens[])')
    p.add_argument('--numero-nfd', dest='numero_nfd', default=None,
                   help='Numero da NF de devolucao (registrar-devolucao-nfd)')
    p.add_argument('--data-devolucao', dest='data_devolucao', default=None,
                   help='Data da NFd: YYYY-MM-DD ou DD/MM/YYYY (registrar-devolucao-nfd)')
    p.add_argument('--chassis-json', dest='chassis_json', default=None,
                   help='Lista de chassis a devolver: ["X","Y"] (registrar-devolucao-nfd; '
                        'ou use --chassi para 1)')

    # Planilha
    p.add_argument('--excel', default=None, help='Caminho do .xlsx')
    p.add_argument('--aba', default=None, help='Nome da aba (default: ativa)')
    p.add_argument('--coluna-chassi', dest='coluna_chassi', default='CHASSI')
    p.add_argument('--coluna-status', dest='coluna_status', default='STATUS')
    p.add_argument('--coluna-modelo', dest='coluna_modelo', default='MODELO')
    p.add_argument('--coluna-cor', dest='coluna_cor', default='COR')
    p.add_argument('--coluna-data', dest='coluna_data', default='DATA DE CHEGADA')
    p.add_argument('--criar-faltantes', action='store_true', dest='criar_faltantes',
                   help='Cria a moto (+ESTOQUE) se o chassi nao existir')

    # Auth e confirmacao
    p.add_argument('--user-id', type=int, required=True, dest='user_id',
                   help='ID do operador (OBRIGATORIO)')
    p.add_argument('--confirmar', action='store_true',
                   help='Efetiva a operacao (sem isso = dry-run)')

    args = p.parse_args()

    selecionado, erro_sel = _selecionar_comando(args)
    if erro_sel:
        out = {'ok': False, 'erro': erro_sel, 'tipo_erro': 'comando_invalido',
               'exit_code': 1}
        print(json.dumps(out, ensure_ascii=False, indent=2))
        return 1
    _, fn_cmd = selecionado

    try:
        with contextlib.redirect_stdout(io.StringIO()):
            app = create_app()
        with app.app_context():
            ok_auth, motivo_auth = _verificar_autorizacao(args.user_id)
            if not ok_auth:
                result = _resp_autorizacao(motivo_auth, args.user_id)
            else:
                try:
                    result = fn_cmd(args)
                except Exception:
                    db.session.rollback()
                    raise
        print(json.dumps(result, default=_json_default, ensure_ascii=False, indent=2))
        return result.get('exit_code', 0)
    except Exception as e:
        err = {'ok': False, 'erro': str(e), 'tipo_erro': 'infra', 'exit_code': 2}
        print(json.dumps(err, ensure_ascii=False), file=sys.stderr)
        print(json.dumps(err, default=_json_default, ensure_ascii=False, indent=2))
        return 2


if __name__ == '__main__':
    sys.exit(main())
