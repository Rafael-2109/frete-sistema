"""
Gera planilha Excel com distâncias e pedágios de Santana de Parnaíba
para todas as cidades de SP usando Google Routes API v2 com TOLLS.

Uso:
    source .venv/bin/activate
    python scripts/gerar_distancias_pedagio_sp.py [--fallback] [--limite N] [--output CAMINHO]

Flags:
    --fallback   Usar Distance Matrix + estimativa em vez de Routes API v2
    --limite N   Processar apenas N cidades (para teste)
    --output     Caminho do arquivo Excel de saída (default: /tmp/distancias_pedagio_sp.xlsx)
"""

import argparse
import os
import sys
import time
from datetime import datetime

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Setup do app Flask para acesso ao banco
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
load_dotenv()

from app import create_app
from app.localidades.models import Cidade

# ─── Constantes ───────────────────────────────────────────────────────────────

ORIGEM_LAT = -23.4430
ORIGEM_LNG = -46.8491
ORIGEM_NOME = "Santana de Parnaíba"

EIXOS = [2, 3, 4, 5, 6, 7, 9]

# Estimativa fallback (baseada em mapa_service.py:762-813)
PERCENTUAL_PEDAGIADO = 0.6
KM_POR_PRACA = 45
VALOR_BASE_PRACA = 8.50  # R$ Cat 1 (carro) média SP

ROUTES_API_URL = "https://routes.googleapis.com/directions/v2:computeRoutes"
DISTANCE_MATRIX_URL = "https://maps.googleapis.com/maps/api/distancematrix/json"

RATE_LIMIT_MS = 200  # ms entre chamadas Routes API


# ─── Funções de API ───────────────────────────────────────────────────────────

def chamar_routes_api(api_key: str, cidade_nome: str) -> dict:
    """
    Chama Google Routes API v2 com extraComputations TOLLS.
    Retorna dict com distancia_km, duracao_min, pedagio_base, num_pracas, fonte.
    """
    headers = {
        "X-Goog-Api-Key": api_key,
        "X-Goog-FieldMask": (
            "routes.distanceMeters,"
            "routes.duration,"
            "routes.travelAdvisory.tollInfo"
        ),
        "Content-Type": "application/json",
    }
    body = {
        "origin": {
            "location": {
                "latLng": {"latitude": ORIGEM_LAT, "longitude": ORIGEM_LNG}
            }
        },
        "destination": {"address": f"{cidade_nome}, SP, Brazil"},
        "travelMode": "DRIVE",
        "extraComputations": ["TOLLS"],
        "routeModifiers": {
            "vehicleInfo": {"emissionType": "DIESEL"}
        },
    }

    resp = requests.post(ROUTES_API_URL, headers=headers, json=body, timeout=30)
    resp.raise_for_status()
    data = resp.json()

    if not data.get("routes"):
        return {"erro": "Sem rota encontrada"}

    rota = data["routes"][0]
    distancia_m = rota.get("distanceMeters", 0)
    duracao_s = rota.get("duration", "0s")

    # duration vem como "1234s"
    if isinstance(duracao_s, str):
        duracao_s = int(duracao_s.replace("s", ""))

    distancia_km = round(distancia_m / 1000, 1)
    duracao_min = round(duracao_s / 60, 0)

    # Extrair pedágio
    pedagio_base = 0.0
    num_pracas = 0
    fonte = "API"

    toll_info = (
        rota.get("travelAdvisory", {}).get("tollInfo", {})
    )
    if toll_info:
        tolls = toll_info.get("estimatedPrice", [])
        for toll in tolls:
            valor = float(toll.get("units", "0")) + float(
                toll.get("nanos", 0)
            ) / 1e9
            pedagio_base += valor
        pedagio_base = round(pedagio_base, 2)
        # Número de praças não vem direto da API, estimar pela distância
        if distancia_km > 0:
            dist_ped = distancia_km * PERCENTUAL_PEDAGIADO
            num_pracas = max(1, int(dist_ped / KM_POR_PRACA)) if pedagio_base > 0 else 0
    else:
        # Sem toll info — usar estimativa
        pedagio_base, num_pracas = _estimar_pedagio(distancia_km)
        fonte = "Estimado"

    return {
        "distancia_km": distancia_km,
        "duracao_min": int(duracao_min),
        "pedagio_base": pedagio_base,
        "num_pracas": num_pracas,
        "fonte": fonte,
    }


def chamar_distance_matrix_batch(api_key: str, cidades: list) -> list:
    """
    Chama Distance Matrix API em batches de 25 destinos.
    Retorna lista de dicts com distancia_km, duracao_min para cada cidade.
    """
    resultados = []

    for i in range(0, len(cidades), 25):
        batch = cidades[i : i + 25]
        destinos = "|".join(f"{c.nome}, SP, Brazil" for c in batch)

        params = {
            "origins": f"{ORIGEM_LAT},{ORIGEM_LNG}",
            "destinations": destinos,
            "key": api_key,
            "language": "pt-BR",
        }

        resp = requests.get(DISTANCE_MATRIX_URL, params=params, timeout=30)
        resp.raise_for_status()
        data = resp.json()

        if data.get("status") != "OK":
            for c in batch:
                resultados.append({"cidade": c, "erro": data.get("status")})
            continue

        elementos = data.get("rows", [{}])[0].get("elements", [])
        for j, elem in enumerate(elementos):
            cidade = batch[j]
            if elem.get("status") != "OK":
                resultados.append({"cidade": cidade, "erro": elem.get("status")})
                continue

            dist_m = elem.get("distance", {}).get("value", 0)
            dur_s = elem.get("duration", {}).get("value", 0)
            dist_km = round(dist_m / 1000, 1)
            dur_min = round(dur_s / 60, 0)

            ped_base, n_pracas = _estimar_pedagio(dist_km)

            resultados.append({
                "cidade": cidade,
                "distancia_km": dist_km,
                "duracao_min": int(dur_min),
                "pedagio_base": ped_base,
                "num_pracas": n_pracas,
                "fonte": "Estimado",
            })

        if i + 25 < len(cidades):
            time.sleep(0.2)

    return resultados


def _estimar_pedagio(distancia_km: float) -> tuple:
    """
    Estima pedágio baseado na distância (fallback).
    Retorna (valor_base, num_pracas).
    FONTE: app/carteira/services/mapa_service.py:762-813
    """
    if distancia_km <= 0:
        return (0.0, 0)

    dist_pedagiada = distancia_km * PERCENTUAL_PEDAGIADO
    num_pracas = max(1, int(dist_pedagiada / KM_POR_PRACA))
    valor_base = round(num_pracas * VALOR_BASE_PRACA, 2)
    return (valor_base, num_pracas)


# ─── Teste de conectividade ──────────────────────────────────────────────────

def testar_routes_api(api_key: str) -> bool:
    """Testa Routes API v2 com Campinas. Retorna True se OK."""
    print("Testando Routes API v2 com TOLLS (Santana → Campinas)...")
    try:
        resultado = chamar_routes_api(api_key, "Campinas")
        if "erro" in resultado:
            print(f"  FALHA: {resultado['erro']}")
            return False
        print(
            f"  OK: {resultado['distancia_km']} km, "
            f"R$ {resultado['pedagio_base']:.2f} pedágio base, "
            f"fonte: {resultado['fonte']}"
        )
        return True
    except Exception as e:
        print(f"  ERRO: {e}")
        return False


# ─── Geração do Excel ─────────────────────────────────────────────────────────

def gerar_excel(dados: list, caminho: str):
    """Gera arquivo Excel formatado com os dados de distância e pedágio."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Distâncias e Pedágios SP"

    # Header
    colunas = [
        "Cidade",
        "Código IBGE",
        "Microrregião",
        "Mesorregião",
        "Distância (km)",
        "Tempo (min)",
        "Nº Praças Est.",
    ]
    for eixo in EIXOS:
        colunas.append(f"Pedágio {eixo} Eixos")
    colunas.append("Fonte Pedágio")

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    erro_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Escrever header
    for col_idx, titulo in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Escrever dados
    for row_idx, item in enumerate(dados, 2):
        ws.cell(row=row_idx, column=1, value=item.get("cidade_nome", ""))
        ws.cell(row=row_idx, column=2, value=item.get("codigo_ibge", ""))
        ws.cell(row=row_idx, column=3, value=item.get("microrregiao", ""))
        ws.cell(row=row_idx, column=4, value=item.get("mesorregiao", ""))

        if item.get("erro"):
            ws.cell(row=row_idx, column=5, value="ERRO")
            ws.cell(row=row_idx, column=5).fill = erro_fill
            ws.cell(row=row_idx, column=6, value=item["erro"])
            ws.cell(row=row_idx, column=len(colunas), value="ERRO")
            continue

        dist = item.get("distancia_km", 0)
        tempo = item.get("duracao_min", 0)
        ped_base = item.get("pedagio_base", 0)
        n_pracas = item.get("num_pracas", 0)

        ws.cell(row=row_idx, column=5, value=dist)
        ws.cell(row=row_idx, column=6, value=tempo)
        ws.cell(row=row_idx, column=7, value=n_pracas)

        # Pedágio por eixo: base * N eixos
        for i, eixo in enumerate(EIXOS):
            valor = round(ped_base * eixo, 2)
            cell = ws.cell(row=row_idx, column=8 + i, value=valor)
            cell.number_format = '#,##0.00'

        ws.cell(row=row_idx, column=8 + len(EIXOS), value=item.get("fonte", ""))

    # Formatar colunas numéricas
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.0'

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'

    # Auto-width
    for col_idx in range(1, len(colunas) + 1):
        letra = get_column_letter(col_idx)
        max_len = len(str(colunas[col_idx - 1]))
        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 20), min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letra].width = min(max_len + 3, 25)

    # Congelar header
    ws.freeze_panes = "A2"

    # Aba de notas
    ws_notas = wb.create_sheet("Notas")
    notas = [
        ["Notas sobre os dados"],
        [""],
        [f"Origem: {ORIGEM_NOME} (CD Nacom Goya) — Lat: {ORIGEM_LAT}, Lng: {ORIGEM_LNG}"],
        [f"Data de geração: {datetime.now().strftime('%d/%m/%Y %H:%M')}"],
        [""],
        ["Pedágio por eixo:"],
        ["  - Pedágio N Eixos = Pedágio Base (Cat 1) × N"],
        ["  - Cat 1 = carro (2 eixos rodagem simples) = valor base da praça"],
        ["  - Para veículos comerciais de 2 eixos com rodagem dupla, o valor é 2× base"],
        ["  - Regra ARTESP/ANTT para veículos comerciais: N × valor base por praça"],
        [""],
        ["Fonte 'API': Pedágio real obtido via Google Routes API v2 com TOLLS"],
        ["Fonte 'Estimado': Estimativa baseada em 60% da rota pedagiada, 1 praça a cada 45km, R$ 8,50/praça Cat 1"],
        [""],
        ["Configurações de eixos padrão da frota:"],
        ["  2 eixos — Toco / VUC"],
        ["  3 eixos — Truck"],
        ["  4 eixos — Truck com reboque simples"],
        ["  5 eixos — Carreta 2 eixos"],
        ["  6 eixos — Carreta 3 eixos"],
        ["  7 eixos — Bi-trem / Rodotrem"],
        ["  9 eixos — Tri-trem / Rodotrem especial"],
    ]
    for row_idx, linha in enumerate(notas, 1):
        for col_idx, valor in enumerate(linha, 1):
            cell = ws_notas.cell(row=row_idx, column=col_idx, value=valor)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
    ws_notas.column_dimensions["A"].width = 90

    wb.save(caminho)
    print(f"\nExcel salvo em: {caminho}")
    print(f"Total de linhas: {len(dados)}")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Gerar Excel de distâncias e pedágios SP")
    parser.add_argument("--fallback", action="store_true", help="Usar Distance Matrix + estimativa")
    parser.add_argument("--limite", type=int, default=0, help="Processar apenas N cidades (teste)")
    parser.add_argument("--output", default="/tmp/distancias_pedagio_sp.xlsx", help="Caminho de saída")
    args = parser.parse_args()

    api_key = os.environ.get("GOOGLE_MAPS_API_KEY")
    if not api_key:
        print("ERRO: GOOGLE_MAPS_API_KEY não encontrada no .env")
        sys.exit(1)

    app = create_app()
    with app.app_context():
        # Buscar cidades SP
        query = Cidade.query.filter_by(uf="SP").order_by(Cidade.nome)
        if args.limite > 0:
            query = query.limit(args.limite)
        cidades = query.all()
        print(f"Cidades SP encontradas: {len(cidades)}")

        if not cidades:
            print("ERRO: Nenhuma cidade SP no banco")
            sys.exit(1)

        usar_routes = not args.fallback

        # Teste de conectividade
        if usar_routes:
            routes_ok = testar_routes_api(api_key)
            if not routes_ok:
                print("Routes API indisponível. Usando fallback (Distance Matrix + estimativa).")
                usar_routes = False

        dados = []

        if usar_routes:
            # ─── Routes API v2 (cidade a cidade) ──────────────────────
            total = len(cidades)
            erros = 0
            t0 = time.time()

            for i, cidade in enumerate(cidades):
                try:
                    resultado = chamar_routes_api(api_key, cidade.nome)

                    if "erro" in resultado:
                        dados.append({
                            "cidade_nome": cidade.nome,
                            "codigo_ibge": cidade.codigo_ibge,
                            "microrregiao": cidade.microrregiao or "",
                            "mesorregiao": cidade.mesorregiao or "",
                            "erro": resultado["erro"],
                        })
                        erros += 1
                    else:
                        dados.append({
                            "cidade_nome": cidade.nome,
                            "codigo_ibge": cidade.codigo_ibge,
                            "microrregiao": cidade.microrregiao or "",
                            "mesorregiao": cidade.mesorregiao or "",
                            **resultado,
                        })
                except Exception as e:
                    dados.append({
                        "cidade_nome": cidade.nome,
                        "codigo_ibge": cidade.codigo_ibge,
                        "microrregiao": cidade.microrregiao or "",
                        "mesorregiao": cidade.mesorregiao or "",
                        "erro": str(e),
                    })
                    erros += 1

                # Progresso a cada 50 cidades
                if (i + 1) % 50 == 0 or (i + 1) == total:
                    elapsed = time.time() - t0
                    rate = (i + 1) / elapsed if elapsed > 0 else 0
                    eta = (total - i - 1) / rate if rate > 0 else 0
                    print(
                        f"  [{i + 1}/{total}] "
                        f"{cidade.nome} — "
                        f"{erros} erros — "
                        f"ETA: {int(eta)}s"
                    )

                # Rate limiting
                time.sleep(RATE_LIMIT_MS / 1000)

        else:
            # ─── Distance Matrix batch + estimativa ───────────────────
            print("Usando Distance Matrix API (batch de 25)...")
            resultados_batch = chamar_distance_matrix_batch(api_key, cidades)

            for item in resultados_batch:
                cidade = item.pop("cidade", None)
                if cidade:
                    item["cidade_nome"] = cidade.nome
                    item["codigo_ibge"] = cidade.codigo_ibge
                    item["microrregiao"] = cidade.microrregiao or ""
                    item["mesorregiao"] = cidade.mesorregiao or ""
                dados.append(item)

        # Gerar Excel
        gerar_excel(dados, args.output)

        # Resumo
        ok = sum(1 for d in dados if not d.get("erro"))
        err = sum(1 for d in dados if d.get("erro"))
        api_count = sum(1 for d in dados if d.get("fonte") == "API")
        est_count = sum(1 for d in dados if d.get("fonte") == "Estimado")
        print(f"\nResumo: {ok} OK, {err} erros | {api_count} API, {est_count} estimados")


if __name__ == "__main__":
    main()
