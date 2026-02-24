#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Gap Analysis Odoo <-> Local -> Excel Multi-Aba para Aprovacao
============================================================

Gera Excel com 5 abas para revisao humana antes de importacao:
1. Resumo — totais por tipo e empresa
2. Faltantes Receber — IDs elegiveis no Odoo sem correspondencia local
3. Faltantes Pagar — idem para contas a pagar
4. Devolucoes detectadas — titulos com balance<=0 / credit<=0 no Odoo
5. Devolucoes no sistema — registros indevidos ja importados (NFs 93556/93557)

Estrategia resiliente (mesmos patterns de exportar_titulos_gap_odoo.py):
- search() no Odoo → so IDs (evita display_name crash)
- db.session.close() antes de queries locais (evita SSL drop)
- read() em batches com campos explicitos
- Partner info via read() separado em res.partner

Uso:
    # Com IDs de producao pre-exportados do Render via MCP
    python scripts/exportar_gap_odoo_excel.py --production /tmp

    # Com conexao direta ao banco (local ou producao)
    python scripts/exportar_gap_odoo_excel.py

    # Limitar para teste rapido
    python scripts/exportar_gap_odoo_excel.py --production /tmp --limite 50

Autor: Sistema de Fretes
Data: 2026-02-23
"""

import argparse
import json
import logging
import sys
import os
from datetime import datetime

# Setup path para imports do app
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from app import create_app, db
from app.utils.timezone import agora_utc_naive
from app.odoo.utils.connection import get_odoo_connection

logger = logging.getLogger(__name__)

# ===================================================================
# CONSTANTES
# ===================================================================

# Odoo company_ids elegiveis para sincronizacao
# FONTE: .claude/references/odoo/IDS_FIXOS.md:10-15
ODOO_COMPANY_IDS_ELEGIVEIS = [1, 3, 4]

# Mapeamento Odoo company_id → sigla (dados vindos do Odoo)
# FONTE: .claude/references/odoo/IDS_FIXOS.md:10-15
ODOO_COMPANY_SIGLA = {1: 'FB', 3: 'SC', 4: 'CD'}

# Mapeamento local empresa → sigla (dados vindos do banco local)
# FONTE: contas_a_receber.json:12 e contas_a_pagar.json:12 — "1=FB, 2=SC, 3=CD"
LOCAL_EMPRESA_SIGLA = {1: 'FB', 2: 'SC', 3: 'CD'}

# Mapeamento Odoo company_id → local empresa (cross-reference)
# FONTE: .claude/references/odoo/IDS_FIXOS.md
ODOO_TO_LOCAL_EMPRESA = {1: 1, 3: 2, 4: 3}

# CNPJs raiz do grupo Nacom (para detectar intercompany)
# FONTE: .claude/references/odoo/IDS_FIXOS.md:18-25
CNPJS_RAIZ_GRUPO = ['61.724.241', '18.467.441']

# Batch sizes
READ_BATCH = 500
PARTNER_BATCH = 500

# Campos para read() no account.move.line
# partner_id incluido com fallback (evita display_name crash)
CAMPOS_TITULO = [
    'id', 'name', 'ref',
    'x_studio_nf_e', 'l10n_br_cobranca_parcela',
    'company_id', 'move_id',
    'date', 'date_maturity',
    'credit', 'balance', 'amount_residual',
    'l10n_br_paga', 'reconciled',
    'account_type', 'parent_state',
    'write_date', 'create_date',
]

CAMPOS_TITULO_COM_PARTNER = CAMPOS_TITULO + ['partner_id']

# Campos para res.partner
CAMPOS_PARTNER = ['id', 'name', 'l10n_br_cnpj', 'l10n_br_razao_social']

# NFs de devolucoes ja importadas indevidamente no sistema
# FONTE: contas_a_pagar (Render) — NFs de credit note Nadir Figueiredo
NFS_DEVOLUCAO_NO_SISTEMA = ['93556', '93557']


def configurar_logging(verbose=False):
    """Configura logging para o script."""
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format='%(asctime)s [%(levelname)s] %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S',
    )


# ===================================================================
# HELPERS
# ===================================================================

def _extrair_company_id(record):
    """Extrai company_id numerico de um record Odoo."""
    company = record.get('company_id', [None, None])
    if isinstance(company, (list, tuple)):
        return company[0]
    return company


def _extrair_partner_id(record):
    """Extrai partner_id numerico de um record Odoo."""
    pid = record.get('partner_id')
    if pid and isinstance(pid, (list, tuple)):
        return pid[0]
    elif pid and isinstance(pid, int):
        return pid
    return None


def _is_intercompany(cnpj):
    """Verifica se o CNPJ pertence ao grupo Nacom (transacao intercompany)."""
    if not cnpj:
        return False
    return any(cnpj.startswith(raiz) for raiz in CNPJS_RAIZ_GRUPO)


def _normalizar_parcela_como_sync(parcela_odoo):
    """
    Normaliza parcela Odoo para formato local (mesma logica do sync).

    FONTE: sincronizacao_contas_receber_service.py:484
    FONTE: sincronizacao_contas_pagar_service.py:494
    Ambos usam: parcela_to_str(row.get('l10n_br_cobranca_parcela')) or '1'
    """
    from app.financeiro.parcela_utils import parcela_to_str
    return parcela_to_str(parcela_odoo) or '1'


# ===================================================================
# PHASE 1: CARREGAR IDS LOCAIS (PRODUCAO OU DB)
# ===================================================================

def carregar_ids_locais(tipo, production_dir=None):
    """
    Carrega odoo_line_ids do banco local ou de arquivo pre-exportado.

    Args:
        tipo: 'receber' ou 'pagar'
        production_dir: diretorio com render_{tipo}_ids.json

    Returns:
        set[int]: Conjunto de odoo_line_ids existentes
    """
    if production_dir:
        ids_file = os.path.join(production_dir, f'render_{tipo}_ids.json')
        logger.info(f"  Lendo IDs de producao: {ids_file}")
        with open(ids_file, 'r') as f:
            ids_list = json.load(f)
        ids_set = set(ids_list)
        logger.info(f"  {len(ids_set):,} IDs de PRODUCAO ({tipo.upper()})")
    else:
        from sqlalchemy import text
        tabela = 'contas_a_pagar' if tipo == 'pagar' else 'contas_a_receber'
        rows = db.session.execute(
            text(f"SELECT odoo_line_id FROM {tabela} WHERE odoo_line_id IS NOT NULL")
        ).fetchall()
        ids_set = {r[0] for r in rows}
        logger.info(f"  {len(ids_set):,} IDs locais ({tipo.upper()})")
        db.session.close()

    return ids_set


def carregar_chaves_negocio_locais(tipo, production_dir=None):
    """
    Carrega chaves de negocio (empresa, titulo_nf, parcela) do banco local ou
    de arquivo pre-exportado do Render.

    Usada na Phase 7 para cross-reference contra dados do Odoo,
    detectando falsos positivos (mesmo NF+empresa+parcela ja existe
    via outro odoo_line_id).

    IMPORTANTE: Em modo --production, o Flask conecta ao banco LOCAL (localhost),
    que pode ter dados diferentes do Render. Por isso, usar os arquivos
    render_{tipo}_keys.json pre-exportados via MCP.

    Args:
        tipo: 'receber' ou 'pagar'
        production_dir: diretorio com render_{tipo}_keys.json

    Returns:
        set[tuple]: {(empresa_int, titulo_nf_str, parcela_str), ...}
    """
    if production_dir:
        keys_file = os.path.join(production_dir, f'render_{tipo}_keys.json')
        logger.info(f"  Lendo chaves de producao: {keys_file}")
        with open(keys_file, 'r') as f:
            keys_list = json.load(f)
        keys_set = {(r[0], r[1], r[2]) for r in keys_list}
        logger.info(f"  {len(keys_set):,} chaves de PRODUCAO ({tipo.upper()})")
    else:
        from sqlalchemy import text
        tabela = 'contas_a_pagar' if tipo == 'pagar' else 'contas_a_receber'
        rows = db.session.execute(
            text(f"SELECT empresa, titulo_nf, parcela FROM {tabela}")
        ).fetchall()
        keys_set = {(r[0], r[1], r[2]) for r in rows}
        logger.info(f"  {len(keys_set):,} chaves locais ({tipo.upper()})")
        db.session.close()

    return keys_set


# ===================================================================
# PHASE 2: BUSCAR IDS ELEGIVEIS NO ODOO
# ===================================================================

def buscar_ids_odoo_elegiveis(connection, tipo):
    """
    Busca TODOS os IDs elegiveis no Odoo com filtros completos.

    Criterios de elegibilidade (do plano):
    - parent_state = 'posted'
    - company_id em [1, 3, 4]
    - x_studio_nf_e != False (preenchido)
    - date_maturity != False (preenchido)
    - balance > 0 (receber) / credit > 0 (pagar)

    Nota: intercompany filtrado depois (requer read de res.partner).

    Args:
        connection: OdooConnection autenticada
        tipo: 'receber' ou 'pagar'

    Returns:
        list[int]: Lista de IDs elegiveis
    """
    account_type = 'asset_receivable' if tipo == 'receber' else 'liability_payable'

    domain = [
        ['account_type', '=', account_type],
        ['parent_state', '=', 'posted'],
        ['company_id', 'in', ODOO_COMPANY_IDS_ELEGIVEIS],
        ['x_studio_nf_e', '!=', False],
        ['date_maturity', '!=', False],
    ]

    # Filtro de saldo (direction) — exclui devolucoes
    if tipo == 'receber':
        domain.append(['balance', '>', 0])
    else:
        domain.append(['credit', '>', 0])

    logger.info(f"  Domain: {domain}")

    all_ids = connection.execute_kw(
        'account.move.line', 'search', [domain],
        timeout_override=180,
    )

    logger.info(f"  {len(all_ids):,} IDs elegiveis no Odoo ({tipo.upper()})")
    return all_ids


# ===================================================================
# PHASE 3: BUSCAR IDS DE DEVOLUCOES NO ODOO
# ===================================================================

def buscar_ids_devolucoes_odoo(connection, tipo):
    """
    Busca IDs de devolucoes no Odoo (titulos com saldo invertido).

    Criterios: mesmos do elegivel, mas com saldo invertido:
    - Receber: balance < 0 (credit notes — cliente devolveu)
    - Pagar: balance > 0 (debit notes — fornecedor devolveu)
      NOTA: Odoo armazena credit >= 0 sempre, entao credit < 0 nao existe.
      Devolucoes de compra criam linhas payable com debit > 0 (balance > 0).

    Args:
        connection: OdooConnection autenticada
        tipo: 'receber' ou 'pagar'

    Returns:
        list[int]: Lista de IDs de devolucoes
    """
    account_type = 'asset_receivable' if tipo == 'receber' else 'liability_payable'

    domain = [
        ['account_type', '=', account_type],
        ['parent_state', '=', 'posted'],
        ['company_id', 'in', ODOO_COMPANY_IDS_ELEGIVEIS],
        ['x_studio_nf_e', '!=', False],
        ['date_maturity', '!=', False],
    ]

    # Saldo invertido = devolucao
    if tipo == 'receber':
        # Credit note: balance negativo na conta receivable
        domain.append(['balance', '<', 0])
    else:
        # Debit note / vendor refund: balance positivo na conta payable
        # (credit=0, debit>0 → balance>0)
        domain.append(['balance', '>', 0])

    logger.info(f"  Domain devolucoes: {domain}")

    all_ids = connection.execute_kw(
        'account.move.line', 'search', [domain],
        timeout_override=180,
    )

    logger.info(f"  {len(all_ids):,} devolucoes no Odoo ({tipo.upper()})")
    return all_ids


# ===================================================================
# PHASE 4: BUSCAR DADOS COMPLETOS
# ===================================================================

def buscar_dados_titulos(connection, ids_to_read, limite=None):
    """
    read() em batches para obter dados completos dos titulos.

    Tenta com partner_id primeiro, fallback sem ele.

    Args:
        connection: OdooConnection autenticada
        ids_to_read: lista de IDs
        limite: limite de IDs (para teste)

    Returns:
        tuple: (records, usou_partner)
    """
    if limite:
        ids_to_read = ids_to_read[:limite]

    total = len(ids_to_read)
    if total == 0:
        return [], True

    n_batches = (total + READ_BATCH - 1) // READ_BATCH
    logger.info(f"  read() de {total:,} titulos em {n_batches} batches...")

    usou_partner = True
    campos = CAMPOS_TITULO_COM_PARTNER

    all_records = []
    for i in range(0, total, READ_BATCH):
        batch_ids = ids_to_read[i:i + READ_BATCH]
        batch_num = i // READ_BATCH + 1

        try:
            records = connection.read('account.move.line', batch_ids, campos)
            all_records.extend(records)

            if batch_num % 10 == 0 or batch_num == n_batches:
                logger.info(f"  Batch {batch_num}/{n_batches}: total {len(all_records):,}")

        except Exception as e:
            if usou_partner and batch_num == 1:
                logger.warning(f"  read() com partner_id falhou: {e}")
                logger.info("  Tentando sem partner_id...")
                usou_partner = False
                campos = CAMPOS_TITULO

                try:
                    records = connection.read(
                        'account.move.line', batch_ids, campos,
                    )
                    all_records.extend(records)
                except Exception as e2:
                    logger.error(f"  Falha total no batch {batch_num}: {e2}")
                    continue
            else:
                logger.error(f"  Erro no batch {batch_num}: {e}")
                continue

    logger.info(f"  Total lidos: {len(all_records):,} de {total:,}")
    return all_records, usou_partner


def buscar_partners(connection, records):
    """
    Busca dados de parceiros (CNPJ, nome) via read() em res.partner.

    Args:
        connection: OdooConnection autenticada
        records: lista de dicts do Odoo

    Returns:
        dict: {partner_id: {id, name, l10n_br_cnpj, l10n_br_razao_social}}
    """
    partner_ids = set()
    for r in records:
        pid = _extrair_partner_id(r)
        if pid:
            partner_ids.add(pid)

    if not partner_ids:
        return {}

    logger.info(f"  {len(partner_ids):,} parceiros unicos para buscar")

    partner_map = {}
    partner_list = list(partner_ids)

    for i in range(0, len(partner_list), PARTNER_BATCH):
        batch = partner_list[i:i + PARTNER_BATCH]
        try:
            partners = connection.read('res.partner', batch, CAMPOS_PARTNER)
            for p in partners:
                partner_map[p['id']] = p
        except Exception as e:
            logger.warning(
                f"  Falha ao buscar partners batch "
                f"{i // PARTNER_BATCH + 1}: {e}"
            )

    logger.info(f"  {len(partner_map):,} parceiros obtidos")
    return partner_map


# ===================================================================
# PHASE 5: TRANSFORMAR E FILTRAR
# ===================================================================

def transformar_registros(records, partner_map, tipo):
    """
    Transforma records Odoo em dicts padronizados e filtra intercompany.

    Args:
        records: lista de dicts do Odoo
        partner_map: {partner_id: {...}}
        tipo: 'receber' ou 'pagar'

    Returns:
        tuple: (registros_filtrados, stats)
    """
    stats = {
        'total_input': len(records),
        'intercompany': 0,
    }

    resultados = []
    for r in records:
        company_id = _extrair_company_id(r)
        partner_id_num = _extrair_partner_id(r)
        p_info = partner_map.get(partner_id_num, {})
        partner_cnpj = p_info.get('l10n_br_cnpj') or ''

        # Filtrar intercompany
        if _is_intercompany(partner_cnpj):
            stats['intercompany'] += 1
            continue

        # Valor original depende do tipo
        if tipo == 'pagar':
            valor_original = float(r.get('credit', 0) or 0)
        else:
            valor_original = float(r.get('balance', 0) or 0)

        amount_residual = float(r.get('amount_residual', 0) or 0)

        # Move info
        move = r.get('move_id', [None, None])
        move_nome = (
            move[1] if isinstance(move, (list, tuple)) and len(move) > 1
            else str(move)
        )

        # Status
        paga = bool(r.get('l10n_br_paga'))
        reconciled = bool(r.get('reconciled'))
        if paga:
            status = 'PAGO'
        elif reconciled and abs(amount_residual) < 0.01:
            status = 'RECONCILIADO'
        elif abs(amount_residual) < 0.01:
            status = 'QUITADO'
        else:
            status = 'EM ABERTO'

        # Parcela: guardar raw para cross-reference, normalizar para display
        parcela_raw = r.get('l10n_br_cobranca_parcela')
        parcela_norm = _normalizar_parcela_como_sync(parcela_raw)

        resultados.append({
            'odoo_line_id': r.get('id'),
            'empresa': ODOO_COMPANY_SIGLA.get(company_id, '?'),
            'company_id': company_id,
            'nf': str(r.get('x_studio_nf_e') or '').strip(),
            'parcela': parcela_norm,
            'parcela_raw': parcela_raw,
            'parceiro_nome': p_info.get('name') or '',
            'parceiro_cnpj': partner_cnpj,
            'valor_original': round(abs(valor_original), 2),
            'valor_residual': round(abs(amount_residual), 2),
            'vencimento': r.get('date_maturity') or '',
            'emissao': r.get('date') or '',
            'status': status,
            'move': move_nome,
            'paga_l10n': paga,
            'reconciliado': reconciled,
        })

    logger.info(
        f"  {len(resultados):,} registros apos filtro "
        f"(intercompany removidos: {stats['intercompany']})"
    )

    return resultados, stats


# ===================================================================
# PHASE 6: BUSCAR DEVOLUCOES NO SISTEMA LOCAL
# ===================================================================

def buscar_devolucoes_no_sistema(production_dir=None):
    """
    Busca os registros de devolucao ja importados indevidamente no sistema.

    FONTE: NFs 93556 e 93557 em contas_a_pagar (Nadir Figueiredo)

    Args:
        production_dir: se fornecido, nao consulta DB (dados hardcoded do Render)

    Returns:
        list[dict]: registros de devolucao
    """
    # Dados ja conhecidos do Render (query previa)
    # FONTE: mcp__render__query_render_postgres — 2026-02-23
    registros = [
        {
            'id_local': 16447,
            'tipo': 'pagar',
            'empresa': 'FB',
            'titulo_nf': '93557',
            'parcela': '1',
            'odoo_line_id': 3135612,
            'cnpj': '61.067.161/0018-35',
            'raz_social': 'NADIR FIGUEIREDO S.A.',
            'valor_original': 0.0,
            'valor_residual': 0.0,
            'vencimento': '2026-02-23',
            'emissao': '2026-02-23',
            'parcela_paga': True,
            'reconciliado': True,
            'status_sistema': 'PAGO',
            'metodo_baixa': 'ODOO_DIRETO',
            'observacao': 'Devolucao importada indevidamente (credit note)',
        },
        {
            'id_local': 16448,
            'tipo': 'pagar',
            'empresa': 'FB',
            'titulo_nf': '93556',
            'parcela': '1',
            'odoo_line_id': 3135585,
            'cnpj': '61.067.161/0018-35',
            'raz_social': 'NADIR FIGUEIREDO S.A.',
            'valor_original': 0.0,
            'valor_residual': 837.16,
            'vencimento': '2026-02-23',
            'emissao': '2026-02-23',
            'parcela_paga': True,
            'reconciliado': False,
            'status_sistema': 'PAGO',
            'metodo_baixa': 'ODOO_DIRETO',
            'observacao': 'Devolucao importada indevidamente (credit note)',
        },
    ]

    if not production_dir:
        # Tentar buscar do banco real para dados mais atualizados
        try:
            from sqlalchemy import text
            rows = db.session.execute(
                text("""
                    SELECT id, empresa, titulo_nf, parcela, odoo_line_id,
                           cnpj, raz_social, valor_original, valor_residual,
                           vencimento, emissao, parcela_paga, reconciliado,
                           status_sistema, metodo_baixa
                    FROM contas_a_pagar
                    WHERE titulo_nf IN ('93556', '93557')
                    ORDER BY titulo_nf
                """)
            ).fetchall()

            if rows:
                registros = []
                for row in rows:
                    registros.append({
                        'id_local': row[0],
                        'tipo': 'pagar',
                        'empresa': LOCAL_EMPRESA_SIGLA.get(row[1], '?'),
                        'titulo_nf': row[2],
                        'parcela': row[3],
                        'odoo_line_id': row[4],
                        'cnpj': row[5],
                        'raz_social': row[6],
                        'valor_original': float(row[7] or 0),
                        'valor_residual': float(row[8] or 0),
                        'vencimento': str(row[9] or ''),
                        'emissao': str(row[10] or ''),
                        'parcela_paga': bool(row[11]),
                        'reconciliado': bool(row[12]),
                        'status_sistema': row[13] or '',
                        'metodo_baixa': row[14] or '',
                        'observacao': 'Devolucao importada indevidamente',
                    })
            db.session.close()
        except Exception as e:
            logger.warning(f"  Falha ao buscar devolucoes do DB: {e}. Usando dados hardcoded.")

    return registros


# ===================================================================
# PHASE 7: GERAR EXCEL COM OPENPYXL
# ===================================================================

def gerar_excel(
    faltantes_receber,
    faltantes_pagar,
    devolucoes_odoo_receber,
    devolucoes_odoo_pagar,
    devolucoes_no_sistema,
    resumo_data,
    output_path,
):
    """
    Gera Excel com 5 abas usando openpyxl.

    Args:
        faltantes_receber: lista de dicts
        faltantes_pagar: lista de dicts
        devolucoes_odoo_receber: lista de dicts
        devolucoes_odoo_pagar: lista de dicts
        devolucoes_no_sistema: lista de dicts
        resumo_data: dict com totais
        output_path: caminho do arquivo Excel
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_fill_green = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
    header_fill_red = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    header_fill_orange = PatternFill(start_color='BF8F00', end_color='BF8F00', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    number_format_br = '#,##0.00'
    date_format = 'DD/MM/YYYY'

    def _apply_header_style(ws, row, fill=None):
        """Aplica estilo ao header de uma planilha."""
        fill = fill or header_fill
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def _auto_width(ws, min_width=10, max_width=40):
        """Ajusta largura das colunas automaticamente."""
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_length = 0
            for cell in col_cells:
                try:
                    length = len(str(cell.value or ''))
                    if length > max_length:
                        max_length = length
                except Exception:
                    pass
            adjusted = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

    def _write_data_rows(ws, start_row, data, columns, formats=None):
        """Escreve dados nas linhas, com formatacao opcional."""
        formats = formats or {}
        for row_idx, item in enumerate(data, start_row):
            for col_idx, col_key in enumerate(columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=item.get(col_key, ''))
                cell.border = thin_border

                fmt = formats.get(col_key)
                if fmt == 'number':
                    cell.number_format = number_format_br
                elif fmt == 'date':
                    # Tentar converter string de data
                    val = item.get(col_key, '')
                    if val and isinstance(val, str) and len(val) >= 10:
                        try:
                            dt = datetime.strptime(val[:10], '%Y-%m-%d')
                            cell.value = dt
                            cell.number_format = date_format
                        except ValueError:
                            pass

    # ---------------------------------------------------------------
    # ABA 1: RESUMO
    # ---------------------------------------------------------------
    ws_resumo = wb.active
    ws_resumo.title = 'Resumo'

    # Titulo
    ws_resumo.merge_cells('A1:H1')
    title_cell = ws_resumo['A1']
    title_cell.value = 'Gap Analysis Odoo <-> Local — Resumo'
    title_cell.font = Font(bold=True, size=14, color='2F5496')
    title_cell.alignment = Alignment(horizontal='center')

    ws_resumo.merge_cells('A2:H2')
    ws_resumo['A2'].value = f"Gerado em: {agora_utc_naive().strftime('%d/%m/%Y %H:%M:%S')}"
    ws_resumo['A2'].font = Font(italic=True, color='666666')
    ws_resumo['A2'].alignment = Alignment(horizontal='center')

    # Tabela de resumo por tipo
    row = 4
    headers_resumo = [
        'Tipo', 'Odoo Elegiveis', 'Local (com ID)',
        'Faltantes (bruto)', 'Falsos Positivos', 'Faltantes (final)',
        'Intercompany', '% Cobertura',
    ]
    for col_idx, h in enumerate(headers_resumo, 1):
        cell = ws_resumo.cell(row=row, column=col_idx, value=h)
    _apply_header_style(ws_resumo, row)

    for tipo_data in resumo_data.get('por_tipo', []):
        row += 1
        ws_resumo.cell(row=row, column=1, value=tipo_data['tipo']).border = thin_border
        ws_resumo.cell(row=row, column=2, value=tipo_data['odoo_elegiveis']).border = thin_border
        ws_resumo.cell(row=row, column=3, value=tipo_data['local_com_id']).border = thin_border
        ws_resumo.cell(row=row, column=4, value=tipo_data['faltantes_bruto']).border = thin_border
        ws_resumo.cell(row=row, column=5, value=tipo_data['falsos_positivos']).border = thin_border
        ws_resumo.cell(row=row, column=6, value=tipo_data['faltantes']).border = thin_border
        ws_resumo.cell(row=row, column=7, value=tipo_data['intercompany']).border = thin_border
        cell_pct = ws_resumo.cell(row=row, column=8, value=tipo_data['cobertura_pct'])
        cell_pct.number_format = '0.0%'
        cell_pct.border = thin_border

    # Totais por empresa
    row += 2
    ws_resumo.cell(row=row, column=1, value='Faltantes por Empresa:').font = Font(bold=True, size=11)

    row += 1
    headers_empresa = ['Empresa', 'Faltantes Receber', 'Faltantes Pagar', 'Total']
    for col_idx, h in enumerate(headers_empresa, 1):
        ws_resumo.cell(row=row, column=col_idx, value=h)
    _apply_header_style(ws_resumo, row, fill=header_fill_green)

    for emp_data in resumo_data.get('por_empresa', []):
        row += 1
        ws_resumo.cell(row=row, column=1, value=emp_data['empresa']).border = thin_border
        ws_resumo.cell(row=row, column=2, value=emp_data['receber']).border = thin_border
        ws_resumo.cell(row=row, column=3, value=emp_data['pagar']).border = thin_border
        ws_resumo.cell(row=row, column=4, value=emp_data['total']).border = thin_border

    # Devolucoes resumo
    row += 2
    ws_resumo.cell(row=row, column=1, value='Devolucoes:').font = Font(bold=True, size=11)
    row += 1
    ws_resumo.cell(
        row=row, column=1,
        value=f"Detectadas no Odoo (nao devem ser importadas): "
              f"{resumo_data.get('devolucoes_odoo', 0)}"
    )
    row += 1
    ws_resumo.cell(
        row=row, column=1,
        value=f"Ja no sistema (importadas indevidamente): "
              f"{resumo_data.get('devolucoes_sistema', 0)}"
    )

    # Cross-reference (falsos positivos)
    row += 2
    ws_resumo.cell(
        row=row, column=1,
        value='Cross-reference (falsos positivos removidos):'
    ).font = Font(bold=True, size=11)
    row += 1
    ws_resumo.cell(
        row=row, column=1,
        value='Titulos onde NF+empresa+parcela ja existe no local via outro odoo_line_id:'
    ).font = Font(italic=True, color='666666')
    for td in resumo_data.get('por_tipo', []):
        row += 1
        ws_resumo.cell(
            row=row, column=1,
            value=f"  {td['tipo']}: {td['falsos_positivos']:,} falsos positivos removidos "
                  f"(de {td['faltantes_bruto']:,} bruto → {td['faltantes']:,} final)"
        )

    _auto_width(ws_resumo)

    # ---------------------------------------------------------------
    # ABA 2: FALTANTES RECEBER
    # ---------------------------------------------------------------
    ws_receber = wb.create_sheet('Faltantes Receber')

    columns_faltantes = [
        'odoo_line_id', 'empresa', 'nf', 'parcela', 'parceiro_nome',
        'parceiro_cnpj', 'valor_original', 'vencimento', 'emissao', 'status',
    ]
    headers_faltantes = [
        'Odoo Line ID', 'Empresa', 'NF-e', 'Parcela', 'Cliente',
        'CNPJ', 'Valor Original', 'Vencimento', 'Emissao', 'Status',
    ]
    formats_faltantes = {
        'valor_original': 'number',
        'vencimento': 'date',
        'emissao': 'date',
    }

    for col_idx, h in enumerate(headers_faltantes, 1):
        ws_receber.cell(row=1, column=col_idx, value=h)
    _apply_header_style(ws_receber, 1)

    _write_data_rows(ws_receber, 2, faltantes_receber, columns_faltantes, formats_faltantes)
    _auto_width(ws_receber)

    # ---------------------------------------------------------------
    # ABA 3: FALTANTES PAGAR
    # ---------------------------------------------------------------
    ws_pagar = wb.create_sheet('Faltantes Pagar')

    headers_faltantes_pagar = [
        'Odoo Line ID', 'Empresa', 'NF-e', 'Parcela', 'Fornecedor',
        'CNPJ', 'Valor Original', 'Vencimento', 'Emissao', 'Status',
    ]

    for col_idx, h in enumerate(headers_faltantes_pagar, 1):
        ws_pagar.cell(row=1, column=col_idx, value=h)
    _apply_header_style(ws_pagar, 1)

    _write_data_rows(ws_pagar, 2, faltantes_pagar, columns_faltantes, formats_faltantes)
    _auto_width(ws_pagar)

    # ---------------------------------------------------------------
    # ABA 4: DEVOLUCOES DETECTADAS
    # ---------------------------------------------------------------
    ws_dev = wb.create_sheet('Devolucoes detectadas')

    columns_dev = [
        'odoo_line_id', 'empresa', 'nf', 'parcela', 'parceiro_nome',
        'parceiro_cnpj', 'valor_original', 'vencimento', 'emissao', 'status',
    ]
    headers_dev = [
        'Odoo Line ID', 'Empresa', 'NF-e', 'Parcela', 'Parceiro',
        'CNPJ', 'Valor (abs)', 'Vencimento', 'Emissao', 'Status',
    ]

    # Adicionar coluna tipo (Receber/Pagar)
    columns_dev_typed = ['tipo'] + columns_dev
    headers_dev_typed = ['Tipo'] + headers_dev

    for col_idx, h in enumerate(headers_dev_typed, 1):
        ws_dev.cell(row=1, column=col_idx, value=h)
    _apply_header_style(ws_dev, 1, fill=header_fill_orange)

    # Combinar receber e pagar
    all_devolucoes = []
    for item in devolucoes_odoo_receber:
        item_copy = dict(item)
        item_copy['tipo'] = 'Receber'
        all_devolucoes.append(item_copy)
    for item in devolucoes_odoo_pagar:
        item_copy = dict(item)
        item_copy['tipo'] = 'Pagar'
        all_devolucoes.append(item_copy)

    formats_dev = {
        'valor_original': 'number',
        'vencimento': 'date',
        'emissao': 'date',
    }
    _write_data_rows(ws_dev, 2, all_devolucoes, columns_dev_typed, formats_dev)
    _auto_width(ws_dev)

    # ---------------------------------------------------------------
    # ABA 5: DEVOLUCOES NO SISTEMA
    # ---------------------------------------------------------------
    ws_dev_sys = wb.create_sheet('Devolucoes no sistema')

    columns_dev_sys = [
        'id_local', 'tipo', 'empresa', 'titulo_nf', 'parcela',
        'odoo_line_id', 'cnpj', 'raz_social', 'valor_original', 'valor_residual',
        'vencimento', 'emissao', 'status_sistema', 'metodo_baixa', 'observacao',
    ]
    headers_dev_sys = [
        'ID Local', 'Tipo', 'Empresa', 'NF-e', 'Parcela',
        'Odoo Line ID', 'CNPJ', 'Razao Social', 'Valor Original', 'Valor Residual',
        'Vencimento', 'Emissao', 'Status', 'Metodo Baixa', 'Observacao',
    ]

    for col_idx, h in enumerate(headers_dev_sys, 1):
        ws_dev_sys.cell(row=1, column=col_idx, value=h)
    _apply_header_style(ws_dev_sys, 1, fill=header_fill_red)

    formats_dev_sys = {
        'valor_original': 'number',
        'valor_residual': 'number',
    }
    _write_data_rows(ws_dev_sys, 2, devolucoes_no_sistema, columns_dev_sys, formats_dev_sys)
    _auto_width(ws_dev_sys)

    # Salvar
    wb.save(output_path)
    logger.info(f"  Excel salvo em: {output_path}")

    return output_path


# ===================================================================
# ORQUESTRADOR PRINCIPAL
# ===================================================================

def processar_tipo_faltantes(connection, tipo, production_dir=None, limite=None):
    """
    Processa faltantes para um tipo (receber ou pagar).

    Fluxo:
    1. Carregar IDs locais (producao ou DB)
    2. Buscar IDs elegiveis no Odoo
    3. Set difference → IDs faltantes
    4. Read dados completos dos faltantes
    5. Buscar partners
    6. Filtrar intercompany
    7. Retornar dados transformados

    Args:
        connection: OdooConnection autenticada
        tipo: 'receber' ou 'pagar'
        production_dir: diretorio com IDs pre-exportados
        limite: limite de IDs faltantes a processar

    Returns:
        tuple: (registros_filtrados, stats)
    """
    tipo_label = tipo.upper()
    logger.info(f"\n{'=' * 60}")
    logger.info(f"FALTANTES: CONTAS A {tipo_label}")
    logger.info(f"{'=' * 60}")

    # Phase 1: IDs locais
    logger.info(f"\n[1/7] Carregando IDs locais...")
    ids_locais = carregar_ids_locais(tipo, production_dir)

    # Phase 2: IDs elegiveis Odoo
    logger.info(f"\n[2/7] Buscando IDs elegiveis no Odoo...")
    ids_odoo = buscar_ids_odoo_elegiveis(connection, tipo)

    # Phase 3: Set difference
    logger.info(f"\n[3/7] Calculando gap...")
    ids_odoo_set = set(ids_odoo)
    gap_ids = [id_ for id_ in ids_odoo if id_ not in ids_locais]
    logger.info(
        f"  Gap: {len(gap_ids):,} titulos no Odoo que nao existem localmente "
        f"(de {len(ids_odoo_set):,} elegiveis)"
    )

    if not gap_ids:
        logger.info("  Nenhum gap encontrado!")
        return [], {
            'odoo_elegiveis': len(ids_odoo_set),
            'local_com_id': len(ids_locais),
            'faltantes': 0,
            'faltantes_antes_crossref': 0,
            'falsos_positivos': 0,
            'intercompany': 0,
        }

    # Phase 4: Dados completos
    logger.info(f"\n[4/7] Buscando dados completos dos faltantes...")
    records, _usou_partner = buscar_dados_titulos(connection, gap_ids, limite)

    if not records:
        logger.error("  Nenhum registro retornado pelo read()!")
        return [], {
            'odoo_elegiveis': len(ids_odoo_set),
            'local_com_id': len(ids_locais),
            'faltantes': len(gap_ids),
            'faltantes_antes_crossref': len(gap_ids),
            'falsos_positivos': 0,
            'intercompany': 0,
        }

    # Phase 5: Partners
    logger.info(f"\n[5/7] Buscando dados de parceiros...")
    partner_map = buscar_partners(connection, records)

    # Phase 6: Transformar e filtrar intercompany
    logger.info(f"\n[6/7] Transformando e filtrando...")
    registros, filter_stats = transformar_registros(records, partner_map, tipo)

    # Phase 7: Cross-reference contra chaves de negocio locais
    # Remove falsos positivos: mesma NF+empresa+parcela ja existe via outro odoo_line_id
    logger.info(f"\n[7/7] Cross-reference contra chaves de negocio locais...")
    db.session.close()
    chaves_locais = carregar_chaves_negocio_locais(tipo, production_dir)

    verdadeiros_faltantes = []
    falsos_positivos = 0
    for reg in registros:
        empresa_local = ODOO_TO_LOCAL_EMPRESA.get(reg['company_id'])
        chave = (empresa_local, reg['nf'], reg['parcela'])
        if chave in chaves_locais:
            falsos_positivos += 1
        else:
            verdadeiros_faltantes.append(reg)

    logger.info(f"  Falsos positivos removidos: {falsos_positivos:,}")
    logger.info(f"  Verdadeiros faltantes: {len(verdadeiros_faltantes):,}")

    stats = {
        'odoo_elegiveis': len(ids_odoo_set),
        'local_com_id': len(ids_locais),
        'faltantes': len(verdadeiros_faltantes),
        'faltantes_antes_crossref': len(registros),
        'falsos_positivos': falsos_positivos,
        'intercompany': filter_stats['intercompany'],
    }

    return verdadeiros_faltantes, stats


def processar_devolucoes_odoo(connection, tipo, limite=None):
    """
    Processa devolucoes de um tipo no Odoo.

    Args:
        connection: OdooConnection autenticada
        tipo: 'receber' ou 'pagar'
        limite: limite de IDs a processar

    Returns:
        tuple: (registros, total_ids)
    """
    tipo_label = tipo.upper()
    logger.info(f"\n[Dev] Buscando devolucoes {tipo_label} no Odoo...")

    # Buscar IDs
    ids_dev = buscar_ids_devolucoes_odoo(connection, tipo)

    if not ids_dev:
        logger.info(f"  Nenhuma devolucao {tipo_label} encontrada")
        return [], 0

    total_ids = len(ids_dev)

    # Read dados (limitados se necessario)
    records, _usou_partner = buscar_dados_titulos(connection, ids_dev, limite)

    if not records:
        return [], total_ids

    # Partners
    partner_map = buscar_partners(connection, records)

    # Transformar (sem filtro intercompany — queremos mostrar todas)
    resultados = []
    for r in records:
        company_id = _extrair_company_id(r)
        partner_id_num = _extrair_partner_id(r)
        p_info = partner_map.get(partner_id_num, {})

        # Devolucoes: usar balance (valor absoluto) para ambos os tipos
        # Receber: balance < 0 → abs() para mostrar positivo
        # Pagar: balance > 0 → valor direto (debit entry = vendor refund)
        balance = float(r.get('balance', 0) or 0)

        resultados.append({
            'odoo_line_id': r.get('id'),
            'empresa': ODOO_COMPANY_SIGLA.get(company_id, '?'),
            'nf': str(r.get('x_studio_nf_e') or '').strip(),
            'parcela': r.get('l10n_br_cobranca_parcela') or '',
            'parceiro_nome': p_info.get('name') or '',
            'parceiro_cnpj': p_info.get('l10n_br_cnpj') or '',
            'valor_original': round(abs(balance), 2),
            'vencimento': r.get('date_maturity') or '',
            'emissao': r.get('date') or '',
            'status': 'DEVOLUCAO',
        })

    return resultados, total_ids


def main():
    parser = argparse.ArgumentParser(
        description='Gap Analysis Odoo <-> Local → Excel Multi-Aba',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos:
  # Com IDs de producao pre-exportados do Render
  python scripts/exportar_gap_odoo_excel.py --production /tmp

  # Apenas receber
  python scripts/exportar_gap_odoo_excel.py --production /tmp --tipo receber

  # Limitar para teste rapido (50 faltantes + 20 devolucoes)
  python scripts/exportar_gap_odoo_excel.py --production /tmp --limite 50

  # Verbose
  python scripts/exportar_gap_odoo_excel.py --production /tmp -v
        """,
    )
    parser.add_argument(
        '--tipo', choices=['pagar', 'receber'],
        help='Processar apenas um tipo (default: ambos)',
    )
    parser.add_argument(
        '--production', type=str, default=None, metavar='DIR',
        help='Diretorio com render_pagar_ids.json e render_receber_ids.json',
    )
    parser.add_argument(
        '--limite', type=int, default=None,
        help='Limite de IDs faltantes por tipo (para teste)',
    )
    parser.add_argument(
        '--limite-dev', type=int, default=200,
        help='Limite de devolucoes por tipo para leitura detalhada (default: 200)',
    )
    parser.add_argument(
        '--output', type=str, default=None,
        help='Caminho do arquivo Excel de saida',
    )
    parser.add_argument(
        '--verbose', '-v', action='store_true',
        help='Logging detalhado (DEBUG)',
    )

    args = parser.parse_args()
    configurar_logging(args.verbose)

    inicio = agora_utc_naive()
    timestamp = inicio.strftime('%Y%m%d_%H%M%S')
    output_path = args.output or f'/tmp/gap_analysis_odoo_{timestamp}.xlsx'

    logger.info("=" * 60)
    logger.info("GAP ANALYSIS ODOO <-> LOCAL → EXCEL")
    logger.info(f"Inicio: {inicio.strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(f"Tipo: {args.tipo or 'ambos'}")
    if args.production:
        logger.info(f"MODO PRODUCAO: IDs de {args.production}")
    if args.limite:
        logger.info(f"Limite faltantes: {args.limite}")
    logger.info(f"Limite devolucoes: {args.limite_dev}")
    logger.info(f"Output: {output_path}")
    logger.info("=" * 60)

    # Criar app Flask
    app = create_app()

    with app.app_context():
        # Conectar ao Odoo
        logger.info("\nConectando ao Odoo...")
        connection = get_odoo_connection()
        if not connection.authenticate():
            logger.error("Falha na autenticacao com Odoo")
            sys.exit(1)
        logger.info("Conectado ao Odoo")

        tipos = [args.tipo] if args.tipo else ['receber', 'pagar']

        # ----- FALTANTES -----
        faltantes_receber = []
        faltantes_pagar = []
        stats_receber = {}
        stats_pagar = {}

        if 'receber' in tipos:
            faltantes_receber, stats_receber = processar_tipo_faltantes(
                connection, 'receber', args.production, args.limite,
            )

        if 'pagar' in tipos:
            # db.session.close() antes de queries longas ao Odoo
            db.session.close()
            faltantes_pagar, stats_pagar = processar_tipo_faltantes(
                connection, 'pagar', args.production, args.limite,
            )

        # ----- DEVOLUCOES ODOO -----
        logger.info(f"\n{'=' * 60}")
        logger.info("DEVOLUCOES NO ODOO")
        logger.info(f"{'=' * 60}")

        devolucoes_odoo_receber = []
        devolucoes_odoo_pagar = []
        total_dev_receber = 0
        total_dev_pagar = 0

        if 'receber' in tipos:
            db.session.close()
            devolucoes_odoo_receber, total_dev_receber = processar_devolucoes_odoo(
                connection, 'receber', args.limite_dev,
            )

        if 'pagar' in tipos:
            db.session.close()
            devolucoes_odoo_pagar, total_dev_pagar = processar_devolucoes_odoo(
                connection, 'pagar', args.limite_dev,
            )

        # ----- DEVOLUCOES NO SISTEMA -----
        logger.info(f"\n{'=' * 60}")
        logger.info("DEVOLUCOES JA NO SISTEMA LOCAL")
        logger.info(f"{'=' * 60}")

        db.session.close()
        devolucoes_no_sistema = buscar_devolucoes_no_sistema(args.production)
        logger.info(f"  {len(devolucoes_no_sistema)} registros de devolucao no sistema")

        # ----- MONTAR RESUMO -----
        logger.info(f"\n{'=' * 60}")
        logger.info("MONTANDO RESUMO")
        logger.info(f"{'=' * 60}")

        # Por tipo
        por_tipo = []
        if stats_receber:
            odoo_el = stats_receber.get('odoo_elegiveis', 0)
            local_id = stats_receber.get('local_com_id', 0)
            cobertura = local_id / odoo_el if odoo_el > 0 else 0
            por_tipo.append({
                'tipo': 'Receber',
                'odoo_elegiveis': odoo_el,
                'local_com_id': local_id,
                'faltantes_bruto': stats_receber.get('faltantes_antes_crossref', 0),
                'falsos_positivos': stats_receber.get('falsos_positivos', 0),
                'faltantes': stats_receber.get('faltantes', 0),
                'intercompany': stats_receber.get('intercompany', 0),
                'cobertura_pct': cobertura,
            })

        if stats_pagar:
            odoo_el = stats_pagar.get('odoo_elegiveis', 0)
            local_id = stats_pagar.get('local_com_id', 0)
            cobertura = local_id / odoo_el if odoo_el > 0 else 0
            por_tipo.append({
                'tipo': 'Pagar',
                'odoo_elegiveis': odoo_el,
                'local_com_id': local_id,
                'faltantes_bruto': stats_pagar.get('faltantes_antes_crossref', 0),
                'falsos_positivos': stats_pagar.get('falsos_positivos', 0),
                'faltantes': stats_pagar.get('faltantes', 0),
                'intercompany': stats_pagar.get('intercompany', 0),
                'cobertura_pct': cobertura,
            })

        # Por empresa
        empresa_receber = {}
        for item in faltantes_receber:
            emp = item.get('empresa', '?')
            empresa_receber[emp] = empresa_receber.get(emp, 0) + 1

        empresa_pagar = {}
        for item in faltantes_pagar:
            emp = item.get('empresa', '?')
            empresa_pagar[emp] = empresa_pagar.get(emp, 0) + 1

        todas_empresas = sorted(set(list(empresa_receber.keys()) + list(empresa_pagar.keys())))
        por_empresa = []
        for emp in todas_empresas:
            rec = empresa_receber.get(emp, 0)
            pag = empresa_pagar.get(emp, 0)
            por_empresa.append({
                'empresa': emp,
                'receber': rec,
                'pagar': pag,
                'total': rec + pag,
            })

        resumo_data = {
            'por_tipo': por_tipo,
            'por_empresa': por_empresa,
            'devolucoes_odoo': total_dev_receber + total_dev_pagar,
            'devolucoes_sistema': len(devolucoes_no_sistema),
        }

        # Log resumo
        for td in por_tipo:
            logger.info(
                f"  {td['tipo']}: Odoo={td['odoo_elegiveis']:,} | "
                f"Local={td['local_com_id']:,} | "
                f"Faltantes bruto={td['faltantes_bruto']:,} | "
                f"Falsos positivos={td['falsos_positivos']:,} | "
                f"Faltantes final={td['faltantes']:,} | "
                f"Intercompany={td['intercompany']} | "
                f"Cobertura={td['cobertura_pct']:.1%}"
            )

        # ----- GERAR EXCEL -----
        logger.info(f"\n{'=' * 60}")
        logger.info("GERANDO EXCEL")
        logger.info(f"{'=' * 60}")

        gerar_excel(
            faltantes_receber=faltantes_receber,
            faltantes_pagar=faltantes_pagar,
            devolucoes_odoo_receber=devolucoes_odoo_receber,
            devolucoes_odoo_pagar=devolucoes_odoo_pagar,
            devolucoes_no_sistema=devolucoes_no_sistema,
            resumo_data=resumo_data,
            output_path=output_path,
        )

        # ----- RESULTADO FINAL -----
        fim = agora_utc_naive()
        duracao = (fim - inicio).total_seconds()

        file_size = os.path.getsize(output_path)
        file_size_str = (
            f"{file_size / 1024:.1f} KB"
            if file_size < 1024 * 1024
            else f"{file_size / 1024 / 1024:.1f} MB"
        )

        print(f"\n{'=' * 60}")
        print("EXCEL GERADO COM SUCESSO!")
        print(f"{'=' * 60}")
        print(f"  Arquivo: {output_path}")
        print(f"  Tamanho: {file_size_str}")
        print(f"  Abas:")
        print(f"    1. Resumo")
        print(f"    2. Faltantes Receber: {len(faltantes_receber):,} registros")
        print(f"    3. Faltantes Pagar:   {len(faltantes_pagar):,} registros")
        print(
            f"    4. Devolucoes Odoo:   "
            f"{len(devolucoes_odoo_receber) + len(devolucoes_odoo_pagar):,} "
            f"registros (de {total_dev_receber + total_dev_pagar:,} total)"
        )
        print(f"    5. Devolucoes no sistema: {len(devolucoes_no_sistema)} registros")
        print(f"  Tempo: {duracao:.0f}s")
        print(f"{'=' * 60}")
        print(f"\n  Proximo passo: Revisar o Excel e aprovar antes de importar.")
        print(f"  Comando de importacao: python scripts/reconciliar_titulos_odoo.py --importar")


if __name__ == '__main__':
    main()
