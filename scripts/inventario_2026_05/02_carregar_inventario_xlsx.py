"""F7.2 — Carrega planilha do inventario fisico em /tmp.

Formato esperado:
- 1 aba por company: FB, CD, LF (case-sensitive, exatamente esses nomes)
- Header na linha 1 (ordem das colunas pode variar entre abas)
- Colunas obrigatorias por nome (case-insensitive, ignora espacos extras):
    CODIGO, LOTE, QTD
- Colunas opcionais:
    VALIDADE (datetime ou string parsavel; "S/INF"/"S/ INF" = sem validade)
    DESCRICAO/PRODUTO/MEDIDA/LOCAL (ignoradas)

Validacoes:
- cod_produto: deve comecar com 1, 2, 3 ou 4 (P002). Outliers (letra inicial,
  ex.: 'CHAVE-X') sao PULADOS silenciosamente com contagem (decisao
  usuario 2026-05-17 — limpa erros de digitacao da planilha).
- qtd_contada: >= 0 (negativo invalido).
- lote: convertido sempre para STRING (planilha mistura int/str em LF);
  vazio → string vazia (script 03 resolve via mais-novo-disponivel).
- validade: datetime → ISO date; "S/INF"/"S/ INF" → None; outras strings
  tentam parse (formatos suportados: ISO YYYY-MM-DD, BR DD/MM/YYYY).

Uso:
    python scripts/inventario_2026_05/02_carregar_inventario_xlsx.py \\
        --xlsx '/mnt/c/Users/.../COMPILADO INV. 16.05.2026.xlsx' [--dry-run]

Spec: docs/superpowers/plans/2026-05-17-ajuste-inventario-nacom-lf.md Task 7.2
Ajustes 2026-05-17: planilha real tem cabecalhos divergentes por aba +
lote int/str misturado + validade.
"""
import argparse
import json
import os
import sys
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

# sys.path para `from app import ...`
_THIS = Path(__file__).resolve()
sys.path.insert(0, str(_THIS.parents[2]))

import openpyxl  # noqa: E402

from app import create_app  # noqa: E402
from app.odoo.constants.operacoes_fiscais import CODIGO_PARA_COMPANY_ID  # noqa: E402
from app.utils.timezone import agora_utc_naive  # noqa: E402
TIPOS_ACEITOS = {'1', '2', '3', '4'}
OUTPUT_JSON = '/tmp/inventario_fisico_2026_05.json'

# Aliases case-insensitive de header → chave canonica
HEADER_ALIASES = {
    'codigo': 'cod_produto',
    'cod': 'cod_produto',
    'cod_produto': 'cod_produto',
    'lote': 'lote',
    'qtd': 'qtd',
    'quantidade': 'qtd',
    'qtd_contada': 'qtd',
    'validade': 'validade',
    'venc': 'validade',
    'vencimento': 'validade',
}


def detectar_colunas(ws) -> dict:
    """Mapeia headers da linha 1 para indices 0-based.

    Retorna dict com chaves: cod_produto, lote, qtd, validade
    (validade pode estar None se a aba nao tiver coluna validade).
    """
    indices = {'cod_produto': None, 'lote': None, 'qtd': None, 'validade': None}
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c).value
        if cell is None:
            continue
        key_normalizada = str(cell).strip().lower()
        canonica = HEADER_ALIASES.get(key_normalizada)
        if canonica and indices.get(canonica) is None:
            indices[canonica] = c - 1  # 0-based
    return indices


def parsear_validade(valor) -> tuple:
    """Normaliza validade do Excel para ISO date string ou None.

    Returns: (validade_iso: str|None, erro: str|None)
    """
    if valor is None:
        return None, None
    if isinstance(valor, datetime):
        return valor.date().isoformat(), None
    if isinstance(valor, date):
        return valor.isoformat(), None
    if isinstance(valor, str):
        v_clean = valor.strip()
        if not v_clean:
            return None, None
        # "S/INF", "S/ INF", "SEM INF" etc — sem validade
        up = v_clean.upper()
        if up.startswith('S/') or 'SEM INF' in up or up in ('-', 'N/A'):
            return None, None
        # Tentar parse ISO YYYY-MM-DD ou BR DD/MM/YYYY
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S'):
            try:
                d = datetime.strptime(v_clean, fmt).date()
                return d.isoformat(), None
            except ValueError:
                continue
        return None, f'validade nao parsavel: {valor!r}'
    return None, f'validade tipo inesperado: {type(valor).__name__} ({valor!r})'


def carregar_xlsx(path: str) -> dict:
    """Le planilha + valida. Retorna dict {timestamp, origem, companies, erros}.

    Skipped (silencioso, com contagem):
    - linhas totalmente vazias
    - cod_produto comecando com nao-digito (erro de digitacao na planilha)
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {
        'timestamp': agora_utc_naive().isoformat(),
        'origem': path,
        'companies': {},
    }
    erros = []
    skipped_outliers_total = 0

    for codigo, cid in CODIGO_PARA_COMPANY_ID.items():
        if codigo not in wb.sheetnames:
            erros.append(f'Sheet {codigo!r} ausente em {path}')
            continue
        ws = wb[codigo]
        idx = detectar_colunas(ws)

        # Validar que colunas obrigatorias foram detectadas
        ausentes = [
            k for k in ('cod_produto', 'lote', 'qtd') if idx[k] is None
        ]
        if ausentes:
            erros.append(
                f'Aba {codigo!r}: colunas ausentes {ausentes}. '
                f'Headers encontrados: '
                f'{[ws.cell(1, c).value for c in range(1, ws.max_column + 1)]}'
            )
            continue

        linhas = []
        skipped_outliers = 0
        skipped_vazias = 0
        for row_n, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if not row or all(c is None for c in row):
                skipped_vazias += 1
                continue

            # Extrair pelos indices detectados
            cod_raw = row[idx['cod_produto']] if idx['cod_produto'] < len(row) else None
            lote_raw = row[idx['lote']] if idx['lote'] < len(row) else None
            qtd_raw = row[idx['qtd']] if idx['qtd'] < len(row) else None
            validade_raw = (
                row[idx['validade']]
                if idx['validade'] is not None and idx['validade'] < len(row)
                else None
            )

            # cod_produto: sempre string
            cod = str(cod_raw).strip() if cod_raw is not None else ''
            if not cod:
                continue  # linha sem codigo, skip silencioso
            # Outlier: primeiro char nao-digito → skip silencioso
            if not cod[0].isdigit():
                skipped_outliers += 1
                skipped_outliers_total += 1
                continue
            if cod[0] not in TIPOS_ACEITOS:
                # Comeca com digito mas fora 1-4 (ex.: 5, 9) → erro real
                erros.append(
                    f'{codigo} linha {row_n}: cod_produto {cod!r} '
                    'nao comeca com 1-4'
                )
                continue

            # qtd: numerica >= 0
            try:
                qtd_dec = Decimal(str(qtd_raw)) if qtd_raw is not None else Decimal('0')
            except (InvalidOperation, ValueError):
                erros.append(
                    f'{codigo} linha {row_n}: qtd nao numerica ({qtd_raw!r})'
                )
                continue
            if qtd_dec < 0:
                erros.append(
                    f'{codigo} linha {row_n}: qtd negativa ({qtd_dec})'
                )
                continue

            # lote: sempre str (planilha LF mistura int/str)
            lote_str = str(lote_raw).strip() if lote_raw is not None else ''

            # validade
            validade_iso, validade_erro = parsear_validade(validade_raw)
            if validade_erro:
                erros.append(
                    f'{codigo} linha {row_n}: {validade_erro}'
                )
                # nao continua — registra erro mas mantem linha sem validade

            linhas.append({
                'cod_produto': cod,
                'lote_inventariado': lote_str,
                'qtd_inventario': str(qtd_dec),
                'tipo_produto': int(cod[0]),
                'validade_inv': validade_iso,
                'linha_origem': row_n,
            })

        result['companies'][cid] = {'codigo': codigo, 'linhas': linhas}
        print(
            f'  {codigo}: {len(linhas)} linhas validas '
            f'(outliers pulados: {skipped_outliers}, '
            f'vazias: {skipped_vazias})'
        )

    if skipped_outliers_total:
        print(
            f'\nTOTAL outliers pulados (cod nao-digito): {skipped_outliers_total} '
            '(decisao usuario 2026-05-17: erros de digitacao da planilha)'
        )

    if erros:
        print(f'\nERROS ({len(erros)}):')
        for e in erros[:30]:
            print(f'  - {e}')
        if len(erros) > 30:
            print(f'  ... +{len(erros) - 30} erros adicionais')
        result['erros'] = erros
    return result


def main(path: str, dry_run: bool) -> None:
    app = create_app()
    with app.app_context():
        if not os.path.exists(path):
            raise FileNotFoundError(f'Planilha nao encontrada: {path}')

        result = carregar_xlsx(path)

        if result.get('erros') and not dry_run:
            print(
                '\nABORTANDO: ha erros de validacao. Corrija a planilha '
                'e rode novamente. (Use --dry-run para gerar parcial.)'
            )
            sys.exit(2)

        if not dry_run:
            with open(OUTPUT_JSON, 'w') as f:
                json.dump(result, f, default=str, indent=2)
            print(f'\nSnapshot: {OUTPUT_JSON}')
        else:
            print(f'\n[DRY RUN] nao gravou {OUTPUT_JSON}')


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        '--xlsx', required=True,
        help='caminho do .xlsx do inventario fisico',
    )
    parser.add_argument(
        '--dry-run', action='store_true',
        help='nao grava JSON; apenas valida e imprime estatisticas',
    )
    args = parser.parse_args()
    main(args.xlsx, args.dry_run)
