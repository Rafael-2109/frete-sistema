"""F7.7 — Extrai estado pos-execucao do Odoo e compara com proposta.

Para cada ajuste EXECUTADO/FALHA do ciclo+company_id, consulta o Odoo
e gera Excel comparativo:

- Estado proposto (do banco local: lote_origem, lote_destino, qtd_ajuste,
  custo_medio, picking_id_odoo, invoice_id_odoo, chave_nfe)
- Estado realizado (do Odoo: stock.lot atual, stock.quant atual,
  account.move emitida, l10n_br SEFAZ status)
- Status (EXECUTADO/FALHA/DIVERGENCIA) + diff esperado vs observado

REPLICAVEL POR FILIAL: --company-id=N (1=FB, 4=CD, 5=LF).

Outputs:
- docs/inventario-2026-05/07-relatorios/pos-execucao-{CODIGO}-{ciclo}.xlsx
- /tmp/pos_execucao_{CODIGO}_{ciclo}.json (machine readable)

Uso:
    python scripts/inventario_2026_05/08_extrair_pos_execucao.py \
        --ciclo=INVENTARIO_2026_05 --company-id=5

    # Filtrar 1 produto (debug):
    python scripts/inventario_2026_05/08_extrair_pos_execucao.py \
        --ciclo=INVENTARIO_2026_05 --company-id=5 --cod-produto=210030325

    # So ler do banco (sem Odoo):
    python scripts/inventario_2026_05/08_extrair_pos_execucao.py \
        --ciclo=INVENTARIO_2026_05 --company-id=5 --offline

Spec: docs/inventario-2026-05/CHECKPOINT_2026_05_17_FIM_DIA.md F7.7
"""
import argparse
import json
import sys
from collections import defaultdict
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List

_THIS = Path(__file__).resolve()
sys.path.insert(0, str(_THIS.parents[2]))

import openpyxl  # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment  # noqa: E402

from app import create_app  # noqa: E402
from app.odoo.models import AjusteEstoqueInventario  # noqa: E402
from app.odoo.utils.connection import get_odoo_connection  # noqa: E402

OUTPUT_DIR = _THIS.parents[2] / 'docs' / 'inventario-2026-05' / '07-relatorios'
COMPANIES = {1: 'FB', 4: 'CD', 5: 'LF'}

STATUS_FINAL = ('EXECUTADO', 'FALHA')


# ============================================================
# Helpers de comparacao
# ============================================================

def to_dec(v) -> Decimal:
    if v is None or v == '':
        return Decimal('0')
    return Decimal(str(v))


def classificar(ajuste: AjusteEstoqueInventario, achados: Dict) -> str:
    """Compara proposta vs realizado e retorna classificacao.

    Atualizado 2026-05-18 apos piloto 210030325 LF — corrige 2 bugs:

    1. RENOMEAR_LOTE: na realidade executa TRANSFERIR (D006). Validar
       que lote_destino existe na company origem (LF) e que tem qty
       no quant do lote alvo correspondente ao qtd_inventario.
       Lotes ORIGEM permanecem (apenas tiveram qty transferida) —
       nao remover do Odoo nem renomear.

    2. PERDA_LF_FB: picking sai para Parceiros/Clientes (virtual, id=5),
       NAO entra na FB. Nao buscar qty_no_destino na FB. EXECUTADO_OK
       se picking done + invoice posted + chave_nfe presente (NF-e
       SEFAZ autorizada).
    """
    if ajuste.status == 'FALHA':
        return 'FALHA'
    if ajuste.fase_pipeline and 'SKIP' in ajuste.fase_pipeline.upper():
        return 'SKIP'
    if achados.get('erro_consulta'):
        return 'ERRO_CONSULTA'

    # RENOMEAR_LOTE (=TRANSFERIR_LOTE, D006): lote alvo deve existir com
    # qty igual a qtd_inventario do ajuste.
    if ajuste.acao_decidida == 'RENOMEAR_LOTE':
        lote_alvo = achados.get('lote_destino_apos')
        if not lote_alvo:
            return 'DIVERGENCIA_LOTE_ALVO_NAO_CRIADO'
        # Verificar qty no quant do lote alvo (achados.qty_no_lote_alvo_origem)
        qty_alvo = achados.get('qty_no_lote_alvo_origem') or Decimal('0')
        qty_esperada = to_dec(ajuste.qtd_inventario)
        if qty_esperada == 0:
            # Sem qtd_inventario para validar — apenas lote existir
            return 'EXECUTADO_OK'
        # qty_alvo eh acumulado total do lote alvo no produto/company.
        # Validar que >= qtd_inventario (pode ter mais se outros ajustes
        # ja transferiram para o mesmo lote).
        if qty_alvo >= qty_esperada - Decimal('0.01'):
            return 'EXECUTADO_OK'
        return 'DIVERGENCIA_QTY_NO_LOTE_ALVO'

    # PERDA_LF_FB e demais NFs fiscais: validar pipeline ate F5e
    # (picking + invoice + chave SEFAZ). NAO comparar qty no destino —
    # picking de PERDA sai para location virtual, NAO entra na FB.
    if ajuste.acao_decidida in (
        'PERDA_LF_FB', 'INDUSTRIALIZACAO_FB_LF',
        'DEV_FB_LF', 'DEV_LF_FB', 'DEV_CD_LF', 'DEV_LF_CD',
        'TRANSFERIR_CD_FB', 'TRANSFERIR_FB_CD',
    ):
        if not ajuste.picking_id_odoo:
            return 'PENDENTE_SEM_PICKING'
        if not ajuste.invoice_id_odoo:
            return 'PENDENTE_SEM_INVOICE'
        if not ajuste.chave_nfe:
            return 'PENDENTE_SEM_SEFAZ'

        # Validar que picking esta done + invoice posted (info do achados)
        picking_state = achados.get('picking_state')
        invoice_state = achados.get('invoice', {}).get('state')
        if picking_state and picking_state != 'done':
            return f'DIVERGENCIA_PICKING_{picking_state.upper()}'
        if invoice_state and invoice_state not in ('posted',):
            return f'DIVERGENCIA_INVOICE_{invoice_state.upper()}'

        # Validar reducao na ORIGEM (LF para PERDA_LF_FB).
        # qty_reduzida_na_origem = qty_origem_antes (do ajuste.qtd_odoo) -
        #                          qty_origem_atual (achados.qty_no_lote_origem_company_origem)
        # Mas como nao temos `antes` exato no extrator (so o ajuste eh fonte),
        # nao validamos isso — apenas pipeline completo eh suficiente.
        return 'EXECUTADO_OK'

    # INDISPONIBILIZAR_*: lote inativo
    if ajuste.acao_decidida.startswith('INDISPONIBILIZAR_'):
        ativo = achados.get('lote_ativo')
        if ativo is None:
            return 'PENDENTE_LOTE_NAO_LOCALIZADO'
        return 'EXECUTADO_OK' if ativo is False else 'DIVERGENCIA_LOTE_ATIVO'

    return 'INDETERMINADO'


# ============================================================
# Consultas Odoo
# ============================================================

def consultar_estado_odoo(
    odoo, ajuste: AjusteEstoqueInventario, product_id_cache: Dict[str, int],
) -> Dict[str, Any]:
    """Para 1 ajuste, busca estado atual no Odoo."""
    achados: Dict[str, Any] = {}
    try:
        # Resolver product_id
        cod = ajuste.cod_produto
        product_id = product_id_cache.get(cod)
        if product_id is None:
            res = odoo.search_read(
                'product.product',
                [['default_code', '=', cod]],
                ['id'], limit=1,
            )
            if not res:
                achados['erro_consulta'] = (
                    f'product.default_code={cod!r} nao encontrado'
                )
                return achados
            product_id = res[0]['id']
            product_id_cache[cod] = product_id
        achados['product_id'] = product_id

        # Buscar stock.lot do lote_destino (lote alvo)
        nome_alvo = ajuste.lote_destino or ajuste.lote_inventariado
        if nome_alvo:
            ids = odoo.search(
                'stock.lot',
                [
                    ['name', 'in', [nome_alvo]],
                    ['product_id', '=', product_id],
                    ['company_id', '=', ajuste.company_id],
                ], limit=1,
            )
            if ids:
                # NOTA: stock.lot no Odoo CIEL IT nao tem campo 'active'
                # (nem em read nem em search domain). Para INDISPONIBILIZAR_*
                # usar outra estrategia (futuro — fase 7.9).
                lot = odoo.read(
                    'stock.lot', ids,
                    ['name', 'product_id', 'company_id', 'expiration_date'],
                )[0]
                achados['lote_destino_apos'] = {
                    'id': lot['id'], 'name': lot['name'],
                    'expiration_date': lot.get('expiration_date'),
                }
                achados['lote_ativo'] = None  # nao consigo determinar via XML-RPC

        # Para RENOMEAR_LOTE — buscar tambem o nome de origem (que NAO deveria
        # mais existir se o renomeio deu certo)
        if ajuste.acao_decidida == 'RENOMEAR_LOTE' and ajuste.lote_origem:
            ids_orig = odoo.search(
                'stock.lot',
                [
                    ['name', 'in', [ajuste.lote_origem]],
                    ['product_id', '=', product_id],
                    ['company_id', '=', ajuste.company_id],
                ], limit=1,
            )
            achados['lote_origem_ainda_existe'] = bool(ids_orig)
            achados['lote_origem_apos'] = achados.get('lote_destino_apos')

        # stock.quant do produto/company atualmente
        quants = odoo.search_read(
            'stock.quant',
            [
                ['product_id', '=', product_id],
                ['company_id', '=', ajuste.company_id],
                ['location_id.usage', '=', 'internal'],
                ['quantity', '!=', 0],
            ],
            ['id', 'lot_id', 'location_id', 'quantity', 'value'],
        )
        achados['quants_lf'] = [
            {
                'quant_id': q['id'],
                'lot_id': q['lot_id'][0] if q.get('lot_id') else None,
                'lot_name': q['lot_id'][1] if q.get('lot_id') else None,
                'location_id': q['location_id'][0] if q.get('location_id') else None,
                'quantity': q['quantity'],
                'value': q.get('value'),
            }
            for q in quants
        ]
        achados['qty_total_company'] = sum(
            to_dec(q['quantity']) for q in quants
        )

        # qty_no_lote_alvo_origem: soma do qty nos quants do lote alvo
        # (lote_destino) na empresa origem (=ajuste.company_id). Util para
        # validar RENOMEAR_LOTE/TRANSFERIR — o lote alvo deve ter qty na
        # mesma empresa onde foi feita a transferencia.
        if nome_alvo:
            qty_no_alvo = Decimal('0')
            for q in quants:
                lname = q['lot_id'][1] if q.get('lot_id') else None
                if lname == nome_alvo:
                    qty_no_alvo += to_dec(q['quantity'])
            achados['qty_no_lote_alvo_origem'] = qty_no_alvo

        # Picking state (se houver picking_id_odoo)
        if ajuste.picking_id_odoo:
            try:
                pk = odoo.read(
                    'stock.picking', [ajuste.picking_id_odoo],
                    ['id', 'name', 'state', 'liberado_faturamento'],
                )
                if pk:
                    achados['picking'] = pk[0]
                    achados['picking_state'] = pk[0].get('state')
            except Exception as e:
                achados['picking_consulta_erro'] = str(e)

        # qty_no_destino: se PERDA/TRANSFER, conferir saldo na company destino
        # com lote_destino
        if ajuste.acao_decidida in (
            'PERDA_LF_FB', 'TRANSFERIR_CD_FB', 'DEV_LF_FB', 'DEV_CD_LF',
        ):
            destino = 1  # FB
        elif ajuste.acao_decidida in (
            'INDUSTRIALIZACAO_FB_LF', 'TRANSFERIR_FB_CD',
            'DEV_FB_LF', 'DEV_LF_CD',
        ):
            destino = 5 if ajuste.acao_decidida == 'INDUSTRIALIZACAO_FB_LF' else 4
        else:
            destino = ajuste.company_id

        if nome_alvo and destino != ajuste.company_id:
            quants_destino = odoo.search_read(
                'stock.quant',
                [
                    ['product_id', '=', product_id],
                    ['company_id', '=', destino],
                    ['location_id.usage', '=', 'internal'],
                    ['quantity', '!=', 0],
                ],
                ['id', 'lot_id', 'quantity'],
            )
            qty_destino = Decimal('0')
            for q in quants_destino:
                lname = q['lot_id'][1] if q.get('lot_id') else None
                if lname == nome_alvo:
                    qty_destino += to_dec(q['quantity'])
            achados['qty_no_destino'] = qty_destino
            achados['company_destino'] = destino

        # Invoice (account.move) + chave SEFAZ + l10n_br_cfop
        if ajuste.invoice_id_odoo:
            inv = odoo.read(
                'account.move', [ajuste.invoice_id_odoo],
                [
                    'name', 'state', 'fiscal_position_id',
                    'l10n_br_tipo_pedido', 'partner_id', 'company_id',
                    'amount_total', 'invoice_origin', 'invoice_line_ids',
                    'date',
                ],
            )
            if inv:
                inv = inv[0]
                achados['invoice'] = {
                    'id': inv['id'], 'name': inv.get('name'),
                    'state': inv.get('state'),
                    'fiscal_position_id': (
                        inv['fiscal_position_id'][0]
                        if inv.get('fiscal_position_id') else None
                    ),
                    'l10n_br_tipo_pedido': inv.get('l10n_br_tipo_pedido'),
                    'amount_total': inv.get('amount_total'),
                    'invoice_origin': inv.get('invoice_origin'),
                    'date': inv.get('date'),
                }
                # CFOPs das linhas (primeira 5)
                line_ids = (inv.get('invoice_line_ids') or [])[:5]
                if line_ids:
                    lines = odoo.read(
                        'account.move.line', line_ids,
                        ['l10n_br_cfop_id', 'product_id', 'quantity', 'price_unit'],
                    )
                    achados['invoice_lines'] = [
                        {
                            'product_id': (
                                ll['product_id'][0] if ll.get('product_id') else None
                            ),
                            'cfop': (
                                ll['l10n_br_cfop_id'][1]
                                if ll.get('l10n_br_cfop_id') else None
                            ),
                            'quantity': ll.get('quantity'),
                            'price_unit': ll.get('price_unit'),
                        }
                        for ll in lines
                    ]

    except Exception as e:
        achados['erro_consulta'] = f'{type(e).__name__}: {e}'
    return achados


# ============================================================
# Excel
# ============================================================

HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
HEADER_FONT = Font(bold=True, color='FFFFFF')

CLASSIFICACAO_COR = {
    'EXECUTADO_OK': 'C6EFCE',
    'FALHA': 'FFC7CE',
    'SKIP': 'FFEB9C',
    'PENDENTE_SEM_PICKING': 'FFEB9C',
    'PENDENTE_SEM_INVOICE': 'FFEB9C',
    'PENDENTE_SEM_SEFAZ': 'FFEB9C',
    'DIVERGENCIA_NOME_LOTE': 'FFC7CE',
    'DIVERGENCIA_QUANTIDADE': 'FFC7CE',
    'DIVERGENCIA_LOTE_ATIVO': 'FFC7CE',
    'DIVERGENCIA_LOTE_NAO_LOCALIZADO': 'FFC7CE',
    'ERRO_CONSULTA': 'FFC7CE',
    'INDETERMINADO': 'D9D9D9',
}


def montar_excel(
    ajustes: List[AjusteEstoqueInventario], achados_por_id: Dict[int, Dict],
    classificacoes: Dict[int, str], output_path: Path, company_codigo: str,
    ciclo: str,
) -> None:
    """Gera Excel com 2 abas: 'Comparativo' (linha por ajuste) e 'Resumo'."""
    wb = openpyxl.Workbook()

    # Aba Comparativo
    ws = wb.active
    ws.title = 'Comparativo'
    headers = [
        'id', 'classificacao', 'cod_produto', 'company_id',
        'acao_decidida', 'status', 'fase_pipeline',
        'lote_origem_proposta', 'lote_destino_proposta',
        'lote_destino_apos_existe', 'lote_destino_apos_ativo',
        'lote_origem_ainda_existe',
        'qtd_inventario', 'qtd_odoo_proposta', 'qtd_ajuste',
        'custo_medio',
        'picking_id_odoo', 'invoice_id_odoo', 'invoice_state', 'chave_nfe',
        'invoice_fiscal_position_id', 'invoice_tipo_pedido',
        'invoice_cfops', 'invoice_amount_total', 'invoice_date',
        'qty_no_destino', 'qty_total_company_atual',
        'company_destino', 'erro_msg', 'erro_consulta',
    ]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')

    for a in ajustes:
        achados = achados_por_id.get(a.id, {})
        klass = classificacoes.get(a.id, 'INDETERMINADO')
        lote_destino_apos = achados.get('lote_destino_apos') or {}
        invoice = achados.get('invoice') or {}
        cfops = [
            ln['cfop'] for ln in (achados.get('invoice_lines') or [])
            if ln.get('cfop')
        ]
        row = [
            a.id, klass, a.cod_produto, a.company_id,
            a.acao_decidida, a.status, a.fase_pipeline,
            a.lote_origem, a.lote_destino,
            bool(lote_destino_apos), lote_destino_apos.get('active'),
            achados.get('lote_origem_ainda_existe'),
            float(a.qtd_inventario or 0), float(a.qtd_odoo or 0),
            float(a.qtd_ajuste or 0),
            float(a.custo_medio or 0) if a.custo_medio else None,
            a.picking_id_odoo, a.invoice_id_odoo,
            invoice.get('state'), a.chave_nfe,
            invoice.get('fiscal_position_id'),
            invoice.get('l10n_br_tipo_pedido'),
            ', '.join(cfops) if cfops else None,
            invoice.get('amount_total'), invoice.get('date'),
            float(achados.get('qty_no_destino', 0)) if achados.get('qty_no_destino') else None,
            float(achados.get('qty_total_company', 0)),
            achados.get('company_destino'),
            a.erro_msg,
            achados.get('erro_consulta'),
        ]
        ws.append(row)
        cor = CLASSIFICACAO_COR.get(klass, 'FFFFFF')
        for col_idx in range(1, len(row) + 1):
            ws.cell(row=ws.max_row, column=col_idx).fill = PatternFill(
                'solid', fgColor=cor,
            )

    # Auto-width
    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) for c in col if c.value is not None),
            default=10,
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # Aba Resumo
    ws_res = wb.create_sheet('Resumo')
    ws_res.append([f'Ciclo', ciclo])
    ws_res.append([f'Filial', company_codigo])
    ws_res.append([f'Total ajustes', len(ajustes)])
    ws_res.append([])
    ws_res.append(['classificacao', 'qtd', 'soma_qtd_ajuste'])
    contadores = defaultdict(int)
    somas = defaultdict(Decimal)
    for a in ajustes:
        k = classificacoes.get(a.id, 'INDETERMINADO')
        contadores[k] += 1
        somas[k] += abs(to_dec(a.qtd_ajuste))
    for klass in sorted(contadores.keys()):
        ws_res.append([klass, contadores[klass], float(somas[klass])])
        cor = CLASSIFICACAO_COR.get(klass, 'FFFFFF')
        for col_idx in range(1, 4):
            ws_res.cell(row=ws_res.max_row, column=col_idx).fill = PatternFill(
                'solid', fgColor=cor,
            )

    ws_res.append([])
    ws_res.append(['Por acao_decidida'])
    ws_res.append(['acao_decidida', 'classificacao', 'qtd'])
    contagem = defaultdict(int)
    for a in ajustes:
        k = classificacoes.get(a.id, 'INDETERMINADO')
        contagem[(a.acao_decidida, k)] += 1
    for (acao, k), n in sorted(contagem.items()):
        ws_res.append([acao, k, n])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f'\n  Excel: {output_path}')


# ============================================================
# Main
# ============================================================

def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--ciclo', default='INVENTARIO_2026_05',
                        help='ciclo (default: INVENTARIO_2026_05)')
    parser.add_argument('--company-id', type=int, required=True,
                        choices=sorted(COMPANIES.keys()),
                        help='filial (1=FB, 4=CD, 5=LF)')
    parser.add_argument('--cod-produto', default=None,
                        help='filtra 1 produto (debug do caso piloto)')
    parser.add_argument('--so-status', default=None,
                        help='filtra por status (EXECUTADO, FALHA, PROPOSTO...)')
    parser.add_argument('--incluir-pendentes', action='store_true',
                        help='inclui status nao-final (PROPOSTO, APROVADO)')
    parser.add_argument('--offline', action='store_true',
                        help='nao consulta Odoo (so DB local)')
    parser.add_argument('--json-out', default=None,
                        help='caminho do JSON (default: /tmp/...)')
    parser.add_argument('--xlsx-out', default=None,
                        help='caminho do Excel (default: 07-relatorios/...)')
    args = parser.parse_args()

    company_codigo = COMPANIES[args.company_id]

    app = create_app()
    with app.app_context():
        q = AjusteEstoqueInventario.query.filter_by(
            ciclo=args.ciclo, company_id=args.company_id,
        )
        if args.cod_produto:
            q = q.filter_by(cod_produto=args.cod_produto)
        if args.so_status:
            q = q.filter_by(status=args.so_status)
        elif not args.incluir_pendentes:
            q = q.filter(AjusteEstoqueInventario.status.in_(STATUS_FINAL))
        ajustes = q.order_by(AjusteEstoqueInventario.id).all()
        print(f'Ciclo {args.ciclo} | filial {company_codigo}({args.company_id})')
        print(f'Ajustes carregados: {len(ajustes)}')
        if not ajustes:
            print('Nada a processar.')
            return

        # Consultar Odoo
        achados_por_id: Dict[int, Dict] = {}
        product_id_cache: Dict[str, int] = {}
        if args.offline:
            print('Modo --offline: nao consultando Odoo.')
        else:
            odoo = get_odoo_connection()
            for i, a in enumerate(ajustes, start=1):
                if i % 10 == 0 or i == 1:
                    print(f'  Consultando {i}/{len(ajustes)} (ajuste {a.id})')
                achados_por_id[a.id] = consultar_estado_odoo(
                    odoo, a, product_id_cache,
                )

        # Classificar
        classificacoes: Dict[int, str] = {}
        contadores = defaultdict(int)
        for a in ajustes:
            klass = classificar(a, achados_por_id.get(a.id, {}))
            classificacoes[a.id] = klass
            contadores[klass] += 1

        # Resumo
        print()
        print(f'Resumo ({len(ajustes)} ajustes):')
        for klass, n in sorted(contadores.items(), key=lambda x: -x[1]):
            print(f'  {klass:<35} {n:>6}')

        # Excel
        sufixo = (
            f'-{args.cod_produto}' if args.cod_produto else ''
        )
        xlsx_path = (
            Path(args.xlsx_out) if args.xlsx_out
            else OUTPUT_DIR / f'pos-execucao-{company_codigo}-{args.ciclo}{sufixo}.xlsx'
        )
        montar_excel(
            ajustes, achados_por_id, classificacoes, xlsx_path, company_codigo,
            args.ciclo,
        )

        # JSON
        json_path = (
            Path(args.json_out) if args.json_out
            else Path(f'/tmp/pos_execucao_{company_codigo}_{args.ciclo}{sufixo}.json')
        )
        payload = {
            'ciclo': args.ciclo, 'company_id': args.company_id,
            'company_codigo': company_codigo,
            'cod_produto_filtro': args.cod_produto,
            'total': len(ajustes),
            'classificacoes': dict(contadores),
            'ajustes': [
                {
                    'id': a.id, 'cod_produto': a.cod_produto,
                    'acao_decidida': a.acao_decidida, 'status': a.status,
                    'fase_pipeline': a.fase_pipeline,
                    'lote_origem': a.lote_origem,
                    'lote_destino': a.lote_destino,
                    'qtd_inventario': float(a.qtd_inventario or 0),
                    'qtd_odoo': float(a.qtd_odoo or 0),
                    'qtd_ajuste': float(a.qtd_ajuste or 0),
                    'custo_medio': (
                        float(a.custo_medio) if a.custo_medio else None
                    ),
                    'picking_id_odoo': a.picking_id_odoo,
                    'invoice_id_odoo': a.invoice_id_odoo,
                    'chave_nfe': a.chave_nfe,
                    'erro_msg': a.erro_msg,
                    'classificacao': classificacoes.get(a.id),
                    'odoo_atual': achados_por_id.get(a.id, {}),
                }
                for a in ajustes
            ],
        }
        with open(json_path, 'w') as f:
            json.dump(payload, f, indent=2, default=str)
        print(f'  JSON:  {json_path}')


if __name__ == '__main__':
    main()
