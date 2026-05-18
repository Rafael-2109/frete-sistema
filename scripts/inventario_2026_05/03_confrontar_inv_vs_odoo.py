"""F7.3 — Confronto inventario fisico x estoque Odoo.

Aplica regras de lote:
- P6: prioridade de escolha de lote alvo = (1) lote inventariado,
      (2) MIGRACAO, (3) mais antigo (menor quant_id)
- P9: qty igual + lote diferente → APENAS_LOTE (caso de rename)

Inputs:
- /tmp/estoque_odoo_2026_05.json (gerado por 01)
- /tmp/inventario_fisico_2026_05.json (gerado por 02)

Outputs:
- docs/inventario-2026-05/07-relatorios/diff-inv-vs-odoo-{FB,CD,LF}.xlsx
- /tmp/diff_inventario_2026_05.json (consumido por 04_propor_ajustes)

Uso:
    python scripts/inventario_2026_05/03_confrontar_inv_vs_odoo.py [--dry-run]

Spec: docs/superpowers/plans/2026-05-17-ajuste-inventario-nacom-lf.md Task 7.3
"""
import argparse
import json
import os
import sys
from collections import defaultdict
from decimal import Decimal
from pathlib import Path

# sys.path para `from app import ...`
_THIS = Path(__file__).resolve()
sys.path.insert(0, str(_THIS.parents[2]))

import openpyxl  # noqa: E402

from app import create_app  # noqa: E402
from app.odoo.utils.connection import get_odoo_connection  # noqa: E402
from app.utils.timezone import agora_utc_naive  # noqa: E402

OUTPUT_DIR = str(_THIS.parents[2] / 'docs' / 'inventario-2026-05' / '07-relatorios')
COMPANIES = {1: 'FB', 4: 'CD', 5: 'LF'}

INPUT_ESTOQUE = '/tmp/estoque_odoo_2026_05.json'
INPUT_INV = '/tmp/inventario_fisico_2026_05.json'
OUTPUT_JSON = '/tmp/diff_inventario_2026_05.json'


def escolher_lote_alvo(
    quants_produto: list,
    lote_inv: str,
    usar_mais_novo: bool = False,
) -> dict:
    """P6: escolhe lote para ajustar entre os quants do produto+company.

    Prioridade:
        (1) lote inventariado (se especificado e existir nos quants)
        (2) MIGRACAO
        (3) mais antigo (menor quant_id) OU mais novo (maior quant_id)

    Args:
        usar_mais_novo: regra 3. Default False (P6 original: mais
            antigo). True quando inventario fisico NAO trouxe lote
            (decisao usuario 2026-05-17 — assume ultimo lote produzido
            eh o presente em estoque).
    """
    if not quants_produto:
        return {}

    if lote_inv:
        for q in quants_produto:
            if (q.get('lote_nome') or '') == lote_inv:
                return q

    for q in quants_produto:
        if (q.get('lote_nome') or '').upper() == 'MIGRACAO':
            return q

    # Regra 3: mais antigo (default) ou mais novo (path inventario sem lote)
    ordered = sorted(
        quants_produto, key=lambda q: q['quant_id'], reverse=usar_mais_novo
    )
    return ordered[0]


def _custo_medio_cod(quants: list) -> Decimal:
    """Custo medio ponderado (D004): usado quando lote alvo nao tem
    custo no Odoo (lote novo). Soma value/quantity dos quants com qty>0
    e value != 0.
    """
    total_val = Decimal('0')
    total_qty = Decimal('0')
    for q in quants:
        qty = Decimal(str(q.get('quantity', 0) or 0))
        val = Decimal(str(q.get('value', 0) or 0))
        if qty > 0 and val != 0:
            total_val += val
            total_qty += qty
    return total_val / total_qty if total_qty > 0 else Decimal('0')


def _checar_validade(
    validade_inv: str | None,
    lote_odoo_dict: dict,
) -> tuple:
    """Cross-check validade_inv (do inventario) vs expiration_date (do Odoo).

    Returns:
        (divergente: bool, msg: str|None)
        divergente=True se ambas existem e datas (YYYY-MM-DD) diferem.
        msg = None se nao diverge ou faltou dado.
    """
    if not validade_inv:
        return False, None
    val_odoo = lote_odoo_dict.get('expiration_date')
    if not val_odoo:
        return False, None
    # Odoo retorna 'YYYY-MM-DD HH:MM:SS' — comparar so a parte data
    val_odoo_date = str(val_odoo)[:10]
    if validade_inv != val_odoo_date:
        return True, f'validade_inv={validade_inv} vs odoo={val_odoo_date}'
    return False, None


def confrontar_company(
    quants_odoo: list, linhas_inv: list, cid: int
) -> tuple:
    """Retorna tupla (diffs, outliers) por (cod_produto, lote).

    Refator 2026-05-17 (planilha real):
    - Separa linhas_inv em (com lote) e (sem lote).
    - Linhas SEM lote: agregado e abatido contra lote MAIS NOVO do Odoo
      (P6.regra3 com usar_mais_novo=True).
    - Cross-check validade_inv vs Odoo expiration_date — flag
      validade_divergente=True quando ambas existem e diferem.

    Outliers (2026-05-17): cods nao-digito (ex: X105000022 produtos
    arquivados, COMP-ICMS-ENTRADA produtos fiscais) sao skipados do diff
    porque o pipeline fiscal depende de tipo_produto=int(cod[0]) para
    determinar CFOP/fiscal_position. Sao retornados em lista separada
    para revisao manual (mesma decisao F7.2 com 'C'/'S' em CD).
    """
    odoo_por_cod = defaultdict(list)
    for q in quants_odoo:
        if q.get('cod_produto'):
            odoo_por_cod[q['cod_produto']].append(q)

    inv_com_lote_por_cod = defaultdict(list)
    inv_sem_lote_por_cod = defaultdict(list)
    for linha in linhas_inv:
        cod = linha['cod_produto']
        if (linha.get('lote_inventariado') or '').strip():
            inv_com_lote_por_cod[cod].append(linha)
        else:
            inv_sem_lote_por_cod[cod].append(linha)

    # FIX 2026-05-17: agregar inv_com_lote por (cod, lote_inventariado).
    # Planilha real tem multiplas linhas mesmo cod+lote (ex: 23 pallets
    # cod=201030011 lote=160100-26 em CD). Sem agregacao, F7.3 gerava 23
    # diffs duplicados e F7.4 dedup descartava 22 (perda de 1895 ajustes).
    for cod_k, linhas_raw in list(inv_com_lote_por_cod.items()):
        agg_por_lote = {}
        for linha in linhas_raw:
            li = linha.get('lote_inventariado', '') or ''
            if li not in agg_por_lote:
                agg_por_lote[li] = {
                    'cod_produto': cod_k,
                    'lote_inventariado': li,
                    'qtd_inventario': Decimal('0'),
                    'tipo_produto': linha.get('tipo_produto'),
                    'validade_inv': None,
                    'linha_origem': [],
                }
            agg_por_lote[li]['qtd_inventario'] += Decimal(
                str(linha['qtd_inventario'])
            )
            lo = linha.get('linha_origem')
            if lo is not None:
                agg_por_lote[li]['linha_origem'].append(lo)
            if linha.get('validade_inv') and not agg_por_lote[li]['validade_inv']:
                agg_por_lote[li]['validade_inv'] = linha['validade_inv']
        for li, a in agg_por_lote.items():
            a['qtd_inventario'] = str(a['qtd_inventario'])
        inv_com_lote_por_cod[cod_k] = list(agg_por_lote.values())

    diffs = []
    outliers = []
    cods_processados = set()
    cods_todos = (
        set(inv_com_lote_por_cod.keys())
        | set(inv_sem_lote_por_cod.keys())
        | set(odoo_por_cod.keys())
    )

    for cod in cods_todos:
        if cod in cods_processados:
            continue
        cods_processados.add(cod)

        # Skip outliers: cod nao-digito (arquivados X* ou fiscais
        # COMP-ICMS-*) e tipo_produto fora de (1,2,3,4) (sem mapeamento
        # fiscal em operacoes_fiscais.py). Registra para revisao manual.
        motivo_outlier = None
        if not cod or not cod[0].isdigit():
            motivo_outlier = (
                'cod_nao_digito (produto arquivado X* ou fiscal)'
            )
        elif int(cod[0]) not in (1, 2, 3, 4):
            motivo_outlier = (
                f'tipo_produto={cod[0]} fora de (1,2,3,4) — sem '
                'mapeamento fiscal'
            )
        if motivo_outlier:
            quants_out = odoo_por_cod.get(cod, [])
            inv_out = (
                inv_com_lote_por_cod.get(cod, [])
                + inv_sem_lote_por_cod.get(cod, [])
            )
            total_odoo_out = sum(
                Decimal(str(q['quantity'])) for q in quants_out
            )
            total_inv_out = sum(
                Decimal(r['qtd_inventario']) for r in inv_out
            )
            total_val_out = sum(
                Decimal(str(q.get('value', 0) or 0)) for q in quants_out
            )
            outliers.append({
                'cod_produto': cod,
                'company_id': cid,
                'qtd_odoo': str(total_odoo_out),
                'valor_odoo': str(total_val_out),
                'qtd_inventario': str(total_inv_out),
                'em_inventario': bool(inv_out),
                'motivo': motivo_outlier,
            })
            continue

        quants = odoo_por_cod.get(cod, [])
        inv_com_lote = inv_com_lote_por_cod.get(cod, [])
        inv_sem_lote = inv_sem_lote_por_cod.get(cod, [])

        total_odoo = sum(Decimal(str(q['quantity'])) for q in quants)
        total_inv_com_lote = sum(
            Decimal(linha['qtd_inventario']) for linha in inv_com_lote
        )
        total_inv_sem_lote = sum(
            Decimal(linha['qtd_inventario']) for linha in inv_sem_lote
        )
        total_inv = total_inv_com_lote + total_inv_sem_lote

        # ============================================================
        # P9: mesma quantidade total + lotes diferentes → APENAS_LOTE
        # (apenas quando TODO o inv tem lote — sem lote nao se aplica)
        # ============================================================
        if (
            total_odoo == total_inv
            and total_odoo > 0
            and not inv_sem_lote
        ):
            lotes_odoo = {(q.get('lote_nome') or '') for q in quants}
            lotes_inv = {
                linha.get('lote_inventariado', '') for linha in inv_com_lote
            }
            if lotes_odoo != lotes_inv:
                diffs.append({
                    'cod_produto': cod,
                    'tipo_produto': int(cod[0]),
                    'company_id': cid,
                    'lote_inventariado': ','.join(sorted(lotes_inv)),
                    'lote_odoo': ','.join(sorted(lotes_odoo)),
                    'lote_origem': ','.join(sorted(lotes_odoo)),
                    'lote_destino': ','.join(sorted(lotes_inv)),
                    'qtd_inventario': str(total_inv),
                    'qtd_odoo': str(total_odoo),
                    'qtd_ajuste': '0',
                    'tipo_divergencia': 'APENAS_LOTE',
                })
                continue
            else:
                continue  # sem divergencia

        # ============================================================
        # D004: LF com saldo nos DOIS lados e lotes disjuntos
        # Renomear lotes Odoo (FIFO) ate cobrir saldo inv + diferenca
        # liquida vai como PERDA (sobra) ou INDUSTRIALIZACAO (falta).
        # Lote destino na FB para fantasmas = MIGRACAO (D005).
        # Custo medio dos lotes (D004): usado quando lote inv nao existe
        # no Odoo (custo zero local) ou diferenca liquida sem origem.
        # ============================================================
        lotes_odoo_set = {(q.get('lote_nome') or '') for q in quants}
        lotes_inv_set = {
            (l.get('lote_inventariado') or '') for l in inv_com_lote
        }
        custo_medio_cod = _custo_medio_cod(quants)
        if (
            cid == 5  # LF apenas por enquanto
            and total_odoo > 0
            and total_inv > 0
            and not lotes_odoo_set.intersection(lotes_inv_set)
            and not inv_sem_lote  # inv sem lote tratado abaixo
        ):
            # 1. Renomear FIFO ate cobrir min(odoo, inv)
            target_rename = min(total_odoo, total_inv)
            qty_renomeada = Decimal('0')
            lote_inv_alvo = next(iter(lotes_inv_set))
            # Estado: quants_restantes[i] = qty residual apos rename
            quants_sorted = sorted(quants, key=lambda x: x['quant_id'])
            qty_restante_por_lote = {}
            for q in quants_sorted:
                if qty_renomeada >= target_rename:
                    qty_restante_por_lote[id(q)] = (
                        q, Decimal(str(q['quantity']))
                    )
                    continue
                lote_o = q.get('lote_nome') or ''
                qty_q = Decimal(str(q['quantity']))
                if qty_q <= 0:
                    qty_restante_por_lote[id(q)] = (q, qty_q)
                    continue
                qty_take = min(qty_q, target_rename - qty_renomeada)
                custo_q = (
                    Decimal(str(q.get('value', 0) or 0)) / qty_q
                    if qty_q else custo_medio_cod
                )
                diffs.append({
                    'cod_produto': cod,
                    'tipo_produto': int(cod[0]),
                    'company_id': cid,
                    'lote_inventariado': lote_inv_alvo,
                    'lote_odoo': lote_o,
                    'lote_origem': lote_o,
                    'lote_destino': lote_inv_alvo,
                    'qtd_inventario': str(qty_take),
                    'qtd_odoo': str(qty_take),
                    'qtd_ajuste': '0',
                    'custo_medio': str(custo_q),
                    'tipo_divergencia': 'RENOMEAR_LOTE_PARCIAL',
                })
                qty_renomeada += qty_take
                qty_restante_por_lote[id(q)] = (q, qty_q - qty_take)

            # 2. Diferenca liquida
            diferenca = total_inv - total_odoo
            if diferenca > 0:
                # Falta na LF: vem da FB (INDUSTRIALIZACAO_FB_LF)
                diffs.append({
                    'cod_produto': cod,
                    'tipo_produto': int(cod[0]),
                    'company_id': cid,
                    'lote_inventariado': lote_inv_alvo,
                    'lote_odoo': '',
                    'lote_origem': 'MIGRACAO',  # vem do MIGRACAO da FB
                    'lote_destino': lote_inv_alvo,
                    'qtd_inventario': str(diferenca),
                    'qtd_odoo': '0',
                    'qtd_ajuste': str(diferenca),
                    'custo_medio': str(custo_medio_cod),
                    'tipo_divergencia': 'INVENTARIO_SEM_ODOO',
                })
            elif diferenca < 0:
                # Sobra na LF: vai para FB (PERDA_LF_FB) com lote_destino=MIGRACAO
                # Gera 1 diff por lote residual (preserva rastreio fiscal +
                # respeita limite VARCHAR(60) em lote_odoo/lote_origem)
                lotes_restantes = [
                    (q.get('lote_nome') or '', qty_r, q)
                    for (q, qty_r) in qty_restante_por_lote.values()
                    if qty_r > 0
                ]
                for lote_o, qty_r, q_origem in lotes_restantes:
                    custo_q = (
                        Decimal(str(q_origem.get('value', 0) or 0))
                        / Decimal(str(q_origem['quantity']))
                        if q_origem.get('quantity') else custo_medio_cod
                    )
                    diffs.append({
                        'cod_produto': cod,
                        'tipo_produto': int(cod[0]),
                        'company_id': cid,
                        'lote_inventariado': '',
                        'lote_odoo': lote_o[:60],
                        'lote_origem': lote_o[:60],
                        'lote_destino': 'MIGRACAO',  # FB
                        'qtd_inventario': '0',
                        'qtd_odoo': str(qty_r),
                        'qtd_ajuste': str(-qty_r),
                        'custo_medio': str(custo_q),
                        'tipo_divergencia': 'QUANTIDADE',
                    })
            continue  # nao cair no fluxo geral abaixo

        # ============================================================
        # Inv COM lote: 1 linha por lote_odoo (matching exato)
        # Lotes_odoo sem contraparte: gera diff QUANTIDADE (qty_inv=0)
        # APENAS se inv_sem_lote vazio — caso contrario sera absorvido
        # pelo bloco "Inv SEM lote" abaixo (evita diff duplicado).
        # ============================================================
        for q in quants:
            lote_odoo = q.get('lote_nome') or ''
            inv_match = next(
                (
                    linha for linha in inv_com_lote
                    if linha.get('lote_inventariado') == lote_odoo
                ),
                None,
            )
            # Skip se nao ha match E ha inv_sem_lote (sera tratado abaixo)
            if inv_match is None and inv_sem_lote:
                continue
            qty_inv = (
                Decimal(inv_match['qtd_inventario'])
                if inv_match else Decimal('0')
            )
            qty_odoo = Decimal(str(q['quantity']))
            if qty_inv != qty_odoo:
                custo_medio = (
                    str(Decimal(str(q.get('value', 0) or 0)) / qty_odoo)
                    if qty_odoo else '0'
                )
                diff_rec = {
                    'cod_produto': cod,
                    'tipo_produto': int(cod[0]),
                    'company_id': cid,
                    'lote_inventariado': (
                        (inv_match or {}).get('lote_inventariado', '')
                    ),
                    'lote_odoo': lote_odoo,
                    'qtd_inventario': str(qty_inv),
                    'qtd_odoo': str(qty_odoo),
                    'qtd_ajuste': str(qty_inv - qty_odoo),
                    'custo_medio': custo_medio,
                    'tipo_divergencia': 'QUANTIDADE',
                }
                # Cross-check validade (ajuste 2026-05-17)
                if inv_match:
                    div, msg = _checar_validade(
                        inv_match.get('validade_inv'), q
                    )
                    if div:
                        diff_rec['validade_divergente'] = True
                        diff_rec['validade_msg'] = msg
                        print(
                            f'  AVISO validade {cod} lote={lote_odoo}: {msg}'
                        )
                diffs.append(diff_rec)

        # Inv COM lote sem contraparte Odoo
        for linha in inv_com_lote:
            lote_inv = linha.get('lote_inventariado', '')
            tem_match = any(
                (q.get('lote_nome') or '') == lote_inv for q in quants
            )
            if not tem_match:
                diffs.append({
                    'cod_produto': cod,
                    'tipo_produto': int(cod[0]),
                    'company_id': cid,
                    'lote_inventariado': lote_inv,
                    'lote_odoo': '',
                    'qtd_inventario': linha['qtd_inventario'],
                    'qtd_odoo': '0',
                    'qtd_ajuste': linha['qtd_inventario'],
                    'tipo_divergencia': 'INVENTARIO_SEM_ODOO',
                })

        # ============================================================
        # Inv SEM lote: agregar qty + abater do lote MAIS NOVO do Odoo
        # (decisao usuario 2026-05-17 — assume ultimo lote produzido)
        # ============================================================
        if total_inv_sem_lote > 0:
            # Calcular qty Odoo nao-coberta pelo inv-com-lote
            qty_odoo_coberta = sum(
                Decimal(str(q['quantity']))
                for q in quants
                if any(
                    linha.get('lote_inventariado') == (q.get('lote_nome') or '')
                    for linha in inv_com_lote
                )
            )
            qty_odoo_livre = total_odoo - qty_odoo_coberta
            qty_ajuste = total_inv_sem_lote - qty_odoo_livre

            if qty_ajuste != 0:
                # Escolhe lote alvo = MAIS NOVO do Odoo
                quants_livres = [
                    q for q in quants
                    if not any(
                        linha.get('lote_inventariado') == (q.get('lote_nome') or '')
                        for linha in inv_com_lote
                    )
                ]
                lote_alvo = escolher_lote_alvo(
                    quants_livres if quants_livres else quants,
                    lote_inv='',
                    usar_mais_novo=True,
                )
                custo_medio_alvo = (
                    str(
                        Decimal(str(lote_alvo.get('value', 0) or 0))
                        / Decimal(str(lote_alvo['quantity']))
                    )
                    if lote_alvo and lote_alvo.get('quantity') else '0'
                )
                diffs.append({
                    'cod_produto': cod,
                    'tipo_produto': int(cod[0]),
                    'company_id': cid,
                    'lote_inventariado': '',
                    'lote_odoo': lote_alvo.get('lote_nome') or '',
                    'qtd_inventario': str(total_inv_sem_lote),
                    'qtd_odoo': str(qty_odoo_livre),
                    'qtd_ajuste': str(qty_ajuste),
                    'custo_medio': custo_medio_alvo,
                    'tipo_divergencia': 'QUANTIDADE_LOTE_INFERIDO',
                    'lote_inferido': True,
                    'linhas_inv_origem': [
                        linha.get('linha_origem') for linha in inv_sem_lote
                    ],
                })

    return diffs, outliers


def main(dry_run: bool) -> None:
    app = create_app()
    with app.app_context():
        os.makedirs(OUTPUT_DIR, exist_ok=True)

        for src in (INPUT_ESTOQUE, INPUT_INV):
            if not os.path.exists(src):
                raise FileNotFoundError(
                    f'Input ausente: {src}. Rode os scripts 01 e 02 antes.'
                )

        with open(INPUT_ESTOQUE) as f:
            estoque = json.load(f)
        with open(INPUT_INV) as f:
            inv = json.load(f)

        odoo = get_odoo_connection()
        total_diffs = []
        total_outliers = []
        cod_to_name: dict = {}  # populado dentro do loop

        for cid_str, c_estoque in estoque['companies'].items():
            cid = int(cid_str)
            quants_raw = c_estoque.get('quants', [])

            # Batch read para resolver cod_produto + lote_nome (P4)
            product_ids = list({
                q['product_id'][0]
                for q in quants_raw if q.get('product_id')
            })
            lot_ids = list({
                q['lot_id'][0]
                for q in quants_raw if q.get('lot_id')
            })
            produtos = (
                {p['id']: p for p in odoo.read(
                    'product.product', product_ids, ['default_code', 'name']
                )}
                if product_ids else {}
            )
            # Mapear cod -> nome (consolidado entre companies; codigo eh unico)
            cod_to_name.update({
                (p.get('default_code') or '').strip(): p.get('name') or ''
                for p in produtos.values()
                if p.get('default_code')
            })
            lotes = (
                {lo['id']: lo for lo in odoo.read(
                    'stock.lot', lot_ids, ['name', 'expiration_date']
                )}
                if lot_ids else {}
            )

            quants_odoo = []
            for q in quants_raw:
                pid = q['product_id'][0] if q.get('product_id') else None
                lid = q['lot_id'][0] if q.get('lot_id') else None
                lote_obj = lotes.get(lid, {}) if lid else {}
                quants_odoo.append({
                    'quant_id': q['id'],
                    'cod_produto': (
                        produtos.get(pid, {}).get('default_code', '') or ''
                    ),
                    'lote_nome': lote_obj.get('name') or '',
                    'expiration_date': lote_obj.get('expiration_date'),
                    'quantity': q['quantity'],
                    'value': q.get('value', 0),
                })

            linhas_inv = inv['companies'].get(
                str(cid), {}
            ).get('linhas', [])
            diffs, outliers = confrontar_company(quants_odoo, linhas_inv, cid)
            codigo = COMPANIES.get(cid, str(cid))
            print(
                f'\n{codigo} (company_id={cid}): {len(diffs)} divergencias'
            )
            if outliers:
                print(
                    f'  [AVISO] {len(outliers)} outliers cod nao-digito '
                    f'(skipados, ver outliers-cod-nao-digito.xlsx): '
                    f'{[o["cod_produto"] for o in outliers]}'
                )
            total_diffs.extend(diffs)
            total_outliers.extend(outliers)

            # Excel — colunas adicionadas em 2026-05-17:
            # lote_inferido (P6 mais novo), validade_divergente +
            # validade_msg (cross-check Odoo) + valor_movimentacao.
            # Numeros gravados como float; number_format pt-BR aplicado
            # (Excel BR exibe `,` como decimal e `.` como milhar).
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = codigo
            ws.append([
                'cod_produto', 'nome_produto', 'tipo_produto', 'company_id',
                'lote_inventariado', 'lote_odoo',
                'lote_origem', 'lote_destino',
                'qtd_inventario', 'qtd_odoo', 'qtd_ajuste',
                'custo_medio', 'valor_movimentacao', 'tipo_divergencia',
                'lote_inferido', 'validade_divergente', 'validade_msg',
            ])

            def _f(v, default=0.0):
                try:
                    return float(v) if v not in ('', None) else default
                except (ValueError, TypeError):
                    return default

            for d in diffs:
                qty_inv = _f(d.get('qtd_inventario'))
                qty_odoo = _f(d.get('qtd_odoo'))
                qty_aj = _f(d.get('qtd_ajuste'))
                custo = _f(d.get('custo_medio'))
                valor_mov = abs(qty_aj * custo)
                nome = cod_to_name.get(d['cod_produto'], '')
                ws.append([
                    d['cod_produto'], nome,
                    d['tipo_produto'], d['company_id'],
                    d['lote_inventariado'], d['lote_odoo'],
                    d.get('lote_origem', ''), d.get('lote_destino', ''),
                    qty_inv, qty_odoo, qty_aj,
                    custo, valor_mov, d['tipo_divergencia'],
                    'SIM' if d.get('lote_inferido') else '',
                    'SIM' if d.get('validade_divergente') else '',
                    d.get('validade_msg', ''),
                ])

            # number_format pt-BR — exibe virgula como decimal em Excel BR
            fmt_qty = '#,##0.0000'
            fmt_money = '#,##0.00'
            n_rows = ws.max_row
            for row in range(2, n_rows + 1):
                # nome_produto deslocou as colunas:
                # qty_inv, qty_odoo, qty_ajuste, custo (cols 9,10,11,12)
                for col in (9, 10, 11, 12):
                    ws.cell(row=row, column=col).number_format = fmt_qty
                ws.cell(row=row, column=13).number_format = fmt_money

            xlsx_path = os.path.join(
                OUTPUT_DIR, f'diff-inv-vs-odoo-{codigo}.xlsx'
            )
            wb.save(xlsx_path)
            print(f'  {xlsx_path}')

        # Excel separado para outliers (cod nao-digito skipados — produtos
        # arquivados X* ou fiscais COMP-ICMS-*). Revisao manual.
        if total_outliers:
            wb_out = openpyxl.Workbook()
            ws_out = wb_out.active
            ws_out.title = 'Outliers'
            ws_out.append([
                'cod_produto', 'nome_produto', 'company_id',
                'qtd_odoo', 'valor_odoo', 'qtd_inventario',
                'em_inventario', 'motivo',
            ])
            for o in total_outliers:
                nome = cod_to_name.get(o['cod_produto'], '')
                ws_out.append([
                    o['cod_produto'], nome, o['company_id'],
                    _f(o['qtd_odoo']), _f(o['valor_odoo']),
                    _f(o['qtd_inventario']),
                    'SIM' if o['em_inventario'] else 'NAO',
                    o['motivo'],
                ])
            n_out = ws_out.max_row
            for row in range(2, n_out + 1):
                ws_out.cell(row=row, column=4).number_format = '#,##0.0000'
                ws_out.cell(row=row, column=5).number_format = '#,##0.00'
                ws_out.cell(row=row, column=6).number_format = '#,##0.0000'
            outliers_path = os.path.join(
                OUTPUT_DIR, 'outliers-cod-nao-digito.xlsx'
            )
            wb_out.save(outliers_path)
            print(
                f'\n[OUTLIERS] {len(total_outliers)} cods nao-digito '
                f'salvos em {outliers_path}'
            )

        if not dry_run:
            with open(OUTPUT_JSON, 'w') as f:
                json.dump(
                    {
                        'diffs': total_diffs,
                        'outliers': total_outliers,
                        'timestamp': agora_utc_naive().isoformat(),
                    },
                    f, default=str, indent=2,
                )
            print(
                f'\nTotal: {len(total_diffs)} divergencias salvas em '
                f'{OUTPUT_JSON} (+ {len(total_outliers)} outliers)'
            )
        else:
            print(
                f'\n[DRY RUN] nao gravou {OUTPUT_JSON} '
                f'(total: {len(total_diffs)} divergencias, '
                f'{len(total_outliers)} outliers)'
            )


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--dry-run', action='store_true')
    args = parser.parse_args()
    main(args.dry_run)
