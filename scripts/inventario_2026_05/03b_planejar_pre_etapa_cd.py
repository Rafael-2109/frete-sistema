"""F7.5b (D007) — Planejar pre-etapa CD.

Le /tmp/estoque_odoo_2026_05.json + /tmp/inventario_fisico_2026_05.json,
chama PreEtapaEstoqueService por produto do CD, e gera plano consolidado.

Decisao D007: substitui TRANSFERIR_CD_FB + INDISPONIBILIZAR_LOTE/LOCAL
do CD por transferencias internas + residual FB→CD + ajuste positivo
puro. Onda 5 (nova).

Inputs:
- /tmp/estoque_odoo_2026_05.json  (gerado por script 01)
- /tmp/inventario_fisico_2026_05.json  (gerado por script 02)

Outputs:
- /tmp/plano_pre_etapa_cd.json
- docs/inventario-2026-05/07-relatorios/plano-pre-etapa-cd.xlsx

Uso:
    python scripts/inventario_2026_05/03b_planejar_pre_etapa_cd.py [--dry-run]

Spec: docs/inventario-2026-05/00-decisoes/D007-pre-etapa-cd-fb-minimizar-nf.md
"""
import argparse
import json
import os
import sys
from collections import defaultdict
from pathlib import Path

_THIS = Path(__file__).resolve()
sys.path.insert(0, str(_THIS.parents[2]))

import openpyxl  # noqa: E402

from app import create_app  # noqa: E402
from app.odoo.services.pre_etapa_estoque_service import (  # noqa: E402
    PreEtapaEstoqueService,
)
from app.odoo.utils.connection import get_odoo_connection  # noqa: E402
from app.utils.timezone import agora_utc_naive  # noqa: E402

OUTPUT_DIR = str(
    _THIS.parents[2] / 'docs' / 'inventario-2026-05' / '07-relatorios'
)

INPUT_ESTOQUE = '/tmp/estoque_odoo_2026_05.json'
INPUT_INV = '/tmp/inventario_fisico_2026_05.json'
OUTPUT_JSON = '/tmp/plano_pre_etapa_cd.json'
OUTPUT_XLSX = 'plano-pre-etapa-cd.xlsx'

CD_CID = 4
FB_CID = 1
CD_LOCATION = 32  # CD/Estoque (constants/locations.py)
FB_LOCATION = 8   # FB/Estoque


def enriquecer_quants(odoo, quants_raw, label=''):
    """Adiciona cod_produto, lote_nome aos quants do JSON.

    NAO busca reserved_quantity (custoso e nao bloqueante — warning em
    runtime via StockInternalTransferService). Assume 0.
    """
    product_ids = list({
        q['product_id'][0] for q in quants_raw if q.get('product_id')
    })
    lot_ids = list({
        q['lot_id'][0] for q in quants_raw if q.get('lot_id')
    })

    print(f'  {label}: lendo {len(product_ids)} produtos + {len(lot_ids)} lotes...')
    produtos = {p['id']: p for p in odoo.read(
        'product.product', product_ids, ['default_code', 'name'],
    )} if product_ids else {}
    lotes = {lo['id']: lo for lo in odoo.read(
        'stock.lot', lot_ids, ['name'],
    )} if lot_ids else {}

    out = []
    for q in quants_raw:
        pid = q['product_id'][0] if q.get('product_id') else None
        lid = q['lot_id'][0] if q.get('lot_id') else None
        produto = produtos.get(pid, {})
        lote = lotes.get(lid, {}) if lid else {}
        out.append({
            'quant_id': q['id'],
            'product_id': pid,
            'cod_produto': (produto.get('default_code') or '').strip(),
            'nome_produto': produto.get('name') or '',
            'lot_id': lid,
            'lote_nome': lote.get('name') or '',
            'location_id': (
                q['location_id'][0] if q.get('location_id') else None
            ),
            'quantity': float(q['quantity']),
            'reserved_quantity': 0.0,  # nao bloqueante — runtime valida
            'value': float(q.get('value', 0)),
        })
    return out


def plano_to_dicts(plano, lista_global):
    """Append dataclasses do plano (1 produto) na lista consolidada global."""
    for t in plano.transferencias_internas:
        lista_global['transferencias_internas'].append({
            'cod_produto': t.cod_produto,
            'company_id': t.company_id,
            'location_id': t.location_id,
            'qty': t.qty,
            'lot_id_origem': t.lot_id_origem,
            'lote_origem_nome': t.lote_origem_nome,
            'lote_destino_nome': t.lote_destino_nome,
            'tipo': t.tipo,
            'custo_medio': str(t.custo_medio),
        })
    for r in plano.residual_fb_cd:
        lista_global['residual_fb_cd'].append({
            'cod_produto': r.cod_produto,
            'qty': r.qty,
            'custo_medio': str(r.custo_medio),
            'lote_origem_fb_sugerido': r.lote_origem_fb_sugerido,
            'lote_destino_cd_nome': r.lote_destino_cd_nome,
        })
    for a in plano.ajustes_positivos_puros:
        lista_global['ajustes_positivos_puros'].append({
            'cod_produto': a.cod_produto,
            'company_id': a.company_id,
            'location_id': a.location_id,
            'qty': a.qty,
            'lote_destino_nome': a.lote_destino_nome,
            'custo_medio': str(a.custo_medio),
        })
    for w in plano.warnings:
        lista_global['warnings'].append(w)


def gerar_excel(plano_total, path, cod_to_name):
    """Gera Excel com 4 abas: Internas, Residual FB-CD, Positivos Puros, Warnings."""
    wb = openpyxl.Workbook()

    # Aba 1: Transferencias internas
    ws = wb.active
    ws.title = 'Internas'
    ws.append([
        'cod_produto', 'nome_produto', 'company_id', 'location_id', 'qty',
        'lot_id_origem', 'lote_origem', 'lote_destino', 'tipo', 'custo_medio',
        'valor_movimentacao',
    ])
    for t in plano_total['transferencias_internas']:
        cm = float(t['custo_medio'])
        ws.append([
            t['cod_produto'], cod_to_name.get(t['cod_produto'], ''),
            t['company_id'], t['location_id'], t['qty'],
            t['lot_id_origem'] or '', t['lote_origem_nome'],
            t['lote_destino_nome'], t['tipo'], cm, t['qty'] * cm,
        ])
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=5).number_format = '#,##0.0000'
        ws.cell(row=row, column=10).number_format = '#,##0.00'
        ws.cell(row=row, column=11).number_format = '#,##0.00'

    # Aba 2: Residual FB→CD
    ws2 = wb.create_sheet('Residual FB-CD')
    ws2.append([
        'cod_produto', 'nome_produto', 'qty', 'custo_medio',
        'lote_origem_fb_sugerido', 'lote_destino_cd_nome',
        'valor_movimentacao',
    ])
    for r in plano_total['residual_fb_cd']:
        cm = float(r['custo_medio'])
        ws2.append([
            r['cod_produto'], cod_to_name.get(r['cod_produto'], ''),
            r['qty'], cm,
            r['lote_origem_fb_sugerido'], r['lote_destino_cd_nome'],
            r['qty'] * cm,
        ])
    for row in range(2, ws2.max_row + 1):
        ws2.cell(row=row, column=3).number_format = '#,##0.0000'
        ws2.cell(row=row, column=4).number_format = '#,##0.00'
        ws2.cell(row=row, column=7).number_format = '#,##0.00'

    # Aba 3: Ajustes positivos puros
    ws3 = wb.create_sheet('Positivos Puros')
    ws3.append([
        'cod_produto', 'nome_produto', 'company_id', 'location_id', 'qty',
        'lote_destino', 'custo_medio', 'valor_movimentacao',
    ])
    for a in plano_total['ajustes_positivos_puros']:
        cm = float(a['custo_medio'])
        ws3.append([
            a['cod_produto'], cod_to_name.get(a['cod_produto'], ''),
            a['company_id'], a['location_id'], a['qty'],
            a['lote_destino_nome'], cm, a['qty'] * cm,
        ])
    for row in range(2, ws3.max_row + 1):
        ws3.cell(row=row, column=5).number_format = '#,##0.0000'
        ws3.cell(row=row, column=7).number_format = '#,##0.00'
        ws3.cell(row=row, column=8).number_format = '#,##0.00'

    # Aba 4: Warnings
    ws4 = wb.create_sheet('Warnings')
    ws4.append(['mensagem'])
    for w in plano_total['warnings']:
        ws4.append([w])

    wb.save(path)


def main(dry_run):
    app = create_app()
    with app.app_context():
        os.makedirs(OUTPUT_DIR, exist_ok=True)

        for src in (INPUT_ESTOQUE, INPUT_INV):
            if not os.path.exists(src):
                raise FileNotFoundError(
                    f'Input ausente: {src}. Rode scripts 01 e 02 antes.'
                )

        print('Carregando JSONs...')
        with open(INPUT_ESTOQUE) as f:
            estoque = json.load(f)
        with open(INPUT_INV) as f:
            inv = json.load(f)

        odoo = get_odoo_connection()

        print('\nEnriquecendo quants do Odoo (CD + FB)...')
        quants_cd = enriquecer_quants(
            odoo, estoque['companies'].get('4', {}).get('quants', []),
            label='CD',
        )
        quants_fb = enriquecer_quants(
            odoo, estoque['companies'].get('1', {}).get('quants', []),
            label='FB',
        )

        linhas_inv_cd = inv['companies'].get('4', {}).get('linhas', [])

        # Agregacao por cod
        cd_por_cod = defaultdict(list)
        for q in quants_cd:
            if q['cod_produto']:
                cd_por_cod[q['cod_produto']].append(q)
        fb_por_cod = defaultdict(list)
        for q in quants_fb:
            if q['cod_produto']:
                fb_por_cod[q['cod_produto']].append(q)
        inv_cd_por_cod = defaultdict(list)
        for linha in linhas_inv_cd:
            inv_cd_por_cod[linha['cod_produto']].append(linha)

        cod_to_name = {
            q['cod_produto']: q['nome_produto']
            for q in (quants_cd + quants_fb)
            if q['cod_produto']
        }

        # Universo de produtos: tudo que tem inv CD OU quants CD
        cods_universo = set(inv_cd_por_cod.keys()) | set(cd_por_cod.keys())
        cods_validos = [
            c for c in cods_universo
            if c and c[0].isdigit() and int(c[0]) in (1, 2, 3, 4)
        ]
        cods_outliers = [c for c in cods_universo if c not in cods_validos]
        print(f'\n{len(cods_validos)} produtos validos a planejar (CD).')
        if cods_outliers:
            print(
                f'  Outliers skipados (cod nao-digito ou tipo fora 1-4): '
                f'{len(cods_outliers)} (ex: {cods_outliers[:5]})'
            )

        # Planejar produto a produto
        svc = PreEtapaEstoqueService(odoo=odoo)
        plano_total = {
            'transferencias_internas': [],
            'residual_fb_cd': [],
            'ajustes_positivos_puros': [],
            'warnings': [],
        }
        n_processados = 0
        n_sem_mudanca = 0
        for i, cod in enumerate(sorted(cods_validos), 1):
            if i % 100 == 0:
                print(f'  ... planejando {i}/{len(cods_validos)}')
            quants_cd_cod = cd_por_cod.get(cod, [])
            inv_cd_cod = inv_cd_por_cod.get(cod, [])
            quants_fb_cod = fb_por_cod.get(cod, [])

            if not quants_cd_cod and not inv_cd_cod:
                continue

            plano = svc.planejar(
                cod_produto=cod,
                company_id=CD_CID,
                location_id=CD_LOCATION,
                quants_odoo=quants_cd_cod,
                linhas_inv=inv_cd_cod,
                quants_fb_disponivel=quants_fb_cod,
            )
            if not (
                plano.transferencias_internas
                or plano.residual_fb_cd
                or plano.ajustes_positivos_puros
            ):
                n_sem_mudanca += 1
                continue
            n_processados += 1
            # Prefix cod nos warnings para rastrear
            plano.warnings = [f'[{cod}] {w}' for w in plano.warnings]
            plano_to_dicts(plano, plano_total)

        # Salvar JSON
        if not dry_run:
            with open(OUTPUT_JSON, 'w') as f:
                json.dump(
                    {
                        **plano_total,
                        'produtos_processados': n_processados,
                        'produtos_sem_mudanca': n_sem_mudanca,
                        'timestamp': agora_utc_naive().isoformat(),
                        'company_id': CD_CID,
                    },
                    f, default=str, indent=2,
                )
            print(f'\nPlano salvo em {OUTPUT_JSON}')
            xlsx_path = os.path.join(OUTPUT_DIR, OUTPUT_XLSX)
            gerar_excel(plano_total, xlsx_path, cod_to_name)
            print(f'Excel salvo em {xlsx_path}')
        else:
            print(f'\n[DRY-RUN] nao salvou {OUTPUT_JSON}')

        # Resumo
        n_pos = sum(
            1 for t in plano_total['transferencias_internas']
            if t['tipo'] == 'POS'
        )
        n_neg = sum(
            1 for t in plano_total['transferencias_internas']
            if t['tipo'] == 'NEG'
        )
        print('\n=========== RESUMO ===========')
        print(f'Produtos processados:        {n_processados}')
        print(f'Produtos sem mudanca:        {n_sem_mudanca}')
        print(f'Transf internas POS:         {n_pos}')
        print(f'Transf internas NEG:         {n_neg}')
        print(f'Residual FB→CD (NF):         {len(plano_total["residual_fb_cd"])}')
        print(f'Ajustes positivos puros:     {len(plano_total["ajustes_positivos_puros"])}')
        print(f'Warnings (reserva, etc):     {len(plano_total["warnings"])}')


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--dry-run', action='store_true')
    args = parser.parse_args()
    main(args.dry_run)
