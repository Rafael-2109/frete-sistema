"""Serviço da Contagem Cíclica: cria base, casa o reupload por QUANT, calcula
e classifica os ajustes, gera preview e contabiliza.

Regras (spec §6):
- A planilha reenviada É o escopo: linha presente = inventariada; ausente = intocada.
- Linha presente com CONTAGEM vazia ⇒ 0 (zera o fantasma).
- Linha que não casa com nenhum quant da base ⇒ LOTE_NOVO (qtd_esperada=0).
- DOIS ajustes com semânticas DISTINTAS (não confundir):
    • ajuste            = contagem − qtd_esperada → delta a APLICAR NO ODOO (skills).
      Define a `classe`. Derivado SEMPRE da contagem; ignora a coluna AJUSTE.
    • ajuste_inventario = valor LITERAL da coluna AJUSTE (autoritativo; vazio = 0)
      → delta somado ao último inventário na coluna INV/MOV do Confronto.
- Classes (precedência): SEM_AJUSTE → LOTE_NOVO → NEGATIVO → RESERVA_FANTASMA → NORMAL.

Funções puras (`calcular_linha`, `classificar`) testáveis sem Odoo/DB.
Spec: docs/superpowers/specs/2026-05-31-inventario-ciclico-contagem-ajustes-design.md
"""
import unicodedata
from decimal import Decimal, InvalidOperation
from typing import IO, Dict, List, Optional

import openpyxl

from app import db
from app.inventario.models import ContagemInventario, ContagemInventarioItem
from app.inventario.services.extracao_quant_service import (
    ExtracaoQuantService, norm_cod, norm_lote,
)
from app.utils.timezone import agora_utc_naive


ZERO = Decimal('0')

HEADER_ALIASES = {
    'location_name': 'location_name', 'local': 'location_name',
    'location': 'location_name', 'localizacao': 'location_name',
    'cod': 'cod_produto', 'codigo': 'cod_produto', 'cod_produto': 'cod_produto',
    'lote': 'lote',
    'contagem': 'contagem', 'contado': 'contagem', 'qtd_contada': 'contagem',
    # Coluna AJUSTE (autoritativa) → ajuste_inventario (delta p/ a coluna INV/MOV
    # do Confronto). Opcional; vazia = 0. NÃO confundir com `contagem`.
    'ajuste': 'ajuste',
}


def _norm_header(h) -> str:
    if not h:
        return ''
    s = str(h).strip().lower()
    s = unicodedata.normalize('NFKD', s)
    return ''.join(c for c in s if not unicodedata.combining(c))


def _to_decimal(v) -> Optional[Decimal]:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return Decimal(str(v).strip().replace(',', '.'))
    except (InvalidOperation, ValueError):
        raise ValueError(f'valor numérico inválido: {v!r}')


def classificar(qtd_esperada: Decimal, reservado_esperado: Decimal,
                ajuste: Decimal, is_lote_novo: bool) -> str:
    """Classe que define qual átomo gestor-estoque-odoo aplica (spec §6.3)."""
    if ajuste == ZERO:
        return 'SEM_AJUSTE'
    if is_lote_novo:
        return 'LOTE_NOVO'
    if qtd_esperada < ZERO:
        return 'NEGATIVO'
    if ajuste < ZERO and reservado_esperado > ZERO:
        return 'RESERVA_FANTASMA'
    return 'NORMAL'


def calcular_linha(item_base: Optional[dict], contagem: Optional[Decimal],
                   ajuste_manual: Optional[Decimal] = None) -> dict:
    """Lógica PURA: dado o item da base (ou None p/ lote novo), a contagem e o
    ajuste manual da coluna AJUSTE, devolve qtd_esperada, reservado_esperado e os
    DOIS ajustes (com a classe).

    - contagem None ⇒ 0 (linha presente, célula vazia) ⇒ `ajuste` = 0 − qtd_esperada.
    - `ajuste`            = contagem − qtd_esperada → delta p/ Odoo. Define a classe.
    - `ajuste_inventario` = `ajuste_manual` (coluna AJUSTE), autoritativo; None/vazio
      ⇒ 0 (sem impacto no Confronto). NUNCA derivado da contagem.
    """
    cont = contagem if contagem is not None else ZERO
    is_nova = item_base is None
    qtd_esp = ZERO if is_nova else Decimal(str(item_base.get('qtd_esperada') or 0))
    res_esp = ZERO if is_nova else Decimal(str(item_base.get('reservado_esperado') or 0))
    ajuste = cont - qtd_esp
    ajuste_inventario = ajuste_manual if ajuste_manual is not None else ZERO
    classe = classificar(qtd_esp, res_esp, ajuste, is_nova)
    return {
        'contagem': cont, 'qtd_esperada': qtd_esp,
        'reservado_esperado': res_esp, 'ajuste': ajuste,
        'ajuste_inventario': ajuste_inventario,
        'classe': classe, 'is_nova': is_nova,
    }


class ContagemService:

    # ---------------------------------------------------------------- código
    @staticmethod
    def _gerar_codigo(empresa: str) -> str:
        hoje = agora_utc_naive().date()
        prefixo = f'CONT-{hoje.isoformat()}-{empresa}-'
        n = (ContagemInventario.query
             .filter(ContagemInventario.codigo.like(prefixo + '%')).count())
        return f'{prefixo}{n + 1:02d}'

    # ------------------------------------------------- criar + gerar base (T0)
    @staticmethod
    def criar_e_gerar_base(empresa: str, filtro_locais: Optional[List[str]] = None,
                           filtro_codigos: Optional[List[str]] = None,
                           incluir_indisponivel: bool = False,
                           descricao: Optional[str] = None,
                           criado_por: Optional[str] = None,
                           odoo=None) -> ContagemInventario:
        """Extrai os quants do Odoo e persiste a contagem + itens (BASE_GERADA).

        Atômico: se a extração falhar, a exceção sobe e o caller faz rollback —
        a contagem só existe quando a base é gerada (sem RASCUNHO órfão).
        """
        empresa = (empresa or '').strip().upper()
        linhas = ExtracaoQuantService.extrair(
            empresa, filtro_locais=filtro_locais, filtro_codigos=filtro_codigos,
            incluir_indisponivel=incluir_indisponivel, odoo=odoo,
        )

        contagem = ContagemInventario(
            codigo=ContagemService._gerar_codigo(empresa),
            empresa=empresa,
            filtro_locais=filtro_locais or None,
            filtro_codigos=filtro_codigos or None,
            incluir_indisponivel=bool(incluir_indisponivel),
            data_base=agora_utc_naive(),
            status='BASE_GERADA',
            descricao=descricao,
            tot_itens=len(linhas),
            criado_em=agora_utc_naive(),
            criado_por=criado_por,
        )
        db.session.add(contagem)
        db.session.flush()  # obtém contagem.id

        for r in linhas:
            db.session.add(ContagemInventarioItem(
                contagem_id=contagem.id,
                location_name=r['location_name'],
                location_id=r.get('location_id'),
                local_tipo=r.get('local_tipo'),
                is_migracao=bool(r.get('is_migracao')),
                cod_produto=r['cod_produto'],
                nome_produto=r.get('nome_produto'),
                lote=r.get('lote') or '',
                company_id=r.get('company_id'),
                qtd_esperada=r.get('qtd', ZERO),
                reservado_esperado=r.get('reservado', ZERO),
                contagem=None,
                ajuste=ZERO,
                classe=None,
            ))
        db.session.flush()  # commit fica para o caller (route)
        return contagem

    # ------------------------------------------------------------ parse xlsx
    @staticmethod
    def parse_planilha(file_storage: IO) -> List[dict]:
        """Lê o xlsx preenchido. Retorna
        [{location_name, cod_produto, lote, contagem, ajuste_manual}].

        - `contagem` (coluna CONTAGEM) → físico; gera o `ajuste` p/ Odoo.
        - `ajuste_manual` (coluna AJUSTE, opcional) → autoritativo p/ a coluna
          INV/MOV do Confronto. None se a coluna não existe ou a célula está vazia.
          Aceita negativo (ao contrário da contagem).

        Levanta ValueError com mensagem clara em caso de header/valor inválido.
        """
        wb = openpyxl.load_workbook(file_storage, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            raise ValueError('Planilha vazia.')

        col_idx: Dict[str, int] = {}
        for i, h in enumerate(header):
            key = HEADER_ALIASES.get(_norm_header(h))
            if key and key not in col_idx:
                col_idx[key] = i

        faltando = [c for c in ('location_name', 'cod_produto', 'contagem')
                    if c not in col_idx]
        if faltando:
            raise ValueError(
                f'Colunas obrigatórias ausentes: {faltando}. '
                f'Esperado: location_name, cod (cod_produto), CONTAGEM (lote opcional).')

        out: List[dict] = []
        for nrow, row in enumerate(rows, start=2):
            loc = row[col_idx['location_name']]
            cod_raw = row[col_idx['cod_produto']]
            if (loc is None or not str(loc).strip()) and \
               (cod_raw is None or not str(cod_raw).strip()):
                continue  # linha em branco
            cod = norm_cod(cod_raw)
            if not cod:
                continue
            lote = norm_lote(row[col_idx['lote']]) if 'lote' in col_idx else ''
            try:
                contagem = _to_decimal(row[col_idx['contagem']])
            except ValueError as exc:
                raise ValueError(f'Linha {nrow}: {exc}')
            if contagem is not None and contagem < ZERO:
                raise ValueError(
                    f'Linha {nrow}: contagem negativa ({contagem}) não é permitida '
                    f'(físico não é negativo).')
            # Coluna AJUSTE (opcional, autoritativa p/ o Confronto). Aceita negativo.
            ajuste_manual = None
            if 'ajuste' in col_idx:
                try:
                    ajuste_manual = _to_decimal(row[col_idx['ajuste']])
                except ValueError as exc:
                    raise ValueError(f'Linha {nrow}: coluna AJUSTE — {exc}')
            out.append({
                'location_name': str(loc).strip(),
                'cod_produto': cod, 'lote': lote, 'contagem': contagem,
                'ajuste_manual': ajuste_manual,
                '_nrow': nrow,
            })
        return out

    # --------------------------------------------------- casar + calcular
    @staticmethod
    def _itens_base_dict(contagem_id: int) -> Dict[tuple, ContagemInventarioItem]:
        itens = ContagemInventarioItem.query.filter_by(contagem_id=contagem_id).all()
        return {(it.location_name, it.cod_produto, it.lote or ''): it for it in itens}

    @staticmethod
    def _processar(contagem: ContagemInventario, linhas: List[dict]) -> dict:
        """Casa as linhas da planilha com a base (por quant) e calcula ajuste/classe.

        Retorna {'linhas': [...], 'resumo': {...}} SEM gravar.
        Última linha vence em caso de tripla duplicada (com aviso).
        """
        base = ContagemService._itens_base_dict(contagem.id)
        vistos = {}
        avisos = []
        resultado: List[dict] = []

        for ln in linhas:
            key = (ln['location_name'], ln['cod_produto'], ln['lote'] or '')
            if key in vistos:
                avisos.append(f"Linha {ln['_nrow']}: tripla duplicada "
                              f"{key} — última prevalece.")
            vistos[key] = ln

        for key, ln in vistos.items():
            it = base.get(key)
            item_base = None if it is None else {
                'qtd_esperada': it.qtd_esperada,
                'reservado_esperado': it.reservado_esperado,
            }
            calc = calcular_linha(item_base, ln['contagem'], ln.get('ajuste_manual'))
            resultado.append({
                'location_name': ln['location_name'],
                'cod_produto': ln['cod_produto'],
                'lote': ln['lote'] or '',
                'nome_produto': (it.nome_produto if it else None),
                'location_id': (it.location_id if it else None),
                'company_id': (it.company_id if it else None),
                'qtd_esperada': calc['qtd_esperada'],
                'reservado_esperado': calc['reservado_esperado'],
                'contagem': calc['contagem'],
                'ajuste': calc['ajuste'],
                'ajuste_inventario': calc['ajuste_inventario'],
                'classe': calc['classe'],
                'is_nova': calc['is_nova'],
            })

        resumo = ContagemService._resumo(resultado)
        resumo['avisos'] = avisos
        return {'linhas': resultado, 'resumo': resumo}

    @staticmethod
    def _resumo(linhas: List[dict]) -> dict:
        """Resumo dos DOIS impactos (não confundir):
        - tot_ajuste_pos/neg     → impacto no ODOO (campo `ajuste`).
        - tot_ajuste_inv_pos/neg → impacto na coluna INV/MOV do Confronto
          (campo `ajuste_inventario`, vindo da coluna AJUSTE).
        """
        por_classe: Dict[str, int] = {}
        tot_pos = tot_neg = ZERO
        tot_inv_pos = tot_inv_neg = ZERO
        tot_com_ajuste = qt_novos = 0
        for r in linhas:
            por_classe[r['classe']] = por_classe.get(r['classe'], 0) + 1
            aj = r['ajuste']
            if aj != ZERO:
                tot_com_ajuste += 1
            if aj > ZERO:
                tot_pos += aj
            elif aj < ZERO:
                tot_neg += aj
            aji = r.get('ajuste_inventario') or ZERO
            if aji > ZERO:
                tot_inv_pos += aji
            elif aji < ZERO:
                tot_inv_neg += aji
            if r['is_nova']:
                qt_novos += 1
        return {
            'tot_itens': len(linhas),
            'tot_com_ajuste': tot_com_ajuste,
            'tot_ajuste_pos': tot_pos,
            'tot_ajuste_neg': tot_neg,
            'tot_ajuste_inv_pos': tot_inv_pos,
            'tot_ajuste_inv_neg': tot_inv_neg,
            'qt_lotes_novos': qt_novos,
            'por_classe': por_classe,
        }

    @staticmethod
    def preview_reupload(contagem_id: int, file_storage: IO) -> dict:
        """Calcula o impacto SEM gravar (preview de impacto antes de confirmar)."""
        contagem = ContagemInventario.query.get(contagem_id)
        if contagem is None:
            raise ValueError(f'Contagem {contagem_id} não encontrada.')
        linhas = ContagemService.parse_planilha(file_storage)
        return ContagemService._processar(contagem, linhas)

    @staticmethod
    def confirmar_reupload(contagem_id: int, file_storage: IO) -> dict:
        """Grava contagem/ajuste/classe nos itens (e cria lotes novos);
        status ⇒ CONTABILIZADA; atualiza resumo. Commit fica para o caller."""
        contagem = ContagemInventario.query.get(contagem_id)
        if contagem is None:
            raise ValueError(f'Contagem {contagem_id} não encontrada.')
        linhas = ContagemService.parse_planilha(file_storage)
        proc = ContagemService._processar(contagem, linhas)

        base = ContagemService._itens_base_dict(contagem.id)
        for r in proc['linhas']:
            key = (r['location_name'], r['cod_produto'], r['lote'] or '')
            it = base.get(key)
            if it is None:  # LOTE_NOVO — cria item
                it = ContagemInventarioItem(
                    contagem_id=contagem.id,
                    location_name=r['location_name'],
                    location_id=r.get('location_id'),
                    local_tipo=None,
                    is_migracao=False,
                    cod_produto=r['cod_produto'],
                    nome_produto=r.get('nome_produto'),
                    lote=r['lote'] or '',
                    company_id=r.get('company_id'),
                    qtd_esperada=ZERO,
                    reservado_esperado=ZERO,
                )
                db.session.add(it)
            it.contagem = r['contagem']
            it.ajuste = r['ajuste']
            it.ajuste_inventario = r['ajuste_inventario']
            it.classe = r['classe']

        res = proc['resumo']
        contagem.status = 'CONTABILIZADA'
        contagem.tot_com_ajuste = res['tot_com_ajuste']
        contagem.tot_ajuste_pos = res['tot_ajuste_pos']
        contagem.tot_ajuste_neg = res['tot_ajuste_neg']
        contagem.qt_lotes_novos = res['qt_lotes_novos']
        db.session.flush()
        return proc
