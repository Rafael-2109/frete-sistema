"""Parser do xlsx de inventário base (3 abas FB/CD/LF).

Reaproveita lógica do scripts/inventario_2026_05/02_carregar_inventario_xlsx.py.
"""
from decimal import Decimal, InvalidOperation
from typing import IO, Dict
import openpyxl
from app import db
from app.inventario.models import InventarioBase


TIPOS_CODIGO_ACEITOS = {'1', '2', '3', '4'}

HEADER_ALIASES = {
    'codigo': 'cod_produto',
    'cod': 'cod_produto',
    'cod_produto': 'cod_produto',
    'lote': 'lote',
    'qtd': 'qtd',
    'quantidade': 'qtd',
    'qtd_contada': 'qtd',
    'descricao': 'nome_produto',
    'produto': 'nome_produto',
    'nome_produto': 'nome_produto',
}

EMPRESAS_ESPERADAS = ('FB', 'CD', 'LF')


class InventarioLoader:
    """Carrega xlsx em InventarioBase, substituindo linhas do ciclo."""

    @staticmethod
    def carregar(ciclo_id: int, file_storage: IO, criado_por: str) -> Dict:
        """Parse xlsx + DELETE linhas antigas do ciclo + INSERT novas.

        Returns: {'inseridos': N, 'pulados': M, 'erros': [str, ...]}
        """
        wb = openpyxl.load_workbook(file_storage, data_only=True)
        abas_presentes = set(wb.sheetnames)
        abas_faltando = set(EMPRESAS_ESPERADAS) - abas_presentes
        if abas_faltando:
            return {
                'inseridos': 0, 'pulados': 0,
                'erros': [f'Abas faltando: {sorted(abas_faltando)}. '
                          f'Esperado: FB, CD, LF.'],
            }

        InventarioBase.query.filter_by(ciclo_id=ciclo_id).delete()
        db.session.flush()

        # Agrega (cod, empresa) -> (qtd_total, nome) para lidar com duplicatas
        agregado = {}  # (empresa, cod) -> {'qtd': Decimal, 'nome': str}
        inseridos = pulados = 0
        erros = []
        for empresa in EMPRESAS_ESPERADAS:
            ws = wb[empresa]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration:
                erros.append(f'Aba {empresa} vazia.')
                continue

            col_idx = {}
            for i, h in enumerate(header):
                if not h:
                    continue
                key = HEADER_ALIASES.get(str(h).strip().lower())
                if key:
                    col_idx[key] = i

            if 'cod_produto' not in col_idx or 'qtd' not in col_idx:
                erros.append(f'Aba {empresa}: faltam colunas CODIGO/QTD.')
                continue

            for nrow, row in enumerate(rows_iter, start=2):
                cod = row[col_idx['cod_produto']]
                if cod is None:
                    continue
                cod = str(cod).strip()
                if not cod:
                    continue
                if not cod[0] in TIPOS_CODIGO_ACEITOS:
                    pulados += 1
                    erros.append(f'Aba {empresa} linha {nrow}: '
                                 f'cod_produto={cod} pulado (não começa com 1-4)')
                    continue

                qtd_raw = row[col_idx['qtd']]
                try:
                    qtd = Decimal(str(qtd_raw or 0))
                except (InvalidOperation, ValueError):
                    erros.append(f'Aba {empresa} linha {nrow}: qtd inválida={qtd_raw}')
                    continue
                if qtd < 0:
                    erros.append(f'Aba {empresa} linha {nrow}: '
                                 f'qtd negativa={qtd} para cod={cod}')
                    continue

                nome = row[col_idx['nome_produto']] if 'nome_produto' in col_idx else None
                nome_str = str(nome).strip() if nome else None
                key = (empresa, cod)
                if key in agregado:
                    agregado[key]['qtd'] += qtd
                    if nome_str and not agregado[key]['nome']:
                        agregado[key]['nome'] = nome_str
                else:
                    agregado[key] = {'qtd': qtd, 'nome': nome_str}

        # Persist agregado
        for (empresa, cod), v in agregado.items():
            db.session.add(InventarioBase(
                ciclo_id=ciclo_id, cod_produto=cod,
                nome_produto=v['nome'], empresa=empresa, qtd=v['qtd'],
            ))
            inseridos += 1

        db.session.flush()
        return {'inseridos': inseridos, 'pulados': pulados, 'erros': erros}
