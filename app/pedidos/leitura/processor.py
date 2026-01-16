"""
Processador principal para pedidos em PDF de redes de atacarejo
Coordena extração, validação e inserção no banco

Fluxo:
1. Upload PDF → Armazena no S3
2. Identificador → Detecta Rede + Tipo
3. Extrator específico → Extrai dados
4. Conversão De-Para → Código Nacom
5. Validação de preços → vs TabelaRede
6. Revisão/Aprovação → Interface
7. Inserção Odoo → sale.order via XML-RPC
"""

from typing import Dict, List, Any, Optional
import pandas as pd
from decimal import Decimal

from .atacadao import AtacadaoExtractor
from .atacadao_pedido import AtacadaoPedidoExtractor
from .assai import AssaiExtractor
from .identificador import identificar_documento, IdentificacaoDocumento


class PedidoProcessor:
    """Processador principal de pedidos de redes de atacarejo"""

    # Mapeamento de formatos para extratores
    # Chave: REDE_TIPO (ex: ATACADAO_PROPOSTA)
    EXTRACTORS = {
        # Atacadão
        'atacadao_proposta': AtacadaoExtractor,
        'atacadao_pedido': AtacadaoPedidoExtractor,
        # Aliases para compatibilidade
        'atacadao': AtacadaoExtractor,  # Default para proposta
        'atacadão': AtacadaoExtractor,
        # Assaí/Sendas
        'assai_pedido': AssaiExtractor,
        'assai_proposta': AssaiExtractor,  # Usa mesmo extractor
        'assai': AssaiExtractor,  # Default
        'sendas': AssaiExtractor,  # Alias
        'sendas_pedido': AssaiExtractor,
        # Futuros extratores
        # 'tenda_proposta': TendaPropostaExtractor,
        # 'tenda_pedido': TendaPedidoExtractor,
    }

    def __init__(self):
        self.errors = []
        self.warnings = []
        self.processed_count = 0
        self.failed_count = 0
        self.identificacao: Optional[IdentificacaoDocumento] = None

    def process_file(self,
                     file_path: str,
                     formato: str = 'auto',
                     validate: bool = True) -> Dict[str, Any]:
        """
        Processa um arquivo PDF de rede de atacarejo

        Args:
            file_path: Caminho do arquivo PDF
            formato: Formato do PDF ('atacadao_proposta', 'atacadao_pedido', etc) ou 'auto' para detectar
            validate: Se deve validar os dados

        Returns:
            Dict com resultado do processamento incluindo identificação
        """
        result = {
            'success': False,
            'data': [],
            'errors': [],
            'warnings': [],
            'summary': {},
            'identificacao': None  # Novo: informações de identificação
        }

        try:
            # Detecta formato se necessário
            if formato == 'auto':
                self.identificacao = identificar_documento(file_path)
                result['identificacao'] = {
                    'rede': self.identificacao.rede,
                    'tipo': self.identificacao.tipo,
                    'numero_documento': self.identificacao.numero_documento,
                    'confianca': self.identificacao.confianca
                }

                # Monta chave do extrator
                formato = f"{self.identificacao.rede.lower()}_{self.identificacao.tipo.lower()}"

                if self.identificacao.rede == 'DESCONHECIDA':
                    result['errors'].append("Não foi possível identificar a rede do documento")
                    return result

                if self.identificacao.tipo == 'DESCONHECIDO':
                    result['errors'].append("Não foi possível identificar o tipo do documento (Proposta/Pedido)")
                    return result

            # Obtém extrator apropriado
            extractor_class = self.EXTRACTORS.get(formato.lower())
            if not extractor_class:
                # Tenta fallback para formato sem tipo
                formato_base = formato.split('_')[0] if '_' in formato else formato
                extractor_class = self.EXTRACTORS.get(formato_base.lower())

            if not extractor_class:
                result['errors'].append(f"Formato '{formato}' não suportado")
                return result

            # Extrai dados
            extractor = extractor_class()
            data = extractor.extract(file_path)

            if not data:
                result['errors'].append("Nenhum dado foi extraído do PDF")
                result['errors'].extend(extractor.errors)
                return result

            # Valida dados se necessário
            if validate:
                data = self._validate_data(data)

            # Prepara resultado
            result['success'] = True
            result['data'] = data
            result['errors'] = extractor.errors
            result['warnings'] = extractor.warnings
            result['summary'] = self._generate_summary(data)

            # Adiciona informações de identificação ao summary
            if self.identificacao:
                result['summary']['rede'] = self.identificacao.rede
                result['summary']['tipo_documento'] = self.identificacao.tipo
                result['summary']['numero_documento'] = self.identificacao.numero_documento

        except Exception as e:
            import traceback
            result['errors'].append(f"Erro ao processar arquivo: {str(e)}")
            result['traceback'] = traceback.format_exc()

        return result

    def _validate_data(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Valida e limpa os dados extraídos
        """
        validated = []

        for item in data:
            # Remove valores None
            item = {k: v for k, v in item.items() if v is not None}

            # Validações básicas
            if not item.get('codigo'):
                self.warnings.append(f"Item sem código: {item.get('descricao', 'sem descrição')}")
                continue

            if item.get('quantidade', 0) <= 0:
                self.warnings.append(f"Quantidade inválida para {item['codigo']}")
                continue

            valor_unitario = item.get('valor_unitario', 0)
            if isinstance(valor_unitario, Decimal):
                valor_unitario = float(valor_unitario)
            if valor_unitario <= 0:
                self.warnings.append(f"Valor inválido para {item['codigo']}")
                continue

            validated.append(item)

        return validated

    def _generate_summary(self, data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Gera resumo dos dados processados agrupados por CNPJ
        """
        try:
            if not data:
                return {}

            df = pd.DataFrame(data)

            summary = {
                'total_itens': len(data),
                'total_filiais': df['cnpj_filial'].nunique() if 'cnpj_filial' in df else 0,
                'total_produtos': df['codigo'].nunique() if 'codigo' in df else 0,
                'quantidade_total': int(df['quantidade'].sum()) if 'quantidade' in df else 0,
                'valor_total': float(df['valor_total'].sum()) if 'valor_total' in df else 0.0
            }

            # Conta itens sem De-Para
            if 'nosso_codigo' in df.columns:
                sem_depara = df['nosso_codigo'].isna().sum()
                summary['sem_depara'] = int(sem_depara)
                summary['com_depara'] = len(df) - int(sem_depara)

            # Resumo por filial com dados completos do cliente
            if 'cnpj_filial' in df:
                summary['por_filial'] = []
                for cnpj in df['cnpj_filial'].unique():
                    # Sanitiza CNPJ - se for NaN, converte para None
                    if pd.isna(cnpj):
                        cnpj = None
                        # Filtra itens sem CNPJ
                        filial_df = df[df['cnpj_filial'].isna()]
                    else:
                        filial_df = df[df['cnpj_filial'] == cnpj]

                    # Pega os dados do primeiro item
                    if not filial_df.empty:
                        primeiro_item = filial_df.iloc[0].to_dict()
                        # Sanitiza valores NaN no primeiro_item também
                        primeiro_item = {k: (None if pd.isna(v) else v) for k, v in primeiro_item.items()}
                    else:
                        primeiro_item = {}

                    # Converte produtos para garantir que Decimals sejam float
                    produtos_list = []
                    for produto in filial_df.to_dict('records'):
                        produto_clean = {}
                        for k, v in produto.items():
                            if hasattr(v, 'quantize'):  # É Decimal
                                produto_clean[k] = float(v)
                            elif pd.isna(v):  # É NaN/None
                                produto_clean[k] = None
                            else:
                                produto_clean[k] = v
                        produtos_list.append(produto_clean)

                    # Conta itens sem De-Para nesta filial
                    sem_depara_filial = filial_df['nosso_codigo'].isna().sum() if 'nosso_codigo' in filial_df.columns else 0

                    # Busca número do pedido (Atacadão: por filial, Assaí: mesmo para todas)
                    # IMPORTANTE: Para Atacadão, o campo "Numero:" (numero_comprador) é o
                    # número do pedido do cliente, diferente por filial.
                    # Para Assaí, o campo "Pedido:" (numero_pedido) é o mesmo para todas as filiais.
                    rede = self.identificacao.rede.upper() if self.identificacao else ''

                    if rede == 'ATACADAO':
                        # Atacadão: prioriza "Numero:" (numero_comprador) - único por filial
                        numero_pedido = (
                            primeiro_item.get('numero_comprador') or   # Campo "Numero: 322506"
                            primeiro_item.get('pedido_edi') or         # Fallback: Pedido EDI
                            primeiro_item.get('numero_pedido') or      # Fallback
                            primeiro_item.get('proposta') or           # Atacadão Proposta
                            ''
                        )
                    else:
                        # Assaí/Sendas: usa "Pedido:" (numero_pedido) - mesmo para todas
                        numero_pedido = (
                            primeiro_item.get('numero_pedido') or      # Campo "Pedido: 21046597"
                            primeiro_item.get('pedido_edi') or         # Fallback
                            primeiro_item.get('proposta') or           # Fallback
                            primeiro_item.get('numero_comprador') or   # Fallback
                            ''
                        )

                    filial_info = {
                        'cnpj': cnpj,
                        'numero_loja': primeiro_item.get('numero_loja', ''),  # Número da filial (ex: "007")
                        'numero_pedido_cliente': numero_pedido,  # Número do pedido/proposta
                        'nome_cliente': primeiro_item.get('nome_cliente', ''),
                        'local': primeiro_item.get('local_entrega', ''),
                        'municipio': primeiro_item.get('municipio', '') or primeiro_item.get('cidade', ''),
                        'estado': primeiro_item.get('estado', '') or primeiro_item.get('uf', ''),
                        'itens': len(filial_df),
                        'sem_depara': int(sem_depara_filial),
                        'quantidade': int(filial_df['quantidade'].sum()) if 'quantidade' in filial_df else 0,
                        'valor': float(filial_df['valor_total'].sum()) if 'valor_total' in filial_df else 0.0,
                        'produtos': produtos_list
                    }
                    summary['por_filial'].append(filial_info)

            return summary
        except Exception as e:
            import traceback
            print(f"Erro ao gerar summary: {e}")
            print(traceback.format_exc())
            return {
                'total_itens': len(data) if data else 0,
                'total_filiais': 0,
                'total_produtos': 0,
                'quantidade_total': 0,
                'valor_total': 0.0,
                'por_filial': [],
                'error': str(e)
            }

    def export_to_excel(self, data: List[Dict[str, Any]], output_path: str,
                        validacao_precos: Optional[Dict[str, Any]] = None):
        """
        Exporta dados para Excel com 2 abas:
        - Aba 1 "Pedido Completo": Todos itens com preço tabela, preço pedido e diferença
        - Aba 2 "Divergências": Apenas itens divergentes com % de diferença

        Args:
            data: Lista de itens extraídos do PDF
            output_path: Caminho do arquivo de saída
            validacao_precos: Dict com validação de preços (opcional)
                - por_filial: lista com cnpj, validacoes (preco_documento, preco_tabela, diferenca_percentual)
        """
        if not data:
            raise ValueError("Sem dados para exportar")

        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # Cria mapeamento de validação: (cnpj, nosso_codigo) → dados de preço
        mapa_validacao = {}
        if validacao_precos and validacao_precos.get('por_filial'):
            for filial in validacao_precos['por_filial']:
                cnpj = filial.get('cnpj', '')
                for val in filial.get('validacoes', []):
                    codigo = val.get('codigo', '')
                    if codigo:
                        chave = (cnpj, codigo)
                        mapa_validacao[chave] = {
                            'preco_documento': val.get('preco_documento'),
                            'preco_tabela': val.get('preco_tabela'),
                            'diferenca': val.get('diferenca', 0),
                            'diferenca_percentual': val.get('diferenca_percentual', 0),
                            'divergente': val.get('divergente', False)
                        }

        # Prepara dados para Aba 1: Pedido Completo
        dados_completos = []
        dados_divergencias = []

        for item in data:
            cnpj = item.get('cnpj_filial', '')
            nosso_codigo = item.get('nosso_codigo', '')
            chave = (cnpj, nosso_codigo)

            # Busca dados de validação
            val_data = mapa_validacao.get(chave, {})
            preco_pedido = item.get('valor_unitario', 0)
            preco_tabela = val_data.get('preco_tabela')
            diferenca = val_data.get('diferenca', 0)
            diferenca_pct = val_data.get('diferenca_percentual', 0)
            divergente = val_data.get('divergente', False)

            # Se não encontrou na validação, tenta calcular
            if preco_tabela is None and preco_pedido:
                preco_tabela = preco_pedido  # Sem tabela cadastrada
                diferenca = 0
                diferenca_pct = 0

            # Linha para Aba 1
            # Prioriza nossa_descricao (descrição do cadastro) sobre descricao (do PDF)
            descricao_exibir = item.get('nossa_descricao') or item.get('descricao', '')

            # Número do pedido do cliente (Atacadão: numero_comprador, Assaí: numero_pedido)
            numero_pedido_cliente = (
                item.get('numero_comprador') or
                item.get('numero_pedido') or
                item.get('pedido_edi') or
                item.get('proposta') or
                ''
            )

            linha_completa = {
                'CNPJ Filial': cnpj,
                'Nº Pedido Cliente': numero_pedido_cliente,
                'Nome Cliente': item.get('nome_cliente', ''),
                'UF': item.get('estado', '') or item.get('uf', ''),
                'Código Rede': item.get('codigo', ''),
                'Nosso Código': nosso_codigo or '-',
                'Descrição': descricao_exibir,
                'Quantidade': item.get('quantidade', 0),
                'Preço Pedido': preco_pedido,
                'Preço Tabela': preco_tabela,
                'Diferença (R$)': diferenca if preco_tabela else '-',
                'Diferença (%)': diferenca_pct if preco_tabela else '-',
                'Valor Total': item.get('valor_total', 0)
            }
            dados_completos.append(linha_completa)

            # Se divergente, adiciona à Aba 2
            if divergente:
                linha_divergencia = {
                    'CNPJ Filial': cnpj,
                    'Nº Pedido Cliente': numero_pedido_cliente,
                    'Nome Cliente': item.get('nome_cliente', ''),
                    'UF': item.get('estado', '') or item.get('uf', ''),
                    'Código Rede': item.get('codigo', ''),
                    'Nosso Código': nosso_codigo or '-',
                    'Descrição': descricao_exibir,
                    'Quantidade': item.get('quantidade', 0),
                    'Preço Pedido': preco_pedido,
                    'Preço Tabela': preco_tabela,
                    'Diferença (R$)': diferenca,
                    'Diferença (%)': diferenca_pct,
                    'Valor Total Pedido': (item.get('quantidade', 0) or 0) * (preco_pedido or 0),
                    'Valor Total Tabela': (item.get('quantidade', 0) or 0) * (preco_tabela or 0)
                }
                dados_divergencias.append(linha_divergencia)

        # Cria DataFrames
        df_completo = pd.DataFrame(dados_completos)
        df_divergencias = pd.DataFrame(dados_divergencias) if dados_divergencias else pd.DataFrame()

        # Estilos
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        divergencia_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # ========== Aba 1: Pedido Completo ==========
            df_completo.to_excel(writer, sheet_name='Pedido Completo', index=False)
            ws_completo = writer.sheets['Pedido Completo']

            # Formata header
            for col_idx, col in enumerate(df_completo.columns, 1):
                cell = ws_completo.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            # Formata dados e ajusta largura
            for col_idx, column in enumerate(df_completo.columns, 1):
                col_letter = get_column_letter(col_idx)
                max_length = len(str(column))

                for row_idx, value in enumerate(df_completo[column], 2):
                    cell = ws_completo.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border

                    # Formata valores monetários
                    if column in ['Preço Pedido', 'Preço Tabela', 'Diferença (R$)', 'Valor Total']:
                        if isinstance(value, (int, float)) and value != '-':
                            cell.number_format = 'R$ #,##0.00'
                            cell.alignment = Alignment(horizontal='right')
                    elif column == 'Diferença (%)':
                        if isinstance(value, (int, float)) and value != '-':
                            cell.number_format = '0.00"%"'
                            cell.alignment = Alignment(horizontal='right')
                            # Destaca divergências em vermelho
                            if value != 0:
                                cell.fill = divergencia_fill
                    elif column == 'Quantidade':
                        cell.alignment = Alignment(horizontal='right')

                    # Calcula largura máxima
                    if value is not None and value != '-':
                        max_length = max(max_length, len(str(value)))

                ws_completo.column_dimensions[col_letter].width = min(max_length + 2, 40)

            # ========== Aba 2: Divergências ==========
            if not df_divergencias.empty:
                df_divergencias.to_excel(writer, sheet_name='Divergências', index=False)
                ws_div = writer.sheets['Divergências']

                # Formata header
                for col_idx, col in enumerate(df_divergencias.columns, 1):
                    cell = ws_div.cell(row=1, column=col_idx)
                    cell.fill = PatternFill(start_color='C62828', end_color='C62828', fill_type='solid')
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border

                # Formata dados e ajusta largura
                for col_idx, column in enumerate(df_divergencias.columns, 1):
                    col_letter = get_column_letter(col_idx)
                    max_length = len(str(column))

                    for row_idx, value in enumerate(df_divergencias[column], 2):
                        cell = ws_div.cell(row=row_idx, column=col_idx)
                        cell.border = thin_border

                        # Formata valores monetários
                        if column in ['Preço Pedido', 'Preço Tabela', 'Diferença (R$)',
                                      'Valor Total Pedido', 'Valor Total Tabela']:
                            if isinstance(value, (int, float)):
                                cell.number_format = 'R$ #,##0.00'
                                cell.alignment = Alignment(horizontal='right')
                        elif column == 'Diferença (%)':
                            if isinstance(value, (int, float)):
                                cell.number_format = '0.00"%"'
                                cell.alignment = Alignment(horizontal='right')
                        elif column == 'Quantidade':
                            cell.alignment = Alignment(horizontal='right')

                        # Calcula largura máxima
                        if value is not None:
                            max_length = max(max_length, len(str(value)))

                    ws_div.column_dimensions[col_letter].width = min(max_length + 2, 40)
            else:
                # Cria aba vazia com mensagem
                df_vazia = pd.DataFrame({'Mensagem': ['Nenhuma divergência de preço encontrada']})
                df_vazia.to_excel(writer, sheet_name='Divergências', index=False)

        return output_path

    def export_to_csv(self, data: List[Dict[str, Any]], output_path: str):
        """
        Exporta dados para CSV
        """
        if not data:
            raise ValueError("Sem dados para exportar")

        df = pd.DataFrame(data)
        df.to_csv(output_path, index=False, encoding='utf-8-sig')

        return output_path
