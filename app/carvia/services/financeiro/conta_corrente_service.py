"""ContaCorrenteService — gerenciamento de saldo CC por transportadora.

Porta para o CarVia o conceito de "Conta Corrente Considerado x Pago" do
modulo Nacom. Cada movimentacao representa a diferenca entre o que CarVia
acordou pagar (`valor_considerado`) e o que efetivamente pagou (`valor_pago`)
de um CarviaSubcontrato.

Sinal canonico (CORRIGIDO no porte — Nacom original tem inversao historica):
- valor_pago > valor_considerado -> DEBITO  (CarVia pagou MAIS, transp. NOS DEVE,
                                              saldo aumenta)
- valor_pago < valor_considerado -> CREDITO (CarVia pagou MENOS, DEVEMOS a transp.,
                                              saldo diminui)

Saldo = SUM(valor_debito) - SUM(valor_credito) -> positivo = transp. nos deve.

Diferencas em relacao ao Nacom:
1. Logica centralizada aqui (Nacom tem 5 hooks inline em routes.py)
2. Service stateless (todos metodos sao staticmethod)
3. Suporta filtro de data no extrato (Nacom nao tem)
4. Inclui export Excel via openpyxl (Nacom nao tem)

Ref:
- .claude/plans/wobbly-tumbling-treasure.md
- /tmp/subagent-findings/conta_corrente_nacom.md
- app/fretes/models.py:471-510 (ContaCorrenteTransportadora Nacom)
"""

import logging
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Dict, List, Optional

from app import db
from app.utils.timezone import agora_utc_naive

logger = logging.getLogger(__name__)


class ContaCorrenteService:
    """Servico de movimentacoes de conta corrente de transportadoras."""

    # =================================================================
    # Logica de quando lancar (espelha Frete.deve_lancar_conta_corrente)
    # =================================================================
    @staticmethod
    def deve_lancar(frete) -> tuple:
        """Avalia se frete deve lancar movimentacao na CC automaticamente.

        Regras (espelho Nacom `Frete.deve_lancar_conta_corrente`):
        - Sem `valor_pago` ou sem `valor_considerado`: nao lanca
        - diff dentro da tolerancia (R$ 5,00): lanca direto
        - diff > tolerancia: requer aprovacao explicita (lancar_diferenca=True)

        Args:
            frete: CarviaFrete (unidade de analise — paridade Nacom)

        Returns:
            (bool, str) — pode lancar, motivo
        """
        from app.carvia.services.documentos.aprovacao_frete_service import (
            TOLERANCIA_APROVACAO,
        )

        if frete.valor_pago is None or frete.valor_considerado is None:
            return False, 'valor_pago ou valor_considerado nao informado'

        valor_pago = Decimal(str(frete.valor_pago))
        valor_considerado = Decimal(str(frete.valor_considerado))
        diff = abs(valor_pago - valor_considerado)

        if diff == 0:
            return False, 'valores iguais — sem movimentacao'

        if diff <= TOLERANCIA_APROVACAO:
            return True, f'dentro da tolerancia (diff R$ {diff:.2f})'

        # Acima da tolerancia — exige aprovacao explicita (caller decide)
        return False, f'acima da tolerancia (diff R$ {diff:.2f}) — requer aprovacao'

    # =================================================================
    # Criacao de movimentacao
    # =================================================================
    @staticmethod
    def lancar_movimentacao(
        frete_id: int,
        descricao: str,
        usuario: str,
        fatura_transportadora_id: Optional[int] = None,
        observacoes: Optional[str] = None,
    ) -> Dict:
        """Cria uma movimentacao na CC com base na diferenca do frete.

        Calcula tipo (DEBITO/CREDITO) pelo sinal de
        `valor_pago - valor_considerado`. NAO commita — chamador deve commitar.

        Args:
            frete_id: ID do CarviaFrete (paridade Nacom Frete.id)
        """
        from app.carvia.models import (
            CarviaContaCorrenteTransportadora,
            CarviaFrete,
        )

        frete = db.session.get(CarviaFrete, frete_id)
        if not frete:
            return {'sucesso': False, 'erro': 'Frete nao encontrado'}

        if frete.valor_pago is None or frete.valor_considerado is None:
            return {
                'sucesso': False,
                'erro': 'Frete precisa ter valor_pago e valor_considerado para gerar CC',
            }

        valor_pago = Decimal(str(frete.valor_pago))
        valor_considerado = Decimal(str(frete.valor_considerado))
        diff_assinada = valor_pago - valor_considerado
        diff_abs = abs(diff_assinada)

        if diff_abs == 0:
            return {'sucesso': False, 'erro': 'Sem diferenca para lancar'}

        # CarVia pagou MAIS que o considerado -> transportadora RECEBEU A MAIS ->
        #     transp. NOS DEVE a diferenca -> DEBITO (a receber).
        # CarVia pagou MENOS que o considerado -> faltou pagar ->
        #     DEVEMOS a transp. -> CREDITO (a pagar).
        # Saldo = SUM(debito) - SUM(credito); positivo = transp. nos deve.
        if diff_assinada > 0:
            tipo = 'DEBITO'
            valor_debito = diff_abs
            valor_credito = Decimal('0')
        else:
            tipo = 'CREDITO'
            valor_credito = diff_abs
            valor_debito = Decimal('0')

        try:
            mov = CarviaContaCorrenteTransportadora(
                transportadora_id=frete.transportadora_id,
                frete_id=frete.id,
                subcontrato_id=None,  # deprecated — fonte e frete_id
                fatura_transportadora_id=(
                    fatura_transportadora_id or frete.fatura_transportadora_id
                ),
                tipo_movimentacao=tipo,
                valor_diferenca=diff_abs,
                valor_debito=valor_debito,
                valor_credito=valor_credito,
                descricao=descricao,
                observacoes=observacoes,
                status='ATIVO',
                criado_em=agora_utc_naive(),
                criado_por=usuario,
            )
            db.session.add(mov)
            db.session.flush()

            logger.info(
                f'CC mov criada | frete={frete_id} | tipo={tipo} | '
                f'valor={diff_abs} | mov_id={mov.id}'
            )
            return {
                'sucesso': True,
                'movimentacao_id': mov.id,
                'tipo': tipo,
                'valor': float(diff_abs),
            }

        except Exception as e:
            logger.exception(f'Erro ao criar mov CC para frete {frete_id}: {e}')
            return {'sucesso': False, 'erro': str(e)}

    # =================================================================
    # Cancelamento (estorno)
    # =================================================================
    @staticmethod
    def cancelar_movimentacoes(frete_id: int, motivo: str, usuario: str) -> int:
        """Marca todas as mov ATIVA do frete como DESCONSIDERADO.

        Usado em hooks de cancelamento (frete.status='CANCELADO',
        desanexar aprovacoes de fatura). NAO commita — chamador deve commitar.
        Retorna qtd de registros alterados.
        """
        from app.carvia.models import CarviaContaCorrenteTransportadora

        movs = CarviaContaCorrenteTransportadora.query.filter_by(
            frete_id=frete_id, status='ATIVO'
        ).all()

        if not movs:
            return 0

        agora = agora_utc_naive()
        for m in movs:
            m.status = 'DESCONSIDERADO'
            m.observacoes = (
                (m.observacoes or '') + f'\n[AUTO {agora.isoformat()}] {motivo}'
            ).strip()

        logger.info(
            f'{len(movs)} mov CC desconsideradas | frete={frete_id} | '
            f'usuario={usuario} | motivo={motivo}'
        )
        return len(movs)

    @staticmethod
    def desconsiderar_movimentacao(mov_id: int, motivo: str, usuario: str) -> Dict:
        """Marca uma unica movimentacao como DESCONSIDERADO.

        Usado pelo endpoint POST /carvia/contas-correntes/<mov_id>/desconsiderar.
        Faz commit.
        """
        from app.carvia.models import CarviaContaCorrenteTransportadora

        mov = db.session.get(CarviaContaCorrenteTransportadora, mov_id)
        if not mov:
            return {'sucesso': False, 'erro': 'Movimentacao nao encontrada'}

        if mov.status != 'ATIVO':
            return {
                'sucesso': False,
                'erro': f'Movimentacao ja {mov.status}',
            }

        try:
            agora = agora_utc_naive()
            mov.status = 'DESCONSIDERADO'
            mov.observacoes = (
                (mov.observacoes or '') + f'\n[{agora.isoformat()}] {usuario}: {motivo}'
            ).strip()
            db.session.commit()
            logger.info(
                f'Mov CC desconsiderada | mov={mov_id} | usuario={usuario}'
            )
            return {'sucesso': True, 'movimentacao_id': mov_id}
        except Exception as e:
            db.session.rollback()
            logger.exception(f'Erro ao desconsiderar mov {mov_id}: {e}')
            return {'sucesso': False, 'erro': str(e)}

    # =================================================================
    # Calculo de saldo
    # =================================================================
    @staticmethod
    def calcular_saldo(transportadora_id: int) -> Dict:
        """Calcula saldo agregado de uma transportadora (status='ATIVO').

        Saldo = SUM(valor_debito) - SUM(valor_credito)
        Positivo = transportadora deve para CarVia
        Negativo = CarVia deve para transportadora
        """
        from sqlalchemy import func
        from app.carvia.models import CarviaContaCorrenteTransportadora

        agregado = (
            db.session.query(
                func.coalesce(func.sum(CarviaContaCorrenteTransportadora.valor_debito), 0).label('total_debito'),
                func.coalesce(func.sum(CarviaContaCorrenteTransportadora.valor_credito), 0).label('total_credito'),
                func.count(CarviaContaCorrenteTransportadora.id).label('qtd'),
            )
            .filter(
                CarviaContaCorrenteTransportadora.transportadora_id == transportadora_id,
                CarviaContaCorrenteTransportadora.status == 'ATIVO',
            )
            .first()
        )

        total_debito = float(agregado.total_debito or 0)
        total_credito = float(agregado.total_credito or 0)
        saldo = round(total_debito - total_credito, 2)

        return {
            'transportadora_id': transportadora_id,
            'saldo': saldo,
            'total_debito': round(total_debito, 2),
            'total_credito': round(total_credito, 2),
            'qtd_movimentacoes': int(agregado.qtd or 0),
            'classificacao': (
                'DEVEDOR' if saldo > 0
                else 'CREDOR' if saldo < 0
                else 'ZERADO'
            ),
        }

    @staticmethod
    def listar_saldos_todas_transportadoras() -> List[Dict]:
        """Lista todas transportadoras com saldo CC ATIVO != 0 (ou com mov).

        Usado pela tela /carvia/contas-correntes.
        """
        from sqlalchemy import func
        from app.carvia.models import CarviaContaCorrenteTransportadora
        from app.transportadoras.models import Transportadora

        agregados = (
            db.session.query(
                CarviaContaCorrenteTransportadora.transportadora_id,
                func.coalesce(func.sum(CarviaContaCorrenteTransportadora.valor_debito), 0).label('total_debito'),
                func.coalesce(func.sum(CarviaContaCorrenteTransportadora.valor_credito), 0).label('total_credito'),
                func.count(CarviaContaCorrenteTransportadora.id).label('qtd'),
            )
            .filter(CarviaContaCorrenteTransportadora.status == 'ATIVO')
            .group_by(CarviaContaCorrenteTransportadora.transportadora_id)
            .all()
        )

        if not agregados:
            return []

        transp_ids = [a.transportadora_id for a in agregados]
        transps = {
            t.id: t for t in Transportadora.query.filter(
                Transportadora.id.in_(transp_ids)
            ).all()
        }

        resultados = []
        for ag in agregados:
            transp = transps.get(ag.transportadora_id)
            total_debito = float(ag.total_debito or 0)
            total_credito = float(ag.total_credito or 0)
            saldo = round(total_debito - total_credito, 2)
            resultados.append({
                'transportadora_id': ag.transportadora_id,
                'transportadora_nome': transp.razao_social if transp else '?',
                'transportadora_cnpj': transp.cnpj if transp else None,
                'saldo': saldo,
                'total_debito': round(total_debito, 2),
                'total_credito': round(total_credito, 2),
                'qtd_movimentacoes': int(ag.qtd or 0),
                'classificacao': (
                    'DEVEDOR' if saldo > 0
                    else 'CREDOR' if saldo < 0
                    else 'ZERADO'
                ),
            })

        # Ordena por saldo absoluto descendente
        resultados.sort(key=lambda x: abs(x['saldo']), reverse=True)
        return resultados

    # =================================================================
    # Extrato
    # =================================================================
    @staticmethod
    def listar_extrato(
        transportadora_id: int,
        data_inicio: Optional[date] = None,
        data_fim: Optional[date] = None,
        status: Optional[str] = 'ATIVO',
    ) -> List[Dict]:
        """Lista movimentacoes de uma transportadora com filtros.

        Paridade Nacom: usa CarviaFrete como unidade de analise. CarviaSubcontrato
        e join opcional apenas para exibir cte_numero na UI.

        status: 'ATIVO' | 'COMPENSADO' | 'DESCONSIDERADO' | None (todos)
        """
        from app.carvia.models import (
            CarviaContaCorrenteTransportadora,
            CarviaFrete,
            CarviaOperacao,
        )

        query = (
            db.session.query(
                CarviaContaCorrenteTransportadora,
                CarviaFrete,
                CarviaOperacao,
            )
            .outerjoin(
                CarviaFrete,
                CarviaFrete.id == CarviaContaCorrenteTransportadora.frete_id,
            )
            .outerjoin(
                CarviaOperacao,
                CarviaOperacao.id == CarviaFrete.operacao_id,
            )
            .filter(
                CarviaContaCorrenteTransportadora.transportadora_id == transportadora_id
            )
        )

        if status and status != 'TODOS':
            query = query.filter(CarviaContaCorrenteTransportadora.status == status)

        if data_inicio:
            query = query.filter(
                CarviaContaCorrenteTransportadora.criado_em >= data_inicio
            )
        if data_fim:
            query = query.filter(
                CarviaContaCorrenteTransportadora.criado_em <= data_fim
            )

        query = query.order_by(CarviaContaCorrenteTransportadora.criado_em.desc())

        resultados = []
        for mov, frete, op in query.all():
            # Obter primeiro sub do frete apenas para exibir cte_numero (UI)
            primary_sub_cte_numero = None
            if frete is not None:
                primary_sub = frete.subcontratos.first()
                if primary_sub:
                    primary_sub_cte_numero = primary_sub.cte_numero
            resultados.append({
                'mov_id': mov.id,
                'criado_em': mov.criado_em,
                'tipo': mov.tipo_movimentacao,
                'valor_diferenca': float(mov.valor_diferenca),
                'valor_debito': float(mov.valor_debito),
                'valor_credito': float(mov.valor_credito),
                'descricao': mov.descricao,
                'observacoes': mov.observacoes,
                'status': mov.status,
                'frete_id': frete.id if frete else None,
                'frete_cte_numero': primary_sub_cte_numero,
                'frete_valor_cotado': float(frete.valor_cotado or 0) if frete else None,
                'frete_valor_considerado': float(frete.valor_considerado or 0) if frete else None,
                'frete_valor_pago': float(frete.valor_pago or 0) if frete else None,
                'op_id': op.id if op else None,
                'op_cte_numero': op.cte_numero if op else None,
                'op_cidade_destino': op.cidade_destino if op else None,
                'op_uf_destino': op.uf_destino if op else None,
                'fatura_transportadora_id': mov.fatura_transportadora_id,
                'criado_por': mov.criado_por,
            })

        return resultados

    # =================================================================
    # Export Excel
    # =================================================================
    @staticmethod
    def exportar_excel(
        transportadora_id: int,
        data_inicio: Optional[date] = None,
        data_fim: Optional[date] = None,
        status: Optional[str] = 'ATIVO',
    ) -> bytes:
        """Gera planilha Excel do extrato com totais."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from app.transportadoras.models import Transportadora

        transp = db.session.get(Transportadora, transportadora_id)
        nome_transp = transp.razao_social if transp else f'Transp #{transportadora_id}'

        movs = ContaCorrenteService.listar_extrato(
            transportadora_id, data_inicio, data_fim, status
        )

        wb = Workbook()
        ws = wb.active
        ws.title = 'Conta Corrente'

        # Header
        ws['A1'] = f'Extrato CC — {nome_transp}'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:J1')

        ws['A2'] = (
            f'Periodo: '
            f'{data_inicio.strftime("%d/%m/%Y") if data_inicio else "(inicio)"} '
            f'a {data_fim.strftime("%d/%m/%Y") if data_fim else "(hoje)"} '
            f'| Status: {status or "TODOS"}'
        )
        ws.merge_cells('A2:J2')

        # Cabecalho da tabela
        cab = [
            'Data', 'Frete', 'CTe', 'Operacao', 'Destino',
            'Tipo', 'Diferenca', 'Debito', 'Credito', 'Status',
        ]
        for col_idx, valor in enumerate(cab, start=1):
            cell = ws.cell(row=4, column=col_idx, value=valor)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='4472C4')
            cell.alignment = Alignment(horizontal='center')

        # Dados
        total_debito = 0.0
        total_credito = 0.0
        for idx, m in enumerate(movs, start=5):
            ws.cell(row=idx, column=1, value=m['criado_em'].strftime('%d/%m/%Y') if m['criado_em'] else '')
            ws.cell(row=idx, column=2, value=f"Frete #{m['frete_id']}" if m['frete_id'] else '-')
            ws.cell(row=idx, column=3, value=m['frete_cte_numero'] or '-')
            ws.cell(row=idx, column=4, value=m['op_cte_numero'] or '-')
            ws.cell(row=idx, column=5, value=(
                f"{m['op_cidade_destino']}/{m['op_uf_destino']}"
                if m['op_cidade_destino'] else '-'
            ))
            ws.cell(row=idx, column=6, value=m['tipo'])
            ws.cell(row=idx, column=7, value=m['valor_diferenca'])
            ws.cell(row=idx, column=8, value=m['valor_debito'])
            ws.cell(row=idx, column=9, value=m['valor_credito'])
            ws.cell(row=idx, column=10, value=m['status'])
            total_debito += m['valor_debito']
            total_credito += m['valor_credito']

        # Linha de totais
        total_row = len(movs) + 5
        ws.cell(row=total_row, column=1, value='TOTAIS').font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=round(total_debito, 2)).font = Font(bold=True)
        ws.cell(row=total_row, column=9, value=round(total_credito, 2)).font = Font(bold=True)

        saldo_row = total_row + 1
        ws.cell(row=saldo_row, column=1, value='SALDO').font = Font(bold=True)
        ws.cell(row=saldo_row, column=8, value=round(total_debito - total_credito, 2)).font = Font(bold=True, color='C00000')

        # Largura colunas
        widths = [12, 10, 12, 12, 22, 12, 12, 12, 12, 14]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w

        # Format moeda
        for row in range(5, total_row + 1):
            for col in (7, 8, 9):
                ws.cell(row=row, column=col).number_format = 'R$ #,##0.00'

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
