"""Helper para gerar Excel com DUPLO CABECALHO hierarquico.

Linha 1 (merged) = GRUPO de granularidade (ex: "PRODUTO", "NF", "CTe", "FATURA",
                   "CONCILIACAO 1", ...)
Linha 2          = CAMPO especifico dentro do grupo (ex: "codigo", "numero",
                   "valor", "data")

Regra de ouro: cada export mostra campos da ENTIDADE propria + AGRUPAMENTOS
SUPERIORES. NUNCA mostra agrupamentos inferiores (NF nao lista produtos, CTe
nao lista NFs, etc) — exceto onde a granularidade do proprio export EXIGE
(ex: export NF em 1 linha por produto; export CTe em 1 linha por NF).

API:
    colunas = [
        ColunaGrupo(grupo='NF', campos=[
            Campo('numero_nf', 'Numero NF'),
            Campo('valor_total', 'Valor', fmt='money'),
        ]),
        ColunaGrupo(grupo='CTe', campos=[
            Campo('cte_numero', 'Numero'),
            ...
        ]),
    ]
    gerar_excel_duplo_cabecalho(colunas, linhas_dict, sheet_name, entity_name)

Formato numerico/data segue padrao pt-BR (vem de _fmt_date/_fmt_datetime do
export.py original).
"""
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any, Callable, Dict, List, Optional

from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


@dataclass
class Campo:
    """Campo dentro de um grupo de cabecalho."""
    key: str                          # chave no dict de dados da linha
    label: str                        # texto visivel na linha 2
    fmt: Optional[str] = None         # 'money' | 'int' | 'date' | None
    width: Optional[int] = None       # largura fixa (caracteres); auto se None


@dataclass
class ColunaGrupo:
    """Grupo de colunas (ex: NF, CTe, FATURA). Aparece merged na linha 1."""
    grupo: str                        # texto visivel na linha 1 (mesclado)
    campos: List[Campo] = field(default_factory=list)


# Estilos
_HDR_FILL_GRUPO = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
_HDR_FILL_CAMPO = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
_HDR_FONT = Font(bold=True, color='FFFFFF', size=11)
_HDR_FONT_CAMPO = Font(bold=True, color='FFFFFF', size=10)
_HDR_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)


def _fmt_value(value: Any, fmt: Optional[str]) -> Any:
    """Converte value para tipo nativo openpyxl + numformat por fmt."""
    if value is None or value == '':
        return ''
    if fmt == 'money':
        try:
            return float(value)
        except (TypeError, ValueError):
            return ''
    if fmt == 'int':
        try:
            return int(value)
        except (TypeError, ValueError):
            return ''
    if fmt == 'date':
        if hasattr(value, 'strftime'):
            return value.strftime('%d/%m/%Y')
        return str(value)
    if fmt == 'datetime':
        if hasattr(value, 'strftime'):
            return value.strftime('%d/%m/%Y %H:%M')
        return str(value)
    return value


def _numberformat(fmt: Optional[str]) -> Optional[str]:
    """Retorna format string openpyxl por tipo."""
    if fmt == 'money':
        return 'R$ #,##0.00'
    if fmt == 'int':
        return '#,##0'
    return None


def gerar_excel_duplo_cabecalho(
    colunas: List[ColunaGrupo],
    linhas: List[Dict[str, Any]],
    sheet_name: str,
    entity_name: str,
    timestamp_fn: Optional[Callable[[], str]] = None,
):
    """Gera Excel com linha 1 (grupos merged) + linha 2 (campos) + dados.

    Args:
        colunas: lista ordenada de ColunaGrupo (cada grupo com N campos)
        linhas:  lista de dicts {key_do_campo: valor}
        sheet_name: nome da aba
        entity_name: usado no nome do arquivo `carvia_{entity_name}_YYYYMMDD_HHMM.xlsx`
        timestamp_fn: funcao que retorna timestamp string (default: agora_utc_naive)

    Returns:
        flask send_file response.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # limite openpyxl

    # ---- Linha 1: grupos merged ----
    col_idx = 1
    for grupo in colunas:
        n_cols = len(grupo.campos)
        if n_cols == 0:
            continue
        start = col_idx
        end = col_idx + n_cols - 1

        cell = ws.cell(row=1, column=start, value=grupo.grupo)
        cell.fill = _HDR_FILL_GRUPO
        cell.font = _HDR_FONT
        cell.alignment = _HDR_ALIGN
        cell.border = _BORDER

        if n_cols > 1:
            ws.merge_cells(
                start_row=1, end_row=1, start_column=start, end_column=end
            )
            # Estilo nas celulas mescladas (merge mantem so a esquerda)
            for c in range(start + 1, end + 1):
                mc = ws.cell(row=1, column=c)
                mc.fill = _HDR_FILL_GRUPO
                mc.border = _BORDER

        col_idx = end + 1

    # ---- Linha 2: campos ----
    col_idx = 1
    for grupo in colunas:
        for campo in grupo.campos:
            cell = ws.cell(row=2, column=col_idx, value=campo.label)
            cell.fill = _HDR_FILL_CAMPO
            cell.font = _HDR_FONT_CAMPO
            cell.alignment = _HDR_ALIGN
            cell.border = _BORDER
            col_idx += 1

    ws.freeze_panes = 'A3'  # congela cabecalhos

    # ---- Linhas de dados ----
    for idx, linha in enumerate(linhas, start=3):
        col_idx = 1
        for grupo in colunas:
            for campo in grupo.campos:
                raw = linha.get(campo.key)
                cell = ws.cell(row=idx, column=col_idx, value=_fmt_value(raw, campo.fmt))
                num_fmt = _numberformat(campo.fmt)
                if num_fmt:
                    cell.number_format = num_fmt
                cell.border = _BORDER
                col_idx += 1

    # ---- Larguras ----
    col_idx = 1
    for grupo in colunas:
        for campo in grupo.campos:
            letter = get_column_letter(col_idx)
            if campo.width:
                ws.column_dimensions[letter].width = campo.width
            else:
                # auto: max(label, conteudo amostrado)
                max_len = len(campo.label)
                for linha in linhas[:200]:  # amostra
                    v = linha.get(campo.key)
                    if v is not None:
                        l = len(str(_fmt_value(v, campo.fmt)))
                        if l > max_len:
                            max_len = l
                ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 50)
            col_idx += 1

    # ---- Altura linha 1 (grupos) ----
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 28

    # ---- Gerar arquivo ----
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    if timestamp_fn is None:
        from app.utils.timezone import agora_utc_naive
        timestamp = agora_utc_naive().strftime('%Y%m%d_%H%M')
    else:
        timestamp = timestamp_fn()
    filename = f'carvia_{entity_name}_{timestamp}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


# ---------------------------------------------------------------------- #
#  Colunas dinamicas (N x M campos por CTe Complementar / Conciliacao)
# ---------------------------------------------------------------------- #

def grupo_dinamico(prefixo_grupo: str, n: int, campos_por_slot: List[Campo]) -> List[ColunaGrupo]:
    """Gera N grupos dinamicos com os mesmos campos internos.

    Ex.: para CTe Complementares:
        grupo_dinamico('CTe Comp', n=3, campos_por_slot=[
            Campo('numero_comp_{i}', 'Numero'),
            Campo('valor_{i}', 'Valor', fmt='money'),
            Campo('motivo_{i}', 'Motivo'),
        ])
      ->  [ColunaGrupo('CTe Comp 1', [numero_comp_1, valor_1, motivo_1]),
           ColunaGrupo('CTe Comp 2', [numero_comp_2, valor_2, motivo_2]),
           ColunaGrupo('CTe Comp 3', [numero_comp_3, valor_3, motivo_3])]

    Cada linha de dados precisa popular as chaves `key_{i}` para i=1..n.
    """
    grupos = []
    for i in range(1, n + 1):
        campos_i = [
            Campo(
                key=c.key.replace('{i}', str(i)),
                label=c.label,
                fmt=c.fmt,
                width=c.width,
            )
            for c in campos_por_slot
        ]
        grupos.append(ColunaGrupo(grupo=f'{prefixo_grupo} {i}', campos=campos_i))
    return grupos
