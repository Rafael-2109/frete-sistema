"""
Rotas Resultado por Frete — receita (CTe) vs custo (subcontrato + coleta), rateado por moto.

Tela  : GET /carvia/resultado-frete
Export: GET /carvia/api/exportar/resultado-frete  (xlsx 2 abas: Resumo + Detalhe NF)
"""

from datetime import datetime, timedelta

from flask import render_template, request, flash, redirect, url_for
from flask_login import login_required, current_user

from app.utils.auth_decorators import require_admin


def _parse_date(arg, default):
    """Converte query-string em date; retorna default se ausente ou invalido."""
    v = request.args.get(arg)
    if not v:
        return default
    try:
        return datetime.strptime(v, '%Y-%m-%d').date()
    except ValueError:
        return default


def register_resultado_frete_routes(bp):

    @bp.route('/resultado-frete')  # type: ignore
    @login_required
    @require_admin
    def resultado_frete():  # type: ignore
        """Tela Resultado por Frete — resumo por eixo + detalhe NF."""
        if not getattr(current_user, 'sistema_carvia', False):
            flash('Acesso negado. Voce nao tem permissao para o sistema CarVia.', 'danger')
            return redirect(url_for('main.dashboard'))

        from app.carvia.services.financeiro.resultado_frete_service import ResultadoFreteService
        from app.utils.timezone import agora_brasil_naive

        hoje = agora_brasil_naive().date()
        data_inicio = _parse_date('data_inicio', hoje - timedelta(days=30))
        data_fim = _parse_date('data_fim', hoje)
        uf = request.args.get('uf') or None
        eixo = request.args.get('eixo') or 'cte'
        if eixo not in ('cte', 'embarque', 'uf_mes'):
            eixo = 'cte'

        svc = ResultadoFreteService()
        resumo = svc.resumo(eixo, data_inicio, data_fim, uf)
        detalhe = svc.detalhe_por_nf(data_inicio, data_fim, uf)

        return render_template(
            'carvia/resultado_frete/index.html',
            resumo=resumo, detalhe=detalhe, eixo=eixo,
            data_inicio=data_inicio.isoformat(), data_fim=data_fim.isoformat(),
            uf=uf or '',
        )

    @bp.route('/api/exportar/resultado-frete')  # type: ignore
    @login_required
    @require_admin
    def exportar_resultado_frete():  # type: ignore
        """Exporta xlsx com 2 abas: Resumo (por eixo) e Detalhe NF."""
        if not getattr(current_user, 'sistema_carvia', False):
            flash('Acesso negado.', 'danger')
            return redirect(url_for('main.dashboard'))

        from io import BytesIO
        from flask import send_file
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from app.carvia.services.financeiro.resultado_frete_service import ResultadoFreteService
        from app.utils.timezone import agora_brasil_naive, agora_utc_naive

        hoje = agora_brasil_naive().date()
        data_inicio = _parse_date('data_inicio', hoje - timedelta(days=30))
        data_fim = _parse_date('data_fim', hoje)
        uf = request.args.get('uf') or None
        eixo = request.args.get('eixo') or 'cte'
        if eixo not in ('cte', 'embarque', 'uf_mes'):
            eixo = 'cte'

        svc = ResultadoFreteService()
        resumo = svc.resumo(eixo, data_inicio, data_fim, uf)
        detalhe = svc.detalhe_por_nf(data_inicio, data_fim, uf)

        hdr_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        hdr_font = Font(bold=True, color='FFFFFF')
        center = Alignment(horizontal='center', vertical='center')

        def _write_sheet(ws, headers, rows):
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = center
            for ridx, row in enumerate(rows, start=2):
                for c, val in enumerate(row, start=1):
                    ws.cell(row=ridx, column=c, value=val)
            ws.freeze_panes = 'A2'

        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Resumo'
        _write_sheet(ws1,
            [eixo.upper(), 'Receita', 'Custo Sub', 'Custo Coleta', 'Custo Total',
             'Resultado', 'Motos', 'R$/Moto Receita', 'R$/Moto Custo', 'R$/Moto Result', 'Margem %'],
            [[g['label'], g['receita'], g['custo_sub'], g['custo_coleta'], g['custo_total'],
              g['resultado'], g['motos'], g['receita_moto'], g['custo_moto'],
              g['resultado_moto'], g['margem_pct']] for g in resumo],
        )
        ws2 = wb.create_sheet('Detalhe NF')
        _write_sheet(ws2,
            ['NF', 'Cidade', 'UF', 'CTe', 'Embarque', 'Motos', 'Receita',
             'Custo Sub', 'Flag Custo', 'Custo Coleta', 'Resultado', 'R$/Moto'],
            [[d['numero_nf'], d['cidade'], d['uf'], d['cte_numero'], d['embarque_id'],
              d['motos'], d['receita'], d['custo_sub'], d['custo_sub_flag'],
              d['custo_coleta'], d['resultado'], d['resultado_moto']] for d in detalhe],
        )

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        ts = agora_utc_naive().strftime('%Y%m%d_%H%M')
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'carvia_resultado_frete_{ts}.xlsx',
        )
