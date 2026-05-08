"""Builder de XLSX no formato canonico HORA a partir de HoraPedido.

Reproduz o layout do print/XLSX que os operadores enviam via WhatsApp:

    PEDIDO DE VENDA - SCOOTER ELETRICA      (titulo mesclado)
    OPERACAO SP - LOJA FRANQUIA             (subtitulo mesclado)

    CLIENTE:   <razao social - apelido loja>
    CNPJ:      <cnpj>                  I.E:    <ie>
    ENDERECO:  <endereco>
    BAIRRO:    <bairro>
    CIDADE:    <cidade>                ESTADO: <uf>     CEP: <cep>
    TELEFONE:  <tel>                   EMAIL:  <email>
    CONTATO:   <data DD/MM/AAAA>

    | PRODUTO | CHASSI | COR | (MOTOR) | PALLET | VALOR UNITARIO |
    | ...     | ...    | ... |   ...   |  ...   | xxx,xx (verde) |
    |         |        |     |         | TOTAL  | =SUM(...) (amarelo) |

Coluna MOTOR e opcional — controlada por `has_motor`. Linha TOTAL usa formula
=SUM() ao inves de valor hardcoded para casar com o XLSX que o Cowork gera
e permitir auditoria/recalculo no Excel.

Uso primario:
    from app.hora.services.pedido_xlsx_builder import build_xlsx_de_pedido
    bytes_xlsx = build_xlsx_de_pedido(pedido)  # HoraPedido

Uso secundario (testes / dict direto sem ORM):
    from app.hora.services.pedido_xlsx_builder import build_xlsx_de_dict
    bytes_xlsx = build_xlsx_de_dict({
        'cliente': '...', 'cnpj': '...', 'has_motor': True,
        'produtos': [['JET MAX', 'LYDA...', 'BRANCO', '', 15, 7550.00], ...],
        'contato': '07/05/2026',
        ...
    })

Adaptado de pb_pkg/pedido_builder.py do Cowork (Downloads/pb_pkg).
Diferencas:
- Retorna bytes em vez de gravar em disco (callers fazem upload ao S3).
- Aceita HoraPedido diretamente (build_xlsx_de_pedido) — monta dict internamente.
- Usa CNPJ matriz HORA hardcoded (62.634.044/0001-20) como default.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


# CNPJ da matriz HORA (Tatuape) — emitente fiscal de todos os pedidos.
CNPJ_MATRIZ_HORA_FORMATADO = '62.634.044/0001-20'

# ---------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------
_THIN = Side(border_style='thin', color='000000')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill('solid', start_color='D9D9D9')
_GREEN_FILL = PatternFill('solid', start_color='C6EFCE')
_YELLOW_FILL = PatternFill('solid', start_color='FFFF00')
_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
_RIGHT = Alignment(horizontal='right', vertical='center')
_BOLD = Font(name='Arial', bold=True, size=11)
_NORMAL = Font(name='Arial', size=11)
_TITLE = Font(name='Arial', bold=True, size=12)


def _formatar_cnpj(cnpj_digitos: str) -> str:
    """14 digitos → '00.000.000/0000-00'. Se ja vier formatado, retorna como esta."""
    if not cnpj_digitos:
        return ''
    digitos = ''.join(c for c in cnpj_digitos if c.isdigit())
    if len(digitos) != 14:
        return cnpj_digitos
    return f'{digitos[:2]}.{digitos[2:5]}.{digitos[5:8]}/{digitos[8:12]}-{digitos[12:]}'


def _data_br(d) -> str:
    """date/datetime → 'DD/MM/AAAA'. None → ''."""
    if d is None:
        return ''
    if isinstance(d, str):
        return d
    return d.strftime('%d/%m/%Y')


def _cliente_para_xlsx(pedido) -> str:
    """Razao social HORA + apelido da loja destino.

    Ex.: 'HORA COMERCIO DE MOTOCICLETAS ELETRICAS LTDA - MOTOCHEFE TATUAPE SP'

    Reusa apelido_detectado quando disponivel (preserva texto original do
    print). Senao, monta a partir da loja_destino (rotulo_display).
    """
    base = 'HORA COMERCIO DE MOTOCICLETAS ELETRICAS LTDA'
    if pedido.apelido_detectado:
        return f'{base} - MOTOCHEFE {pedido.apelido_detectado}'
    if pedido.loja_destino:
        rotulo = pedido.loja_destino.rotulo_display
        if rotulo:
            return f'{base} - {rotulo}'
    return base


def _itens_para_produtos(pedido) -> tuple[list, bool]:
    """Converte itens do pedido em lista de tuplas posicionais.

    Retorna (produtos, has_motor):
      - has_motor=False (default): [PRODUTO, CHASSI, COR, PALLET, VALOR_UNIT]
      - has_motor=True (futuro): [PRODUTO, CHASSI, COR, MOTOR, PALLET, VALOR_UNIT]

    Hoje HoraPedidoItem nao guarda numero_motor (so HoraNfEntradaItem guarda
    via texto_original). Por isso has_motor=False sempre — coluna so
    aparece quando o LLM extrair motor da imagem (caso de pedido pre-NF
    com motor declarado, raro). Quando reconstruir XLSX a partir de pedido
    ja persistido, fica sem MOTOR — ok.

    Ignora itens-peca (peca_id IS NOT NULL) — pedido_xlsx_builder cobre
    so itens-moto.
    """
    has_motor = False
    produtos = []
    for item in pedido.itens:
        if item.peca_id is not None:
            continue  # ignora pecas no XLSX (template HORA so tem motos)

        modelo_nome = ''
        if item.modelo:
            modelo_nome = item.modelo.nome_modelo or ''

        chassi = item.numero_chassi or ''
        cor = item.cor or ''
        pallet = ''  # HoraPedidoItem nao tem campo pallet — fica vazio
        valor = float(item.preco_compra_esperado or 0)

        if has_motor:
            produtos.append([modelo_nome, chassi, cor, '', pallet, valor])
        else:
            produtos.append([modelo_nome, chassi, cor, pallet, valor])
    return produtos, has_motor


def _dict_de_pedido(pedido, *, loja=None) -> dict:
    """Monta dict no schema do builder a partir de HoraPedido + loja destino.

    Campos de endereco/cidade/uf vem da `loja_destino` quando disponivel.
    Quando o pedido foi importado por imagem, esses dados foram extraidos
    pelo LLM e podem nao estar 100% nos campos da loja. Para auditoria
    fiel, preferiria reler do PedidoExtraido — mas no momento da geracao
    em background o token ja foi descartado. Solucao: usar dados da loja.
    """
    loja = loja or pedido.loja_destino

    produtos, has_motor = _itens_para_produtos(pedido)

    # Endereco/cidade/uf da loja (se houver). Senao deixa vazio — operador
    # ja viu os dados originais no preview ao confirmar.
    endereco = ''
    bairro = ''
    cidade = ''
    estado = 'SP'
    cep = ''
    telefone = ''
    if loja:
        # Campos comuns em HoraLoja (ver app/hora/models/cadastros.py).
        # Defensivo: cada loja pode ter campos diferentes preenchidos.
        endereco = getattr(loja, 'endereco_logradouro', '') or getattr(loja, 'endereco', '') or ''
        bairro = getattr(loja, 'endereco_bairro', '') or getattr(loja, 'bairro', '') or ''
        cidade = getattr(loja, 'cidade', '') or ''
        uf_loja = getattr(loja, 'uf', '') or getattr(loja, 'estado', '') or 'SP'
        estado = uf_loja or 'SP'
        cep = getattr(loja, 'cep', '') or ''
        telefone = getattr(loja, 'telefone', '') or ''

    return {
        'cliente': _cliente_para_xlsx(pedido),
        'cnpj': _formatar_cnpj(pedido.cnpj_destino) or CNPJ_MATRIZ_HORA_FORMATADO,
        'ie': '',  # nao guardado no pedido
        'endereco': endereco,
        'bairro': bairro,
        'cidade': cidade,
        'estado': estado,
        'cep': cep,
        'telefone': telefone,
        'email': '',
        'contato': _data_br(pedido.data_pedido),
        'has_motor': has_motor,
        'produtos': produtos,
    }


def build_xlsx_de_pedido(pedido) -> bytes:
    """Gera XLSX no formato canonico HORA a partir de HoraPedido.

    Usado pelo worker de background para origem='IMAGEM' — produz XLSX
    equivalente ao print original.

    Levanta ValueError se o pedido nao tem itens-moto.
    """
    d = _dict_de_pedido(pedido)
    return build_xlsx_de_dict(d)


def build_xlsx_de_dict(data: dict) -> bytes:
    """Gera XLSX a partir de dict no schema documentado. Retorna bytes.

    Schema:
        cliente, cnpj, ie, endereco, bairro, cidade, estado, cep, telefone,
        email, contato (DD/MM/AAAA), has_motor (bool), produtos (lista de listas).

    Quando has_motor=True, cada produto deve ter 6 elementos:
        [PRODUTO, CHASSI, COR, MOTOR, PALLET, VALOR_UNITARIO]
    Quando has_motor=False, 5 elementos:
        [PRODUTO, CHASSI, COR, PALLET, VALOR_UNITARIO]
    """
    _validate(data)

    cliente = data['cliente']
    cnpj = data.get('cnpj') or CNPJ_MATRIZ_HORA_FORMATADO
    ie = data.get('ie', '')
    endereco = data.get('endereco', '')
    bairro = data.get('bairro', '')
    cidade = data.get('cidade', '')
    estado = data.get('estado', 'SP')
    cep = data.get('cep', '')
    telefone = data.get('telefone', '')
    email = data.get('email', '')
    contato = data.get('contato', '')
    has_motor = bool(data.get('has_motor', False))
    produtos = data['produtos']

    if has_motor:
        headers = ['PRODUTO', 'CHASSI', 'COR', 'MOTOR', 'PALLET', 'VALOR UNITARIO']
        widths = [14, 22, 18, 10, 10, 18]
    else:
        headers = ['PRODUTO', 'CHASSI', 'COR', 'PALLET', 'VALOR UNITARIO']
        widths = [14, 22, 18, 10, 18]

    ncols = len(headers)
    last_col = get_column_letter(ncols)

    wb = Workbook()
    # Remove a aba default e cria nova com nome desejado — evita o tipo
    # Optional[Worksheet] que wb.active retorna.
    default_ws = wb.active
    if default_ws is not None:
        wb.remove(default_ws)
    ws = wb.create_sheet(title='PEDIDO')

    # Titulos
    ws.merge_cells(f'A1:{last_col}1')
    ws.cell(row=1, column=1, value='PEDIDO DE VENDA - SCOOTER ELETRICA')
    ws.cell(row=1, column=1).font = _TITLE
    ws.cell(row=1, column=1).alignment = _CENTER

    ws.merge_cells(f'A2:{last_col}2')
    ws.cell(row=2, column=1, value='OPERAÇÃO SP - LOJA FRANQUIA')
    ws.cell(row=2, column=1).font = _TITLE
    ws.cell(row=2, column=1).alignment = _CENTER

    cidade_label = f'CIDADE: {cidade}' if cidade else 'CIDADE:'
    estado_label = f'ESTADO: {estado}' if estado else 'ESTADO:'
    cep_label = f'CEP: {cep}' if cep else 'CEP:'

    info_rows = [
        ('CLIENTE:', cliente, None, None),
        ('CNPJ:', cnpj, 'I.E:', ie),
        ('ENDEREÇO:', endereco, None, None),
        ('BAIRRO:', bairro, None, None),
        (cidade_label, '', estado_label, cep_label),
        ('TELEFONE:', telefone, 'EMAIL:', email),
        ('CONTATO:', contato, None, None),
    ]

    start_row = 3
    for i, (lbl, val, lbl2, val2) in enumerate(info_rows):
        r = start_row + i
        if lbl2 is None:
            ws.cell(row=r, column=1, value=lbl).font = _BOLD
            ws.cell(row=r, column=1).alignment = _LEFT
            if ncols >= 2:
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=ncols)
            ws.cell(row=r, column=2, value=val).font = _NORMAL
            ws.cell(row=r, column=2).alignment = _LEFT
        else:
            mid = ncols // 2 + 1
            label_already_has_value = lbl.startswith('CIDADE')
            ws.cell(row=r, column=1, value=lbl).font = _BOLD
            ws.cell(row=r, column=1).alignment = _LEFT
            if mid - 1 >= 2:
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=mid - 1)
            ws.cell(
                row=r, column=2,
                value=('' if label_already_has_value else val),
            ).font = _NORMAL
            ws.cell(row=r, column=2).alignment = _LEFT

            ws.cell(row=r, column=mid, value=lbl2).font = _BOLD
            ws.cell(row=r, column=mid).alignment = _LEFT
            if ncols > mid:
                ws.merge_cells(start_row=r, start_column=mid + 1, end_row=r, end_column=ncols)
                ws.cell(row=r, column=mid + 1, value=val2).font = _NORMAL
                ws.cell(row=r, column=mid + 1).alignment = _LEFT

        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).border = _BORDER

    # Header da tabela
    sep_row = start_row + len(info_rows)
    header_row = sep_row + 1
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=idx, value=h)
        cell.font = _BOLD
        cell.alignment = _CENTER
        cell.fill = _HEADER_FILL
        cell.border = _BORDER

    # Linhas de produto
    first_data_row = header_row + 1
    for i, p in enumerate(produtos):
        r = first_data_row + i
        if len(p) != ncols:
            raise ValueError(
                f'Produto na linha {i+1} tem {len(p)} colunas, esperado {ncols}: {p}'
            )
        for idx, val in enumerate(p, start=1):
            cell = ws.cell(row=r, column=idx, value=val)
            cell.font = _NORMAL
            cell.border = _BORDER
            if idx == ncols:  # VALOR UNITARIO
                cell.number_format = '#,##0.00'
                cell.fill = _GREEN_FILL
                cell.alignment = _RIGHT
            elif headers[idx - 1] in ('PRODUTO', 'COR', 'MOTOR'):
                cell.alignment = _LEFT
            else:
                cell.alignment = _RIGHT

    last_data_row = first_data_row + len(produtos) - 1

    # Linha TOTAL com formula
    total_row = last_data_row + 1
    for c in range(1, ncols - 1):
        ws.cell(row=total_row, column=c, value=None).border = _BORDER

    total_label_col = ncols - 1
    total_cell = ws.cell(row=total_row, column=total_label_col, value='TOTAL')
    total_cell.font = _BOLD
    total_cell.alignment = _CENTER
    total_cell.fill = _YELLOW_FILL
    total_cell.border = _BORDER

    sum_col = get_column_letter(ncols)
    sum_cell = ws.cell(
        row=total_row, column=ncols,
        value=f'=SUM({sum_col}{first_data_row}:{sum_col}{last_data_row})',
    )
    sum_cell.font = _BOLD
    sum_cell.number_format = '#,##0.00'
    sum_cell.fill = _YELLOW_FILL
    sum_cell.alignment = _RIGHT
    sum_cell.border = _BORDER

    # Larguras de coluna
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def suggest_filename(contato_ddmmaaaa: str, loja_codigo: str) -> str:
    """Gera nome 'DD.MMH-XX.xlsx' (Cowork convention).

    XX = codigo de loja (PG=Praia Grande, TAT=Tatuape, BR=Bragança, etc).
    Usado quando precisar nomear o XLSX antes do upload S3 (apenas
    cosmetico — a S3 key e gerada via FileStorage().save_file).
    """
    try:
        dt = datetime.strptime((contato_ddmmaaaa or '').strip(), '%d/%m/%Y')
    except ValueError:
        dt = datetime.now()
    return f"{dt.strftime('%d.%m')}H-{(loja_codigo or 'XX').upper()}.xlsx"


def _validate(data: dict) -> None:
    if not isinstance(data, dict):
        raise ValueError('data deve ser um dict')
    if 'cliente' not in data or not data['cliente']:
        raise ValueError("Campo obrigatorio: 'cliente'")
    if 'produtos' not in data or not isinstance(data['produtos'], list):
        raise ValueError("Campo obrigatorio: 'produtos' (lista)")
    if not data['produtos']:
        raise ValueError("'produtos' nao pode estar vazio")
    has_motor = bool(data.get('has_motor', False))
    expected_cols = 6 if has_motor else 5
    for i, p in enumerate(data['produtos']):
        if not isinstance(p, (list, tuple)):
            raise ValueError(f'Produto {i} deve ser lista/tupla')
        if len(p) != expected_cols:
            raise ValueError(
                f'Produto {i} tem {len(p)} colunas, esperado {expected_cols} '
                f'(has_motor={has_motor})'
            )
