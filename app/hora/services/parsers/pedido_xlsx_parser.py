"""Parser de pedido HORA em Excel (formato WhatsApp).

Formato canônico HORA (ex.: "13.04H HORA BRAGANÇA.xlsx"):
    Linhas 1-9: header com CLIENTE, CNPJ, CIDADE, ESTADO, data (mistura A:B e C:D)
    Linha 11: cabeçalho da tabela: PRODUTO | CHASSI | COR | PALLET | VALOR UNITARIO
    Linhas 12-N: um chassi por linha
    Final: totalizador + observações (a ignorar)

Formato legado ocasional (ex.: "PEDIDO MC 1901.xlsx"):
    Header com PEDIDO DE VENDA <nº>
    Tabela com chassi VAZIO (pedido pré-NF)
    Colunas em ordem PRODUTO | COR | CHASSI | PALLET | VALOR

Estratégia (6 camadas):
    C1: scan dinâmico da linha de header da tabela (≥3 tokens-chave).
    C2: mapeamento fuzzy coluna → campo via ALIASES.
    C3: extração do corpo até TOTAL/soma/linha vazia dupla.
    C4: regex em metadados (CNPJ, número, data, cliente) nas linhas acima.
    C5: normalização de chassi, preço, modelo.
    C6: fallback LLM (opcional, desligado por padrão).

Chassi vazio é tratado como PENDENTE e gera aviso. Ingestão decide aceitar ou não.
"""
from __future__ import annotations

import logging
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Dict, List, Optional, Tuple

from app import db

logger = logging.getLogger(__name__)


# ----------------------------- Config -----------------------------

# CNPJ da matriz HORA (Tatuapé). Todos os pedidos/NFs da HORA são emitidos para
# este CNPJ — a loja física de destino é identificada por apelido/nome. Sua
# presença no XLSX serve como triagem: se não aparece, não é pedido HORA.
CNPJ_MATRIZ_HORA = '62634044000120'

ALIASES_CAMPO = {
    'modelo': ['PRODUTO', 'MODELO', 'DESCRICAO'],
    'cor': ['COR', 'COLOR'],
    'chassi': ['CHASSI', 'CHASSIS', 'VIN', 'N CHASSI', 'NUMERO CHASSI'],
    'preco': ['VALOR', 'VALOR UNITARIO', 'VALOR UNIT', 'PRECO', 'PRECO UNITARIO'],
    'pallet': ['PALLET', 'PALETE'],
    'qtd': ['QTD', 'QUANTIDADE', 'QUANT'],
}

# Mínimo de tokens-chave numa linha para considerá-la o header da tabela.
MIN_TOKENS_HEADER = 3

# Marcadores de fim do corpo da tabela. Verificados como palavra standalone
# (não prefixo) — "SOMA AM" é modelo real, não totalizador. "RET" é modelo.
TOKENS_FIM_CORPO = frozenset({'TOTAL', 'TOTAIS', 'OBS', 'OBSERVACAO', 'OBSERVACOES'})


# ----------------------------- Estruturas -----------------------------

@dataclass
class ItemPedidoExtraido:
    numero_chassi: Optional[str]
    modelo: Optional[str]
    cor: Optional[str]
    preco_compra_esperado: Optional[Decimal]
    linha_origem: int
    aviso: Optional[str] = None


@dataclass
class PedidoExtraido:
    numero_pedido: Optional[str]
    cnpj_destino: Optional[str]
    data_pedido: Optional[date]
    cliente_nome: Optional[str]
    cidade: Optional[str]
    uf: Optional[str]
    apelido_detectado: Optional[str] = None
    # Texto bruto capturado do cabeçalho que parece identificar a loja destino
    # (ex: "HORA BRAGANÇA", "MOTOCHEFE TATUAPÉ"). Usado p/ sugerir loja_destino.
    cnpjs_candidatos: List[str] = field(default_factory=list)
    itens: List[ItemPedidoExtraido] = field(default_factory=list)
    avisos: List[str] = field(default_factory=list)
    header_row: Optional[int] = None
    metodo_extracao: str = 'REGEX'


class PedidoParseError(Exception):
    """Erro irrecuperável ao parsear pedido XLSX."""


# ----------------------------- Helpers -----------------------------

def _normalizar_token(s: str) -> str:
    """Upper + strip acento + colapsa espaços + remove pontuação leve."""
    if not s:
        return ''
    s = str(s).strip().upper()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = re.sub(r'[^\w\s]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def _classificar_celula_header(celula: str) -> Optional[str]:
    """Dado texto de uma célula, retorna qual campo (modelo/cor/chassi/preco/...) ela representa.

    Match por igualdade ou substring sobre o token normalizado.
    """
    token = _normalizar_token(celula)
    if not token:
        return None
    for campo, aliases in ALIASES_CAMPO.items():
        for alias in aliases:
            if token == alias or alias in token or token in alias:
                return campo
    return None


def _eh_totalizador(row: Tuple) -> bool:
    """Detecta linha-totalizador (alguma célula tem TOKEN exato ou OBS: prefixo).

    Usa match por palavra standalone — "SOMA AM" (modelo) e "RET" (modelo) não
    devem disparar. "TOTAL" isolado ou "OBS: xxx" sim.
    """
    for c in row:
        if c is None:
            continue
        token_full = _normalizar_token(str(c))
        if not token_full:
            continue
        palavras = token_full.split()
        if not palavras:
            continue
        primeira = palavras[0]
        # TOTAL/TOTAIS/OBS/OBSERVAC* sozinhos ou no início de uma frase-observação.
        if primeira in TOKENS_FIM_CORPO:
            return True
    return False


def _linha_vazia(row: Tuple) -> bool:
    return all(c is None or (isinstance(c, str) and not c.strip()) for c in row)


def _normalizar_chassi(raw) -> Optional[str]:
    """Chassi upper, sem espaços, sem apóstrofe inicial (Excel text-force)."""
    if raw is None:
        return None
    s = str(raw).strip()
    s = s.lstrip("'").strip()
    s = s.upper().replace(' ', '')
    if not s or s in {'NONE', 'NULL'}:
        return None
    if len(s) > 30:
        logger.warning("Chassi excede 30 chars: %r — truncando", s)
        s = s[:30]
    return s


def _normalizar_preco(raw) -> Optional[Decimal]:
    """Aceita int, float, '7170', '7.170,50', '7,170.50'."""
    if raw is None or raw == '':
        return None
    if isinstance(raw, (int, float)):
        try:
            return Decimal(str(raw))
        except InvalidOperation:
            return None
    s = str(raw).strip().replace('R$', '').replace(' ', '')
    if not s:
        return None
    # Formato brasileiro: 1.234,56 → 1234.56
    # Heurística: se tem tanto '.' quanto ',', assume que ',' é decimal.
    if ',' in s and '.' in s:
        if s.rfind(',') > s.rfind('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _normalizar_modelo(raw) -> Optional[str]:
    if raw is None:
        return None
    s = str(raw).strip().upper()
    s = re.sub(r'\s+', ' ', s)
    return s or None


def _normalizar_cnpj(raw) -> Optional[str]:
    if raw is None:
        return None
    digitos = ''.join(c for c in str(raw) if c.isdigit())
    return digitos if len(digitos) == 14 else None


# ----------------------------- Scanners -----------------------------

def _achar_header_row(linhas: List[Tuple]) -> Tuple[Optional[int], Dict[int, str]]:
    """Procura a linha que tem ≥MIN_TOKENS_HEADER células classificadas como campo.

    Retorna (linha_index, {col_idx: campo}).
    """
    for i, row in enumerate(linhas):
        mapa = {}
        for col, celula in enumerate(row):
            campo = _classificar_celula_header(celula)
            if campo and campo != 'pallet':  # pallet é só colunagem, não ajuda a identificar tabela
                mapa[col] = campo
        # Conta campos distintos (não apenas colunas)
        campos_distintos = set(mapa.values())
        if len(campos_distintos) >= MIN_TOKENS_HEADER:
            return i, mapa
    return None, {}


def _extrair_metadados(linhas_antes_header: List[Tuple], nome_arquivo: Optional[str]) -> Dict:
    """Varre linhas antes do header aplicando regex por célula (não fixa em coluna)."""
    meta = {
        'numero_pedido': None,
        'cnpj_destino': None,
        'data_pedido': None,
        'cliente_nome': None,
        'cidade': None,
        'uf': None,
        'apelido_detectado': None,
        'cnpjs_candidatos': [],  # todos CNPJs encontrados (ordem de aparição)
    }

    # Padrões de apelido de loja (genéricos — match contra banco faz refinamento).
    # "HORA BRAGANÇA", "MOTOCHEFE TATUAPÉ", "LOJA PRAIA GRANDE", etc.
    re_apelido = re.compile(
        r'\b(?:HORA|MOTOCHEFE|LOJA|FRANQUIA)\s+([A-Z\u00C0-\u00FF][A-Z\u00C0-\u00FF\s]+?)(?:\s*[-—|,]|\s*$)',
        re.IGNORECASE,
    )
    # Também captura "BRAGANÇA" isolado em linhas tipo "OPERAÇÃO SP - LOJA FRANQUIA"
    re_cidade_known = re.compile(
        r'\b(BRAGAN[ÇC]A|TATUAP[EÉ]|PRAIA\s+GRANDE|SANTOS|S[AÃ]O\s+PAULO)\b',
        re.IGNORECASE,
    )

    # CNPJ formatado (com . / -) OU raw 14 dígitos puros.
    # O raw pega número do pedido com 14+ dígitos? Não — número de pedido tem <10 dígitos
    # tipicamente, então o filtro len==14 é seguro.
    re_cnpj_formatado = re.compile(r'\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}')
    re_cnpj_raw = re.compile(r'(?<!\d)(\d{14})(?!\d)')
    re_pedido_num = re.compile(r'PEDIDO[^\d]*?(\d{3,})', re.IGNORECASE)
    re_data = re.compile(r'(\d{1,2}/\d{1,2}/\d{2,4})')
    re_uf = re.compile(r'(?:ESTADO|UF)[:\s]*([A-Z]{2})', re.IGNORECASE)

    candidatos_vistos = set()

    for row in linhas_antes_header:
        for celula in row:
            if celula is None:
                continue

            # Data/datetime direto de célula Excel
            if not meta['data_pedido']:
                if isinstance(celula, datetime):
                    meta['data_pedido'] = celula.date()
                elif isinstance(celula, date):
                    meta['data_pedido'] = celula

            texto = str(celula).strip()
            if not texto:
                continue

            # Coleta TODOS os CNPJs encontrados (formatados + raw).
            for m in re_cnpj_formatado.finditer(texto):
                cnpj = _normalizar_cnpj(m.group(0))
                if cnpj and cnpj not in candidatos_vistos:
                    candidatos_vistos.add(cnpj)
                    meta['cnpjs_candidatos'].append(cnpj)
            for m in re_cnpj_raw.finditer(texto):
                cnpj = _normalizar_cnpj(m.group(1))
                if cnpj and cnpj not in candidatos_vistos:
                    candidatos_vistos.add(cnpj)
                    meta['cnpjs_candidatos'].append(cnpj)

            # Primeiro CNPJ = cnpj_destino default (pode ser sobrescrito via lookup).
            if not meta['cnpj_destino'] and meta['cnpjs_candidatos']:
                meta['cnpj_destino'] = meta['cnpjs_candidatos'][0]

            if not meta['numero_pedido']:
                m = re_pedido_num.search(texto)
                if m:
                    meta['numero_pedido'] = m.group(1)

            if not meta['data_pedido']:
                m = re_data.search(texto)
                if m:
                    try:
                        data_str = m.group(1)
                        partes = data_str.split('/')
                        if len(partes[2]) == 2:
                            partes[2] = '20' + partes[2]
                        meta['data_pedido'] = date(
                            int(partes[2]), int(partes[1]), int(partes[0]),
                        )
                    except (ValueError, IndexError):
                        pass

            if not meta['cliente_nome']:
                m = re.search(r'CLIENTE[:\s]+(.+)', texto, re.IGNORECASE)
                if m:
                    meta['cliente_nome'] = m.group(1).strip()

            if not meta['cidade']:
                m = re.search(r'CIDADE[:\s]+(.+)', texto, re.IGNORECASE)
                if m:
                    meta['cidade'] = m.group(1).strip()

            if not meta['uf']:
                m = re_uf.search(texto)
                if m:
                    meta['uf'] = m.group(1).upper()

            # Detecta apelido/cidade-loja no cabeçalho (prioridade: cidade conhecida)
            if not meta['apelido_detectado']:
                m = re_cidade_known.search(texto)
                if m:
                    meta['apelido_detectado'] = m.group(1).upper()
                else:
                    m2 = re_apelido.search(texto)
                    if m2:
                        candidato = m2.group(1).strip().rstrip(':').strip()
                        # Rejeita termos genéricos
                        if (2 < len(candidato) < 50 and
                                candidato.upper() not in {'FRANQUIA', 'MATRIZ', 'FILIAL', 'LOJA'}):
                            meta['apelido_detectado'] = candidato.upper()

    # Fallback: tenta extrair apelido do nome do arquivo ("13.04H HORA BRAGANÇA" → "BRAGANÇA")
    if not meta['apelido_detectado'] and nome_arquivo:
        base = re.sub(r'\.xlsx?$', '', nome_arquivo, flags=re.IGNORECASE)
        m = re_cidade_known.search(base)
        if m:
            meta['apelido_detectado'] = m.group(1).upper()

    # Fallbacks
    if not meta['numero_pedido'] and nome_arquivo:
        # "13.04H HORA BRAGANÇA" → "13.04H-HORA-BRAGANÇA"
        # "PEDIDO MC 1901" → pega o 1901
        base = re.sub(r'\.xlsx?$', '', nome_arquivo, flags=re.IGNORECASE).strip()
        m = re.search(r'(\d{3,})', base)
        if m:
            # Se tem "HORA" no nome, usa formato HORA-<nome>
            if 'HORA' in base.upper():
                meta['numero_pedido'] = re.sub(r'\s+', '-', base.strip())
            else:
                meta['numero_pedido'] = m.group(1)
        else:
            meta['numero_pedido'] = re.sub(r'\s+', '-', base.strip())

    if not meta['data_pedido']:
        meta['data_pedido'] = date.today()

    return meta


def _extrair_itens(
    linhas_corpo: List[Tuple],
    col_map: Dict[int, str],
    linha_inicial: int,
) -> List[ItemPedidoExtraido]:
    itens: List[ItemPedidoExtraido] = []
    vazia_anterior = False

    for offset, row in enumerate(linhas_corpo):
        linha_absoluta = linha_inicial + offset

        if _linha_vazia(row):
            if vazia_anterior:
                break  # duas vazias em sequência = fim
            vazia_anterior = True
            continue
        vazia_anterior = False

        if _eh_totalizador(row):
            break

        # Extrai campos conforme mapa de colunas
        dados = {'modelo': None, 'cor': None, 'chassi': None, 'preco': None}
        for col, campo in col_map.items():
            if campo in dados and col < len(row):
                dados[campo] = row[col]

        # Heurística: se a linha tem só 1-2 células preenchidas em colunas NÃO
        # mapeadas como item (ex.: totalizador disfarçado), pular.
        celulas_preenchidas_mapeadas = sum(
            1 for col, campo in col_map.items()
            if campo in dados and col < len(row) and row[col] not in (None, '')
        )
        if celulas_preenchidas_mapeadas == 0:
            continue

        chassi = _normalizar_chassi(dados['chassi'])
        modelo = _normalizar_modelo(dados['modelo'])
        cor = _normalizar_modelo(dados['cor'])
        preco = _normalizar_preco(dados['preco'])

        aviso = None
        if not chassi:
            aviso = 'chassi_pendente'

        # Se não tem NEM modelo NEM chassi, descarta (linha vazia disfarçada)
        if not modelo and not chassi:
            continue

        itens.append(ItemPedidoExtraido(
            numero_chassi=chassi,
            modelo=modelo,
            cor=cor,
            preco_compra_esperado=preco,
            linha_origem=linha_absoluta,
            aviso=aviso,
        ))

    return itens


# ----------------------------- API pública -----------------------------

def parse_pedido_xlsx(
    xlsx_bytes: bytes,
    nome_arquivo: Optional[str] = None,
) -> PedidoExtraido:
    """Parseia XLSX de pedido HORA e retorna PedidoExtraido.

    Args:
        xlsx_bytes: conteúdo binário do .xlsx.
        nome_arquivo: nome para logging/fallback (opcional).

    Raises:
        PedidoParseError: quando não consegue identificar header da tabela.

    Returns:
        PedidoExtraido com metadados + itens + avisos.
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise PedidoParseError(f"openpyxl não instalado: {exc}")

    try:
        wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
    except Exception as exc:
        raise PedidoParseError(f"não foi possível abrir XLSX: {exc}")

    # Usa a primeira aba com conteúdo
    ws = None
    for name in wb.sheetnames:
        candidato = wb[name]
        if candidato.max_row and candidato.max_row > 5:
            ws = candidato
            break
    if ws is None:
        raise PedidoParseError("XLSX sem aba com conteúdo suficiente")

    # Lê todas as linhas (limitadas, para evitar planilhas gigantes)
    linhas = list(ws.iter_rows(max_row=min(500, ws.max_row), values_only=True))

    header_idx, col_map = _achar_header_row(linhas)
    if header_idx is None:
        raise PedidoParseError(
            f"Header da tabela não encontrado em {nome_arquivo or 'arquivo'}. "
            f"Esperado linha com ≥{MIN_TOKENS_HEADER} de: PRODUTO/MODELO, CHASSI, COR, VALOR."
        )

    # Metadados: linhas antes do header
    meta = _extrair_metadados(linhas[:header_idx], nome_arquivo)

    # Corpo: linhas depois do header
    itens = _extrair_itens(
        linhas_corpo=linhas[header_idx + 1:],
        col_map=col_map,
        linha_inicial=header_idx + 2,  # +2 porque: +1 pula header, +1 para 1-based display
    )

    avisos: List[str] = []
    chassis_pendentes = sum(1 for i in itens if i.aviso == 'chassi_pendente')
    if chassis_pendentes:
        avisos.append(
            f'{chassis_pendentes}/{len(itens)} itens sem chassi — '
            f'formato parece ser pedido pré-NF (solicitação).'
        )
    if not itens:
        raise PedidoParseError("Nenhum item extraído após header")

    # Validação mínima: números que queremos de toda forma
    if not meta.get('numero_pedido'):
        avisos.append('numero_pedido não identificado — use o nome do arquivo')
    if not meta.get('cnpj_destino'):
        avisos.append('CNPJ destino não identificado — preencher manualmente')

    return PedidoExtraido(
        numero_pedido=meta.get('numero_pedido'),
        cnpj_destino=meta.get('cnpj_destino'),
        data_pedido=meta.get('data_pedido'),
        cliente_nome=meta.get('cliente_nome'),
        cidade=meta.get('cidade'),
        uf=meta.get('uf'),
        apelido_detectado=meta.get('apelido_detectado'),
        cnpjs_candidatos=meta.get('cnpjs_candidatos', []),
        itens=itens,
        avisos=avisos,
        header_row=header_idx + 1,
        metodo_extracao='REGEX',
    )


def resolver_loja_por_apelido(apelido_detectado: Optional[str]) -> Tuple[Optional[int], str]:
    """Tenta casar apelido detectado no header do XLSX contra HoraLoja.apelido/nome.

    Match case-insensitive contra apelido, nome_fantasia, razao_social e cidade.
    Retorna (loja_id, mensagem).
    """
    if not apelido_detectado:
        return None, 'Nenhum apelido/cidade identificado no cabeçalho.'

    from app.hora.models import HoraLoja
    alvo = apelido_detectado.upper().strip()

    # Busca em ordem de prioridade: apelido → nome_fantasia → razão → cidade
    for campo in (HoraLoja.apelido, HoraLoja.nome_fantasia,
                  HoraLoja.razao_social, HoraLoja.nome, HoraLoja.cidade):
        candidatas = (
            HoraLoja.query
            .filter(HoraLoja.ativa.is_(True))
            .filter(db.func.upper(campo).like(f'%{alvo}%'))
            .all()
        )
        if len(candidatas) == 1:
            loja = candidatas[0]
            return loja.id, f'Loja sugerida: {loja.rotulo_display} (match em {campo.key})'
        elif len(candidatas) > 1:
            nomes = ', '.join(l.rotulo_display for l in candidatas[:3])
            return None, f'{len(candidatas)} lojas casam com "{alvo}": {nomes}... — selecione manualmente.'

    return None, f'Nenhuma loja ativa casa com "{alvo}". Cadastre a loja ou selecione manualmente.'


def cnpj_matriz_presente(cnpjs_candidatos: List[str]) -> bool:
    """Triagem: verifica se o CNPJ da matriz HORA (Tatuapé) aparece no XLSX.

    Todos os pedidos legítimos da HORA são emitidos para o CNPJ da matriz —
    a loja física de entrega é identificada separadamente por nome/apelido.
    A ausência deste CNPJ indica que o arquivo NÃO é pedido HORA (provavelmente
    foi aberto da pasta errada ou é pedido de outro cliente).

    Args:
        cnpjs_candidatos: lista de CNPJs encontrados no cabeçalho do XLSX.

    Returns:
        True se o CNPJ da matriz HORA está entre os candidatos.
    """
    if not cnpjs_candidatos:
        return False
    return CNPJ_MATRIZ_HORA in cnpjs_candidatos


def resolver_loja_por_cnpj(cnpjs_candidatos: List[str]) -> Tuple[Optional[str], str]:
    """Tenta casar cada CNPJ candidato contra lojas HORA cadastradas.

    Returns:
        (cnpj_resolvido, mensagem). cnpj_resolvido=None se zero ou múltiplos matches.
    """
    from app.hora.models import HoraLoja

    if not cnpjs_candidatos:
        return None, 'Nenhum CNPJ encontrado no arquivo.'

    cnpjs_cadastrados_match = [
        c for c in cnpjs_candidatos
        if HoraLoja.query.filter_by(cnpj=c, ativa=True).first() is not None
    ]

    if len(cnpjs_cadastrados_match) == 1:
        cnpj = cnpjs_cadastrados_match[0]
        return cnpj, f'Loja resolvida: {cnpj}'
    elif len(cnpjs_cadastrados_match) > 1:
        return None, (
            f'{len(cnpjs_cadastrados_match)} lojas HORA casam com CNPJs do arquivo '
            f'({", ".join(cnpjs_cadastrados_match)}) — selecione manualmente.'
        )
    else:
        return None, (
            f'Nenhum dos {len(cnpjs_candidatos)} CNPJs do arquivo bate com loja '
            f'HORA cadastrada. CNPJs encontrados: {", ".join(cnpjs_candidatos)}'
        )
