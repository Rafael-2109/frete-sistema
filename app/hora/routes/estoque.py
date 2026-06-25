"""Rotas de Estoque HORA: KPIs, listagem chassi-a-chassi, rastreamento de moto."""
from __future__ import annotations

from io import BytesIO

from flask import (
    Response,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    url_for,
)

from app.hora.decorators import require_hora_perm
from app.hora.models import HoraLoja, HoraModelo, HoraMoto
from app.hora.routes import hora_bp
from app.hora.services import estoque_service
from app.hora.services.auth_helper import (
    chassi_acessivel,
    lojas_permitidas_ids,
    usuario_tem_acesso_a_loja,
)
from app.utils.timezone import agora_utc_naive


def _int_arg(nome: str):
    v = (request.args.get(nome) or '').strip()
    return int(v) if v.isdigit() else None


def _coletar_filtros_estoque() -> dict:
    """Le os filtros da query string compartilhados pela tela e pela exportacao.

    Encapsula a regra de forcar `incluir_fora_estoque=True` quando ha filtro por
    documento (pedido/NF/venda) ou por chassi, garantindo paridade EXATA entre o
    que a tela mostra e o que o Excel exporta. Retorna kwargs prontos para
    `estoque_service.listar_estoque` (exceto `lojas_permitidas_ids`).
    """
    chassi = (request.args.get('chassi') or '').strip() or None
    pedido_id = _int_arg('pedido_id')
    nf_entrada_id = _int_arg('nf_entrada_id')
    venda_id = _int_arg('venda_id')

    incluir_fora_estoque = request.args.get('incluir_fora_estoque', '0') == '1'
    if pedido_id or nf_entrada_id or venda_id or chassi:
        incluir_fora_estoque = True

    return {
        'loja_id': _int_arg('loja_id'),
        'modelo_id': _int_arg('modelo_id'),
        'cor': (request.args.get('cor') or '').strip() or None,
        'chassi': chassi,
        'pedido_id': pedido_id,
        'nf_entrada_id': nf_entrada_id,
        'venda_id': venda_id,
        'status': (request.args.get('status') or '').strip().upper() or None,
        'incluir_avariadas': request.args.get('incluir_avariadas', '1') == '1',
        'incluir_faltando_peca': request.args.get('incluir_faltando_peca', '1') == '1',
        'incluir_fora_estoque': incluir_fora_estoque,
    }


@hora_bp.route('/estoque')
@require_hora_perm('estoque', 'ver')
def estoque_lista():
    permitidas = lojas_permitidas_ids()

    # Filtros por documento forcam `incluir_fora_estoque=True` (ver
    # _coletar_filtros_estoque) — permite ver vendidas ao filtrar por venda,
    # NF entrada cujo chassi ja saiu, etc. Status explicito (MOTO_FALTANDO,
    # AVARIADA, etc.) ignora os flags — listar_estoque trata internamente.
    filtros = _coletar_filtros_estoque()
    loja_id = filtros['loja_id']

    if loja_id and not usuario_tem_acesso_a_loja(loja_id):
        flash('Acesso negado a essa loja.', 'danger')
        return redirect(url_for('hora.estoque_lista'))

    motos = estoque_service.listar_estoque(
        lojas_permitidas_ids=permitidas,
        **filtros,
    )
    kpis_loja_modelo_cor = estoque_service.kpis_loja_modelo_cor(
        lojas_permitidas_ids=permitidas,
    )

    lojas_q = HoraLoja.query.filter_by(ativa=True)
    if permitidas is not None:
        if not permitidas:
            lojas = []
        else:
            lojas_q = lojas_q.filter(HoraLoja.id.in_(permitidas))
            lojas = lojas_q.order_by(HoraLoja.nome).all()
    else:
        lojas = lojas_q.order_by(HoraLoja.nome).all()

    opcoes = estoque_service.opcoes_filtro_estoque(lojas_permitidas_ids=permitidas)
    opcoes_docs = estoque_service.opcoes_documentos_filtro(
        lojas_permitidas_ids=permitidas,
    )

    # Para preencher input quando vier filtrado por modelo_id, precisamos do nome.
    modelo_selecionado_nome = None
    if filtros['modelo_id']:
        m = HoraModelo.query.get(filtros['modelo_id'])
        if m:
            modelo_selecionado_nome = m.nome_modelo

    return render_template(
        'hora/estoque_lista.html',
        motos=motos,
        kpis_loja_modelo_cor=kpis_loja_modelo_cor,
        lojas=lojas,
        modelos=opcoes['modelos'],
        cores=opcoes['cores'],
        pedidos_filtro=opcoes_docs['pedidos'],
        nfs_entrada_filtro=opcoes_docs['nfs_entrada'],
        vendas_filtro=opcoes_docs['vendas'],
        filtro_loja_id=filtros['loja_id'],
        filtro_modelo_id=filtros['modelo_id'],
        filtro_modelo_nome=modelo_selecionado_nome,
        filtro_cor=filtros['cor'],
        filtro_chassi=filtros['chassi'],
        filtro_pedido_id=filtros['pedido_id'],
        filtro_nf_entrada_id=filtros['nf_entrada_id'],
        filtro_venda_id=filtros['venda_id'],
        filtro_status=filtros['status'],
        incluir_avariadas=filtros['incluir_avariadas'],
        incluir_faltando_peca=filtros['incluir_faltando_peca'],
        incluir_fora_estoque=filtros['incluir_fora_estoque'],
    )


@hora_bp.route('/estoque/exportar.xlsx')
@require_hora_perm('estoque_exportar', 'ver')
def estoque_exportar_xlsx():
    """Exporta o estoque em Excel: Chassi, Modelo, Cor e Loja (onde se encontra).

    Respeita EXATAMENTE os mesmos filtros da tela de estoque (query string) —
    o botao "Exportar Excel" submete o proprio form de filtros via `formaction`.
    Escopo de loja respeitado: usuario nao-admin so exporta as lojas permitidas.

    A coluna "Loja" reflete a loja do ultimo evento do chassi (onde a moto se
    encontra). Para chassis fora de estoque/aguardando NF (so quando o filtro
    "Mostrar fora de estoque" esta ativo) a loja pode vir vazia.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        flash('openpyxl nao disponivel no servidor.', 'danger')
        return redirect(url_for('hora.estoque_lista'))

    permitidas = lojas_permitidas_ids()
    filtros = _coletar_filtros_estoque()
    loja_id = filtros['loja_id']

    if loja_id and not usuario_tem_acesso_a_loja(loja_id):
        flash('Acesso negado a essa loja.', 'danger')
        return redirect(url_for('hora.estoque_lista'))

    motos = estoque_service.listar_estoque(
        lojas_permitidas_ids=permitidas,
        **filtros,
    )
    if not motos:
        flash('Nenhuma moto encontrada para os filtros aplicados.', 'warning')
        return redirect(url_for('hora.estoque_lista', **{
            k: v for k, v in request.args.items() if v
        }))

    # (titulo da coluna, chave no dict de listar_estoque)
    colunas = [
        ('Chassi', 'chassi'),
        ('Modelo', 'modelo_nome'),
        ('Cor', 'cor'),
        ('Loja', 'loja_nome'),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Estoque'
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(
        start_color='4F81BD', end_color='4F81BD', fill_type='solid',
    )

    for col_idx, (titulo, _chave) in enumerate(colunas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, m in enumerate(motos, start=2):
        for col_idx, (_titulo, chave) in enumerate(colunas, start=1):
            ws.cell(row=row_idx, column=col_idx, value=m.get(chave))

    # Auto-width best-effort (mesmo padrao de vendas_exportar_xlsx).
    for col_idx, (titulo, chave) in enumerate(colunas, start=1):
        max_len = max(
            [len(str(m.get(chave) or '')) for m in motos] + [len(titulo)]
        )
        ws.column_dimensions[
            ws.cell(row=1, column=col_idx).column_letter
        ].width = min(max_len + 2, 60)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    ts = agora_utc_naive().strftime('%Y%m%d_%H%M%S')
    filename = f'estoque_hora_{ts}.xlsx'
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


# ------------------------------------------------------------------
# Autocomplete endpoints (JSON) — usados pelos inputs da tela de estoque.
# ------------------------------------------------------------------

@hora_bp.route('/estoque/autocomplete/chassi')
@require_hora_perm('estoque', 'ver')
def estoque_autocomplete_chassi():
    q = request.args.get('q') or ''
    return jsonify(estoque_service.autocomplete_chassi(
        q=q,
        lojas_permitidas_ids=lojas_permitidas_ids(),
        limit=20,
    ))


@hora_bp.route('/estoque/autocomplete/modelo')
@require_hora_perm('estoque', 'ver')
def estoque_autocomplete_modelo():
    q = request.args.get('q') or ''
    return jsonify(estoque_service.autocomplete_modelo(
        q=q,
        lojas_permitidas_ids=lojas_permitidas_ids(),
        limit=20,
    ))


@hora_bp.route('/estoque/autocomplete/cor')
@require_hora_perm('estoque', 'ver')
def estoque_autocomplete_cor():
    q = request.args.get('q') or ''
    return jsonify(estoque_service.autocomplete_cor(
        q=q,
        lojas_permitidas_ids=lojas_permitidas_ids(),
        limit=20,
    ))


@hora_bp.route('/estoque/chassi/<numero_chassi>')
@require_hora_perm('estoque', 'ver')
def estoque_chassi_detalhe(numero_chassi: str):
    chassi = numero_chassi.strip().upper()
    moto = HoraMoto.query.get_or_404(chassi)

    rastreio = estoque_service.rastreamento_completo(chassi)

    # Autorizacao: usuario deve ter acesso a ALGUMA loja onde este chassi tem
    # evento OU esta em pedido/NF entrada/venda. Admin (lojas_permitidas_ids()
    # is None) sempre passa.
    permitidas = lojas_permitidas_ids()
    if permitidas is not None:
        if not permitidas:
            flash('Sem acesso a nenhuma loja.', 'danger')
            return redirect(url_for('hora.estoque_lista'))
        if not chassi_acessivel(chassi, permitidas):
            flash('Acesso negado a esse chassi.', 'danger')
            return redirect(url_for('hora.estoque_lista'))

    return render_template(
        'hora/estoque_chassi_detalhe.html',
        moto=moto,
        rastreio=rastreio,
    )
