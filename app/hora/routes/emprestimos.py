"""Rotas de empréstimo de moto entre nossa loja HORA e loja externa."""
from __future__ import annotations

from datetime import date as _date
from io import BytesIO

from flask import Response, flash, redirect, render_template, request, url_for
from flask_login import current_user

from app.hora.decorators import require_hora_perm
from app.hora.models import (
    HoraEmprestimoMoto, HoraLoja, HoraModelo,
    EMPRESTIMO_TIPO_SAIDA, EMPRESTIMO_TIPO_ENTRADA,
    EMPRESTIMO_STATUS_EM_ABERTO, EMPRESTIMO_STATUS_RESSARCIDO,
    EMPRESTIMO_STATUS_CANCELADO,
)
from app.hora.routes import hora_bp
from app.hora.services import emprestimo_service
from app.hora.services.auth_helper import (
    lojas_permitidas_ids, usuario_tem_acesso_a_loja,
)
from app.utils.timezone import agora_utc_naive


def _operador() -> str:
    return getattr(current_user, 'nome', None) or 'desconhecido'


def _lojas_ativas_permitidas():
    permitidas = lojas_permitidas_ids()
    q = HoraLoja.query.filter_by(ativa=True)
    if permitidas is not None:
        if not permitidas:
            return []
        q = q.filter(HoraLoja.id.in_(permitidas))
    return q.order_by(HoraLoja.nome).all()


# ------------------------------------------------------------------------
# Listagem
# ------------------------------------------------------------------------

@hora_bp.route('/emprestimos')
@require_hora_perm('emprestimos', 'ver')
def emprestimos_lista():
    from datetime import datetime as _dt

    try:
        page = int(request.args.get('page', 1))
    except (TypeError, ValueError):
        page = 1
    try:
        per_page = int(request.args.get('per_page', 50))
    except (TypeError, ValueError):
        per_page = 50

    status_filtro = (request.args.get('status') or '').strip().upper() or None
    if status_filtro and status_filtro not in (
        EMPRESTIMO_STATUS_EM_ABERTO,
        EMPRESTIMO_STATUS_RESSARCIDO,
        EMPRESTIMO_STATUS_CANCELADO,
    ):
        status_filtro = None

    tipo_filtro = (request.args.get('tipo') or '').strip().upper() or None
    if tipo_filtro and tipo_filtro not in (
        EMPRESTIMO_TIPO_SAIDA, EMPRESTIMO_TIPO_ENTRADA,
    ):
        tipo_filtro = None

    chassi = (request.args.get('chassi') or '').strip() or None
    loja_hora_id_str = (request.args.get('loja_hora_id') or '').strip()
    loja_hora_id = int(loja_hora_id_str) if loja_hora_id_str.isdigit() else None
    modelo_id_str = (request.args.get('modelo_id') or '').strip()
    modelo_id = int(modelo_id_str) if modelo_id_str.isdigit() else None
    loja_externa = (request.args.get('loja_externa') or '').strip() or None
    data_ini_str = (request.args.get('data_inicio') or '').strip()
    data_fim_str = (request.args.get('data_fim') or '').strip()

    try:
        data_inicio = _dt.strptime(data_ini_str, '%Y-%m-%d').date() if data_ini_str else None
        data_fim = _dt.strptime(data_fim_str, '%Y-%m-%d').date() if data_fim_str else None
    except ValueError:
        flash('Data invalida (use formato YYYY-MM-DD).', 'warning')
        data_inicio = None
        data_fim = None

    permitidas = lojas_permitidas_ids()
    if loja_hora_id and not usuario_tem_acesso_a_loja(loja_hora_id):
        flash('Acesso negado a essa loja.', 'danger')
        return redirect(url_for('hora.emprestimos_lista'))

    pagination = emprestimo_service.paginar_emprestimos(
        page=page, per_page=per_page,
        lojas_permitidas_ids=permitidas,
        status=status_filtro,
        tipo=tipo_filtro,
        chassi=chassi,
        loja_hora_id=loja_hora_id,
        modelo_id=modelo_id,
        loja_externa=loja_externa,
        data_inicio=data_inicio,
        data_fim=data_fim,
    )

    lojas_ativas = _lojas_ativas_permitidas()
    modelos_ativos = HoraModelo.query.filter_by(ativo=True).order_by(HoraModelo.nome_modelo).all()

    return render_template(
        'hora/emprestimos_lista.html',
        pagination=pagination,
        emprestimos=(pagination.items if pagination else []),
        status_filtro=status_filtro,
        tipo_filtro=tipo_filtro,
        per_page=per_page,
        filtro_chassi=chassi,
        filtro_loja_hora_id=loja_hora_id,
        filtro_modelo_id=modelo_id,
        filtro_loja_externa=loja_externa,
        filtro_data_inicio=data_ini_str,
        filtro_data_fim=data_fim_str,
        lojas_ativas=lojas_ativas,
        modelos_ativos=modelos_ativos,
    )


# ------------------------------------------------------------------------
# Novo emprestimo
# ------------------------------------------------------------------------

@hora_bp.route('/emprestimos/novo', methods=['GET', 'POST'])
@require_hora_perm('emprestimos', 'criar')
def emprestimos_novo():
    lojas_ativas = _lojas_ativas_permitidas()
    modelos = HoraModelo.query.filter_by(ativo=True).order_by(HoraModelo.nome_modelo).all()

    if request.method == 'POST':
        tipo = (request.form.get('tipo') or '').strip().upper()
        loja_str = (request.form.get('loja_hora_id') or '').strip()
        modelo_str = (request.form.get('modelo_id') or '').strip()
        chassi = (request.form.get('chassi') or '').strip()
        loja_externa_nome = (request.form.get('loja_externa_nome') or '').strip()
        loja_externa_cnpj = (request.form.get('loja_externa_cnpj') or '').strip()
        data_str = (request.form.get('data_emprestimo') or '').strip()
        observacoes = (request.form.get('observacoes') or '').strip()

        try:
            loja_id = int(loja_str)
            modelo_id = int(modelo_str)
        except (TypeError, ValueError):
            flash('Loja e modelo obrigatorios.', 'danger')
            return render_template('hora/emprestimo_novo.html',
                                   lojas=lojas_ativas, modelos=modelos)

        if lojas_permitidas_ids() is not None and not usuario_tem_acesso_a_loja(loja_id):
            flash('Acesso negado para essa loja.', 'danger')
            return redirect(url_for('hora.emprestimos_lista'))

        try:
            data_emp = _date.fromisoformat(data_str) if data_str else _date.today()
        except ValueError:
            flash(f'Data invalida: {data_str!r}', 'danger')
            return render_template('hora/emprestimo_novo.html',
                                   lojas=lojas_ativas, modelos=modelos)

        try:
            emp = emprestimo_service.criar_emprestimo(
                tipo=tipo,
                loja_hora_id=loja_id,
                loja_externa_nome=loja_externa_nome,
                loja_externa_cnpj=loja_externa_cnpj or None,
                modelo_id=modelo_id,
                data_emprestimo=data_emp,
                chassi=chassi,
                observacoes=observacoes or None,
                operador=_operador(),
            )
            flash(
                f'Emprestimo #{emp.id} ({emp.tipo}) criado para "{emp.loja_externa_nome}".',
                'success',
            )
            return redirect(url_for('hora.emprestimos_detalhe', emprestimo_id=emp.id))
        except emprestimo_service.EmprestimoError as exc:
            flash(f'Erro: {exc}', 'danger')

    return render_template(
        'hora/emprestimo_novo.html',
        lojas=lojas_ativas,
        modelos=modelos,
    )


# ------------------------------------------------------------------------
# Detalhe + acoes
# ------------------------------------------------------------------------

@hora_bp.route('/emprestimos/<int:emprestimo_id>')
@require_hora_perm('emprestimos', 'ver')
def emprestimos_detalhe(emprestimo_id: int):
    emp = HoraEmprestimoMoto.query.get_or_404(emprestimo_id)
    if not usuario_tem_acesso_a_loja(emp.loja_hora_id):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('hora.emprestimos_lista'))
    return render_template('hora/emprestimo_detalhe.html', emp=emp)


@hora_bp.route('/emprestimos/<int:emprestimo_id>/ressarcir', methods=['POST'])
@require_hora_perm('emprestimos', 'editar')
def emprestimos_ressarcir(emprestimo_id: int):
    emp = HoraEmprestimoMoto.query.get_or_404(emprestimo_id)
    if not usuario_tem_acesso_a_loja(emp.loja_hora_id):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('hora.emprestimos_lista'))

    chassi = (request.form.get('chassi') or '').strip()
    data_str = (request.form.get('data_ressarcimento') or '').strip()
    observacoes = (request.form.get('observacoes_extra') or '').strip()
    try:
        data_r = _date.fromisoformat(data_str) if data_str else _date.today()
    except ValueError:
        flash(f'Data invalida: {data_str!r}', 'danger')
        return redirect(url_for('hora.emprestimos_detalhe', emprestimo_id=emp.id))

    try:
        emprestimo_service.ressarcir_emprestimo(
            emprestimo_id=emp.id,
            chassi_ressarcimento=chassi,
            data_ressarcimento=data_r,
            observacoes_extra=observacoes or None,
            operador=_operador(),
        )
        flash(f'Emprestimo #{emp.id} ressarcido com sucesso.', 'success')
    except emprestimo_service.EmprestimoError as exc:
        flash(f'Erro: {exc}', 'danger')
    return redirect(url_for('hora.emprestimos_detalhe', emprestimo_id=emp.id))


@hora_bp.route('/emprestimos/<int:emprestimo_id>/cancelar', methods=['POST'])
@require_hora_perm('emprestimos', 'apagar')
def emprestimos_cancelar(emprestimo_id: int):
    emp = HoraEmprestimoMoto.query.get_or_404(emprestimo_id)
    if not usuario_tem_acesso_a_loja(emp.loja_hora_id):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('hora.emprestimos_lista'))

    motivo = (request.form.get('motivo') or '').strip()
    try:
        emprestimo_service.cancelar_emprestimo(
            emprestimo_id=emp.id,
            motivo=motivo,
            operador=_operador(),
        )
        flash(f'Emprestimo #{emp.id} cancelado.', 'warning')
    except emprestimo_service.EmprestimoError as exc:
        flash(f'Erro: {exc}', 'danger')
    return redirect(url_for('hora.emprestimos_detalhe', emprestimo_id=emp.id))


# ------------------------------------------------------------------------
# Export Excel (1 aba — Emprestimos com seus 2 chassis)
# ------------------------------------------------------------------------

@hora_bp.route('/emprestimos/exportar.xlsx')
@require_hora_perm('emprestimos', 'ver')
def emprestimos_exportar_xlsx():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        flash('openpyxl nao disponivel.', 'danger')
        return redirect(url_for('hora.emprestimos_lista'))

    permitidas = lojas_permitidas_ids()
    q = HoraEmprestimoMoto.query.order_by(
        HoraEmprestimoMoto.data_emprestimo.desc(),
        HoraEmprestimoMoto.id.desc(),
    )
    if permitidas is not None:
        if not permitidas:
            flash('Sem lojas permitidas.', 'warning')
            return redirect(url_for('hora.emprestimos_lista'))
        q = q.filter(HoraEmprestimoMoto.loja_hora_id.in_(permitidas))

    status = (request.args.get('status') or '').strip().upper() or None
    if status:
        q = q.filter(HoraEmprestimoMoto.status == status)
    tipo = (request.args.get('tipo') or '').strip().upper() or None
    if tipo:
        q = q.filter(HoraEmprestimoMoto.tipo == tipo)

    emprestimos = q.all()
    if not emprestimos:
        flash('Nenhum emprestimo encontrado.', 'warning')
        return redirect(url_for('hora.emprestimos_lista'))

    cab = [
        'id', 'tipo', 'status',
        'loja_hora', 'loja_externa_nome', 'loja_externa_cnpj',
        'modelo',
        'chassi_saida', 'chassi_entrada',
        'data_emprestimo', 'data_ressarcimento',
        'observacoes',
        'criado_em', 'criado_por',
        'ressarcido_em', 'ressarcido_por',
        'cancelado_em', 'cancelado_por', 'cancelamento_motivo',
    ]

    def _dt(v):
        return v.strftime('%d/%m/%Y %H:%M') if v else None

    def _data(v):
        return v.strftime('%d/%m/%Y') if v else None

    linhas = []
    for e in emprestimos:
        linhas.append({
            'id': e.id,
            'tipo': e.tipo,
            'status': e.status,
            'loja_hora': e.loja_hora.rotulo_display if e.loja_hora else None,
            'loja_externa_nome': e.loja_externa_nome,
            'loja_externa_cnpj': e.loja_externa_cnpj,
            'modelo': e.modelo.nome_modelo if e.modelo else None,
            'chassi_saida': e.chassi_saida,
            'chassi_entrada': e.chassi_entrada,
            'data_emprestimo': _data(e.data_emprestimo),
            'data_ressarcimento': _data(e.data_ressarcimento),
            'observacoes': e.observacoes,
            'criado_em': _dt(e.criado_em),
            'criado_por': e.criado_por,
            'ressarcido_em': _dt(e.ressarcido_em),
            'ressarcido_por': e.ressarcido_por,
            'cancelado_em': _dt(e.cancelado_em),
            'cancelado_por': e.cancelado_por,
            'cancelamento_motivo': e.cancelamento_motivo,
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Emprestimos'
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    for col_idx, campo in enumerate(cab, start=1):
        c = ws.cell(row=1, column=col_idx, value=campo)
        c.font = header_font
        c.fill = header_fill
    for row_idx, item in enumerate(linhas, start=2):
        for col_idx, campo in enumerate(cab, start=1):
            ws.cell(row=row_idx, column=col_idx, value=item.get(campo))
    for col_idx, campo in enumerate(cab, start=1):
        max_len = max(
            [len(str(item.get(campo) or '')) for item in linhas] + [len(campo)]
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 60)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = agora_utc_naive().strftime('%Y%m%d_%H%M%S')
    filename = f'emprestimos_motos_{ts}.xlsx'
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )
