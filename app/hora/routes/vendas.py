"""Rotas de Pedido de Venda (HORA -> consumidor final).

Inclui: upload DANFE legado, criacao manual via TagPlus, listagem, detalhe,
edicao de header e itens, transicoes (confirmar, cancelar), definicao de loja,
resolucao de divergencias, exportacao em Excel.
"""
from __future__ import annotations

from decimal import Decimal, InvalidOperation
from io import BytesIO

from flask import Response, flash, jsonify, redirect, render_template, request, url_for
from flask_login import current_user

from app.hora.decorators import require_hora_perm
from app.hora.models import (
    HoraLoja,
    HoraVenda,
    HoraVendaDivergencia,
    VENDA_STATUS_COTACAO,
)
from app.hora.models.tagplus import HoraTagPlusFormaPagamentoMap
from app.hora.routes import hora_bp
from app.hora.services import venda_service
from app.hora.services.auth_helper import (
    lojas_permitidas_ids,
    usuario_tem_acesso_a_loja,
)
from app.hora.services.parsers.danfe_adapter import DanfeParseError
from app.utils.timezone import agora_utc_naive


def _lojas_ativas_permitidas():
    permitidas = lojas_permitidas_ids()
    q = HoraLoja.query.filter_by(ativa=True)
    if permitidas is not None:
        if not permitidas:
            return []
        q = q.filter(HoraLoja.id.in_(permitidas))
    return q.order_by(HoraLoja.nome).all()


def _operador_atual() -> str:
    return getattr(current_user, 'nome', None) or 'desconhecido'


# ------------------------------------------------------------------------
# Listagem
# ------------------------------------------------------------------------

@hora_bp.route('/vendas')
@require_hora_perm('vendas', 'ver')
def vendas_lista():
    try:
        page = int(request.args.get('page', 1))
    except (TypeError, ValueError):
        page = 1
    try:
        per_page = int(request.args.get('per_page', 50))
    except (TypeError, ValueError):
        per_page = 50

    status_filtro = (request.args.get('status') or '').strip().upper() or None
    if status_filtro and status_filtro not in (
        'COTACAO', 'CONFIRMADO', 'FATURADO', 'CANCELADO',
    ):
        status_filtro = None

    busca = (request.args.get('busca') or '').strip() or None
    loja_id_str = (request.args.get('loja_id') or '').strip()
    loja_id = int(loja_id_str) if loja_id_str.isdigit() else None
    data_ini_str = (request.args.get('data_inicio') or '').strip()
    data_fim_str = (request.args.get('data_fim') or '').strip()
    filtro_chassi = (request.args.get('chassi') or '').strip().upper() or None

    from datetime import datetime as _dt
    try:
        data_inicio = _dt.strptime(data_ini_str, '%Y-%m-%d').date() if data_ini_str else None
        data_fim = _dt.strptime(data_fim_str, '%Y-%m-%d').date() if data_fim_str else None
    except ValueError:
        flash('Data invalida (use formato YYYY-MM-DD).', 'warning')
        data_inicio = None
        data_fim = None

    if loja_id and not usuario_tem_acesso_a_loja(loja_id):
        flash('Acesso negado a essa loja.', 'danger')
        return redirect(url_for('hora.vendas_lista'))

    permitidas = lojas_permitidas_ids()
    pagination = venda_service.paginar_vendas(
        page=page, per_page=per_page,
        lojas_permitidas_ids=permitidas,
        status=status_filtro,
        busca=busca,
        loja_id=loja_id,
        data_inicio=data_inicio,
        data_fim=data_fim,
        chassi=filtro_chassi,
    )

    # Lojas para filtro
    lojas_q = HoraLoja.query.filter_by(ativa=True)
    if permitidas is not None:
        if not permitidas:
            lojas_lista = []
        else:
            lojas_q = lojas_q.filter(HoraLoja.id.in_(permitidas))
            lojas_lista = lojas_q.order_by(HoraLoja.nome).all()
    else:
        lojas_lista = lojas_q.order_by(HoraLoja.nome).all()

    return render_template(
        'hora/vendas_lista.html',
        pagination=pagination,
        vendas=(pagination.items if pagination else []),
        status_filtro=status_filtro,
        per_page=per_page,
        filtro_busca=busca,
        filtro_loja_id=loja_id,
        filtro_data_inicio=data_ini_str,
        filtro_data_fim=data_fim_str,
        filtro_chassi=filtro_chassi,
        lojas_ativas=lojas_lista,
    )


# ------------------------------------------------------------------------
# Upload DANFE -> cria venda
# ------------------------------------------------------------------------

@hora_bp.route('/vendas/upload', methods=['GET', 'POST'])
@require_hora_perm('vendas', 'criar')
def vendas_upload():
    """Upload DANFE PDF emitido pela loja HORA — aceita 1 ou N arquivos
    (BACKFILL em lote). Parseia via adapter e cria HoraVenda + itens +
    eventos VENDIDA + divergencias para cada PDF.

    Nao pede loja nem cliente — tudo vem da NF. Operador preenche vendedor e
    forma_pagamento depois na tela de detalhe.
    """
    if request.method == 'POST':
        # Aceita campo `pdf` (legacy 1 arquivo) e `pdfs` (multiplo).
        arquivos = request.files.getlist('pdfs') or []
        if not arquivos:
            unico = request.files.get('pdf')
            if unico and unico.filename:
                arquivos = [unico]

        arquivos = [a for a in arquivos if a and a.filename]
        if not arquivos:
            flash('Selecione ao menos 1 arquivo PDF.', 'danger')
            return render_template('hora/venda_upload.html', resultados=[])

        resultados = []
        operador = _operador_atual()
        for arq in arquivos:
            entry = {
                'arquivo': arq.filename,
                'status': None,         # 'sucesso' | 'duplicado' | 'erro'
                'venda_id': None,
                'numero_nf': None,
                'qtd_chassis': 0,
                'qtd_divergencias': 0,
                'mensagem': '',
            }
            try:
                pdf_bytes = arq.read()
                venda = venda_service.importar_nf_saida_pdf(
                    pdf_bytes=pdf_bytes,
                    nome_arquivo_origem=arq.filename,
                    criado_por=operador,
                )
                entry.update({
                    'status': 'sucesso',
                    'venda_id': venda.id,
                    'numero_nf': venda.nf_saida_numero,
                    'qtd_chassis': len(venda.itens),
                    'qtd_divergencias': len(venda.divergencias_abertas),
                    'mensagem': (
                        f'NF {venda.nf_saida_numero} importada — '
                        f'{len(venda.itens)} chassi(s) para {venda.nome_cliente}.'
                    ),
                })
            except venda_service.NfSaidaJaImportada as exc:
                entry['status'] = 'duplicado'
                entry['mensagem'] = str(exc)
            except (ValueError, DanfeParseError) as exc:
                entry['status'] = 'erro'
                entry['mensagem'] = f'Erro ao importar: {exc}'
            except Exception as exc:  # pragma: no cover
                entry['status'] = 'erro'
                entry['mensagem'] = f'Erro inesperado: {exc}'
            resultados.append(entry)

        # Resumo flash + tela de resultados.
        n_ok = sum(1 for r in resultados if r['status'] == 'sucesso')
        n_dup = sum(1 for r in resultados if r['status'] == 'duplicado')
        n_err = sum(1 for r in resultados if r['status'] == 'erro')
        n_div = sum(r['qtd_divergencias'] for r in resultados)
        if len(resultados) == 1 and n_ok == 1 and n_div == 0:
            # Caso 1-arquivo bem sucedido: mantem comportamento legado
            # (redireciona pro detalhe).
            return redirect(url_for(
                'hora.vendas_detalhe', venda_id=resultados[0]['venda_id'],
            ))
        flash(
            f'Backfill: {n_ok} ok · {n_dup} duplicado(s) · {n_err} erro(s) '
            f'· {n_div} divergencia(s) total.',
            'warning' if (n_err or n_dup or n_div) else 'success',
        )
        return render_template('hora/venda_upload.html', resultados=resultados)

    return render_template('hora/venda_upload.html', resultados=[])


# ------------------------------------------------------------------------
# Detalhe
# ------------------------------------------------------------------------

@hora_bp.route('/vendas/<int:venda_id>')
@require_hora_perm('vendas', 'ver')
def vendas_detalhe(venda_id: int):
    venda = HoraVenda.query.get_or_404(venda_id)
    if venda.loja_id and not usuario_tem_acesso_a_loja(venda.loja_id):
        flash('Acesso negado: venda de loja fora do seu escopo.', 'danger')
        return redirect(url_for('hora.vendas_lista'))
    # Venda sem loja (loja_id=NULL) so e acessivel para admin (lojas_permitidas_ids=None).
    if not venda.loja_id and lojas_permitidas_ids() is not None:
        flash(
            'Esta venda ainda nao tem loja definida — apenas administradores podem abrir.',
            'warning',
        )
        return redirect(url_for('hora.vendas_lista'))

    # Sempre popular lojas_ativas — operador pode trocar loja em vendas
    # backfilladas que caem na matriz por causa do CNPJ emitente (regra
    # fiscal: NFe HORA sai sempre com CNPJ da matriz). Filtrado por escopo.
    lojas_ativas = _lojas_ativas_permitidas()

    # Formas de pagamento dinamicas: mapeamentos cadastrados em
    # HoraTagPlusFormaPagamentoMap (mesma fonte usada no formulario de pedido
    # de venda manual). 'NAO_INFORMADO' eh sentinela (default da coluna) e
    # sempre aparece como primeira opcao.
    formas_pagamento = (
        HoraTagPlusFormaPagamentoMap.query
        .order_by(HoraTagPlusFormaPagamentoMap.forma_pagamento_hora)
        .all()
    )

    return render_template(
        'hora/venda_detalhe.html',
        venda=venda,
        lojas_ativas=lojas_ativas,
        formas_pagamento=formas_pagamento,
    )


# ------------------------------------------------------------------------
# Download PDF
# ------------------------------------------------------------------------

@hora_bp.route('/vendas/<int:venda_id>/download-pdf')
@require_hora_perm('vendas', 'ver')
def vendas_download_pdf(venda_id: int):
    venda = HoraVenda.query.get_or_404(venda_id)
    if venda.loja_id and not usuario_tem_acesso_a_loja(venda.loja_id):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('hora.vendas_lista'))
    if not venda.arquivo_pdf_s3_key:
        flash('PDF desta venda nao esta armazenado.', 'warning')
        return redirect(url_for('hora.vendas_detalhe', venda_id=venda.id))

    from app.utils.file_storage import FileStorage
    url = FileStorage().get_file_url(venda.arquivo_pdf_s3_key)
    if not url:
        flash('Falha ao gerar URL do PDF.', 'danger')
        return redirect(url_for('hora.vendas_detalhe', venda_id=venda.id))
    return redirect(url)


# ------------------------------------------------------------------------
# Editar (vendedor, forma_pagamento, contato, observacoes)
# ------------------------------------------------------------------------

@hora_bp.route('/vendas/<int:venda_id>/editar', methods=['POST'])
@require_hora_perm('vendas', 'editar')
def vendas_editar(venda_id: int):
    """Edita campos do header. Conjunto de campos permitidos varia por status:
    - COTACAO: tudo (incluindo cliente/endereco)
    - CONFIRMADO: contato/endereco/operacionais
    - FATURADO: so observacoes
    - CANCELADO: nada
    """
    venda = HoraVenda.query.get_or_404(venda_id)
    if venda.loja_id and not usuario_tem_acesso_a_loja(venda.loja_id):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('hora.vendas_lista'))
    if not venda.loja_id and lojas_permitidas_ids() is not None:
        flash('Pedido sem loja definida — apenas admin edita.', 'warning')
        return redirect(url_for('hora.vendas_lista'))

    # consumidor_final: checkbox + hidden flag de "campo presente no form".
    # Sem o flag, nao distinguimos "operador deixou tudo como esta" de
    # "operador desmarcou o checkbox". Quando flag=='1', interpretamos a
    # presenca/ausencia do name 'consumidor_final' (value='1') no POST.
    if request.form.get('consumidor_final_flag') == '1':
        consumidor_final_edit = request.form.get('consumidor_final') == '1'
    else:
        consumidor_final_edit = None

    try:
        venda_service.editar_venda(
            venda_id=venda.id,
            vendedor=request.form.get('vendedor'),
            forma_pagamento=request.form.get('forma_pagamento'),
            telefone_cliente=request.form.get('telefone_cliente'),
            email_cliente=request.form.get('email_cliente'),
            observacoes=request.form.get('observacoes'),
            nome_cliente=request.form.get('nome_cliente'),
            cpf_cliente=request.form.get('cpf_cliente'),
            cep=request.form.get('cep'),
            endereco_logradouro=request.form.get('endereco_logradouro'),
            endereco_numero=request.form.get('endereco_numero'),
            endereco_complemento=request.form.get('endereco_complemento'),
            endereco_bairro=request.form.get('endereco_bairro'),
            endereco_cidade=request.form.get('endereco_cidade'),
            endereco_uf=request.form.get('endereco_uf'),
            modalidade_frete=request.form.get('modalidade_frete'),
            numero_parcelas=request.form.get('numero_parcelas'),
            intervalo_parcelas_dias=request.form.get('intervalo_parcelas_dias'),
            consumidor_final=consumidor_final_edit,
            usuario=_operador_atual(),
        )
        flash('Pedido atualizado.', 'success')
    except (ValueError, venda_service.TransicaoInvalidaError) as exc:
        flash(f'Erro: {exc}', 'danger')
    return redirect(url_for('hora.vendas_detalhe', venda_id=venda.id))


# ------------------------------------------------------------------------
# Editar formas de pagamento (multi-formas — INCOMPLETO ou COTACAO)
# ------------------------------------------------------------------------

@hora_bp.route('/vendas/<int:venda_id>/pagamentos/editar', methods=['POST'])
@require_hora_perm('vendas', 'editar')
def vendas_pagamentos_editar(venda_id: int):
    """Substitui a lista de formas de pagamento de um pedido.

    Permitido em status INCOMPLETO ou COTACAO. Apos a edicao, o service
    re-avalia status (INCOMPLETO se soma!=total ou aut_id faltando, senao
    COTACAO). UI deve mostrar o novo status na resposta.

    Form-array esperado (cada index = 1 forma):
      pagamento_forma[]    forma_pagamento_hora (string)
      pagamento_valor[]    valor (string com ',' BR ou '.' US)
      pagamento_parcelas[] numero_parcelas (int, default 1)
      pagamento_aut_id[]   aut_id (string, opcional)
    """
    venda = HoraVenda.query.get_or_404(venda_id)
    if venda.loja_id and not usuario_tem_acesso_a_loja(venda.loja_id):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('hora.vendas_lista'))
    if not venda.loja_id and lojas_permitidas_ids() is not None:
        flash('Pedido sem loja definida — apenas admin edita.', 'warning')
        return redirect(url_for('hora.vendas_lista'))

    formas_lista = request.form.getlist('pagamento_forma')
    valores_lista = request.form.getlist('pagamento_valor')
    parcelas_lista = request.form.getlist('pagamento_parcelas')
    aut_ids_lista = request.form.getlist('pagamento_aut_id')

    pagamentos: list[dict] = []
    for i, forma_raw in enumerate(formas_lista):
        forma = (forma_raw or '').strip().upper()
        if not forma:
            continue
        valor_raw = valores_lista[i] if i < len(valores_lista) else '0'
        valor_str = (valor_raw or '').strip()
        if ',' in valor_str:
            valor_str = valor_str.replace('.', '').replace(',', '.')
        try:
            valor = Decimal(valor_str) if valor_str else Decimal('0')
        except (InvalidOperation, ValueError):
            continue
        try:
            par = int((parcelas_lista[i] if i < len(parcelas_lista) else '1') or '1')
        except ValueError:
            par = 1
        aut = (aut_ids_lista[i] if i < len(aut_ids_lista) else '').strip() or None
        pagamentos.append({
            'forma_pagamento_hora': forma,
            'valor': valor,
            'numero_parcelas': par,
            'aut_id': aut,
        })

    try:
        res = venda_service.editar_pagamentos(
            venda_id=venda.id, pagamentos=pagamentos, usuario=_operador_atual(),
        )
    except venda_service.TransicaoInvalidaError as exc:
        flash(f'Erro: {exc}', 'danger')
        return redirect(url_for('hora.vendas_detalhe', venda_id=venda.id))
    except ValueError as exc:
        flash(f'Dados invalidos: {exc}', 'danger')
        return redirect(url_for('hora.vendas_detalhe', venda_id=venda.id))

    if res.status == 'INCOMPLETO':
        flash(
            f'Pagamentos atualizados — pedido continua INCOMPLETO. Verifique '
            f'soma vs valor total e AUT/ID das formas que exigem.',
            'warning',
        )
    else:
        flash(
            f'Pagamentos atualizados — pedido pronto para confirmacao.',
            'success',
        )
    return redirect(url_for('hora.vendas_detalhe', venda_id=venda.id))


# ------------------------------------------------------------------------
# Confirmar pedido (COTACAO -> CONFIRMADO)
# ------------------------------------------------------------------------

@hora_bp.route('/vendas/<int:venda_id>/confirmar', methods=['POST'])
@require_hora_perm('vendas', 'aprovar')
def vendas_confirmar(venda_id: int):
    venda = HoraVenda.query.get_or_404(venda_id)
    if venda.loja_id and not usuario_tem_acesso_a_loja(venda.loja_id):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('hora.vendas_lista'))
    if not venda.loja_id and lojas_permitidas_ids() is not None:
        flash('Pedido sem loja definida — apenas admin confirma.', 'warning')
        return redirect(url_for('hora.vendas_lista'))

    try:
        venda_service.confirmar_venda(
            venda_id=venda.id, usuario=_operador_atual(),
        )
        flash(
            f'Pedido #{venda.id} confirmado. Pronto para emissao de NFe.',
            'success',
        )
    except (ValueError, venda_service.TransicaoInvalidaError) as exc:
        flash(f'Erro: {exc}', 'danger')
    return redirect(url_for('hora.vendas_detalhe', venda_id=venda.id))


# ------------------------------------------------------------------------
# Voltar para cotacao (CONFIRMADO -> COTACAO, para reabrir edicao)
# ------------------------------------------------------------------------

@hora_bp.route('/vendas/<int:venda_id>/voltar-cotacao', methods=['POST'])
@require_hora_perm('vendas', 'editar')
def vendas_voltar_cotacao(venda_id: int):
    venda = HoraVenda.query.get_or_404(venda_id)
    if venda.loja_id and not usuario_tem_acesso_a_loja(venda.loja_id):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('hora.vendas_lista'))
    try:
        venda_service.voltar_para_cotacao(
            venda_id=venda.id, usuario=_operador_atual(),
        )
        flash(
            f'Pedido #{venda.id} voltou para COTACAO. Edite e confirme novamente.',
            'success',
        )
    except (ValueError, venda_service.TransicaoInvalidaError) as exc:
        flash(f'Erro: {exc}', 'danger')
    return redirect(url_for('hora.vendas_detalhe', venda_id=venda.id))


# ------------------------------------------------------------------------
# Edicao de itens (apenas em COTACAO)
# ------------------------------------------------------------------------

def _parse_decimal_form(valor_raw: str) -> Decimal:
    valor_raw = (valor_raw or '').strip()
    try:
        return (
            Decimal(valor_raw.replace('.', '').replace(',', '.'))
            if ',' in valor_raw else Decimal(valor_raw)
        )
    except (InvalidOperation, ValueError):
        raise ValueError(f'Valor invalido: {valor_raw!r}')


@hora_bp.route('/vendas/<int:venda_id>/itens/adicionar', methods=['POST'])
@require_hora_perm('vendas', 'editar')
def vendas_item_adicionar(venda_id: int):
    venda = HoraVenda.query.get_or_404(venda_id)
    if venda.loja_id and not usuario_tem_acesso_a_loja(venda.loja_id):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('hora.vendas_lista'))
    if venda.status != VENDA_STATUS_COTACAO:
        flash('Itens so podem ser adicionados em pedidos COTACAO.', 'warning')
        return redirect(url_for('hora.vendas_detalhe', venda_id=venda.id))

    chassi = (request.form.get('numero_chassi') or '').strip()
    valor_raw = request.form.get('valor_final', '')
    try:
        valor = _parse_decimal_form(valor_raw)
        venda_service.adicionar_item_pedido(
            venda_id=venda.id,
            numero_chassi=chassi,
            valor_final=valor,
            usuario=_operador_atual(),
        )
        flash(f'Item adicionado ao pedido #{venda.id}.', 'success')
    except (ValueError, venda_service.ChassiIndisponivelError,
            venda_service.TransicaoInvalidaError) as exc:
        flash(f'Erro: {exc}', 'danger')
    return redirect(url_for('hora.vendas_detalhe', venda_id=venda.id))


@hora_bp.route('/vendas/<int:venda_id>/itens/<int:item_id>/remover', methods=['POST'])
@require_hora_perm('vendas', 'editar')
def vendas_item_remover(venda_id: int, item_id: int):
    venda = HoraVenda.query.get_or_404(venda_id)
    if venda.loja_id and not usuario_tem_acesso_a_loja(venda.loja_id):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('hora.vendas_lista'))
    try:
        venda_service.remover_item_pedido(
            venda_id=venda.id, item_id=item_id,
            usuario=_operador_atual(),
        )
        flash('Item removido.', 'success')
    except (ValueError, venda_service.TransicaoInvalidaError) as exc:
        flash(f'Erro: {exc}', 'danger')
    return redirect(url_for('hora.vendas_detalhe', venda_id=venda.id))


@hora_bp.route('/vendas/<int:venda_id>/itens/<int:item_id>/editar', methods=['POST'])
@require_hora_perm('vendas', 'editar')
def vendas_item_editar(venda_id: int, item_id: int):
    venda = HoraVenda.query.get_or_404(venda_id)
    if venda.loja_id and not usuario_tem_acesso_a_loja(venda.loja_id):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('hora.vendas_lista'))

    novo_chassi = (request.form.get('novo_chassi') or '').strip() or None
    valor_raw = (request.form.get('valor_final') or '').strip()
    novo_valor = None
    if valor_raw:
        try:
            novo_valor = _parse_decimal_form(valor_raw)
        except ValueError as exc:
            flash(f'Erro: {exc}', 'danger')
            return redirect(url_for('hora.vendas_detalhe', venda_id=venda.id))

    try:
        venda_service.editar_item_pedido(
            venda_id=venda.id, item_id=item_id,
            novo_chassi=novo_chassi, novo_valor=novo_valor,
            usuario=_operador_atual(),
        )
        flash('Item atualizado.', 'success')
    except (ValueError, venda_service.ChassiIndisponivelError,
            venda_service.TransicaoInvalidaError) as exc:
        flash(f'Erro: {exc}', 'danger')
    return redirect(url_for('hora.vendas_detalhe', venda_id=venda.id))


# ------------------------------------------------------------------------
# Definir loja (resolve CNPJ_DESCONHECIDO)
# ------------------------------------------------------------------------

@hora_bp.route('/vendas/<int:venda_id>/definir-loja', methods=['POST'])
@require_hora_perm('vendas', 'editar')
def vendas_definir_loja(venda_id: int):
    """Define ou TROCA a loja do pedido.

    - Vendas sem loja (CNPJ_DESCONHECIDO): apenas admin pode definir.
    - Vendas com loja: usuario com permissao `vendas/editar` na loja
      atual pode trocar (caso tipico do backfill TagPlus, onde todas
      caem na matriz por causa do CNPJ emitente).
    """
    venda = HoraVenda.query.get_or_404(venda_id)

    if venda.loja_id is None:
        # Apenas admin abre venda com loja=NULL — coerente com vendas_detalhe.
        if lojas_permitidas_ids() is not None:
            flash('Apenas administradores podem definir loja de venda sem loja.', 'danger')
            return redirect(url_for('hora.vendas_lista'))
    else:
        # Trocar loja: usuario precisa ter acesso a loja atual.
        if not usuario_tem_acesso_a_loja(venda.loja_id):
            flash('Acesso negado.', 'danger')
            return redirect(url_for('hora.vendas_lista'))

    loja_str = (request.form.get('loja_id') or '').strip()
    if not loja_str.isdigit():
        flash('Selecione a loja.', 'danger')
        return redirect(url_for('hora.vendas_detalhe', venda_id=venda.id))

    nova_loja_id = int(loja_str)
    # Usuario nao-admin nao pode mover venda para loja fora do seu escopo.
    permitidas = lojas_permitidas_ids()
    if permitidas is not None and nova_loja_id not in permitidas:
        flash(
            'Voce nao tem permissao na loja destino — peca para um administrador.',
            'danger',
        )
        return redirect(url_for('hora.vendas_detalhe', venda_id=venda.id))

    try:
        venda_atualizada = venda_service.definir_loja_venda(
            venda_id=venda.id,
            loja_id=nova_loja_id,
            usuario=_operador_atual(),
        )
        if venda.loja_id is None and venda_atualizada.loja_id == nova_loja_id:
            flash('Loja definida e divergencia CNPJ_DESCONHECIDO resolvida.', 'success')
        else:
            flash('Loja do pedido atualizada.', 'success')
    except ValueError as exc:
        flash(f'Erro: {exc}', 'danger')
    return redirect(url_for('hora.vendas_detalhe', venda_id=venda.id))


# ------------------------------------------------------------------------
# Cancelar venda
# ------------------------------------------------------------------------

@hora_bp.route('/vendas/<int:venda_id>/cancelar', methods=['POST'])
@require_hora_perm('vendas', 'apagar')
def vendas_cancelar(venda_id: int):
    venda = HoraVenda.query.get_or_404(venda_id)
    if venda.loja_id and not usuario_tem_acesso_a_loja(venda.loja_id):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('hora.vendas_lista'))
    if not venda.loja_id and lojas_permitidas_ids() is not None:
        flash('Venda sem loja definida — apenas admin cancela.', 'warning')
        return redirect(url_for('hora.vendas_lista'))

    motivo = (request.form.get('motivo') or '').strip()
    try:
        venda_service.cancelar_venda(
            venda_id=venda.id,
            motivo=motivo,
            usuario=_operador_atual(),
        )
        flash(
            f'Pedido #{venda.id} cancelado. Chassis devolvidos ao estoque.',
            'success',
        )
    except (ValueError, venda_service.TransicaoInvalidaError) as exc:
        flash(f'Erro: {exc}', 'danger')
    return redirect(url_for('hora.vendas_detalhe', venda_id=venda.id))


# ------------------------------------------------------------------------
# Descartar NF de teste (pos janela 24h SEFAZ) — admin only
# ------------------------------------------------------------------------

@hora_bp.route('/vendas/<int:venda_id>/descartar-teste', methods=['POST'])
@require_hora_perm('vendas_descarte', 'apagar')
def vendas_descartar_teste(venda_id: int):
    venda = HoraVenda.query.get_or_404(venda_id)
    motivo = (request.form.get('motivo') or '').strip()
    try:
        venda_service.descartar_venda_teste(
            venda_id=venda.id,
            motivo=motivo,
            usuario=_operador_atual(),
        )
        flash(
            f'Pedido #{venda.id} descartado (NF teste). NFe permanece valida na SEFAZ; '
            'chassis devolvidos ao estoque.',
            'success',
        )
    except (ValueError, venda_service.TransicaoInvalidaError) as exc:
        flash(f'Erro: {exc}', 'danger')
    return redirect(url_for('hora.vendas_detalhe', venda_id=venda.id))


# ------------------------------------------------------------------------
# Resolver divergencia (marca como tratada)
# ------------------------------------------------------------------------

@hora_bp.route(
    '/vendas/<int:venda_id>/divergencias/<int:div_id>/resolver',
    methods=['POST'],
)
@require_hora_perm('vendas', 'editar')
def vendas_resolver_divergencia(venda_id: int, div_id: int):
    venda = HoraVenda.query.get_or_404(venda_id)
    is_ajax = request.is_json or request.headers.get('Accept') == 'application/json'

    if venda.loja_id and not usuario_tem_acesso_a_loja(venda.loja_id):
        if is_ajax:
            return jsonify({'ok': False, 'erro': 'acesso negado'}), 403
        flash('Acesso negado.', 'danger')
        return redirect(url_for('hora.vendas_lista'))

    # Venda sem loja (loja_id=NULL) so resolve por admin — mesma regra das
    # outras rotas desta venda (vendas_detalhe/editar/cancelar).
    if not venda.loja_id and lojas_permitidas_ids() is not None:
        if is_ajax:
            return jsonify({'ok': False, 'erro': 'venda sem loja — apenas admin'}), 403
        flash('Venda sem loja definida — apenas admin resolve divergencias.', 'warning')
        return redirect(url_for('hora.vendas_lista'))

    div = HoraVendaDivergencia.query.get_or_404(div_id)
    if div.venda_id != venda.id:
        flash('Divergencia nao pertence a essa venda.', 'danger')
        return redirect(url_for('hora.vendas_detalhe', venda_id=venda.id))

    try:
        venda_service.resolver_divergencia(
            divergencia_id=div_id,
            usuario=_operador_atual(),
        )
        flash('Divergencia marcada como resolvida.', 'success')
    except ValueError as exc:
        flash(f'Erro: {exc}', 'danger')
    return redirect(url_for('hora.vendas_detalhe', venda_id=venda.id))


# ------------------------------------------------------------------------
# Exportar pedidos de venda + motos + NF (Excel, 2 abas)
# ------------------------------------------------------------------------

@hora_bp.route('/vendas/exportar.xlsx')
@require_hora_perm('vendas', 'ver')
def vendas_exportar_xlsx():
    """Exporta pedidos de venda em Excel com 2 abas:

    - **Pedidos**: 1 linha por HoraVenda com header completo (cliente,
      endereço, NF, forma de pagamento, modalidade frete, parcelas, valor).
    - **Motos**: 1 linha por HoraVendaItem (chassi) com modelo, cor,
      motor, ano, preços e referencia da venda/NF.

    Respeita filtros da query string (mesmo padrao da listagem):
      ?status=COTACAO|CONFIRMADO|FATURADO|CANCELADO  (opcional)
      ?since=YYYY-MM-DD                              (data_venda inclusivo)
      ?until=YYYY-MM-DD                              (data_venda inclusivo)
      ?origem=TAGPLUS_API|DANFE|MANUAL|TAGPLUS       (opcional)

    Escopo: usuario nao-admin so exporta vendas das lojas permitidas.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        flash('openpyxl nao disponivel no servidor.', 'danger')
        return redirect(url_for('hora.vendas_lista'))

    from datetime import date as _date

    # ---- Filtros ----
    status_filtro = (request.args.get('status') or '').strip().upper() or None
    if status_filtro and status_filtro not in (
        'COTACAO', 'CONFIRMADO', 'FATURADO', 'CANCELADO',
    ):
        status_filtro = None

    origem_filtro = (request.args.get('origem') or '').strip().upper() or None

    def _parse_data(name: str):
        v = (request.args.get(name) or '').strip()
        if not v:
            return None
        try:
            return _date.fromisoformat(v)
        except ValueError:
            return None

    since = _parse_data('since')
    until = _parse_data('until')

    # ---- Query base com escopo ----
    permitidas = lojas_permitidas_ids()
    query = HoraVenda.query.order_by(
        HoraVenda.data_venda.desc(), HoraVenda.id.desc(),
    )
    if permitidas is not None:
        if not permitidas:
            flash('Sem lojas permitidas — nada a exportar.', 'warning')
            return redirect(url_for('hora.vendas_lista'))
        query = query.filter(HoraVenda.loja_id.in_(permitidas))
    if status_filtro:
        query = query.filter(HoraVenda.status == status_filtro)
    if origem_filtro:
        query = query.filter(HoraVenda.origem_criacao == origem_filtro)
    if since:
        query = query.filter(HoraVenda.data_venda >= since)
    if until:
        query = query.filter(HoraVenda.data_venda <= until)

    vendas = query.all()
    if not vendas:
        flash('Nenhum pedido encontrado para os filtros aplicados.', 'warning')
        return redirect(url_for('hora.vendas_lista'))

    # ---- Helpers ----
    def _dec(v):
        if v is None:
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    def _dt(v):
        if not v:
            return None
        try:
            return v.strftime('%d/%m/%Y %H:%M')
        except AttributeError:
            return str(v)

    def _data(v):
        if not v:
            return None
        try:
            return v.strftime('%d/%m/%Y')
        except AttributeError:
            return str(v)

    # ---- Aba 1: Pedidos ----
    cab_pedidos = [
        'venda_id', 'status', 'origem',
        'nf_numero', 'nf_chave_44', 'nf_emitida_em',
        'loja', 'cnpj_emitente',
        'data_venda',
        'cliente_nome', 'cliente_cpf', 'cliente_telefone', 'cliente_email',
        'cep', 'logradouro', 'numero', 'complemento', 'bairro', 'cidade', 'uf',
        'forma_pagamento', 'numero_parcelas', 'intervalo_parcelas_dias',
        'modalidade_frete',
        'valor_total', 'qtd_itens',
        'vendedor', 'observacoes',
        'qtd_divergencias_abertas',
        'criado_em', 'cancelado_em', 'cancelado_por',
    ]

    linhas_pedidos = []
    for v in vendas:
        loja_lbl = v.loja.rotulo_display if v.loja else None
        linhas_pedidos.append({
            'venda_id': v.id,
            'status': v.status,
            'origem': v.origem_criacao,
            'nf_numero': v.nf_saida_numero,
            'nf_chave_44': v.nf_saida_chave_44,
            'nf_emitida_em': _dt(v.nf_saida_emitida_em),
            'loja': loja_lbl,
            'cnpj_emitente': v.cnpj_emitente,
            'data_venda': _data(v.data_venda),
            'cliente_nome': v.nome_cliente,
            'cliente_cpf': v.cpf_cliente,
            'cliente_telefone': v.telefone_cliente,
            'cliente_email': v.email_cliente,
            'cep': v.cep,
            'logradouro': v.endereco_logradouro,
            'numero': v.endereco_numero,
            'complemento': v.endereco_complemento,
            'bairro': v.endereco_bairro,
            'cidade': v.endereco_cidade,
            'uf': v.endereco_uf,
            'forma_pagamento': v.forma_pagamento,
            'numero_parcelas': v.numero_parcelas,
            'intervalo_parcelas_dias': v.intervalo_parcelas_dias,
            'modalidade_frete': v.modalidade_frete,
            'valor_total': _dec(v.valor_total),
            'qtd_itens': len(v.itens),
            'vendedor': v.vendedor,
            'observacoes': v.observacoes,
            'qtd_divergencias_abertas': len(v.divergencias_abertas),
            'criado_em': _dt(v.criado_em),
            'cancelado_em': _dt(v.cancelado_em),
            'cancelado_por': v.cancelado_por,
        })

    # ---- Aba 2: Motos (1 linha por chassi) ----
    cab_motos = [
        'venda_id', 'nf_numero', 'data_venda', 'status',
        'loja', 'cliente_nome', 'cliente_cpf',
        'numero_chassi', 'modelo', 'cor', 'numero_motor', 'ano_modelo',
        'preco_tabela_referencia', 'desconto_aplicado', 'preco_final',
    ]

    linhas_motos = []
    for v in vendas:
        loja_lbl = v.loja.rotulo_display if v.loja else None
        for it in v.itens:
            moto = it.moto
            modelo_nome = (
                moto.modelo.nome_modelo
                if (moto and moto.modelo) else None
            )
            linhas_motos.append({
                'venda_id': v.id,
                'nf_numero': v.nf_saida_numero,
                'data_venda': _data(v.data_venda),
                'status': v.status,
                'loja': loja_lbl,
                'cliente_nome': v.nome_cliente,
                'cliente_cpf': v.cpf_cliente,
                'numero_chassi': it.numero_chassi,
                'modelo': modelo_nome,
                'cor': moto.cor if moto else None,
                'numero_motor': moto.numero_motor if moto else None,
                'ano_modelo': moto.ano_modelo if moto else None,
                'preco_tabela_referencia': _dec(it.preco_tabela_referencia),
                'desconto_aplicado': _dec(it.desconto_aplicado),
                'preco_final': _dec(it.preco_final),
            })

    # ---- Monta workbook ----
    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')

    def _escrever_aba(ws, titulo: str, cabecalho: list, linhas: list):
        ws.title = titulo[:31]
        for col_idx, campo in enumerate(cabecalho, start=1):
            cell = ws.cell(row=1, column=col_idx, value=campo)
            cell.font = header_font
            cell.fill = header_fill
        for row_idx, item in enumerate(linhas, start=2):
            for col_idx, campo in enumerate(cabecalho, start=1):
                ws.cell(row=row_idx, column=col_idx, value=item.get(campo))
        # Auto-width best-effort.
        for col_idx, campo in enumerate(cabecalho, start=1):
            max_len = max(
                [len(str(item.get(campo) or '')) for item in linhas] + [len(campo)]
            )
            ws.column_dimensions[
                ws.cell(row=1, column=col_idx).column_letter
            ].width = min(max_len + 2, 60)

    ws_pedidos = wb.active
    _escrever_aba(ws_pedidos, 'Pedidos', cab_pedidos, linhas_pedidos)
    ws_motos = wb.create_sheet('Motos')
    _escrever_aba(ws_motos, 'Motos', cab_motos, linhas_motos)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    ts = agora_utc_naive().strftime('%Y%m%d_%H%M%S')
    sufixo = []
    if status_filtro:
        sufixo.append(status_filtro.lower())
    if origem_filtro:
        sufixo.append(origem_filtro.lower())
    if since:
        sufixo.append(f'desde-{since.isoformat()}')
    if until:
        sufixo.append(f'ate-{until.isoformat()}')
    sufixo_str = ('_' + '_'.join(sufixo)) if sufixo else ''
    filename = f'pedidos_venda{sufixo_str}_{ts}.xlsx'

    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


# ========================================================================
# Itens PECA em pedido de venda (XOR moto)
# ========================================================================

@hora_bp.route('/vendas/<int:venda_id>/itens-peca/novo', methods=['POST'])
@require_hora_perm('vendas', 'editar')
def venda_adicionar_item_peca(venda_id: int):
    """Adiciona peca em pedido de venda (apenas em COTACAO)."""
    venda = HoraVenda.query.get_or_404(venda_id)
    from app.hora.services.auth_helper import usuario_tem_acesso_a_loja
    if venda.loja_id and not usuario_tem_acesso_a_loja(venda.loja_id):
        flash('Acesso negado: pedido de loja fora do seu escopo.', 'danger')
        return redirect(url_for('hora.vendas_lista'))

    try:
        peca_id_str = (request.form.get('peca_id') or '').strip()
        if not peca_id_str.isdigit():
            raise ValueError('Selecione uma peca')
        qtd_str = (request.form.get('qtd') or '').strip().replace(',', '.')
        valor_str = (request.form.get('valor_unitario_final') or '').strip().replace(',', '.')
        if not qtd_str or not valor_str:
            raise ValueError('Quantidade e valor unitario obrigatorios')
        from app.hora.services import venda_service
        venda_service.adicionar_item_peca(
            venda_id=venda_id,
            peca_id=int(peca_id_str),
            qtd=Decimal(qtd_str),
            valor_unitario_final=Decimal(valor_str),
            usuario=current_user.nome if hasattr(current_user, 'nome') else None,
        )
        flash('Peca adicionada ao pedido.', 'success')
    except (ValueError, InvalidOperation) as exc:
        flash(f'Erro ao adicionar peca: {exc}', 'danger')

    return redirect(url_for('hora.venda_detalhe', venda_id=venda_id))


@hora_bp.route('/vendas/<int:venda_id>/itens-peca/<int:item_id>/remover', methods=['POST'])
@require_hora_perm('vendas', 'editar')
def venda_remover_item_peca(venda_id: int, item_id: int):
    venda = HoraVenda.query.get_or_404(venda_id)
    from app.hora.services.auth_helper import usuario_tem_acesso_a_loja
    if venda.loja_id and not usuario_tem_acesso_a_loja(venda.loja_id):
        flash('Acesso negado: pedido de loja fora do seu escopo.', 'danger')
        return redirect(url_for('hora.vendas_lista'))
    try:
        from app.hora.services import venda_service
        venda_service.remover_item_peca(
            venda_id=venda_id, item_id=item_id,
            usuario=current_user.nome if hasattr(current_user, 'nome') else None,
        )
        flash('Peca removida do pedido.', 'success')
    except ValueError as exc:
        flash(f'Erro: {exc}', 'danger')
    return redirect(url_for('hora.venda_detalhe', venda_id=venda_id))
