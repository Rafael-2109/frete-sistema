"""Rotas TagPlus: configuracao, OAuth, mapeamentos, webhook e operacoes NFe.

Detalhes: app/hora/EMISSAO_NFE_ENGENHARIA.md secoes 7 (OAuth), 10 (rotas).

Convencoes:
  - `tagplus.*` (admin/ops): protegido por `require_hora_perm('tagplus', acao)`.
  - `nfe/*` (acoes na venda): protegido por `require_hora_perm('vendas', acao)`.
  - `webhook` e callback OAuth: publicos (validacao por secret/state).
"""
from __future__ import annotations

import hmac
import logging
import secrets

from flask import (
    Response,
    abort,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    url_for,
)
from flask_login import current_user

from app import db
from app.hora.decorators import require_hora_perm
from app.hora.models import (
    HoraModelo,
    HoraVenda,
)
from app.hora.models.tagplus import (
    HoraTagPlusConta,
    HoraTagPlusFormaPagamentoMap,
    HoraTagPlusNfeEmissao,
    HoraTagPlusProdutoMap,
    NFE_STATUS_CANCELADA,
    NFE_STATUS_VALIDOS,
)
from app.hora.routes import hora_bp
from app.hora.services.tagplus.api_client import ApiClient
from app.hora.services.tagplus.cancelador_nfe import (
    CanceladorNfe,
    CancelamentoBloqueadoError,
    CancelamentoEmProcessamentoError,
)
from app.hora.services.tagplus.cce_service import CceError, CceService
from app.hora.services.tagplus.crypto import encrypt
from app.hora.services.tagplus.emissor_nfe import (
    EmissaoBloqueadaError,
    EmissorNfeHora,
)
from app.hora.services.tagplus.oauth_client import OAuthClient

logger = logging.getLogger(__name__)


def _operador() -> str:
    return getattr(current_user, 'nome', None) or 'desconhecido'


def _parse_itens_form(form) -> list:
    """Monta a lista de itens (N motos) a partir de chassi[]/valor[] do form.

    Valor em formato BR ('1.000,00') ou US; chassi vazio e ignorado; valor
    invalido vira None (o service valida).

    Args:
        form: werkzeug MultiDict (ou qualquer objeto com .getlist()).

    Returns:
        Lista de dicts {'numero_chassi': str, 'valor_final': Decimal|None}.
    """
    from decimal import Decimal, InvalidOperation

    chassis = form.getlist('chassi')
    valores = form.getlist('valor')
    itens = []
    for i, ch in enumerate(chassis):
        ch = (ch or '').strip()
        if not ch:
            continue
        v_raw = (valores[i] if i < len(valores) else '') or ''
        v_str = v_raw.strip()
        if ',' in v_str:
            v_str = v_str.replace('.', '').replace(',', '.')
        try:
            v = Decimal(v_str) if v_str else None
        except (InvalidOperation, ValueError):
            v = None
        itens.append({'numero_chassi': ch, 'valor_final': v})
    return itens


# ============================================================
# 1) Conta TagPlus — singleton (config + OAuth)
# ============================================================

@hora_bp.route('/tagplus/conta', methods=['GET'])
@require_hora_perm('tagplus', 'ver')
def tagplus_conta():
    """Mostra conta ativa ou form para cadastrar a primeira."""
    conta = HoraTagPlusConta.query.filter_by(ativo=True).first()
    return render_template(
        'hora/tagplus/conta_form.html',
        conta=conta,
    )


@hora_bp.route('/tagplus/conta', methods=['POST'])
@require_hora_perm('tagplus', 'editar')
def tagplus_conta_salvar():
    """Cria/atualiza conta ativa. Encripta client_secret antes de persistir."""
    client_id = (request.form.get('client_id') or '').strip()
    client_secret = (request.form.get('client_secret') or '').strip()
    redirect_uri = (request.form.get('redirect_uri') or '').strip() or None
    ambiente = (request.form.get('ambiente') or 'producao').strip()
    scope = (request.form.get('scope_contratado') or
             'write:nfes read:clientes write:clientes read:produtos').strip()

    if not client_id:
        flash('client_id obrigatorio.', 'danger')
        return redirect(url_for('hora.tagplus_conta'))

    conta = HoraTagPlusConta.query.filter_by(ativo=True).first()
    if not conta:
        if not client_secret:
            flash('client_secret obrigatorio na primeira configuracao.', 'danger')
            return redirect(url_for('hora.tagplus_conta'))
        conta = HoraTagPlusConta(
            client_id=client_id,
            client_secret_encrypted=encrypt(client_secret),
            webhook_secret=secrets.token_urlsafe(32),
            ativo=True,
        )
        db.session.add(conta)
    else:
        conta.client_id = client_id
        if client_secret:
            conta.client_secret_encrypted = encrypt(client_secret)

    conta.redirect_uri = redirect_uri
    conta.ambiente = ambiente if ambiente in ('producao', 'homologacao') else 'producao'
    conta.scope_contratado = scope

    db.session.commit()
    flash('Conta TagPlus salva.', 'success')
    return redirect(url_for('hora.tagplus_conta'))


@hora_bp.route('/tagplus/conta/oauth')
@require_hora_perm('tagplus', 'editar')
def tagplus_oauth_iniciar():
    """Gera state CSRF e redireciona para developers.tagplus.com.br/authorize."""
    conta = HoraTagPlusConta.query.filter_by(ativo=True).first()
    if not conta:
        flash('Cadastre a conta antes de iniciar OAuth.', 'danger')
        return redirect(url_for('hora.tagplus_conta'))

    state = secrets.token_urlsafe(32)
    conta.oauth_state_last = state
    db.session.commit()

    url = OAuthClient(conta).get_authorization_url(state=state)
    return redirect(url)


@hora_bp.route('/tagplus/conta/callback')
def tagplus_oauth_callback():
    """Callback do TagPlus apos autorizacao. NAO requer login (publico)
    porque o portal redireciona direto. Validacao por `state`.
    """
    code = request.args.get('code')
    state = request.args.get('state')
    if not code or not state:
        return render_template(
            'hora/tagplus/oauth_result.html',
            ok=False, mensagem='Faltam parametros code/state.',
        ), 400

    conta = HoraTagPlusConta.query.filter_by(ativo=True).first()
    if not conta:
        return render_template(
            'hora/tagplus/oauth_result.html',
            ok=False, mensagem='Conta ativa nao encontrada.',
        ), 400

    if not conta.oauth_state_last or not hmac.compare_digest(state, conta.oauth_state_last):
        return render_template(
            'hora/tagplus/oauth_result.html',
            ok=False, mensagem='State invalido (possivel CSRF). Refaca o fluxo.',
        ), 400

    try:
        OAuthClient(conta).exchange_code(code)
    except Exception as exc:
        logger.exception('Falha exchange_code TagPlus: %s', exc)
        return render_template(
            'hora/tagplus/oauth_result.html',
            ok=False,
            mensagem='Falha na troca de tokens. Veja os logs do servidor para detalhes.',
        ), 500

    # Sucesso: consome state apenas apos exchange_code dar OK.
    conta.oauth_state_last = None
    db.session.commit()

    return render_template('hora/tagplus/oauth_result.html', ok=True, mensagem='Tokens obtidos com sucesso.')


@hora_bp.route('/tagplus/conta/refresh', methods=['POST'])
@require_hora_perm('tagplus', 'editar')
def tagplus_oauth_refresh():
    conta = HoraTagPlusConta.query.filter_by(ativo=True).first()
    if not conta:
        flash('Conta nao configurada.', 'danger')
        return redirect(url_for('hora.tagplus_conta'))
    try:
        OAuthClient(conta)._do_refresh()  # noqa: SLF001
        flash('Token renovado.', 'success')
    except Exception as exc:
        flash(f'Falha ao renovar: {exc}', 'danger')
    return redirect(url_for('hora.tagplus_conta'))


@hora_bp.route('/tagplus/conta/checklist')
@require_hora_perm('tagplus', 'ver')
def tagplus_checklist():
    """Diagnostico de configuracao. Chama endpoints leves do TagPlus."""
    conta = HoraTagPlusConta.query.filter_by(ativo=True).first()
    checks: list[dict] = []
    api_ok = False

    if not conta:
        checks.append({'item': 'Conta TagPlus ativa', 'status': 'erro', 'mensagem': 'Nao configurada.'})
    else:
        checks.append({
            'item': 'Conta TagPlus ativa', 'status': 'ok',
            'mensagem': f'ambiente={conta.ambiente} scope={conta.scope_contratado}',
        })
        if not conta.token:
            checks.append({'item': 'Token OAuth', 'status': 'erro', 'mensagem': 'Faltando — refazer OAuth.'})
        else:
            checks.append({
                'item': 'Token OAuth', 'status': 'ok',
                'mensagem': f'expira em {conta.token.expires_at}',
            })
            # Scope mismatch: contratado vs efetivo no token vigente.
            scope_efet = (conta.token.scope_efetivo or '').strip()
            scope_contr = (conta.scope_contratado or '').strip()
            if scope_efet and scope_contr and scope_efet != scope_contr:
                checks.append({
                    'item': 'Scope OAuth atualizado',
                    'status': 'erro',
                    'mensagem': (
                        f'scope contratado={scope_contr!r} difere do scope '
                        f'vigente do token={scope_efet!r}. '
                        'Refresh nao basta — refazer Authorization Code Flow '
                        '(Conta > Reautorizar OAuth) para ativar novos scopes '
                        '(ex.: read:pedidos para backfill de pedidos).'
                    ),
                })
            elif scope_efet:
                checks.append({
                    'item': 'Scope OAuth atualizado', 'status': 'ok',
                    'mensagem': f'scope vigente={scope_efet!r} (em sincronia com contratado).',
                })
            else:
                checks.append({
                    'item': 'Scope OAuth atualizado', 'status': 'aviso',
                    'mensagem': (
                        'scope vigente nao registrado (token gerado antes do '
                        'tracking). Refazer OAuth para sincronizar.'
                    ),
                })
            try:
                client = ApiClient(conta)
                # Probe leve dentro do scope `read:produtos` — `/usuario_atual` nao existe
                # na API TagPlus (ver scripts/doc_tagplus.md). `?per_page=1` reduz payload.
                r = client.get('/produtos', params={'per_page': 1})
                if r.status_code == 200:
                    api_ok = True
                    body = r.json() if r.content else []
                    qtd = len(body) if isinstance(body, list) else 'n/a'
                    checks.append({
                        'item': 'GET /produtos (probe)', 'status': 'ok',
                        'mensagem': f'API OK (retornou {qtd} item).',
                    })
                else:
                    checks.append({
                        'item': 'GET /produtos (probe)', 'status': 'erro',
                        'mensagem': f'HTTP {r.status_code}: {r.text[:200]}',
                    })
            except Exception as exc:
                checks.append({'item': 'GET /produtos (probe)', 'status': 'erro', 'mensagem': str(exc)})

    # Mapeamentos — produtos: valida quantidade E formato do ID (deve ser inteiro
    # numerico, nao codigo string como MT-JET — TagPlus rejeita com 400).
    qtd_modelos = HoraModelo.query.filter_by(ativo=True).count()
    todos_maps = HoraTagPlusProdutoMap.query.all()
    qtd_modelos_mapeados = len(todos_maps)
    ids_invalidos = [
        m for m in todos_maps
        if not (m.tagplus_produto_id or '').strip().isdigit()
    ]
    if ids_invalidos:
        nomes = ', '.join(
            f'modelo_id={m.modelo_id}({m.tagplus_produto_id!r})'
            for m in ids_invalidos[:5]
        )
        sufixo = f' [+{len(ids_invalidos) - 5} mais]' if len(ids_invalidos) > 5 else ''
        checks.append({
            'item': 'Mapeamento de produtos',
            'status': 'erro',
            'mensagem': (
                f'{qtd_modelos_mapeados}/{qtd_modelos} mapeados, mas '
                f'{len(ids_invalidos)} com ID nao-numerico (TagPlus rejeita 400). '
                f'Exemplos: {nomes}{sufixo}. Use o botao "Listar do TagPlus" '
                f'para obter os IDs inteiros corretos.'
            ),
        })
    else:
        checks.append({
            'item': 'Mapeamento de produtos',
            'status': 'ok' if qtd_modelos_mapeados >= qtd_modelos and qtd_modelos > 0 else 'aviso',
            'mensagem': f'{qtd_modelos_mapeados}/{qtd_modelos} modelos mapeados (IDs validos).',
        })

    qtd_formas = HoraTagPlusFormaPagamentoMap.query.count()
    checks.append({
        'item': 'Mapeamento de formas de pagamento',
        'status': 'ok' if qtd_formas >= 1 else 'aviso',
        'mensagem': f'{qtd_formas} forma(s) de pagamento mapeada(s).',
    })

    return render_template(
        'hora/tagplus/checklist.html',
        conta=conta, checks=checks, api_ok=api_ok,
    )


# ============================================================
# 2) Mapeamentos
# ============================================================

@hora_bp.route('/tagplus/conta/mapeamento', methods=['GET'])
@require_hora_perm('tagplus', 'ver')
def tagplus_produto_map_lista():
    """Lista mapeamentos modelo HORA -> produto TagPlus."""
    modelos = HoraModelo.query.order_by(HoraModelo.nome_modelo).all()
    maps = {m.modelo_id: m for m in HoraTagPlusProdutoMap.query.all()}
    rows = [{'modelo': m, 'map': maps.get(m.id)} for m in modelos]
    return render_template('hora/tagplus/produto_map.html', rows=rows)


@hora_bp.route('/tagplus/conta/mapeamento/listar-produtos', methods=['GET'])
@require_hora_perm('tagplus', 'ver')
def tagplus_produto_map_listar_api():
    """Proxy GET /produtos no TagPlus com paginacao automatica.

    Por default: itera page=1..N agregando ate uma pagina retornar <100
    (ultima). Cap em 20 paginas (=2000 produtos) para defesa.

    Query params:
      - q: filtro de busca livre (LIKE no TagPlus)
      - page: pagina especifica (apenas se quiser modo single-page; default agrega todas)
    """
    conta = HoraTagPlusConta.query.filter_by(ativo=True).first()
    if not conta:
        return jsonify({'ok': False, 'error': 'Conta nao configurada.'}), 404
    if not conta.token:
        return jsonify({'ok': False, 'error': 'Sem token OAuth — autorizar em /hora/tagplus/conta.'}), 400

    q = (request.args.get('q') or '').strip()
    page_param = request.args.get('page')
    paginas_a_buscar = (
        [max(1, int(page_param))] if page_param else list(range(1, 21))
    )

    client = ApiClient(conta)
    body: list = []
    paginas_lidas = 0
    for p in paginas_a_buscar:
        params = {'per_page': 100, 'page': p}
        if q:
            params['q'] = q
        r = client.get('/produtos', params=params)
        if r.status_code != 200:
            if not body:
                return jsonify({
                    'ok': False,
                    'http_status': r.status_code,
                    'error': r.text[:300],
                }), 502
            # Ja tinhamos paginas anteriores OK: retorna parcial com aviso.
            break
        try:
            chunk = r.json()
        except ValueError:
            chunk = []
        if isinstance(chunk, dict):
            chunk = chunk.get('data') or chunk.get('produtos') or chunk.get('results') or []
        if not isinstance(chunk, list) or not chunk:
            break
        body.extend(chunk)
        paginas_lidas += 1
        if len(chunk) < 100:
            break  # ultima pagina

    # Normaliza: pega so campos uteis para a UI.
    produtos = []
    if isinstance(body, list):
        for p in body:
            if not isinstance(p, dict):
                continue
            produtos.append({
                'id': p.get('id'),
                'codigo': p.get('codigo') or p.get('cod_secundario'),
                'nome': p.get('descricao_produto') or p.get('descricao') or p.get('nome'),
                'ativo': p.get('ativo'),
            })

    return jsonify({
        'ok': True,
        'paginas_lidas': paginas_lidas,
        'modo': 'pagina_unica' if page_param else 'agregado',
        'count': len(produtos),
        'produtos': produtos,
    })


@hora_bp.route('/tagplus/conta/mapeamento', methods=['POST'])
@require_hora_perm('tagplus', 'editar')
def tagplus_produto_map_salvar():
    """Upsert dos mapeamentos a partir do form (1 linha por modelo)."""
    for modelo in HoraModelo.query.all():
        prefix = f'modelo_{modelo.id}_'
        tagplus_id = (request.form.get(prefix + 'tagplus_id') or '').strip()
        codigo = (request.form.get(prefix + 'codigo') or '').strip() or None
        cfop = (request.form.get(prefix + 'cfop') or '5.403').strip() or '5.403'

        existente = HoraTagPlusProdutoMap.query.filter_by(modelo_id=modelo.id).first()

        if not tagplus_id:
            # Limpa se vazio.
            if existente:
                db.session.delete(existente)
            continue

        if len(tagplus_id) > 50:
            flash(f'tagplus_id invalido para modelo {modelo.nome_modelo} (>50 chars): {tagplus_id}', 'danger')
            continue

        # TagPlus exige ID inteiro (numerico). Strings como `MT-JET` sao rejeitadas
        # com HTTP 400 ("IDs nao existem na base de dados"). Bloqueamos no save.
        if not tagplus_id.isdigit():
            flash(
                f'tagplus_id deve ser numerico para {modelo.nome_modelo}: '
                f'{tagplus_id!r}. Use o botao "Listar do TagPlus" para obter o ID inteiro. '
                f'Codigos string vao no campo `tagplus_codigo` (debug).',
                'danger',
            )
            continue

        if existente:
            existente.tagplus_produto_id = tagplus_id
            existente.tagplus_codigo = codigo
            existente.cfop_default = cfop
        else:
            db.session.add(HoraTagPlusProdutoMap(
                modelo_id=modelo.id,
                tagplus_produto_id=tagplus_id,
                tagplus_codigo=codigo,
                cfop_default=cfop,
            ))

    db.session.commit()
    flash('Mapeamentos salvos.', 'success')
    return redirect(url_for('hora.tagplus_produto_map_lista'))


@hora_bp.route('/tagplus/conta/mapeamento/listar-produtos.xlsx', methods=['GET'])
@require_hora_perm('tagplus', 'ver')
def tagplus_produto_map_listar_excel():
    """Exporta lista de produtos do TagPlus em Excel.

    Pagina automaticamente GET /produtos ate esgotar (per_page=100, max=20 paginas).
    Util para o operador conferir IDs/codigos em massa antes de mapear.
    """
    conta = HoraTagPlusConta.query.filter_by(ativo=True).first()
    if not conta:
        flash('Conta TagPlus nao configurada.', 'danger')
        return redirect(url_for('hora.tagplus_produto_map_lista'))
    if not conta.token:
        flash('Sem token OAuth — autorizar em /hora/tagplus/conta.', 'danger')
        return redirect(url_for('hora.tagplus_produto_map_lista'))

    client = ApiClient(conta)
    produtos: list[dict] = []
    for page in range(1, 21):  # cap em 20*100 = 2000 produtos
        r = client.get('/produtos', params={'per_page': 100, 'page': page})
        if r.status_code != 200:
            flash(
                f'GET /produtos page={page} retornou HTTP {r.status_code}. '
                f'Exportacao parcial com {len(produtos)} produto(s).',
                'warning',
            )
            break
        try:
            body = r.json()
        except ValueError:
            break
        if isinstance(body, dict):
            body = body.get('data') or body.get('produtos') or body.get('results') or []
        if not isinstance(body, list) or not body:
            break
        for p in body:
            if isinstance(p, dict):
                produtos.append({
                    'id': p.get('id'),
                    'codigo': p.get('codigo') or p.get('cod_secundario'),
                    'descricao': (
                        p.get('descricao_produto') or p.get('descricao') or p.get('nome')
                    ),
                    'ativo': p.get('ativo'),
                    'preco_venda': p.get('preco_venda') or p.get('valor_venda'),
                })
        if len(body) < 100:
            break  # ultima pagina

    return _gerar_xlsx_resposta(
        nome_base='tagplus_produtos',
        cabecalho=['id', 'codigo', 'descricao', 'ativo', 'preco_venda'],
        linhas=produtos,
        titulo_aba='Produtos TagPlus',
    )


@hora_bp.route('/tagplus/conta/formas-pagamento', methods=['GET'])
@require_hora_perm('tagplus', 'ver')
def tagplus_forma_map_lista():
    """Lista todas as formas mapeadas. Operador adiciona novas via botao
    "Adicionar nova forma" no template (sem sugestoes hardcoded)."""
    todas = HoraTagPlusFormaPagamentoMap.query.order_by(
        HoraTagPlusFormaPagamentoMap.forma_pagamento_hora
    ).all()
    rows = [{'forma': m.forma_pagamento_hora, 'map': m} for m in todas]
    return render_template('hora/tagplus/forma_pag_map.html', rows=rows)


@hora_bp.route('/tagplus/conta/formas-pagamento/listar', methods=['GET'])
@require_hora_perm('tagplus', 'ver')
def tagplus_forma_map_listar_api():
    """Lista formas de pagamento do TagPlus.

    Estrategia em 2 passos:
      1. Tenta endpoints diretos: /formas_pagamento, /formas_pgto, /forma_pagamento.
         (No scope atual `write:nfes read:clientes write:clientes read:produtos`,
         estes retornam 403 — o endpoint existe mas falta scope).

      2. Fallback automatico: extrai IDs unicos de `faturas[].id_forma_pagamento`
         das NFes recentes (scope `write:nfes` cobre o read implicito).
    """
    conta = HoraTagPlusConta.query.filter_by(ativo=True).first()
    if not conta:
        return jsonify({'ok': False, 'error': 'Conta nao configurada.'}), 404
    if not conta.token:
        return jsonify({'ok': False, 'error': 'Sem token OAuth — autorizar em /hora/tagplus/conta.'}), 400

    client = ApiClient(conta)

    # ---- Passo 1: endpoints diretos ----
    candidatos = ['/formas_pagamento', '/formas_pgto', '/forma_pagamento']
    ultima_resposta = None
    for path in candidatos:
        r = client.get(path, params={'per_page': 100})
        ultima_resposta = (path, r.status_code, r.text[:300])
        if r.status_code == 200:
            try:
                body = r.json()
            except ValueError:
                body = []
            if isinstance(body, dict):
                body = body.get('data') or body.get('results') or body.get('formas_pagamento') or []
            return jsonify({
                'ok': True,
                'fonte': 'endpoint_direto',
                'endpoint': path,
                'count': len(body) if isinstance(body, list) else 0,
                'formas': body,
            })

    # ---- Passo 2: fallback via NFes recentes ----
    # Tenta /nfes (scope write:nfes contratado garante read).
    r_nfes = client.get('/nfes', params={'per_page': 50, 'fields': 'id,faturas,status'})
    if r_nfes.status_code == 200:
        try:
            nfes = r_nfes.json() or []
        except ValueError:
            nfes = []
        if isinstance(nfes, dict):
            nfes = nfes.get('data') or nfes.get('nfes') or nfes.get('results') or []

        # Extrai IDs unicos (id -> primeiro objeto encontrado).
        formas_por_id: dict = {}
        for nf in nfes if isinstance(nfes, list) else []:
            faturas = nf.get('faturas') if isinstance(nf, dict) else None
            if not isinstance(faturas, list):
                continue
            for f in faturas:
                if not isinstance(f, dict):
                    continue
                # Schema observado em doc: faturas[].id_forma_pagamento OR faturas[].forma_pagamento (int).
                fid = f.get('id_forma_pagamento') or f.get('forma_pagamento')
                if isinstance(fid, dict):
                    obj_id = fid.get('id')
                    if obj_id is None:
                        continue
                    formas_por_id.setdefault(obj_id, {
                        'id': obj_id,
                        'descricao': fid.get('descricao') or fid.get('nome') or f'forma {obj_id}',
                        'origem_nfe_id': nf.get('id'),
                    })
                elif isinstance(fid, int):
                    formas_por_id.setdefault(fid, {
                        'id': fid,
                        'descricao': f'forma {fid} (sem descricao — extraida de NFe {nf.get("id")})',
                        'origem_nfe_id': nf.get('id'),
                    })

        if formas_por_id:
            return jsonify({
                'ok': True,
                'fonte': 'nfes_existentes',
                'count': len(formas_por_id),
                'formas': sorted(formas_por_id.values(), key=lambda x: x['id']),
                'aviso': (
                    f'Endpoint direto retornou 403 (provavel falta de scope). '
                    f'IDs extraidos de {len(nfes) if isinstance(nfes, list) else 0} NFes '
                    f'recentes. Para descricoes completas, ampliar scope OAuth ou '
                    f'consultar GET /nfes/{{id}} individualmente.'
                ),
            })

    # ---- Falha total ----
    return jsonify({
        'ok': False,
        'error': 'Endpoint direto retornou 403 e nao ha NFes para extrair IDs.',
        'tentativas_diretas': candidatos,
        'ultima_resposta_direta': {
            'endpoint': ultima_resposta[0] if ultima_resposta else None,
            'http_status': ultima_resposta[1] if ultima_resposta else None,
            'body': ultima_resposta[2] if ultima_resposta else None,
        },
        'fallback_nfes': {
            'http_status': r_nfes.status_code,
            'count': 0,
        },
        'instrucoes': (
            'Opcoes: (a) emitir 1 NFe de teste no portal TagPlus com cada forma '
            'utilizada e clicar de novo neste botao para '
            'extrair os IDs automaticamente das NFes; (b) ampliar o scope no '
            'campo "Scopes" em /hora/tagplus/conta (adicionar `read:formas_pagamento` '
            'ou variante como `read:formas_pgto`), salvar e clicar "Re-autorizar OAuth" '
            '— scope e enviado em ?scope= na URL /authorize, nao configurado no portal.'
        ),
    }), 502


@hora_bp.route('/tagplus/conta/formas-pagamento/listar.xlsx', methods=['GET'])
@require_hora_perm('tagplus', 'ver')
def tagplus_forma_map_listar_excel():
    """Exporta formas de pagamento do TagPlus em Excel.

    Reusa estrategia 2-passos do listar (endpoint direto -> fallback NFes).
    """
    conta = HoraTagPlusConta.query.filter_by(ativo=True).first()
    if not conta:
        flash('Conta TagPlus nao configurada.', 'danger')
        return redirect(url_for('hora.tagplus_forma_map_lista'))
    if not conta.token:
        flash('Sem token OAuth — autorizar em /hora/tagplus/conta.', 'danger')
        return redirect(url_for('hora.tagplus_forma_map_lista'))

    client = ApiClient(conta)
    linhas: list[dict] = []
    fonte = 'desconhecida'

    # Passo 1: endpoints diretos
    for path in ('/formas_pagamento', '/formas_pgto', '/forma_pagamento'):
        r = client.get(path, params={'per_page': 100})
        if r.status_code == 200:
            try:
                body = r.json()
            except ValueError:
                body = []
            if isinstance(body, dict):
                body = (
                    body.get('data')
                    or body.get('results')
                    or body.get('formas_pagamento')
                    or []
                )
            if isinstance(body, list):
                for f in body:
                    if isinstance(f, dict):
                        linhas.append({
                            'id': f.get('id') or f.get('codigo'),
                            'descricao': f.get('descricao') or f.get('nome') or f.get('tipo'),
                            'origem_nfe_id': '',
                        })
                fonte = f'endpoint_direto: {path}'
                break

    # Passo 2: fallback via NFes
    if not linhas:
        r_nfes = client.get('/nfes', params={'per_page': 50, 'fields': 'id,faturas,status'})
        if r_nfes.status_code == 200:
            try:
                nfes = r_nfes.json() or []
            except ValueError:
                nfes = []
            if isinstance(nfes, dict):
                nfes = nfes.get('data') or nfes.get('nfes') or nfes.get('results') or []
            formas_por_id: dict = {}
            for nf in nfes if isinstance(nfes, list) else []:
                faturas = nf.get('faturas') if isinstance(nf, dict) else None
                if not isinstance(faturas, list):
                    continue
                for f in faturas:
                    if not isinstance(f, dict):
                        continue
                    fid = f.get('id_forma_pagamento') or f.get('forma_pagamento')
                    if isinstance(fid, dict):
                        obj_id = fid.get('id')
                        if obj_id is None:
                            continue
                        formas_por_id.setdefault(obj_id, {
                            'id': obj_id,
                            'descricao': fid.get('descricao') or fid.get('nome') or '',
                            'origem_nfe_id': nf.get('id'),
                        })
                    elif isinstance(fid, int):
                        formas_por_id.setdefault(fid, {
                            'id': fid,
                            'descricao': '(extraida de NFe)',
                            'origem_nfe_id': nf.get('id'),
                        })
            linhas = sorted(formas_por_id.values(), key=lambda x: x.get('id') or 0)
            fonte = 'nfes_existentes'

    if not linhas:
        flash('Nao foi possivel obter formas (endpoint direto e fallback NFes falharam).', 'warning')
        return redirect(url_for('hora.tagplus_forma_map_lista'))

    return _gerar_xlsx_resposta(
        nome_base='tagplus_formas_pagamento',
        cabecalho=['id', 'descricao', 'origem_nfe_id'],
        linhas=linhas,
        titulo_aba=f'Formas ({fonte})',
    )


def _gerar_xlsx_resposta(
    nome_base: str,
    cabecalho: list[str],
    linhas: list[dict],
    titulo_aba: str,
):
    """Helper: monta XLSX em memoria e devolve Response com download.

    `linhas` e lista de dicts onde cada chave deve estar em `cabecalho`.
    """
    from io import BytesIO
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        flash('openpyxl nao instalado no servidor.', 'danger')
        return redirect(url_for('hora.tagplus_checklist'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (titulo_aba or 'Dados')[:31]

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    for col_idx, campo in enumerate(cabecalho, start=1):
        cell = ws.cell(row=1, column=col_idx, value=campo)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, item in enumerate(linhas, start=2):
        for col_idx, campo in enumerate(cabecalho, start=1):
            ws.cell(row=row_idx, column=col_idx, value=item.get(campo))

    # Auto-width best-effort
    for col_idx, campo in enumerate(cabecalho, start=1):
        max_len = max(
            [len(str(item.get(campo) or '')) for item in linhas] + [len(campo)]
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from app.utils.timezone import agora_utc_naive
    ts = agora_utc_naive().strftime('%Y%m%d_%H%M%S')
    filename = f'{nome_base}_{ts}.xlsx'
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@hora_bp.route('/tagplus/conta/formas-pagamento', methods=['POST'])
@require_hora_perm('tagplus', 'editar')
def tagplus_forma_map_salvar():
    """Processa dinamicamente todas as linhas no form (form-array indexed).

    Form fields esperados (idx = 0..N):
      forma[idx]            -> nome HORA (ex: PIX, CARTAO_DEBITO, etc.)
      tagplus_id[idx]       -> ID inteiro no TagPlus (vazio = remover)
      descricao[idx]        -> texto livre (opcional)
    """
    from app.hora.models.tagplus import TIPOS_PAGAMENTO_VALIDOS

    formas_form = request.form.getlist('forma[]')
    ids_form = request.form.getlist('tagplus_id[]')
    descs_form = request.form.getlist('descricao[]')
    tipos_form = request.form.getlist('tipo_pagamento[]')
    aut_form = request.form.getlist('exige_aut_id[]')

    # Garante alinhamento de listas (linhas adicionadas por JS podem nao ter
    # o tipo_pagamento se template antigo cachado). Pad com None.
    def _at(lst, idx):
        return lst[idx] if idx < len(lst) else ''

    nomes_processados = set()
    for idx, forma_raw in enumerate(formas_form):
        id_raw = _at(ids_form, idx)
        desc_raw = _at(descs_form, idx)
        tipo_raw = _at(tipos_form, idx)
        forma = (forma_raw or '').strip().upper()[:20]
        tagplus_id = (id_raw or '').strip()
        descricao = (desc_raw or '').strip() or None
        tipo_pgto = (tipo_raw or '').strip().upper() or None
        # exige_aut_id chega como '1' (checkbox marcado) ou ausente. Como
        # form-array nao envia o checkbox quando desmarcado, usar lista
        # paralela de "indices marcados" via name=exige_aut_id[] value=idx.
        # Aqui o getlist retorna a lista dos indices marcados.
        exige_aut = str(idx) in aut_form
        if tipo_pgto and tipo_pgto not in TIPOS_PAGAMENTO_VALIDOS:
            flash(
                f'tipo_pagamento invalido para {forma}: {tipo_pgto!r} '
                f'(esperado A_VISTA ou A_PRAZO).', 'danger',
            )
            continue

        if not forma:
            continue
        if forma in nomes_processados:
            flash(f'Forma duplicada no form: {forma}', 'warning')
            continue
        nomes_processados.add(forma)

        existente = HoraTagPlusFormaPagamentoMap.query.filter_by(
            forma_pagamento_hora=forma,
        ).first()

        if not tagplus_id:
            # Vazio = remover (se existir).
            if existente:
                db.session.delete(existente)
            continue

        try:
            tagplus_id_int = int(tagplus_id)
        except ValueError:
            flash(f'tagplus_id invalido para {forma}: {tagplus_id}', 'danger')
            continue

        if existente:
            existente.tagplus_forma_id = tagplus_id_int
            existente.descricao = descricao
            existente.tipo_pagamento = tipo_pgto
            existente.exige_aut_id = exige_aut
        else:
            db.session.add(HoraTagPlusFormaPagamentoMap(
                forma_pagamento_hora=forma,
                tagplus_forma_id=tagplus_id_int,
                descricao=descricao,
                tipo_pagamento=tipo_pgto,
                exige_aut_id=exige_aut,
            ))
    db.session.commit()
    flash(f'Mapeamentos salvos ({len(nomes_processados)} formas processadas).', 'success')
    return redirect(url_for('hora.tagplus_forma_map_lista'))


# ============================================================
# 3) Webhook publico
# ============================================================

@hora_bp.route('/tagplus/webhook', methods=['POST'])
def tagplus_webhook():
    """Receiver publico. Valida X-Hub-Secret e enfileira processamento.

    Responde 200 rapido para nao causar timeout/retry no TagPlus. O processamento
    real (GET /nfes/{id} + atualizar HoraVenda + emitir HoraMotoEvento) roda no
    worker RQ.
    """
    conta = HoraTagPlusConta.query.filter_by(ativo=True).first()
    if not conta:
        return jsonify({'ok': False, 'error': 'sem conta ativa'}), 503

    secret_recebido = request.headers.get('X-Hub-Secret', '')
    if not conta.webhook_secret or not hmac.compare_digest(
        secret_recebido, conta.webhook_secret,
    ):
        logger.warning('Webhook TagPlus com secret invalido (origem=%s)', request.remote_addr)
        return jsonify({'ok': False, 'error': 'secret invalido'}), 401

    try:
        body = request.get_json(force=True, silent=True) or {}
    except Exception:
        body = {}

    event_type = body.get('event_type')
    data = body.get('data', [])
    if not event_type or not isinstance(data, list):
        return jsonify({'ok': False, 'error': 'body invalido'}), 400

    # Enfileira processamento (rapido, libera o TagPlus).
    try:
        from rq import Queue
        from redis import Redis
        import os
        redis_url = os.environ.get('REDIS_URL')
        if redis_url:
            redis_conn = Redis.from_url(redis_url)
            queue = Queue(EmissorNfeHora.QUEUE_NAME, connection=redis_conn)
            queue.enqueue(
                'app.hora.workers.emissao_nfe_worker.processar_webhook',
                conta.id, event_type, data,
            )
        else:
            # Fallback sincrono (dev sem Redis).
            from app.hora.services.tagplus.webhook_handler import WebhookHandler
            WebhookHandler.processar(conta.id, event_type, data)
    except Exception as exc:
        logger.exception('Falha enfileirando webhook: %s', exc)
        # Mesmo assim respondemos 200 — reconciliacao recupera depois.

    return jsonify({'ok': True}), 200


# ============================================================
# 4) Fila de emissoes (gestao)
# ============================================================

@hora_bp.route('/tagplus/emissoes')
@require_hora_perm('tagplus', 'ver')
def tagplus_emissoes_lista():
    from datetime import datetime as _dt
    from sqlalchemy import or_

    status_filtro = (request.args.get('status') or '').strip().upper() or None
    if status_filtro and status_filtro not in NFE_STATUS_VALIDOS:
        status_filtro = None

    try:
        page = int(request.args.get('page', 1))
    except (TypeError, ValueError):
        page = 1
    try:
        per_page = int(request.args.get('per_page', 50))
    except (TypeError, ValueError):
        per_page = 50
    page = max(1, page)
    per_page = max(1, min(per_page, 200))

    busca = (request.args.get('busca') or '').strip() or None
    data_ini_str = (request.args.get('data_inicio') or '').strip()
    data_fim_str = (request.args.get('data_fim') or '').strip()
    try:
        data_inicio = _dt.strptime(data_ini_str, '%Y-%m-%d') if data_ini_str else None
        data_fim = _dt.strptime(data_fim_str, '%Y-%m-%d') if data_fim_str else None
    except ValueError:
        flash('Data invalida (use formato YYYY-MM-DD).', 'warning')
        data_inicio = None
        data_fim = None

    q = HoraTagPlusNfeEmissao.query.order_by(HoraTagPlusNfeEmissao.criado_em.desc())
    if status_filtro:
        q = q.filter_by(status=status_filtro)

    if busca:
        # Busca em numero_nfe, chave_44 (NF) e nome_cliente da venda relacionada.
        b = busca
        q = (
            q.outerjoin(HoraVenda, HoraTagPlusNfeEmissao.venda_id == HoraVenda.id)
            .filter(or_(
                HoraTagPlusNfeEmissao.numero_nfe.ilike(f'%{b}%'),
                HoraTagPlusNfeEmissao.chave_44.ilike(f'%{b}%'),
                HoraVenda.nome_cliente.ilike(f'%{b}%'),
                HoraVenda.cpf_cliente.ilike(f'%{b}%'),
            ))
        )

    if data_inicio:
        q = q.filter(HoraTagPlusNfeEmissao.criado_em >= data_inicio)
    if data_fim:
        # data_fim inclusivo: pegar tudo ate 23:59 do dia
        from datetime import timedelta
        q = q.filter(HoraTagPlusNfeEmissao.criado_em <= data_fim + timedelta(days=1))

    pagination = q.paginate(page=page, per_page=per_page, error_out=False)
    return render_template(
        'hora/tagplus/emissoes_lista.html',
        emissoes=pagination.items,
        pagination=pagination,
        per_page=per_page,
        status_filtro=status_filtro,
        status_validos=NFE_STATUS_VALIDOS,
        filtro_busca=busca,
        filtro_data_inicio=data_ini_str,
        filtro_data_fim=data_fim_str,
    )



# ============================================================
# 4.6) Pedido de Venda — formulario manual para emissao TagPlus
# ============================================================

@hora_bp.route('/tagplus/pedido-venda/novo', methods=['GET'])
@require_hora_perm('vendas', 'criar')
def tagplus_pedido_venda_novo():
    """Formulario para criar HoraVenda manualmente (sem upload de DANFE).

    Operador preenche cliente + endereco + escolhe modelo/cor/chassi e forma de
    pagamento. Submissao cria venda em status COTACAO + redireciona para o
    detalhe da venda; a emissao da NFe e feita la pelo botao "Emitir NFe".

    Decisao de negocio (2026-05-06): qualquer operador HORA-habilitado pode
    criar pedido em qualquer loja com chassi de qualquer loja — listas globais
    (sem filtro de escopo). A loja escolhida no SELECT vira a loja oficial da
    venda.
    """
    # Decisao de negocio (2026-05-06): qualquer operador HORA-habilitado pode
    # criar pedido em qualquer loja, com chassi de qualquer loja — listas
    # globais (sem filtro de escopo). Helper compartilhado com vendas_detalhe.
    from app.hora.routes.vendas import _contexto_lookup_pedido_venda
    ctx = _contexto_lookup_pedido_venda()

    return render_template(
        'hora/tagplus/pedido_venda_novo.html',
        **ctx,
    )


@hora_bp.route('/tagplus/pedido-venda', methods=['POST'])
@require_hora_perm('vendas', 'criar')
def tagplus_pedido_venda_criar():
    """Cria HoraVenda manual e redireciona para o detalhe da venda."""
    from decimal import Decimal, InvalidOperation
    from app.hora.services import venda_service

    def _g(name: str, max_len: int = 255) -> str:
        return (request.form.get(name) or '').strip()[:max_len]

    # Itens: N motos via chassi[]/valor[] (FU-3). _parse_itens_form ignora
    # chassis vazios e deixa valor_final=None quando invalido (service valida).
    itens = _parse_itens_form(request.form)

    # Frete: modalidade restrita a 0 (CIF) ou 1 (FOB) por decisao do dono
    # (2026-05-07). Outros valores (2, 3, 4, 9) ainda podem entrar via DANFE
    # PDF / TagPlus, mas nao por este formulario.
    mod_frete = _g('modalidade_frete', 1) or '0'
    if mod_frete not in ('0', '1'):
        flash(f'Modalidade de frete invalida: {mod_frete!r} (esperado 0 ou 1).', 'danger')
        return redirect(url_for('hora.tagplus_pedido_venda_novo'))

    # Frete CIF (hora_38): tipo_frete_calc + valor_frete sao opcionais e
    # so fazem sentido quando modalidade='0'. Service revalida e zera se
    # combinacao for incoerente. Aqui apenas extrai brutos do form.
    valor_frete_raw = _g('valor_frete', 30) or None
    tipo_frete_calc_raw = (_g('tipo_frete_calc', 10) or '').upper() or None
    if mod_frete != '0':
        # Modalidade FOB descarta dados de frete inadvertidamente enviados.
        valor_frete_raw = None
        tipo_frete_calc_raw = None

    # Parcelamento legacy (cache): pega numero_parcelas da MAIOR forma a prazo,
    # senao 1. Service agora persiste parcelas por forma em hora_venda_pagamento.
    intervalo = 30
    n_parcelas = 1  # cache HoraVenda.numero_parcelas (legacy); real fica em pagamentos.

    # Pagamentos multi-formas (migration hora_34): listas paralelas vindas do
    # form (cada index = uma forma). Listas vazias -> pedido nasce INCOMPLETO.
    formas_lista = request.form.getlist('pagamento_forma')
    valores_lista = request.form.getlist('pagamento_valor')
    parcelas_lista = request.form.getlist('pagamento_parcelas')
    aut_ids_lista = request.form.getlist('pagamento_aut_id')

    pagamentos_in: list[dict] = []
    for i, forma_raw in enumerate(formas_lista):
        forma = (forma_raw or '').strip().upper()
        if not forma:
            continue
        valor_raw_p = valores_lista[i] if i < len(valores_lista) else '0'
        # Normaliza formato BR (1.234,56) ou ingles (1234.56).
        valor_str = (valor_raw_p or '').strip()
        if ',' in valor_str:
            valor_str = valor_str.replace('.', '').replace(',', '.')
        try:
            valor_p = Decimal(valor_str) if valor_str else Decimal('0')
        except (InvalidOperation, ValueError):
            continue
        try:
            par_p = int((parcelas_lista[i] if i < len(parcelas_lista) else '1') or '1')
        except ValueError:
            par_p = 1
        aut_p = (aut_ids_lista[i] if i < len(aut_ids_lista) else '').strip() or None
        pagamentos_in.append({
            'forma_pagamento_hora': forma,
            'valor': valor_p,
            'numero_parcelas': par_p,
            'aut_id': aut_p,
        })
    # Cache legacy: usa parcelas da PRIMEIRA forma como n_parcelas para o header.
    if pagamentos_in:
        n_parcelas = pagamentos_in[0]['numero_parcelas'] or 1

    # Vendedor: SELECT no form, default = usuario logado. Server valida que o
    # nome enviado pertence a um usuario HORA-habilitado (defesa em
    # profundidade contra manipulacao do POST). Se vazio, fallback p/ logado.
    from app.hora.services import permissao_service
    vendedor_raw = _g('vendedor', 100)
    if not vendedor_raw:
        vendedor_final = _operador()
    else:
        nomes_validos = {
            u.nome for u in permissao_service.listar_usuarios_habilitados()
        }
        if vendedor_raw in nomes_validos:
            vendedor_final = vendedor_raw
        else:
            flash(
                f'Vendedor invalido: {vendedor_raw!r} nao esta habilitado no modulo HORA.',
                'danger',
            )
            return redirect(url_for('hora.tagplus_pedido_venda_novo'))

    # Loja: SELECT obrigatorio no form. Lista global (sem escopo); valida
    # apenas que a loja existe, esta ativa e nao e a CNPJ matriz excluida.
    # A loja escolhida vira a loja oficial da venda — pode diferir da loja
    # fisica do chassi (transferencia implicita). Decisao do usuario em
    # 2026-05-06.
    from app.hora.services import cadastro_service
    loja_id_raw = _g('loja_id', 20)
    if not loja_id_raw or not loja_id_raw.isdigit():
        flash('Loja obrigatoria — selecione a loja do pedido.', 'danger')
        return redirect(url_for('hora.tagplus_pedido_venda_novo'))
    loja_id_int = int(loja_id_raw)
    lojas_validas = cadastro_service.listar_lojas_para_pedido_venda(
        lojas_permitidas_ids=None,
    )
    if loja_id_int not in {l.id for l in lojas_validas}:
        flash('Loja invalida (inativa ou nao permitida).', 'danger')
        return redirect(url_for('hora.tagplus_pedido_venda_novo'))

    # consumidor_final: campo removido da UI em 2026-05-07. Payload TagPlus
    # agora envia sempre True (ver payload_builder.py). Nao lemos mais do form.

    try:
        venda = venda_service.criar_venda_manual(
            # 18 chars acomoda mascara de CNPJ "00.000.000/0000-00".
            # Service normaliza para apenas digitos antes de gravar.
            cpf_cliente=_g('cpf', 18),
            nome_cliente=_g('nome', 200),
            cep=_g('cep', 9),
            endereco_logradouro=_g('logradouro', 255),
            endereco_numero=_g('numero_endereco', 20),
            endereco_complemento=_g('complemento', 100),
            endereco_bairro=_g('bairro', 100),
            endereco_cidade=_g('cidade', 100),
            endereco_uf=_g('uf', 2),
            itens=itens,
            forma_pagamento=None,  # Legacy fallback: nao mais usado se ha pagamentos.
            telefone_cliente=_g('telefone', 20) or None,
            email_cliente=_g('email', 120) or None,
            vendedor=vendedor_final,
            observacoes=_g('observacoes', 500) or None,
            modalidade_frete=mod_frete,
            numero_parcelas=n_parcelas,
            intervalo_parcelas_dias=intervalo,
            criado_por=_operador(),
            criado_por_id=getattr(current_user, 'id', None),
            loja_id_override=loja_id_int,
            pagamentos=pagamentos_in,
            valor_frete=valor_frete_raw,
            tipo_frete_calc=tipo_frete_calc_raw,
        )
    except ValueError as exc:
        flash(f'Erro ao criar pedido de venda: {exc}', 'danger')
        return redirect(url_for('hora.tagplus_pedido_venda_novo'))

    if venda.status == 'INCOMPLETO':
        flash(
            f'Pedido #{venda.id} salvo como INCOMPLETO — complete as formas de '
            f'pagamento antes de faturar.',
            'warning',
        )
    else:
        flash(
            f'Pedido de venda #{venda.id} criado para {venda.nome_cliente}. '
            f'Confirme o pedido e emita a NFe.',
            'success',
        )
    return redirect(url_for('hora.vendas_detalhe', venda_id=venda.id))


@hora_bp.route('/tagplus/pedido-venda/api/cores')
@require_hora_perm('vendas', 'criar')
def tagplus_pedido_venda_api_cores():
    """JSON com cores que tem chassi disponivel para o modelo informado."""
    from app.hora.services.estoque_service import cores_disponiveis_por_modelo

    try:
        modelo_id = int(request.args.get('modelo_id', '0'))
    except ValueError:
        return jsonify({'ok': False, 'error': 'modelo_id invalido'}), 400
    if not modelo_id:
        return jsonify({'ok': True, 'cores': []})

    # Sem filtro de escopo + incluir aliases: operador escolheu canonico no
    # SELECT; cores devem cobrir motos vinculadas tambem aos aliases.
    cores = cores_disponiveis_por_modelo(
        modelo_id=modelo_id,
        lojas_permitidas_ids=None,
        incluir_aliases=True,
    )
    return jsonify({'ok': True, 'cores': cores})


@hora_bp.route('/tagplus/pedido-venda/api/chassis')
@require_hora_perm('vendas', 'criar')
def tagplus_pedido_venda_api_chassis():
    """JSON com chassis disponiveis filtrados por modelo + cor."""
    from app.hora.services.estoque_service import chassis_disponiveis_para_venda

    try:
        modelo_id = int(request.args.get('modelo_id', '0'))
    except ValueError:
        return jsonify({'ok': False, 'error': 'modelo_id invalido'}), 400
    cor = (request.args.get('cor') or '').strip().upper() or None
    if not modelo_id:
        return jsonify({'ok': True, 'chassis': []})

    # Sem filtro de escopo + incluir aliases: operador escolheu canonico no
    # SELECT; chassis devem cobrir motos vinculadas tambem aos aliases.
    chassis = chassis_disponiveis_para_venda(
        modelo_id=modelo_id,
        cor=cor,
        lojas_permitidas_ids=None,
        incluir_aliases=True,
    )
    return jsonify({'ok': True, 'chassis': chassis})


@hora_bp.route('/tagplus/pedido-venda/api/preco-modelo')
@require_hora_perm('vendas', 'criar')
def tagplus_pedido_venda_api_preco_modelo():
    """JSON: preço de tabela do modelo + classificacao da forma de pagamento.

    Query string:
      modelo_id        — obrigatorio (int)
      forma_pagamento  — opcional (string da HoraVenda; ex: PIX, CARTAO_CREDITO)

    Retorna:
      {
        ok: True,
        preco: float | null,             # preço de tabela vigente (a vista vs a prazo conforme forma)
        fonte: 'MODELO_A_VISTA'|'MODELO_A_PRAZO'|'TABELA_LEGADA'|'AUSENTE',
        tipo_pagamento: 'A_VISTA'|'A_PRAZO'|null,
        preco_a_vista: float | null,     # campo direto do modelo (informativo)
        preco_a_prazo: float | null,
      }
    """
    from app.hora.services.venda_service import buscar_preco_para_pedido

    try:
        modelo_id = int(request.args.get('modelo_id', '0'))
    except ValueError:
        return jsonify({'ok': False, 'error': 'modelo_id invalido'}), 400
    if not modelo_id:
        return jsonify({'ok': False, 'error': 'modelo_id obrigatorio'}), 400

    forma_pgto = (request.args.get('forma_pagamento') or '').strip().upper() or None

    info = buscar_preco_para_pedido(
        modelo_id=modelo_id,
        forma_pagamento_hora=forma_pgto,
    )

    def _f(v):
        return float(v) if v is not None else None

    return jsonify({
        'ok': True,
        'preco': _f(info['preco']),
        'fonte': info['fonte'],
        'tipo_pagamento': info['tipo_pagamento'],
        'preco_a_vista': _f(info['preco_a_vista']),
        'preco_a_prazo': _f(info['preco_a_prazo']),
    })


# ============================================================
# 4.7) Historico de notificacoes WhatsApp + reenvio manual
# ============================================================

@hora_bp.route('/tagplus/notificacoes', methods=['GET'])
@require_hora_perm('tagplus', 'ver')
def tagplus_notificacoes_lista():
    """Historico das notificacoes WhatsApp (NFe aprovada / pedido confirmado)."""
    from app.hora.models.tagplus import HoraTagPlusNotificacaoWhatsapp
    page = request.args.get('page', 1, type=int)
    pag = (HoraTagPlusNotificacaoWhatsapp.query
           .order_by(HoraTagPlusNotificacaoWhatsapp.criado_em.desc())
           .paginate(page=page, per_page=50, error_out=False))
    return render_template('hora/tagplus/notificacoes.html', pag=pag)


@hora_bp.route('/tagplus/notificacoes/<int:reg_id>/reenviar', methods=['POST'])
@require_hora_perm('tagplus', 'editar')
def tagplus_notificacao_reenviar(reg_id):
    """Reenfileira uma notificacao WhatsApp para reprocessamento."""
    from app.hora.models.tagplus import HoraTagPlusNotificacaoWhatsapp
    from app.hora.services.tagplus.notificacao_whatsapp import reenfileirar
    reg = HoraTagPlusNotificacaoWhatsapp.query.get_or_404(reg_id)
    reenfileirar(reg.id)
    flash(f'Reenvio disparado para {reg.tipo} #{reg.ref_id}.', 'info')
    return redirect(url_for('hora.tagplus_notificacoes_lista'))


# ============================================================
# 5) Acoes na venda (NFe)
# ============================================================

@hora_bp.route('/vendas/<int:venda_id>/nfe')
@require_hora_perm('vendas', 'ver')
def venda_nfe_status(venda_id: int):
    from app.utils.timezone import agora_utc_naive
    venda = HoraVenda.query.get_or_404(venda_id)
    emissao = HoraTagPlusNfeEmissao.query.filter_by(venda_id=venda_id).first()
    return render_template(
        'hora/tagplus/nfe_status.html',
        venda=venda, emissao=emissao,
        now_utc=agora_utc_naive(),
    )


@hora_bp.route('/vendas/<int:venda_id>/nfe/preview', methods=['GET'])
@require_hora_perm('vendas', 'criar')
def venda_nfe_preview(venda_id: int):
    """Tela read-only com totais, custos e margem antes de emitir a NFe.

    Mostra:
      - Itens da venda (chassi, modelo, cor, preco_venda, custo_moto)
      - Totais: Venda, Frete, Custo Moto, Liquido, Margem Bruta, % Margem
      - Botao "Confirmar emissao" que POSTa para venda_nfe_emitir.

    O custo da moto vem do `preco_compra_esperado` no pedido de compra
    (HoraPedidoItem) buscado pelo chassi. Quando o chassi nao tem pedido
    associado, a linha aparece como "—" e o flag `tem_custo_faltante`
    sinaliza que a margem fica distorcida.
    """
    from app.hora.services.venda_preview_service import montar_preview

    venda = HoraVenda.query.get_or_404(venda_id)
    preview = montar_preview(venda)
    return render_template(
        'hora/venda_preview_nfe.html',
        venda=venda,
        preview=preview,
    )


@hora_bp.route('/vendas/<int:venda_id>/nfe/emitir', methods=['POST'])
@require_hora_perm('vendas', 'criar')
def venda_nfe_emitir(venda_id: int):
    try:
        emissao_id = EmissorNfeHora.enfileirar(venda_id)
        flash(f'Emissao enfileirada (id={emissao_id}). Acompanhe o status.', 'success')
    except EmissaoBloqueadaError as exc:
        flash(f'Emissao bloqueada: {exc}', 'warning')
    except RuntimeError as exc:
        flash(f'Erro: {exc}', 'danger')
    return redirect(url_for('hora.venda_nfe_status', venda_id=venda_id))


@hora_bp.route('/vendas/<int:venda_id>/nfe/cancelar', methods=['POST'])
@require_hora_perm('vendas', 'apagar')
def venda_nfe_cancelar(venda_id: int):
    justificativa = (request.form.get('justificativa') or '').strip()
    emissao = HoraTagPlusNfeEmissao.query.filter_by(venda_id=venda_id).first()
    if not emissao:
        flash('Sem emissao para esta venda.', 'warning')
        return redirect(url_for('hora.venda_nfe_status', venda_id=venda_id))
    try:
        CanceladorNfe.cancelar(emissao.id, justificativa, _operador())
        flash('Cancelamento solicitado. Aguardando confirmacao SEFAZ.', 'success')
    except ValueError as exc:
        flash(str(exc), 'danger')
    except CancelamentoEmProcessamentoError as exc:
        # NFe ainda em processamento na SEFAZ — orientar usuario a aguardar.
        flash(str(exc), 'info')
    except CancelamentoBloqueadoError as exc:
        flash(f'Bloqueado: {exc}', 'warning')
    except Exception as exc:
        flash(f'Erro: {exc}', 'danger')
    return redirect(url_for('hora.venda_nfe_status', venda_id=venda_id))


@hora_bp.route('/vendas/<int:venda_id>/nfe/sincronizar', methods=['POST'])
@require_hora_perm('vendas', 'ver')
def venda_nfe_sincronizar(venda_id: int):
    """Forca reconciliacao da emissao com o TagPlus (puxa status atual).

    Util quando webhook nfe_aprovada/cancelada nao chegou no sistema mas o
    portal TagPlus mostra a NFe ja resolvida. Faz GET /nfes/{id} e aplica
    o handler do evento correspondente.
    """
    from app.hora.workers.reconciliacao_worker import reconciliar_uma_emissao

    emissao = HoraTagPlusNfeEmissao.query.filter_by(venda_id=venda_id).first()
    if not emissao:
        flash('Sem emissao para esta venda.', 'warning')
        return redirect(url_for('hora.venda_nfe_status', venda_id=venda_id))

    resultado = reconciliar_uma_emissao(emissao.id)
    if resultado.get('ok'):
        if resultado.get('acao_aplicada'):
            flash(resultado['mensagem'], 'success')
        else:
            flash(resultado['mensagem'], 'info')
    else:
        # 5xx do TagPlus = warning (transitorio), nao erro fatal.
        status_http = resultado.get('status_http') or 0
        cat = 'warning' if 500 <= status_http < 600 else 'danger'
        flash(resultado.get('mensagem') or 'Falha ao sincronizar.', cat)
    return redirect(url_for('hora.venda_nfe_status', venda_id=venda_id))


@hora_bp.route('/vendas/<int:venda_id>/nfe/cce', methods=['POST'])
@require_hora_perm('vendas', 'editar')
def venda_nfe_cce(venda_id: int):
    texto = (request.form.get('texto_correcao') or '').strip()
    emissao = HoraTagPlusNfeEmissao.query.filter_by(venda_id=venda_id).first()
    if not emissao:
        flash('Sem emissao para esta venda.', 'warning')
        return redirect(url_for('hora.venda_nfe_status', venda_id=venda_id))
    try:
        CceService.gerar(emissao.id, texto)
        flash('CC-e emitida.', 'success')
    except ValueError as exc:
        flash(str(exc), 'danger')
    except CceError as exc:
        flash(f'Erro: {exc}', 'danger')
    return redirect(url_for('hora.venda_nfe_status', venda_id=venda_id))


# ============================================================
# 6) Proxy DANFE / XML
# ============================================================

@hora_bp.route('/vendas/<int:venda_id>/nfe/danfe.pdf')
@require_hora_perm('vendas', 'ver')
def venda_nfe_danfe_pdf(venda_id: int):
    """Proxy on-the-fly do DANFE PDF via TagPlus (`/nfes/pdf/recibo_a4/{id}`)."""
    emissao = HoraTagPlusNfeEmissao.query.filter_by(venda_id=venda_id).first()
    if not emissao or not emissao.tagplus_nfe_id:
        abort(404)
    client = ApiClient(emissao.conta)
    r = client.get(f'/nfes/pdf/recibo_a4/{emissao.tagplus_nfe_id}')
    if r.status_code != 200:
        return jsonify({
            'ok': False, 'http_status': r.status_code,
            'detalhe': r.text[:500],
        }), 502
    filename = f'danfe_{emissao.numero_nfe or emissao.tagplus_nfe_id}.pdf'
    return Response(
        r.content,
        mimetype='application/pdf',
        headers={'Content-Disposition': f'inline; filename="{filename}"'},
    )


@hora_bp.route('/vendas/<int:venda_id>/nfe/xml')
@require_hora_perm('vendas', 'ver')
def venda_nfe_xml(venda_id: int):
    """Proxy do XML — primeiro tenta gerar_link_xml, senao baixa direto."""
    emissao = HoraTagPlusNfeEmissao.query.filter_by(venda_id=venda_id).first()
    if not emissao or not emissao.tagplus_nfe_id:
        abort(404)
    client = ApiClient(emissao.conta)

    # Tenta endpoint que retorna URL temporaria.
    r_link = client.get(f'/nfes/gerar_link_xml/{emissao.tagplus_nfe_id}')
    if r_link.status_code == 200:
        try:
            body = r_link.json()
            url = body.get('url') if isinstance(body, dict) else None
            if url:
                return redirect(url)
        except ValueError:
            pass

    # Fallback: tenta baixar XML direto.
    r = client.get(f'/nfes/xml/{emissao.tagplus_nfe_id}')
    if r.status_code != 200:
        return jsonify({
            'ok': False, 'http_status': r.status_code,
            'detalhe': r.text[:500],
        }), 502

    filename = f'nfe_{emissao.numero_nfe or emissao.tagplus_nfe_id}.xml'
    return Response(
        r.content,
        mimetype='application/xml',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@hora_bp.route('/vendas/<int:venda_id>/nfe/xml-cancelamento')
@require_hora_perm('vendas', 'ver')
def venda_nfe_xml_cancelamento(venda_id: int):
    """Proxy do XML do EVENTO de cancelamento (protocolo SEFAZ).

    Diferente do XML da NFe original (`venda_nfe_xml`): aqui retornamos o XML
    do evento de cancelamento, que o contador precisa para escrita fiscal/SPED.

    Endpoint TagPlus (`scripts/doc_tagplus.md:2625-2636`):
      GET /nfes/gerar_xml_sem_assinatura/{id}?cancelada=true
    """
    emissao = HoraTagPlusNfeEmissao.query.filter_by(venda_id=venda_id).first()
    if not emissao or not emissao.tagplus_nfe_id:
        abort(404)
    if emissao.status != NFE_STATUS_CANCELADA:
        return jsonify({
            'ok': False,
            'detalhe': (
                f'Emissao em status {emissao.status} — XML de cancelamento '
                f'so existe quando NFe esta CANCELADA.'
            ),
        }), 409

    client = ApiClient(emissao.conta)
    r = client.get(
        f'/nfes/gerar_xml_sem_assinatura/{emissao.tagplus_nfe_id}',
        params={'cancelada': 'true'},
    )
    if r.status_code != 200:
        return jsonify({
            'ok': False, 'http_status': r.status_code,
            'detalhe': r.text[:500],
        }), 502

    filename = (
        f'cancelamento_{emissao.numero_nfe or emissao.tagplus_nfe_id}.xml'
    )
    return Response(
        r.content,
        mimetype='application/xml',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


# ============================================================================
# Backfill via API TagPlus (puxa NFs emitidas, sem PDF)
# ============================================================================

@hora_bp.route('/tagplus/backfill', methods=['GET', 'POST'])
@require_hora_perm('vendas', 'criar')
def tagplus_backfill():
    """Enfileira backfill TagPlus em RQ (queue hora_backfill).

    GET: tela com filtros (since/until/limite) + lista de jobs anteriores.
    POST: cria HoraTagPlusBackfillJob, enfileira e redireciona para o
          detalhe do job (auto-refresh AJAX).
    """
    from datetime import date, timedelta
    from app.hora.models import HoraTagPlusBackfillJob
    from app.hora.services.tagplus.backfill_service import (
        enfileirar_backfill_job,
    )

    today = date.today()
    default_since = (today - timedelta(days=30)).isoformat()
    default_until = today.isoformat()

    if request.method == 'POST':
        def _parse_data(name: str):
            v = (request.form.get(name) or '').strip()
            if not v:
                return None
            try:
                return date.fromisoformat(v)
            except ValueError:
                flash(f'Data invalida em {name}: {v!r}', 'danger')
                return None

        since = _parse_data('since')
        until = _parse_data('until')
        limite_raw = (request.form.get('limite') or '').strip()
        try:
            limite = int(limite_raw) if limite_raw else None
        except ValueError:
            limite = None
            flash(f'Limite invalido: {limite_raw!r} — ignorado.', 'warning')

        operador = (
            current_user.nome
            if current_user.is_authenticated and getattr(current_user, 'nome', None)
            else (current_user.email if current_user.is_authenticated else None)
        )

        try:
            job_id = enfileirar_backfill_job(
                since=since, until=until, operador=operador, limite=limite,
            )
            flash(
                f'Backfill enfileirado (job #{job_id}). '
                f'Acompanhe o progresso abaixo — pode fechar a aba e voltar depois.',
                'success',
            )
            return redirect(url_for('hora.tagplus_backfill_detalhe', job_id=job_id))
        except RuntimeError as exc:
            flash(f'Backfill nao pode ser enfileirado: {exc}', 'danger')
        except Exception as exc:  # pragma: no cover
            flash(f'Erro inesperado ao enfileirar backfill: {exc}', 'danger')
            logger.exception('Enfileiramento backfill TagPlus falhou')

    jobs = (
        HoraTagPlusBackfillJob.query
        .order_by(HoraTagPlusBackfillJob.id.desc())
        .limit(20)
        .all()
    )

    return render_template(
        'hora/tagplus/backfill.html',
        jobs=jobs,
        default_since=default_since, default_until=default_until,
    )


@hora_bp.route('/tagplus/backfill/<int:job_id>', methods=['GET'])
@require_hora_perm('vendas', 'criar')
def tagplus_backfill_detalhe(job_id: int):
    """Tela de detalhe de um HoraTagPlusBackfillJob com auto-refresh AJAX."""
    from app.hora.models import HoraTagPlusBackfillJob

    job = HoraTagPlusBackfillJob.query.get_or_404(job_id)
    return render_template('hora/tagplus/backfill_detalhe.html', job=job)


@hora_bp.route('/tagplus/backfill/nfe-unica', methods=['POST'])
@require_hora_perm('vendas', 'criar')
def tagplus_backfill_nfe_unica():
    """Importa UMA NFe especifica do TagPlus, sincronamente (modo teste).

    Nao usa RQ — executa direto na request. Util para validar pipeline
    apos um deploy ou investigar por que uma NFe especifica falhou no
    backfill em lote.
    """
    from app.hora.services.tagplus.backfill_service import (
        executar_backfill_unica_nfe,
    )

    nfe_id_raw = (request.form.get('tagplus_nfe_id') or '').strip()
    if not nfe_id_raw:
        flash('Informe o ID TagPlus da NFe.', 'warning')
        return redirect(url_for('hora.tagplus_backfill'))

    try:
        nfe_id = int(nfe_id_raw)
    except ValueError:
        flash(f'ID invalido: {nfe_id_raw!r} (esperado numero inteiro).', 'danger')
        return redirect(url_for('hora.tagplus_backfill'))

    operador = (
        current_user.nome
        if current_user.is_authenticated and getattr(current_user, 'nome', None)
        else (current_user.email if current_user.is_authenticated else None)
    )

    try:
        resultado = executar_backfill_unica_nfe(nfe_id, operador=operador)
    except RuntimeError as exc:
        flash(f'Backfill unica NFe falhou: {exc}', 'danger')
        return redirect(url_for('hora.tagplus_backfill'))

    status = resultado.get('status') or 'desconhecido'
    msg = resultado.get('mensagem') or ''
    venda_id = resultado.get('venda_id')

    # Mapeamento status -> categoria flash.
    if status in ('criado', 'atualizado', 'cancelado'):
        categoria = 'success'
    elif status in ('inalterado', 'pulada_cancelada', 'pulada_status_invalido'):
        categoria = 'info'
    elif status == 'duplicado':
        categoria = 'warning'
    else:
        categoria = 'danger'

    detalhe = (
        f'NFe {nfe_id} → status={status}. {msg}'
        + (f' (venda #{venda_id})' if venda_id else '')
    )
    flash(detalhe, categoria)

    # Se criou/atualizou venda, redireciona direto pra tela do pedido.
    if venda_id and status in ('criado', 'atualizado'):
        return redirect(url_for('hora.vendas_detalhe', venda_id=venda_id))
    return redirect(url_for('hora.tagplus_backfill'))


@hora_bp.route('/tagplus/backfill/<int:job_id>/json', methods=['GET'])
@require_hora_perm('vendas', 'criar')
def tagplus_backfill_job_json(job_id: int):
    """Endpoint AJAX para polling do progresso (a tela de detalhe consome)."""
    from flask import jsonify
    from app.hora.models import HoraTagPlusBackfillJob

    job = HoraTagPlusBackfillJob.query.get_or_404(job_id)

    def _fmt(dt):
        return dt.strftime('%d/%m/%Y %H:%M:%S') if dt else None

    # `erros` (cap 500) gravado incrementalmente em job.relatorio['erros'] —
    # ver backfill_worker._gravar_progresso. Permite que a tela atualize a
    # tabela de erros em tempo real durante o backfill.
    erros_lista = []
    if isinstance(job.relatorio, dict):
        erros_lista = job.relatorio.get('erros') or []

    return jsonify({
        'id': job.id,
        'tipo': job.tipo,
        'status': job.status,
        'iniciado_em': _fmt(job.iniciado_em),
        'finalizado_em': _fmt(job.finalizado_em),
        'total_listadas': job.total_listadas,
        'processadas': job.processadas,
        'n_criado': job.n_criado,
        'n_atualizado': job.n_atualizado,
        'n_inalterado': job.n_inalterado,
        'n_cancelado': job.n_cancelado,
        'n_pulada_cancelada': job.n_pulada_cancelada,
        'n_pulada_invalida': job.n_pulada_invalida,
        'n_dup': job.n_dup,
        'n_erro': job.n_erro,
        'n_divergencias': job.n_divergencias,
        'ultimo_erro': job.ultimo_erro,
        'em_estado_final': job.em_estado_final,
        'erros': erros_lista,
    })


# ============================================================================
# Backfill enriquecedor de pedidos via GET /pedidos/{id} TagPlus
# ============================================================================

@hora_bp.route('/tagplus/backfill-pedidos', methods=['GET', 'POST'])
@require_hora_perm('tagplus', 'editar')
def tagplus_backfill_pedidos():
    """Enfileira backfill de enriquecimento de pedidos em RQ.

    Para cada HoraTagPlusNfeEmissao APROVADA: puxa pedido_os_vinculada via
    GET /nfes/{id} e enriquece HoraVenda com dados do GET /pedidos/{id}
    (vendedor, departamento raw, forma_pagamento via mapa de IDs).

    NAO mexe em loja_id — usar /tagplus/departamento-map depois para
    aplicar o de-para de loja fisica.
    """
    from app.hora.models import (
        BACKFILL_JOB_TIPO_PEDIDO_ENRIQ,
        HoraTagPlusBackfillJob,
        HoraTagPlusNfeEmissao,
        NFE_STATUS_APROVADA,
    )
    from app.hora.services.tagplus.pedido_backfill_service import (
        enfileirar_backfill_pedidos_job,
    )

    if request.method == 'POST':
        limite_raw = (request.form.get('limite') or '').strip()
        try:
            limite = int(limite_raw) if limite_raw else None
        except ValueError:
            limite = None
            flash(f'Limite invalido: {limite_raw!r} — ignorado.', 'warning')

        operador = (
            current_user.nome
            if current_user.is_authenticated and getattr(current_user, 'nome', None)
            else (current_user.email if current_user.is_authenticated else None)
        )

        try:
            job_id = enfileirar_backfill_pedidos_job(
                operador=operador, limite=limite,
            )
            flash(
                f'Backfill de pedidos enfileirado (job #{job_id}). '
                f'Acompanhe o progresso abaixo — pode fechar a aba e voltar depois.',
                'success',
            )
            return redirect(url_for('hora.tagplus_backfill_detalhe', job_id=job_id))
        except RuntimeError as exc:
            flash(f'Backfill nao pode ser enfileirado: {exc}', 'danger')
        except Exception as exc:  # pragma: no cover
            flash(f'Erro inesperado: {exc}', 'danger')
            logger.exception('Enfileiramento backfill PEDIDOS falhou')

    # Universo previsto + jobs anteriores deste tipo.
    total_universo = db.session.query(HoraTagPlusNfeEmissao).filter(
        HoraTagPlusNfeEmissao.status == NFE_STATUS_APROVADA,
        HoraTagPlusNfeEmissao.tagplus_nfe_id.isnot(None),
    ).count()

    jobs = (
        HoraTagPlusBackfillJob.query
        .filter(HoraTagPlusBackfillJob.tipo == BACKFILL_JOB_TIPO_PEDIDO_ENRIQ)
        .order_by(HoraTagPlusBackfillJob.id.desc())
        .limit(20).all()
    )

    return render_template(
        'hora/tagplus/backfill_pedidos.html',
        jobs=jobs, total_universo=total_universo,
    )


# ============================================================================
# Backfill tagplus_pedido_id para vendas legadas (DANFE / sem emissao)
# ============================================================================

@hora_bp.route('/tagplus/backfill-pedidos-legados', methods=['GET', 'POST'])
@require_hora_perm('tagplus', 'editar')
def tagplus_backfill_pedidos_legados():
    """Enfileira backfill de tagplus_pedido_id para HoraVenda FATURADO sem
    pedido vinculado.

    Coberturas (em ordem de tentativa):
      - Vendas com HoraTagPlusNfeEmissao (path rapido, igual ao backfill de
        enriquecimento de pedidos).
      - Vendas DANFE legadas / origem MANUAL — descobre tagplus_nfe_id via
        GET /nfes em janela de datas (since=data_venda-7d, until=data_venda+7d),
        bate por chave_acesso (fallback numero).
    """
    from app.hora.models import (
        BACKFILL_JOB_TIPO_PEDIDO_VENDAS_LEGADAS,
        HoraTagPlusBackfillJob,
    )
    from app.hora.services.tagplus.pedido_backfill_service import (
        contar_universo_vendas_legadas,
        enfileirar_backfill_pedidos_vendas_legadas_job,
    )

    if request.method == 'POST':
        limite_raw = (request.form.get('limite') or '').strip()
        try:
            limite = int(limite_raw) if limite_raw else None
        except ValueError:
            limite = None
            flash(f'Limite invalido: {limite_raw!r} — ignorado.', 'warning')

        operador = (
            current_user.nome
            if current_user.is_authenticated and getattr(current_user, 'nome', None)
            else (current_user.email if current_user.is_authenticated else None)
        )

        try:
            job_id = enfileirar_backfill_pedidos_vendas_legadas_job(
                operador=operador, limite=limite,
            )
            flash(
                f'Backfill de vendas legadas enfileirado (job #{job_id}). '
                f'Acompanhe o progresso abaixo — pode fechar a aba e voltar depois.',
                'success',
            )
            return redirect(url_for('hora.tagplus_backfill_detalhe', job_id=job_id))
        except RuntimeError as exc:
            flash(f'Backfill nao pode ser enfileirado: {exc}', 'danger')
        except Exception as exc:  # pragma: no cover
            flash(f'Erro inesperado: {exc}', 'danger')
            logger.exception('Enfileiramento backfill PEDIDOS-VENDAS-LEGADAS falhou')

    total_universo = contar_universo_vendas_legadas()

    jobs = (
        HoraTagPlusBackfillJob.query
        .filter(HoraTagPlusBackfillJob.tipo == BACKFILL_JOB_TIPO_PEDIDO_VENDAS_LEGADAS)
        .order_by(HoraTagPlusBackfillJob.id.desc())
        .limit(20).all()
    )

    return render_template(
        'hora/tagplus/backfill_pedidos_legados.html',
        jobs=jobs, total_universo=total_universo,
    )


# ============================================================================
# De-para departamento TagPlus -> HoraLoja (revisao pos-backfill)
# ============================================================================

@hora_bp.route('/tagplus/departamento-map', methods=['GET'])
@require_hora_perm('vendas', 'editar')
def tagplus_departamento_map():
    """Lista departamentos coletados pelo backfill com de-para para HoraLoja.

    Sugestao automatica: match normalizado entre departamento_norm e
    apelido da HoraLoja sem prefixo "MOTOCHEFE ".
    """
    from app.hora.models import HoraLoja, HoraTagPlusDepartamentoMap
    from app.hora.services.tagplus.pedido_service import normalizar_departamento

    mapas = (
        HoraTagPlusDepartamentoMap.query
        .order_by(HoraTagPlusDepartamentoMap.qtd_vendas_observadas.desc())
        .all()
    )
    lojas = HoraLoja.query.filter_by(ativa=True).order_by(HoraLoja.id).all()

    # Sugestao automatica: normaliza apelido sem "MOTOCHEFE " e compara.
    def _suggest(mapa: HoraTagPlusDepartamentoMap):
        if mapa.loja_id:
            return None
        for loja in lojas:
            apelido = (loja.apelido or '').replace('MOTOCHEFE ', '').strip()
            if normalizar_departamento(apelido) == mapa.departamento_norm:
                return loja
        return None

    sugestoes = {m.id: _suggest(m) for m in mapas}

    return render_template(
        'hora/tagplus/departamento_map.html',
        mapas=mapas, lojas=lojas, sugestoes=sugestoes,
    )


@hora_bp.route('/tagplus/departamento-map/salvar', methods=['POST'])
@require_hora_perm('vendas', 'editar')
def tagplus_departamento_map_salvar():
    """Atribui loja_id aos departamentos a partir do form (batch)."""
    from app.hora.models import HoraTagPlusDepartamentoMap

    operador = (
        current_user.nome
        if current_user.is_authenticated and getattr(current_user, 'nome', None)
        else (current_user.email if current_user.is_authenticated else None)
    )

    alterados = 0
    for mapa in HoraTagPlusDepartamentoMap.query.all():
        novo_raw = (request.form.get(f'loja_{mapa.id}') or '').strip()
        if not novo_raw:
            continue  # vazio = mantem (nao limpa).
        try:
            novo_loja_id = int(novo_raw)
        except ValueError:
            continue
        if mapa.loja_id != novo_loja_id:
            mapa.loja_id = novo_loja_id
            mapa.revisado_por = operador
            from app.utils.timezone import agora_utc_naive
            mapa.revisado_em = agora_utc_naive()
            alterados += 1
    if alterados:
        db.session.commit()
        flash(f'{alterados} departamento(s) atualizado(s).', 'success')
    else:
        flash('Nenhuma alteracao detectada.', 'info')
    return redirect(url_for('hora.tagplus_departamento_map'))


@hora_bp.route('/tagplus/departamento-map/aplicar', methods=['POST'])
@require_hora_perm('vendas', 'editar')
def tagplus_departamento_map_aplicar():
    """Aplica UPDATE em massa em hora_venda.loja_id baseado no de-para.

    Para cada HoraTagPlusDepartamentoMap com loja_id NOT NULL:
      UPDATE hora_venda SET loja_id = mapa.loja_id
      WHERE tagplus_departamento (norm) = mapa.departamento_norm.

    Defesas:
      - PULA vendas onde UF da loja_destino difere da UF da loja_origem
        (evita mudar CFOP em re-emissao pos-cancelamento — REGRA FISCAL).
      - Registra HoraVendaAuditoria acao=DEFINIU_LOJA por venda atualizada
        (em bulk insert, preserva auditoria do modulo).

    Sem match no mapa, ou mapa.loja_id NULL: vendas mantem loja atual
    (matriz por default — sinalizador para revisao posterior).
    """
    from app.hora.models import (
        HoraLoja, HoraTagPlusDepartamentoMap, HoraVenda, HoraVendaAuditoria,
    )
    from app.hora.services.tagplus.pedido_service import normalizar_departamento
    from app.utils.timezone import agora_utc_naive

    operador = (
        current_user.nome
        if current_user.is_authenticated and getattr(current_user, 'nome', None)
        else (current_user.email if current_user.is_authenticated else None)
    )

    mapas_resolvidos = (
        HoraTagPlusDepartamentoMap.query
        .filter(HoraTagPlusDepartamentoMap.loja_id.isnot(None))
        .all()
    )
    if not mapas_resolvidos:
        flash(
            'Nenhum departamento mapeado para loja — preencha os dropdowns '
            'e clique Salvar mapeamento antes de aplicar.',
            'info',
        )
        return redirect(url_for('hora.tagplus_departamento_map'))

    # Indice por departamento_norm -> mapa, para lookup O(1) no loop unico.
    mapas_por_norm = {m.departamento_norm: m for m in mapas_resolvidos}

    # Cache UF por loja_id (evita N queries por venda).
    lojas = {l.id: l for l in HoraLoja.query.all()}

    # Carrega TODAS as vendas candidatas em UMA query (elimina N+1).
    vendas = HoraVenda.query.filter(
        HoraVenda.tagplus_departamento.isnot(None),
    ).all()

    total_atualizado = 0
    pulada_uf_diferente = 0
    detalhe: dict[str, int] = {}
    agora = agora_utc_naive()
    auditorias = []

    for v in vendas:
        norm = normalizar_departamento(v.tagplus_departamento)
        mapa = mapas_por_norm.get(norm)
        if mapa is None or v.loja_id == mapa.loja_id:
            continue

        # Defesa em profundidade: nao trocar loja se UF difere (CFOP muda).
        loja_origem = lojas.get(v.loja_id) if v.loja_id else None
        loja_destino = lojas.get(mapa.loja_id)
        uf_origem = (loja_origem.uf or '').upper() if loja_origem else None
        uf_destino = (loja_destino.uf or '').upper() if loja_destino else None
        if uf_origem and uf_destino and uf_origem != uf_destino:
            pulada_uf_diferente += 1
            continue

        loja_id_anterior = v.loja_id
        v.loja_id = mapa.loja_id
        total_atualizado += 1
        detalhe[mapa.departamento_raw] = detalhe.get(mapa.departamento_raw, 0) + 1
        auditorias.append({
            'venda_id': v.id,
            'usuario': operador or '',
            'acao': 'DEFINIU_LOJA',
            'detalhe': (
                f'De-para departamento TagPlus {v.tagplus_departamento!r} -> '
                f'loja_id {loja_id_anterior} -> {mapa.loja_id} '
                f'(loja={(loja_destino.apelido or loja_destino.nome) if loja_destino else "?"})'
            ),
            'criado_em': agora,
        })

    # Marca aplicado_em nos mapas que efetivamente afetaram vendas.
    for raw_label in detalhe:
        for m in mapas_resolvidos:
            if m.departamento_raw == raw_label:
                m.aplicado_em = agora
                break

    if total_atualizado:
        # Bulk insert da auditoria — preserva contrato de auditabilidade
        # do modulo HORA. ON DELETE SET NULL garantiu que item_id pode
        # ser NULL aqui (mudanca de header, nao item).
        db.session.bulk_insert_mappings(HoraVendaAuditoria, auditorias)
        db.session.commit()
        msg_detalhe = ', '.join(f'{k}={v}' for k, v in detalhe.items())
        msg = (
            f'{total_atualizado} venda(s) atualizada(s). Detalhe: '
            f'{msg_detalhe}. Operador: {operador}'
        )
        if pulada_uf_diferente:
            msg += (
                f'. {pulada_uf_diferente} venda(s) PULADAS por UF diferente '
                '(evita mudanca de CFOP em re-emissao).'
            )
        flash(msg, 'success')
    else:
        msg = 'Nenhuma venda atualizada'
        if pulada_uf_diferente:
            msg += (
                f' ({pulada_uf_diferente} pulada(s) por UF diferente)'
            )
        flash(msg + '.', 'info')
    return redirect(url_for('hora.tagplus_departamento_map'))


# ============================================================================
# Append-prompt do parser de chassi/motor (aprendizado por feedback)
# ============================================================================

@hora_bp.route('/tagplus/parser-append', methods=['GET'])
@require_hora_perm('tagplus', 'ver')
def tagplus_parser_append():
    """Tela de gestao do append-prompt: ver atual + historico + form de
    recomendacao/teste/gravacao via AJAX."""
    from app.hora.services.parser_append_service import (
        get_append_ativo, listar_historico,
    )
    ativo = get_append_ativo()
    historico = listar_historico(limit=20)
    # Pre-preenchimento opcional via querystring (?detalhes=...).
    detalhes_pre = (request.args.get('detalhes') or '').strip()
    return render_template(
        'hora/tagplus/parser_append.html',
        ativo=ativo,
        historico=historico,
        detalhes_pre=detalhes_pre,
    )


@hora_bp.route('/tagplus/parser-append/recomendar', methods=['POST'])
@require_hora_perm('tagplus', 'editar')
def tagplus_parser_append_recomendar():
    """Recebe `detalhes`, `extracao_atual`, `valor_correto` (JSON) e devolve
    o ACRESCIMO sugerido + o append PROPOSTO (atual + acrescimo)."""
    from flask import jsonify
    from app.hora.services.parser_append_service import (
        recomendar_acrescimo, texto_append_ativo,
    )

    data = request.get_json(silent=True) or {}
    detalhes = (data.get('detalhes') or '').strip()
    extracao_atual = data.get('extracao_atual') or {}
    valor_correto = data.get('valor_correto') or {}

    if not detalhes:
        return jsonify({'ok': False, 'erro': 'detalhes obrigatorio'}), 400
    if not (valor_correto.get('chassi') or valor_correto.get('motor')):
        return jsonify({
            'ok': False,
            'erro': 'informe ao menos chassi OU motor correto',
        }), 400

    append_atual = texto_append_ativo()
    acrescimo = recomendar_acrescimo(
        detalhes=detalhes,
        extracao_atual=extracao_atual,
        valor_correto=valor_correto,
        append_atual=append_atual,
    )
    if not acrescimo:
        return jsonify({
            'ok': False,
            'erro': 'LLM indisponivel ou nao retornou recomendacao',
        }), 502

    proposto = (
        (append_atual.rstrip() + '\n\n' + acrescimo.strip())
        if append_atual else acrescimo.strip()
    )
    return jsonify({
        'ok': True,
        'append_atual': append_atual,
        'acrescimo': acrescimo,
        'append_proposto': proposto,
    })


@hora_bp.route('/tagplus/parser-append/testar', methods=['POST'])
@require_hora_perm('tagplus', 'editar')
def tagplus_parser_append_testar():
    """Roda Haiku com `append_proposto` sobre `detalhes` e devolve a extracao."""
    from flask import jsonify
    from app.hora.services.parser_append_service import testar_append

    data = request.get_json(silent=True) or {}
    detalhes = (data.get('detalhes') or '').strip()
    append_proposto = (data.get('append_proposto') or '').strip()
    if not detalhes:
        return jsonify({'ok': False, 'erro': 'detalhes obrigatorio'}), 400

    resultado = testar_append(detalhes, append_proposto)
    return jsonify({
        'ok': resultado.get('ok', False),
        'chassi': resultado.get('chassi'),
        'motor': resultado.get('motor'),
        'raw_response': resultado.get('_raw_response'),
    })


@hora_bp.route('/tagplus/parser-append/salvar', methods=['POST'])
@require_hora_perm('tagplus', 'editar')
def tagplus_parser_append_salvar():
    """Persiste novo append (texto completo + acrescimo + motivo)."""
    from app.hora.services.parser_append_service import salvar_nova_versao

    texto = (request.form.get('texto_completo') or '').strip()
    acrescimo = (request.form.get('acrescimo_aplicado') or '').strip() or None
    motivo = (request.form.get('motivo') or '').strip() or None
    if not texto:
        flash('Texto do append vazio.', 'danger')
        return redirect(url_for('hora.tagplus_parser_append'))

    autor = (
        current_user.nome
        if current_user.is_authenticated and getattr(current_user, 'nome', None)
        else (current_user.email if current_user.is_authenticated else None)
    )
    try:
        nova = salvar_nova_versao(
            texto, acrescimo_aplicado=acrescimo, motivo=motivo,
            criado_por=autor,
        )
        flash(f'Append v{nova.versao} gravado e ativado.', 'success')
    except Exception as exc:  # pragma: no cover
        logger.exception('Falha ao gravar append')
        flash(f'Erro ao gravar: {exc}', 'danger')
    return redirect(url_for('hora.tagplus_parser_append'))


@hora_bp.route('/tagplus/parser-append/<int:append_id>/reativar', methods=['POST'])
@require_hora_perm('tagplus', 'editar')
def tagplus_parser_append_reativar(append_id: int):
    """Rollback: marca a versao escolhida como ativa (clona como nova versao)."""
    from app.hora.models.tagplus import HoraDanfeParserAppend
    from app.hora.services.parser_append_service import salvar_nova_versao

    alvo = HoraDanfeParserAppend.query.get_or_404(append_id)
    autor = (
        current_user.nome
        if current_user.is_authenticated and getattr(current_user, 'nome', None)
        else (current_user.email if current_user.is_authenticated else None)
    )
    try:
        nova = salvar_nova_versao(
            alvo.texto_append,
            motivo=f'Rollback para v{alvo.versao}',
            criado_por=autor,
        )
        flash(
            f'Append v{nova.versao} (clone de v{alvo.versao}) ativado.',
            'success',
        )
    except Exception as exc:  # pragma: no cover
        flash(f'Erro: {exc}', 'danger')
    return redirect(url_for('hora.tagplus_parser_append'))


# ========================================================================
# Mapeamento TagPlus de PECAS (paralelo a tagplus_produto_map de motos)
# ========================================================================

@hora_bp.route('/tagplus/peca-map')
@require_hora_perm('tagplus', 'ver')
def tagplus_peca_map_lista():
    """Lista de mapeamentos TagPlus -> peca."""
    from app.hora.models import HoraPeca, HoraTagPlusPecaMap
    pecas = HoraPeca.query.filter_by(ativo=True).order_by(HoraPeca.codigo_interno).all()
    rows = []
    for p in pecas:
        m = HoraTagPlusPecaMap.query.filter_by(peca_id=p.id).first()
        rows.append({'peca': p, 'map': m})
    return render_template('hora/tagplus_peca_map_lista.html', rows=rows)


@hora_bp.route('/tagplus/peca-map/<int:peca_id>/salvar', methods=['POST'])
@require_hora_perm('tagplus', 'editar')
def tagplus_peca_map_salvar(peca_id: int):
    from app.hora.services import peca_service
    try:
        tagplus_id = (request.form.get('tagplus_produto_id') or '').strip()
        if not tagplus_id:
            from app.hora.models import HoraTagPlusPecaMap
            existing = HoraTagPlusPecaMap.query.filter_by(peca_id=peca_id).first()
            if existing:
                peca_service.remover_tagplus_map(peca_id)
                flash('Mapeamento removido.', 'info')
        else:
            peca_service.set_tagplus_map(
                peca_id=peca_id,
                tagplus_produto_id=tagplus_id,
                tagplus_codigo=(request.form.get('tagplus_codigo') or '') or None,
                cfop_default=(request.form.get('cfop_default') or '') or None,
            )
            flash('Mapeamento salvo.', 'success')
    except ValueError as exc:
        flash(f'Erro: {exc}', 'danger')
    return redirect(url_for('hora.tagplus_peca_map_lista'))


# ========================================================================
# Backfill catalogo de produtos -> hora_peca + hora_tagplus_peca_map
# ========================================================================

@hora_bp.route('/tagplus/backfill-produtos-pecas', methods=['GET'])
@require_hora_perm('tagplus', 'ver')
def tagplus_backfill_produtos():
    """Tela de form para executar backfill de catalogo de pecas."""
    return render_template('hora/tagplus_backfill_produtos.html', relatorio=None)


@hora_bp.route('/tagplus/backfill-produtos-pecas/executar', methods=['POST'])
@require_hora_perm('tagplus', 'editar')
def tagplus_backfill_produtos_executar():
    from app.hora.services.tagplus import backfill_service
    try:
        relatorio = backfill_service.executar_backfill_produtos_pecas(operador=_operador())
        flash(
            f'Backfill produtos: {relatorio["criadas"]} criadas, '
            f'{relatorio["atualizadas"]} atualizadas, '
            f'{relatorio["puladas_moto"]} motos puladas, '
            f'{relatorio["erros"]} erros.',
            'success' if relatorio['erros'] == 0 else 'warning',
        )
    except Exception as exc:  # pragma: no cover
        flash(f'Erro no backfill: {exc}', 'danger')
        relatorio = {'erros': 1, 'erros_detalhe': [str(exc)]}
    return render_template('hora/tagplus_backfill_produtos.html', relatorio=relatorio)


# ========================================================================
# Backfill delta (NFes legadas com peca ignorada)
# ========================================================================

@hora_bp.route('/tagplus/backfill-pecas-delta', methods=['GET'])
@require_hora_perm('tagplus', 'ver')
def tagplus_backfill_pecas_delta():
    return render_template('hora/tagplus_backfill_pecas_delta.html', relatorio=None)


@hora_bp.route('/tagplus/backfill-pecas-delta/executar', methods=['POST'])
@require_hora_perm('tagplus', 'editar')
def tagplus_backfill_pecas_delta_executar():
    from app.hora.services.tagplus import backfill_service
    try:
        limite_str = (request.form.get('limite') or '50').strip()
        limite = int(limite_str) if limite_str.isdigit() else 50
        relatorio = backfill_service.executar_backfill_pecas_faltantes(
            operador=_operador(), limite=limite,
        )
        flash(
            f'Backfill delta: {relatorio["analisadas"]} analisadas, '
            f'{relatorio["reprocessadas"]} reprocessadas, '
            f'{relatorio["pecas_criadas"]} pecas criadas, '
            f'{relatorio["sem_emissao"]} sem emissao, '
            f'{relatorio["sem_mapping"]} sem mapping, '
            f'{relatorio["erros"]} erros.',
            'success' if relatorio['erros'] == 0 else 'warning',
        )
    except Exception as exc:  # pragma: no cover
        flash(f'Erro no backfill: {exc}', 'danger')
        relatorio = {'erros': 1, 'detalhes': [str(exc)]}
    return render_template('hora/tagplus_backfill_pecas_delta.html', relatorio=relatorio)
