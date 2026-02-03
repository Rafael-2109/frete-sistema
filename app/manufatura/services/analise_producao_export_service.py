"""
Serviço de Exportação Excel para Análise de Produção
Gera planilhas com agrupamento por Ordem ou Dia, com/sem Lista de Materiais (BOM)
"""

import io
import logging
from typing import Dict, List, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.estoque.models import MovimentacaoEstoque
from app.manufatura.services.bom_service import ServicoBOM

logger = logging.getLogger(__name__)


class AnaliseProducaoExportService:
    """Serviço para exportação Excel da Análise de Produção"""

    # ============================================================
    # ESTILOS PADRÃO
    # ============================================================
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
    SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    SUBHEADER_FONT = Font(color="FFFFFF", bold=True, size=9)
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    # Cores por tipo de componente BOM
    FILL_INTERMEDIARIO = PatternFill(start_color="E8D5F5", end_color="E8D5F5", fill_type="solid")
    FILL_COMPONENTE = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    FILL_ACABADO = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    FILL_NEGATIVO = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    # ============================================================
    # PONTO DE ENTRADA
    # ============================================================
    @staticmethod
    def exportar(
        itens: List[Dict[str, Any]],
        agrupamento: str,
        com_bom: bool,
        filtros: Dict[str, str]
    ) -> bytes:
        """
        Gera arquivo Excel com os dados de produção.

        Args:
            itens: Lista de dicts com dados de produção (vindos da query)
            agrupamento: 'ordem' ou 'dia'
            com_bom: True para incluir sheet de Lista de Materiais
            filtros: Dict com filtros aplicados (para referência no nome)

        Returns:
            Bytes do arquivo Excel pronto para download
        """
        _ = filtros  # Reservado para uso futuro (ex: incluir resumo de filtros no Excel)

        wb = Workbook()

        # Sheet principal de produção
        if agrupamento == 'ordem':
            AnaliseProducaoExportService._gerar_sheet_producao_por_ordem(wb, itens)
        else:
            AnaliseProducaoExportService._gerar_sheet_producao_por_dia(wb, itens)

        # Sheet de BOM (se solicitado)
        if com_bom:
            AnaliseProducaoExportService._gerar_sheet_bom(wb, itens, agrupamento)

        # Gerar bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    # ============================================================
    # SHEET: PRODUÇÃO POR ORDEM
    # ============================================================
    @staticmethod
    def _gerar_sheet_producao_por_ordem(wb: Workbook, itens: List[Dict]) -> None:
        """Gera sheet 'Produção por Ordem' com dados agrupados por (ordem_producao, cod_produto)"""
        ws = wb.active
        ws.title = "Produção por Ordem"

        # Headers
        headers = [
            "Ordem Produção", "Código", "Produto",
            "Qtd Total", "Qtd Produções", "Última Data", "Local"
        ]
        AnaliseProducaoExportService._escrever_header(ws, headers, row=1)

        # Dados
        for row_idx, item in enumerate(itens, 2):
            ultima_data = item.get('ultima_data')
            if ultima_data and hasattr(ultima_data, 'strftime'):
                data_fmt = ultima_data.strftime('%d/%m/%Y')
            else:
                data_fmt = str(ultima_data) if ultima_data else '-'

            valores = [
                item.get('ordem_producao', '-') or '-',
                item.get('cod_produto', ''),
                item.get('nome_produto', ''),
                round(float(item.get('qtd_total', 0)), 3),
                int(item.get('qtd_producoes', 0)),
                data_fmt,
                item.get('local_movimentacao', '-') or '-'
            ]
            for col_idx, val in enumerate(valores, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = AnaliseProducaoExportService.THIN_BORDER

        # Larguras
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 18

    # ============================================================
    # SHEET: PRODUÇÃO POR DIA
    # ============================================================
    @staticmethod
    def _gerar_sheet_producao_por_dia(wb: Workbook, itens: List[Dict]) -> None:
        """Gera sheet 'Produção por Dia' com dados agrupados por (data, ordem_producao, cod_produto)"""
        ws = wb.active
        ws.title = "Produção por Dia"

        # Headers
        headers = [
            "Data", "Ordem Produção", "Código", "Produto",
            "Qtd Total", "Qtd Produções", "Local"
        ]
        AnaliseProducaoExportService._escrever_header(ws, headers, row=1)

        # Dados
        for row_idx, item in enumerate(itens, 2):
            data_mov = item.get('data_movimentacao') or item.get('ultima_data')
            if data_mov and hasattr(data_mov, 'strftime'):
                data_fmt = data_mov.strftime('%d/%m/%Y')
            else:
                data_fmt = str(data_mov) if data_mov else '-'

            valores = [
                data_fmt,
                item.get('ordem_producao', '-') or '-',
                item.get('cod_produto', ''),
                item.get('nome_produto', ''),
                round(float(item.get('qtd_total', 0)), 3),
                int(item.get('qtd_producoes', 0)),
                item.get('local_movimentacao', '-') or '-'
            ]
            for col_idx, val in enumerate(valores, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = AnaliseProducaoExportService.THIN_BORDER

        # Larguras
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 18

    # ============================================================
    # SHEET: LISTA DE MATERIAIS (BOM)
    # ============================================================
    @staticmethod
    def _gerar_sheet_bom(wb: Workbook, itens: List[Dict], agrupamento: str) -> None:
        """
        Gera sheet 'Lista de Materiais' com BOM explodido + consumos/ajustes.
        Cada item de produção gera N linhas (uma por componente da BOM).
        """
        ws = wb.create_sheet(title="Lista de Materiais")

        # Headers (colunas iniciais mudam conforme agrupamento)
        if agrupamento == 'ordem':
            headers_inicio = ["Ordem Produção"]
        else:
            headers_inicio = ["Data", "Ordem Produção"]

        headers = headers_inicio + [
            "Prod. Código", "Prod. Nome",
            "Comp. Código", "Comp. Nome", "Tipo", "Nível",
            "Qtd Prevista", "Consumo Registrado", "Ajuste", "Consumo Real"
        ]
        AnaliseProducaoExportService._escrever_header(ws, headers, row=1)

        # Cache de BOM por cod_produto (evita re-explodir a mesma estrutura)
        bom_cache: Dict[str, Dict] = {}

        # Coletar TODOS os operacao_producao_ids para buscar consumos/ajustes em batch
        todos_op_ids = []
        for item in itens:
            op_ids = item.get('operacao_ids', '')
            if op_ids:
                if isinstance(op_ids, str):
                    todos_op_ids.extend([oid.strip() for oid in op_ids.split(',') if oid.strip()])
                elif isinstance(op_ids, list):
                    todos_op_ids.extend(op_ids)

        # Buscar consumos e ajustes em batch
        consumos_por_op = {}
        ajustes_por_op = {}
        if todos_op_ids:
            consumos_por_op, ajustes_por_op = AnaliseProducaoExportService._buscar_consumos_ajustes_batch(
                todos_op_ids
            )

        # Gerar linhas
        row_idx = 2
        for item in itens:
            cod_produto = item.get('cod_produto', '')
            qtd_total = float(item.get('qtd_total', 0))

            if not cod_produto or qtd_total == 0:
                continue

            # Colunas iniciais (variam conforme agrupamento)
            if agrupamento == 'ordem':
                cols_inicio = [item.get('ordem_producao', '-') or '-']
            else:
                data_mov = item.get('data_movimentacao') or item.get('ultima_data')
                if data_mov and hasattr(data_mov, 'strftime'):
                    data_fmt = data_mov.strftime('%d/%m/%Y')
                else:
                    data_fmt = str(data_mov) if data_mov else '-'
                cols_inicio = [data_fmt, item.get('ordem_producao', '-') or '-']

            # Explodir BOM (com cache)
            if cod_produto not in bom_cache:
                try:
                    bom_cache[cod_produto] = ServicoBOM.explodir_bom(cod_produto, 1.0)
                except Exception as e:
                    logger.warning(f"Erro ao explodir BOM de {cod_produto}: {e}")
                    bom_cache[cod_produto] = {'componentes': [], 'tem_estrutura': False}

            bom_base = bom_cache[cod_produto]
            if not bom_base.get('tem_estrutura'):
                continue

            # Juntar consumos/ajustes dos operacao_ids deste item
            op_ids_item = item.get('operacao_ids', '')
            if isinstance(op_ids_item, str):
                op_ids_lista = [oid.strip() for oid in op_ids_item.split(',') if oid.strip()]
            elif isinstance(op_ids_item, list):
                op_ids_lista = op_ids_item
            else:
                op_ids_lista = []

            consumos_item = AnaliseProducaoExportService._agrupar_por_produto(
                consumos_por_op, op_ids_lista
            )
            ajustes_item = AnaliseProducaoExportService._agrupar_por_produto(
                ajustes_por_op, op_ids_lista
            )

            # Achatar BOM e escrever linhas
            componentes = AnaliseProducaoExportService._achatar_bom(
                bom_base, consumos_item, ajustes_item, qtd_total
            )

            for comp in componentes:
                tipo = comp.get('tipo', '')
                valores = cols_inicio + [
                    cod_produto,
                    item.get('nome_produto', ''),
                    comp.get('cod_produto', ''),
                    comp.get('nome_produto', ''),
                    tipo,
                    comp.get('nivel', 0),
                    round(comp.get('consumo_previsto', 0), 3),
                    round(comp.get('consumo_registrado', 0), 3),
                    round(comp.get('ajuste_registrado', 0), 3),
                    round(comp.get('consumo_real', 0), 3)
                ]
                for col_idx, val in enumerate(valores, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = AnaliseProducaoExportService.THIN_BORDER

                    # Cor de fundo por tipo de componente
                    if tipo == 'INTERMEDIARIO':
                        cell.fill = AnaliseProducaoExportService.FILL_INTERMEDIARIO
                    elif tipo == 'COMPONENTE':
                        cell.fill = AnaliseProducaoExportService.FILL_COMPONENTE
                    elif tipo == 'ACABADO':
                        cell.fill = AnaliseProducaoExportService.FILL_ACABADO

                row_idx += 1

        # Larguras (ajuste dinâmico conforme agrupamento)
        if agrupamento == 'ordem':
            # 11 colunas: Ordem | Prod.Cód | Prod.Nome | Comp.Cód | Comp.Nome | Tipo | Nível | Prev | Cons | Aj | Real
            larguras = {'A': 18, 'B': 14, 'C': 30, 'D': 14, 'E': 30, 'F': 16, 'G': 8,
                        'H': 18, 'I': 18, 'J': 18, 'K': 18}
        else:
            # 12 colunas: Data | Ordem | Prod.Cód | Prod.Nome | Comp.Cód | Comp.Nome | Tipo | Nível | Prev | Cons | Aj | Real
            larguras = {'A': 14, 'B': 18, 'C': 14, 'D': 30, 'E': 14, 'F': 30, 'G': 16,
                        'H': 8, 'I': 18, 'J': 18, 'K': 18, 'L': 18}
        for col_letter, width in larguras.items():
            ws.column_dimensions[col_letter].width = width

    # ============================================================
    # HELPERS
    # ============================================================
    @staticmethod
    def _escrever_header(ws, headers: List[str], row: int = 1) -> None:
        """Escreve linha de cabeçalho com estilo padrão"""
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = AnaliseProducaoExportService.HEADER_FILL
            cell.font = AnaliseProducaoExportService.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = AnaliseProducaoExportService.THIN_BORDER

    @staticmethod
    def _buscar_consumos_ajustes_batch(
        operacao_ids: List[str]
    ) -> tuple:
        """
        Busca TODOS os consumos e ajustes em batch para uma lista de operacao_producao_ids.

        Returns:
            (consumos_por_op, ajustes_por_op)
            Cada um é dict: {operacao_producao_id: {cod_produto: qtd}}
        """
        consumos_por_op: Dict[str, Dict[str, float]] = {}
        ajustes_por_op: Dict[str, Dict[str, float]] = {}

        if not operacao_ids:
            return consumos_por_op, ajustes_por_op

        try:
            # Buscar consumos
            consumos = MovimentacaoEstoque.query.filter(
                MovimentacaoEstoque.operacao_producao_id.in_(operacao_ids),
                MovimentacaoEstoque.tipo_movimentacao == 'CONSUMO',
                MovimentacaoEstoque.ativo == True  # noqa: E712
            ).all()

            for c in consumos:
                op_id = c.operacao_producao_id
                cod = c.cod_produto
                if op_id not in consumos_por_op:
                    consumos_por_op[op_id] = {}
                if cod not in consumos_por_op[op_id]:
                    consumos_por_op[op_id][cod] = 0
                consumos_por_op[op_id][cod] += abs(float(c.qtd_movimentacao or 0))

            # Buscar ajustes
            ajustes = MovimentacaoEstoque.query.filter(
                MovimentacaoEstoque.operacao_producao_id.in_(operacao_ids),
                MovimentacaoEstoque.tipo_movimentacao == 'AJUSTE',
                MovimentacaoEstoque.ativo == True  # noqa: E712
            ).all()

            for a in ajustes:
                op_id = a.operacao_producao_id
                cod = a.cod_produto
                if op_id not in ajustes_por_op:
                    ajustes_por_op[op_id] = {}
                if cod not in ajustes_por_op[op_id]:
                    ajustes_por_op[op_id][cod] = 0
                ajustes_por_op[op_id][cod] += float(a.qtd_movimentacao or 0)

        except Exception as e:
            logger.error(f"Erro ao buscar consumos/ajustes em batch: {e}")

        return consumos_por_op, ajustes_por_op

    @staticmethod
    def _agrupar_por_produto(
        dados_por_op: Dict[str, Dict[str, float]],
        op_ids: List[str]
    ) -> Dict[str, float]:
        """
        Soma os valores de vários operacao_producao_ids em um único dict {cod_produto: qtd_total}.
        """
        resultado: Dict[str, float] = {}
        for op_id in op_ids:
            produtos = dados_por_op.get(op_id, {})
            for cod, qtd in produtos.items():
                resultado[cod] = resultado.get(cod, 0) + qtd
        return resultado

    @staticmethod
    def _achatar_bom(
        bom_explodido: Dict,
        consumos: Dict[str, float],
        ajustes: Dict[str, float],
        qtd_produzida: float
    ) -> List[Dict[str, Any]]:
        """
        Achata estrutura BOM recursiva em lista plana com consumos/ajustes.
        A BOM base é explodida para 1 unidade; multiplica pela qtd_produzida.

        Args:
            bom_explodido: Resultado de ServicoBOM.explodir_bom(cod, 1.0)
            consumos: {cod_produto: qtd_consumida} do item de produção
            ajustes: {cod_produto: qtd_ajuste_estoque} do item de produção
            qtd_produzida: Quantidade total produzida (para proporcionalizar BOM)

        Returns:
            Lista de dicts com dados de cada componente
        """
        resultados: List[Dict[str, Any]] = []

        def _recursar(componente_info: Dict, nivel: int = 1):
            for comp in componente_info.get('componentes', []):
                cod = comp.get('cod_produto', '')

                # Proporcionalizar: BOM base é para 1 unidade
                consumo_previsto = float(comp.get('qtd_necessaria', 0)) * qtd_produzida

                # Consumo registrado e ajuste
                consumo_registrado = consumos.get(cod, 0)
                ajuste_estoque = ajustes.get(cod, 0)
                ajuste_registrado = -ajuste_estoque  # Inverter para visão de consumo
                consumo_real = consumo_registrado + ajuste_registrado

                resultados.append({
                    'cod_produto': cod,
                    'nome_produto': comp.get('nome_produto', ''),
                    'tipo': comp.get('tipo', ''),
                    'nivel': nivel,
                    'consumo_previsto': consumo_previsto,
                    'consumo_registrado': consumo_registrado,
                    'ajuste_registrado': ajuste_registrado,
                    'consumo_real': consumo_real
                })

                # Recursão para intermediários
                if comp.get('tem_estrutura') and comp.get('componentes'):
                    _recursar(comp, nivel + 1)

        _recursar(bom_explodido)
        return resultados
