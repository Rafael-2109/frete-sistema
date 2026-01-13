"""
Serviço de Macro Projeção de Componentes
========================================

Calcula projeção macro de necessidades de componentes (produto_comprado=True)
com visão consolidada de Ruptura e Saldo por Mês.

Considera UnificacaoCodigos para evitar duplicação na exibição.

Autor: Sistema de Fretes
Data: 13/01/2026
"""

from datetime import date
from decimal import Decimal
from typing import Dict, List, Any, Set
from collections import defaultdict
from sqlalchemy import func, extract
import logging

from app import db
from app.producao.models import CadastroPalletizacao
from app.estoque.models import MovimentacaoEstoque, UnificacaoCodigos
from app.manufatura.models import PrevisaoDemanda, PedidoCompras
from app.carteira.models import CarteiraPrincipal
from app.separacao.models import Separacao

logger = logging.getLogger(__name__)


class MacroProjecaoComponentesService:
    """
    Serviço para calcular projeção macro de componentes

    Estrutura do relatório:
    - cod_produto
    - nome_produto
    - RUPTURA: estoque_atual, necessidade, saldo
    - SALDO MÊS ATUAL: chegada, necessidade, saldo
    - SALDO MÊS +1: chegada, necessidade, saldo
    """

    def __init__(self):
        self.hoje = date.today()
        self.mes_atual = self.hoje.month
        self.ano_atual = self.hoje.year

        # Calcular mês+1
        if self.mes_atual == 12:
            self.mes_proximo = 1
            self.ano_proximo = self.ano_atual + 1
        else:
            self.mes_proximo = self.mes_atual + 1
            self.ano_proximo = self.ano_atual

    def calcular_projecao_macro(
        self,
        categoria: str = None,
        tipo_materia_prima: str = None,
        linha_producao: str = None
    ) -> Dict[str, Any]:
        """
        Calcula a projeção macro para todos os componentes

        Args:
            categoria: Filtro por categoria_produto
            tipo_materia_prima: Filtro por tipo_materia_prima
            linha_producao: Filtro por linha_producao

        Returns:
            Dict com dados da projeção e metadados
        """
        try:
            # 1. Buscar todos os produtos comprados (componentes)
            query = CadastroPalletizacao.query.filter_by(
                produto_comprado=True,
                ativo=True
            )

            # Aplicar filtros
            if categoria:
                query = query.filter(CadastroPalletizacao.categoria_produto == categoria)
            if tipo_materia_prima:
                query = query.filter(CadastroPalletizacao.tipo_materia_prima == tipo_materia_prima)
            if linha_producao:
                query = query.filter(CadastroPalletizacao.linha_producao == linha_producao)

            produtos_comprados = query.order_by(CadastroPalletizacao.cod_produto).all()

            if not produtos_comprados:
                return {
                    'sucesso': True,
                    'data_calculo': self.hoje.isoformat(),
                    'mes_atual': self.mes_atual,
                    'ano_atual': self.ano_atual,
                    'mes_proximo': self.mes_proximo,
                    'ano_proximo': self.ano_proximo,
                    'total_componentes': 0,
                    'componentes': []
                }

            # 2. Identificar códigos já processados (para evitar duplicação por unificação)
            codigos_processados: Set[str] = set()
            componentes_resultado = []

            for produto in produtos_comprados:
                cod_produto = produto.cod_produto

                # Verificar se este código já foi processado (ou um relacionado)
                codigos_relacionados = UnificacaoCodigos.get_todos_codigos_relacionados(cod_produto)

                # Se algum código relacionado já foi processado, pular
                if any(cod in codigos_processados for cod in codigos_relacionados):
                    continue

                # Marcar todos os códigos relacionados como processados
                codigos_processados.update(codigos_relacionados)

                # 3. Calcular dados para este componente (considerando unificação)
                dados_componente = self._calcular_dados_componente(
                    cod_produto=cod_produto,
                    nome_produto=produto.nome_produto,
                    codigos_relacionados=codigos_relacionados,
                    categoria=produto.categoria_produto,
                    tipo_materia_prima=produto.tipo_materia_prima,
                    linha_producao=produto.linha_producao
                )

                componentes_resultado.append(dados_componente)

            # 4. Ordenar por saldo de ruptura (mais críticos primeiro)
            componentes_resultado.sort(key=lambda x: x['ruptura']['saldo'])

            return {
                'sucesso': True,
                'data_calculo': self.hoje.isoformat(),
                'mes_atual': self.mes_atual,
                'ano_atual': self.ano_atual,
                'mes_proximo': self.mes_proximo,
                'ano_proximo': self.ano_proximo,
                'total_componentes': len(componentes_resultado),
                'componentes': componentes_resultado
            }

        except Exception as e:
            logger.error(f"Erro ao calcular projeção macro: {e}")
            return {
                'sucesso': False,
                'erro': str(e),
                'componentes': []
            }

    def _calcular_dados_componente(
        self,
        cod_produto: str,
        nome_produto: str,
        codigos_relacionados: List[str],
        categoria: str = None,
        tipo_materia_prima: str = None,
        linha_producao: str = None
    ) -> Dict[str, Any]:
        """
        Calcula todos os dados de um componente

        Args:
            cod_produto: Código principal do produto
            nome_produto: Nome do produto
            codigos_relacionados: Lista de códigos unificados
            categoria: Categoria do produto
            tipo_materia_prima: Tipo de matéria-prima
            linha_producao: Linha de produção
        """
        # === RUPTURA ===
        estoque_atual = self._calcular_estoque_atual(codigos_relacionados)
        necessidade_ruptura = self._calcular_necessidade_ruptura(codigos_relacionados)
        saldo_ruptura = estoque_atual - necessidade_ruptura

        # === SALDO MÊS ATUAL ===
        chegada_mes_atual = self._calcular_chegadas_mes(
            codigos_relacionados, self.mes_atual, self.ano_atual
        )
        necessidade_mes_atual = self._calcular_necessidade_mes(
            codigos_relacionados, self.mes_atual, self.ano_atual
        )
        saldo_mes_atual = saldo_ruptura + chegada_mes_atual - necessidade_mes_atual

        # === SALDO MÊS +1 ===
        chegada_mes_proximo = self._calcular_chegadas_mes(
            codigos_relacionados, self.mes_proximo, self.ano_proximo
        )
        necessidade_mes_proximo = self._calcular_necessidade_mes(
            codigos_relacionados, self.mes_proximo, self.ano_proximo
        )
        saldo_mes_proximo = saldo_mes_atual + chegada_mes_proximo - necessidade_mes_proximo

        return {
            'cod_produto': cod_produto,
            'nome_produto': nome_produto,
            'codigos_unificados': codigos_relacionados if len(codigos_relacionados) > 1 else None,
            'categoria': categoria or '',
            'tipo_materia_prima': tipo_materia_prima or '',
            'linha_producao': linha_producao or '',
            'ruptura': {
                'estoque_atual': round(estoque_atual, 2),
                'necessidade': round(necessidade_ruptura, 2),
                'saldo': round(saldo_ruptura, 2)
            },
            'saldo_mes_atual': {
                'mes': self.mes_atual,
                'ano': self.ano_atual,
                'chegada': round(chegada_mes_atual, 2),
                'necessidade': round(necessidade_mes_atual, 2),
                'saldo': round(saldo_mes_atual, 2)
            },
            'saldo_mes_proximo': {
                'mes': self.mes_proximo,
                'ano': self.ano_proximo,
                'chegada': round(chegada_mes_proximo, 2),
                'necessidade': round(necessidade_mes_proximo, 2),
                'saldo': round(saldo_mes_proximo, 2)
            }
        }

    def _calcular_estoque_atual(self, codigos: List[str]) -> float:
        """
        Calcula estoque atual somando movimentações ativas

        Fonte: movimentacao_estoque.qtd_movimentacao WHERE ativo=True
        """
        try:
            resultado = db.session.query(
                func.sum(MovimentacaoEstoque.qtd_movimentacao)
            ).filter(
                MovimentacaoEstoque.cod_produto.in_(codigos),
                MovimentacaoEstoque.ativo == True
            ).scalar()

            return float(resultado or 0)
        except Exception as e:
            logger.error(f"Erro ao calcular estoque para {codigos}: {e}")
            return 0.0

    def _calcular_necessidade_ruptura(self, codigos: List[str]) -> float:
        """
        Calcula necessidade total para ruptura

        Fórmula: Saldo Previsão + Saldo Carteira + Pedidos Programados

        - Saldo Previsão = qtd_demanda_prevista - qtd_demanda_realizada
        - Saldo Carteira = CarteiraPrincipal.qtd_saldo - Separacao.qtd_saldo (não faturado)
        - Pedidos Programados = Separacao.qtd_saldo (mesmo mês/ano, não faturado)
        """
        try:
            # 1. Saldo da Previsão (todos os meses futuros + mês atual)
            saldo_previsao = db.session.query(
                func.sum(PrevisaoDemanda.qtd_demanda_prevista - PrevisaoDemanda.qtd_demanda_realizada)
            ).filter(
                PrevisaoDemanda.cod_produto.in_(codigos),
                db.or_(
                    PrevisaoDemanda.data_ano > self.ano_atual,
                    db.and_(
                        PrevisaoDemanda.data_ano == self.ano_atual,
                        PrevisaoDemanda.data_mes >= self.mes_atual
                    )
                )
            ).scalar() or Decimal('0')

            # 2. Saldo da Carteira (qtd_saldo_produto_pedido)
            saldo_carteira = db.session.query(
                func.sum(CarteiraPrincipal.qtd_saldo_produto_pedido)
            ).filter(
                CarteiraPrincipal.cod_produto.in_(codigos),
                CarteiraPrincipal.qtd_saldo_produto_pedido > 0
            ).scalar() or Decimal('0')

            # 3. Separações não faturadas (desconta da carteira para evitar duplicação)
            separacoes_nao_faturadas = db.session.query(
                func.sum(Separacao.qtd_saldo)
            ).filter(
                Separacao.cod_produto.in_(codigos),
                Separacao.sincronizado_nf == False,
                Separacao.qtd_saldo > 0
            ).scalar() or Decimal('0')

            # 4. Pedidos programados (separações com expedição no mês atual)
            pedidos_programados = db.session.query(
                func.sum(Separacao.qtd_saldo)
            ).filter(
                Separacao.cod_produto.in_(codigos),
                Separacao.sincronizado_nf == False,
                extract('month', Separacao.expedicao) == self.mes_atual,
                extract('year', Separacao.expedicao) == self.ano_atual,
                Separacao.qtd_saldo > 0
            ).scalar() or Decimal('0')

            # Cálculo final da necessidade
            # Saldo carteira já desconta separações (evita duplicação)
            saldo_carteira_liquido = float(saldo_carteira) - float(separacoes_nao_faturadas)
            if saldo_carteira_liquido < 0:
                saldo_carteira_liquido = 0

            necessidade = float(saldo_previsao) + saldo_carteira_liquido + float(pedidos_programados)

            return max(necessidade, 0)

        except Exception as e:
            logger.error(f"Erro ao calcular necessidade ruptura para {codigos}: {e}")
            return 0.0

    def _calcular_chegadas_mes(self, codigos: List[str], mes: int, ano: int) -> float:
        """
        Calcula chegadas previstas para um mês específico

        Fonte: PedidoCompras.data_pedido_previsao no mês/ano
        Considera apenas pedidos não cancelados e não concluídos
        """
        try:
            # Saldo a receber = qtd_produto_pedido - qtd_recebida
            resultado = db.session.query(
                func.sum(PedidoCompras.qtd_produto_pedido - func.coalesce(PedidoCompras.qtd_recebida, 0))
            ).filter(
                PedidoCompras.cod_produto.in_(codigos),
                extract('month', PedidoCompras.data_pedido_previsao) == mes,
                extract('year', PedidoCompras.data_pedido_previsao) == ano,
                PedidoCompras.status_odoo.notin_(['cancel', 'done']),
                PedidoCompras.importado_odoo == True
            ).scalar()

            return max(float(resultado or 0), 0)

        except Exception as e:
            logger.error(f"Erro ao calcular chegadas para {codigos} em {mes}/{ano}: {e}")
            return 0.0

    def _calcular_necessidade_mes(self, codigos: List[str], mes: int, ano: int) -> float:
        """
        Calcula necessidade para um mês específico (futuro)

        Fórmula: previsao_demanda.qtd_demanda_prevista + Separacao.qtd_saldo (mês/ano)

        Para meses futuros não considera demanda realizada nem saldo carteira
        """
        try:
            # 1. Previsão de demanda para o mês
            previsao = db.session.query(
                func.sum(PrevisaoDemanda.qtd_demanda_prevista)
            ).filter(
                PrevisaoDemanda.cod_produto.in_(codigos),
                PrevisaoDemanda.data_mes == mes,
                PrevisaoDemanda.data_ano == ano
            ).scalar() or Decimal('0')

            # 2. Separações programadas para o mês
            separacoes = db.session.query(
                func.sum(Separacao.qtd_saldo)
            ).filter(
                Separacao.cod_produto.in_(codigos),
                Separacao.sincronizado_nf == False,
                extract('month', Separacao.expedicao) == mes,
                extract('year', Separacao.expedicao) == ano,
                Separacao.qtd_saldo > 0
            ).scalar() or Decimal('0')

            return float(previsao) + float(separacoes)

        except Exception as e:
            logger.error(f"Erro ao calcular necessidade para {codigos} em {mes}/{ano}: {e}")
            return 0.0

    def exportar_para_excel(self, dados: Dict[str, Any]) -> bytes:
        """
        Exporta os dados da projeção para Excel

        Args:
            dados: Dicionário retornado por calcular_projecao_macro()

        Returns:
            Bytes do arquivo Excel
        """
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Macro Projeção Componentes"

        # Estilos
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        subheader_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        subheader_font = Font(color="FFFFFF", bold=True, size=9)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        negative_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        positive_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

        # Cabeçalho principal
        mes_atual_nome = self._nome_mes(dados.get('mes_atual', 1))
        mes_proximo_nome = self._nome_mes(dados.get('mes_proximo', 2))

        headers_principais = [
            "Código", "Produto",
            "RUPTURA", "", "",
            f"SALDO {mes_atual_nome}/{dados.get('ano_atual', 2026)}", "", "",
            f"SALDO {mes_proximo_nome}/{dados.get('ano_proximo', 2026)}", "", ""
        ]

        # Linha 1: Grupos
        ws.merge_cells('A1:B1')
        ws['A1'] = "IDENTIFICAÇÃO"
        ws['A1'].fill = header_fill
        ws['A1'].font = header_font
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['B1'].fill = header_fill

        ws.merge_cells('C1:E1')
        ws['C1'] = "RUPTURA"
        ws['C1'].fill = header_fill
        ws['C1'].font = header_font
        ws['C1'].alignment = Alignment(horizontal='center')
        for col in ['D1', 'E1']:
            ws[col].fill = header_fill

        ws.merge_cells('F1:H1')
        ws['F1'] = f"SALDO DEMANDA {mes_atual_nome}/{dados.get('ano_atual', 2026)}"
        ws['F1'].fill = header_fill
        ws['F1'].font = header_font
        ws['F1'].alignment = Alignment(horizontal='center')
        for col in ['G1', 'H1']:
            ws[col].fill = header_fill

        ws.merge_cells('I1:K1')
        ws['I1'] = f"SALDO DEMANDA {mes_proximo_nome}/{dados.get('ano_proximo', 2026)}"
        ws['I1'].fill = header_fill
        ws['I1'].font = header_font
        ws['I1'].alignment = Alignment(horizontal='center')
        for col in ['J1', 'K1']:
            ws[col].fill = header_fill

        # Linha 2: Sub-cabeçalhos
        subheaders = [
            "Código", "Produto",
            "Estoque", "Necessidade", "Saldo",
            "Chegada", "Necessidade", "Saldo",
            "Chegada", "Necessidade", "Saldo"
        ]

        for col_idx, header in enumerate(subheaders, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.fill = subheader_fill
            cell.font = subheader_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Dados
        for row_idx, comp in enumerate(dados.get('componentes', []), 3):
            ruptura = comp.get('ruptura', {})
            mes_atual = comp.get('saldo_mes_atual', {})
            mes_proximo = comp.get('saldo_mes_proximo', {})

            ws.cell(row=row_idx, column=1, value=comp.get('cod_produto', '')).border = thin_border
            ws.cell(row=row_idx, column=2, value=comp.get('nome_produto', '')).border = thin_border

            # Ruptura
            ws.cell(row=row_idx, column=3, value=ruptura.get('estoque_atual', 0)).border = thin_border
            ws.cell(row=row_idx, column=4, value=ruptura.get('necessidade', 0)).border = thin_border
            saldo_ruptura_cell = ws.cell(row=row_idx, column=5, value=ruptura.get('saldo', 0))
            saldo_ruptura_cell.border = thin_border
            if ruptura.get('saldo', 0) < 0:
                saldo_ruptura_cell.fill = negative_fill
            elif ruptura.get('saldo', 0) > 0:
                saldo_ruptura_cell.fill = positive_fill

            # Mês Atual
            ws.cell(row=row_idx, column=6, value=mes_atual.get('chegada', 0)).border = thin_border
            ws.cell(row=row_idx, column=7, value=mes_atual.get('necessidade', 0)).border = thin_border
            saldo_atual_cell = ws.cell(row=row_idx, column=8, value=mes_atual.get('saldo', 0))
            saldo_atual_cell.border = thin_border
            if mes_atual.get('saldo', 0) < 0:
                saldo_atual_cell.fill = negative_fill
            elif mes_atual.get('saldo', 0) > 0:
                saldo_atual_cell.fill = positive_fill

            # Mês Próximo
            ws.cell(row=row_idx, column=9, value=mes_proximo.get('chegada', 0)).border = thin_border
            ws.cell(row=row_idx, column=10, value=mes_proximo.get('necessidade', 0)).border = thin_border
            saldo_proximo_cell = ws.cell(row=row_idx, column=11, value=mes_proximo.get('saldo', 0))
            saldo_proximo_cell.border = thin_border
            if mes_proximo.get('saldo', 0) < 0:
                saldo_proximo_cell.fill = negative_fill
            elif mes_proximo.get('saldo', 0) > 0:
                saldo_proximo_cell.fill = positive_fill

        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 40
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            ws.column_dimensions[col].width = 14

        # Salvar em memória
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    def _nome_mes(self, mes: int) -> str:
        """Retorna nome do mês em português"""
        meses = {
            1: "JAN", 2: "FEV", 3: "MAR", 4: "ABR",
            5: "MAI", 6: "JUN", 7: "JUL", 8: "AGO",
            9: "SET", 10: "OUT", 11: "NOV", 12: "DEZ"
        }
        return meses.get(mes, "???")

    def get_opcoes_filtros(self) -> Dict[str, List[str]]:
        """
        Retorna valores únicos para os filtros dropdown

        Returns:
            Dict com listas de valores únicos por campo
        """
        try:
            # Buscar valores únicos apenas de produtos comprados ativos
            base_query = db.session.query(CadastroPalletizacao).filter(
                CadastroPalletizacao.produto_comprado == True,
                CadastroPalletizacao.ativo == True
            )

            # Categorias
            categorias = db.session.query(
                CadastroPalletizacao.categoria_produto
            ).filter(
                CadastroPalletizacao.produto_comprado == True,
                CadastroPalletizacao.ativo == True,
                CadastroPalletizacao.categoria_produto.isnot(None),
                CadastroPalletizacao.categoria_produto != ''
            ).distinct().order_by(CadastroPalletizacao.categoria_produto).all()

            # Tipos de matéria-prima
            tipos_mp = db.session.query(
                CadastroPalletizacao.tipo_materia_prima
            ).filter(
                CadastroPalletizacao.produto_comprado == True,
                CadastroPalletizacao.ativo == True,
                CadastroPalletizacao.tipo_materia_prima.isnot(None),
                CadastroPalletizacao.tipo_materia_prima != ''
            ).distinct().order_by(CadastroPalletizacao.tipo_materia_prima).all()

            # Linhas de produção
            linhas = db.session.query(
                CadastroPalletizacao.linha_producao
            ).filter(
                CadastroPalletizacao.produto_comprado == True,
                CadastroPalletizacao.ativo == True,
                CadastroPalletizacao.linha_producao.isnot(None),
                CadastroPalletizacao.linha_producao != ''
            ).distinct().order_by(CadastroPalletizacao.linha_producao).all()

            return {
                'categorias': [c[0] for c in categorias],
                'tipos_materia_prima': [t[0] for t in tipos_mp],
                'linhas_producao': [l[0] for l in linhas]
            }

        except Exception as e:
            logger.error(f"Erro ao buscar opções de filtros: {e}")
            return {
                'categorias': [],
                'tipos_materia_prima': [],
                'linhas_producao': []
            }
