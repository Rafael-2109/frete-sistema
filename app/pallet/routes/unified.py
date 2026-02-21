"""
Routes GET da Tela Unificada de Pallets V3

Endpoints:
- GET / → Renderiza template unico
- GET /api/kpis → 6 KPIs + alertas
- GET /api/tabela → Dados paginados com filtros
- GET /api/tabela/row/<nf_id> → HTML de 1 linha (partial update)
- GET /api/nf/<nf_id>/completo → Drill-down painel lateral
- GET /api/filtros/ufs → UFs distintas
- GET /api/filtros/cidades → Cidades por UF
- GET /api/filtros/responsaveis → Autocomplete responsavel
- GET /api/contadores → Contadores por aba
- GET /api/exportar → Exportar XLSX com filtros atuais
"""
import logging
from flask import Blueprint, render_template, request, jsonify, send_file

from app.pallet.services.unified_query_service import UnifiedQueryService

logger = logging.getLogger(__name__)

unified_bp = Blueprint('unified', __name__, url_prefix='')


# =========================================================================
# TEMPLATE
# =========================================================================

@unified_bp.route('/')
def index():
    """Renderiza a tela unificada de gestao de pallets V3."""
    # Carregar UFs para select de filtro (dados estaticos)
    ufs = UnifiedQueryService.listar_ufs()
    return render_template('pallet/v3/unified.html', ufs=ufs)


# =========================================================================
# API: KPIs
# =========================================================================

@unified_bp.route('/api/kpis')
def api_kpis():
    """Retorna os 6 KPIs + alertas em JSON."""
    try:
        kpis = UnifiedQueryService.obter_kpis()
        return jsonify({'sucesso': True, 'dados': kpis})
    except Exception as e:
        logger.error(f"Erro ao obter KPIs: {e}")
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500


# =========================================================================
# API: TABELA PAGINADA
# =========================================================================

@unified_bp.route('/api/tabela')
def api_tabela():
    """Retorna dados paginados da tabela principal."""
    try:
        # Extrair filtros dos query params
        filtros = {
            'busca': request.args.get('busca', ''),
            'status_nf': request.args.get('status_nf', ''),
            'status_credito': request.args.get('status_credito', ''),
            'empresa': request.args.get('empresa', ''),
            'tipo_destinatario': request.args.get('tipo_destinatario', ''),
            'cnpj': request.args.get('cnpj', ''),
            'data_de': request.args.get('data_de', ''),
            'data_ate': request.args.get('data_ate', ''),
            'uf': request.args.get('uf', ''),
            'cidade': request.args.get('cidade', ''),
            'aba': request.args.get('aba', 'visao_geral'),
            'vencido': request.args.get('vencido', ''),
            'docs_pendentes': request.args.get('docs_pendentes', ''),
        }

        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        ordenar_por = request.args.get('ordenar_por', 'data_emissao')
        ordem = request.args.get('ordem', 'desc')

        # Limitar per_page para evitar queries pesadas
        per_page = min(per_page, 100)

        resultado = UnifiedQueryService.listar_paginado(
            filtros=filtros,
            page=page,
            per_page=per_page,
            ordenar_por=ordenar_por,
            ordem=ordem
        )

        return jsonify({'sucesso': True, 'dados': resultado})

    except Exception as e:
        logger.error(f"Erro ao listar tabela: {e}")
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500


# =========================================================================
# API: LINHA INDIVIDUAL (partial update)
# =========================================================================

@unified_bp.route('/api/tabela/row/<int:nf_id>')
def api_tabela_row(nf_id):
    """Retorna dados de uma unica linha (para atualizar apos acao)."""
    try:
        linha = UnifiedQueryService.obter_linha(nf_id)
        if not linha:
            return jsonify({'sucesso': False, 'mensagem': 'NF nao encontrada'}), 404
        return jsonify({'sucesso': True, 'dados': linha})
    except Exception as e:
        logger.error(f"Erro ao obter linha NF #{nf_id}: {e}")
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500


# =========================================================================
# API: DRILL-DOWN (OFFCANVAS)
# =========================================================================

@unified_bp.route('/api/nf/<int:nf_id>/completo')
def api_nf_completo(nf_id):
    """Retorna dados completos de uma NF para o painel lateral."""
    try:
        dados = UnifiedQueryService.obter_completo(nf_id)
        if not dados:
            return jsonify({'sucesso': False, 'mensagem': 'NF nao encontrada'}), 404
        return jsonify({'sucesso': True, 'dados': dados})
    except Exception as e:
        logger.error(f"Erro ao obter NF completa #{nf_id}: {e}")
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500


# =========================================================================
# API: FILTROS DINAMICOS
# =========================================================================

@unified_bp.route('/api/filtros/ufs')
def api_filtros_ufs():
    """Lista UFs distintas para o select de filtro."""
    try:
        ufs = UnifiedQueryService.listar_ufs()
        return jsonify({'sucesso': True, 'dados': ufs})
    except Exception as e:
        logger.error(f"Erro ao listar UFs: {e}")
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500


@unified_bp.route('/api/filtros/cidades')
def api_filtros_cidades():
    """Lista cidades por UF para o select de filtro."""
    uf = request.args.get('uf', '').strip()
    if not uf:
        return jsonify({'sucesso': True, 'dados': []})

    try:
        cidades = UnifiedQueryService.listar_cidades_por_uf(uf)
        return jsonify({'sucesso': True, 'dados': cidades})
    except Exception as e:
        logger.error(f"Erro ao listar cidades da UF {uf}: {e}")
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500


@unified_bp.route('/api/filtros/responsaveis')
def api_filtros_responsaveis():
    """Autocomplete de responsaveis."""
    termo = request.args.get('termo', '').strip()
    if len(termo) < 2:
        return jsonify({'sucesso': True, 'dados': []})

    try:
        resultados = UnifiedQueryService.buscar_responsaveis(termo)
        return jsonify({'sucesso': True, 'dados': resultados})
    except Exception as e:
        logger.error(f"Erro ao buscar responsaveis: {e}")
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500


# =========================================================================
# API: CONTADORES POR ABA
# =========================================================================

@unified_bp.route('/api/contadores')
def api_contadores():
    """Retorna contagem de itens por aba."""
    try:
        contadores = UnifiedQueryService.contar_por_aba()
        return jsonify({'sucesso': True, 'dados': contadores})
    except Exception as e:
        logger.error(f"Erro ao contar por aba: {e}")
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500


# =========================================================================
# EXPORTAR XLSX
# =========================================================================

@unified_bp.route('/api/exportar')
def api_exportar():
    """Exporta dados filtrados da tabela em XLSX."""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from app.utils.timezone import agora_utc_naive

    try:
        # Mesmos filtros da tabela
        filtros = {
            'busca': request.args.get('busca', ''),
            'status_nf': request.args.get('status_nf', ''),
            'status_credito': request.args.get('status_credito', ''),
            'empresa': request.args.get('empresa', ''),
            'tipo_destinatario': request.args.get('tipo_destinatario', ''),
            'cnpj': request.args.get('cnpj', ''),
            'data_de': request.args.get('data_de', ''),
            'data_ate': request.args.get('data_ate', ''),
            'uf': request.args.get('uf', ''),
            'cidade': request.args.get('cidade', ''),
            'aba': request.args.get('aba', 'visao_geral'),
            'vencido': request.args.get('vencido', ''),
            'docs_pendentes': request.args.get('docs_pendentes', ''),
        }

        ordenar_por = request.args.get('ordenar_por', 'data_emissao')
        ordem = request.args.get('ordem', 'desc')

        # Buscar TODOS os itens (sem paginacao, limite seguro)
        resultado = UnifiedQueryService.listar_paginado(
            filtros=filtros,
            page=1,
            per_page=5000,
            ordenar_por=ordenar_por,
            ordem=ordem
        )

        itens = resultado.get('itens', [])

        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Pallets'

        # Estilos
        header_font = Font(bold=True, color='FFFFFF', size=10)
        header_fill = PatternFill(start_color='2B5797', end_color='2B5797', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0'),
        )

        # Cabecalho
        colunas = [
            ('NF Remessa', 14),
            ('Emissao', 12),
            ('Empresa', 10),
            ('Destinatario', 30),
            ('CNPJ Destinatario', 20),
            ('Tipo', 16),
            ('Quantidade', 12),
            ('Saldo', 10),
            ('Dom.A %', 10),
            ('Dom.A Resolvida', 14),
            ('Dom.B %', 10),
            ('Dom.B Resolvida', 14),
            ('Vencimento', 12),
            ('Vencido', 10),
            ('Docs Recebidos', 14),
            ('Total Docs', 12),
            ('Status NF', 12),
            ('Status Credito', 14),
        ]

        for col_idx, (titulo, largura) in enumerate(colunas, 1):
            cell = ws.cell(row=1, column=col_idx, value=titulo)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = largura

        # Dados
        for row_idx, item in enumerate(itens, 2):
            valores = [
                item.get('numero_nf', ''),
                item.get('data_emissao', ''),
                item.get('empresa', ''),
                item.get('nome_destinatario', ''),
                item.get('cnpj_destinatario', ''),
                item.get('tipo_destinatario', ''),
                item.get('quantidade', 0),
                item.get('qtd_saldo', 0),
                item.get('dom_a_pct', 0),
                item.get('dom_a_resolvida', 0),
                item.get('dom_b_pct', 0),
                item.get('dom_b_resolvida', 0),
                item.get('data_vencimento', ''),
                'Sim' if item.get('vencido') else 'Nao',
                item.get('docs_recebidos', 0),
                item.get('total_docs', 0),
                item.get('status', ''),
                item.get('credito_status', ''),
            ]
            for col_idx, valor in enumerate(valores, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                cell.border = thin_border

        # Auto-filtro
        if itens:
            ws.auto_filter.ref = f'A1:{ws.cell(row=1, column=len(colunas)).column_letter}{len(itens) + 1}'

        # Congelar cabecalho
        ws.freeze_panes = 'A2'

        # Gerar BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"pallets_{agora_utc_naive().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            output,
            download_name=filename,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f"Erro ao exportar XLSX: {e}")
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500
