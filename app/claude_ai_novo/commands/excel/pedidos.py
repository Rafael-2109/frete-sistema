#!/usr/bin/env python3
"""
Excel Pedidos - Mini esqueleto especializado para relatórios de pedidos
Versão otimizada integrada com BaseCommand
"""

from claude_ai_novo.commands.base_command import (
    BaseCommand, format_response_advanced, create_excel_summary,
    logging, datetime, db, current_user
)
from pathlib import Path

# Excel generation imports
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Sistema imports (com fallbacks)
try:
    from app.pedidos.models import Pedido
    from app.cotacao.models import CotacaoFrete
    from app.cadastros_agendamento.models import ContatoAgendamento
    from sqlalchemy import func, and_, or_
    SISTEMA_AVAILABLE = True
except ImportError:
    SISTEMA_AVAILABLE = False

logger = logging.getLogger(__name__)

class ExcelPedidos(BaseCommand):
    """Gerador de Excel especializado em pedidos e cotações"""
    
    def __init__(self):
        super().__init__()
        self.tipo = "pedidos"
        
    def is_excel_pedidos_command(self, consulta: str) -> bool:
        """Detecta comandos específicos de Excel para pedidos"""
        if not self._validate_input(consulta):
            return False
        
        keywords_pedidos = [
            'excel pedidos', 'planilha pedidos', 'relatório pedidos',
            'exportar pedidos', 'pedidos excel', 'pedidos em excel',
            'cotações excel', 'carteira excel', 'vendas excel'
        ]
        
        consulta_lower = consulta.lower()
        return any(keyword in consulta_lower for keyword in keywords_pedidos)
    
    def gerar_excel_pedidos(self, consulta: str, filtros=None) -> str:
        """Gera Excel de pedidos com análise de vendas"""
        
        if not EXCEL_AVAILABLE:
            return self._fallback_sem_openpyxl()
        
        if not SISTEMA_AVAILABLE:
            return self._fallback_sem_sistema()
        
        # Sanitizar entrada e extrair filtros avançados
        consulta = self._sanitize_input(consulta)
        if not filtros:
            filtros = self._extract_filters_advanced(consulta)
        
        # Log avançado
        self._log_command(consulta, "excel_pedidos", filtros)
        
        try:
            # Verificar cache primeiro
            cache_key = self._generate_cache_key("excel_pedidos", consulta, filtros)
            cached_result = self._get_cached_result(cache_key, 600)
            
            if cached_result:
                logger.info("✅ Excel pedidos encontrado em cache")
                return cached_result
            
            # Processar geração
            resultado = self._gerar_excel_pedidos_interno(consulta, filtros)
            
            # Armazenar em cache
            self._set_cached_result(cache_key, resultado, 600)
            
            return resultado
            
        except Exception as e:
            return self._handle_error(e, "excel_pedidos", f"Consulta: {consulta[:100]}")
    
    def _gerar_excel_pedidos_interno(self, consulta: str, filtros: dict) -> str:
        """Geração interna do Excel de pedidos"""
        
        # Buscar dados com filtros avançados
        pedidos = self._buscar_dados_pedidos(filtros)
        
        if not pedidos:
            return f"⚠️ **Nenhum pedido encontrado** para os critérios: {', '.join([f'{k}={v}' for k, v in filtros.items()])}"
        
        # Gerar arquivo Excel
        caminho_arquivo = self._criar_excel_pedidos(pedidos, filtros)
        
        # Criar estatísticas
        stats = self._create_summary_stats(pedidos, 'pedidos')
        
        # Criar resumo
        resumo = self._criar_resumo_pedidos(pedidos, filtros, caminho_arquivo, stats)
        
        return format_response_advanced(resumo, "ExcelPedidos", stats)
    
    def _buscar_dados_pedidos(self, filtros: dict) -> list:
        """Busca dados de pedidos com filtros avançados"""
        
        query = db.session.query(Pedido).join(ContatoAgendamento, isouter=True)
        
        # Aplicar filtros
        if filtros.get('cliente'):
            cliente = filtros['cliente']
            query = query.filter(
                or_(
                    Pedido.raz_social_red.ilike(f'%{cliente}%'),
                    Pedido.cnpj_cpf.ilike(f'%{cliente}%')
                )
            )
        
        if filtros.get('uf'):
            uf = filtros['uf']
            query = query.filter(Pedido.cod_uf == uf)
        
        if filtros.get('status'):
            status = filtros['status']
            if status == 'nf_cd':
                # Status NF no CD: flag nf_cd é True (PRIMEIRO - prioridade máxima)
                query = query.filter(Pedido.nf_cd == True)
            elif status == 'faturado':
                # Status FATURADO: Tem NF e não está no CD
                query = query.filter(
                    Pedido.nf.isnot(None),
                    Pedido.nf != '',
                    Pedido.nf_cd == False
                )
            elif status == 'embarcado':
                # Status EMBARCADO: Tem data_embarque mas não tem NF
                query = query.filter(
                    Pedido.data_embarque.isnot(None),
                    or_(Pedido.nf.is_(None), Pedido.nf == ''),
                    Pedido.nf_cd == False
                )
            elif status == 'cotado':
                # Status COTADO: Tem cotação mas não tem data_embarque e não está no CD
                query = query.filter(
                    Pedido.cotacao_id.isnot(None),
                    Pedido.data_embarque.is_(None),
                    Pedido.nf_cd == False
                )
            elif status == 'pendente' or status == 'aberto':
                # Status ABERTO: Não tem cotação e não está no CD
                query = query.filter(
                    Pedido.cotacao_id.is_(None),
                    Pedido.nf_cd == False
                )
        
        if filtros.get('data_inicio') and filtros.get('data_fim'):
            query = query.filter(
                and_(
                    Pedido.data_pedido >= filtros['data_inicio'],
                    Pedido.data_pedido <= filtros['data_fim']
                )
            )
        
        if filtros.get('valor_minimo'):
            query = query.filter(Pedido.valor_saldo_total >= filtros['valor_minimo'])
        
        # Ordenação inteligente
        query = query.order_by(Pedido.data_pedido.desc(), Pedido.id.desc())
        
        # Limitar resultados para performance
        return query.limit(5000).all()
    
    def _criar_excel_pedidos(self, pedidos: list, filtros: dict) -> str:
        """Cria arquivo Excel com dados dos pedidos"""
        
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)
        
        # Aba 1: Dados principais
        ws_dados = wb.create_sheet("Pedidos Detalhados")
        self._criar_aba_pedidos_principal(ws_dados, pedidos)
        
        # Aba 2: Análise de vendas
        ws_vendas = wb.create_sheet("Análise de Vendas")
        self._criar_aba_analise_vendas(ws_vendas, pedidos)
        
        # Aba 3: Status e agendamentos
        ws_status = wb.create_sheet("Status e Agendamentos")
        self._criar_aba_status_agendamentos(ws_status, pedidos)
        
        # Salvar arquivo
        nome_arquivo = f"pedidos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        caminho_completo = self.output_dir / nome_arquivo
        
        wb.save(caminho_completo)
        
        return f"/static/reports/{nome_arquivo}"
    
    def _criar_aba_pedidos_principal(self, ws, pedidos):
        """Cria aba principal com dados completos dos pedidos"""
        
        headers = [
            'Num Pedido', 'Cliente', 'CNPJ', 'UF', 'Cidade', 'Peso Total', 
            'Valor Pedido', 'Data Pedido', 'Data Entrega', 'Status Calculado',
            'Vendedor', 'Observações', 'Agendamento', 'Contato'
        ]
        
        # Aplicar headers com estilo
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._aplicar_estilo_header(cell)
        
        # Dados
        for row, pedido in enumerate(pedidos, 2):
            try:
                ws.cell(row=row, column=1, value=getattr(pedido, 'num_pedido', '') or '')
                ws.cell(row=row, column=2, value=getattr(pedido, 'raz_social_red', '') or '')
                ws.cell(row=row, column=3, value=getattr(pedido, 'cnpj_cpf', '') or '')
                ws.cell(row=row, column=4, value=getattr(pedido, 'cod_uf', '') or '')
                ws.cell(row=row, column=5, value=getattr(pedido, 'nome_cidade', '') or '')
                ws.cell(row=row, column=6, value=float(getattr(pedido, 'peso_total', 0) or 0))
                ws.cell(row=row, column=7, value=float(getattr(pedido, 'valor_saldo_total', 0) or 0))
                
                # Datas
                data_pedido = getattr(pedido, 'data_pedido', None)
                ws.cell(row=row, column=8, value=self._format_date_br(data_pedido))
                
                data_entrega = getattr(pedido, 'expedicao', None)
                ws.cell(row=row, column=9, value=self._format_date_br(data_entrega))
                
                ws.cell(row=row, column=10, value=getattr(pedido, 'status_calculado', '') or '')
                ws.cell(row=row, column=11, value=getattr(pedido, 'usuario_id', '') or '')
                ws.cell(row=row, column=12, value=getattr(pedido, 'observ_ped_1', '') or '')
                
                # Agendamento
                contato = getattr(pedido, 'contato_agendamento', None)
                if contato:
                    ws.cell(row=row, column=13, value=getattr(contato, 'forma', '') or '')
                    ws.cell(row=row, column=14, value=getattr(contato, 'contato', '') or '')
                else:
                    ws.cell(row=row, column=13, value='Não cadastrado')
                    ws.cell(row=row, column=14, value='N/A')
                
            except Exception as e:
                logger.warning(f"Erro ao processar pedido {getattr(pedido, 'num_pedido', 'N/A')}: {e}")
                continue
        
        # Auto-ajustar colunas
        self._auto_ajustar_colunas(ws)
    
    def _criar_aba_analise_vendas(self, ws, pedidos):
        """Cria aba com análise de vendas"""
        
        # Calcular totais
        valor_total = sum(float(getattr(p, 'valor_saldo_total', 0) or 0) for p in pedidos)
        peso_total = sum(float(getattr(p, 'peso_total', 0) or 0) for p in pedidos)
        
        # Headers
        ws.cell(row=1, column=1, value="ANÁLISE DE VENDAS - PEDIDOS")
        self._aplicar_estilo_header(ws.cell(row=1, column=1))
        
        # Resumo geral
        row = 3
        resumo_data = [
            ("Total de Pedidos", len(pedidos)),
            ("Valor Total", valor_total),
            ("Peso Total (kg)", peso_total),
            ("Valor Médio por Pedido", valor_total / len(pedidos) if pedidos else 0),
            ("Peso Médio por Pedido", peso_total / len(pedidos) if pedidos else 0),
            ("R$/kg Médio", valor_total / peso_total if peso_total > 0 else 0)
        ]
        
        for descricao, valor in resumo_data:
            ws.cell(row=row, column=1, value=descricao)
            ws.cell(row=row, column=2, value=valor)
            row += 1
        
        # Análise por status
        row += 2
        ws.cell(row=row, column=1, value="POR STATUS:")
        self._aplicar_estilo_header(ws.cell(row=row, column=1))
        row += 1
        
        status_count = {}
        for pedido in pedidos:
            status = getattr(pedido, 'status_calculado', 'N/A')
            status_count[status] = status_count.get(status, 0) + 1
        
        for status, count in status_count.items():
            ws.cell(row=row, column=1, value=status)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"{(count/len(pedidos)*100):.1f}%" if pedidos else "0%")
            row += 1
        
        self._auto_ajustar_colunas(ws)
    
    def _criar_aba_status_agendamentos(self, ws, pedidos):
        """Cria aba com análise de status e agendamentos"""
        
        headers = ['Status', 'Quantidade', 'Percentual', 'Valor Total', 'Com Agendamento']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._aplicar_estilo_header(cell)
        
        # Agrupar por status
        status_data = {}
        for pedido in pedidos:
            status = getattr(pedido, 'status_calculado', 'N/A')
            valor = float(getattr(pedido, 'valor_saldo_total', 0) or 0)
            tem_agendamento = bool(getattr(pedido, 'contato_agendamento', None))
            
            if status not in status_data:
                status_data[status] = {
                    'count': 0, 'valor': 0, 'com_agendamento': 0
                }
            
            status_data[status]['count'] += 1
            status_data[status]['valor'] += valor
            if tem_agendamento:
                status_data[status]['com_agendamento'] += 1
        
        # Dados por status
        row = 2
        total_pedidos = len(pedidos)
        for status, dados in sorted(status_data.items()):
            ws.cell(row=row, column=1, value=status)
            ws.cell(row=row, column=2, value=dados['count'])
            ws.cell(row=row, column=3, value=f"{(dados['count']/total_pedidos*100):.1f}%" if total_pedidos > 0 else "0%")
            ws.cell(row=row, column=4, value=dados['valor'])
            ws.cell(row=row, column=5, value=f"{dados['com_agendamento']}/{dados['count']}")
            row += 1
        
        self._auto_ajustar_colunas(ws)
    
    def _criar_resumo_pedidos(self, pedidos: list, filtros: dict, caminho_arquivo: str, stats: dict) -> str:
        """Cria resumo detalhado do relatório"""
        
        # Estatísticas
        valor_total = sum(float(getattr(p, 'valor_saldo_total', 0) or 0) for p in pedidos)
        peso_total = sum(float(getattr(p, 'peso_total', 0) or 0) for p in pedidos)
        
        # Contadores por status
        status_count = {}
        com_agendamento = 0
        for pedido in pedidos:
            status = getattr(pedido, 'status_calculado', 'N/A')
            status_count[status] = status_count.get(status, 0) + 1
            if getattr(pedido, 'contato_agendamento', None):
                com_agendamento += 1
        
        # Estados únicos
        estados = set(getattr(p, 'cod_uf', '') for p in pedidos if getattr(p, 'cod_uf', ''))
        
        resumo = f"""📊 **RELATÓRIO DE PEDIDOS GERADO COM SUCESSO!**

📋 **Detalhes do Relatório:**
• **Total de Pedidos:** {len(pedidos)}
• **Período:** {self._format_date_br(filtros.get('data_inicio'))} a {self._format_date_br(filtros.get('data_fim')) if filtros.get('data_fim') else 'Atual'}
• **Estados:** {', '.join(sorted(estados)) if estados else 'Todos'}
• **Cliente:** {filtros.get('cliente', 'Todos')}

💰 **ANÁLISE DE VENDAS:**
• **Valor Total:** {self._format_currency(valor_total)}
• **Peso Total:** {self._format_weight(peso_total)}
• **Valor Médio:** {self._format_currency(valor_total / len(pedidos) if pedidos else 0)}
• **Peso Médio:** {self._format_weight(peso_total / len(pedidos) if pedidos else 0)}
• **R$/kg Médio:** {self._format_currency(valor_total / peso_total if peso_total > 0 else 0)}

📊 **STATUS DOS PEDIDOS:**"""

        for status, count in sorted(status_count.items()):
            percentual = (count / len(pedidos)) * 100 if pedidos else 0
            resumo += f"\n• **{status}:** {count} ({percentual:.1f}%)"

        resumo += f"""

📅 **AGENDAMENTOS:**
• **Com Agendamento:** {com_agendamento}/{len(pedidos)} ({(com_agendamento/len(pedidos)*100):.1f}% if pedidos else 0%)
• **Sem Agendamento:** {len(pedidos) - com_agendamento}/{len(pedidos)}

📁 **ARQUIVO EXCEL:**
**[📊 BAIXAR RELATÓRIO COMPLETO]({caminho_arquivo})**

📈 **ABAS DISPONÍVEIS:**
• **Pedidos Detalhados** - Dados completos com 14 colunas
• **Análise de Vendas** - Resumo financeiro e médias
• **Status e Agendamentos** - Performance por status"""

        return resumo
    
    def _aplicar_estilo_header(self, cell):
        """Aplica estilo ao header"""
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )
    
    def _auto_ajustar_colunas(self, ws):
        """Auto-ajusta largura das colunas"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _fallback_sem_openpyxl(self) -> str:
        """Fallback quando openpyxl não está disponível"""
        return format_response_advanced("""❌ **Excel não disponível**

⚠️ **Problema:** Biblioteca openpyxl não instalada

💡 **Solução:**
```bash
pip install openpyxl
```

📊 **Funcionalidades Excel de Pedidos:**
• Relatórios detalhados com 14 colunas
• Análise de vendas completa
• Agrupamento por status
• Indicadores de agendamento
• Performance por UF e cliente""", "ExcelPedidos")
    
    def _fallback_sem_sistema(self) -> str:
        """Fallback quando sistema não está disponível"""
        return format_response_advanced("""⚠️ **Sistema não disponível**

🔧 **Problema:** Modelos Flask não carregados

💡 **Excel de Pedidos disponível quando sistema ativo:**
• Dados de pedidos e cotações
• Status calculado automaticamente
• Agendamentos de entrega
• Análise de vendas por período""", "ExcelPedidos")

# Instância global
_excel_pedidos = None

def get_excel_pedidos():
    """Retorna instância de ExcelPedidos"""
    global _excel_pedidos
    if _excel_pedidos is None:
        _excel_pedidos = ExcelPedidos()
    return _excel_pedidos 