#!/usr/bin/env python3
"""
Excel Faturamento - Mini esqueleto especializado para relatórios de faturamento
Versão otimizada integrada com BaseCommand
"""

from app.claude_ai_novo.commands.base_command import (
    BaseCommand, format_response_advanced, create_excel_summary,
    logging, datetime, current_user
)
from pathlib import Path
from typing import Dict, Optional, List

# Excel generation imports
try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    from unittest.mock import Mock
    Workbook = Mock()
    OPENPYXL_AVAILABLE = False
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Sistema imports (com fallbacks)
try:
    from app.faturamento.models import RelatorioFaturamentoImportado
except Exception as e:
    logger.error(f'Erro: {e}')
    pass
try:
    from sqlalchemy import func, and_, or_
    SQLALCHEMY_AVAILABLE = True
except ImportError:
    func, and_, or_ = None
    SQLALCHEMY_AVAILABLE = False
    SISTEMA_AVAILABLE = True
except ImportError:
    SISTEMA_AVAILABLE = False

logger = logging.getLogger(__name__)

class ExcelFaturamento(BaseCommand):
    """Gerador de relatórios Excel para faturamento"""
    
    @property
    def db(self):
        """Obtém db com fallback"""
        if not hasattr(self, "_db"):
            from app.claude_ai_novo.utils.flask_fallback import get_db
            self._db = get_db()
        return self._db
    
    def __init__(self):
        super().__init__()
        self.tipo = "faturamento"
        
    def is_excel_faturamento_command(self, consulta: str) -> bool:
        """Detecta comandos específicos de Excel para faturamento"""
        if not self._validate_input(consulta):
            return False
        
        keywords_faturamento = [
            'excel faturamento', 'planilha faturamento', 'relatório faturamento',
            'exportar faturamento', 'faturamento excel', 'faturamento em excel',
            'nf excel', 'nota fiscal excel', 'faturas excel', 'receitas excel'
        ]
        
        consulta_lower = consulta.lower()
        return any(keyword in consulta_lower for keyword in keywords_faturamento)
    
    def gerar_excel_faturamento(self, consulta: str, filtros=None) -> str:
        """Gera Excel de faturamento com análise financeira"""
        
        if not EXCEL_AVAILABLE:
            return self._fallback_sem_openpyxl()
        
        if not SISTEMA_AVAILABLE:
            return self._fallback_sem_sistema()
        
        # Sanitizar entrada e extrair filtros avançados
        consulta = self._sanitize_input(consulta)
        if not filtros:
            filtros = self._extract_filters_advanced(consulta)
        
        # Log avançado
        self._log_command(consulta, "excel_faturamento", filtros)
        
        try:
            # Verificar cache primeiro
            cache_key = self._generate_cache_key("excel_faturamento", consulta, filtros)
            cached_result = self._get_cached_result(cache_key, 600)
            
            if cached_result:
                logger.info("✅ Excel faturamento encontrado em cache")
                return cached_result
            
            # Processar geração
            resultado = self._gerar_excel_faturamento_interno(consulta, filtros)
            
            # Armazenar em cache
            self._set_cached_result(cache_key, resultado, 600)
            
            return resultado
            
        except Exception as e:
            return self._handle_error(e, "excel_faturamento", f"Consulta: {consulta[:100]}")
    
    def _gerar_excel_faturamento_interno(self, consulta: str, filtros: dict) -> str:
        """Geração interna do Excel de faturamento"""
        
        # Buscar dados com filtros avançados
        faturamento = self._buscar_dados_faturamento(filtros)
        
        if not faturamento:
            return f"⚠️ **Nenhum faturamento encontrado** para os critérios: {', '.join([f'{k}={v}' for k, v in filtros.items()])}"
        
        # Gerar arquivo Excel
        caminho_arquivo = self._criar_excel_faturamento(faturamento, filtros)
        
        # Criar estatísticas
        stats = self._create_summary_stats(faturamento, 'faturamento')
        
        # Criar resumo
        resumo = self._criar_resumo_faturamento(faturamento, filtros, caminho_arquivo, stats)
        
        return format_response_advanced(resumo, "ExcelFaturamento", stats)

    def _buscar_dados_faturamento(self, filtros: dict) -> list:
        """Busca dados de faturamento com filtros avançados"""
        
        query = self.db.session.query(RelatorioFaturamentoImportado)
        
        # Aplicar filtros
        if filtros.get('cliente'):
            cliente = filtros['cliente']
            query = query.filter(RelatorioFaturamentoImportado.nome_cliente.ilike(f'%{cliente}%'))
        
        if filtros.get('data_inicio') and filtros.get('data_fim'):
            query = query.filter(
                and_(
                    RelatorioFaturamentoImportado.data_fatura >= filtros['data_inicio'],
                    RelatorioFaturamentoImportado.data_fatura <= filtros['data_fim']
                )
            )
        
        if filtros.get('valor_minimo'):
            query = query.filter(RelatorioFaturamentoImportado.valor_total >= filtros['valor_minimo'])
        
        # Ordenação inteligente
        query = query.order_by(RelatorioFaturamentoImportado.data_fatura.desc(), RelatorioFaturamentoImportado.id.desc())
        
        # Limitar resultados para performance
        return query.limit(5000).all()

    def _criar_excel_faturamento(self, faturamento: list, filtros: dict) -> str:
        """Cria arquivo Excel com dados do faturamento"""
        
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)
        
        # Aba 1: Dados principais
        ws_dados = wb.create_sheet("Faturamento Detalhado")
        self._criar_aba_faturamento_principal(ws_dados, faturamento)
        
        # Aba 2: Análise financeira
        ws_financeiro = wb.create_sheet("Análise Financeira")
        self._criar_aba_analise_financeira(ws_financeiro, faturamento)
        
        # Salvar arquivo
        nome_arquivo = f"faturamento_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        caminho_completo = self.output_dir / nome_arquivo
        
        wb.save(caminho_completo)
        
        return f"/static/reports/{nome_arquivo}"

    def _criar_aba_faturamento_principal(self, ws, faturamento):
        """Cria aba principal com dados completos do faturamento"""
        
        headers = [
            'Número NF', 'Cliente', 'CNPJ', 'Valor Total', 'Data Fatura',
            'Origem', 'Incoterm', 'Status', 'Valor Líquido', 'Impostos'
        ]
        
        # Aplicar headers com estilo
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._aplicar_estilo_header(cell)
        
        # Dados
        for row, fat in enumerate(faturamento, 2):
            try:
                ws.cell(row=row, column=1, value=getattr(fat, 'numero_nf', '') or '')
                ws.cell(row=row, column=2, value=getattr(fat, 'nome_cliente', '') or '')
                ws.cell(row=row, column=3, value=getattr(fat, 'cnpj_cliente', '') or '')
                ws.cell(row=row, column=4, value=float(getattr(fat, 'valor_total', 0) or 0))
                
                data_fatura = getattr(fat, 'data_fatura', None)
                ws.cell(row=row, column=5, value=self._format_date_br(data_fatura))
                
                ws.cell(row=row, column=6, value=getattr(fat, 'origem', '') or '')
                ws.cell(row=row, column=7, value=getattr(fat, 'incoterm', '') or '')
                ws.cell(row=row, column=8, value=getattr(fat, 'status', '') or '')
                
                # Campos adicionais
                valor_liquido = float(getattr(fat, 'valor_liquido', 0) or 0)
                valor_total = float(getattr(fat, 'valor_total', 0) or 0)
                impostos = valor_total - valor_liquido
                
                ws.cell(row=row, column=9, value=valor_liquido)
                ws.cell(row=row, column=10, value=impostos)
                
            except Exception as e:
                logger.warning(f"Erro ao processar faturamento {getattr(fat, 'numero_nf', 'N/A')}: {e}")
                continue
        
        # Auto-ajustar colunas
        self._auto_ajustar_colunas(ws)

    def _criar_aba_analise_financeira(self, ws, faturamento):
        """Cria aba com análise financeira"""
        
        # Calcular totais
        valor_total = sum(float(getattr(f, 'valor_total', 0) or 0) for f in faturamento)
        valor_liquido = sum(float(getattr(f, 'valor_liquido', 0) or 0) for f in faturamento)
        impostos = valor_total - valor_liquido
        
        # Headers
        ws.cell(row=1, column=1, value="ANÁLISE FINANCEIRA - FATURAMENTO")
        self._aplicar_estilo_header(ws.cell(row=1, column=1))
        
        # Resumo geral
        row = 3
        resumo_data = [
            ("Total de Notas Fiscais", len(faturamento)),
            ("", ""),
            ("Valor Total Bruto", valor_total),
            ("Valor Total Líquido", valor_liquido),
            ("Total de Impostos", impostos),
            ("", ""),
            ("Faturamento Médio", valor_total / len(faturamento) if faturamento else 0),
            ("Margem Líquida %", (valor_liquido / valor_total * 100) if valor_total > 0 else 0),
            ("Carga Tributária %", (impostos / valor_total * 100) if valor_total > 0 else 0)
        ]
        
        for descricao, valor in resumo_data:
            ws.cell(row=row, column=1, value=descricao)
            if isinstance(valor, (int, float)):
                ws.cell(row=row, column=2, value=valor)
            row += 1
        
        self._auto_ajustar_colunas(ws)

    def _criar_resumo_faturamento(self, faturamento: list, filtros: dict, caminho_arquivo: str, stats: dict) -> str:
        """Cria resumo detalhado do relatório"""
        
        # Estatísticas financeiras
        valor_total = sum(float(getattr(f, 'valor_total', 0) or 0) for f in faturamento)
        valor_liquido = sum(float(getattr(f, 'valor_liquido', 0) or 0) for f in faturamento)
        impostos = valor_total - valor_liquido
        
        # Clientes únicos
        clientes = set(getattr(f, 'nome_cliente', '') for f in faturamento if getattr(f, 'nome_cliente', ''))
        
        resumo = f"""💰 **RELATÓRIO DE FATURAMENTO GERADO COM SUCESSO!**

📋 **Detalhes do Relatório:**
• **Total de Notas Fiscais:** {len(faturamento)}
• **Período:** {self._format_date_br(filtros.get('data_inicio'))} a {self._format_date_br(filtros.get('data_fim')) if filtros.get('data_fim') else 'Atual'}
• **Cliente:** {filtros.get('cliente', 'Todos')}

💰 **ANÁLISE FINANCEIRA COMPLETA:**
• **Valor Total Bruto:** {self._format_currency(valor_total)}
• **Valor Total Líquido:** {self._format_currency(valor_liquido)}
• **Total de Impostos:** {self._format_currency(impostos)}

📊 **INDICADORES:**
• **Faturamento Médio:** {self._format_currency(valor_total / len(faturamento) if faturamento else 0)}
• **Margem Líquida:** {self._format_percentage(valor_liquido / valor_total * 100 if valor_total > 0 else 0)}
• **Carga Tributária:** {self._format_percentage(impostos / valor_total * 100 if valor_total > 0 else 0)}

🏢 **CLIENTES:**
• **Clientes únicos:** {len(clientes)}
• **Faturamento por cliente:** {self._format_currency(valor_total / len(clientes) if clientes else 0)}

📁 **ARQUIVO EXCEL:**
**[📊 BAIXAR RELATÓRIO COMPLETO]({caminho_arquivo})**

📈 **ABAS DISPONÍVEIS:**
• **Faturamento Detalhado** - Dados completos com 10 colunas
• **Análise Financeira** - Resumo e indicadores tributários"""

        return resumo

    def _aplicar_estilo_header(self, cell):
        """Aplica estilo ao header"""
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )
    
    def _auto_ajustar_colunas(self, ws):
        """Auto-ajusta largura das colunas"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _fallback_sem_openpyxl(self) -> str:
        """Fallback quando openpyxl não está disponível"""
        return format_response_advanced("""❌ **Excel não disponível**

⚠️ **Problema:** Biblioteca openpyxl não instalada

💡 **Solução:**
```bash
pip install openpyxl
```

📊 **Funcionalidades Excel de Faturamento:**
• Relatórios detalhados com 10 colunas
• Análise financeira completa
• Indicadores tributários
• Margem líquida e carga tributária
• Agrupamento por cliente""", "ExcelFaturamento")
    
    def _fallback_sem_sistema(self) -> str:
        """Fallback quando sistema não está disponível"""
        return format_response_advanced("""⚠️ **Sistema não disponível**

🔧 **Problema:** Modelos Flask não carregados

💡 **Excel de Faturamento disponível quando sistema ativo:**
• Dados de notas fiscais
• Análise financeira completa
• Indicadores tributários
• Performance por cliente/período""", "ExcelFaturamento")

# Instância global
_excel_faturamento = None

def get_excel_faturamento():
    """Retorna instância do ExcelFaturamento"""
    global _excel_faturamento
    if _excel_faturamento is None:
        _excel_faturamento = ExcelFaturamento()
    return _excel_faturamento 