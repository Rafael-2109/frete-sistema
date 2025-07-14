#!/usr/bin/env python3
"""
Excel Fretes - Mini esqueleto especializado para relatórios de fretes
Versão otimizada integrada com BaseCommand
"""

from app.claude_ai_novo.commands.base_command import (
    BaseCommand, format_response_advanced, create_excel_summary,
    logging, datetime, current_user
)
from pathlib import Path

# Excel generation imports
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Sistema imports (com fallbacks)
try:
    from app.fretes.models import Frete, DespesaExtra
    from app.transportadoras.models import Transportadora
    from app.embarques.models import Embarque
    from sqlalchemy import func, and_, or_
    SISTEMA_AVAILABLE = True
except ImportError:
    SISTEMA_AVAILABLE = False

logger = logging.getLogger(__name__)

class ExcelFretes(BaseCommand):
    """Gerador de relatórios Excel para fretes"""
    
    @property
    def db(self):
        """Obtém db com fallback"""
        if not hasattr(self, "_db"):
            from app.claude_ai_novo.utils.flask_fallback import get_db
            self._db = get_db()
        return self._db
    
    def __init__(self):
        super().__init__()
        self.tipo = "fretes"
        
    def is_excel_fretes_command(self, consulta: str) -> bool:
        """Detecta comandos específicos de Excel para fretes"""
        if not self._validate_input(consulta):
            return False
        
        keywords_fretes = [
            'excel fretes', 'planilha fretes', 'relatório fretes',
            'exportar fretes', 'fretes excel', 'fretes em excel',
            'transportadoras excel', 'cotações excel', 'cte excel',
            'freteiros excel', 'despesas extras excel'
        ]
        
        consulta_lower = consulta.lower()
        return any(keyword in consulta_lower for keyword in keywords_fretes)
    
    def gerar_excel_fretes(self, consulta: str, filtros=None) -> str:
        """Gera Excel de fretes com análise financeira completa"""
        
        if not EXCEL_AVAILABLE:
            return self._fallback_sem_openpyxl()
        
        if not SISTEMA_AVAILABLE:
            return self._fallback_sem_sistema()
        
        # Sanitizar entrada e extrair filtros avançados
        consulta = self._sanitize_input(consulta)
        if not filtros:
            filtros = self._extract_filters_advanced(consulta)
        
        # Log avançado
        self._log_command(consulta, "excel_fretes", filtros)
        
        try:
            # Verificar cache primeiro
            cache_key = self._generate_cache_key("excel_fretes", consulta, filtros)
            cached_result = self._get_cached_result(cache_key, 600)
            
            if cached_result:
                logger.info("✅ Excel fretes encontrado em cache")
                return cached_result
            
            # Processar geração
            resultado = self._gerar_excel_fretes_interno(consulta, filtros)
            
            # Armazenar em cache
            self._set_cached_result(cache_key, resultado, 600)
            
            return resultado
            
        except Exception as e:
            return self._handle_error(e, "excel_fretes", f"Consulta: {consulta[:100]}")
    
    def _gerar_excel_fretes_interno(self, consulta: str, filtros: dict) -> str:
        """Geração interna do Excel de fretes"""
        
        # Buscar dados com filtros avançados
        fretes = self._buscar_dados_fretes(filtros)
        
        if not fretes:
            return f"⚠️ **Nenhum frete encontrado** para os critérios: {', '.join([f'{k}={v}' for k, v in filtros.items()])}"
        
        # Gerar arquivo Excel
        caminho_arquivo = self._criar_excel_fretes(fretes, filtros)
        
        # Criar estatísticas
        stats = self._create_summary_stats(fretes, 'fretes')
        
        # Criar resumo com estatísticas financeiras
        resumo = self._criar_resumo_fretes(fretes, filtros, caminho_arquivo, stats)
        
        return format_response_advanced(resumo, "ExcelFretes", stats)
    
    def _buscar_dados_fretes(self, filtros: dict) -> list:
        """Busca dados de fretes com filtros avançados"""
        
        query = self.db.session.query(Frete).join(Transportadora, isouter=True)
        
        # Aplicar filtros
        if filtros.get('cliente'):
            cliente = filtros['cliente']
            query = query.filter(
                or_(
                    Frete.nome_cliente.ilike(f'%{cliente}%'),
                    Frete.razao_social_cliente.ilike(f'%{cliente}%')
                )
            )
        
        if filtros.get('uf'):
            uf = filtros['uf']
            query = query.filter(Frete.uf_destino == uf)
        
        if filtros.get('status'):
            status = filtros['status']
            if status == 'pendente':
                query = query.filter(Frete.numero_cte.is_(None))
            elif status == 'aprovado':
                query = query.filter(Frete.numero_cte.isnot(None))
        
        if filtros.get('data_inicio') and filtros.get('data_fim'):
            query = query.filter(
                and_(
                    Frete.data_embarque >= filtros['data_inicio'],
                    Frete.data_embarque <= filtros['data_fim']
                )
            )
        
        if filtros.get('valor_minimo'):
            query = query.filter(Frete.valor_cotado >= filtros['valor_minimo'])
        
        # Ordenação inteligente
        query = query.order_by(Frete.data_embarque.desc(), Frete.id.desc())
        
        # Limitar resultados para performance
        return query.limit(5000).all()
    
    def _criar_excel_fretes(self, fretes: list, filtros: dict) -> str:
        """Cria arquivo Excel com dados dos fretes"""
        
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)
        
        # Aba 1: Dados principais
        ws_dados = wb.create_sheet("Fretes Detalhados")
        self._criar_aba_fretes_principal(ws_dados, fretes)
        
        # Aba 2: Análise financeira
        ws_financeiro = wb.create_sheet("Análise Financeira")
        self._criar_aba_analise_financeira(ws_financeiro, fretes)
        
        # Aba 3: Transportadoras
        ws_transportadoras = wb.create_sheet("Por Transportadora")
        self._criar_aba_transportadoras(ws_transportadoras, fretes)
        
        # Salvar arquivo
        nome_arquivo = f"fretes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        caminho_completo = self.output_dir / nome_arquivo
        
        wb.save(caminho_completo)
        
        return f"/static/reports/{nome_arquivo}"
    
    def _criar_aba_fretes_principal(self, ws, fretes):
        """Cria aba principal com dados completos dos fretes"""
        
        # Headers expandidos com campos solicitados
        headers = [
            'ID', 'Cliente', 'CNPJ', 'Destino', 'UF', 'Peso (kg)', 
            'Valor Cotado', 'Valor Considerado', 'Valor CTe', 'Valor Pago',
            'R$/kg Cotado', 'R$/kg Considerado', 'R$/kg CTe', 'R$/kg Pago',
            'Transportadora', 'É Freteiro', 'Data Embarque', 'CTe', 'Status'
        ]
        
        # Aplicar headers com estilo
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._aplicar_estilo_header(cell)
        
        # Dados
        for row, frete in enumerate(fretes, 2):
            try:
                # Dados básicos
                ws.cell(row=row, column=1, value=frete.id)
                ws.cell(row=row, column=2, value=getattr(frete, 'nome_cliente', '') or '')
                ws.cell(row=row, column=3, value=getattr(frete, 'cnpj_cliente', '') or '')
                ws.cell(row=row, column=4, value=getattr(frete, 'cidade_destino', '') or '')
                ws.cell(row=row, column=5, value=getattr(frete, 'uf_destino', '') or '')
                ws.cell(row=row, column=6, value=float(getattr(frete, 'peso_total', 0) or 0))
                
                # Valores financeiros
                valor_cotado = float(getattr(frete, 'valor_cotado', 0) or 0)
                valor_considerado = float(getattr(frete, 'valor_considerado', 0) or 0)
                valor_cte = float(getattr(frete, 'valor_cte', 0) or 0)
                valor_pago = float(getattr(frete, 'valor_pago', 0) or 0)
                peso = float(getattr(frete, 'peso_total', 0) or 0)
                
                ws.cell(row=row, column=7, value=valor_cotado)
                ws.cell(row=row, column=8, value=valor_considerado)
                ws.cell(row=row, column=9, value=valor_cte)
                ws.cell(row=row, column=10, value=valor_pago)
                
                # R$/kg calculados
                ws.cell(row=row, column=11, value=valor_cotado/peso if peso > 0 else 0)
                ws.cell(row=row, column=12, value=valor_considerado/peso if peso > 0 else 0)
                ws.cell(row=row, column=13, value=valor_cte/peso if peso > 0 else 0)
                ws.cell(row=row, column=14, value=valor_pago/peso if peso > 0 else 0)
                
                # Transportadora
                transportadora = getattr(frete, 'transportadora', None)
                if transportadora:
                    ws.cell(row=row, column=15, value=getattr(transportadora, 'razao_social', '') or '')
                    ws.cell(row=row, column=16, value='Sim' if getattr(transportadora, 'freteiro', False) else 'Não')
                else:
                    ws.cell(row=row, column=15, value='N/A')
                    ws.cell(row=row, column=16, value='N/A')
                
                # Data e status
                data_embarque = getattr(frete, 'data_embarque', None)
                ws.cell(row=row, column=17, value=self._format_date_br(data_embarque))
                ws.cell(row=row, column=18, value=getattr(frete, 'numero_cte', '') or '')
                
                # Status baseado em CTe
                status = 'Aprovado' if getattr(frete, 'numero_cte', None) else 'Pendente'
                ws.cell(row=row, column=19, value=status)
                
            except Exception as e:
                logger.warning(f"Erro ao processar frete {getattr(frete, 'id', 'N/A')}: {e}")
                continue
        
        # Auto-ajustar colunas
        self._auto_ajustar_colunas(ws)
    
    def _criar_aba_analise_financeira(self, ws, fretes):
        """Cria aba com análise financeira detalhada"""
        
        # Calcular totais
        valores = {
            'cotado': sum(float(getattr(f, 'valor_cotado', 0) or 0) for f in fretes),
            'considerado': sum(float(getattr(f, 'valor_considerado', 0) or 0) for f in fretes),
            'cte': sum(float(getattr(f, 'valor_cte', 0) or 0) for f in fretes),
            'pago': sum(float(getattr(f, 'valor_pago', 0) or 0) for f in fretes)
        }
        
        # Headers
        ws.cell(row=1, column=1, value="ANÁLISE FINANCEIRA - FRETES")
        self._aplicar_estilo_header(ws.cell(row=1, column=1))
        
        # Resumo geral
        row = 3
        resumo_data = [
            ("Total de Fretes", len(fretes)),
            ("", ""),
            ("Valor Total Cotado", valores['cotado']),
            ("Valor Total Considerado", valores['considerado']),
            ("Valor Total CTe", valores['cte']),
            ("Valor Total Pago", valores['pago']),
            ("", ""),
            ("Economia (Cotado - Considerado)", valores['cotado'] - valores['considerado']),
            ("Diferença (CTe - Considerado)", valores['cte'] - valores['considerado']),
            ("Saldo a Pagar (CTe - Pago)", valores['cte'] - valores['pago'])
        ]
        
        for descricao, valor in resumo_data:
            ws.cell(row=row, column=1, value=descricao)
            if isinstance(valor, (int, float)):
                ws.cell(row=row, column=2, value=valor)
                if valor < 0:
                    ws.cell(row=row, column=2).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                elif valor > 0 and 'Economia' in descricao:
                    ws.cell(row=row, column=2).fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            row += 1
        
        self._auto_ajustar_colunas(ws)
    
    def _criar_aba_transportadoras(self, ws, fretes):
        """Cria aba com análise por transportadora"""
        
        # Agrupar por transportadora
        transportadoras = {}
        for frete in fretes:
            transp = getattr(frete, 'transportadora', None)
            nome = getattr(transp, 'razao_social', 'Sem Transportadora') if transp else 'Sem Transportadora'
            freteiro = getattr(transp, 'freteiro', False) if transp else False
            
            if nome not in transportadoras:
                transportadoras[nome] = {
                    'fretes': [],
                    'freteiro': freteiro
                }
            transportadoras[nome]['fretes'].append(frete)
        
        # Headers
        headers = ['Transportadora', 'Tipo', 'Qtd Fretes', 'Valor Cotado', 'Valor Considerado', 'Economia']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._aplicar_estilo_header(cell)
        
        # Dados por transportadora
        row = 2
        for nome, dados in sorted(transportadoras.items()):
            fretes_transp = dados['fretes']
            
            cotado = sum(float(getattr(f, 'valor_cotado', 0) or 0) for f in fretes_transp)
            considerado = sum(float(getattr(f, 'valor_considerado', 0) or 0) for f in fretes_transp)
            economia = cotado - considerado
            
            ws.cell(row=row, column=1, value=nome)
            ws.cell(row=row, column=2, value='Freteiro' if dados['freteiro'] else 'Empresa')
            ws.cell(row=row, column=3, value=len(fretes_transp))
            ws.cell(row=row, column=4, value=cotado)
            ws.cell(row=row, column=5, value=considerado)
            ws.cell(row=row, column=6, value=economia)
            
            row += 1
        
        self._auto_ajustar_colunas(ws)
    
    def _criar_resumo_fretes(self, fretes: list, filtros: dict, caminho_arquivo: str, stats: dict) -> str:
        """Cria resumo detalhado do relatório"""
        
        # Estatísticas financeiras
        valores = {
            'cotado': sum(float(getattr(f, 'valor_cotado', 0) or 0) for f in fretes),
            'considerado': sum(float(getattr(f, 'valor_considerado', 0) or 0) for f in fretes),
            'cte': sum(float(getattr(f, 'valor_cte', 0) or 0) for f in fretes),
            'pago': sum(float(getattr(f, 'valor_pago', 0) or 0) for f in fretes)
        }
        
        # Contadores por tipo
        freteiros = sum(1 for f in fretes if getattr(getattr(f, 'transportadora', None), 'freteiro', False))
        empresas = len(fretes) - freteiros
        
        # Estados únicos
        estados = set(getattr(f, 'uf_destino', '') for f in fretes if getattr(f, 'uf_destino', ''))
        
        resumo = f"""📊 **RELATÓRIO DE FRETES GERADO COM SUCESSO!**

📋 **Detalhes do Relatório:**
• **Total de Fretes:** {len(fretes)}
• **Período:** {self._format_date_br(filtros.get('data_inicio'))} a {self._format_date_br(filtros.get('data_fim')) if filtros.get('data_fim') else 'Atual'}
• **Estados:** {', '.join(sorted(estados)) if estados else 'Todos'}
• **Cliente:** {filtros.get('cliente', 'Todos')}

💰 **ANÁLISE FINANCEIRA COMPLETA:**
• **Valor Total Cotado:** {self._format_currency(valores['cotado'])}
• **Valor Total Considerado:** {self._format_currency(valores['considerado'])}
• **Valor Total CTe:** {self._format_currency(valores['cte'])}
• **Valor Total Pago:** {self._format_currency(valores['pago'])}

💡 **INDICADORES:**
• **Economia Negociada:** {self._format_currency(valores['cotado'] - valores['considerado'])} ({self._format_percentage((valores['cotado'] - valores['considerado']) / valores['cotado'] * 100 if valores['cotado'] > 0 else 0)})
• **Diferença CTe:** {self._format_currency(valores['cte'] - valores['considerado'])}
• **Saldo a Pagar:** {self._format_currency(valores['cte'] - valores['pago'])}

🚚 **TRANSPORTADORAS:**
• **Freteiros:** {freteiros} fretes
• **Empresas:** {empresas} fretes
• **R$/kg Médio:** {self._format_currency(valores['considerado'] / sum(float(getattr(f, 'peso_total', 0) or 0) for f in fretes) if fretes else 0)}

📁 **ARQUIVO EXCEL:**
**[📊 BAIXAR RELATÓRIO COMPLETO]({caminho_arquivo})**

📈 **ABAS DISPONÍVEIS:**
• **Fretes Detalhados** - Dados completos com 19 colunas
• **Análise Financeira** - Resumo e indicadores
• **Por Transportadora** - Agrupamento e performance"""

        return resumo
    
    def _aplicar_estilo_header(self, cell):
        """Aplica estilo ao header"""
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )
    
    def _auto_ajustar_colunas(self, ws):
        """Auto-ajusta largura das colunas"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _fallback_sem_openpyxl(self) -> str:
        """Fallback quando openpyxl não está disponível"""
        return format_response_advanced("""❌ **Excel não disponível**

⚠️ **Problema:** Biblioteca openpyxl não instalada

💡 **Solução:**
```bash
pip install openpyxl
```

📊 **Funcionalidades Excel de Fretes:**
• Relatórios detalhados com 19 colunas
• Análise financeira completa (Cotado → Considerado → CTe → Pago)
• Agrupamento por transportadora
• Indicadores R$/kg e economia
• Separação freteiros vs empresas""", "ExcelFretes")
    
    def _fallback_sem_sistema(self) -> str:
        """Fallback quando sistema não está disponível"""
        return format_response_advanced("""⚠️ **Sistema não disponível**

🔧 **Problema:** Modelos Flask não carregados

💡 **Excel de Fretes disponível quando sistema ativo:**
• Dados de fretes e transportadoras
• Campos: freteiro, valor_pago, valor_cte
• Análise financeira em 4 estágios
• Performance por transportadora""", "ExcelFretes")

# Instância global
_excel_fretes = None

def get_excel_fretes():
    """Retorna instância de ExcelFretes"""
    global _excel_fretes
    if _excel_fretes is None:
        _excel_fretes = ExcelFretes()
    return _excel_fretes 