#!/usr/bin/env python3
"""
Excel Entregas - Mini esqueleto especializado para relatórios de entregas
Versão otimizada integrada com BaseCommand
"""

from app.claude_ai_novo.commands.base_command import (
    BaseCommand, format_response_advanced, create_excel_summary,
    logging, datetime, db, current_user
)
from pathlib import Path

# Excel generation imports
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Sistema imports (com fallbacks)
try:
    from app.monitoramento.models import EntregaMonitorada, AgendamentoEntrega
    from app.pedidos.models import Pedido
    from sqlalchemy import func, and_, or_
    SISTEMA_AVAILABLE = True
except ImportError:
    SISTEMA_AVAILABLE = False

logger = logging.getLogger(__name__)

class ExcelEntregas(BaseCommand):
    """Gerador de Excel especializado em entregas e monitoramento"""
    
    def __init__(self):
        super().__init__()
        self.tipo = "entregas"
        
    def is_excel_entregas_command(self, consulta: str) -> bool:
        """Detecta comandos específicos de Excel para entregas"""
        if not self._validate_input(consulta):
            return False
        
        keywords_entregas = [
            'excel entregas', 'planilha entregas', 'relatório entregas',
            'exportar entregas', 'entregas excel', 'entregas em excel',
            'monitoramento excel', 'agendamentos excel', 'canhotos excel'
        ]
        
        consulta_lower = consulta.lower()
        return any(keyword in consulta_lower for keyword in keywords_entregas)
    
    def gerar_excel_entregas(self, consulta: str, filtros=None) -> str:
        """Gera Excel de entregas com análise de performance"""
        
        if not EXCEL_AVAILABLE:
            return self._fallback_sem_openpyxl()
        
        if not SISTEMA_AVAILABLE:
            return self._fallback_sem_sistema()
        
        # Sanitizar entrada e extrair filtros avançados
        consulta = self._sanitize_input(consulta)
        if not filtros:
            filtros = self._extract_filters_advanced(consulta)
        
        # Log avançado
        self._log_command(consulta, "excel_entregas", filtros)
        
        try:
            # Verificar cache primeiro
            cache_key = self._generate_cache_key("excel_entregas", consulta, filtros)
            cached_result = self._get_cached_result(cache_key, 600)
            
            if cached_result:
                logger.info("✅ Excel entregas encontrado em cache")
                return cached_result
            
            # Processar geração
            resultado = self._gerar_excel_entregas_interno(consulta, filtros)
            
            # Armazenar em cache
            self._set_cached_result(cache_key, resultado, 600)
            
            return resultado
            
        except Exception as e:
            return self._handle_error(e, "excel_entregas", f"Consulta: {consulta[:100]}")
    
    def _gerar_excel_entregas_interno(self, consulta: str, filtros: dict) -> str:
        """Geração interna do Excel de entregas"""
        
        # Buscar dados com filtros avançados
        entregas = self._buscar_dados_entregas(filtros)
        
        if not entregas:
            return f"⚠️ **Nenhuma entrega encontrada** para os critérios: {', '.join([f'{k}={v}' for k, v in filtros.items()])}"
        
        # Gerar arquivo Excel
        caminho_arquivo = self._criar_excel_entregas(entregas, filtros)
        
        # Criar estatísticas
        stats = self._create_summary_stats(entregas, 'entregas')
        
        # Criar resumo
        resumo = self._criar_resumo_entregas(entregas, filtros, caminho_arquivo, stats)
        
        return format_response_advanced(resumo, "ExcelEntregas", stats)
    
    def _buscar_dados_entregas(self, filtros: dict) -> list:
        """Busca dados de entregas com filtros avançados"""
        
        query = db.session.query(EntregaMonitorada).join(Pedido, isouter=True)
        
        # Aplicar filtros
        if filtros.get('cliente'):
            cliente = filtros['cliente']
            query = query.filter(
                or_(
                    EntregaMonitorada.nome_cliente.ilike(f'%{cliente}%'),
                    Pedido.nome_cliente.ilike(f'%{cliente}%')
                )
            )
        
        if filtros.get('uf'):
            uf = filtros['uf']
            query = query.filter(EntregaMonitorada.uf_destino == uf)
        
        if filtros.get('status'):
            status = filtros['status']
            if status == 'entregue':
                query = query.filter(EntregaMonitorada.entregue == True)
            elif status == 'pendente':
                query = query.filter(EntregaMonitorada.entregue == False)
            elif status == 'atrasado':
                hoje = datetime.now().date()
                query = query.filter(
                    and_(
                        EntregaMonitorada.entregue == False,
                        EntregaMonitorada.data_entrega_prevista < hoje
                    )
                )
        
        if filtros.get('data_inicio') and filtros.get('data_fim'):
            query = query.filter(
                or_(
                    and_(
                        EntregaMonitorada.data_embarque >= filtros['data_inicio'],
                        EntregaMonitorada.data_embarque <= filtros['data_fim']
                    ),
                    and_(
                        EntregaMonitorada.data_entrega_prevista >= filtros['data_inicio'],
                        EntregaMonitorada.data_entrega_prevista <= filtros['data_fim']
                    )
                )
            )
        
        # Ordenação inteligente
        query = query.order_by(EntregaMonitorada.data_entrega_prevista.desc(), EntregaMonitorada.id.desc())
        
        # Limitar resultados para performance
        return query.limit(5000).all()
    
    def _criar_excel_entregas(self, entregas: list, filtros: dict) -> str:
        """Cria arquivo Excel com dados das entregas"""
        
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)
        
        # Aba 1: Dados principais
        ws_dados = wb.create_sheet("Entregas Detalhadas")
        self._criar_aba_entregas_principal(ws_dados, entregas)
        
        # Aba 2: Performance de entregas
        ws_performance = wb.create_sheet("Performance")
        self._criar_aba_performance(ws_performance, entregas)
        
        # Aba 3: Agendamentos
        ws_agendamentos = wb.create_sheet("Agendamentos")
        self._criar_aba_agendamentos(ws_agendamentos, entregas)
        
        # Salvar arquivo
        nome_arquivo = f"entregas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        caminho_completo = self.output_dir / nome_arquivo
        
        wb.save(caminho_completo)
        
        return f"/static/reports/{nome_arquivo}"
    
    def _criar_aba_entregas_principal(self, ws, entregas):
        """Cria aba principal com dados completos das entregas"""
        
        headers = [
            'NF', 'Cliente', 'CNPJ', 'Destino', 'UF', 'Peso',
            'Data Embarque', 'Data Prevista', 'Data Realizada', 'Lead Time',
            'Status', 'Entregue', 'Canhoto', 'Vendedor', 'Observações'
        ]
        
        # Aplicar headers com estilo
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._aplicar_estilo_header(cell)
        
        # Dados
        for row, entrega in enumerate(entregas, 2):
            try:
                ws.cell(row=row, column=1, value=getattr(entrega, 'numero_nf', '') or '')
                ws.cell(row=row, column=2, value=getattr(entrega, 'nome_cliente', '') or '')
                ws.cell(row=row, column=3, value=getattr(entrega, 'cnpj_cliente', '') or '')
                ws.cell(row=row, column=4, value=getattr(entrega, 'cidade_destino', '') or '')
                ws.cell(row=row, column=5, value=getattr(entrega, 'uf_destino', '') or '')
                ws.cell(row=row, column=6, value=float(getattr(entrega, 'peso_nf', 0) or 0))
                
                # Datas
                data_embarque = getattr(entrega, 'data_embarque', None)
                ws.cell(row=row, column=7, value=self._format_date_br(data_embarque))
                
                data_prevista = getattr(entrega, 'data_entrega_prevista', None)
                ws.cell(row=row, column=8, value=self._format_date_br(data_prevista))
                
                data_realizada = getattr(entrega, 'data_entrega_realizada', None)
                ws.cell(row=row, column=9, value=self._format_date_br(data_realizada))
                
                # Lead time
                lead_time = getattr(entrega, 'lead_time_dias', None)
                ws.cell(row=row, column=10, value=f"{lead_time} dias" if lead_time else 'N/A')
                
                # Status e flags
                ws.cell(row=row, column=11, value=getattr(entrega, 'status_finalizacao', '') or '')
                ws.cell(row=row, column=12, value='Sim' if getattr(entrega, 'entregue', False) else 'Não')
                
                # Canhoto
                possui_canhoto = getattr(entrega, 'possui_canhoto', False) if hasattr(entrega, 'possui_canhoto') else False
                ws.cell(row=row, column=13, value='Sim' if possui_canhoto else 'Não')
                
                ws.cell(row=row, column=14, value=getattr(entrega, 'vendedor_codigo', '') or '')
                ws.cell(row=row, column=15, value=getattr(entrega, 'observacoes', '') or '')
                
            except Exception as e:
                logger.warning(f"Erro ao processar entrega {getattr(entrega, 'numero_nf', 'N/A')}: {e}")
                continue
        
        # Auto-ajustar colunas
        self._auto_ajustar_colunas(ws)
    
    def _criar_aba_performance(self, ws, entregas):
        """Cria aba com análise de performance"""
        
        # Calcular estatísticas
        total_entregas = len(entregas)
        entregues = sum(1 for e in entregas if getattr(e, 'entregue', False))
        pendentes = total_entregas - entregues
        
        # Headers
        ws.cell(row=1, column=1, value="PERFORMANCE DE ENTREGAS")
        self._aplicar_estilo_header(ws.cell(row=1, column=1))
        
        # Resumo geral
        row = 3
        resumo_data = [
            ("Total de Entregas", total_entregas),
            ("Entregas Realizadas", entregues),
            ("Entregas Pendentes", pendentes),
            ("Taxa de Entrega", f"{(entregues/total_entregas*100):.1f}%" if total_entregas > 0 else "0%"),
            ("", ""),
        ]
        
        for descricao, valor in resumo_data:
            ws.cell(row=row, column=1, value=descricao)
            ws.cell(row=row, column=2, value=valor)
            row += 1
        
        # Análise por prazo
        ws.cell(row=row, column=1, value="ANÁLISE DE PRAZOS:")
        self._aplicar_estilo_header(ws.cell(row=row, column=1))
        row += 1
        
        hoje = datetime.now().date()
        no_prazo = 0
        atrasadas = 0
        entregues_prazo = 0
        entregues_atraso = 0
        
        for entrega in entregas:
            data_prevista = getattr(entrega, 'data_entrega_prevista', None)
            data_realizada = getattr(entrega, 'data_entrega_realizada', None)
            entregue = getattr(entrega, 'entregue', False)
            
            if entregue and data_realizada and data_prevista:
                if data_realizada <= data_prevista:
                    entregues_prazo += 1
                else:
                    entregues_atraso += 1
            elif not entregue and data_prevista:
                if data_prevista >= hoje:
                    no_prazo += 1
                else:
                    atrasadas += 1
        
        prazo_data = [
            ("Pendentes no Prazo", no_prazo),
            ("Pendentes Atrasadas", atrasadas),
            ("Entregues no Prazo", entregues_prazo),
            ("Entregues em Atraso", entregues_atraso),
        ]
        
        for descricao, valor in prazo_data:
            ws.cell(row=row, column=1, value=descricao)
            ws.cell(row=row, column=2, value=valor)
            if valor > 0 and 'Atraso' in descricao:
                ws.cell(row=row, column=2).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            elif valor > 0 and 'Prazo' in descricao:
                ws.cell(row=row, column=2).fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            row += 1
        
        self._auto_ajustar_colunas(ws)
    
    def _criar_aba_agendamentos(self, ws, entregas):
        """Cria aba com análise de agendamentos"""
        
        headers = ['NF', 'Cliente', 'Data Prevista', 'Agendamentos', 'Último Status', 'Confirmado']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._aplicar_estilo_header(cell)
        
        # Dados de agendamentos
        row = 2
        for entrega in entregas:
            try:
                # Buscar agendamentos relacionados
                agendamentos = []
                if hasattr(entrega, 'agendamentos_entrega'):
                    agendamentos = getattr(entrega, 'agendamentos_entrega', [])
                
                ws.cell(row=row, column=1, value=getattr(entrega, 'numero_nf', '') or '')
                ws.cell(row=row, column=2, value=getattr(entrega, 'nome_cliente', '') or '')
                
                data_prevista = getattr(entrega, 'data_entrega_prevista', None)
                ws.cell(row=row, column=3, value=self._format_date_br(data_prevista))
                
                ws.cell(row=row, column=4, value=len(agendamentos))
                
                # Último agendamento
                if agendamentos:
                    ultimo = agendamentos[-1]
                    ws.cell(row=row, column=5, value=getattr(ultimo, 'status', '') or '')
                    confirmado = getattr(ultimo, 'confirmado_por', None)
                    ws.cell(row=row, column=6, value='Sim' if confirmado else 'Não')
                else:
                    ws.cell(row=row, column=5, value='Sem agendamento')
                    ws.cell(row=row, column=6, value='N/A')
                
                row += 1
                
            except Exception as e:
                logger.warning(f"Erro ao processar agendamentos da entrega {getattr(entrega, 'numero_nf', 'N/A')}: {e}")
                continue
        
        self._auto_ajustar_colunas(ws)
    
    def _criar_resumo_entregas(self, entregas: list, filtros: dict, caminho_arquivo: str, stats: dict) -> str:
        """Cria resumo detalhado do relatório"""
        
        # Estatísticas
        total_entregas = len(entregas)
        entregues = sum(1 for e in entregas if getattr(e, 'entregue', False))
        pendentes = total_entregas - entregues
        
        # Análise de prazos
        hoje = datetime.now().date()
        atrasadas = 0
        no_prazo = 0
        
        for entrega in entregas:
            data_prevista = getattr(entrega, 'data_entrega_prevista', None)
            entregue = getattr(entrega, 'entregue', False)
            
            if not entregue and data_prevista:
                if data_prevista < hoje:
                    atrasadas += 1
                else:
                    no_prazo += 1
        
        # Estados únicos
        estados = set(getattr(e, 'uf_destino', '') for e in entregas if getattr(e, 'uf_destino', ''))
        
        # Peso total
        peso_total = sum(float(getattr(e, 'peso_nf', 0) or 0) for e in entregas)
        
        resumo = f"""📊 **RELATÓRIO DE ENTREGAS GERADO COM SUCESSO!**

📋 **Detalhes do Relatório:**
• **Total de Entregas:** {total_entregas}
• **Período:** {self._format_date_br(filtros.get('data_inicio'))} a {self._format_date_br(filtros.get('data_fim')) if filtros.get('data_fim') else 'Atual'}
• **Estados:** {', '.join(sorted(estados)) if estados else 'Todos'}
• **Cliente:** {filtros.get('cliente', 'Todos')}

📊 **PERFORMANCE DE ENTREGAS:**
• **Entregas Realizadas:** {entregues}/{total_entregas} ({(entregues/total_entregas*100):.1f}%)
• **Entregas Pendentes:** {pendentes}/{total_entregas} ({(pendentes/total_entregas*100):.1f}%)
• **Taxa de Entrega:** {self._format_percentage(entregues/total_entregas*100 if total_entregas > 0 else 0)}

⏰ **ANÁLISE DE PRAZOS:**
• **Pendentes no Prazo:** {no_prazo}
• **Pendentes Atrasadas:** {atrasadas}
• **Peso Total:** {self._format_weight(peso_total)}

🚛 **INDICADORES:**
• **Peso Médio por Entrega:** {self._format_weight(peso_total / total_entregas if total_entregas > 0 else 0)}
• **Entregas por Dia:** {(total_entregas / 30):.1f} (média mensal)

📁 **ARQUIVO EXCEL:**
**[📊 BAIXAR RELATÓRIO COMPLETO]({caminho_arquivo})**

📈 **ABAS DISPONÍVEIS:**
• **Entregas Detalhadas** - Dados completos com 15 colunas
• **Performance** - Taxa de entrega e análise de prazos
• **Agendamentos** - Status e confirmações de agendamento"""

        return resumo
    
    def _aplicar_estilo_header(self, cell):
        """Aplica estilo ao header"""
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )
    
    def _auto_ajustar_colunas(self, ws):
        """Auto-ajusta largura das colunas"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _fallback_sem_openpyxl(self) -> str:
        """Fallback quando openpyxl não está disponível"""
        return format_response_advanced("""❌ **Excel não disponível**

⚠️ **Problema:** Biblioteca openpyxl não instalada

💡 **Solução:**
```bash
pip install openpyxl
```

📊 **Funcionalidades Excel de Entregas:**
• Relatórios detalhados com 15 colunas
• Análise de performance de entregas
• Controle de prazos e atrasos
• Status de agendamentos
• Indicadores de canhotos""", "ExcelEntregas")
    
    def _fallback_sem_sistema(self) -> str:
        """Fallback quando sistema não está disponível"""
        return format_response_advanced("""⚠️ **Sistema não disponível**

🔧 **Problema:** Modelos Flask não carregados

💡 **Excel de Entregas disponível quando sistema ativo:**
• Dados de entregas monitoradas
• Performance e prazos
• Agendamentos e confirmações
• Status de finalização e canhotos""", "ExcelEntregas")

# Instância global
_excel_entregas = None

def get_excel_entregas():
    """Retorna instância de ExcelEntregas"""
    global _excel_entregas
    if _excel_entregas is None:
        _excel_entregas = ExcelEntregas()
    return _excel_entregas 