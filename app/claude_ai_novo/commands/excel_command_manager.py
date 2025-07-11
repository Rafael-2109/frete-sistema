#!/usr/bin/env python3
"""
Excel Orchestrator - Orquestrador otimizado dos mini esqueletos Excel
Versão avançada integrada com BaseCommand e patterns inteligentes
"""

from app.claude_ai_novo.commands.base_command import (
    BaseCommand, format_response_advanced, create_excel_summary, detect_command_type,
    logging, datetime, db, current_user
)

# Imports condicionais dos mini esqueletos da pasta excel/
try:
    from app.claude_ai_novo.commands.excel.fretes import get_excel_fretes
    FRETES_AVAILABLE = True
except ImportError:
    FRETES_AVAILABLE = False

try:
    from app.claude_ai_novo.commands.excel.pedidos import get_excel_pedidos
    PEDIDOS_AVAILABLE = True
except ImportError:
    PEDIDOS_AVAILABLE = False

try:
    from app.claude_ai_novo.commands.excel.entregas import get_excel_entregas
    ENTREGAS_AVAILABLE = True
except ImportError:
    ENTREGAS_AVAILABLE = False

try:
    from app.claude_ai_novo.commands.excel.faturamento import get_excel_faturamento
    FATURAMENTO_AVAILABLE = True
except ImportError:
    FATURAMENTO_AVAILABLE = False

logger = logging.getLogger(__name__)

class ExcelOrchestrator(BaseCommand):
    """Orquestrador inteligente dos mini esqueletos Excel"""
    
    def __init__(self):
        super().__init__()
        
        # Inicializar mini esqueletos disponíveis
        self.esqueletos = {}
        self._inicializar_esqueletos()
        
        # Configuração de prioridades
        self.prioridades = {
            'faturamento': 1,
            'fretes': 2,  
            'pedidos': 3,
            'entregas': 4
        }
        
        # Patterns de detecção avançada
        self.patterns_deteccao = {
            'faturamento': [
                'faturamento', 'faturas', 'nf', 'nota fiscal', 'invoice',
                'billing', 'receita', 'cobrança', 'valor faturado'
            ],
            'fretes': [
                'fretes', 'frete', 'transportadoras', 'cotações', 'cte',
                'freteiros', 'despesas', 'transporte', 'shipping'
            ],
            'pedidos': [
                'pedidos', 'pedido', 'cotações', 'vendas', 'carteira',
                'orders', 'sales', 'clientes'
            ],
            'entregas': [
                'entregas', 'entrega', 'monitoramento', 'agendamentos',
                'delivery', 'canhotos', 'performance'
            ]
        }
    
    def _inicializar_esqueletos(self):
        """Inicializa mini esqueletos disponíveis"""
        
        if FATURAMENTO_AVAILABLE:
            self.esqueletos['faturamento'] = get_excel_faturamento()
            logger.info("✅ Mini esqueleto Faturamento carregado")
        
        if FRETES_AVAILABLE:
            self.esqueletos['fretes'] = get_excel_fretes()
            logger.info("✅ Mini esqueleto Fretes carregado")
        
        if PEDIDOS_AVAILABLE:
            self.esqueletos['pedidos'] = get_excel_pedidos()
            logger.info("✅ Mini esqueleto Pedidos carregado")
            
        if ENTREGAS_AVAILABLE:
            self.esqueletos['entregas'] = get_excel_entregas()
            logger.info("✅ Mini esqueleto Entregas carregado")
        
        logger.info(f"🎯 Excel Orchestrator iniciado com {len(self.esqueletos)} mini esqueletos")
    
    def is_excel_command(self, consulta: str) -> bool:
        """Detecta se é comando Excel usando base.py"""
        if not self._validate_input(consulta):
            return False
        
        # Usar detecção do base.py
        return detect_command_type(consulta) == 'excel'
    
    def processar_comando_excel(self, consulta: str, user_context=None) -> str:
        """Processa comando Excel com orquestração inteligente"""
        
        if not self._validate_input(consulta):
            return self._handle_error(ValueError("Consulta inválida"), "excel", "Entrada vazia ou inválida")
        
        # Sanitizar entrada
        consulta = self._sanitize_input(consulta)
        
        # Extrair filtros avançados
        filtros = self._extract_filters_advanced(consulta)
        
        # Log avançado
        self._log_command(consulta, "excel_orchestrator", filtros)
        
        try:
            # Verificar cache primeiro
            cache_key = self._generate_cache_key("excel", consulta, filtros)
            cached_result = self._get_cached_result(cache_key, 900)  # 15 min cache
            
            if cached_result:
                logger.info("✅ Resultado Excel encontrado em cache")
                return cached_result
            
            # Processar comando
            resultado = self._processar_excel_interno(consulta, filtros, user_context)
            
            # Armazenar em cache
            self._set_cached_result(cache_key, resultado, 900)
            
            return resultado
            
        except Exception as e:
            return self._handle_error(e, "excel_orchestrator", f"Consulta: {consulta[:100]}")
    
    def _processar_excel_interno(self, consulta: str, filtros: dict, user_context) -> str:
        """Processamento interno com roteamento inteligente"""
        
        # Detectar tipo específico
        tipo_detectado = self._detectar_tipo_excel(consulta)
        
        # Verificar se temos o esqueleto apropriado
        if tipo_detectado not in self.esqueletos:
            return self._fallback_tipo_indisponivel(tipo_detectado, consulta)
        
        # Verificar se é comando específico do esqueleto
        esqueleto = self.esqueletos[tipo_detectado]
        
        try:
            # Roteamento específico por tipo
            if tipo_detectado == 'faturamento':
                return esqueleto.gerar_excel_faturamento(consulta, filtros)
            elif tipo_detectado == 'fretes':
                return esqueleto.gerar_excel_fretes(consulta, filtros)
            elif tipo_detectado == 'pedidos':
                return esqueleto.gerar_excel_pedidos(consulta, filtros)
            elif tipo_detectado == 'entregas':
                return esqueleto.gerar_excel_entregas(consulta, filtros)
            else:
                # Fallback para Excel geral
                return self._gerar_excel_geral_multi(consulta, filtros)
                
        except Exception as e:
            logger.error(f"Erro no esqueleto {tipo_detectado}: {e}")
            # Fallback para outros esqueletos
            return self._tentar_fallback_esqueletos(consulta, filtros, tipo_detectado)
    
    def _detectar_tipo_excel(self, consulta: str) -> str:
        """Detecta tipo específico de Excel com scoring inteligente"""
        
        consulta_lower = consulta.lower()
        scores = {}
        
        # Calcular scores para cada tipo
        for tipo, keywords in self.patterns_deteccao.items():
            score = 0
            for keyword in keywords:
                if keyword in consulta_lower:
                    # Score baseado na especificidade e posição
                    base_score = len(keyword)  # Palavras maiores = mais específicas
                    position_bonus = 1.2 if consulta_lower.index(keyword) < 20 else 1.0  # Início da frase
                    score += base_score * position_bonus
            
            scores[tipo] = score
        
        # Retornar tipo com maior score ou fallback baseado em prioridade
        if scores:
            tipo_detectado = max(scores.keys(), key=lambda k: scores[k])
            if scores[tipo_detectado] > 0:
                logger.info(f"🎯 Tipo detectado: {tipo_detectado} (score: {scores[tipo_detectado]})")
                return tipo_detectado
        
        # Fallback: usar prioridade
        for tipo in sorted(self.prioridades.keys(), key=lambda k: self.prioridades[k]):
            if tipo in self.esqueletos:
                logger.info(f"🔄 Fallback para tipo: {tipo}")
                return tipo
        
        return 'geral'
    
    def _tentar_fallback_esqueletos(self, consulta: str, filtros: dict, tipo_original: str) -> str:
        """Tenta outros esqueletos em caso de erro"""
        
        # Ordenar esqueletos por prioridade (excluindo o que deu erro)
        esqueletos_fallback = [
            tipo for tipo in sorted(self.prioridades.keys(), key=lambda k: self.prioridades[k])
            if tipo in self.esqueletos and tipo != tipo_original
        ]
        
        for tipo_fallback in esqueletos_fallback:
            try:
                logger.info(f"🔄 Tentando fallback: {tipo_fallback}")
                esqueleto = self.esqueletos[tipo_fallback]
                
                if tipo_fallback == 'fretes':
                    return esqueleto.gerar_excel_fretes(consulta, filtros)
                elif tipo_fallback == 'pedidos':
                    return esqueleto.gerar_excel_pedidos(consulta, filtros)
                elif tipo_fallback == 'entregas':
                    return esqueleto.gerar_excel_entregas(consulta, filtros)
                elif tipo_fallback == 'faturamento':
                    return esqueleto.gerar_excel_faturamento(consulta, filtros)
                    
            except Exception as e:
                logger.warning(f"Fallback {tipo_fallback} também falhou: {e}")
                continue
        
        # Último fallback: Excel geral
        return self._gerar_excel_geral_multi(consulta, filtros)
    
    def _gerar_excel_geral_multi(self, consulta: str, filtros: dict) -> str:
        """Gera Excel geral com dados de múltiplos módulos"""
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            return self._fallback_sem_openpyxl()
        
        try:
            wb = Workbook()
            # Remover aba padrão se existir
            if wb.active:
                wb.remove(wb.active)
            
            abas_criadas = 0
            
            # Aba 1: Fretes (se disponível)
            if 'fretes' in self.esqueletos:
                try:
                    ws_fretes = wb.create_sheet("Resumo Fretes")
                    self._criar_aba_resumo_fretes(ws_fretes, filtros)
                    abas_criadas += 1
                except Exception as e:
                    logger.warning(f"Erro ao criar aba fretes: {e}")
            
            # Aba 2: Pedidos (se disponível)
            if 'pedidos' in self.esqueletos:
                try:
                    ws_pedidos = wb.create_sheet("Resumo Pedidos")
                    self._criar_aba_resumo_pedidos(ws_pedidos, filtros)
                    abas_criadas += 1
                except Exception as e:
                    logger.warning(f"Erro ao criar aba pedidos: {e}")
            
            # Aba 3: Entregas (se disponível)
            if 'entregas' in self.esqueletos:
                try:
                    ws_entregas = wb.create_sheet("Resumo Entregas")
                    self._criar_aba_resumo_entregas(ws_entregas, filtros)
                    abas_criadas += 1
                except Exception as e:
                    logger.warning(f"Erro ao criar aba entregas: {e}")
            
            if abas_criadas == 0:
                # Criar aba de status
                ws_status = wb.create_sheet("Status Sistema")
                self._criar_aba_status_sistema(ws_status)
                abas_criadas += 1
            
            # Salvar arquivo
            nome_arquivo = f"relatorio_geral_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            caminho_completo = self.output_dir / nome_arquivo
            
            wb.save(caminho_completo)
            
            # Criar estatísticas
            stats = {
                'total': abas_criadas,
                'abas_criadas': abas_criadas,
                'esqueletos_ativos': len(self.esqueletos)
            }
            
            content = f"""📊 **RELATÓRIO GERAL MULTI-MÓDULO GERADO!**

📋 **Detalhes:**
• **Abas criadas:** {abas_criadas}
• **Esqueletos ativos:** {len(self.esqueletos)}
• **Filtros aplicados:** {', '.join([f'{k}={v}' for k, v in filtros.items()]) if filtros else 'Nenhum'}

📁 **ARQUIVO:**
**[📊 BAIXAR RELATÓRIO COMPLETO](/static/reports/{nome_arquivo})**

🎯 **Conteúdo disponível:**"""

            if 'fretes' in self.esqueletos:
                content += "\n• **Resumo Fretes** - Transportadoras e valores"
            if 'pedidos' in self.esqueletos:
                content += "\n• **Resumo Pedidos** - Carteira e vendas"
            if 'entregas' in self.esqueletos:
                content += "\n• **Resumo Entregas** - Performance e prazos"

            return format_response_advanced(content, "ExcelOrchestrator", stats)
            
        except Exception as e:
            return self._handle_error(e, "excel_geral", "Geração multi-módulo")
    
    def _criar_aba_resumo_fretes(self, ws, filtros):
        """Cria aba resumo de fretes"""
        ws.cell(row=1, column=1, value="RESUMO DE FRETES")
        ws.cell(row=2, column=1, value="Dados resumidos dos fretes disponíveis")
        # Adicionar lógica específica conforme necessário
    
    def _criar_aba_resumo_pedidos(self, ws, filtros):
        """Cria aba resumo de pedidos"""
        ws.cell(row=1, column=1, value="RESUMO DE PEDIDOS")
        ws.cell(row=2, column=1, value="Dados resumidos dos pedidos disponíveis")
        # Adicionar lógica específica conforme necessário
    
    def _criar_aba_resumo_entregas(self, ws, filtros):
        """Cria aba resumo de entregas"""
        ws.cell(row=1, column=1, value="RESUMO DE ENTREGAS")
        ws.cell(row=2, column=1, value="Dados resumidos das entregas disponíveis")
        # Adicionar lógica específica conforme necessário
    
    def _criar_aba_status_sistema(self, ws):
        """Cria aba de status do sistema"""
        ws.cell(row=1, column=1, value="STATUS DO SISTEMA EXCEL")
        
        row = 3
        ws.cell(row=row, column=1, value="Esqueletos Disponíveis:")
        row += 1
        
        for tipo, disponivel in [
            ('Faturamento', FATURAMENTO_AVAILABLE),
            ('Fretes', FRETES_AVAILABLE),
            ('Pedidos', PEDIDOS_AVAILABLE),
            ('Entregas', ENTREGAS_AVAILABLE)
        ]:
            ws.cell(row=row, column=1, value=tipo)
            ws.cell(row=row, column=2, value="✅ Ativo" if disponivel else "❌ Indisponível")
            row += 1
    
    def _fallback_tipo_indisponivel(self, tipo: str, consulta: str) -> str:
        """Fallback quando tipo específico não está disponível"""
        
        content = f"""⚠️ **Mini Esqueleto '{tipo.title()}' Indisponível**

🔧 **Problema:** O módulo {tipo} não está carregado

✅ **Esqueletos Disponíveis:**"""

        for tipo_disp in self.esqueletos.keys():
            content += f"\n• **{tipo_disp.title()}** - Ativo"

        if not self.esqueletos:
            content += "\n• Nenhum esqueleto disponível"

        content += f"""

💡 **Alternativas:**
• Tente: "excel {list(self.esqueletos.keys())[0] if self.esqueletos else 'geral'}"
• Ou: "relatório geral em excel"

🎯 **Sua consulta:** {consulta}"""

        stats = {
            'total': len(self.esqueletos),
            'tipo_solicitado': tipo,
            'esqueletos_ativos': list(self.esqueletos.keys())
        }

        return format_response_advanced(content, "ExcelOrchestrator", stats)
    
    def _fallback_sem_openpyxl(self) -> str:
        """Fallback quando openpyxl não está disponível"""
        return format_response_advanced("""❌ **Excel não disponível**

⚠️ **Problema:** Biblioteca openpyxl não instalada

💡 **Solução:**
```bash
pip install openpyxl
```

🎯 **Funcionalidades disponíveis após instalação:**
• Relatórios de fretes com análise financeira
• Relatórios de pedidos com vendas
• Relatórios de entregas com performance
• Relatório geral multi-módulo""", "ExcelOrchestrator")
    
    def get_status_esqueletos(self) -> dict:
        """Retorna status dos esqueletos para debug"""
        return {
            'esqueletos_carregados': list(self.esqueletos.keys()),
            'total_esqueletos': len(self.esqueletos),
            'disponibilidade': {
                'faturamento': FATURAMENTO_AVAILABLE,
                'fretes': FRETES_AVAILABLE,
                'pedidos': PEDIDOS_AVAILABLE,
                'entregas': ENTREGAS_AVAILABLE
            }
        }

# Instância global
_excel_orchestrator = None

def get_excel_orchestrator():
    """Retorna instância de ExcelOrchestrator"""
    global _excel_orchestrator
    if _excel_orchestrator is None:
        _excel_orchestrator = ExcelOrchestrator()
    return _excel_orchestrator
