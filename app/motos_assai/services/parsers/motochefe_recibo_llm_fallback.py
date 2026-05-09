"""Fallback LLM para recibo Motochefe (PDF ou XLSX→PDF).

Acionado quando determinístico extrai < 80% das linhas declaradas no header.

Estrutura:
- Sonnet 4.6 (modelo unico — Haiku descartado em 2026-05-09: imprecisao em
  recibos com tabelas longas de chassis)
- PDF: enviado como document block direto
- XLSX: converte primeiro para texto plano (sem layout) e envia como text
"""

from __future__ import annotations

import base64
import json
import logging
import os
import re
from decimal import Decimal
from typing import Dict, List, Any

logger = logging.getLogger(__name__)

SONNET_MODEL = 'claude-sonnet-4-6'
LLM_MODELS = [SONNET_MODEL]  # ordem de fallback (atualmente unico)

PROMPT_SYSTEM = """Você é um parser de recibos da Motochefe (contra-prova de entrega de motos elétricas para um CD).

Estrutura do recibo:
- Header com: data, empresa Motochefe, CNPJ, endereço, equipe, conferente, total de motos
- Tabela com linhas: PEDIDO | DESCRIÇÃO DO PRODUTO | CHASSI | MOTOR | COR

Modelos esperados (mas pode haver outros): DOT 1000W, X11 MINI 1000W, SOL 1000W, MIA 1000W.

Retorne JSON puro (sem markdown, sem comentários):

{
  "data_recibo": "DD/MM/YYYY",
  "empresa_motochefe": "...",
  "cnpj_motochefe": "37542484000100",
  "equipe": "HAROLDO SP",
  "conferente": "KAROLINE",
  "total_motos_declarado": 115,
  "chassis": [
    {
      "chassi": "LA2025SA110007354",
      "modelo_texto": "DOT 1000W",
      "motor": "QS60V30H25120801923",
      "cor": "CINZA"
    }
  ]
}

REGRAS:
- chassi sempre em UPPERCASE.
- Se um campo não existir, omita-o (não use null).
- Não inclua texto antes ou depois do JSON.
"""


class MotochefeReciboLlmFallbackError(Exception):
    pass


def parse_pdf_via_llm(pdf_bytes: bytes) -> Dict[str, Any]:
    """Parseia PDF de recibo via LLM. Retorna dict no mesmo formato dos extractors determinísticos."""
    try:
        import anthropic
    except ImportError:
        raise MotochefeReciboLlmFallbackError("anthropic SDK não instalado")

    api_key = os.getenv('ANTHROPIC_API_KEY')
    if not api_key:
        raise MotochefeReciboLlmFallbackError("ANTHROPIC_API_KEY não configurada")

    client = anthropic.Anthropic(api_key=api_key)
    pdf_b64 = base64.b64encode(pdf_bytes).decode('ascii')

    for modelo in LLM_MODELS:
        try:
            data = _chamar_llm_pdf(client, pdf_b64, modelo)
            items = _converter_para_lista_flat(data)
            if items:
                return {'parser_usado': _nome_parser(modelo), 'items': items}
        except Exception as e:
            logger.warning(f'{modelo} falhou: {e}')

    raise MotochefeReciboLlmFallbackError(f"Todos os modelos LLM falharam: {LLM_MODELS}")


def parse_xlsx_via_llm(xlsx_bytes: bytes) -> Dict[str, Any]:
    """Parseia XLSX serializando-o como texto via openpyxl + envia como text."""
    import io
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)

    chunks: List[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        chunks.append(f'== Aba: {sheet_name} ==')
        for row in ws.iter_rows(values_only=True):
            cells = ['' if c is None else str(c) for c in row]
            chunks.append(' | '.join(cells))

    texto_xlsx = '\n'.join(chunks)
    return _parse_text_via_llm(texto_xlsx)


def _parse_text_via_llm(texto: str) -> Dict[str, Any]:
    try:
        import anthropic
    except ImportError:
        raise MotochefeReciboLlmFallbackError("anthropic SDK não instalado")

    api_key = os.getenv('ANTHROPIC_API_KEY')
    if not api_key:
        raise MotochefeReciboLlmFallbackError("ANTHROPIC_API_KEY não configurada")

    client = anthropic.Anthropic(api_key=api_key)

    for modelo in LLM_MODELS:
        try:
            response = client.messages.create(
                model=modelo,
                max_tokens=8192,
                system=PROMPT_SYSTEM,
                messages=[{
                    'role': 'user',
                    'content': f'Texto extraído do XLSX:\n\n{texto}',
                }],
            )
            raw = response.content[0].text.strip()
            data = json.loads(_extrair_json(raw))
            items = _converter_para_lista_flat(data)
            if items:
                return {'parser_usado': _nome_parser(modelo), 'items': items}
        except Exception as e:
            logger.warning(f'{modelo} (xlsx) falhou: {e}')

    raise MotochefeReciboLlmFallbackError(f"Todos os modelos LLM falharam (xlsx): {LLM_MODELS}")


def _nome_parser(modelo: str) -> str:
    """Devolve nome curto persistido em assai_recibo_motochefe.parser_usado."""
    if 'sonnet' in modelo:
        return 'LLM_SONNET'
    if 'haiku' in modelo:
        return 'LLM_HAIKU'
    return f'LLM_{modelo.upper()}'


def _chamar_llm_pdf(client, pdf_b64: str, model: str) -> Dict:
    response = client.messages.create(
        model=model,
        max_tokens=8192,
        system=PROMPT_SYSTEM,
        messages=[{
            'role': 'user',
            'content': [
                {'type': 'document',
                 'source': {'type': 'base64', 'media_type': 'application/pdf', 'data': pdf_b64}},
                {'type': 'text', 'text': 'Extraia o recibo em JSON.'},
            ],
        }],
    )
    raw = response.content[0].text.strip()
    return json.loads(_extrair_json(raw))


def _extrair_json(raw: str) -> str:
    raw = raw.strip()
    if raw.startswith('```'):
        m = re.search(r'```(?:json)?\s*(\{.*\})\s*```', raw, re.DOTALL)
        if m:
            return m.group(1)
    inicio = raw.find('{')
    if inicio < 0:
        raise MotochefeReciboLlmFallbackError(f'Sem JSON: {raw[:200]}')
    return raw[inicio:]


def _converter_para_lista_flat(data: Dict) -> List[Dict]:
    items = []
    header = {
        'data_recibo': data.get('data_recibo'),
        'empresa_motochefe': data.get('empresa_motochefe'),
        'cnpj_motochefe': re.sub(r'\D', '', str(data.get('cnpj_motochefe', '')))[:14] or None,
        'equipe': data.get('equipe'),
        'conferente': data.get('conferente'),
        'total_motos_declarado': data.get('total_motos_declarado'),
    }
    for c in data.get('chassis', []):
        chassi = str(c.get('chassi', '')).strip().upper()
        if not chassi:
            continue
        items.append({
            **header,
            'chassi': chassi,
            'modelo_texto': c.get('modelo_texto') or c.get('modelo') or '',
            'motor': c.get('motor', ''),
            'cor': c.get('cor', ''),
        })
    return items
