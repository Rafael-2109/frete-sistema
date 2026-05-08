"""Extrator determinístico do recibo Motochefe em formato Excel (.xlsx).

Estratégia:
- Carrega workbook com data_only=True (resolve fórmulas)
- Detecta header pela presença de células contendo CHASSI, MOTOR, COR
- Extrai header (EMPRESA, CNPJ, EQUIPE, CONFERENTE) das células acima da tabela
- Itera linhas a partir do header até a primeira linha sem chassi
"""

from __future__ import annotations

import re
from typing import Dict, List, Any, Optional, Tuple

from openpyxl import load_workbook


class MotochefeReciboXlsxExtractor:
    """Parseia recibo Excel da Motochefe (estrutura tabular)."""

    def __init__(self):
        self.errors: List[str] = []
        self.warnings: List[str] = []
        self.formato = 'MOTOCHEFE_RECIBO_XLSX'

    def extract(self, xlsx_path: str) -> List[Dict[str, Any]]:
        items: List[Dict[str, Any]] = []
        try:
            wb = load_workbook(xlsx_path, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Localiza header da tabela (linha onde aparece CHASSI/MOTOR/COR)
                header_row, col_map = self._localizar_header(ws)
                if header_row is None:
                    continue

                # Header textual: campos fora da tabela
                header_data = self._extract_header_excel(ws, header_row)

                # Linhas de chassi
                for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                    chassi_raw = row[col_map['chassi']] if col_map.get('chassi') is not None else None
                    if not chassi_raw:
                        continue
                    chassi = str(chassi_raw).strip().upper()
                    if len(chassi) < 5:
                        continue
                    items.append({
                        **header_data,
                        'chassi': chassi,
                        'modelo_texto': self._cell_str(row, col_map.get('descricao')),
                        'motor': self._cell_str(row, col_map.get('motor')),
                        'cor': self._cell_str(row, col_map.get('cor')),
                    })
        except Exception as e:
            self.errors.append(f'Erro ao processar XLSX: {e}')
        return items

    def _cell_str(self, row, idx: Optional[int]) -> str:
        if idx is None:
            return ''
        v = row[idx]
        return str(v).strip() if v is not None else ''

    def _localizar_header(self, ws) -> Tuple[Optional[int], Dict[str, int]]:
        """Acha linha cabeçalho da tabela. Retorna (linha, {chassi, descricao, motor, cor})."""
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            cells_upper = [str(c or '').strip().upper() for c in row]

            col_map: Dict[str, int] = {}
            for i, c in enumerate(cells_upper):
                if c == 'CHASSI':
                    col_map['chassi'] = i
                elif 'DESCRI' in c or c == 'MODELO' or 'PRODUTO' in c:
                    col_map.setdefault('descricao', i)
                elif c == 'MOTOR':
                    col_map['motor'] = i
                elif c == 'COR':
                    col_map['cor'] = i

            if 'chassi' in col_map:
                return row_idx, col_map

        return None, {}

    def _extract_header_excel(self, ws, header_row: int) -> Dict[str, Any]:
        """Procura campos do header nas linhas anteriores ao cabeçalho da tabela."""
        header = {
            'data_recibo': None,
            'empresa_motochefe': None,
            'cnpj_motochefe': None,
            'equipe': None,
            'conferente': None,
            'total_motos_declarado': None,
        }
        for row in ws.iter_rows(min_row=1, max_row=header_row - 1, values_only=True):
            for cell in row:
                if not cell:
                    continue
                cell_str = str(cell).strip()
                upper = cell_str.upper()
                if upper.startswith('EMPRESA:'):
                    header['empresa_motochefe'] = cell_str.split(':', 1)[1].strip()
                elif upper.startswith('CNPJ:'):
                    cnpj_str = cell_str.split(':', 1)[1].strip()
                    header['cnpj_motochefe'] = re.sub(r'\D', '', cnpj_str)[:14] or None
                elif re.match(r'\d{2}/\d{2}/\d{4}', cell_str):
                    header['data_recibo'] = cell_str[:10]

        # Total e equipe podem estar em linhas APÓS a tabela; procurar tudo
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                cell_str = str(cell).strip()
                # Total: número entre 10 e 9999
                if cell_str.isdigit():
                    n = int(cell_str)
                    if 10 <= n <= 9999 and (header['total_motos_declarado'] is None or n > header['total_motos_declarado']):
                        header['total_motos_declarado'] = n

        return header
