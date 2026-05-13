"""Geração do Excel de solicitação de faturamento Q.P.A.

Estrutura espelhada do template `285.xlsx`:
- Aba **PEDIDO**: header com Nº LOJA, CLIENTE, CNPJ, IE, ENDEREÇO, BAIRRO, UF,
  CIDADE, CEP. Tabela ITEM | CHASSI | MODELO | COR | VALOR. Linha TOTAL.
- Aba **BASE LOJAS**: cópia das 39 lojas Assaí (referência).
"""

from __future__ import annotations

import io
from decimal import Decimal
from typing import Tuple, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.exc import IntegrityError

from app import db
from app.utils.file_storage import FileStorage
from app.utils.timezone import agora_brasil_naive
from app.motos_assai.models import (
    AssaiSeparacao, AssaiSeparacaoItem, AssaiLoja, AssaiMoto, AssaiModelo,
    AssaiPedidoExcel,
    SEPARACAO_STATUS_EM_SEPARACAO, SEPARACAO_STATUS_FECHADA,
    SEPARACAO_STATUS_CARREGADA, SEPARACAO_STATUS_FATURADA,
)


class FaturamentoError(Exception):
    """Erro base do faturamento_service."""


def gerar_excel_qpa(separacao_id: int, gerada_por_id: int) -> Tuple[bytes, Optional[str]]:
    """Gera Excel da solicitação. Retorna (bytes, s3_key).

    Salva em S3 em `motos_assai/solicitacoes/<separacao_id>.xlsx` e atualiza
    `assai_separacao.solicitacao_excel_s3_key`.

    Raises:
        ValueError: se separação não está em status FECHADA ou FATURADA.
    """
    sep = AssaiSeparacao.query.get_or_404(separacao_id)
    # H3: validar status antes de gerar.
    # Aceita: FECHADA, CARREGADA, FATURADA + EM_SEPARACAO (para regeneracao
    # apos substituicao cross-loja durante separacao em andamento).
    statuses_validos = (
        SEPARACAO_STATUS_EM_SEPARACAO,
        SEPARACAO_STATUS_FECHADA,
        SEPARACAO_STATUS_CARREGADA,
        SEPARACAO_STATUS_FATURADA,
    )
    if sep.status not in statuses_validos:
        raise ValueError(
            f'Separação {separacao_id} está {sep.status}, '
            f'esperado um de: {", ".join(statuses_validos)}'
        )
    loja = AssaiLoja.query.get(sep.loja_id)

    items = (
        db.session.query(AssaiSeparacaoItem, AssaiMoto, AssaiModelo)
        .join(AssaiMoto, AssaiMoto.chassi == AssaiSeparacaoItem.chassi)
        .join(AssaiModelo, AssaiModelo.id == AssaiSeparacaoItem.modelo_id)
        .filter(AssaiSeparacaoItem.separacao_id == separacao_id)
        .order_by(AssaiSeparacaoItem.id)
        .all()
    )

    wb = Workbook()

    # ===== Aba PEDIDO =====
    ws = wb.active
    ws.title = 'PEDIDO'

    bold = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14)
    fill_header = PatternFill(start_color='D0D0D0', end_color='D0D0D0', fill_type='solid')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    row = 1
    ws.cell(row=row, column=2, value=f'COLETA DIA {agora_brasil_naive().strftime("%d/%m")} - MONTADO.').font = bold
    row += 1
    ws.cell(row=row, column=2, value='PEDIDO DE VENDA  - SCOOTER ELETRICA').font = title_font
    row += 1
    ws.cell(row=row, column=2, value='OPERAÇÃO VOE X SENDAS').font = bold
    row += 1

    # Header info loja
    info = [
        ('Nº LOJA', loja.numero),
        ('CLIENTE:', loja.razao_social),
        ('CNPJ:', loja.cnpj),
        ('I.E', loja.ie or ''),
        ('ENDEREÇO:', loja.endereco or ''),
        ('BAIRRO:', loja.bairro or ''),
        ('CIDADE:', loja.cidade or ''),
        ('CEP:', loja.cep or ''),
    ]
    for label, value in info:
        ws.cell(row=row, column=2, value=label).font = bold
        ws.cell(row=row, column=3, value=value)
        if label == 'BAIRRO:':
            ws.cell(row=row, column=4, value='UF').font = bold
            ws.cell(row=row, column=5, value=loja.uf or '')
        row += 1

    row += 1  # linha em branco

    # Tabela
    for i, lab in enumerate(['ITEM', 'CHASSI', 'MODELO', 'COR', 'VALOR']):
        c = ws.cell(row=row, column=i + 1, value=lab)
        c.font = bold
        c.fill = fill_header
        c.border = border
        c.alignment = Alignment(horizontal='center')
    row += 1

    total = Decimal('0')
    for idx, (item, moto, modelo) in enumerate(items, start=1):
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=moto.chassi).border = border
        ws.cell(row=row, column=3, value=modelo.codigo).border = border
        ws.cell(row=row, column=4, value=moto.cor or '').border = border
        cell_v = ws.cell(row=row, column=5, value=float(item.valor_unitario_qpa))
        cell_v.border = border
        cell_v.number_format = '#,##0.00'
        total += item.valor_unitario_qpa
        row += 1

    ws.cell(row=row, column=4, value='TOTAL').font = bold
    cell_t = ws.cell(row=row, column=5, value=float(total))
    cell_t.font = bold
    cell_t.number_format = '#,##0.00'

    # Largura colunas
    for col_idx, w in enumerate([6, 22, 14, 12, 14], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ===== Aba BASE LOJAS =====
    ws2 = wb.create_sheet('BASE LOJAS')
    headers = ['Nº Loja', 'Loja', 'Regional', 'CNPJ', 'IE', 'RAZAO SOCIAL',
               'Endereço', 'BAIRRO', 'CEP', 'Cidade', 'UF']
    for i, h in enumerate(headers, start=1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = bold
        c.fill = fill_header

    todas_lojas = AssaiLoja.query.filter_by(ativo=True).order_by(AssaiLoja.numero).all()
    for r, l in enumerate(todas_lojas, start=2):
        valores = [l.numero, l.nome, l.regional or '', l.cnpj, l.ie or '',
                   l.razao_social, l.endereco or '', l.bairro or '',
                   l.cep or '', l.cidade or '', l.uf or '']
        for c_idx, v in enumerate(valores, start=1):
            ws2.cell(row=r, column=c_idx, value=v)

    for col_idx, w in enumerate([8, 26, 26, 18, 14, 36, 36, 22, 12, 22, 4], start=1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = w

    # ===== Salvar =====
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    bytes_xlsx = buf.getvalue()

    nome_arquivo = f'LJ{loja.numero}_solicitacao_{separacao_id}.xlsx'
    buf_s3 = io.BytesIO(bytes_xlsx)
    buf_s3.name = nome_arquivo
    s3_key = FileStorage().save_file(
        buf_s3,
        folder=f'motos_assai/solicitacoes',
        filename=nome_arquivo,
        allowed_extensions=['xlsx'],
    )
    sep.solicitacao_excel_s3_key = s3_key
    db.session.commit()

    return bytes_xlsx, s3_key


def regenerar_excel_qpa(separacao_id: int, operador_id: int, motivo: str) -> AssaiPedidoExcel:
    """Regenera Excel Q.P.A. de uma separação (D-C + S13=a + CR-9/CR-12).

    Wrapper sobre `gerar_excel_qpa` que aplica a regra D-C:
    - Desativa Excel anterior (versao--, ativo=False)
    - Cria nova versao (com lock CR-12 + retry CR-9 via SAVEPOINT)
    - Persiste em AssaiPedidoExcel (historico versionado)

    Concorrencia (CR-12): lock pessimista na sep para serializar regeneracoes
    concorrentes do mesmo Excel. Sem isso, 2 transacoes podem criar 2 Excels
    com mesma versao = IntegrityError.

    Retry (CR-9): IntegrityError ao inserir AssaiPedidoExcel (race entre
    SELECT MAX(versao) e INSERT) trata via SAVEPOINT — recalcula MAX e
    tenta novamente. Sem savepoint, rollback total perderia operacao caller.

    Args:
        separacao_id: id da AssaiSeparacao a regenerar Excel.
        operador_id: usuario que solicitou (auditoria).
        motivo: explicacao para historico (mostrado em /faturamento e auditoria).

    Returns:
        AssaiPedidoExcel ativo recem-criado (commit fica para o caller).

    Raises:
        ValueError: status invalido (gerar_excel_qpa).
    """
    # CR-12: lock pessimista na sep para serializar regeneracoes
    sep_locked = AssaiSeparacao.query.filter_by(id=separacao_id).with_for_update().first()
    if not sep_locked:
        raise FaturamentoError(f'Separação {separacao_id} não encontrada')

    excel_anterior = AssaiPedidoExcel.query.filter_by(
        separacao_id=separacao_id, ativo=True,
    ).first()
    if excel_anterior:
        excel_anterior.ativo = False  # mantem historico

    nova_versao = (excel_anterior.versao + 1) if excel_anterior else 1

    # gerar_excel_qpa retorna (bytes, s3_key) — internamente faz commit do
    # sep.solicitacao_excel_s3_key, mas isso é compatível com nosso flow.
    bytes_xlsx, s3_key = gerar_excel_qpa(separacao_id, operador_id)

    # CR-9 retry: SAVEPOINT isola o INSERT. Se falhar (race no MAX(versao)),
    # recalcula MAX e tenta novamente.
    try:
        with db.session.begin_nested():
            excel_novo = AssaiPedidoExcel(
                pedido_id=sep_locked.pedido_id,
                separacao_id=separacao_id,
                s3_key=s3_key, versao=nova_versao, ativo=True,
                motivo_regeneracao=motivo,
                gerado_por_id=operador_id,
            )
            db.session.add(excel_novo)
    except IntegrityError:
        # Savepoint reverteu o INSERT. Recalcula MAX e tenta de novo.
        max_versao = (db.session.query(db.func.coalesce(db.func.max(AssaiPedidoExcel.versao), 0))
                      .filter_by(separacao_id=separacao_id)
                      .scalar() or 0)
        nova_versao = max_versao + 1
        excel_novo = AssaiPedidoExcel(
            pedido_id=sep_locked.pedido_id,
            separacao_id=separacao_id,
            s3_key=s3_key, versao=nova_versao, ativo=True,
            motivo_regeneracao=f'{motivo} (retry)',
            gerado_por_id=operador_id,
        )
        db.session.add(excel_novo)
        db.session.flush()

    return excel_novo
