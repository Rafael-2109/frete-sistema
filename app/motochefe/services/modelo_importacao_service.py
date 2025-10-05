"""
Service para geração de modelos de importação Excel - MotoChefe
Centraliza a lógica de criação de arquivos modelo com instruções detalhadas
"""
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _formatar_cabecalho(ws):
    """Aplica formatação ao cabeçalho (primeira linha)"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border


def _formatar_exemplos(ws, num_linhas_exemplo=2):
    """Aplica formatação às linhas de exemplo"""
    example_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    example_alignment = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row_idx in range(2, 2 + num_linhas_exemplo):
        for cell in ws[row_idx]:
            cell.fill = example_fill
            cell.alignment = example_alignment
            cell.border = thin_border


def _ajustar_larguras_colunas(ws):
    """Ajusta largura das colunas baseado no conteúdo"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)  # Limite de 50
        ws.column_dimensions[column_letter].width = adjusted_width


def _criar_aba_instrucoes(workbook, titulo, instrucoes_linhas):
    """Cria aba de instruções formatada"""
    ws_instrucoes = workbook.create_sheet("INSTRUÇÕES")

    # Título
    ws_instrucoes['A1'] = titulo
    ws_instrucoes['A1'].font = Font(bold=True, size=14, color="4472C4")
    ws_instrucoes['A1'].alignment = Alignment(horizontal="left", vertical="center")
    ws_instrucoes.merge_cells('A1:D1')

    # Linha em branco
    linha_atual = 3

    # Instruções
    for instrucao in instrucoes_linhas:
        ws_instrucoes[f'A{linha_atual}'] = instrucao
        ws_instrucoes[f'A{linha_atual}'].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Negrito para títulos (linhas que terminam com :)
        if instrucao.strip().endswith(':'):
            ws_instrucoes[f'A{linha_atual}'].font = Font(bold=True, size=11)

        linha_atual += 1

    # Ajustar largura da coluna A
    ws_instrucoes.column_dimensions['A'].width = 100
    ws_instrucoes.row_dimensions[1].height = 20


# ====================================================================
# MODELOS ESPECÍFICOS POR ENTIDADE
# ====================================================================

def gerar_modelo_equipes():
    """Gera modelo de importação para Equipes de Vendas"""
    # Dados de exemplo
    data = {
        'Equipe': ['Equipe Sul', 'Equipe Norte']
    }

    df = pd.DataFrame(data)
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Modelo')

        workbook = writer.book
        ws_modelo = writer.sheets['Modelo']

        # Formatação
        _formatar_cabecalho(ws_modelo)
        _formatar_exemplos(ws_modelo, num_linhas_exemplo=2)
        _ajustar_larguras_colunas(ws_modelo)

        # Instruções
        instrucoes = [
            "📋 INSTRUÇÕES PARA IMPORTAÇÃO DE EQUIPES DE VENDAS",
            "",
            "✅ COLUNAS OBRIGATÓRIAS:",
            "  • Equipe - Nome da equipe de vendas (texto)",
            "",
            "📌 REGRAS DE VALIDAÇÃO:",
            "  • Nome da equipe não pode estar vazio",
            "  • Equipes duplicadas (mesmo nome) serão ignoradas",
            "  • Linhas vazias serão ignoradas automaticamente",
            "",
            "💡 EXEMPLO DE PREENCHIMENTO:",
            "  Equipe Sul",
            "  Equipe Norte",
            "  Equipe Centro-Oeste",
            "",
            "⚠️ ERROS COMUNS A EVITAR:",
            "  • Não altere o nome da coluna 'Equipe'",
            "  • Não deixe células em branco se houver equipe",
            "  • Evite caracteres especiais no início do nome",
            "",
            "🔄 PROCESSO DE IMPORTAÇÃO:",
            "  1. Preencha a planilha com os dados das equipes",
            "  2. Salve o arquivo no formato .xlsx",
            "  3. Na tela de Equipes, clique em 'Importar'",
            "  4. Selecione o arquivo e clique em 'Importar'",
            "  5. O sistema mostrará quantas equipes foram importadas",
        ]

        _criar_aba_instrucoes(workbook, "INSTRUÇÕES - IMPORTAÇÃO DE EQUIPES", instrucoes)

    output.seek(0)
    return output


def gerar_modelo_vendedores():
    """Gera modelo de importação para Vendedores"""
    data = {
        'Vendedor': ['João Silva', 'Maria Santos'],
        'Equipe': ['Equipe Sul', 'Equipe Norte']
    }

    df = pd.DataFrame(data)
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Modelo')

        workbook = writer.book
        ws_modelo = writer.sheets['Modelo']

        _formatar_cabecalho(ws_modelo)
        _formatar_exemplos(ws_modelo, num_linhas_exemplo=2)
        _ajustar_larguras_colunas(ws_modelo)

        instrucoes = [
            "📋 INSTRUÇÕES PARA IMPORTAÇÃO DE VENDEDORES",
            "",
            "✅ COLUNAS OBRIGATÓRIAS:",
            "  • Vendedor - Nome completo do vendedor (texto)",
            "",
            "📌 COLUNAS OPCIONAIS:",
            "  • Equipe - Nome da equipe de vendas (texto)",
            "    ⚠️ Se informada, a equipe DEVE existir previamente no sistema",
            "    ⚠️ Se não informada, vendedor será cadastrado sem equipe",
            "",
            "📌 REGRAS DE VALIDAÇÃO:",
            "  • Nome do vendedor não pode estar vazio",
            "  • Vendedores duplicados (mesmo nome) serão criados normalmente",
            "  • Se a equipe informada não existir, o vendedor será criado sem equipe",
            "  • Linhas vazias serão ignoradas automaticamente",
            "",
            "💡 EXEMPLO DE PREENCHIMENTO:",
            "  Vendedor: João Silva | Equipe: Equipe Sul",
            "  Vendedor: Maria Santos | Equipe: Equipe Norte",
            "  Vendedor: Pedro Oliveira | Equipe: (vazio - sem equipe)",
            "",
            "⚠️ ERROS COMUNS A EVITAR:",
            "  • Não altere os nomes das colunas",
            "  • Não deixe a coluna 'Vendedor' vazia",
            "  • Certifique-se que a equipe existe antes de importar",
            "  • Evite caracteres especiais no início do nome",
            "",
            "🔄 PROCESSO DE IMPORTAÇÃO:",
            "  1. Certifique-se que as equipes existem no sistema",
            "  2. Preencha a planilha com os dados dos vendedores",
            "  3. Salve o arquivo no formato .xlsx",
            "  4. Na tela de Vendedores, clique em 'Importar'",
            "  5. Selecione o arquivo e clique em 'Importar'",
            "  6. O sistema mostrará quantos vendedores foram importados",
        ]

        _criar_aba_instrucoes(workbook, "INSTRUÇÕES - IMPORTAÇÃO DE VENDEDORES", instrucoes)

    output.seek(0)
    return output


def gerar_modelo_transportadoras():
    """Gera modelo de importação para Transportadoras"""
    data = {
        'Transportadora': ['Transportadora Rápida Ltda', 'Logística Express S.A.'],
        'CNPJ': ['12.345.678/0001-99', '98.765.432/0001-11'],
        'Telefone': ['(11) 3333-4444', '(21) 5555-6666']
    }

    df = pd.DataFrame(data)
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Modelo')

        workbook = writer.book
        ws_modelo = writer.sheets['Modelo']

        _formatar_cabecalho(ws_modelo)
        _formatar_exemplos(ws_modelo, num_linhas_exemplo=2)
        _ajustar_larguras_colunas(ws_modelo)

        instrucoes = [
            "📋 INSTRUÇÕES PARA IMPORTAÇÃO DE TRANSPORTADORAS",
            "",
            "✅ COLUNAS OBRIGATÓRIAS:",
            "  • Transportadora - Razão social ou nome fantasia (texto)",
            "",
            "📌 COLUNAS OPCIONAIS:",
            "  • CNPJ - CNPJ da transportadora (texto com ou sem formatação)",
            "  • Telefone - Telefone de contato (texto com ou sem formatação)",
            "",
            "📌 REGRAS DE VALIDAÇÃO:",
            "  • Nome da transportadora não pode estar vazio",
            "  • Transportadoras duplicadas (mesmo nome) serão ignoradas",
            "  • CNPJ pode ser informado com ou sem pontuação (XX.XXX.XXX/XXXX-XX ou XXXXXXXXXXXXXX)",
            "  • Telefone pode ser informado com ou sem formatação ((XX) XXXX-XXXX ou XXXXXXXXXX)",
            "  • Linhas vazias serão ignoradas automaticamente",
            "",
            "💡 EXEMPLO DE PREENCHIMENTO:",
            "  Transportadora: Transportadora Rápida Ltda | CNPJ: 12.345.678/0001-99 | Telefone: (11) 3333-4444",
            "  Transportadora: Logística Express S.A. | CNPJ: 98765432000111 | Telefone: 2155556666",
            "  Transportadora: Frete Fácil ME | CNPJ: (vazio) | Telefone: (vazio)",
            "",
            "⚠️ ERROS COMUNS A EVITAR:",
            "  • Não altere os nomes das colunas",
            "  • Não deixe a coluna 'Transportadora' vazia",
            "  • Evite caracteres especiais no início do nome",
            "  • CNPJ e Telefone são opcionais, mas recomenda-se preencher",
            "",
            "🔄 PROCESSO DE IMPORTAÇÃO:",
            "  1. Preencha a planilha com os dados das transportadoras",
            "  2. Salve o arquivo no formato .xlsx",
            "  3. Na tela de Transportadoras, clique em 'Importar'",
            "  4. Selecione o arquivo e clique em 'Importar'",
            "  5. O sistema mostrará quantas transportadoras foram importadas",
        ]

        _criar_aba_instrucoes(workbook, "INSTRUÇÕES - IMPORTAÇÃO DE TRANSPORTADORAS", instrucoes)

    output.seek(0)
    return output


def gerar_modelo_clientes():
    """Gera modelo de importação para Clientes"""
    data = {
        'Cliente': ['Mercado Central Ltda', 'Distribuidora Norte S.A.'],
        'CNPJ': ['11.222.333/0001-44', '55.666.777/0001-88'],
        'Endereço': ['Rua das Flores', 'Av. Principal'],
        'Número': ['123', '456'],
        'Complemento': ['Loja 1', 'Galpão B'],
        'Bairro': ['Centro', 'Industrial'],
        'Cidade': ['São Paulo', 'Rio de Janeiro'],
        'Estado': ['SP', 'RJ'],
        'CEP': ['01234-567', '20000-000'],
        'Telefone': ['(11) 2222-3333', '(21) 4444-5555'],
        'Email': ['contato@mercado.com.br', 'vendas@distribuidora.com.br']
    }

    df = pd.DataFrame(data)
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Modelo')

        workbook = writer.book
        ws_modelo = writer.sheets['Modelo']

        _formatar_cabecalho(ws_modelo)
        _formatar_exemplos(ws_modelo, num_linhas_exemplo=2)
        _ajustar_larguras_colunas(ws_modelo)

        instrucoes = [
            "📋 INSTRUÇÕES PARA IMPORTAÇÃO DE CLIENTES",
            "",
            "✅ COLUNAS OBRIGATÓRIAS:",
            "  • Cliente - Razão social ou nome fantasia (texto)",
            "  • CNPJ - CNPJ do cliente (texto com ou sem formatação)",
            "",
            "📌 COLUNAS OPCIONAIS:",
            "  • Endereço - Logradouro (rua, avenida, etc)",
            "  • Número - Número do endereço",
            "  • Complemento - Complemento (apto, sala, loja, etc)",
            "  • Bairro - Nome do bairro",
            "  • Cidade - Nome da cidade",
            "  • Estado - Sigla do estado (2 letras: SP, RJ, MG, etc)",
            "  • CEP - CEP com ou sem formatação (XXXXX-XXX ou XXXXXXXX)",
            "  • Telefone - Telefone de contato",
            "  • Email - Email de contato",
            "",
            "📌 REGRAS DE VALIDAÇÃO:",
            "  • Nome do cliente e CNPJ são obrigatórios",
            "  • CNPJ deve ser único - CNPJs duplicados serão ignorados",
            "  • CNPJ pode ser informado com ou sem pontuação",
            "  • Estado deve ter exatamente 2 letras (SP, RJ, MG, etc)",
            "  • CEP pode ser informado com ou sem traço",
            "  • Todos os campos opcionais podem ficar em branco",
            "  • Linhas vazias serão ignoradas automaticamente",
            "",
            "💡 EXEMPLO DE PREENCHIMENTO:",
            "  Cliente: Mercado Central Ltda | CNPJ: 11.222.333/0001-44 | Cidade: São Paulo | Estado: SP",
            "  Cliente: Distribuidora Norte S.A. | CNPJ: 55666777000188 | CEP: 20000000",
            "",
            "⚠️ ERROS COMUNS A EVITAR:",
            "  • Não altere os nomes das colunas",
            "  • Não deixe 'Cliente' ou 'CNPJ' vazios",
            "  • Não use Estado com mais de 2 letras (use SP, não São Paulo)",
            "  • Certifique-se que o CNPJ está correto e não duplicado",
            "  • Evite caracteres especiais no início do nome",
            "",
            "🔄 PROCESSO DE IMPORTAÇÃO:",
            "  1. Preencha a planilha com os dados dos clientes",
            "  2. Salve o arquivo no formato .xlsx",
            "  3. Na tela de Clientes, clique em 'Importar'",
            "  4. Selecione o arquivo e clique em 'Importar'",
            "  5. O sistema mostrará quantos clientes foram importados",
            "  6. Clientes com CNPJ duplicado serão automaticamente ignorados",
        ]

        _criar_aba_instrucoes(workbook, "INSTRUÇÕES - IMPORTAÇÃO DE CLIENTES", instrucoes)

    output.seek(0)
    return output


def gerar_modelo_despesas():
    """Gera modelo de importação para Despesas Mensais"""
    data = {
        'Tipo': ['Aluguel', 'Energia Elétrica'],
        'Descrição': ['Aluguel sede Outubro/2025', 'Conta de luz Outubro/2025'],
        'Valor': ['5.000,00', '850,50'],  # Formato brasileiro
        'Mês': [10, 10],
        'Ano': [2025, 2025],
        'Status': ['PENDENTE', 'PENDENTE']
    }

    df = pd.DataFrame(data)
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Modelo')

        workbook = writer.book
        ws_modelo = writer.sheets['Modelo']

        _formatar_cabecalho(ws_modelo)
        _formatar_exemplos(ws_modelo, num_linhas_exemplo=2)
        _ajustar_larguras_colunas(ws_modelo)

        instrucoes = [
            "📋 INSTRUÇÕES PARA IMPORTAÇÃO DE DESPESAS MENSAIS",
            "",
            "✅ COLUNAS OBRIGATÓRIAS:",
            "  • Tipo - Tipo da despesa (texto): Aluguel, Energia, Telefone, etc",
            "  • Valor - Valor da despesa (formato brasileiro): 1.500,00",
            "  • Mês - Mês de competência (número de 1 a 12)",
            "  • Ano - Ano de competência (número com 4 dígitos): 2025",
            "",
            "📌 COLUNAS OPCIONAIS:",
            "  • Descrição - Descrição detalhada da despesa (texto)",
            "  • Status - Status da despesa (texto): PENDENTE, PAGO",
            "    ⚠️ Se não informado, será definido como 'PENDENTE'",
            "",
            "📌 REGRAS DE VALIDAÇÃO:",
            "  • Tipo, Valor, Mês e Ano são obrigatórios",
            "  • Valor deve usar vírgula como decimal e ponto como milhar: 1.500,00",
            "  • Mês deve ser número de 1 a 12",
            "  • Ano deve ser número com 4 dígitos (2025, 2026, etc)",
            "  • Status aceita: PENDENTE ou PAGO (padrão: PENDENTE)",
            "  • Linhas vazias serão ignoradas automaticamente",
            "",
            "💡 EXEMPLO DE PREENCHIMENTO:",
            "  Tipo: Aluguel | Valor: 5.000,00 | Mês: 10 | Ano: 2025 | Status: PENDENTE",
            "  Tipo: Energia Elétrica | Valor: 850,50 | Mês: 10 | Ano: 2025",
            "  Tipo: Telefone | Valor: 250,00 | Mês: 11 | Ano: 2025 | Descrição: Linha fixa + celular",
            "",
            "⚠️ ERROS COMUNS A EVITAR:",
            "  • Não altere os nomes das colunas",
            "  • Use VÍRGULA (,) como separador decimal e PONTO (.) como separador de milhar",
            "  • Mês deve ser NÚMERO (10), não texto (Outubro)",
            "  • Não deixe Tipo, Valor, Mês ou Ano vazios",
            "  • Status deve ser exatamente PENDENTE ou PAGO (maiúsculas)",
            "",
            "🔄 PROCESSO DE IMPORTAÇÃO:",
            "  1. Preencha a planilha com os dados das despesas",
            "  2. Salve o arquivo no formato .xlsx",
            "  3. Na tela de Despesas, clique em 'Importar'",
            "  4. Selecione o arquivo e clique em 'Importar'",
            "  5. O sistema mostrará quantas despesas foram importadas",
        ]

        _criar_aba_instrucoes(workbook, "INSTRUÇÕES - IMPORTAÇÃO DE DESPESAS MENSAIS", instrucoes)

    output.seek(0)
    return output


def gerar_modelo_modelos():
    """Gera modelo de importação para Modelos de Motos"""
    data = {
        'Modelo': ['X12', 'X11'],
        'Descrição': ['X12 autopropelida 1000W', 'X11 2000W Roda aro 10"'],
        'Potência': ['1000W', '2000W'],
        'Autopropelido': ['Sim', 'Não'],
        'Preço Tabela': ['12.500,00', '18.900,00']  # Formato brasileiro
    }

    df = pd.DataFrame(data)
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Modelo')

        workbook = writer.book
        ws_modelo = writer.sheets['Modelo']

        _formatar_cabecalho(ws_modelo)
        _formatar_exemplos(ws_modelo, num_linhas_exemplo=2)
        _ajustar_larguras_colunas(ws_modelo)

        instrucoes = [
            "📋 INSTRUÇÕES PARA IMPORTAÇÃO DE MODELOS DE MOTOS",
            "",
            "✅ COLUNAS OBRIGATÓRIAS:",
            "  • Modelo - Nome do modelo da moto (texto)",
            "  • Potência - Potência do motor (texto): 160cc, 250cc, etc",
            "  • Preço Tabela - Preço de tabela (formato brasileiro): 12.500,00",
            "",
            "📌 COLUNAS OPCIONAIS:",
            "  • Descrição - Descrição do modelo (texto)",
            "  • Autopropelido - Se a moto é autopropelida (Sim/Não)",
            "    ⚠️ Aceita: Sim, Yes, True, 1 para SIM",
            "    ⚠️ Qualquer outro valor será considerado NÃO",
            "",
            "📌 REGRAS DE VALIDAÇÃO:",
            "  • Modelo, Potência e Preço Tabela são obrigatórios",
            "  • Modelos duplicados (mesmo nome) serão ignorados",
            "  • Preço deve usar vírgula como decimal e ponto como milhar: 12.500,00",
            "  • Autopropelido é opcional, padrão é NÃO",
            "  • Linhas vazias serão ignoradas automaticamente",
            "",
            "💡 EXEMPLO DE PREENCHIMENTO:",
            "  Modelo: X12 | Potência: 1000W | Preço: 12.500,00 | Autopropelido: Sim",
            "  Modelo: X11 | Potência: 2000W | Preço: 18.900,00 | Autopropelido: Não",
            "",
            "⚠️ ERROS COMUNS A EVITAR:",
            "  • Não altere os nomes das colunas",
            "  • Use VÍRGULA (,) como separador decimal e PONTO (.) como separador de milhar",
            "  • Não deixe Modelo, Potência ou Preço Tabela vazios",
            "  • Evite cadastrar modelos duplicados",
            "  • Potência é campo de texto livre (pode usar cc, cv, HP, etc)",
            "",
            "🔄 PROCESSO DE IMPORTAÇÃO:",
            "  1. Preencha a planilha com os dados dos modelos",
            "  2. Salve o arquivo no formato .xlsx",
            "  3. Na tela de Modelos, clique em 'Importar'",
            "  4. Selecione o arquivo e clique em 'Importar'",
            "  5. O sistema mostrará quantos modelos foram importados",
            "  6. Modelos com nome duplicado serão automaticamente ignorados",
        ]

        _criar_aba_instrucoes(workbook, "INSTRUÇÕES - IMPORTAÇÃO DE MODELOS DE MOTOS", instrucoes)

    output.seek(0)
    return output


def gerar_modelo_motos():
    """Gera modelo de importação para Motos (Chassi)"""
    data = {
        'Chassi': ['9BD123456789ABC01', '9BD987654321XYZ02'],
        'Motor': ['MOT123456', 'MOT987654'],
        'Modelo': ['HONDA CG 160', 'YAMAHA FAZER 250'],
        'Cor': ['Preta', 'Vermelha'],
        'Ano': [2024, 2025],
        'NF Entrada': ['12345', '67890'],
        'Data NF': ['01/10/2025', '15/10/2025'],
        'Data Entrada': ['05/10/2025', '20/10/2025'],
        'Fornecedor': ['Honda Motos Ltda', 'Yamaha Brasil S.A.'],
        'Custo': ['11.000,00', '16.500,00'],  # Formato brasileiro
        'Pallet': ['P001', 'P002']
    }

    df = pd.DataFrame(data)
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Modelo')

        workbook = writer.book
        ws_modelo = writer.sheets['Modelo']

        _formatar_cabecalho(ws_modelo)
        _formatar_exemplos(ws_modelo, num_linhas_exemplo=2)
        _ajustar_larguras_colunas(ws_modelo)

        instrucoes = [
            "📋 INSTRUÇÕES PARA IMPORTAÇÃO DE MOTOS (CHASSI)",
            "",
            "✅ COLUNAS OBRIGATÓRIAS:",
            "  • Chassi - Número do chassi único (texto): 9BD123456789ABC01",
            "  • Motor - Número do motor único (texto): MOT123456",
            "  • Modelo - Nome do modelo (texto) - DEVE EXISTIR NO SISTEMA",
            "  • Cor - Cor da moto (texto): Preta, Vermelha, Azul, etc",
            "  • NF Entrada - Número da nota fiscal de entrada",
            "  • Fornecedor - Nome do fornecedor",
            "  • Custo - Custo de aquisição (formato brasileiro): 11.000,00",
            "",
            "📌 COLUNAS OPCIONAIS:",
            "  • Ano - Ano de fabricação (número): 2024, 2025",
            "  • Data NF - Data da nota fiscal (formato: DD/MM/AAAA)",
            "  • Data Entrada - Data de entrada no estoque (formato: DD/MM/AAAA)",
            "  • Pallet - Identificação do pallet (texto)",
            "",
            "📌 REGRAS DE VALIDAÇÃO:",
            "  • Chassi e Motor devem ser ÚNICOS - duplicados serão rejeitados",
            "  • Modelo informado DEVE EXISTIR previamente no sistema",
            "  • Custo deve usar vírgula como decimal e ponto como milhar: 11.000,00",
            "  • Data NF e Data Entrada devem estar no formato DD/MM/AAAA",
            "  • Se Data Entrada não for informada, será usada a data atual",
            "  • Linhas vazias serão ignoradas automaticamente",
            "",
            "💡 EXEMPLO DE PREENCHIMENTO:",
            "  Chassi: 9BD123456789ABC01 | Motor: MOT123456 | Modelo: HONDA CG 160",
            "  Cor: Preta | Ano: 2024 | NF: 12345 | Custo: 11.000,00",
            "",
            "⚠️ ERROS COMUNS A EVITAR:",
            "  • Não altere os nomes das colunas",
            "  • Certifique-se que o Modelo existe no sistema ANTES de importar",
            "  • Chassi e Motor devem ser únicos (não podem repetir)",
            "  • Use PONTO (.) como separador decimal no Custo",
            "  • Datas devem estar no formato DD/MM/AAAA (ex: 15/10/2025)",
            "  • Não deixe Chassi, Motor, Modelo, Cor, NF, Fornecedor ou Custo vazios",
            "",
            "🔄 PROCESSO DE IMPORTAÇÃO:",
            "  1. Certifique-se que os Modelos existem no sistema",
            "  2. Preencha a planilha com os dados das motos",
            "  3. Salve o arquivo no formato .xlsx",
            "  4. Na tela de Motos, clique em 'Importar'",
            "  5. Selecione o arquivo e clique em 'Importar'",
            "  6. O sistema mostrará quantas motos foram importadas e os erros (se houver)",
            "  7. Motos com Chassi ou Motor duplicado serão automaticamente rejeitadas",
            "  8. Motos com Modelo inexistente serão automaticamente rejeitadas",
        ]

        _criar_aba_instrucoes(workbook, "INSTRUÇÕES - IMPORTAÇÃO DE MOTOS (CHASSI)", instrucoes)

    output.seek(0)
    return output
