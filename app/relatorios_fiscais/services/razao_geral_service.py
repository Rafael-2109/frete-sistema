"""
Service: Razão Geral (General Ledger) do Odoo
==============================================

Extrai dados contábeis do Odoo (account.move.line) e gera Excel
formatado do relatório Razão Geral com saldo inicial para contas
patrimoniais.

Modelos Odoo utilizados:
- account.move.line (linhas de lançamento)
- account.account (plano de contas)

Autor: Sistema de Fretes
Data: 2026-01-23
"""

import logging
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ============================================================
# CONSTANTES
# ============================================================

# Empresas disponíveis para o relatório
EMPRESAS_RAZAO_GERAL = [
    {'id': 1, 'nome': 'NACOM GOYA - FB', 'sigla': 'FB'},
    {'id': 3, 'nome': 'NACOM GOYA - SC', 'sigla': 'SC'},
    {'id': 4, 'nome': 'NACOM GOYA - CD', 'sigla': 'CD'},
    {'id': 5, 'nome': 'LA FAMIGLIA - LF', 'sigla': 'LF'},
]

# Account types patrimoniais (recebem saldo inicial)
ACCOUNT_TYPES_PATRIMONIAIS = [
    'asset_receivable', 'asset_cash', 'asset_current',
    'asset_non_current', 'asset_prepayments', 'asset_fixed',
    'liability_payable', 'liability_credit_card',
    'liability_current', 'liability_non_current',
    'equity', 'equity_unaffected'
]

# Campos a buscar no account.move.line
CAMPOS_MOVE_LINE = [
    'date', 'move_name', 'account_id', 'partner_id',
    'ref', 'name', 'debit', 'credit', 'balance',
    'journal_id', 'matching_number'
]

# Configuração de paginação
BATCH_SIZE = 3000

# Estilos do Excel
FONT_TITULO = Font(name='Calibri', size=14, bold=True)
FONT_CABECALHO = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
FONT_NORMAL = Font(name='Calibri', size=10)
FONT_SALDO_INICIAL = Font(name='Calibri', size=10, italic=True, color='555555')

FILL_CABECALHO = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

BORDER_THIN = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

COLUNAS = ['Conta Contábil', 'Data', 'Lançamento', 'Diário', 'Parceiro',
           'Referência', 'Label', 'Débito', 'Crédito', 'Saldo Acumulado', 'Conciliação']

LARGURAS_COLUNAS = [30, 12, 22, 12, 35, 25, 40, 15, 15, 18, 15]


# ============================================================
# FUNÇÕES DE BUSCA NO ODOO
# ============================================================

def buscar_paginado(connection, model, domain, fields, batch_size=BATCH_SIZE, timeout=120):
    """
    Busca todos os registros com paginação via offset.
    Usa execute_kw diretamente para suportar limit/offset/order.

    Args:
        connection: OdooConnection autenticada
        model: Nome do modelo Odoo
        domain: Filtro de busca
        fields: Campos a retornar
        batch_size: Registros por página
        timeout: Timeout em segundos

    Returns:
        list: Todos os registros encontrados
    """
    todos = []
    offset = 0

    while True:
        kwargs = {
            'fields': fields,
            'limit': batch_size,
            'offset': offset,
            'order': 'account_id, date, id'
        }

        lote = connection.execute_kw(
            model, 'search_read', [domain], kwargs,
            timeout_override=timeout
        )

        if not lote:
            break

        todos.extend(lote)
        logger.info(f"   ... {len(todos)} registros buscados")

        if len(lote) < batch_size:
            break

        offset += batch_size

    return todos


def calcular_saldos_iniciais(connection, account_ids_patrimoniais, data_inicio, company_ids):
    """
    Calcula saldos iniciais para contas patrimoniais via read_group.
    Uma ÚNICA query agrupando por account_id.

    Args:
        connection: OdooConnection autenticada
        account_ids_patrimoniais: Lista de IDs de contas patrimoniais
        data_inicio: Data início do período (string YYYY-MM-DD)
        company_ids: Lista de company_ids

    Returns:
        dict: {account_id: {'debit': X, 'credit': Y, 'balance': Z}}
    """
    if not account_ids_patrimoniais:
        return {}

    domain = [
        ['account_id', 'in', account_ids_patrimoniais],
        ['date', '<', data_inicio],
        ['parent_state', '=', 'posted'],
        ['company_id', 'in', company_ids]
    ]

    result = connection.execute_kw(
        'account.move.line',
        'read_group',
        [domain],
        {
            'fields': ['account_id', 'debit:sum', 'credit:sum', 'balance:sum'],
            'groupby': ['account_id'],
            'lazy': False
        },
        timeout_override=180
    )

    saldos = {}
    for grupo in result:
        acc_id = grupo['account_id']
        if isinstance(acc_id, (list, tuple)):
            acc_id = acc_id[0]

        saldos[acc_id] = {
            'debit': grupo.get('debit', 0) or 0,
            'credit': grupo.get('credit', 0) or 0,
            'balance': grupo.get('balance', 0) or 0
        }

    return saldos


def buscar_movimentos_razao(connection, data_ini, data_fim, company_ids, conta_filter=None):
    """
    Busca todos os movimentos contábeis do período.

    Args:
        connection: OdooConnection autenticada
        data_ini: Data inicial (string YYYY-MM-DD)
        data_fim: Data final (string YYYY-MM-DD)
        company_ids: Lista de company_ids
        conta_filter: Código parcial da conta para filtrar (opcional)

    Returns:
        tuple: (registros, contas_info, saldos_iniciais)
    """
    # 1. Montar domain
    domain = [
        ['date', '>=', data_ini],
        ['date', '<=', data_fim],
        ['parent_state', '=', 'posted'],
        ['company_id', 'in', company_ids]
    ]

    # Se filtro de conta, buscar account_ids primeiro
    account_ids_filtro = None
    if conta_filter and conta_filter.strip():
        contas_filtradas = connection.execute_kw(
            'account.account', 'search_read',
            [[['code', 'ilike', conta_filter.strip()]]],
            {'fields': ['id'], 'limit': 1000}
        )
        if contas_filtradas:
            account_ids_filtro = [c['id'] for c in contas_filtradas]
            domain.append(['account_id', 'in', account_ids_filtro])
        else:
            return [], {}, {}

    # 2. Buscar movimentos paginado
    logger.info(f"Buscando account.move.line: {data_ini} a {data_fim}, companies={company_ids}")
    registros = buscar_paginado(connection, 'account.move.line', domain, CAMPOS_MOVE_LINE, BATCH_SIZE)
    logger.info(f"Total: {len(registros)} linhas contábeis")

    if not registros:
        return [], {}, {}

    # 3. Extrair account_ids únicos
    account_ids_unicos = list(set(
        r['account_id'][0] for r in registros
        if r.get('account_id') and isinstance(r['account_id'], (list, tuple))
    ))

    # 4. Buscar dados das contas
    contas = connection.execute_kw(
        'account.account', 'search_read',
        [[['id', 'in', account_ids_unicos]]],
        {'fields': ['id', 'code', 'name', 'account_type'], 'limit': 5000}
    )
    contas_info = {c['id']: c for c in contas}

    # 5. Identificar contas patrimoniais
    contas_patrimoniais = [
        c['id'] for c in contas
        if c.get('account_type') in ACCOUNT_TYPES_PATRIMONIAIS
    ]

    # 6. Calcular saldos iniciais
    saldos_iniciais = calcular_saldos_iniciais(
        connection, contas_patrimoniais, data_ini, company_ids
    )

    return registros, contas_info, saldos_iniciais


# ============================================================
# FUNÇÕES DE GERAÇÃO DO EXCEL
# ============================================================

def _formatar_data_br(data_str):
    """Converte YYYY-MM-DD para DD/MM/YYYY"""
    if not data_str:
        return ''
    try:
        dt = datetime.strptime(str(data_str), '%Y-%m-%d')
        return dt.strftime('%d/%m/%Y')
    except (ValueError, TypeError):
        return str(data_str)


def _normalizar_many2one(campo):
    """Extrai display_name de campo many2one [id, name]"""
    if campo and isinstance(campo, (list, tuple)) and len(campo) > 1:
        return campo[1]
    return ''


def _tratar_false(valor, tipo='str'):
    """Trata valores False do Odoo"""
    if valor is False or valor is None:
        return '' if tipo == 'str' else 0.0
    return valor


def gerar_excel_razao(registros, contas_info, saldos_iniciais, data_ini='', data_fim='', company_ids=None):
    """
    Gera Excel do Razão Geral com coluna de Conta Contábil.
    Retorna BytesIO pronto para download.

    Args:
        registros: Lista de account.move.line do Odoo
        contas_info: Dict {account_id: {code, name, account_type}}
        saldos_iniciais: Dict {account_id: {debit, credit, balance}}
        data_ini: Data inicial (para título)
        data_fim: Data final (para título)
        company_ids: Lista de companies (para título)

    Returns:
        BytesIO: Arquivo Excel em memória
    """
    if company_ids is None:
        company_ids = []

    # Agrupar movimentos por conta
    dados_agrupados = {}
    for reg in registros:
        acc_id = reg['account_id'][0] if isinstance(reg.get('account_id'), (list, tuple)) else None
        if acc_id is None:
            continue

        if acc_id not in dados_agrupados:
            dados_agrupados[acc_id] = []

        dados_agrupados[acc_id].append({
            'date': _tratar_false(reg.get('date')),
            'move_name': _tratar_false(reg.get('move_name')),
            'partner_name': _normalizar_many2one(reg.get('partner_id')),
            'ref': _tratar_false(reg.get('ref')),
            'name': _tratar_false(reg.get('name')),
            'debit': float(_tratar_false(reg.get('debit'), 'num')),
            'credit': float(_tratar_false(reg.get('credit'), 'num')),
            'balance': float(_tratar_false(reg.get('balance'), 'num')),
            'journal_name': _normalizar_many2one(reg.get('journal_id')),
            'matching_number': _tratar_false(reg.get('matching_number'))
        })

    # Ordenar movimentos por data dentro de cada conta
    for acc_id in dados_agrupados:
        dados_agrupados[acc_id].sort(key=lambda x: (x['date'], x['move_name']))

    # --- Gerar Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = 'Razão Geral'

    # Título
    titulo = f'RAZÃO GERAL'
    if data_ini and data_fim:
        titulo += f' - Período: {_formatar_data_br(data_ini)} a {_formatar_data_br(data_fim)}'
    ws.append([titulo])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUNAS))
    ws.cell(row=1, column=1).font = FONT_TITULO
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    # Subtítulo
    empresas_nomes = [e['nome'] for e in EMPRESAS_RAZAO_GERAL if e['id'] in (company_ids or [])]
    subtitulo = f'Empresas: {", ".join(empresas_nomes)}' if empresas_nomes else 'Todas as empresas'
    ws.append([subtitulo + ' | Status: Posted'])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUNAS))
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')

    ws.append([])  # Linha em branco

    # Cabeçalho das colunas
    ws.append(COLUNAS)
    row_num = 4
    for col_idx in range(1, len(COLUNAS) + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.font = FONT_CABECALHO
        cell.fill = FILL_CABECALHO
        cell.alignment = Alignment(horizontal='center')
        cell.border = BORDER_THIN

    # Larguras das colunas
    for col_idx, largura in enumerate(LARGURAS_COLUNAS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    row_num += 1

    # Contas ordenadas por código
    contas_ordenadas = sorted(
        dados_agrupados.keys(),
        key=lambda x: contas_info.get(x, {}).get('code', '999999')
    )

    for account_id in contas_ordenadas:
        conta = contas_info.get(account_id, {})
        code = conta.get('code', '???')
        name = conta.get('name', 'Conta desconhecida')
        conta_label = f'{code} - {name}'
        movimentos = dados_agrupados[account_id]

        # Saldo inicial (se patrimonial)
        saldo_acumulado = 0.0
        if account_id in saldos_iniciais:
            si = saldos_iniciais[account_id]
            saldo_acumulado = si['balance']

            ws.append([
                conta_label, '', '', '', '', '', 'Saldo Inicial',
                si['debit'], si['credit'], saldo_acumulado, ''
            ])
            for col_idx in range(1, len(COLUNAS) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = FONT_SALDO_INICIAL
                cell.border = BORDER_THIN
            ws.cell(row=row_num, column=8).number_format = '#,##0.00'
            ws.cell(row=row_num, column=9).number_format = '#,##0.00'
            ws.cell(row=row_num, column=10).number_format = '#,##0.00'
            row_num += 1

        # Movimentos do período
        for mov in movimentos:
            saldo_acumulado += mov['balance']

            ws.append([
                conta_label,
                _formatar_data_br(mov['date']),
                mov['move_name'],
                mov['journal_name'],
                mov['partner_name'],
                mov['ref'],
                mov['name'],
                mov['debit'],
                mov['credit'],
                saldo_acumulado,
                mov['matching_number']
            ])

            for col_idx in range(1, len(COLUNAS) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = FONT_NORMAL
                cell.border = BORDER_THIN

            ws.cell(row=row_num, column=8).number_format = '#,##0.00'
            ws.cell(row=row_num, column=9).number_format = '#,##0.00'
            ws.cell(row=row_num, column=10).number_format = '#,##0.00'
            row_num += 1

    # Salvar em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info(f"Excel gerado: {row_num - 5} linhas de dados")
    return output
