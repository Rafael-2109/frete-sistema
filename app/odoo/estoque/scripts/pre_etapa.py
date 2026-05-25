"""PreEtapaEstoqueService — planejador da pre-etapa CD/FB (D007).

Substitui NFs inter-filial CD↔FB (R$ 32,9 mi) e INDISPONIBILIZAR_*
(R$ 60,5 mi) por transferencias INTERNAS na company + residual minimo.

Algoritmo (por produto):
    1. Agrega linhas inv por lote (lote vazio → 'P-15/05')
    2. Calcula custo medio ponderado (D004)
    3. Identifica doadores:
       - Quants de lote NAO em inv: doam total
       - Quants de lote em inv com saldo > desejada: doam excesso
    4. Aloca doadores FIFO (quant_id) para cobrir deficits dos lotes alvo (POS)
    5. Sobras restantes dos doadores → MIGRAÇÃO da propria company (NEG)
    6. Deficits restantes:
       - Tenta puxar da FB (residual_fb_cd) — gera NF FB→CD (CFOP 5152)
       - Senao: ajuste positivo puro (inventory adjustment, sem origem)

Executor (`09b_executar_pre_etapa.py`) consome o plano via:
- StockInternalTransferService.transferir_entre_lotes (D006) para internas
- StockPickingService.criar_transferencia para residual_fb_cd
- stock.quant.action_apply_inventory direto para ajuste positivo puro

Restricao temporal (D007):
- ENQUANTO FB nao tiver passado pela propria pre-etapa, qualquer lote FB
  pode doar para CD.
- APOS pre-etapa FB, so o lote MIGRAÇÃO da FB pode doar (preserva
  corretude FB) — caller deve filtrar `quants_fb_disponivel` ao chamar.

Spec: docs/inventario-2026-05/00-decisoes/D007-pre-etapa-cd-fb-minimizar-nf.md
"""
import hashlib
import logging
import os
import subprocess
from collections import defaultdict
from dataclasses import dataclass, field
from decimal import Decimal
from typing import Any, Dict, List, Optional

from app.odoo.utils.connection import get_odoo_connection

logger = logging.getLogger(__name__)

# ============================================================
# Constantes do workflow Pre-Etapa (Skill 6 — capinada 2026-05-24)
# ============================================================

ACOES_INTERNAS_POR_CID: Dict[int, Dict[str, str]] = {
    4: {  # CD (Onda 5)
        'POS': 'AJUSTE_CD_TRANSF_INTERNA_POS',
        'NEG': 'AJUSTE_CD_TRANSF_INTERNA_NEG',
        'PURO': 'AJUSTE_CD_POSITIVO_PURO',
    },
    1: {  # FB (Onda 6 futura)
        'POS': 'AJUSTE_FB_TRANSF_INTERNA_POS',
        'NEG': 'AJUSTE_FB_TRANSF_INTERNA_NEG',
        'PURO': 'AJUSTE_FB_POSITIVO_PURO',
    },
}
ONDA_NUM_POR_CID: Dict[int, int] = {4: 5, 1: 6}
ACAO_RESIDUAL_FB_CD: str = 'TRANSFERIR_FB_CD'  # CFOP 5152 (so CD)
COMPANY_LOCATIONS_PRE_ETAPA: Dict[int, int] = {4: 32, 1: 8}  # principal por CID

# Mapeamento curto para auditoria (VARCHAR(20) constraint em
# operacao_odoo_auditoria.acao). Gerado a partir de ACOES_INTERNAS_POR_CID
# para garantir sincronia (CR-PATTERN-2 v9 — fonte unica). Preserva os
# nomes usados pelo 09b legacy (compatibilidade com registros historicos
# em operacao_odoo_auditoria).
_PREFIXO_CID_AUDIT = {4: 'cd', 1: 'fb'}
_ABREV_OP_AUDIT = {'POS': 'pre_pos', 'NEG': 'pre_neg', 'PURO': 'pos_puro'}
ACAO_AUDIT_CURTA: Dict[str, str] = {
    acao: f'{_PREFIXO_CID_AUDIT[cid]}_{_ABREV_OP_AUDIT[op_long]}'
    for cid, acoes in ACOES_INTERNAS_POR_CID.items()
    for op_long, acao in acoes.items()
}


@dataclass
class TransferenciaInternaPlanejada:
    """Transferencia interna planejada (lote → lote, mesma company)."""
    cod_produto: str
    company_id: int
    location_id: int
    qty: float
    lot_id_origem: Optional[int]
    lote_origem_nome: str
    lote_destino_nome: str
    tipo: str  # 'POS' (preencher lote alvo) ou 'NEG' (consolidar em MIGRAÇÃO)
    custo_medio: Decimal


@dataclass
class ResidualFbCdPlanejado:
    """Residual que precisa vir da FB para o CD via NF (CFOP 5152)."""
    cod_produto: str
    qty: float
    custo_medio: Decimal
    lote_origem_fb_sugerido: str
    lote_destino_cd_nome: str


@dataclass
class AjustePositivoPuroPlanejado:
    """Ajuste positivo direto (sem doador), via inventory adjustment."""
    cod_produto: str
    company_id: int
    location_id: int
    qty: float
    lote_destino_nome: str
    custo_medio: Decimal


@dataclass
class PlanoPreEtapa:
    """Plano consolidado da pre-etapa por produto."""
    transferencias_internas: List[TransferenciaInternaPlanejada] = field(
        default_factory=list
    )
    residual_fb_cd: List[ResidualFbCdPlanejado] = field(default_factory=list)
    ajustes_positivos_puros: List[AjustePositivoPuroPlanejado] = field(
        default_factory=list
    )
    warnings: List[str] = field(default_factory=list)


class PreEtapaEstoqueService:
    """Planejador da pre-etapa CD/FB (D007)."""

    LOTE_DEFAULT_SEM_NOME = 'P-15/05'
    LOTE_CONSOLIDADOR_NEG = 'MIGRAÇÃO'

    def __init__(self, odoo=None):
        self.odoo = odoo or get_odoo_connection()

    # ============================================================
    # Helpers de agregacao
    # ============================================================

    @staticmethod
    def _custo_medio_cod(quants: List[Dict[str, Any]]) -> Decimal:
        """Media ponderada `value/quantity` por produto (D004)."""
        total_val = Decimal('0')
        total_qty = Decimal('0')
        for q in quants:
            qty = Decimal(str(q.get('quantity', 0) or 0))
            val = Decimal(str(q.get('value', 0) or 0))
            if qty > 0 and val != 0:
                total_val += val
                total_qty += qty
        return total_val / total_qty if total_qty > 0 else Decimal('0')

    def _agregar_inv_por_lote(
        self, linhas_inv: List[Dict[str, Any]]
    ) -> Dict[str, float]:
        """Soma qtd_inventario por nome de lote.

        Linhas sem lote sao agregadas em LOTE_DEFAULT_SEM_NOME ('P-15/05').
        """
        agg: Dict[str, float] = {}
        for linha in linhas_inv:
            nome = (linha.get('lote_inventariado') or '').strip()
            if not nome:
                nome = self.LOTE_DEFAULT_SEM_NOME
            qty = float(linha.get('qtd_inventario', 0) or 0)
            agg[nome] = agg.get(nome, 0.0) + qty
        return agg

    # ============================================================
    # Algoritmo principal
    # ============================================================

    def planejar(
        self,
        cod_produto: str,
        company_id: int,
        location_id: int,
        quants_odoo: List[Dict[str, Any]],
        linhas_inv: List[Dict[str, Any]],
        quants_fb_disponivel: Optional[List[Dict[str, Any]]] = None,
    ) -> PlanoPreEtapa:
        """Plano da pre-etapa para 1 produto.

        Args:
            cod_produto: default_code Odoo (informacional no plano).
            company_id: 4 (CD) ou 1 (FB).
            location_id: location interna principal da company
                (32 CD, 8 FB, 42 LF) — usada apenas para ajuste positivo puro.
            quants_odoo: lista de dicts da company com chaves:
                quant_id, lot_id, lote_nome (str, '' se sem lote),
                quantity, reserved_quantity, location_id, value.
            linhas_inv: lista de dicts do inventario com chaves:
                lote_inventariado (str, '' se sem lote), qtd_inventario.
            quants_fb_disponivel: para CD, quants da FB disponiveis para
                cobrir residual positivo (opcional). Se None ou vazia,
                deficits restantes viram ajuste positivo puro. Caller
                aplica restricao temporal (qualquer lote ou so MIGRAÇÃO).

        Returns:
            PlanoPreEtapa com 4 listas: transferencias_internas (POS+NEG),
            residual_fb_cd, ajustes_positivos_puros, warnings.
        """
        plano = PlanoPreEtapa()

        # 1. Agregar inv por lote
        inv_por_lote = self._agregar_inv_por_lote(linhas_inv)

        # 2. Agregar Odoo por lote
        odoo_por_lote: Dict[str, List[Dict[str, Any]]] = {}
        for q in quants_odoo:
            nome = (q.get('lote_nome') or '').strip()
            odoo_por_lote.setdefault(nome, []).append(q)

        # 3. Custo medio ponderado
        custo_medio = self._custo_medio_cod(quants_odoo)

        # 4. Saldo total atual por lote
        saldo_por_lote: Dict[str, float] = {
            nome: sum(float(q['quantity']) for q in quants)
            for nome, quants in odoo_por_lote.items()
        }

        # 5. Identificar doadores
        # Estrutura: {'quant': dict, 'qty_disponivel': float, 'qty_original': float}
        doadores: List[Dict[str, Any]] = []
        for nome, quants in odoo_por_lote.items():
            if nome == self.LOTE_CONSOLIDADOR_NEG and nome not in inv_por_lote:
                # MIGRAÇÃO ja consolidador: nao move (a menos que inv pida)
                continue
            if nome not in inv_por_lote:
                # Lote nao-alvo: doa total
                for q in quants:
                    qty_q = float(q['quantity'])
                    if qty_q > 0:
                        doadores.append({
                            'quant': q,
                            'qty_disponivel': qty_q,
                            'qty_original': qty_q,
                        })
            else:
                # Lote alvo com possivel excesso
                saldo = saldo_por_lote[nome]
                qty_desejada = inv_por_lote[nome]
                excesso = saldo - qty_desejada
                if excesso > 0:
                    excesso_rest = excesso
                    quants_ord = sorted(
                        quants, key=lambda x: x['quant_id']
                    )
                    for q in quants_ord:
                        if excesso_rest <= 0:
                            break
                        qty_doar = min(float(q['quantity']), excesso_rest)
                        if qty_doar > 0:
                            doadores.append({
                                'quant': q,
                                'qty_disponivel': qty_doar,
                                'qty_original': qty_doar,
                            })
                            excesso_rest -= qty_doar

        # 6. Calcular deficits
        deficits: Dict[str, float] = {}
        for nome, qty_desejada in inv_por_lote.items():
            saldo = saldo_por_lote.get(nome, 0.0)
            if saldo < qty_desejada:
                deficits[nome] = qty_desejada - saldo

        # 7. Doadores FIFO preenchem deficits (POS)
        doadores.sort(key=lambda d: d['quant']['quant_id'])
        # Lotes alvo ordenados para determinismo
        for lote_alvo in sorted(deficits.keys()):
            deficit_rest = deficits[lote_alvo]
            for doador in doadores:
                if deficit_rest <= 0:
                    break
                if doador['qty_disponivel'] <= 0:
                    continue
                qty_take = min(doador['qty_disponivel'], deficit_rest)
                q = doador['quant']
                plano.transferencias_internas.append(
                    TransferenciaInternaPlanejada(
                        cod_produto=cod_produto,
                        company_id=company_id,
                        location_id=int(q.get('location_id') or location_id),
                        qty=qty_take,
                        lot_id_origem=q.get('lot_id'),
                        lote_origem_nome=(q.get('lote_nome') or ''),
                        lote_destino_nome=lote_alvo,
                        tipo='POS',
                        custo_medio=custo_medio,
                    )
                )
                doador['qty_disponivel'] -= qty_take
                deficit_rest -= qty_take
            deficits[lote_alvo] = deficit_rest

        # 8. Sobras dos doadores → MIGRAÇÃO (NEG)
        for doador in doadores:
            qty_sobra = doador['qty_disponivel']
            if qty_sobra <= 0:
                continue
            q = doador['quant']
            origem_nome = (q.get('lote_nome') or '')
            # Skip se ja e MIGRAÇÃO (nao mover para si mesmo)
            if origem_nome == self.LOTE_CONSOLIDADOR_NEG:
                continue
            plano.transferencias_internas.append(
                TransferenciaInternaPlanejada(
                    cod_produto=cod_produto,
                    company_id=company_id,
                    location_id=int(q.get('location_id') or location_id),
                    qty=qty_sobra,
                    lot_id_origem=q.get('lot_id'),
                    lote_origem_nome=origem_nome,
                    lote_destino_nome=self.LOTE_CONSOLIDADOR_NEG,
                    tipo='NEG',
                    custo_medio=custo_medio,
                )
            )
            doador['qty_disponivel'] = 0.0

        # 9. Warnings de reserva (apos calcular doacoes finais)
        for doador in doadores:
            q = doador['quant']
            qty_doada = doador['qty_original'] - doador['qty_disponivel']
            if qty_doada <= 0:
                continue
            reservada = float(q.get('reserved_quantity', 0) or 0)
            qty_restante = float(q['quantity']) - qty_doada
            if qty_restante < reservada:
                plano.warnings.append(
                    f'quant {q.get("quant_id")} ({cod_produto}, lote='
                    f'{q.get("lote_nome") or "(sem lote)"}): tem {reservada} '
                    f'un reservadas em pickings ativos; saldo apos doacao '
                    f'({qty_restante}) ficaria < reservada. Pode bloquear '
                    f'a execucao via StockInternalTransferService.'
                )

        # 10. Deficits restantes: FB → ajuste positivo puro
        for lote_alvo in sorted(deficits.keys()):
            deficit = deficits[lote_alvo]
            if deficit <= 0:
                continue
            if quants_fb_disponivel:
                disponivel_fb = sum(
                    float(q.get('quantity', 0) or 0)
                    for q in quants_fb_disponivel
                )
                qty_de_fb = min(disponivel_fb, deficit)
                if qty_de_fb > 0:
                    doador_fb = max(
                        quants_fb_disponivel,
                        key=lambda q: float(q.get('quantity', 0) or 0),
                    )
                    plano.residual_fb_cd.append(ResidualFbCdPlanejado(
                        cod_produto=cod_produto,
                        qty=qty_de_fb,
                        custo_medio=custo_medio,
                        lote_origem_fb_sugerido=(
                            doador_fb.get('lote_nome') or ''
                        ),
                        lote_destino_cd_nome=lote_alvo,
                    ))
                    deficit -= qty_de_fb
            if deficit > 0:
                plano.ajustes_positivos_puros.append(
                    AjustePositivoPuroPlanejado(
                        cod_produto=cod_produto,
                        company_id=company_id,
                        location_id=location_id,
                        qty=deficit,
                        lote_destino_nome=lote_alvo,
                        custo_medio=custo_medio,
                    )
                )

        return plano


# ============================================================
# Helpers I/O e workflow (capinada de 03b + 04b em 2026-05-24)
# ============================================================
# Convencao: funcoes top-level orquestram o PreEtapaEstoqueService + I/O
# (JSON, XLSX, banco local) reproduzindo 03b_planejar_pre_etapa_cd.py e
# 04b_propor_pre_etapa_cd.py. Lazy import de Flask/openpyxl onde aplicavel
# para que tests do service puro (PreEtapaEstoqueService) nao exijam
# app_context.


def enriquecer_quants_para_planejar(
    odoo,
    quants_raw: List[Dict[str, Any]],
    label: str = '',
) -> List[Dict[str, Any]]:
    """Enriquece quants do JSON (script 01 output) com cod_produto, lote_nome, value.

    NAO busca reserved_quantity (custoso, nao bloqueante — runtime valida via
    PreEtapaEstoqueService warnings). Espelha enriquecer_quants do 03b.

    Args:
        odoo: conexao Odoo.
        quants_raw: lista de dicts vindos do JSON do script 01 com keys
            id, product_id (tuple), lot_id (tuple), location_id (tuple),
            quantity, value.
        label: rotulo para log (ex.: 'CD', 'FB').

    Returns:
        lista enriquecida com keys: quant_id, product_id, cod_produto,
        nome_produto, lot_id, lote_nome, location_id, quantity,
        reserved_quantity (sempre 0), value.
    """
    product_ids = list({
        q['product_id'][0] for q in quants_raw if q.get('product_id')
    })
    lot_ids = list({
        q['lot_id'][0] for q in quants_raw if q.get('lot_id')
    })

    if label:
        logger.info(
            f'enriquecer_quants[{label}]: lendo {len(product_ids)} produtos '
            f'+ {len(lot_ids)} lotes...'
        )
    produtos = {p['id']: p for p in odoo.read(
        'product.product', product_ids, ['default_code', 'name'],
    )} if product_ids else {}
    lotes = {lo['id']: lo for lo in odoo.read(
        'stock.lot', lot_ids, ['name'],
    )} if lot_ids else {}

    out = []
    for q in quants_raw:
        pid = q['product_id'][0] if q.get('product_id') else None
        lid = q['lot_id'][0] if q.get('lot_id') else None
        produto = produtos.get(pid, {})
        lote = lotes.get(lid, {}) if lid else {}
        out.append({
            'quant_id': q['id'],
            'product_id': pid,
            'cod_produto': (produto.get('default_code') or '').strip(),
            'nome_produto': produto.get('name') or '',
            'lot_id': lid,
            'lote_nome': lote.get('name') or '',
            'location_id': (
                q['location_id'][0] if q.get('location_id') else None
            ),
            'quantity': float(q['quantity']),
            'reserved_quantity': 0.0,  # nao bloqueante — runtime valida
            'value': float(q.get('value', 0)),
        })
    return out


def _serializar_plano_em_dicts(
    plano: PlanoPreEtapa,
    dest: Dict[str, List[Any]],
) -> None:
    """Append dataclasses do PlanoPreEtapa em dest consolidado (json-safe).

    Decimals -> str. Espelha plano_to_dicts do 03b.
    """
    for t in plano.transferencias_internas:
        dest['transferencias_internas'].append({
            'cod_produto': t.cod_produto,
            'company_id': t.company_id,
            'location_id': t.location_id,
            'qty': t.qty,
            'lot_id_origem': t.lot_id_origem,
            'lote_origem_nome': t.lote_origem_nome,
            'lote_destino_nome': t.lote_destino_nome,
            'tipo': t.tipo,
            'custo_medio': str(t.custo_medio),
        })
    for r in plano.residual_fb_cd:
        dest['residual_fb_cd'].append({
            'cod_produto': r.cod_produto,
            'qty': r.qty,
            'custo_medio': str(r.custo_medio),
            'lote_origem_fb_sugerido': r.lote_origem_fb_sugerido,
            'lote_destino_cd_nome': r.lote_destino_cd_nome,
        })
    for a in plano.ajustes_positivos_puros:
        dest['ajustes_positivos_puros'].append({
            'cod_produto': a.cod_produto,
            'company_id': a.company_id,
            'location_id': a.location_id,
            'qty': a.qty,
            'lote_destino_nome': a.lote_destino_nome,
            'custo_medio': str(a.custo_medio),
        })
    for w in plano.warnings:
        dest['warnings'].append(w)


def gerar_excel_plano_pre_etapa(
    plano_total: Dict[str, List[Any]],
    path: str,
    cod_to_name: Dict[str, str],
) -> None:
    """Gera Excel 4 abas (Internas, Residual FB-CD, Positivos Puros, Warnings).

    Espelha gerar_excel do 03b. Lazy import openpyxl.
    """
    import openpyxl  # lazy

    wb = openpyxl.Workbook()

    # Aba 1: Transferencias internas
    ws = wb.active
    ws.title = 'Internas'
    ws.append([
        'cod_produto', 'nome_produto', 'company_id', 'location_id', 'qty',
        'lot_id_origem', 'lote_origem', 'lote_destino', 'tipo', 'custo_medio',
        'valor_movimentacao',
    ])
    for t in plano_total['transferencias_internas']:
        cm = float(t['custo_medio'])
        ws.append([
            t['cod_produto'], cod_to_name.get(t['cod_produto'], ''),
            t['company_id'], t['location_id'], t['qty'],
            t['lot_id_origem'] or '', t['lote_origem_nome'],
            t['lote_destino_nome'], t['tipo'], cm, t['qty'] * cm,
        ])
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=5).number_format = '#,##0.0000'
        ws.cell(row=row, column=10).number_format = '#,##0.00'
        ws.cell(row=row, column=11).number_format = '#,##0.00'

    # Aba 2: Residual FB->CD
    ws2 = wb.create_sheet('Residual FB-CD')
    ws2.append([
        'cod_produto', 'nome_produto', 'qty', 'custo_medio',
        'lote_origem_fb_sugerido', 'lote_destino_cd_nome',
        'valor_movimentacao',
    ])
    for r in plano_total['residual_fb_cd']:
        cm = float(r['custo_medio'])
        ws2.append([
            r['cod_produto'], cod_to_name.get(r['cod_produto'], ''),
            r['qty'], cm,
            r['lote_origem_fb_sugerido'], r['lote_destino_cd_nome'],
            r['qty'] * cm,
        ])
    for row in range(2, ws2.max_row + 1):
        ws2.cell(row=row, column=3).number_format = '#,##0.0000'
        ws2.cell(row=row, column=4).number_format = '#,##0.00'
        ws2.cell(row=row, column=7).number_format = '#,##0.00'

    # Aba 3: Ajustes positivos puros
    ws3 = wb.create_sheet('Positivos Puros')
    ws3.append([
        'cod_produto', 'nome_produto', 'company_id', 'location_id', 'qty',
        'lote_destino', 'custo_medio', 'valor_movimentacao',
    ])
    for a in plano_total['ajustes_positivos_puros']:
        cm = float(a['custo_medio'])
        ws3.append([
            a['cod_produto'], cod_to_name.get(a['cod_produto'], ''),
            a['company_id'], a['location_id'], a['qty'],
            a['lote_destino_nome'], cm, a['qty'] * cm,
        ])
    for row in range(2, ws3.max_row + 1):
        ws3.cell(row=row, column=5).number_format = '#,##0.0000'
        ws3.cell(row=row, column=7).number_format = '#,##0.00'
        ws3.cell(row=row, column=8).number_format = '#,##0.00'

    # Aba 4: Warnings
    ws4 = wb.create_sheet('Warnings')
    ws4.append(['mensagem'])
    for w in plano_total['warnings']:
        ws4.append([w])

    wb.save(path)


def planejar_pre_etapa_batch_company(
    odoo,
    company_id: int,
    location_id: int,
    quants_company_raw: List[Dict[str, Any]],
    linhas_inv_raw: List[Dict[str, Any]],
    quants_complementar_raw: Optional[List[Dict[str, Any]]] = None,
    cods_filter: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """Roda PreEtapaEstoqueService.planejar() em batch para 1 company.

    Espelha main() de 03b_planejar_pre_etapa_cd. Retorna dict consolidado.

    Args:
        odoo: conexao Odoo.
        company_id: 4 (CD) ou 1 (FB).
        location_id: location principal (32 CD, 8 FB) — usada em positivos puros.
        quants_company_raw: lista raw da company alvo (saida do script 01).
        linhas_inv_raw: linhas do inventario (saida do script 02).
        quants_complementar_raw: para CD, quants raw da FB para cobrir residual.
            Para FB (Onda 6), None ou vazia.
        cods_filter: cods especificos (None = todos validos com cod[0] in 1-4).

    Returns:
        dict com keys: transferencias_internas, residual_fb_cd,
        ajustes_positivos_puros, warnings, produtos_processados,
        produtos_sem_mudanca, cod_to_name, outliers_skipados, company_id.
    """
    quants_alvo = enriquecer_quants_para_planejar(
        odoo, quants_company_raw,
        label=f'cid={company_id}',
    )
    quants_complementar = (
        enriquecer_quants_para_planejar(
            odoo, quants_complementar_raw, label='complementar',
        )
        if quants_complementar_raw else []
    )

    alvo_por_cod: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for q in quants_alvo:
        if q['cod_produto']:
            alvo_por_cod[q['cod_produto']].append(q)
    complementar_por_cod: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for q in quants_complementar:
        if q['cod_produto']:
            complementar_por_cod[q['cod_produto']].append(q)
    inv_por_cod: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for linha in linhas_inv_raw:
        cod = linha.get('cod_produto', '')
        if cod:
            inv_por_cod[cod].append(linha)

    cod_to_name = {
        q['cod_produto']: q['nome_produto']
        for q in (quants_alvo + quants_complementar)
        if q['cod_produto']
    }

    cods_universo = set(inv_por_cod.keys()) | set(alvo_por_cod.keys())
    if cods_filter:
        cods_universo = cods_universo & set(cods_filter)
    cods_validos = [
        c for c in cods_universo
        if c and c[0].isdigit() and int(c[0]) in (1, 2, 3, 4)
    ]
    outliers = sorted([c for c in cods_universo if c not in cods_validos])

    svc = PreEtapaEstoqueService(odoo=odoo)
    plano_total: Dict[str, List[Any]] = {
        'transferencias_internas': [],
        'residual_fb_cd': [],
        'ajustes_positivos_puros': [],
        'warnings': [],
    }
    n_processados = 0
    n_sem_mudanca = 0
    for cod in sorted(cods_validos):
        quants_cod = alvo_por_cod.get(cod, [])
        inv_cod = inv_por_cod.get(cod, [])
        complementar_cod = complementar_por_cod.get(cod, [])

        if not quants_cod and not inv_cod:
            continue

        plano = svc.planejar(
            cod_produto=cod,
            company_id=company_id,
            location_id=location_id,
            quants_odoo=quants_cod,
            linhas_inv=inv_cod,
            quants_fb_disponivel=complementar_cod or None,
        )
        if not (
            plano.transferencias_internas
            or plano.residual_fb_cd
            or plano.ajustes_positivos_puros
        ):
            n_sem_mudanca += 1
            continue
        n_processados += 1
        # Prefix cod nos warnings para rastrear
        plano.warnings = [f'[{cod}] {w}' for w in plano.warnings]
        _serializar_plano_em_dicts(plano, plano_total)

    return {
        **plano_total,
        'produtos_processados': n_processados,
        'produtos_sem_mudanca': n_sem_mudanca,
        'cod_to_name': cod_to_name,
        'outliers_skipados': outliers,
        'company_id': company_id,
    }


def _calcular_hash_onda(ajustes: List[Any]) -> str:
    """Hash sha256 do payload da onda (anti-replay). Espelha 04b.

    Acessa campos via getattr tolerante (CR-F2): se o ORM evoluir e renomear
    `lote_odoo`/`qtd_ajuste`/`acao_decidida`, AttributeError silenciaria a
    garantia anti-replay (listar/aprovar veriam FALHA_BANCO sem indicar a
    causa). getattr com default ''/0 mantem o hash calculavel e estavel.
    """
    h = hashlib.sha256()
    for a in sorted(ajustes, key=lambda x: getattr(x, 'id', 0) or 0):
        h.update(
            f'{getattr(a, "id", "")}|{getattr(a, "cod_produto", "")}|'
            f'{getattr(a, "company_id", "")}|{getattr(a, "lote_odoo", "")}|'
            f'{getattr(a, "qtd_ajuste", "")}|{getattr(a, "acao_decidida", "")}'.encode()
        )
    return h.hexdigest()


def _fazer_backup_pg_dump(
    backup_dir: str,
    db_host: str = 'localhost',
    db_user: str = 'frete_user',
    db_name: str = 'frete_sistema',
    db_password: Optional[str] = None,
) -> str:
    """pg_dump --data-only da tabela ajuste_estoque_inventario.

    Helper opcional (default OFF na CLI); usado antes de DELETE+INSERT em
    --modo propor para cinto+suspensorio. Espelha fazer_backup do 04b.
    """
    from app.utils.timezone import agora_brasil_naive
    os.makedirs(backup_dir, exist_ok=True)
    ts = agora_brasil_naive().strftime('%Y%m%d_%H%M%S')
    path = os.path.join(backup_dir, f'ajuste_estoque_inventario_pre_etapa_{ts}.sql')

    env = os.environ.copy()
    if db_password:
        env['PGPASSWORD'] = db_password
    elif 'PGPASSWORD' not in env:
        raise RuntimeError(
            'PGPASSWORD nao definido. Passe db_password= ou defina no env.'
        )
    try:
        result = subprocess.run(
            [
                'pg_dump',
                '-h', db_host, '-U', db_user, '-d', db_name,
                '--data-only', '--table=ajuste_estoque_inventario',
                '--inserts', '--no-comments', '-f', path,
            ],
            env=env, capture_output=True, text=True,
        )
    except FileNotFoundError as exc:
        # CR-F6: mensagem actionable em vez de "[Errno 2] No such file"
        raise RuntimeError(
            'pg_dump nao encontrado no PATH. Instale postgresql-client '
            '(apt-get install postgresql-client OU brew install libpq).'
        ) from exc
    if result.returncode != 0:
        raise RuntimeError(f'pg_dump falhou: {result.stderr}')
    return path


def propor_ajustes_pre_etapa(
    plano_total: Dict[str, Any],
    ciclo: str,
    company_id: int,
    usuario: str,
    dry_run: bool = True,
) -> Dict[str, Any]:
    """DELETE PROPOSTO + INSERT novos ajustes da Onda (5 CD / 6 FB futura).

    Cria 4 acoes:
    - AJUSTE_{CD|FB}_TRANSF_INTERNA_POS / NEG (transferencias internas)
    - AJUSTE_{CD|FB}_POSITIVO_PURO
    - TRANSFERIR_FB_CD (somente para company_id=4 CD, CFOP 5152)

    Espelha cmd_propor de 04b_propor_pre_etapa_cd. Lazy import de
    Flask/SQLAlchemy/models.

    Args:
        plano_total: dict (saida de planejar_pre_etapa_batch_company).
        ciclo: identificador (ex.: INVENTARIO_2026_05).
        company_id: 4 (CD) ou 1 (FB).
        usuario: criado_por.
        dry_run: True = rollback no fim (NAO modifica banco); False = commit.

    Returns:
        dict com keys: n_antes, n_apos, n_deletados (None se dry_run),
        contador (4 acoes), dry_run, ciclo, company_id.
    """
    from sqlalchemy import text  # lazy

    from app import db  # lazy
    from app.odoo.models import AjusteEstoqueInventario  # lazy

    if company_id not in ACOES_INTERNAS_POR_CID:
        raise ValueError(
            f'company_id={company_id} nao suportado. Use 4 (CD) ou 1 (FB).'
        )
    acoes_cid = ACOES_INTERNAS_POR_CID[company_id]

    def tipo_de_cod(cod: str) -> int:
        return int(cod[0])

    def _cod_valido(cod: str) -> bool:
        """Guard CR-F4: cods do plano JSON podem ter sido manualmente editados;
        outliers (cod sem inicial 1-4) viriam quebrar tipo_de_cod com ValueError.
        Espelha o filtro de planejar_pre_etapa_batch_company."""
        return bool(cod) and cod[0].isdigit() and int(cod[0]) in (1, 2, 3, 4)

    contador = {
        acoes_cid['POS']: 0,
        acoes_cid['NEG']: 0,
        acoes_cid['PURO']: 0,
    }
    # Residual FB->CD so existe para CD (company_id=4)
    if company_id == 4:
        contador[ACAO_RESIDUAL_FB_CD] = 0
    cods_ignorados_outlier: List[str] = []

    # Wrap em SAVEPOINT (CR-F1): isola transacao do caller. Sem isso, um
    # rollback() em dry_run nuke transacoes em aberto do caller (Flask route,
    # subagente web, etc.). Padrao herdado de [[gotcha_commit_service_vaza_savepoint]].
    savepoint = db.session.begin_nested()
    try:
        # 1. Contar PROPOSTO antes (referencia)
        n_antes = db.session.execute(text(
            "SELECT COUNT(*) FROM ajuste_estoque_inventario "
            "WHERE ciclo=:c AND company_id=:cid AND status='PROPOSTO'"
        ), {'c': ciclo, 'cid': company_id}).scalar() or 0

        # 2. DELETE PROPOSTO da company (se !dry_run)
        n_deletados: Optional[int] = None
        if not dry_run:
            result = db.session.execute(text(
                "DELETE FROM ajuste_estoque_inventario "
                "WHERE ciclo=:c AND company_id=:cid AND status='PROPOSTO'"
            ), {'c': ciclo, 'cid': company_id})
            n_deletados = result.rowcount

        # 3. INSERTs
        for t in plano_total.get('transferencias_internas', []):
            cod = t['cod_produto']
            if not _cod_valido(cod):
                logger.warning(f'propor: cod outlier ignorado em transferencia_interna: {cod!r}')
                cods_ignorados_outlier.append(cod)
                continue
        if t['tipo'] == 'POS':
            acao = acoes_cid['POS']
            lote_inv = t['lote_destino_nome']
            qtd_inv = t['qty']
            qtd_odoo_val = t['qty']
        else:  # NEG
            acao = acoes_cid['NEG']
            lote_inv = ''  # lote_origem (que sai) nao esta no inv
            qtd_inv = 0
            qtd_odoo_val = t['qty']

        rec = AjusteEstoqueInventario(
            ciclo=ciclo,
            cod_produto=cod,
            tipo_produto=tipo_de_cod(cod),
            company_id=company_id,
            lote_inventariado=lote_inv,
            lote_odoo=t['lote_origem_nome'],
            lote_origem=t['lote_origem_nome'] or None,
            lote_destino=t['lote_destino_nome'],
            qtd_inventario=Decimal(str(qtd_inv)),
            qtd_odoo=Decimal(str(qtd_odoo_val)),
            qtd_ajuste=Decimal('0'),  # interno: nao muda saldo total
            custo_medio=Decimal(t['custo_medio']),
            acao_decidida=acao,
            criado_por=usuario,
        )
        db.session.add(rec)
        contador[acao] += 1

        # Residual FB->CD (so para CD)
        if company_id == 4:
            for r in plano_total.get('residual_fb_cd', []):
                cod = r['cod_produto']
                if not _cod_valido(cod):
                    logger.warning(f'propor: cod outlier ignorado em residual_fb_cd: {cod!r}')
                    cods_ignorados_outlier.append(cod)
                    continue
                rec = AjusteEstoqueInventario(
                    ciclo=ciclo,
                    cod_produto=cod,
                    tipo_produto=tipo_de_cod(cod),
                    company_id=company_id,
                    lote_inventariado=r['lote_destino_cd_nome'],
                    lote_odoo='',
                    lote_origem=r['lote_origem_fb_sugerido'] or 'MIGRAÇÃO',
                    lote_destino=r['lote_destino_cd_nome'],
                    qtd_inventario=Decimal(str(r['qty'])),
                    qtd_odoo=Decimal('0'),
                    qtd_ajuste=Decimal(str(r['qty'])),
                    custo_medio=Decimal(r['custo_medio']),
                    acao_decidida=ACAO_RESIDUAL_FB_CD,
                    criado_por=usuario,
                )
                db.session.add(rec)
                contador[ACAO_RESIDUAL_FB_CD] += 1

        # Ajustes positivos puros
        for a in plano_total.get('ajustes_positivos_puros', []):
            cod = a['cod_produto']
            if not _cod_valido(cod):
                logger.warning(f'propor: cod outlier ignorado em positivo_puro: {cod!r}')
                cods_ignorados_outlier.append(cod)
                continue
            rec = AjusteEstoqueInventario(
                ciclo=ciclo,
                cod_produto=cod,
                tipo_produto=tipo_de_cod(cod),
                company_id=company_id,
                lote_inventariado=a['lote_destino_nome'],
                lote_odoo='',
                lote_origem=None,
                lote_destino=a['lote_destino_nome'],
                qtd_inventario=Decimal(str(a['qty'])),
                qtd_odoo=Decimal('0'),
                qtd_ajuste=Decimal(str(a['qty'])),
                custo_medio=Decimal(a['custo_medio']),
                acao_decidida=acoes_cid['PURO'],
                criado_por=usuario,
            )
            db.session.add(rec)
            contador[acoes_cid['PURO']] += 1

        # 4. Commit ou rollback no SAVEPOINT (isolado do caller)
        n_apos = n_antes
        if dry_run:
            savepoint.rollback()
        else:
            savepoint.commit()
            n_apos = db.session.execute(text(
                "SELECT COUNT(*) FROM ajuste_estoque_inventario "
                "WHERE ciclo=:c AND company_id=:cid AND status='PROPOSTO'"
            ), {'c': ciclo, 'cid': company_id}).scalar() or 0
    except Exception:
        # Erro inesperado: descartar savepoint e propagar (caller decide commit/rollback do parent)
        try:
            savepoint.rollback()
        except Exception:  # noqa: BLE001
            pass
        raise

    return {
        'n_antes': n_antes,
        'n_apos': n_apos,
        'n_deletados': n_deletados,
        'contador': contador,
        'total_inserts': sum(contador.values()),
        'cods_ignorados_outlier': cods_ignorados_outlier,  # CR-F4: visibilidade
        'dry_run': dry_run,
        'ciclo': ciclo,
        'company_id': company_id,
    }


def listar_onda_pre_etapa(
    ciclo: str,
    company_id: int,
    onda_num: Optional[int] = None,
) -> Dict[str, Any]:
    """Lista ajustes PROPOSTO da Onda + hash sha256 (READ-only).

    Espelha cmd_listar de 04b. onda_num default = inferido de company_id
    (CD=Onda5, FB=Onda6). Hash usado em aprovar_onda_pre_etapa.

    Returns:
        dict com keys: total, hash, por_acao (dict acao -> {n, valor}),
        valor_total, ciclo, company_id, onda_num.
    """
    from app.odoo.models import AjusteEstoqueInventario  # lazy

    if onda_num is None:
        onda_num = ONDA_NUM_POR_CID.get(company_id)
    if company_id not in ACOES_INTERNAS_POR_CID:
        raise ValueError(
            f'company_id={company_id} nao suportado.'
        )
    acoes_set = set(ACOES_INTERNAS_POR_CID[company_id].values())
    # Residual FB->CD so para CD na onda 5
    if company_id == 4:
        acoes_set.add(ACAO_RESIDUAL_FB_CD)

    ajustes = (
        AjusteEstoqueInventario.query
        .filter_by(ciclo=ciclo, status='PROPOSTO', company_id=company_id)
        .filter(AjusteEstoqueInventario.acao_decidida.in_(list(acoes_set)))
        .all()
    )
    total = len(ajustes)
    if total == 0:
        return {
            'total': 0,
            'hash': None,
            'por_acao': {},
            'valor_total': '0.00',
            'ciclo': ciclo,
            'company_id': company_id,
            'onda_num': onda_num,
        }

    por_acao: Dict[str, Dict[str, Any]] = {}
    for a in ajustes:
        por_acao.setdefault(
            a.acao_decidida, {'n': 0, 'valor': Decimal('0')},
        )
        por_acao[a.acao_decidida]['n'] += 1
        por_acao[a.acao_decidida]['valor'] += abs(
            (a.qtd_ajuste or Decimal('0')) * (a.custo_medio or Decimal('0'))
        )
    # Serializar Decimals -> str
    por_acao_ser = {
        acao: {'n': d['n'], 'valor': str(d['valor'].quantize(Decimal('0.01')))}
        for acao, d in por_acao.items()
    }

    valor_total = sum(
        (abs((a.qtd_ajuste or Decimal('0')) * (a.custo_medio or Decimal('0')))
         for a in ajustes),
        Decimal('0'),
    )
    h = _calcular_hash_onda(ajustes)

    return {
        'total': total,
        'hash': h,
        'por_acao': por_acao_ser,
        'valor_total': str(valor_total.quantize(Decimal('0.01'))),
        'ciclo': ciclo,
        'company_id': company_id,
        'onda_num': onda_num,
    }


def aprovar_onda_pre_etapa(
    ciclo: str,
    company_id: int,
    hash_esperado: str,
    usuario: str,
    onda_num: Optional[int] = None,
    dry_run: bool = True,
) -> Dict[str, Any]:
    """Aprova ajustes da Onda com hash check (anti-replay).

    Bloqueia se hash atual != esperado. Espelha cmd_aprovar de 04b.

    Args:
        ciclo: identificador.
        company_id: 4 ou 1.
        hash_esperado: hash sha256 retornado por listar_onda_pre_etapa.
        usuario: aprovado_por.
        onda_num: default inferido de company_id.
        dry_run: True = NAO altera status; False = APROVADO + aprovado_em.

    Returns:
        dict com keys: sucesso, ajustes_aprovados, hash_atual, hash_esperado,
        ts_aprovacao (None se dry_run), ciclo, company_id, onda_num,
        erro (se sucesso=False).
    """
    from app import db  # lazy
    from app.odoo.models import AjusteEstoqueInventario  # lazy
    from app.utils.timezone import agora_utc_naive  # lazy

    if onda_num is None:
        onda_num = ONDA_NUM_POR_CID.get(company_id)
    if company_id not in ACOES_INTERNAS_POR_CID:
        raise ValueError(f'company_id={company_id} nao suportado.')
    acoes_set = set(ACOES_INTERNAS_POR_CID[company_id].values())
    if company_id == 4:
        acoes_set.add(ACAO_RESIDUAL_FB_CD)

    ajustes = (
        AjusteEstoqueInventario.query
        .filter_by(ciclo=ciclo, status='PROPOSTO', company_id=company_id)
        .filter(AjusteEstoqueInventario.acao_decidida.in_(list(acoes_set)))
        .all()
    )
    if not ajustes:
        return {
            'sucesso': False,
            'erro': 'NENHUM_PROPOSTO',
            'ajustes_aprovados': 0,
            'hash_atual': None,
            'hash_esperado': hash_esperado,
            'ts_aprovacao': None,
            'ciclo': ciclo,
            'company_id': company_id,
            'onda_num': onda_num,
        }

    h_atual = _calcular_hash_onda(ajustes)
    if h_atual != hash_esperado:
        return {
            'sucesso': False,
            'erro': 'HASH_DIVERGENTE',
            'ajustes_aprovados': 0,
            'hash_atual': h_atual,
            'hash_esperado': hash_esperado,
            'ts_aprovacao': None,
            'ciclo': ciclo,
            'company_id': company_id,
            'onda_num': onda_num,
        }

    if dry_run:
        return {
            'sucesso': True,
            'erro': None,
            'ajustes_aprovados': len(ajustes),
            'hash_atual': h_atual,
            'hash_esperado': hash_esperado,
            'ts_aprovacao': None,
            'ciclo': ciclo,
            'company_id': company_id,
            'onda_num': onda_num,
            'dry_run': True,
        }

    agora = agora_utc_naive()
    for a in ajustes:
        a.status = 'APROVADO'
        a.aprovado_em = agora
        a.aprovado_por = usuario
    db.session.commit()
    return {
        'sucesso': True,
        'erro': None,
        'ajustes_aprovados': len(ajustes),
        'hash_atual': h_atual,
        'hash_esperado': hash_esperado,
        'ts_aprovacao': agora.isoformat(),
        'ciclo': ciclo,
        'company_id': company_id,
        'onda_num': onda_num,
        'dry_run': False,
    }
