"""
Leitor de Comprovantes de Pagamento de Boleto - SICOOB
======================================================

Extrai informações estruturadas de PDFs de comprovantes de pagamento
gerados pelo Internet Banking do Sicoob.

Dependências:
    pip install pypdfium2 tesserocr pillow

Configuração:
    - Requer dados treinados do Tesseract (por.traineddata)
    - Definir TESSDATA_PREFIX apontando para o diretório com os dados

Uso:
    python scripts/leitor_comprovantes_sicoob.py <caminho_do_pdf>
    python scripts/leitor_comprovantes_sicoob.py comprovantes.pdf --formato json
    python scripts/leitor_comprovantes_sicoob.py comprovantes.pdf --formato csv
"""

import os

# Configurar TESSDATA antes de importar tesserocr
TESSDATA_DIR = os.environ.get(
    'TESSDATA_PREFIX',
    os.path.expanduser('~/tessdata')
)
os.environ['TESSDATA_PREFIX'] = TESSDATA_DIR

import re  # noqa: E402
import json  # noqa: E402
import csv  # noqa: E402
from io import StringIO  # noqa: E402
from dataclasses import dataclass, asdict  # noqa: E402
from typing import Optional  # noqa: E402

import pypdfium2 as pdfium  # noqa: E402
import tesserocr  # noqa: E402
from PIL import Image  # noqa: E402


# ---------------------------------------------------------------------------
# Modelo de dados
# ---------------------------------------------------------------------------

@dataclass
class ComprovanteBoleto:
    """Representa um comprovante de pagamento de boleto SICOOB."""
    # Cabeçalho
    data_comprovante: Optional[str] = None
    cooperativa: Optional[str] = None
    conta: Optional[str] = None
    cliente: Optional[str] = None
    linha_digitavel: Optional[str] = None
    numero_documento: Optional[str] = None
    nosso_numero: Optional[str] = None
    numero_agendamento: Optional[str] = None
    instituicao_emissora: Optional[str] = None
    tipo_documento: Optional[str] = None

    # Beneficiário
    beneficiario_razao_social: Optional[str] = None
    beneficiario_nome_fantasia: Optional[str] = None
    beneficiario_cnpj_cpf: Optional[str] = None

    # Pagador
    pagador_razao_social: Optional[str] = None
    pagador_nome_fantasia: Optional[str] = None
    pagador_cnpj_cpf: Optional[str] = None

    # Datas
    data_realizado: Optional[str] = None
    data_pagamento: Optional[str] = None
    data_vencimento: Optional[str] = None

    # Valores
    valor_documento: Optional[str] = None
    valor_desconto_abatimento: Optional[str] = None
    valor_juros_multa: Optional[str] = None
    valor_pago: Optional[str] = None

    # Status
    situacao: Optional[str] = None
    autenticacao: Optional[str] = None

    # Metadados
    pagina: Optional[int] = None


# ---------------------------------------------------------------------------
# Mapeamento de campos (label OCR → atributo do dataclass)
# ---------------------------------------------------------------------------

# Mapeamento direto: label normalizado → campo do dataclass
MAPEAMENTO_CAMPOS = {
    'cooperativa': 'cooperativa',
    'conta': 'conta',
    'cliente': 'cliente',
    'linha digitavel': 'linha_digitavel',
    'linha digitável': 'linha_digitavel',
    'numero do documento': 'numero_documento',
    'número do documento': 'numero_documento',
    'nosso numero': 'nosso_numero',
    'nosso número': 'nosso_numero',
    'numero do agendamento': 'numero_agendamento',
    'número do agendamento': 'numero_agendamento',
    'instituicao emissora': 'instituicao_emissora',
    'instituição emissora': 'instituicao_emissora',
    'tipo documento': 'tipo_documento',
    'situacao': 'situacao',
    'situação': 'situacao',
    'autenticacao': 'autenticacao',
    'autenticação': 'autenticacao',
}

# Campos que dependem do contexto (seção atual)
# "beneficiario final" mapeia para os MESMOS campos de "beneficiario" — sobrescreve
# automaticamente porque a seção "Beneficiário final" vem DEPOIS no PDF.
CAMPOS_CONTEXTUAIS = {
    'nome/razao social': {
        'beneficiario': 'beneficiario_razao_social',
        'beneficiario final': 'beneficiario_razao_social',
        'pagador': 'pagador_razao_social',
    },
    'nome/razão social': {
        'beneficiario': 'beneficiario_razao_social',
        'beneficiario final': 'beneficiario_razao_social',
        'pagador': 'pagador_razao_social',
    },
    'nome fantasia': {
        'beneficiario': 'beneficiario_nome_fantasia',
        'beneficiario final': 'beneficiario_nome_fantasia',
        'pagador': 'pagador_nome_fantasia',
    },
    'cpf/cnpj': {
        'beneficiario': 'beneficiario_cnpj_cpf',
        'beneficiario final': 'beneficiario_cnpj_cpf',
        'pagador': 'pagador_cnpj_cpf',
    },
}

# Campos de datas
CAMPOS_DATAS = {
    'realizado': 'data_realizado',
    'pagamento': 'data_pagamento',
    'vencimento': 'data_vencimento',
}

# Campos de valores
CAMPOS_VALORES = {
    'documento': 'valor_documento',
    'desconto/abatimento': 'valor_desconto_abatimento',
    'juros/multa': 'valor_juros_multa',
    'pago': 'valor_pago',
}

# Seções que mudam o contexto
# "Beneficiário final" aparece em alguns PDFs (FIDC/cessão) — ao encontrar,
# SOBRESCREVE os campos do beneficiário original (que era o intermediário).
SECOES = {
    'beneficiário', 'beneficiario',
    'beneficiário final', 'beneficiario final',
    'pagador', 'datas', 'valores',
}


# ---------------------------------------------------------------------------
# Mapeamento de campos — Layout V2 ("Comprovante de Pagamento de Boleto")
# ---------------------------------------------------------------------------
# Layout V2 usa labels auto-descritivos (sem seções separadas).
# Ex: "CPF/CNPJ Beneficiário:" em vez de seção "Beneficiário" + "CPF/CNPJ:".

MAPEAMENTO_CAMPOS_V2 = {
    # --- Campos iguais ao V1 ---
    'linha digitavel': 'linha_digitavel',
    'linha digitável': 'linha_digitavel',
    'nosso numero': 'nosso_numero',
    'nosso número': 'nosso_numero',
    'instituicao emissora': 'instituicao_emissora',
    'instituição emissora': 'instituicao_emissora',
    'tipo documento': 'tipo_documento',
    'situacao': 'situacao',
    'situação': 'situacao',
    'autenticacao': 'autenticacao',
    'autenticação': 'autenticacao',

    # --- Labels V2 específicos: Nº documento / agendamento ---
    'nº documento': 'numero_documento',
    'n documento': 'numero_documento',
    'n° documento': 'numero_documento',
    'no. agendamento': 'numero_agendamento',
    'no agendamento': 'numero_agendamento',
    'nº agendamento': 'numero_agendamento',
    'n. agendamento': 'numero_agendamento',
    'n° agendamento': 'numero_agendamento',

    # --- Beneficiário (label auto-descritivo, sem seção) ---
    'nome/razao social do beneficiario': 'beneficiario_razao_social',
    'nome/razão social do beneficiário': 'beneficiario_razao_social',
    'nome fantasia beneficiario': 'beneficiario_nome_fantasia',
    'nome fantasia beneficiário': 'beneficiario_nome_fantasia',
    'cpf/cnpj beneficiario': 'beneficiario_cnpj_cpf',
    'cpf/cnpj beneficiário': 'beneficiario_cnpj_cpf',

    # --- Pagador (label auto-descritivo, sem seção) ---
    'nome/razao social do pagador': 'pagador_razao_social',
    'nome/razão social do pagador': 'pagador_razao_social',
    'nome fantasia pagador': 'pagador_nome_fantasia',
    'cpf/cnpj pagador': 'pagador_cnpj_cpf',

    # --- Datas (prefixo "Data" no label) ---
    'data agendamento': 'data_realizado',
    'data pagamento': 'data_pagamento',
    'data vencimento': 'data_vencimento',

    # --- Valores ---
    'valor documento': 'valor_documento',
    '(-) desconto / abatimento': 'valor_desconto_abatimento',
    '(-) desconto/abatimento': 'valor_desconto_abatimento',
    # "(+) Outros acréscimos" → valor_juros_multa (campo mais próximo no dataclass)
    '(+) outros acrescimos': 'valor_juros_multa',
    '(+) outros acréscimos': 'valor_juros_multa',
    'valor pago': 'valor_pago',
}


# ---------------------------------------------------------------------------
# Funções de extração
# ---------------------------------------------------------------------------

def renderizar_pagina(pdf_path: str, pagina_idx: int, escala: float = 4.0) -> Image.Image:
    """Renderiza uma página do PDF como imagem PIL."""
    pdf = pdfium.PdfDocument(pdf_path)
    page = pdf[pagina_idx]
    bitmap = page.render(scale=escala)
    return bitmap.to_pil()


def extrair_texto_ocr(imagem: Image.Image, idioma: str = 'por') -> str:
    """Executa OCR em uma imagem PIL e retorna o texto."""
    return tesserocr.image_to_text(imagem, lang=idioma)


def normalizar_label(label: str) -> str:
    """Normaliza um label removendo ':' e espaços extras."""
    return label.strip().rstrip(':').strip().lower()


def _detectar_formato(linhas: list[str]) -> str:
    """
    Detecta o formato do OCR.

    Formato A (labels separadas): Labels vêm ANTES do cabeçalho SICOOB,
        valores vêm DEPOIS. OCR leu coluna esquerda toda, depois coluna direita.

    Formato B (inline): Cabeçalho SICOOB vem PRIMEIRO, seguido de linhas
        no formato "Label: Valor". OCR leu da esquerda para direita, linha a linha.
    """
    for i, linha in enumerate(linhas):
        if 'SICOOB - SISTEMA DE COOPERATIVAS' in linha.upper():
            # Se SICOOB aparece nas primeiras 3 linhas → Formato B (inline)
            if i <= 2:
                return 'B'
            # Se aparece mais tarde → Formato A (labels separadas)
            return 'A'
    return 'B'  # fallback


def _detectar_layout_boleto(linhas: list[str]) -> str:
    """
    Detecta se o comprovante usa Layout V1 (seções) ou V2 (labels auto-descritivos).

    V1 (novo, ~2026): Seções separadas ("Beneficiário:", "Pagador:", "Datas:", "Valores:")
        com labels genéricos dentro de cada seção. 1 comprovante por página.

    V2 (antigo, ~2024-2025): Labels auto-descritivos (ex: "CPF/CNPJ Beneficiário:").
        1 ou 2 comprovantes por página.

    Estratégia: Verificar marcadores EXCLUSIVOS do V2 (match positivo).
    Se nenhum marcador V2 encontrado → V1 (default seguro).

    NOTA: Ambos layouts contêm "COMPROVANTE DE PAGAMENTO DE BOLETO" no texto,
    por isso NÃO usamos esse cabeçalho para diferenciação.
    """
    texto_junto = ' '.join(linhas).upper()

    # --- Marcadores EXCLUSIVOS do V2 (NÃO existem no V1) ---
    # V2 usa labels auto-descritivos; V1 usa labels genéricos dentro de seções.
    marcadores_v2 = [
        'CPF/CNPJ BENEFICI',               # V1 usa "CPF/CNPJ:" genérico
        'NOME/RAZÃO SOCIAL DO BENEFICI',    # V1 usa "Nome/Razão Social:" genérico
        'NOME/RAZAO SOCIAL DO BENEFICI',    # Variante sem acento (OCR)
        'DATA AGENDAMENTO',                 # V1 não tem este label
        'VALOR DOCUMENTO',                  # V1 usa "Documento:" (sem "Valor")
        'VALOR PAGO',                       # V1 usa "Pago:" (sem "Valor")
        'NO. AGENDAMENTO',                  # V1 usa "Número do agendamento:"
        'PLATAFORMA DE SERVI',              # Rodapé exclusivo V2
    ]

    if any(m in texto_junto for m in marcadores_v2):
        return 'V2'

    # Marcador V2 por linha: "Coop.:" no início (V1 usa "Cooperativa:")
    for linha in linhas:
        if re.match(r'^Coop\.?\s*:', linha.strip(), re.IGNORECASE):
            return 'V2'

    # Default: V1 (layout com seções)
    return 'V1'


def _split_v2_page_text(linhas: list[str]) -> list[list[str]]:
    """
    Detecta se uma página V2 contém 2 comprovantes e divide em segmentos.

    Páginas V2 podem conter 1 ou 2 comprovantes empilhados verticalmente.
    Detecta pela SEGUNDA ocorrência de "Coop.:" (primeira linha específica
    do cabeçalho V2 que aparece após "COMPROVANTE DE PAGAMENTO DE BOLETO").

    Returns:
        Lista com 1 ou 2 listas de linhas (cada uma = 1 comprovante).
    """
    # Encontrar índices das linhas que iniciam um comprovante V2
    # O marcador mais confiável: "Coop.:" ou "Coop:" no início da linha
    # (formato V2: "Coop.: CÓDIGO / NOME" — dois-pontos obrigatório para
    #  não casar com "Cooperativa" do V1)
    indices_coop = []
    for i, linha in enumerate(linhas):
        if re.match(r'^Coop\.?\s*:\s*.+$', linha.strip(), re.IGNORECASE):
            indices_coop.append(i)

    if len(indices_coop) <= 1:
        # Página com 1 comprovante (ou nenhum Coop encontrado)
        return [linhas]

    # 2 comprovantes: dividir no início do segundo
    # O segundo comprovante começa algumas linhas ANTES do segundo "Coop"
    # (com o header "COMPROVANTE DE PAGAMENTO DE BOLETO", "SICOOB", etc.)
    split_idx = indices_coop[1]

    # Retroceder para incluir linhas de cabeçalho do segundo comprovante
    for j in range(split_idx - 1, max(split_idx - 6, 0), -1):
        upper = linhas[j].strip().upper()
        if any(kw in upper for kw in [
            'COMPROVANTE DE', 'PAGAMENTO DE BOLETO',
            'SICOOB', 'SISBR', 'PLATAFORMA',
        ]):
            split_idx = j
            break

    segmento1 = linhas[:split_idx]
    segmento2 = linhas[split_idx:]

    return [segmento1, segmento2]


def _parse_formato_a(linhas: list[str], comprovante: ComprovanteBoleto) -> None:
    """
    Parse do Formato A: labels e valores em blocos separados.
    O OCR leu toda a coluna esquerda, depois toda a coluna direita.
    """
    # Encontrar separador (SICOOB - SISTEMA...)
    idx_separador = None
    for i, linha in enumerate(linhas):
        if 'SICOOB - SISTEMA DE COOPERATIVAS' in linha.upper():
            idx_separador = i
            break

    if idx_separador is None:
        return

    bloco_labels_raw = linhas[:idx_separador]
    bloco_valores_raw = linhas[idx_separador:]

    # Filtrar cabeçalho SICOOB dos valores
    valores_inicio = 0
    for i, linha in enumerate(bloco_valores_raw):
        if any(kw in linha.upper() for kw in [
            'SICOOB', 'SISBR', 'COMPROVANTE', 'PAGAMENTO DE BOLETO'
        ]):
            valores_inicio = i + 1
            continue
        else:
            valores_inicio = i
            break

    bloco_valores = bloco_valores_raw[valores_inicio:]

    # Filtrar OUVIDORIA, versão e hora solta
    bloco_valores = [
        v for v in bloco_valores
        if not any(kw in v.upper() for kw in ['OUVIDORIA', 'V1.0'])
        and not re.match(r'^\d{2}:\d{2}:\d{2}$', v.strip())
    ]

    # Processar labels
    labels = []
    for linha in bloco_labels_raw:
        if comprovante.data_comprovante is None and re.match(r'^\d{2}/\d{2}/\d{4}$', linha.strip()):
            comprovante.data_comprovante = linha.strip()
            continue
        label_limpo = normalizar_label(linha)
        if label_limpo:
            labels.append(label_limpo)

    # Parear labels com valores
    idx_valor = 0
    secao_atual = None

    for label in labels:
        if label in SECOES:
            secao_atual = label.replace('á', 'a').replace('é', 'e')
            continue

        if idx_valor >= len(bloco_valores):
            break

        valor = bloco_valores[idx_valor].strip()
        idx_valor += 1

        _atribuir_campo(comprovante, label, valor, secao_atual)


def _parse_formato_b(linhas: list[str], comprovante: ComprovanteBoleto) -> None:
    """
    Parse do Formato B: labels e valores inline ("Label: Valor").
    O OCR leu linha por linha, da esquerda para direita.
    """
    # Pular cabeçalho SICOOB e extrair data do comprovante da linha mista
    secao_atual = None
    valores_pendentes = []  # Para labels sem valor na mesma linha

    for linha in linhas:
        # Ignorar linhas de cabeçalho/rodapé
        if any(kw in linha.upper() for kw in [
            'SICOOB - SISTEMA', 'SISBR - SISTEMA', 'COMPROVANTE DE',
            'OUVIDORIA', 'V1.0'
        ]):
            continue

        # Linha com data + "PAGAMENTO DE BOLETO" + hora
        match_data_cabecalho = re.match(
            r'^(\d{2}/\d{2}/\d{4})\s+PAGAMENTO DE BOLETO\s+(\d{2}:\d{2}:\d{2})$',
            linha.strip()
        )
        if match_data_cabecalho:
            comprovante.data_comprovante = match_data_cabecalho.group(1)
            continue

        # Hora solta (rodapé)
        if re.match(r'^\d{2}:\d{2}:\d{2}$', linha.strip()):
            continue

        # Detectar seções
        label_normalizado = normalizar_label(linha)
        if label_normalizado in SECOES:
            secao_atual = label_normalizado.replace('á', 'a').replace('é', 'e')
            continue

        # Tentar formato "Label: Valor" (com dois-pontos seguido de valor)
        match_inline = re.match(r'^(.+?):\s+(.+)$', linha.strip())
        if match_inline:
            label = normalizar_label(match_inline.group(1))
            valor = match_inline.group(2).strip()

            # Verificar se não é label sem valor (label que termina em ":" sozinha)
            _atribuir_campo(comprovante, label, valor, secao_atual)
            continue

        # Linha que é apenas um label com ":" (sem valor)
        if linha.strip().endswith(':'):
            label = normalizar_label(linha)
            valores_pendentes.append((label, secao_atual))
            continue

        # Linha que é apenas um valor (sem label) — consumir label pendente
        if valores_pendentes:
            label, secao_contexto = valores_pendentes.pop(0)
            _atribuir_campo(comprovante, label, linha.strip(), secao_contexto)
            continue


def _normalizar_label_v2(label: str) -> str:
    """
    Normaliza um label V2 para lookup no MAPEAMENTO_CAMPOS_V2.

    Além da normalização padrão (strip, lower, remover ':'), também
    normaliza prefixos (+)/(-) que o OCR pode corromper.
    """
    label = normalizar_label(label)

    # Normalizar prefixos (-) e (+) — OCR pode gerar variantes:
    # "(- )", "(—)", "(–)", "(-)", etc.
    label = re.sub(r'^\(\s*[-–—]\s*\)\s*', '(-) ', label)
    label = re.sub(r'^\(\s*\+\s*\)\s*', '(+) ', label)

    return label


def _parse_layout_v2(linhas: list[str], comprovante: ComprovanteBoleto) -> None:
    """
    Parse do Layout V2: labels auto-descritivos, sem seções.

    Sub-cabeçalho: "Comprovante de Pagamento de Boleto"
    Formato: cada campo é "Label: Valor" com label já descrevendo a quem pertence.
    Ex: "CPF/CNPJ Beneficiário: 37.706.133/0001-98"
    """
    valores_pendentes = []  # Para labels sem valor na mesma linha

    for linha in linhas:
        stripped = linha.strip()
        if not stripped:
            continue

        # --- Linhas de cabeçalho/rodapé: ignorar ---
        upper = stripped.upper()
        if any(kw in upper for kw in [
            'SICOOB - SISTEMA', 'SISBR - SISTEMA',
            'PLATAFORMA DE SERVI', 'COMPROVANTE DE',
            'PAGAMENTO DE BOLETO', 'OUVIDORIA', 'V1.0',
        ]):
            continue

        # --- Cabeçalho V2: "Coop.: CÓDIGO / NOME" ---
        match_coop = re.match(r'^Coop\.?\s*:?\s*(.+)$', stripped, re.IGNORECASE)
        if match_coop:
            comprovante.cooperativa = match_coop.group(1).strip()
            continue

        # --- Cabeçalho V2: "Conta: CÓDIGO / NOME_CLIENTE" ---
        match_conta = re.match(r'^Conta\s*:\s*(.+)$', stripped, re.IGNORECASE)
        if match_conta:
            valor_conta = match_conta.group(1).strip()
            partes = valor_conta.split('/', 1)
            comprovante.conta = partes[0].strip()
            if len(partes) > 1:
                comprovante.cliente = partes[1].strip()
            continue

        # --- Cabeçalho V2: "Data: DD/MM/YYYY" (pode ter "Hora: HH:MM:SS" na mesma linha) ---
        match_data_hora = re.match(
            r'^Data\s*:\s*(\d{2}/\d{2}/\d{4})\s*(Hora\s*:\s*\d{2}:\d{2}:\d{2})?$',
            stripped, re.IGNORECASE,
        )
        if match_data_hora:
            comprovante.data_comprovante = match_data_hora.group(1)
            continue

        # --- Hora solta (rodapé ou cabeçalho) ---
        if re.match(r'^(Hora\s*:\s*)?\d{2}:\d{2}:\d{2}$', stripped, re.IGNORECASE):
            continue

        # --- Labels ignorados (sem campo correspondente no dataclass) ---
        if 'autorizou pagar valor diferente' in stripped.lower():
            continue

        # --- Tentar formato "Label: Valor" (inline) ---
        # \s* (zero ou mais espaços) — OCR pode gerar "Label:Valor" sem espaço
        match_inline = re.match(r'^(.+?):\s*(.+)$', stripped)
        if match_inline:
            label = _normalizar_label_v2(match_inline.group(1))
            valor = match_inline.group(2).strip()

            if label in MAPEAMENTO_CAMPOS_V2:
                # Tratar placeholders "--" como valor ausente (FIDC/cessão)
                if valor.strip() in ('--', '-', '---'):
                    valor = None
                setattr(comprovante, MAPEAMENTO_CAMPOS_V2[label], valor)
            continue

        # --- Label V2 com ":" sozinho na linha (valor na próxima linha) ---
        if stripped.endswith(':'):
            label = _normalizar_label_v2(stripped)
            valores_pendentes.append(label)
            continue

        # --- Linha que é apenas um valor (sem label) — consumir label pendente ---
        if valores_pendentes:
            label = valores_pendentes.pop(0)
            if label in MAPEAMENTO_CAMPOS_V2:
                # Tratar placeholders "--" como valor ausente (FIDC/cessão)
                val = stripped if stripped not in ('--', '-', '---') else None
                setattr(comprovante, MAPEAMENTO_CAMPOS_V2[label], val)
            continue


def _atribuir_campo(
    comprovante: ComprovanteBoleto,
    label: str,
    valor: str,
    secao_atual: Optional[str]
) -> bool:
    """
    Atribui um valor ao campo correto do comprovante baseado no label e seção.

    Returns:
        True se o campo foi atribuído, False caso contrário.
    """
    # 1. Campos diretos
    if label in MAPEAMENTO_CAMPOS:
        setattr(comprovante, MAPEAMENTO_CAMPOS[label], valor)
        return True

    # 2. Campos contextuais (dependem da seção)
    if label in CAMPOS_CONTEXTUAIS:
        if secao_atual and secao_atual in CAMPOS_CONTEXTUAIS[label]:
            setattr(comprovante, CAMPOS_CONTEXTUAIS[label][secao_atual], valor)
            return True

    # 3. Campos de datas
    if label in CAMPOS_DATAS and secao_atual in ('datas', None):
        setattr(comprovante, CAMPOS_DATAS[label], valor)
        return True

    # 4. Campos de valores
    if label in CAMPOS_VALORES and secao_atual in ('valores', None):
        setattr(comprovante, CAMPOS_VALORES[label], valor)
        return True

    return False


def _validar_comprovante(comprovante: ComprovanteBoleto) -> ComprovanteBoleto:
    """
    Valida campos do comprovante após OCR para detectar parsing incorreto.

    Campos cujo valor não bate com o padrão esperado são anulados (None)
    para evitar dados incorretos no banco (ex: CNPJ no campo de razão social).
    """
    # numero_agendamento deve ser numérico
    # Layout V2 pode trazer com separador de milhar (ex: "17.375.820") — limpar antes
    if comprovante.numero_agendamento:
        agendamento_limpo = comprovante.numero_agendamento.replace('.', '').replace(' ', '').strip()
        if re.match(r'^\d+$', agendamento_limpo):
            comprovante.numero_agendamento = agendamento_limpo
        else:
            print(f"  ⚠️  p.{comprovante.pagina}: numero_agendamento inválido: '{comprovante.numero_agendamento}' → None")
            comprovante.numero_agendamento = None

    # numero_documento e nosso_numero: tratar placeholders como None
    # Comprovantes FIDC/cessão usam "--" ou "-" quando campo não se aplica
    for campo in ('numero_documento', 'nosso_numero'):
        valor = getattr(comprovante, campo, None)
        if valor and valor.strip() in ('--', '-', '---', ''):
            print(f"  ℹ️  p.{comprovante.pagina}: {campo} placeholder: '{valor}' → None")
            setattr(comprovante, campo, None)

    # CNPJ/CPF deve conter apenas dígitos, pontos, barras, hífens
    for campo in ('beneficiario_cnpj_cpf', 'pagador_cnpj_cpf'):
        valor = getattr(comprovante, campo, None)
        if valor and not re.match(r'^[\d./ -]+$', valor):
            print(f"  ⚠️  p.{comprovante.pagina}: {campo} inválido: '{valor}' → None")
            setattr(comprovante, campo, None)

    # Valores devem conter R$ ou padrão numérico (com separadores BR)
    for campo in ('valor_documento', 'valor_desconto_abatimento', 'valor_juros_multa', 'valor_pago'):
        valor = getattr(comprovante, campo, None)
        if valor and not re.match(r'^R?\$?\s*[\d.,]+$', valor):
            print(f"  ⚠️  p.{comprovante.pagina}: {campo} inválido: '{valor}' → None")
            setattr(comprovante, campo, None)

    return comprovante


def _eh_comprovante_boleto(texto: str) -> bool:
    """
    Verifica se o texto OCR corresponde a um comprovante de PAGAMENTO DE BOLETO.

    O cabeçalho "SICOOB - SISTEMA DE COOPERATIVAS..." aparece em TODOS os
    comprovantes Sicoob (boleto, PIX, TED, etc.). O que diferencia é o
    sub-cabeçalho "PAGAMENTO DE BOLETO" — presente APENAS em boletos.

    Se o texto não contém "PAGAMENTO DE BOLETO", não é um comprovante de
    boleto e deve ser pulado pelo parser.
    """
    return 'PAGAMENTO DE BOLETO' in texto.upper()


def parse_comprovantes_from_text(texto: str, pagina: int) -> list[ComprovanteBoleto]:
    """
    Faz o parsing do texto OCR e retorna 1 ou mais comprovantes.

    Para V2: detecta se há 2 comprovantes na mesma página e divide em segmentos.
    Para V1: sempre retorna 1 comprovante.

    Args:
        texto: Texto OCR da página do PDF
        pagina: Número da página (1-indexed)

    Returns:
        Lista de ComprovanteBoleto extraídos (1 ou 2 para V2, sempre 1 para V1).
    """
    linhas = texto.strip().split('\n')
    linhas = [ln.strip() for ln in linhas if ln.strip()]

    if not linhas:
        return [ComprovanteBoleto(pagina=pagina)]

    # Detectar layout do comprovante (V1 com seções ou V2 auto-descritivo)
    layout = _detectar_layout_boleto(linhas)

    if layout == 'V2':
        # V2 pode ter 1 ou 2 comprovantes por página
        segmentos = _split_v2_page_text(linhas)
        comprovantes = []
        for seg in segmentos:
            comp = ComprovanteBoleto(pagina=pagina)
            _parse_layout_v2(seg, comp)
            comp = _validar_comprovante(comp)
            comprovantes.append(comp)
        return comprovantes
    else:
        # Layout V1: sempre 1 comprovante por página
        comprovante = ComprovanteBoleto(pagina=pagina)
        formato = _detectar_formato(linhas)
        if formato == 'A':
            _parse_formato_a(linhas, comprovante)
        else:
            _parse_formato_b(linhas, comprovante)
        comprovante = _validar_comprovante(comprovante)
        return [comprovante]


def parse_comprovante(texto: str, pagina: int) -> ComprovanteBoleto:
    """
    Faz o parsing do texto OCR de um comprovante e retorna os dados estruturados.

    Backward-compatible: retorna o primeiro comprovante da página.
    Para obter todos os comprovantes (V2 pode ter 2), use parse_comprovantes_from_text().
    """
    resultados = parse_comprovantes_from_text(texto, pagina)
    return resultados[0] if resultados else ComprovanteBoleto(pagina=pagina)


def extrair_comprovantes(pdf_path: str) -> list[ComprovanteBoleto]:
    """
    Extrai todos os comprovantes de um PDF.

    Args:
        pdf_path: Caminho do arquivo PDF

    Returns:
        Lista de ComprovanteBoleto extraídos
    """
    if not os.path.exists(pdf_path):
        raise FileNotFoundError(f"Arquivo não encontrado: {pdf_path}")

    pdf = pdfium.PdfDocument(pdf_path)
    total_paginas = len(pdf)
    comprovantes = []

    print(f"Processando: {pdf_path}")
    print(f"Total de páginas: {total_paginas}")
    print("-" * 60)

    pulados = 0

    for i in range(total_paginas):
        print(f"  Página {i + 1}/{total_paginas}... ", end="", flush=True)

        # Renderizar e OCR
        imagem = renderizar_pagina(pdf_path, i)
        texto = extrair_texto_ocr(imagem)

        # Verificar se é comprovante de boleto
        if not _eh_comprovante_boleto(texto):
            print("PULADA (não é boleto)")
            pulados += 1
            continue

        # Parsear (V2 pode ter 1 ou 2 comprovantes por página)
        resultados = parse_comprovantes_from_text(texto, pagina=i + 1)
        comprovantes.extend(resultados)

        # Feedback
        for comp in resultados:
            if comp.beneficiario_razao_social:
                print(f"OK → {comp.beneficiario_razao_social}")
            else:
                print("OK (sem beneficiário identificado)")

    print("-" * 60)
    print(f"Total extraído: {len(comprovantes)} comprovante(s)")
    if pulados:
        print(f"Páginas puladas (não são boleto): {pulados}")

    return comprovantes


def extrair_comprovantes_from_bytes(pdf_bytes: bytes) -> list[ComprovanteBoleto]:
    """
    Extrai comprovantes a partir de bytes do PDF (uso em APIs web).

    Args:
        pdf_bytes: Conteúdo binário do PDF

    Returns:
        Lista de ComprovanteBoleto extraídos
    """
    pdf = pdfium.PdfDocument(pdf_bytes)
    total_paginas = len(pdf)
    comprovantes = []

    for i in range(total_paginas):
        page = pdf[i]
        bitmap = page.render(scale=4.0)
        imagem = bitmap.to_pil()
        texto = extrair_texto_ocr(imagem)

        # Pular páginas que não são comprovante de boleto (PIX, TED, etc.)
        if not _eh_comprovante_boleto(texto):
            continue

        # V2 pode ter 1 ou 2 comprovantes por página
        resultados = parse_comprovantes_from_text(texto, pagina=i + 1)
        comprovantes.extend(resultados)

    return comprovantes


# ---------------------------------------------------------------------------
# Formatação de saída
# ---------------------------------------------------------------------------

def formatar_tabela(comprovantes: list[ComprovanteBoleto]) -> str:
    """Formata os comprovantes em uma tabela legível."""
    saida = []

    for i, c in enumerate(comprovantes):
        saida.append(f"\n{'='*70}")
        saida.append(f" COMPROVANTE {i+1} (Página {c.pagina})")
        saida.append(f"{'='*70}")

        saida.append(f"\n  📅 Data:                  {c.data_comprovante or 'N/A'}")
        saida.append(f"  🏦 Cooperativa:           {c.cooperativa or 'N/A'}")
        saida.append(f"  💳 Conta:                 {c.conta or 'N/A'}")
        saida.append(f"  👤 Cliente:               {c.cliente or 'N/A'}")
        saida.append(f"  📝 Linha Digitável:       {c.linha_digitavel or 'N/A'}")
        saida.append(f"  📄 Nº Documento:          {c.numero_documento or 'N/A'}")
        saida.append(f"  🔢 Nosso Número:          {c.nosso_numero or 'N/A'}")
        saida.append(f"  📋 Nº Agendamento:        {c.numero_agendamento or 'N/A'}")
        saida.append(f"  🏛️  Instituição Emissora:  {c.instituicao_emissora or 'N/A'}")
        saida.append(f"  📌 Tipo Documento:        {c.tipo_documento or 'N/A'}")

        saida.append(f"\n  --- BENEFICIÁRIO ---")
        saida.append(f"  🏢 Razão Social:          {c.beneficiario_razao_social or 'N/A'}")
        saida.append(f"  🏪 Nome Fantasia:         {c.beneficiario_nome_fantasia or 'N/A'}")
        saida.append(f"  📋 CNPJ/CPF:              {c.beneficiario_cnpj_cpf or 'N/A'}")

        saida.append(f"\n  --- PAGADOR ---")
        saida.append(f"  🏢 Razão Social:          {c.pagador_razao_social or 'N/A'}")
        saida.append(f"  🏪 Nome Fantasia:         {c.pagador_nome_fantasia or 'N/A'}")
        saida.append(f"  📋 CNPJ/CPF:              {c.pagador_cnpj_cpf or 'N/A'}")

        saida.append(f"\n  --- DATAS ---")
        saida.append(f"  ✅ Realizado:              {c.data_realizado or 'N/A'}")
        saida.append(f"  💰 Pagamento:             {c.data_pagamento or 'N/A'}")
        saida.append(f"  📅 Vencimento:            {c.data_vencimento or 'N/A'}")

        saida.append(f"\n  --- VALORES ---")
        saida.append(f"  💵 Documento:             {c.valor_documento or 'N/A'}")
        saida.append(f"  🔽 Desconto/Abatimento:   {c.valor_desconto_abatimento or 'N/A'}")
        saida.append(f"  📈 Juros/Multa:           {c.valor_juros_multa or 'N/A'}")
        saida.append(f"  💰 PAGO:                  {c.valor_pago or 'N/A'}")

        saida.append(f"\n  --- STATUS ---")
        saida.append(f"  📊 Situação:              {c.situacao or 'N/A'}")
        saida.append(f"  🔐 Autenticação:          {c.autenticacao or 'N/A'}")

    return '\n'.join(saida)


def formatar_json(comprovantes: list[ComprovanteBoleto]) -> str:
    """Formata os comprovantes como JSON."""
    dados = [asdict(c) for c in comprovantes]
    return json.dumps(dados, ensure_ascii=False, indent=2)


def formatar_csv(comprovantes: list[ComprovanteBoleto]) -> str:
    """Formata os comprovantes como CSV."""
    if not comprovantes:
        return ""

    output = StringIO()
    campos = list(asdict(comprovantes[0]).keys())
    writer = csv.DictWriter(output, fieldnames=campos, delimiter=';')
    writer.writeheader()
    for c in comprovantes:
        writer.writerow(asdict(c))

    return output.getvalue()


def formatar_resumo(comprovantes: list[ComprovanteBoleto]) -> str:
    """Formata um resumo consolidado dos pagamentos."""
    saida = []
    saida.append(f"\n{'='*70}")
    saida.append(f" RESUMO DE PAGAMENTOS ({len(comprovantes)} comprovante(s))")
    saida.append(f"{'='*70}\n")

    total_pago = 0.0
    for c in comprovantes:
        valor_str = (c.valor_pago or '0').replace('R$', '').replace('.', '').replace(',', '.').strip()
        try:
            valor = float(valor_str)
        except ValueError:
            valor = 0.0
        total_pago += valor

        saida.append(
            f"  Pg.{c.pagina:>2} | {c.beneficiario_razao_social or 'N/A':<45} | "
            f"{c.beneficiario_cnpj_cpf or 'N/A':<20} | {c.valor_pago or 'N/A':>12} | "
            f"Venc: {c.data_vencimento or 'N/A'}"
        )

    saida.append(f"\n{'─'*70}")
    saida.append(f"  TOTAL PAGO: R$ {total_pago:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
    saida.append(f"{'─'*70}")

    return '\n'.join(saida)


def formatar_xlsx(comprovantes: list[ComprovanteBoleto], caminho_saida: str) -> str:
    """Gera arquivo Excel (.xlsx) com os comprovantes."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Comprovantes'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2E5090', end_color='2E5090', fill_type='solid')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    colunas = [
        ('pagina', 'Página'),
        ('data_comprovante', 'Data'),
        ('beneficiario_razao_social', 'Beneficiário'),
        ('beneficiario_cnpj_cpf', 'CNPJ Beneficiário'),
        ('pagador_razao_social', 'Pagador'),
        ('pagador_cnpj_cpf', 'CNPJ Pagador'),
        ('numero_documento', 'Nº Documento'),
        ('nosso_numero', 'Nosso Número'),
        ('numero_agendamento', 'Nº Agendamento'),
        ('valor_documento', 'Valor Documento'),
        ('valor_desconto_abatimento', 'Desconto'),
        ('valor_juros_multa', 'Juros/Multa'),
        ('valor_pago', 'Valor Pago'),
        ('data_pagamento', 'Data Pagamento'),
        ('data_vencimento', 'Data Vencimento'),
        ('data_realizado', 'Realizado'),
        ('situacao', 'Situação'),
        ('autenticacao', 'Autenticação'),
        ('cooperativa', 'Cooperativa'),
        ('conta', 'Conta'),
        ('cliente', 'Cliente'),
        ('linha_digitavel', 'Linha Digitável'),
        ('instituicao_emissora', 'Inst. Emissora'),
        ('tipo_documento', 'Tipo Documento'),
    ]

    # Header
    for col_idx, (_, nome) in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_idx, value=nome)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Dados
    for row_idx, comp in enumerate(comprovantes, 2):
        dados = asdict(comp)
        for col_idx, (campo, _) in enumerate(colunas, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=dados.get(campo, ''))
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if campo == 'pagina' else 'left')

    # Auto-ajustar largura
    for col_idx, (campo, nome) in enumerate(colunas, 1):
        max_len = len(nome)
        for row in range(2, len(comprovantes) + 2):
            val = str(ws.cell(row=row, column=col_idx).value or '')
            max_len = max(max_len, len(val))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 3, 45)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    wb.save(caminho_saida)
    return f"Arquivo Excel salvo: {caminho_saida}"


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    import argparse

    parser = argparse.ArgumentParser(
        description='Leitor de Comprovantes de Pagamento de Boleto - SICOOB',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos:
  python leitor_comprovantes_sicoob.py comprovantes.pdf
  python leitor_comprovantes_sicoob.py comprovantes.pdf --formato json
  python leitor_comprovantes_sicoob.py comprovantes.pdf --formato csv --saida pagamentos.csv
  python leitor_comprovantes_sicoob.py comprovantes.pdf --formato resumo
  python leitor_comprovantes_sicoob.py comprovantes.pdf --formato xlsx --saida comprovantes.xlsx
        """
    )
    parser.add_argument('pdf', help='Caminho do arquivo PDF')
    parser.add_argument(
        '--formato', '-f',
        choices=['tabela', 'json', 'csv', 'resumo', 'xlsx'],
        default='tabela',
        help='Formato de saída (padrão: tabela)'
    )
    parser.add_argument(
        '--saida', '-o',
        help='Arquivo de saída (padrão: stdout). Obrigatório para xlsx.'
    )
    parser.add_argument(
        '--escala',
        type=float,
        default=4.0,
        help='Escala de renderização para OCR (padrão: 4.0 = ~288 DPI)'
    )

    args = parser.parse_args()

    # Validar xlsx requer --saida
    if args.formato == 'xlsx' and not args.saida:
        parser.error("--saida é obrigatório quando --formato=xlsx")

    # Extrair
    comprovantes = extrair_comprovantes(args.pdf)

    # Formato xlsx tem tratamento especial (binário)
    if args.formato == 'xlsx':
        print(formatar_resumo(comprovantes))
        msg = formatar_xlsx(comprovantes, args.saida)
        print(f"\n{msg}")
        return

    # Formatos texto
    formatadores = {
        'tabela': formatar_tabela,
        'json': formatar_json,
        'csv': formatar_csv,
        'resumo': formatar_resumo,
    }

    resultado = formatadores[args.formato](comprovantes)

    # Sempre mostrar resumo no stderr quando salvando em arquivo
    if args.saida:
        print(formatar_resumo(comprovantes))

    # Saída
    if args.saida:
        with open(args.saida, 'w', encoding='utf-8') as f:
            f.write(resultado)
        print(f"\nArquivo salvo: {args.saida}")
    else:
        print(resultado)


if __name__ == '__main__':
    main()
