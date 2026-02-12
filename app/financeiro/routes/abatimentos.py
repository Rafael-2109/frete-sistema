"""
Rotas de Abatimentos
APIs CRUD de abatimentos por conta + Listagem geral
"""

from flask import render_template, request, jsonify, Response
from flask_login import login_required, current_user
from datetime import datetime
from io import BytesIO

from app import db
from app.utils.timezone import agora_utc_naive
from app.financeiro.routes import financeiro_bp


# ========================================
# ROTAS API: ABATIMENTOS
# ========================================

@financeiro_bp.route('/contas-receber/api/<int:conta_id>/abatimentos', methods=['GET'])
@login_required
def api_listar_abatimentos(conta_id):
    """
    Lista abatimentos de uma conta a receber + desconto contratual
    """
    try:
        from app.financeiro.models import ContasAReceber, ContasAReceberTipo

        conta = ContasAReceber.query.get_or_404(conta_id)

        # Buscar tipos de abatimento
        tipos_abatimento = ContasAReceberTipo.query.filter_by(
            tabela='contas_a_receber_abatimento',
            campo='tipo',
            ativo=True
        ).all()

        # Montar lista de abatimentos
        abatimentos_list = []

        # Primeiro: Desconto Contratual (não editável)
        if conta.desconto and conta.desconto > 0:
            desconto_percentual_str = f" ({conta.desconto_percentual * 100:.1f}%)" if conta.desconto_percentual else ""
            abatimentos_list.append({
                'id': None,  # Não é um registro real
                'tipo': f'Desconto Contratual{desconto_percentual_str}',
                'tipo_id': None,
                'motivo': 'Desconto definido no Odoo',
                'doc_motivo': None,
                'valor': conta.desconto,
                'previsto': False,
                'data': conta.emissao.isoformat() if conta.emissao else None,
                'data_vencimento': None,
                'editavel': False  # Flag para o frontend
            })

        # Depois: Abatimentos reais
        for ab in conta.abatimentos.all():
            abatimentos_list.append({
                'id': ab.id,
                'tipo': ab.tipo.tipo if ab.tipo else None,
                'tipo_id': ab.tipo_id,
                'motivo': ab.motivo,
                'doc_motivo': ab.doc_motivo,
                'valor': ab.valor,
                'previsto': ab.previsto,
                'data': ab.data.isoformat() if ab.data else None,
                'data_vencimento': ab.data_vencimento.isoformat() if ab.data_vencimento else None,
                'editavel': True
            })

        # Calcular totais separados
        desconto_contratual = conta.desconto or 0
        total_abatimentos_sem_desconto = sum(ab.valor or 0 for ab in conta.abatimentos.all())
        total_abatimentos = desconto_contratual + total_abatimentos_sem_desconto
        valor_original = conta.valor_original or 0
        valor_titulo = conta.valor_titulo or 0

        return jsonify({
            'success': True,
            'abatimentos': abatimentos_list,
            'titulo_nf': conta.titulo_nf,
            'valor_original': valor_original,
            'desconto_contratual': desconto_contratual,
            'total_abatimentos_sem_desconto': total_abatimentos_sem_desconto,
            'total_abatimentos': total_abatimentos,
            'valor_titulo': valor_titulo,
            'tipos_abatimento': [{'id': t.id, 'tipo': t.tipo} for t in tipos_abatimento]
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@financeiro_bp.route('/contas-receber/api/<int:conta_id>/abatimentos', methods=['POST'])
@login_required
def api_criar_abatimento(conta_id):
    """
    Cria novo abatimento para uma conta a receber
    """
    try:
        from app.financeiro.models import ContasAReceber, ContasAReceberAbatimento

        conta = ContasAReceber.query.get_or_404(conta_id)
        data = request.get_json()

        abatimento = ContasAReceberAbatimento(
            conta_a_receber_id=conta_id,
            tipo_id=data.get('tipo_id'),
            motivo=data.get('motivo', ''),
            doc_motivo=data.get('doc_motivo', ''),
            valor=float(data.get('valor', 0)),
            previsto=data.get('previsto', True),
            criado_por=current_user.nome
        )

        # Converter datas
        if data.get('data'):
            abatimento.data = datetime.strptime(data['data'], '%Y-%m-%d').date()
        if data.get('data_vencimento'):
            abatimento.data_vencimento = datetime.strptime(data['data_vencimento'], '%Y-%m-%d').date()

        db.session.add(abatimento)

        # Recalcular valor do título
        conta.atualizar_valor_titulo()
        conta.atualizado_por = current_user.nome
        conta.atualizado_em = agora_utc_naive()

        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Abatimento criado com sucesso!',
            'abatimento_id': abatimento.id,
            'novo_valor_titulo': conta.valor_titulo
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@financeiro_bp.route('/contas-receber/api/abatimentos/<int:abatimento_id>', methods=['PUT'])
@login_required
def api_atualizar_abatimento(abatimento_id):
    """
    Atualiza abatimento existente
    """
    try:
        from app.financeiro.models import ContasAReceberAbatimento

        abatimento = ContasAReceberAbatimento.query.get_or_404(abatimento_id)
        data = request.get_json()

        abatimento.tipo_id = data.get('tipo_id')
        abatimento.motivo = data.get('motivo', '')
        abatimento.doc_motivo = data.get('doc_motivo', '')
        abatimento.valor = float(data.get('valor', 0))
        abatimento.previsto = data.get('previsto', True)
        abatimento.atualizado_por = current_user.nome
        abatimento.atualizado_em = agora_utc_naive()

        # Converter datas
        if data.get('data'):
            abatimento.data = datetime.strptime(data['data'], '%Y-%m-%d').date()
        else:
            abatimento.data = None
        if data.get('data_vencimento'):
            abatimento.data_vencimento = datetime.strptime(data['data_vencimento'], '%Y-%m-%d').date()
        else:
            abatimento.data_vencimento = None

        # Recalcular valor do título da conta
        conta = abatimento.conta_a_receber
        conta.atualizar_valor_titulo()
        conta.atualizado_por = current_user.nome
        conta.atualizado_em = agora_utc_naive()

        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Abatimento atualizado com sucesso!',
            'novo_valor_titulo': conta.valor_titulo
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@financeiro_bp.route('/contas-receber/api/abatimentos/<int:abatimento_id>', methods=['DELETE'])
@login_required
def api_excluir_abatimento(abatimento_id):
    """
    Exclui abatimento
    """
    try:
        from app.financeiro.models import ContasAReceberAbatimento

        abatimento = ContasAReceberAbatimento.query.get_or_404(abatimento_id)
        conta = abatimento.conta_a_receber

        db.session.delete(abatimento)

        # Recalcular valor do título
        conta.atualizar_valor_titulo()
        conta.atualizado_por = current_user.nome
        conta.atualizado_em = agora_utc_naive()

        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Abatimento excluído com sucesso!',
            'novo_valor_titulo': conta.valor_titulo
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ========================================
# BUSCAR NF-PARCELA PARA AUTOCOMPLETE
# ========================================

@financeiro_bp.route('/contas-receber/api/buscar-titulos')
@login_required
def api_buscar_titulos():
    """
    API para autocomplete de NF-Parcela
    Retorna lista de títulos que correspondem à busca
    """
    from app.financeiro.models import ContasAReceber

    termo = request.args.get('q', '').strip()
    if len(termo) < 2:
        return jsonify({'success': True, 'titulos': []})

    # Buscar NFs que contenham o termo
    query = ContasAReceber.query.filter(
        ContasAReceber.titulo_nf.ilike(f'%{termo}%')
    ).order_by(
        ContasAReceber.titulo_nf,
        ContasAReceber.parcela
    ).limit(20)

    titulos = query.all()

    return jsonify({
        'success': True,
        'titulos': [{
            'id': t.id,
            'titulo_nf': t.titulo_nf,
            'parcela': t.parcela,
            'display': f"{t.titulo_nf}-{t.parcela}",
            'cliente': t.raz_social_red or t.raz_social or '-',
            'empresa': t.empresa
        } for t in titulos]
    })


@financeiro_bp.route('/contas-receber/api/titulo/<int:conta_id>')
@login_required
def api_detalhe_titulo(conta_id):
    """
    API para obter detalhes de um título específico
    Usado para carregar dinamicamente cliente e empresa ao selecionar NF-Parcela
    """
    from app.financeiro.models import ContasAReceber

    try:
        conta = ContasAReceber.query.get_or_404(conta_id)

        return jsonify({
            'success': True,
            'titulo': {
                'id': conta.id,
                'titulo_nf': conta.titulo_nf,
                'parcela': conta.parcela,
                'cliente': conta.raz_social_red or conta.raz_social or '-',
                'empresa': conta.empresa,
                'cnpj': conta.cnpj,
                'valor_titulo': conta.valor_titulo
            }
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ========================================
# EXPORTAR ABATIMENTOS PARA EXCEL
# ========================================

@financeiro_bp.route('/contas-receber/abatimentos/exportar-excel')
@login_required
def exportar_abatimentos_excel():
    """
    Exporta abatimentos para Excel com os mesmos filtros da listagem
    """
    from app.financeiro.models import ContasAReceberAbatimento, ContasAReceber

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        # Filtros (mesmos da listagem)
        titulo_nf = request.args.get('titulo_nf', '')
        cliente = request.args.get('cliente', '')
        tipo_id = request.args.get('tipo_id', '')
        previsto = request.args.get('previsto', '')
        data_de = request.args.get('data_de', '')
        data_ate = request.args.get('data_ate', '')

        # Query base com join
        query = db.session.query(
            ContasAReceberAbatimento,
            ContasAReceber.titulo_nf,
            ContasAReceber.parcela,
            ContasAReceber.raz_social_red,
            ContasAReceber.empresa
        ).join(
            ContasAReceber, ContasAReceberAbatimento.conta_a_receber_id == ContasAReceber.id
        )

        # Aplicar filtros
        if titulo_nf:
            query = query.filter(ContasAReceber.titulo_nf.ilike(f'%{titulo_nf}%'))
        if cliente:
            query = query.filter(ContasAReceber.raz_social_red.ilike(f'%{cliente}%'))
        if tipo_id:
            query = query.filter(ContasAReceberAbatimento.tipo_id == int(tipo_id))
        if previsto:
            query = query.filter(ContasAReceberAbatimento.previsto == (previsto == 'true'))
        if data_de:
            query = query.filter(ContasAReceberAbatimento.data >= datetime.strptime(data_de, '%Y-%m-%d').date())
        if data_ate:
            query = query.filter(ContasAReceberAbatimento.data <= datetime.strptime(data_ate, '%Y-%m-%d').date())

        # Ordenar por data desc
        query = query.order_by(ContasAReceberAbatimento.criado_em.desc())

        resultados = query.all()

        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Abatimentos"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="343a40", end_color="343a40", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Cabeçalhos
        headers = ["Emp", "NF-Parcela", "Cliente", "Tipo", "Motivo", "Documento", "Valor", "Previsto", "Data", "Criado Em"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Dados
        for row, item in enumerate(resultados, 2):
            abat = item[0]
            titulo_nf_val = item[1]
            parcela = item[2]
            cliente_val = item[3]
            empresa = item[4]

            ws.cell(row=row, column=1, value=empresa).border = thin_border
            ws.cell(row=row, column=2, value=f"{titulo_nf_val}-{parcela}").border = thin_border
            ws.cell(row=row, column=3, value=cliente_val or '-').border = thin_border
            ws.cell(row=row, column=4, value=abat.tipo.tipo if abat.tipo else '-').border = thin_border
            ws.cell(row=row, column=5, value=abat.motivo or '-').border = thin_border
            ws.cell(row=row, column=6, value=abat.doc_motivo or '-').border = thin_border
            ws.cell(row=row, column=7, value=abat.valor or 0).border = thin_border
            ws.cell(row=row, column=7).number_format = 'R$ #,##0.00'
            ws.cell(row=row, column=8, value='Sim' if abat.previsto else 'Não').border = thin_border
            ws.cell(row=row, column=9, value=abat.data.strftime('%d/%m/%Y') if abat.data else '-').border = thin_border
            ws.cell(row=row, column=10, value=abat.criado_em.strftime('%d/%m/%Y %H:%M') if abat.criado_em else '-').border = thin_border

        # Ajustar larguras
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 16

        # Salvar em BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Nome do arquivo
        nome_arquivo = f"abatimentos_{agora_utc_naive().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment;filename={nome_arquivo}'}
        )

    except ImportError:
        return jsonify({'success': False, 'error': 'Biblioteca openpyxl não instalada'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ========================================
# CRUD: ABATIMENTOS (LISTAGEM GERAL)
# ========================================

@financeiro_bp.route('/contas-receber/abatimentos')
@login_required
def listar_abatimentos():
    """
    Listagem geral de todos os abatimentos cadastrados no sistema
    """
    from app.financeiro.models import ContasAReceberAbatimento, ContasAReceber, ContasAReceberTipo

    # Filtros
    titulo_nf = request.args.get('titulo_nf', '')
    cliente = request.args.get('cliente', '')
    tipo_id = request.args.get('tipo_id', '')
    previsto = request.args.get('previsto', '')
    data_de = request.args.get('data_de', '')
    data_ate = request.args.get('data_ate', '')

    # Query base com join
    query = db.session.query(
        ContasAReceberAbatimento,
        ContasAReceber.titulo_nf,
        ContasAReceber.parcela,
        ContasAReceber.raz_social_red,
        ContasAReceber.empresa
    ).join(
        ContasAReceber, ContasAReceberAbatimento.conta_a_receber_id == ContasAReceber.id
    )

    # Aplicar filtros
    if titulo_nf:
        query = query.filter(ContasAReceber.titulo_nf.ilike(f'%{titulo_nf}%'))
    if cliente:
        query = query.filter(ContasAReceber.raz_social_red.ilike(f'%{cliente}%'))
    if tipo_id:
        query = query.filter(ContasAReceberAbatimento.tipo_id == int(tipo_id))
    if previsto:
        query = query.filter(ContasAReceberAbatimento.previsto == (previsto == 'true'))
    if data_de:
        query = query.filter(ContasAReceberAbatimento.data >= datetime.strptime(data_de, '%Y-%m-%d').date())
    if data_ate:
        query = query.filter(ContasAReceberAbatimento.data <= datetime.strptime(data_ate, '%Y-%m-%d').date())

    # Ordenar por data desc
    query = query.order_by(ContasAReceberAbatimento.criado_em.desc())

    # Paginação
    page = request.args.get('page', 1, type=int)
    per_page = 50
    paginacao = query.paginate(page=page, per_page=per_page, error_out=False)

    # Tipos para filtro
    tipos = ContasAReceberTipo.query.filter_by(
        tabela='contas_a_receber_abatimento',
        campo='tipo',
        ativo=True
    ).order_by(ContasAReceberTipo.tipo).all()

    return render_template(
        'financeiro/crud_abatimentos.html',
        abatimentos=paginacao.items,
        paginacao=paginacao,
        tipos=tipos
    )
