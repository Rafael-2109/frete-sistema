"""
Exportador Excel do Validador de Titulos x Bancos.

Gera um .xlsx com uma aba por resultado, para download e conferencia offline.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font, PatternFill

_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_VERIFICAR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_OK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


def _sanitizar(valor):
    """Remove caracteres ilegais (NUL etc.) que o openpyxl rejeita em celulas."""
    if isinstance(valor, str):
        return ILLEGAL_CHARACTERS_RE.sub("", valor)
    return valor


def _append(ws, valores):
    """append sanitizando cada celula textual (blindagem contra caractere ilegal)."""
    ws.append([_sanitizar(v) for v in valores])


def _escrever_cabecalho(ws, colunas):
    ws.append(colunas)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT


def gerar_excel(rv) -> BytesIO:
    """Gera o Excel a partir de um ResultadoValidacao. Retorna BytesIO pronto p/ download."""
    wb = Workbook()

    # --- RESUMO ---
    ws = wb.active
    ws.title = "RESUMO"
    _escrever_cabecalho(ws, ["Indicador", "Quantidade"])
    rotulos = {
        "qtd_boletos": "Boletos lidos (4 bancos)",
        "qtd_identificados": "Boletos identificados",
        "qtd_nao_identificados": "Boletos NAO identificados",
        "qtd_recompras_cp": "Recompras no CP-NACOM",
        "qtd_faturamento": "Notas faturadas",
        "qtd_duplicados": "Titulos em 2+ bancos",
        "qtd_duplicados_com_recompra": "  ... com recompra (OK)",
        "qtd_duplicados_sem_recompra": "  ... SEM recompra (verificar)",
        "qtd_faturado_sem_boleto": "Faturado SEM boleto",
        "qtd_boleto_sem_nota": "Boleto SEM nota",
    }
    for chave, rotulo in rotulos.items():
        if chave in rv.resumo:
            _append(ws, [rotulo, rv.resumo[chave]])

    # --- DUPLICADOS ---
    ws = wb.create_sheet("DUPLICADOS")
    _escrever_cabecalho(ws, ["NF-PARC", "BANCOS", "QTD BANCOS", "RECOMPRA"])
    for d in rv.resultado.duplicados:
        flag = "RECOMPRA OK" if d["tem_recompra"] else "VERIFICAR"
        _append(ws, [d["nf_parc"], ", ".join(d["bancos"]), d["qtd_bancos"], flag])
        cell = ws.cell(row=ws.max_row, column=4)
        cell.fill = _OK_FILL if d["tem_recompra"] else _VERIFICAR_FILL

    # --- FATURADO SEM BOLETO ---
    ws = wb.create_sheet("FATURADO SEM BOLETO")
    _escrever_cabecalho(ws, ["NF-PARC", "EMPRESA", "CLIENTE", "CNPJ", "UF", "VENCIMENTO", "VALOR", "TIPO"])
    for f in rv.resultado.faturado_sem_boleto:
        _append(ws, [
            f.get("nf_parc"), f.get("empresa"), f.get("cliente"), f.get("cnpj"),
            f.get("uf"), _data_str(f.get("vencimento")), f.get("valor"), f.get("tipo_titulo"),
        ])

    # --- BOLETO SEM NOTA ---
    ws = wb.create_sheet("BOLETO SEM NOTA")
    _escrever_cabecalho(ws, ["NF-PARC", "BANCOS"])
    for b in rv.resultado.boleto_sem_nota:
        _append(ws, [b["nf_parc"], ", ".join(b["bancos"])])

    # --- NAO IDENTIFICADOS ---
    ws = wb.create_sheet("NAO IDENTIFICADOS")
    _escrever_cabecalho(ws, ["BANCO", "ORIGINAL", "VALOR", "VENCIMENTO"])
    for n in rv.resultado.nao_identificados:
        _append(ws, [n.get("banco"), n.get("original"), n.get("valor"),
                     _data_str(n.get("vencimento"))])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _data_str(valor):
    """Formata data/valor de vencimento para texto simples."""
    if valor is None:
        return ""
    try:
        return valor.strftime("%d/%m/%Y")
    except AttributeError:
        return str(valor)
