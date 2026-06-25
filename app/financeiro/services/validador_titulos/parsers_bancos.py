"""
Parsers das bases de banco do Validador de Titulos x Bancos.

Cada banco (SRM, GRAFENO, AGIS, VORTX) exporta os boletos num layout proprio.
Este modulo isola essa diversidade atras de uma interface unica:

    extrair_boletos(linhas, banco) -> List[dict]   # logica pura (matriz -> boletos)
    parsear_arquivo(arquivo, banco) -> List[dict]   # le o arquivo e extrai

Cada boleto: {nf_parc, banco, original, valor, vencimento, pagador}.
A coluna identificadora e localizada POR NOME (tolerante a acento/encoding/ordem),
nunca por posicao fixa — o cabecalho dos bancos GRAFENO/VORTX vem depois de
linhas de resumo, entao a linha de cabecalho e detectada dinamicamente.
"""

import csv
import logging
import os
import unicodedata
from typing import List, Optional

from app.financeiro.services.validador_titulos.normalizador import montar_nf_parc

logger = logging.getLogger(__name__)

EXTENSOES_SUPORTADAS = (".xlsx", ".xlsm", ".xlsb", ".csv", ".txt")


# Configuracao por banco. Nomes-alvo de coluna sao casados de forma normalizada
# (sem acento, sem pontuacao, minusculo), entao "Seu Numero" casa "Seu_Número".
CONFIG_BANCOS = {
    "SRM": {
        "col_id": "Nro Documento",
        "col_valor": "Valor",
        "col_venc": "Data Vencimento",
        "col_pagador": "Nome do Sacado",
    },
    "GRAFENO": {
        "col_id": "Seu Numero",
        "col_valor": "Valor Cobranca",
        "col_venc": "Data Vencimento",
        "col_pagador": "Pagador",
    },
    "VORTX": {
        "col_id": "Seu Numero",
        "col_valor": "Valor Cobranca",
        "col_venc": "Data Vencimento",
        "col_pagador": "Pagador",
    },
    "AGIS": {
        "col_id": "Recebivel",
        "col_valor": "Valor Aberto",
        "col_venc": None,
        "col_pagador": None,
    },
}

BANCOS_SUPORTADOS = list(CONFIG_BANCOS.keys())


def _norm_col(valor) -> str:
    """Normaliza nome de coluna: sem acento, sem pontuacao, minusculo."""
    if valor is None:
        return ""
    txt = str(valor)
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    return "".join(c for c in txt.lower() if c.isalnum())


def _parse_valor(valor) -> Optional[float]:
    """Converte valor monetario (BR 'R$ 16.252,62' ou '1745.17') em float."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    txt = str(valor).strip().replace("R$", "").replace(" ", "")
    if not txt:
        return None
    if "," in txt:
        # formato brasileiro: ponto = milhar, virgula = decimal
        txt = txt.replace(".", "").replace(",", ".")
    try:
        return float(txt)
    except ValueError:
        logger.warning("Valor nao numerico ignorado: %r", valor)
        return None


def _achar_indices(linhas, config):
    """Localiza a linha de cabecalho e os indices das colunas configuradas.

    Retorna (idx_header, mapa_colunas). Levanta ValueError se a coluna
    identificadora nao for encontrada em nenhuma linha.
    """
    alvo_id = _norm_col(config["col_id"])
    for idx, linha in enumerate(linhas):
        normalizadas = [_norm_col(c) for c in linha]
        col_id = _achar_coluna(normalizadas, alvo_id)
        if col_id is not None:
            mapa = {"id": col_id}
            for chave, nome in (
                ("valor", config.get("col_valor")),
                ("venc", config.get("col_venc")),
                ("pagador", config.get("col_pagador")),
            ):
                if nome:
                    mapa[chave] = _achar_coluna(normalizadas, _norm_col(nome))
                else:
                    mapa[chave] = None
            return idx, mapa
    raise ValueError(
        f"Coluna identificadora '{config['col_id']}' nao encontrada — "
        f"o arquivo enviado nao parece ser desse banco."
    )


def _achar_coluna(colunas_normalizadas, alvo) -> Optional[int]:
    """Indice da coluna cujo nome normalizado e igual ou CONTEM o alvo.

    So `alvo in c` (cabecalho mais descritivo que o alvo). A direcao inversa
    (`c in alvo`) foi removida: deixava uma celula curta de resumo casar a
    coluna errada (ex: 'v' casando 'valor').
    """
    for i, c in enumerate(colunas_normalizadas):
        if c == alvo:
            return i
    for i, c in enumerate(colunas_normalizadas):
        if c and alvo in c:
            return i
    return None


def _celula(linha, idx):
    if idx is None or idx >= len(linha):
        return None
    return linha[idx]


def extrair_boletos(linhas: List[list], banco: str) -> List[dict]:
    """
    Interpreta a matriz ja lida do arquivo do `banco` e devolve a lista de boletos.

    - Detecta a linha de cabecalho pela coluna identificadora (tolerante a resumo no topo).
    - Ignora linhas com identificador vazio.
    - Identificador presente mas nao normalizavel entra com nf_parc=None
      (vira "nao identificado" no comparador, para conferencia manual).
    """
    banco = banco.upper()
    if banco not in CONFIG_BANCOS:
        raise ValueError(f"Banco nao suportado: {banco}. Use um de {BANCOS_SUPORTADOS}.")

    config = CONFIG_BANCOS[banco]
    idx_header, mapa = _achar_indices(linhas, config)

    boletos = []
    for linha in linhas[idx_header + 1:]:
        identificador = _celula(linha, mapa["id"])
        ident_txt = "" if identificador is None else str(identificador).strip()
        if not ident_txt:
            continue
        boletos.append({
            "nf_parc": montar_nf_parc(ident_txt),
            "banco": banco,
            "original": ident_txt,
            "valor": _parse_valor(_celula(linha, mapa["valor"])),
            "vencimento": _celula(linha, mapa["venc"]),
            "pagador": _celula(linha, mapa["pagador"]),
        })
    return boletos


def ler_arquivo(caminho: str, aba: Optional[str] = None) -> List[list]:
    """Le um arquivo xlsx/xlsm/xlsb/csv/txt e devolve a matriz de linhas.

    Se `aba` for informado (planilhas Excel), seleciona a aba pelo nome de forma
    tolerante (ignora espacos/hifens/maiusculas). CSV ignora `aba`.
    """
    ext = os.path.splitext(caminho)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return _ler_xlsx(caminho, aba)
    if ext == ".xlsb":
        return _ler_xlsb(caminho, aba)
    if ext in (".csv", ".txt"):
        return _ler_csv(caminho)
    raise ValueError(
        f"Extensao nao suportada: {ext}. Use um de {EXTENSOES_SUPORTADAS}."
    )


def _escolher_aba(nomes: List[str], aba: Optional[str]) -> str:
    """Resolve o nome real da aba a partir de um alvo tolerante a formatacao."""
    if aba is None:
        return nomes[0]
    alvo = _norm_col(aba)
    for nome in nomes:
        if _norm_col(nome) == alvo:
            return nome
    raise ValueError(
        f"Aba '{aba}' nao encontrada no arquivo. Abas disponiveis: {nomes}."
    )


def _ler_xlsx(caminho: str, aba: Optional[str] = None) -> List[list]:
    from openpyxl import load_workbook
    wb = load_workbook(caminho, read_only=True, data_only=True)
    try:
        nome = _escolher_aba(wb.sheetnames, aba)
        ws = wb[nome]
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _ler_xlsb(caminho: str, aba: Optional[str] = None) -> List[list]:
    from pyxlsb import open_workbook
    with open_workbook(caminho) as wb:
        nome = _escolher_aba(wb.sheets, aba)
        with wb.get_sheet(nome) as ws:
            return [[c.v for c in row] for row in ws.rows()]


def _ler_csv(caminho: str) -> List[list]:
    # Detecta encoding (utf-8 com fallback latin-1) e delimitador (; , tab)
    dados = None
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(caminho, "r", encoding=enc, newline="") as f:
                dados = f.read()
            break
        except UnicodeDecodeError:
            continue
    if dados is None:
        raise ValueError("Nao foi possivel decodificar o CSV (encoding desconhecido).")
    amostra = dados[:4096]
    try:
        dialect = csv.Sniffer().sniff(amostra, delimiters=";,\t")
    except csv.Error:
        # NAO mutar csv.excel (classe global -> poluiria o processo inteiro);
        # subclasse local preserva os demais atributos do excel.
        class _ExcelPontoVirgula(csv.excel):
            delimiter = ";"
        dialect = _ExcelPontoVirgula
    return [row for row in csv.reader(dados.splitlines(), dialect)]


def parsear_arquivo(caminho: str, banco: str) -> List[dict]:
    """Le o arquivo do `banco` e devolve a lista de boletos normalizados."""
    return extrair_boletos(ler_arquivo(caminho), banco)
