#!/usr/bin/env python3
"""
Solução REAL para normalizar planilhas Excel para o portal Sendas.

PROBLEMA REAL IDENTIFICADO:
- O Sendas espera uma estrutura ESPECÍFICA de estilos (14 cellXfs, 4 fonts, 4 fills, 5 borders)
- Recriar o arquivo do zero perde essa estrutura
- A solução é usar um arquivo que funciona como template

SOLUÇÃO:
- Copiar um arquivo que funciona como template
- Substituir apenas os dados, mantendo a estrutura de estilos
"""

import os
import shutil
import logging
from typing import Tuple
import openpyxl
from openpyxl import load_workbook
import pandas as pd

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def normalizar_usando_template(arquivo_entrada: str, arquivo_saida: str, template_path: str = None) -> Tuple[bool, str]:
    """
    Normaliza arquivo Excel usando um template que funciona.
    
    Esta é a solução correta: mantém a estrutura de estilos que o Sendas espera.
    
    Args:
        arquivo_entrada: Caminho do arquivo com dados para normalizar
        arquivo_saida: Caminho para salvar o arquivo normalizado
        template_path: Caminho do arquivo template (opcional)
        
    Returns:
        (sucesso, caminho_arquivo) - True se sucesso, False caso contrário
    """
    try:
        logger.info("🔧 Normalizando usando template...")
        
        # Se não forneceu template, usar o template padrão limpo
        if not template_path:
            # Template oficial limpo (sem dados, apenas estrutura)
            template_path = os.path.join(os.path.dirname(__file__), 'template_sendas_limpo.xlsx')
            
            if not os.path.exists(template_path):
                # Fallback para outros possíveis templates
                possible_templates = [
                    '/tmp/sendas_funciona.xlsx',
                    os.path.join(os.path.dirname(__file__), 'template_sendas.xlsx'),
                ]
                
                for template in possible_templates:
                    if os.path.exists(template):
                        template_path = template
                        logger.info(f"✅ Template alternativo encontrado: {template_path}")
                        break
                
                if not os.path.exists(template_path):
                    logger.error("❌ Template não encontrado")
                    logger.error(f"   Esperado em: {os.path.join(os.path.dirname(__file__), 'template_sendas_limpo.xlsx')}")
                    return False, arquivo_entrada
            else:
                logger.info(f"✅ Usando template oficial: {template_path}")
        
        # Verificar se template existe
        if not os.path.exists(template_path):
            logger.error(f"❌ Template não encontrado: {template_path}")
            return False, arquivo_entrada
        
        # Ler dados do arquivo de entrada
        logger.info("📖 Lendo dados do arquivo de entrada...")
        df_entrada = pd.read_excel(arquivo_entrada, header=None, engine='openpyxl')
        
        # Copiar template para arquivo de saída
        logger.info(f"📋 Copiando template para {arquivo_saida}...")
        shutil.copy2(template_path, arquivo_saida)
        
        # Abrir arquivo copiado com openpyxl
        logger.info("✏️ Substituindo dados no template...")
        wb = load_workbook(arquivo_saida)
        
        # Pegar a primeira planilha
        if wb.worksheets:
            ws = wb.worksheets[0]
        else:
            logger.error("❌ Template não tem planilhas")
            return False, arquivo_entrada
        
        # Limpar dados existentes (mantendo formatação)
        # Primeiro, descobrir o range de dados
        logger.info("🧹 Limpando dados antigos do template...")
        
        # Limpar células mas manter a formatação
        # Começar da linha 3 (onde geralmente começam os dados)
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                # Manter o estilo mas limpar o valor
                cell.value = None
        
        # Inserir novos dados
        logger.info("📝 Inserindo novos dados...")
        for row_idx, row_data in df_entrada.iterrows():
            excel_row = row_idx + 1  # Excel começa em 1
            
            for col_idx, value in enumerate(row_data):
                excel_col = col_idx + 1  # Excel começa em 1
                
                # Pular valores NaN
                if pd.notna(value):
                    try:
                        # Obter célula (mantém formatação existente)
                        cell = ws.cell(row=excel_row, column=excel_col)
                        
                        # Definir valor mantendo o estilo original
                        if isinstance(value, (int, float)):
                            cell.value = value
                        else:
                            cell.value = str(value)
                            
                    except Exception as e:
                        logger.debug(f"Erro ao definir célula ({excel_row}, {excel_col}): {e}")
                        continue
        
        # Salvar arquivo
        logger.info("💾 Salvando arquivo normalizado...")
        wb.save(arquivo_saida)
        wb.close()
        
        # Verificar se foi criado
        if os.path.exists(arquivo_saida):
            tamanho = os.path.getsize(arquivo_saida)
            logger.info(f"✅ Arquivo normalizado criado: {arquivo_saida} ({tamanho/1024:.1f} KB)")
            return True, arquivo_saida
        else:
            logger.error("❌ Arquivo normalizado não foi criado")
            return False, arquivo_entrada
            
    except Exception as e:
        logger.error(f"❌ Erro ao normalizar com template: {e}")
        import traceback
        logger.debug(traceback.format_exc())
        return False, arquivo_entrada


def normalizar_planilha_sendas(arquivo_entrada: str, arquivo_saida: str = None) -> Tuple[bool, str]:
    """
    Função principal de normalização usando template.
    
    Args:
        arquivo_entrada: Caminho do arquivo Excel para normalizar
        arquivo_saida: Caminho de saída (opcional)
        
    Returns:
        (sucesso, caminho_arquivo_normalizado)
    """
    
    # Validar entrada
    if not os.path.exists(arquivo_entrada):
        logger.error(f"❌ Arquivo não encontrado: {arquivo_entrada}")
        return False, arquivo_entrada
    
    # Gerar nome de saída se não fornecido
    if not arquivo_saida:
        base_name = os.path.splitext(os.path.basename(arquivo_entrada))[0]
        arquivo_saida = f"/tmp/{base_name}_normalizado.xlsx"
    
    logger.info("=" * 60)
    logger.info("🚀 NORMALIZAÇÃO COM TEMPLATE")
    logger.info(f"📥 Arquivo entrada: {arquivo_entrada}")
    logger.info(f"📤 Arquivo saída: {arquivo_saida}")
    logger.info("=" * 60)
    
    # Usar template para normalizar
    return normalizar_usando_template(arquivo_entrada, arquivo_saida)


# Teste
if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        arquivo_teste = sys.argv[1]
    else:
        arquivo_teste = "/tmp/sendas_nao_funciona.xlsx"
    
    if os.path.exists(arquivo_teste):
        logger.info(f"\n🧪 Testando normalização com template: {arquivo_teste}")
        sucesso, arquivo_normalizado = normalizar_planilha_sendas(arquivo_teste)
        
        if sucesso:
            logger.info(f"\n✅ Normalização com template bem-sucedida!")
            logger.info(f"📁 Arquivo normalizado: {arquivo_normalizado}")
        else:
            logger.error("\n❌ Normalização falhou")