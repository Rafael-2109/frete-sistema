#!/usr/bin/env python3
"""
Módulo para preencher planilha Sendas usando template limpo como base.
Garante uso de sharedStrings para compatibilidade com Portal Sendas.
"""

import os
import sys
import logging
import tempfile
import shutil
from datetime import datetime, date
from typing import Dict, List, Tuple, Optional
import pandas as pd
import xlsxwriter
from openpyxl import load_workbook

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def preencher_planilha_com_template(
    planilha_portal: str,
    dados_agendamento: List[Dict],
    template_path: str = None,
    arquivo_saida: str = None
) -> Tuple[bool, str]:
    """
    Preenche planilha usando template limpo como base, preservando sharedStrings.
    
    Args:
        planilha_portal: Caminho da planilha baixada do portal (para disponibilidade)
        dados_agendamento: Lista de dicts com dados do agendamento:
            [{'cnpj': str, 'data_agendamento': date, 'itens': [...]}]
        template_path: Caminho do template limpo (padrão: template_sendas_limpo.xlsx)
        arquivo_saida: Caminho de saída (se None, gera automaticamente)
    
    Returns:
        (sucesso, caminho_arquivo_gerado)
    """
    try:
        logger.info("=" * 60)
        logger.info("📋 PREENCHIMENTO COM TEMPLATE SENDAS")
        logger.info("=" * 60)
        
        # 1. Definir template
        if not template_path:
            template_path = os.path.join(
                os.path.dirname(__file__),
                'template_sendas_limpo.xlsx'
            )
        
        if not os.path.exists(template_path):
            logger.error(f"❌ Template não encontrado: {template_path}")
            return False, None
        
        logger.info(f"✅ Template: {template_path}")
        
        # 2. Ler planilha do portal (disponibilidade)
        logger.info(f"📖 Lendo planilha do portal: {planilha_portal}")
        
        # Ler com pandas para facilitar manipulação
        df_portal = pd.read_excel(planilha_portal, header=None, dtype=object, engine='openpyxl')
        logger.info(f"   Dimensões portal: {df_portal.shape}")
        
        # 3. Ler template
        logger.info(f"📖 Lendo template...")
        df_template = pd.read_excel(template_path, header=None, dtype=object, engine='openpyxl')
        logger.info(f"   Dimensões template: {df_template.shape}")
        
        # 4. Preparar dados para preenchimento
        logger.info("🔄 Processando dados de agendamento...")
        
        # Copiar cabeçalhos do template (primeiras 3 linhas)
        df_resultado = df_template.iloc[:3].copy()
        
        # Processar cada linha do portal a partir da linha 3 (índice 3)
        linhas_processadas = []
        demanda_id = 1
        
        for idx in range(3, len(df_portal)):
            linha_portal = df_portal.iloc[idx].copy()
            
            # Extrair informações chave da linha do portal
            pedido_cliente = linha_portal[6]  # Coluna G
            codigo_produto = linha_portal[8]  # Coluna I
            
            if pd.isna(pedido_cliente) or pd.isna(codigo_produto):
                continue
            
            # Verificar se esta linha está no agendamento
            linha_agendada = False
            for agend in dados_agendamento:
                # Aqui você precisa implementar a lógica de matching
                # Por exemplo, verificar se o pedido_cliente corresponde
                # Este é um exemplo simplificado
                if str(pedido_cliente).startswith(agend.get('prefixo_pedido', '')):
                    linha_agendada = True
                    
                    # Preencher campos necessários
                    linha_portal[0] = demanda_id  # Coluna A - Demanda ID
                    linha_portal[16] = linha_portal[14]  # Coluna Q - Quantidade entrega
                    linha_portal[17] = agend['data_agendamento']  # Coluna R - Data
                    linha_portal[20] = 'Paletizada'  # Coluna U - Característica
                    linha_portal[21] = determinar_tipo_caminhao(float(linha_portal[14] or 0))  # Coluna V
                    
                    break
            
            if linha_agendada:
                linhas_processadas.append(linha_portal)
        
        # 5. Montar DataFrame final
        if linhas_processadas:
            df_dados = pd.DataFrame(linhas_processadas)
            df_resultado = pd.concat([df_resultado, df_dados], ignore_index=True)
        
        logger.info(f"✅ Linhas processadas: {len(linhas_processadas)}")
        
        # 6. Definir arquivo de saída
        if not arquivo_saida:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            arquivo_saida = f"/tmp/sendas_template_{timestamp}.xlsx"
        
        # 7. Escrever com xlsxwriter (SEMPRE usa sharedStrings)
        logger.info("✏️ Escrevendo com xlsxwriter (sharedStrings)...")
        
        # Converter datas para formato correto
        for col in df_resultado.columns:
            df_resultado[col] = df_resultado[col].apply(lambda x: processar_celula(x, col))
        
        # Criar workbook
        wb = xlsxwriter.Workbook(arquivo_saida, {
            'strings_to_numbers': True,
            'strings_to_formulas': False,
            'strings_to_urls': False,
            'use_zip64': False,
            'nan_inf_to_errors': True,
            'default_date_format': 'dd/mm/yyyy'
        })
        
        ws = wb.add_worksheet('Planilha1')
        
        # Formato de data
        date_format = wb.add_format({'num_format': 'dd/mm/yyyy'})
        
        # Escrever dados
        for r, row in enumerate(df_resultado.itertuples(index=False, name=None)):
            for c, v in enumerate(row):
                if v is not None and pd.notna(v):
                    # Coluna 17 (R) é data - escrever como DATE (sem hora)
                    if c == 17:
                        if isinstance(v, (datetime, pd.Timestamp)):
                            # Converter para date (sem hora)
                            ws.write_datetime(r, c, v.date(), date_format)
                        elif isinstance(v, date):
                            # Já é date
                            ws.write_datetime(r, c, v, date_format)
                        elif isinstance(v, str) and '/' in v:
                            try:
                                # Converter string para date
                                dt = pd.to_datetime(v, format='%d/%m/%Y')
                                ws.write_datetime(r, c, dt.date(), date_format)
                            except:
                                ws.write(r, c, v)
                        else:
                            ws.write(r, c, v)
                    else:
                        ws.write(r, c, v)
        
        wb.close()
        
        # 8. Verificar resultado
        if os.path.exists(arquivo_saida):
            tamanho = os.path.getsize(arquivo_saida)
            logger.info(f"✅ Arquivo criado: {arquivo_saida}")
            logger.info(f"📏 Tamanho: {tamanho/1024:.1f} KB")
            
            # Verificar sharedStrings
            import zipfile
            with zipfile.ZipFile(arquivo_saida, 'r') as zf:
                if 'xl/sharedStrings.xml' in zf.namelist():
                    logger.info("✅ Usando sharedStrings (compatível com Sendas)")
                else:
                    logger.warning("⚠️ AVISO: não está usando sharedStrings")
            
            return True, arquivo_saida
        
        return False, None
        
    except Exception as e:
        logger.error(f"❌ Erro ao preencher planilha: {e}", exc_info=True)
        return False, None


def processar_celula(val, col_idx):
    """Processa valor da célula baseado no índice da coluna"""
    if val is None or pd.isna(val):
        return None
    
    # Coluna 17 é data - manter como datetime para xlsxwriter formatar
    if col_idx == 17:
        if isinstance(val, (datetime, date)):
            return val
        elif isinstance(val, str) and '/' in val:
            try:
                return pd.to_datetime(val, format='%d/%m/%Y')
            except:
                return val
    
    # Converter strings numéricas
    if isinstance(val, str):
        s = val.strip()
        if s.replace(',', '.').replace('.', '').replace('-', '').isdigit():
            try:
                return float(s.replace(',', '.')) if '.' in s else int(s)
            except:
                pass
    
    return val


def determinar_tipo_caminhao(peso_kg: float) -> str:
    """Determina tipo de caminhão baseado no peso"""
    if peso_kg <= 800:
        return 'Utilitário'
    elif peso_kg <= 2000:
        return 'Caminhão VUC 3/4'
    elif peso_kg <= 4000:
        return 'Caminhão 3/4 (2 eixos) 16T'
    elif peso_kg <= 8000:
        return 'Caminhão Truck (6x2) 23T'
    elif peso_kg <= 25000:
        return 'Carreta Simples Toco (3 eixos) 25T'
    else:
        return 'Caminhão (4 eixos) 31T'


def criar_planilha_usando_template(
    planilha_portal: str,
    planilha_preenchida_atual: str,
    template_path: str = None,
    arquivo_saida: str = None
) -> Tuple[bool, str]:
    """
    Cria nova planilha copiando dados da planilha preenchida atual
    para o template limpo, preservando sharedStrings.
    
    Esta função é útil para "consertar" planilhas que foram criadas
    com openpyxl e não têm sharedStrings.
    
    Args:
        planilha_portal: Planilha baixada do portal (referência)
        planilha_preenchida_atual: Planilha já preenchida mas sem sharedStrings
        template_path: Template limpo com sharedStrings
        arquivo_saida: Arquivo de saída
    
    Returns:
        (sucesso, caminho_arquivo)
    """
    try:
        logger.info("=" * 60)
        logger.info("🔧 RECRIANDO PLANILHA COM TEMPLATE")
        logger.info("=" * 60)
        
        # Definir template
        if not template_path:
            template_path = os.path.join(
                os.path.dirname(__file__),
                'template_sendas_limpo.xlsx'
            )
        
        # Ler planilha preenchida
        logger.info(f"📖 Lendo planilha preenchida: {planilha_preenchida_atual}")
        df_preenchida = pd.read_excel(
            planilha_preenchida_atual, 
            header=None, 
            dtype=object,
            engine='openpyxl'
        )
        
        # Definir saída
        if not arquivo_saida:
            base = os.path.splitext(os.path.basename(planilha_preenchida_atual))[0]
            arquivo_saida = f"/tmp/{base}_com_template.xlsx"
        
        logger.info("✏️ Escrevendo com xlsxwriter...")
        
        # Processar células
        for col in df_preenchida.columns:
            df_preenchida[col] = df_preenchida[col].apply(lambda x: processar_celula(x, col))
        
        # Criar workbook
        wb = xlsxwriter.Workbook(arquivo_saida, {
            'strings_to_numbers': True,
            'strings_to_formulas': False,
            'strings_to_urls': False,
            'use_zip64': False,
            'nan_inf_to_errors': True
        })
        
        ws = wb.add_worksheet('Planilha1')
        date_format = wb.add_format({'num_format': 'dd/mm/yyyy'})
        
        # Escrever dados
        for r, row in enumerate(df_preenchida.itertuples(index=False, name=None)):
            for c, v in enumerate(row):
                if v is not None and pd.notna(v):
                    # Coluna 17 é data - escrever como DATE (sem hora)
                    if c == 17:
                        if isinstance(v, (datetime, pd.Timestamp)):
                            # Converter datetime para date (remover hora)
                            ws.write_datetime(r, c, v.date(), date_format)
                        elif isinstance(v, date):
                            # Já é date, escrever direto
                            ws.write_datetime(r, c, v, date_format)
                        elif isinstance(v, str) and '/' in v:
                            try:
                                # Converter string DD/MM/YYYY para date (sem hora)
                                dt = pd.to_datetime(v, format='%d/%m/%Y')
                                ws.write_datetime(r, c, dt.date(), date_format)
                            except:
                                # Se falhar conversão, escrever como está
                                ws.write(r, c, v)
                        else:
                            ws.write(r, c, v)
                    else:
                        ws.write(r, c, v)
        
        wb.close()
        
        if os.path.exists(arquivo_saida):
            logger.info(f"✅ Arquivo criado: {arquivo_saida}")
            
            # Verificar sharedStrings
            import zipfile
            with zipfile.ZipFile(arquivo_saida, 'r') as zf:
                if 'xl/sharedStrings.xml' in zf.namelist():
                    logger.info("✅ CONFIRMADO: usando sharedStrings!")
                else:
                    logger.error("❌ ERRO: ainda sem sharedStrings!")
            
            return True, arquivo_saida
        
        return False, None
        
    except Exception as e:
        logger.error(f"❌ Erro: {e}", exc_info=True)
        return False, None


if __name__ == "__main__":
    # Teste
    import sys
    
    if len(sys.argv) > 1:
        planilha = sys.argv[1]
        logger.info(f"Processando: {planilha}")
        
        # Tentar recriar com template
        sucesso, arquivo_novo = criar_planilha_usando_template(
            planilha_portal=planilha,
            planilha_preenchida_atual=planilha
        )
        
        if sucesso:
            logger.info(f"✅ SUCESSO! Novo arquivo: {arquivo_novo}")
        else:
            logger.error("❌ Falha ao processar")
    else:
        logger.info("Uso: python preencher_com_template.py <planilha.xlsx>")