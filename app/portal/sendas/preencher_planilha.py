#!/usr/bin/env python3
"""
Módulo para preencher planilha de agendamento do Portal Sendas (Trizy)
Preenche os campos necessários da planilha baixada com dados do banco
SEGUE ESTRITAMENTE mapeamento_planilha.md
BUSCA EXATAMENTE COMO programacao_em_lote
"""

import os
import sys
from datetime import datetime, date
from decimal import Decimal
from typing import Dict, Any
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from copy import copy
from sqlalchemy import and_, func
import logging
import tempfile
import pandas as pd
import xlsxwriter

# Adicionar o caminho do projeto
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))))

from app import create_app, db
from app.separacao.models import Separacao
from app.carteira.models import CarteiraPrincipal
from app.producao.models import CadastroPalletizacao
from app.portal.sendas.models import FilialDeParaSendas, ProdutoDeParaSendas

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class PreencherPlanilhaSendas:
    """Classe para preencher planilha de agendamento do Sendas"""
    
    # Mapeamento de peso para tipo de caminhão (conforme mapeamento_planilha.md)
    CAMINHOES = [
        ('Utilitário', 800),
        ('Caminhão VUC 3/4', 2000),
        ('Caminhão 3/4 (2 eixos) 16T', 4000),
        ('Caminhão Truck (6x2) 23T', 8000),
        ('Carreta Simples Toco (3 eixos) 25T', 25000),
        ('Caminhão (4 eixos) 31T', float('inf'))  # Acima de 25000
    ]
    
    def __init__(self, app=None):
        """Inicializa o preenchedor de planilhas"""
        # Não criar novo contexto - usar o contexto existente do Flask
        from flask import current_app, has_app_context
        
        if has_app_context():
            # Se já estamos em um contexto Flask, usar ele
            self.app = current_app._get_current_object()
        else:
            # Se não há contexto (ex: script standalone), criar app
            self.app = app or create_app()
            # NÃO fazer push do contexto aqui - deixar para quando for necessário
    
    def _converter_para_xlsxwriter(self, arquivo_origem: str, arquivo_destino: str) -> bool:
        """
        Converte arquivo Excel para usar xlsxwriter (garante sharedStrings e datas corretas)
        
        Args:
            arquivo_origem: Arquivo criado com openpyxl
            arquivo_destino: Arquivo de saída com xlsxwriter
            
        Returns:
            bool: True se sucesso, False se erro
        """
        try:
            # Ler arquivo com pandas
            df = pd.read_excel(arquivo_origem, header=None, dtype=object, engine='openpyxl')
            
            # Criar workbook com xlsxwriter
            wb = xlsxwriter.Workbook(arquivo_destino, {
                'strings_to_numbers': True,
                'strings_to_formulas': False,
                'strings_to_urls': False,
                'use_zip64': False,
                'nan_inf_to_errors': True,
                'default_date_format': 'dd/mm/yyyy'
            })
            
            ws = wb.add_worksheet('Planilha1')
            date_format = wb.add_format({'num_format': 'dd/mm/yyyy'})
            
            # Escrever dados
            for r, row in enumerate(df.itertuples(index=False, name=None)):
                for c, v in enumerate(row):
                    if v is not None and pd.notna(v):
                        # Coluna 17 (R) é data - escrever como datetime
                        if c == 17:
                            if isinstance(v, (datetime, pd.Timestamp)):
                                # Converter para date (sem hora)
                                ws.write_datetime(r, c, v.date() if hasattr(v, 'date') else v, date_format)
                            elif isinstance(v, date):
                                ws.write_datetime(r, c, v, date_format)
                            elif isinstance(v, str) and '/' in v:
                                try:
                                    # Converter string para date
                                    dt = pd.to_datetime(v, format='%d/%m/%Y')
                                    ws.write_datetime(r, c, dt.date(), date_format)
                                except:
                                    ws.write(r, c, v)
                            else:
                                ws.write(r, c, v)
                        else:
                            # Para outras colunas, escrever normalmente
                            ws.write(r, c, v)
            
            wb.close()
            return True
            
        except Exception as e:
            logger.error(f"  ❌ Erro ao converter para xlsxwriter: {e}")
            return False
    
    def determinar_tipo_caminhao(self, peso_total_kg: float) -> str:
        """
        Determina o tipo de caminhão baseado no peso total
        
        Args:
            peso_total_kg: Peso total em KG
            
        Returns:
            Tipo de caminhão adequado
        """
        for tipo_caminhao, peso_maximo in self.CAMINHOES:
            if peso_total_kg <= peso_maximo:
                return tipo_caminhao
        return 'Caminhão (4 eixos) 31T'  # Default para pesos muito grandes
    
    def buscar_dados_cnpj(self, cnpj: str) -> Dict[str, Any]:
        """
        Busca dados EXATAMENTE como programacao_em_lote faz:
        1. CarteiraPrincipal (pedidos com pedido_cliente)
        2. Separacao (sincronizado_nf=False)
        3. NFs no CD (sincronizado_nf=True e nf_cd=True)
        
        Args:
            cnpj: CNPJ para buscar dados
            
        Returns:
            Dicionário com dados do CNPJ
        """
        from flask import has_app_context
        
        # Se não tem contexto e precisa dele, criar temporariamente
        if not has_app_context() and hasattr(self.app, 'app_context'):
            with self.app.app_context():
                return self._buscar_dados_cnpj_impl(cnpj)
        else:
            return self._buscar_dados_cnpj_impl(cnpj)
    
    def _executar_query_com_reconexao(self, query, erro_msg="Erro ao executar query"):
        """
        Executa uma query com tratamento de reconexão automática
        
        Args:
            query: Query SQLAlchemy para executar
            erro_msg: Mensagem de erro personalizada
            
        Returns:
            Resultado da query
        """
        try:
            return query.all()
        except Exception as e:
            if "SSL connection" in str(e) or "OperationalError" in str(e.__class__.__name__):
                logger.warning(f"  ⚠️ Conexão SSL perdida, reconectando...")
                try:
                    # Limpar sessão e reconectar
                    db.session.rollback()
                    db.session.close()
                    db.session.remove()
                    
                    # Tentar novamente
                    resultado = query.all()
                    logger.info("  ✅ Reconexão bem-sucedida")
                    return resultado
                except Exception as e2:
                    logger.error(f"  ❌ {erro_msg}: {e2}")
                    raise
            else:
                logger.error(f"  ❌ {erro_msg}: {e}")
                raise
    
    def _buscar_dados_cnpj_impl(self, cnpj: str) -> Dict[str, Any]:
        """Implementação real da busca de dados"""
        logger.info(f"📊 Buscando dados para CNPJ: {cnpj}")
        
        dados = {
            'cnpj': cnpj,
            'itens': [],
            'peso_total': Decimal('0'),
            'pallets_total': Decimal('0'),
            'valor_total': Decimal('0')
        }
        
        # 1. BUSCAR EM CARTEIRAPRINCIPAL (onde pedido_cliente está preenchido!)
        logger.info("  📦 Buscando em CarteiraPrincipal...")
        
        # Buscar itens agrupados por pedido com reconexão automática
        query_pedidos = db.session.query(
            CarteiraPrincipal.num_pedido,
            CarteiraPrincipal.pedido_cliente,
            func.sum(CarteiraPrincipal.qtd_saldo_produto_pedido * CarteiraPrincipal.preco_produto_pedido).label('valor_total')
        ).filter(
            and_(
                CarteiraPrincipal.cnpj_cpf == cnpj,
                CarteiraPrincipal.ativo == True
            )
        ).group_by(
            CarteiraPrincipal.num_pedido,
            CarteiraPrincipal.pedido_cliente
        )
        
        pedidos = self._executar_query_com_reconexao(query_pedidos, "Erro ao buscar pedidos da CarteiraPrincipal")
        
        logger.info(f"    Encontrados {len(pedidos)} pedidos")
        
        # Para cada pedido, buscar os itens detalhados
        for pedido in pedidos:
            itens_pedido = db.session.query(
                CarteiraPrincipal.num_pedido,
                CarteiraPrincipal.cod_produto,
                CarteiraPrincipal.nome_produto,
                CarteiraPrincipal.qtd_saldo_produto_pedido,
                CarteiraPrincipal.preco_produto_pedido,
                CarteiraPrincipal.pedido_cliente,
                CarteiraPrincipal.expedicao,
                CarteiraPrincipal.agendamento,
                CarteiraPrincipal.protocolo,
                CarteiraPrincipal.observ_ped_1
            ).filter(
                and_(
                    CarteiraPrincipal.cnpj_cpf == cnpj,
                    CarteiraPrincipal.num_pedido == pedido.num_pedido,
                    CarteiraPrincipal.ativo == True
                )
            ).all()
            
            for item in itens_pedido:
                # Buscar dados de palletização
                pallet_info = db.session.query(
                    CadastroPalletizacao.peso_bruto,
                    CadastroPalletizacao.palletizacao
                ).filter(
                    CadastroPalletizacao.cod_produto == item.cod_produto
                ).first()
                
                peso_item = Decimal('0')
                pallets_item = Decimal('0')
                
                if pallet_info:
                    qtd_decimal = Decimal(str(item.qtd_saldo_produto_pedido))
                    peso_item = qtd_decimal * Decimal(str(pallet_info.peso_bruto or 0))
                    if pallet_info.palletizacao and pallet_info.palletizacao > 0:
                        pallets_item = qtd_decimal / Decimal(str(pallet_info.palletizacao))
                
                dados['itens'].append({
                    'num_pedido': item.num_pedido,
                    'pedido_cliente': item.pedido_cliente,  # CAMPO IMPORTANTE!
                    'cod_produto': item.cod_produto,
                    'nome_produto': item.nome_produto,
                    'quantidade': float(item.qtd_saldo_produto_pedido),
                    'peso': float(peso_item),
                    'pallets': float(pallets_item),
                    'valor': float(Decimal(str(item.qtd_saldo_produto_pedido)) * Decimal(str(item.preco_produto_pedido or 0))),
                    'expedicao': item.expedicao,
                    'agendamento': item.agendamento,
                    'protocolo': item.protocolo,
                    'observacoes': item.observ_ped_1,
                    'origem': 'carteira'
                })
                
                dados['peso_total'] += peso_item
                dados['pallets_total'] += pallets_item
                dados['valor_total'] += Decimal(str(item.qtd_saldo_produto_pedido)) * Decimal(str(item.preco_produto_pedido or 0))
        
        # 2. BUSCAR SEPARAÇÕES (sincronizado_nf=False)
        logger.info("  📋 Buscando em Separacao...")
        
        separacoes = db.session.query(
            Separacao.num_pedido,
            Separacao.cod_produto,
            Separacao.qtd_saldo,
            Separacao.peso,
            Separacao.pallet,
            Separacao.valor_saldo,
            Separacao.expedicao,
            Separacao.agendamento,
            Separacao.protocolo,
            Separacao.observ_ped_1,
            Separacao.pedido_cliente
        ).filter(
            and_(
                Separacao.cnpj_cpf == cnpj,
                Separacao.sincronizado_nf == False
            )
        ).all()
        
        logger.info(f"    Encontradas {len(separacoes)} separações")
        
        # Criar set de chaves já adicionadas da carteira para evitar duplicatas
        chaves_carteira = set()
        for item in dados['itens']:
            if item['origem'] == 'carteira':
                chaves_carteira.add(f"{item['num_pedido']}_{item['cod_produto']}")
        
        for sep in separacoes:
            # Pular se já foi adicionado da carteira
            chave = f"{sep.num_pedido}_{sep.cod_produto}"
            if chave in chaves_carteira:
                continue
            
            # Buscar nome do produto
            nome_produto = None
            pallet_cadastro = db.session.query(CadastroPalletizacao.nome_produto).filter(
                CadastroPalletizacao.cod_produto == sep.cod_produto
            ).first()
            if pallet_cadastro:
                nome_produto = pallet_cadastro.nome_produto
            
            dados['itens'].append({
                'num_pedido': sep.num_pedido,
                'pedido_cliente': sep.pedido_cliente,  # Pode estar NULL mas vamos incluir
                'cod_produto': sep.cod_produto,
                'nome_produto': nome_produto or f"Produto {sep.cod_produto}",
                'quantidade': float(sep.qtd_saldo or 0),
                'peso': float(sep.peso or 0),
                'pallets': float(sep.pallet or 0),
                'valor': float(sep.valor_saldo or 0),
                'expedicao': sep.expedicao,
                'agendamento': sep.agendamento,
                'protocolo': sep.protocolo,
                'observacoes': sep.observ_ped_1,
                'origem': 'separacao'
            })
            
            dados['peso_total'] += Decimal(str(sep.peso or 0))
            dados['pallets_total'] += Decimal(str(sep.pallet or 0))
            dados['valor_total'] += Decimal(str(sep.valor_saldo or 0))
        
        # 3. BUSCAR NFs NO CD (sincronizado_nf=True e nf_cd=True)
        logger.info("  📄 Buscando NFs no CD...")
        
        nfs_cd = db.session.query(
            Separacao.num_pedido,
            Separacao.numero_nf,
            Separacao.cod_produto,
            Separacao.qtd_saldo,
            Separacao.peso,
            Separacao.pallet,
            Separacao.valor_saldo,
            Separacao.expedicao,
            Separacao.agendamento,
            Separacao.protocolo,
            Separacao.observ_ped_1,
            Separacao.pedido_cliente
        ).filter(
            and_(
                Separacao.cnpj_cpf == cnpj,
                Separacao.sincronizado_nf == True,
                Separacao.nf_cd == True
            )
        ).all()
        
        logger.info(f"    Encontradas {len(nfs_cd)} NFs no CD")
        
        for nf in nfs_cd:
            # Buscar nome do produto
            nome_produto = None
            pallet_cadastro = db.session.query(CadastroPalletizacao.nome_produto).filter(
                CadastroPalletizacao.cod_produto == nf.cod_produto
            ).first()
            if pallet_cadastro:
                nome_produto = pallet_cadastro.nome_produto
            
            dados['itens'].append({
                'num_pedido': nf.num_pedido,
                'pedido_cliente': nf.pedido_cliente,  # Pode estar NULL mas vamos incluir
                'numero_nf': nf.numero_nf,
                'cod_produto': nf.cod_produto,
                'nome_produto': nome_produto or f"Produto {nf.cod_produto}",
                'quantidade': float(nf.qtd_saldo or 0),
                'peso': float(nf.peso or 0),
                'pallets': float(nf.pallet or 0),
                'valor': float(nf.valor_saldo or 0),
                'expedicao': nf.expedicao,
                'agendamento': nf.agendamento,
                'protocolo': nf.protocolo,
                'observacoes': nf.observ_ped_1,
                'origem': 'nf_cd'
            })
            
            dados['peso_total'] += Decimal(str(nf.peso or 0))
            dados['pallets_total'] += Decimal(str(nf.pallet or 0))
            dados['valor_total'] += Decimal(str(nf.valor_saldo or 0))
        
        logger.info(f"  ✅ Total: {len(dados['itens'])} itens, Peso: {dados['peso_total']:.2f} kg")
        
        return dados
    
    def preencher_planilha(self, arquivo_origem: str, cnpj: str, 
                          data_agendamento: date = None,
                          arquivo_destino: str = None) -> str:
        """
        Preenche a planilha de agendamento com dados do CNPJ
        MANTÉM os dados existentes e PREENCHE apenas campos específicos vazios
        SEGUE ESTRITAMENTE mapeamento_planilha.md
        
        Args:
            arquivo_origem: Caminho da planilha original baixada
            cnpj: CNPJ para buscar dados
            data_agendamento: Data de agendamento (se None, usa hoje + 2 dias)
            arquivo_destino: Caminho para salvar (se None, gera automaticamente)
            
        Returns:
            Caminho do arquivo preenchido
        """
        logger.info("=" * 60)
        logger.info(f"📝 PREENCHENDO PLANILHA PARA CNPJ: {cnpj}")
        logger.info("=" * 60)
        
        # Validar que data de agendamento foi fornecida (OBRIGATÓRIA)
        if not data_agendamento:
            logger.error(f"❌ Data de agendamento é OBRIGATÓRIA para CNPJ {cnpj}")
            return {
                'sucesso': False,
                'mensagem': 'Data de agendamento não fornecida',
                'arquivo': None
            }
        
        # Buscar dados do CNPJ (EXATAMENTE como programacao_em_lote)
        dados_cnpj = self.buscar_dados_cnpj(cnpj)
        
        if not dados_cnpj['itens']:
            logger.warning(f"⚠️ Nenhum item encontrado para o CNPJ {cnpj}")
            return ''
        
        # Criar índice de dados por pedido_cliente e código produto para busca rápida
        # IMPORTANTE: Vamos traduzir NOSSOS códigos para códigos SENDAS
        indice_dados = {}
        for item in dados_cnpj['itens']:
            # Extrair pedido_cliente base (sem filial)
            pedido_cliente = item.get('pedido_cliente')
            if pedido_cliente:
                # Converter para string e remover parte após "-" se existir
                pedido_cliente = str(pedido_cliente)
                if '-' in pedido_cliente:
                    pedido_cliente = pedido_cliente.split('-')[0]
                
                # Traduzir NOSSO código para código SENDAS
                codigo_nosso = item['cod_produto']
                codigo_sendas = ProdutoDeParaSendas.obter_codigo_sendas(codigo_nosso, cnpj)
                
                if not codigo_sendas:
                    # Se não tem DE-PARA, usar nosso código mesmo
                    codigo_sendas = codigo_nosso
                    logger.debug(f"  ⚠️ Sem DE-PARA para produto {codigo_nosso}, usando direto")
                
                # Criar índice com código SENDAS
                chave = f"{pedido_cliente}_{codigo_sendas}"
                indice_dados[chave] = item
                logger.debug(f"  📌 Índice criado: {chave} (nosso: {codigo_nosso} → sendas: {codigo_sendas})")
        
        logger.info(f"  📊 Índice criado com {len(indice_dados)} chaves de busca usando pedido_cliente")
        
        # Carregar planilha
        logger.info(f"📂 Carregando planilha: {arquivo_origem}")
        wb = load_workbook(arquivo_origem)
        ws = wb.active
        
        # Identificar linha inicial (conforme mapeamento)
        linha_inicial = 4  # Dados começam na linha 4
        
        # Detectar última linha com dados (máximo 2000 linhas)
        logger.info("  🔍 Detectando última linha com dados...")
        ultima_linha_com_dados = linha_inicial - 1
        max_linha_verificar = min(2000, ws.max_row)
        
        for row in range(linha_inicial, max_linha_verificar + 1):
            # Verificar se há algum dado nas colunas principais
            if (ws.cell(row=row, column=7).value or    # Coluna G - Código pedido
            ws.cell(row=row, column=8).value or    # Coluna H - Código produto
            ws.cell(row=row, column=15).value):    # Coluna O - Saldo disponível
                ultima_linha_com_dados = row
            # Se encontrou 10 linhas vazias consecutivas, parar
            elif row > ultima_linha_com_dados + 10:
                break
        
        logger.info(f"  📊 Última linha com dados: {ultima_linha_com_dados}")
        
        # Observação única para identificar agendamento
        observacao_unica = f"AGEND_{cnpj[-4:]}_{data_agendamento.strftime('%Y%m%d')}"
        
        # Calcular tipo de caminhão baseado no peso total
        tipo_caminhao = self.determinar_tipo_caminhao(float(dados_cnpj['peso_total']))
        
        # Controle de demanda_id (mesmo ID para todo o CNPJ)
        demanda_id = 1
        
        # Processar linhas existentes (NÃO LIMPAR, APENAS PREENCHER CAMPOS VAZIOS)
        logger.info(f"  📝 Processando linhas {linha_inicial} até {ultima_linha_com_dados}...")
        
        linhas_preenchidas = 0
        linhas_nao_encontradas = 0
        linhas_debug = []
        
        for row in range(linha_inicial, ultima_linha_com_dados + 1):
            # EXTRAIR dados conforme mapeamento_planilha.md
            unidade_destino = ws.cell(row=row, column=4).value       # Coluna D - Nome COMPLETO da filial
            codigo_pedido_cliente = ws.cell(row=row, column=7).value  # Coluna G - Pedido Cliente
            codigo_produto_sendas = ws.cell(row=row, column=8).value  # Coluna H - Código Produto Cliente
            saldo_disponivel = ws.cell(row=row, column=15).value     # Coluna O - Saldo disponível
            
            if not codigo_pedido_cliente or not codigo_produto_sendas:
                continue  # Pular linhas sem dados essenciais
            
            # 1. TRADUZIR NOSSO CNPJ PARA CÓDIGO DA FILIAL
            filial_esperada = FilialDeParaSendas.cnpj_to_filial(cnpj)
            
            if not filial_esperada:
                logger.warning(f"    ⚠️ CNPJ {cnpj} não tem mapeamento de filial!")
                continue
            
            # Verificar se a unidade da planilha corresponde à nossa filial
            # Comparar de forma flexível (pode estar truncado)
            if unidade_destino:
                unidade_str = str(unidade_destino).upper().strip()
                filial_str = filial_esperada.upper().strip()
                
                # Verificar se é a mesma filial (match exato ou início)
                if not (unidade_str == filial_str or unidade_str.startswith(filial_str[:20])):
                    # Esta linha é de outra filial, pular
                    logger.debug(f"    ⏭️ Linha {row}: Filial diferente ('{unidade_destino}' != '{filial_esperada}')")
                    continue
            else:
                # Sem unidade de destino, não podemos verificar
                logger.debug(f"    ⚠️ Linha {row}: Sem unidade de destino")
                continue
            
            # 2. Controle do demanda_id (todas as linhas do mesmo CNPJ têm o mesmo ID)
            # Como estamos filtrando por um único CNPJ, demanda_id permanece o mesmo
            
            # 3. EXTRAIR pedido_cliente (antes do "-")
            pedido_cliente = str(codigo_pedido_cliente)
            if '-' in pedido_cliente:
                pedido_cliente = pedido_cliente.split('-')[0]
            
            # 4. NÃO TRADUZIR! Usar código Sendas direto da planilha
            codigo_sendas = str(codigo_produto_sendas)
            
            # 5. BUSCAR item nos nossos dados usando código SENDAS
            chave_busca = f"{pedido_cliente}_{codigo_sendas}"
            item_encontrado = indice_dados.get(chave_busca)
            
            if item_encontrado:
                logger.debug(f"    ✓ Linha {row}: MATCH! Pedido {pedido_cliente} produto {codigo_nosso}")
                
                # PREENCHER APENAS CAMPOS VAZIOS ESPECÍFICOS (conforme mapeamento_planilha.md)
                
                # Coluna A - Demanda (se vazio)
                if not ws.cell(row=row, column=1).value:
                    ws.cell(row=row, column=1).value = demanda_id
                
                # Coluna Q - Quantidade entrega (se vazio)
                if not ws.cell(row=row, column=17).value:
                    # Respeitar saldo disponível (coluna O)
                    quantidade_nossa = item_encontrado['quantidade']
                    if saldo_disponivel:
                        quantidade = min(quantidade_nossa, float(saldo_disponivel))
                    else:
                        quantidade = quantidade_nossa
                    ws.cell(row=row, column=17).value = quantidade
                
                # Coluna R - Data sugerida de entrega (se vazio)
                if not ws.cell(row=row, column=18).value:
                    # Gravar como datetime real do Excel, não como string
                    celula_data = ws.cell(row=row, column=18)
                    celula_data.value = data_agendamento  # Python date/datetime é automaticamente convertido
                    celula_data.number_format = 'DD/MM/YYYY'  # Formato de exibição no Excel
                
                # Coluna U - Característica da carga (se vazio) - SEMPRE "Paletizada"
                if not ws.cell(row=row, column=21).value:
                    ws.cell(row=row, column=21).value = 'Paletizada'
                
                # Coluna V - Característica do veículo (se vazio)
                if not ws.cell(row=row, column=22).value:
                    ws.cell(row=row, column=22).value = tipo_caminhao
                
                # Coluna X - Observação/Fornecedor (se vazio)
                if not ws.cell(row=row, column=24).value:
                    ws.cell(row=row, column=24).value = observacao_unica
                
                linhas_preenchidas += 1
            else:
                logger.debug(f"    ❌ Linha {row}: NÃO encontrado pedido {pedido_cliente} produto {codigo_nosso}")
                linhas_nao_encontradas += 1
        
        # Destacar células preenchidas com cor suave
        fill_preenchido = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        for row in range(linha_inicial, ultima_linha_com_dados + 1):
            # Verificar se a linha foi preenchida (tem nossa observação)
            if ws.cell(row=row, column=24).value == observacao_unica:
                # Destacar apenas colunas que preenchemos
                for col in [1, 17, 18, 21, 22, 24]:  # Colunas A, Q, R, U, V, X
                    cell = ws.cell(row=row, column=col)
                    cell.fill = fill_preenchido
        
        # Salvar planilha
        if not arquivo_destino:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            arquivo_destino = os.path.join(
                os.path.dirname(arquivo_origem),
                f"sendas_{cnpj[-4:]}_{timestamp}_preenchido.xlsx"
            )
        
        # Salvar temporariamente com openpyxl
        temp_file = tempfile.mktemp(suffix='_temp.xlsx')
        wb.save(temp_file)
        
        # Converter para xlsxwriter (garante sharedStrings e datas corretas)
        logger.info("  🔄 Convertendo para formato compatível com Sendas...")
        sucesso = self._converter_para_xlsxwriter(temp_file, arquivo_destino)
        
        # Remover arquivo temporário
        if os.path.exists(temp_file):
            os.remove(temp_file)
        
        if not sucesso:
            logger.error("  ❌ Erro na conversão para xlsxwriter")
            return None
            
        logger.info(f"  💾 Planilha salva: {arquivo_destino}")
        
        # Resumo
        logger.info("\n" + "=" * 60)
        logger.info("📊 RESUMO DO PREENCHIMENTO:")
        logger.info("=" * 60)
        logger.info(f"  CNPJ: {cnpj}")
        logger.info(f"  Data Agendamento: {data_agendamento.strftime('%d/%m/%Y')}")
        logger.info(f"  Linhas processadas: {ultima_linha_com_dados - linha_inicial + 1}")
        logger.info(f"  Linhas preenchidas: {linhas_preenchidas}")
        logger.info(f"  Linhas não encontradas: {linhas_nao_encontradas}")
        logger.info(f"  Peso Total: {dados_cnpj['peso_total']:.2f} kg")
        logger.info(f"  Tipo Caminhão: {tipo_caminhao}")
        logger.info(f"  Observação: {observacao_unica}")
        
        if linhas_debug and logger.isEnabledFor(logging.DEBUG):
            logger.debug("\n📋 DETALHES DE DEBUG:")
            for msg in linhas_debug[:10]:  # Mostrar apenas primeiras 10
                logger.debug(f"  {msg}")
        
        logger.info("=" * 60)
        
        return arquivo_destino
    
    def preencher_multiplos_cnpjs(self, arquivo_origem, lista_cnpjs_agendamento, arquivo_destino=None):
        """
        Preenche planilha para MÚLTIPLOS CNPJs de uma vez e REMOVE linhas não agendadas
        
        Args:
            arquivo_origem: Caminho da planilha baixada do portal
            lista_cnpjs_agendamento: Lista de dict com {'cnpj': str, 'data_agendamento': date}
            arquivo_destino: Caminho opcional para salvar
        
        Returns:
            str: Caminho do arquivo salvo ou None se erro
        """
        logger.info("\n" + "=" * 80)
        logger.info("🔄 PROCESSAMENTO DE MÚLTIPLOS CNPJs - PORTAL SENDAS")
        logger.info("=" * 80)
        logger.info(f"📋 Total de CNPJs a processar: {len(lista_cnpjs_agendamento)}")
        
        if not os.path.exists(arquivo_origem):
            logger.error(f"❌ Arquivo não encontrado: {arquivo_origem}")
            return None
        
        # Carregar planilha
        wb = openpyxl.load_workbook(arquivo_origem)
        ws = wb.active
        
        # Encontrar linha inicial e final de dados
        linha_inicial = 4  # Dados começam na linha 4
        ultima_linha_com_dados = ws.max_row
        
        # Coletar dados de TODOS os CNPJs
        todos_dados = {}
        peso_total_geral = Decimal('0')
        
        for idx, agendamento in enumerate(lista_cnpjs_agendamento, 1):
            cnpj = agendamento['cnpj']
            data_agendamento = agendamento['data_agendamento']
            
            logger.info(f"\n  [{idx}/{len(lista_cnpjs_agendamento)}] Coletando dados do CNPJ: {cnpj}")
            
            # Buscar dados do CNPJ
            dados_cnpj = self.buscar_dados_cnpj(cnpj)
            if not dados_cnpj or not dados_cnpj['itens']:
                logger.warning(f"    ⚠️ Sem dados para CNPJ {cnpj}")
                continue
            
            todos_dados[cnpj] = {
                'data_agendamento': data_agendamento,
                'dados': dados_cnpj,
                'filial': FilialDeParaSendas.cnpj_to_filial(cnpj)
            }
            peso_total_geral += dados_cnpj['peso_total']
        
        if not todos_dados:
            logger.error("❌ Nenhum dado encontrado para processar")
            return None
        
        # Determinar tipo de caminhão baseado no peso total
        tipo_caminhao = self.determinar_tipo_caminhao(float(peso_total_geral))
        
        # Criar observação única para identificar este agendamento
        timestamp_obs = datetime.now().strftime('%Y%m%d_%H%M%S')
        observacao_unica = f"AG_MULTI_{timestamp_obs}"
        
        logger.info(f"\n📝 Preenchendo planilha com {len(todos_dados)} CNPJs...")
        
        # Rastrear linhas preenchidas
        linhas_preenchidas = set()
        demanda_id = 1
        
        # PASSO 1: Preencher dados de todos os CNPJs
        for cnpj, info in todos_dados.items():
            dados_cnpj = info['dados']
            data_agendamento = info['data_agendamento']
            filial_esperada = info['filial']
            
            if not filial_esperada:
                logger.warning(f"  ⚠️ CNPJ {cnpj} sem mapeamento de filial")
                continue
            
            logger.info(f"  📌 Processando CNPJ {cnpj} - Filial: {filial_esperada}")
            
            # Criar índice para busca rápida
            indice_dados = {}
            for item in dados_cnpj['itens']:
                pedido_cliente = str(item.get('pedido_cliente', ''))
                if pedido_cliente and pedido_cliente != 'None':
                    # Mapear código nosso para código Sendas
                    codigo_nosso = str(item['cod_produto'])
                    codigo_sendas = ProdutoDeParaSendas.obter_codigo_sendas(codigo_nosso)
                    if codigo_sendas:
                        chave = f"{pedido_cliente}_{codigo_sendas}"
                        indice_dados[chave] = item
            
            # Processar linhas da planilha
            for row in range(linha_inicial, ultima_linha_com_dados + 1):
                unidade_destino = ws.cell(row=row, column=4).value
                codigo_pedido_cliente = ws.cell(row=row, column=7).value
                codigo_produto_sendas = ws.cell(row=row, column=8).value
                saldo_disponivel = ws.cell(row=row, column=15).value
                
                if not codigo_pedido_cliente or not codigo_produto_sendas:
                    continue
                
                # Verificar se é a filial correta
                if unidade_destino:
                    unidade_str = str(unidade_destino).upper().strip()
                    filial_str = filial_esperada.upper().strip()
                    
                    if not (unidade_str == filial_str or unidade_str.startswith(filial_str[:20])):
                        continue  # Filial diferente, pular
                
                # Extrair pedido_cliente
                pedido_cliente = str(codigo_pedido_cliente)
                if '-' in pedido_cliente:
                    pedido_cliente = pedido_cliente.split('-')[0]
                
                # Buscar item
                codigo_sendas = str(codigo_produto_sendas)
                chave_busca = f"{pedido_cliente}_{codigo_sendas}"
                item_encontrado = indice_dados.get(chave_busca)
                
                if item_encontrado:
                    # PREENCHER CAMPOS
                    ws.cell(row=row, column=1).value = demanda_id  # Coluna A - Demanda
                    
                    # Quantidade respeitando saldo
                    quantidade = item_encontrado['quantidade']
                    if saldo_disponivel:
                        quantidade = min(quantidade, float(saldo_disponivel))
                    ws.cell(row=row, column=17).value = quantidade  # Coluna Q
                    
                    # Data de agendamento como objeto date nativo
                    celula_data = ws.cell(row=row, column=18)  # Coluna R
                    celula_data.value = data_agendamento
                    celula_data.number_format = 'DD/MM/YYYY'
                    
                    ws.cell(row=row, column=21).value = 'Paletizada'  # Coluna U
                    ws.cell(row=row, column=22).value = tipo_caminhao  # Coluna V
                    ws.cell(row=row, column=24).value = observacao_unica  # Coluna X
                    
                    linhas_preenchidas.add(row)
            
            demanda_id += 1  # Incrementar demanda_id para próximo CNPJ
        
        logger.info(f"  ✅ {len(linhas_preenchidas)} linhas preenchidas no total")
        
        # PASSO 2: REMOVER LINHAS NÃO PREENCHIDAS
        logger.info("\n🗑️ Removendo linhas não agendadas...")
        
        # Criar nova planilha apenas com linhas preenchidas
        wb_novo = openpyxl.Workbook()
        ws_novo = wb_novo.active
        
        # Copiar cabeçalhos (linhas 1-3)
        for row in range(1, linha_inicial):
            for col in range(1, ws.max_column + 1):
                cell_original = ws.cell(row=row, column=col)
                cell_novo = ws_novo.cell(row=row, column=col)
                cell_novo.value = cell_original.value
                # Copiar formato se existir
                if cell_original.has_style:
                    cell_novo.font = copy(cell_original.font)
                    cell_novo.fill = copy(cell_original.fill)
                    cell_novo.border = copy(cell_original.border)
                    cell_novo.alignment = copy(cell_original.alignment)
                    cell_novo.number_format = cell_original.number_format
        
        # Copiar apenas linhas preenchidas
        nova_linha = linha_inicial
        linhas_removidas = 0
        
        for row in range(linha_inicial, ultima_linha_com_dados + 1):
            if row in linhas_preenchidas:
                # Copiar linha preenchida
                for col in range(1, ws.max_column + 1):
                    cell_original = ws.cell(row=row, column=col)
                    cell_novo = ws_novo.cell(row=nova_linha, column=col)
                    cell_novo.value = cell_original.value
                    # Copiar formato
                    if cell_original.has_style:
                        cell_novo.font = copy(cell_original.font)
                        cell_novo.fill = copy(cell_original.fill)
                        cell_novo.border = copy(cell_original.border)
                        cell_novo.alignment = copy(cell_original.alignment)
                        cell_novo.number_format = cell_original.number_format
                nova_linha += 1
            else:
                linhas_removidas += 1
        
        logger.info(f"  🗑️ {linhas_removidas} linhas removidas (não agendadas)")
        logger.info(f"  ✅ {len(linhas_preenchidas)} linhas mantidas (agendadas)")
        
        # Salvar nova planilha
        if not arquivo_destino:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            arquivo_destino = os.path.join(
                tempfile.gettempdir(),
                f"sendas_multi_{timestamp}.xlsx"
            )
        
        # Salvar temporariamente com openpyxl
        temp_file = tempfile.mktemp(suffix='_temp.xlsx')
        wb_novo.save(temp_file)
        
        # Converter para xlsxwriter (garante sharedStrings e datas corretas)
        logger.info("  🔄 Convertendo para formato compatível com Sendas...")
        sucesso = self._converter_para_xlsxwriter(temp_file, arquivo_destino)
        
        # Remover arquivo temporário
        if os.path.exists(temp_file):
            os.remove(temp_file)
        
        if not sucesso:
            logger.error("  ❌ Erro na conversão para xlsxwriter")
            return None
            
        logger.info(f"\n💾 Planilha salva: {arquivo_destino}")
        
        # Resumo final
        logger.info("\n" + "=" * 80)
        logger.info("📊 RESUMO DO PROCESSAMENTO MÚLTIPLO:")
        logger.info("=" * 80)
        logger.info(f"  CNPJs processados: {len(todos_dados)}")
        logger.info(f"  Linhas preenchidas: {len(linhas_preenchidas)}")
        logger.info(f"  Linhas removidas: {linhas_removidas}")
        logger.info(f"  Peso total geral: {peso_total_geral:.2f} kg")
        logger.info(f"  Tipo caminhão: {tipo_caminhao}")
        logger.info(f"  Observação: {observacao_unica}")
        logger.info("=" * 80)
        
        return arquivo_destino


def main():
    """Função principal para teste"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Preencher planilha de agendamento Sendas')
    parser.add_argument('arquivo', help='Caminho da planilha baixada')
    parser.add_argument('cnpj', help='CNPJ para buscar dados')
    parser.add_argument('--data', help='Data de agendamento (DD/MM/YYYY)', default=None)
    parser.add_argument('--output', help='Arquivo de saída', default=None)
    
    args = parser.parse_args()
    
    # Parse da data se fornecida
    data_agendamento = None
    if args.data:
        try:
            data_agendamento = datetime.strptime(args.data, '%d/%m/%Y').date()
        except ValueError:
            logger.error("❌ Formato de data inválido. Use DD/MM/YYYY")
            return
    
    # Criar preenchedor
    preenchedor = PreencherPlanilhaSendas()
    
    # Preencher planilha
    try:
        arquivo_preenchido = preenchedor.preencher_planilha(
            arquivo_origem=args.arquivo,
            cnpj=args.cnpj,
            data_agendamento=data_agendamento,
            arquivo_destino=args.output
        )
        
        if arquivo_preenchido:
            logger.info(f"\n✅ Planilha preenchida com sucesso: {arquivo_preenchido}")
        else:
            logger.error("\n❌ Falha ao preencher planilha")
            
    except Exception as e:
        logger.error(f"\n❌ Erro ao preencher planilha: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()