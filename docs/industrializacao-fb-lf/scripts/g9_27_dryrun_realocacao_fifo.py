#!/usr/bin/env python3
"""G9 DRY-RUN realocacao FIFO (READ-ONLY) — adiantamentos FB contra faturas ENTSI em aberto.

Simula casar cada adiantamento (sobra) contra as faturas a-pagar EM ABERTO mais antigas (FIFO),
SEMPRE min(adiantamento, residual) -> nunca pgto > fatura. Marca os que sao DUPLICATA provavel
(X ~= soma faturas-alvo do ref): esses NAO devem ser realocados (estornar). NAO escreve no Odoo.
"""
import sys
import re
from collections import defaultdict
sys.path.insert(0, '/home/rafaelnascimento/projetos/frete_sistema')
from app.odoo.utils.connection import get_odoo_connection
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

CTX = {'allowed_company_ids': [1, 5]}
OUT = '/home/rafaelnascimento/projetos/frete_sistema/docs/industrializacao-fb-lf/G9_DryRun_Realocacao_FIFO.xlsx'
ACC_FB, P_LF = 11038, 35
RE_FAT = re.compile(r'[A-Z]{3,6}/\d{4}/\d{2}/\d{4}')
HDR = PatternFill('solid', fgColor='1F4E78'); HF = Font(bold=True, color='FFFFFF')
TIT = Font(bold=True, size=13, color='1F4E78'); B = Font(bold=True); NUM = '#,##0.00'
BAD = PatternFill('solid', fgColor='FCE4D6'); WARN = PatternFill('solid', fgColor='FFF2CC'); OKF = PatternFill('solid', fgColor='E2EFDA')


def main():
    o = get_odoo_connection()
    assert o.authenticate(), "FALHA AUTH"

    base_fb = [('account_id', '=', ACC_FB), ('partner_id', '=', P_LF), ('parent_state', '=', 'posted')]
    adi = o.execute_kw('account.move.line', 'search_read',
                       [base_fb + [('debit', '>', 0), ('reconciled', '=', False), ('amount_residual', '>', 0)]],
                       {'fields': ['date', 'amount_residual', 'ref', 'move_id'], 'order': 'date', 'context': CTX})

    # S = soma das faturas-alvo (p/ flag de duplicata)
    alvo_names = set()
    for l in adi:
        alvo_names |= set(RE_FAT.findall(l.get('ref') or ''))
    fm = o.execute_kw('account.move', 'search_read', [[('name', 'in', list(alvo_names)), ('company_id', '=', 1)]], {'fields': ['name'], 'context': CTX})
    mid2name = {m['id']: m['name'] for m in fm}
    flines = o.execute_kw('account.move.line', 'search_read', [[('move_id', 'in', list(mid2name)), ('account_id', '=', ACC_FB), ('credit', '>', 0)]], {'fields': ['move_id', 'credit'], 'context': CTX})
    fcred = defaultdict(float)
    for l in flines:
        fcred[mid2name[l['move_id'][0]]] += l['credit'] or 0

    # faturas a-pagar EM ABERTO (FIFO por data) — pilha de residual disponivel
    abertas = o.execute_kw('account.move.line', 'search_read',
                           [base_fb + [('credit', '>', 0), ('reconciled', '=', False), ('amount_residual', '<', 0)]],
                           {'fields': ['move_id', 'amount_residual', 'date'], 'order': 'date, id', 'context': CTX})
    pilha = [{'nome': a['move_id'][1], 'date': a['date'], 'resid': -(a['amount_residual'] or 0)} for a in abertas]

    rows = []
    tot_realoc = tot_dup = 0.0
    for l in adi:
        X = l['amount_residual']
        S = sum(fcred.get(a, 0) for a in RE_FAT.findall(l.get('ref') or ''))
        dup = S > 0 and abs(S - X) <= max(1.0, 0.01 * X)
        # FIFO contra a pilha (sem consumir se for duplicata — so simula o potencial)
        aplica = 0.0; det = []
        if not dup:
            for f in pilha:
                if f['resid'] <= 0.005:
                    continue
                usar = min(X - aplica, f['resid'])
                if usar > 0.005:
                    aplica += usar; f['resid'] -= usar
                    det.append(f"{f['nome']} ({f['date']}): R$ {usar:,.2f}")
                if aplica >= X - 0.005:
                    break
        alerta = ('DUPLICATA provável — ESTORNAR (não realocar)' if dup else
                  'realocável FIFO' if aplica >= X - 0.01 else 'realocável parcial' if aplica > 0.01 else 'sem fatura aberta')
        if dup:
            tot_dup += X
        else:
            tot_realoc += aplica
        rows.append({'date': l['date'], 'pag': l['move_id'][1], 'X': X, 'S': S, 'aplica': aplica,
                     'alerta': alerta, 'det': '; '.join(det[:4]) + (f'  (+{len(det)-4})' if len(det) > 4 else '')})

    # ===== Excel =====
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'DryRun Realocacao FIFO'
    for col, w in zip('ABCDEFG', [12, 22, 15, 15, 15, 34, 64]):
        ws.column_dimensions[col].width = w
    ws['A1'] = 'DRY-RUN — realocação FIFO dos adiantamentos contra faturas ENTSI EM ABERTO'; ws['A1'].font = TIT
    ws['A2'] = (f'realocável (não-duplicata): R$ {tot_realoc:,.2f} | duplicata provável (estornar, NÃO realocar): R$ {tot_dup:,.2f} | '
                f'SIMULAÇÃO — nada efetivado. Regra min(): nunca pgto > residual da fatura.')
    ws['A2'].font = Font(italic=True, size=9, color='808080')
    hr = 4
    for c, h in enumerate(['Data', 'Adiantamento', 'Valor (X)', 'Soma alvo (S)', 'Realocaria', 'Recomendação', 'Faturas em aberto que cobriria (FIFO)'], 1):
        cell = ws.cell(row=hr, column=c, value=h); cell.fill = HDR; cell.font = HF; cell.alignment = Alignment(horizontal='center', wrap_text=True)
    r = hr + 1
    for row in rows:
        for c, v in enumerate([row['date'], row['pag'], row['X'], row['S'], row['aplica'], row['alerta'], row['det']], 1):
            cell = ws.cell(row=r, column=c, value=v)
            if c in (3, 4, 5):
                cell.number_format = NUM
        ws.cell(row=r, column=6).fill = BAD if 'DUPLICATA' in row['alerta'] else OKF if row['alerta'] == 'realocável FIFO' else WARN
        r += 1
    ws.freeze_panes = 'A5'
    wb.save(OUT)

    print(f"Excel: {OUT}")
    print(f"  realocável (não-duplicata): R$ {tot_realoc:,.2f} | duplicata provável (estornar): R$ {tot_dup:,.2f}")
    cnt = defaultdict(lambda: [0, 0.0])
    for row in rows:
        cnt[row['alerta']][0] += 1; cnt[row['alerta']][1] += row['X']
    for k, c in sorted(cnt.items()):
        print(f"    {k:46s}: {c[0]:3d} | R$ {c[1]:,.2f}")


if __name__ == '__main__':
    main()
