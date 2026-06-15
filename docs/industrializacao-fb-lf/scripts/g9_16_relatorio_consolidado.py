#!/usr/bin/env python3
"""G9 RELATORIO CONSOLIDADO (READ-ONLY) — resumo executivo da regularizacao p/ a Contadora."""
import sys
sys.path.insert(0, '/home/rafaelnascimento/projetos/frete_sistema')
from app.odoo.utils.connection import get_odoo_connection
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

CTX = {'allowed_company_ids': [1, 5]}
OUT = '/home/rafaelnascimento/projetos/frete_sistema/docs/industrializacao-fb-lf/G9_Consolidado_Regularizacao.xlsx'
# snapshots iniciais (medidos no inicio da sessao, antes da regularizacao)
ATIVA_INI, PASSIVA_INI = 61930965.26, -37749509.88


def main():
    o = get_odoo_connection()
    assert o.authenticate(), "FALHA AUTH"

    def rg(m, d, f, g=[]):
        return o.execute_kw(m, 'read_group', [d, f, g], {'context': CTX, 'lazy': False})

    def saldo(acc):
        g = rg('account.move.line', [('account_id', '=', acc), ('parent_state', '=', 'posted')], ['debit:sum', 'credit:sum'])
        return (g[0]['debit'] or 0) - (g[0]['credit'] or 0)

    ativa = saldo(22800)
    passiva = saldo(26667)
    # baixado via G9
    gA = rg('account.move.line', [('account_id', '=', 22800), ('journal_id', '=', 893), ('parent_state', '=', 'posted')], ['credit:sum'])
    baixado = gA[0]['credit'] or 0
    nfs = gA[0]['__count']
    # por mes (entries G9 na ATIVA FB)
    gm = rg('account.move.line', [('account_id', '=', 22800), ('journal_id', '=', 893), ('parent_state', '=', 'posted')], ['credit:sum'], ['date:month'])
    # conta corrente
    def cc(sign):
        op = '>' if sign > 0 else '<'
        g = rg('account.move.line', [('account_id', '=', 26085), ('partner_id', '=', 1), ('parent_state', '=', 'posted'), ('reconciled', '=', False), ('amount_residual', op, 0)], ['amount_residual:sum'])
        return g[0].get('amount_residual', 0) or 0, g[0]['__count']
    areceber, n_ar = cc(1)
    creditos, n_cr = cc(-1)

    wb = openpyxl.Workbook()
    HDR = PatternFill('solid', fgColor='1F4E78'); HF = Font(bold=True, color='FFFFFF')
    TIT = Font(bold=True, size=13, color='1F4E78'); B = Font(bold=True); NUM = '#,##0.00'

    ws = wb.active; ws.title = 'Resumo'
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 50; ws.column_dimensions['C'].width = 20; ws.column_dimensions['D'].width = 20
    ws['B1'] = 'Regularização industrialização FB↔LF — consolidado (01/2025 a hoje)'; ws['B1'].font = TIT
    rows = [
        ('', '', ''),
        ('CONTAS DE CONTROLE', 'Antes', 'Depois'),
        ('5101010001 REMESSA IND. (ATIVA) — FB', ATIVA_INI, ativa),
        ('5101020001 REMESSA IND. (PASSIVA) — LF', PASSIVA_INI, passiva),
        ('', '', ''),
        ('Total baixado (insumos regularizados)', baixado, ''),
        ('NFs regularizadas (01/2025→hoje)', nfs, ''),
        ('(saldo remanescente da ATIVA = exercício 2024, fora do escopo)', '', ''),
        ('', '', ''),
        ('CONTA CORRENTE FB na LF (após FIFO)', '', ''),
        ('A receber (serviço genuíno)', areceber, f'{n_ar} títulos'),
        ('Créditos em aberto (excedentes)', -creditos, f'{n_cr} linhas'),
        ('', '', ''),
        ('MÉTODO', '', ''),
        ('FASE 1 — ajuste: D PASSIVA/C Clientes (LF) + D CPV/C ATIVA (FB)', '', ''),
        ('FASE 2 — FIFO: créditos/excedentes compensam a-receber mais antigos', '', ''),
        ('Excedente = compensação na conta corrente (nunca devolução)', '', ''),
    ]
    r = 3
    for a, b, c in rows:
        ws.cell(row=r, column=2, value=a)
        if a.isupper() or a.startswith('Total') or a.startswith('A receber'):
            ws.cell(row=r, column=2).font = B
        if a in ('CONTAS DE CONTROLE',):
            for col in (2, 3, 4):
                ws.cell(row=r, column=col).fill = HDR; ws.cell(row=r, column=col).font = HF
            ws.cell(row=r, column=3, value=b); ws.cell(row=r, column=4, value=c)
        else:
            for col, v in ((3, b), (4, c)):
                if isinstance(v, (int, float)):
                    cell = ws.cell(row=r, column=col, value=v)
                    cell.number_format = NUM if isinstance(v, float) else '0'
                    cell.font = B
                elif v:
                    ws.cell(row=r, column=col, value=v)
        r += 1

    ws2 = wb.create_sheet('Por mês')
    for c, h in enumerate(['Mês', 'Insumos regularizados (R$)'], 1):
        cell = ws2.cell(row=1, column=c, value=h); cell.fill = HDR; cell.font = HF
    ws2.column_dimensions['A'].width = 16; ws2.column_dimensions['B'].width = 26
    r = 2
    for row in sorted(gm, key=lambda x: x.get('date:month') or ''):
        ws2.cell(row=r, column=1, value=row.get('date:month'))
        ws2.cell(row=r, column=2, value=row.get('credit', 0) or 0).number_format = NUM
        r += 1
    ws2.cell(row=r + 1, column=1, value='TOTAL').font = B
    cc_ = ws2.cell(row=r + 1, column=2, value=baixado); cc_.number_format = NUM; cc_.font = B

    wb.save(OUT)
    print(f"Excel: {OUT}")
    print(f"  ATIVA FB: {ATIVA_INI:,.2f} -> {ativa:,.2f} | PASSIVA LF: {PASSIVA_INI:,.2f} -> {passiva:,.2f}")
    print(f"  baixado {baixado:,.2f} em {nfs} NFs | a receber liquido {areceber:,.2f} ({n_ar}) | creditos abertos {-creditos:,.2f} ({n_cr})")


if __name__ == '__main__':
    main()
