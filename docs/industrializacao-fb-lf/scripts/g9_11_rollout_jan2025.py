#!/usr/bin/env python3
"""G9 ROLLOUT jan/2025 (reversao pura per-doc) — com DESCONCILIACAO (diretriz Contadora).

DRY-RUN default. So efetiva com --confirmar.
Fluxo por NF de retorno (j847, 5902) de jan/2025:
  1) se o recebivel esta conciliado -> DESCONCILIAR (remove_move_reconcile) -> recebivel reabre, pagamentos soltos
  2) entry LF: D 5101020001 PASSIVA / C <conta recebivel>   (NAO reconcilia — a Contadora reconcilia depois)
  3) entry FB: D 3201000001 CPV / C 5101010001 ATIVA
Idempotente por REF. Periodo jan/2025 ABERTO (lock 2024-12-31). Gera Excel.
NOTA: a reconciliacao final (pagamentos x recebivel ajustado + excedente) e' feita pela Contadora.
"""
import sys
sys.path.insert(0, '/home/rafaelnascimento/projetos/frete_sistema')
from app.odoo.utils.connection import get_odoo_connection
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

CTX_BOTH = {'allowed_company_ids': [1, 5]}
CTX_LF = {'allowed_company_ids': [5]}
CTX_FB = {'allowed_company_ids': [1]}
OPS_5902 = [2864, 2710]
ACC_PASSIVA_LF = 26667
ACC_CPV_FB = 22527
ACC_ATIVA_FB = 22800
J_DIV_LF = 894
J_DIV_FB = 893
D1, D2 = '2025-01-01', '2025-01-31'
OUT = '/home/rafaelnascimento/projetos/frete_sistema/docs/industrializacao-fb-lf/G9_NFs_corrigidas_jan2025.xlsx'
CONFIRMAR = '--confirmar' in sys.argv
LIMIT = None
for i, a in enumerate(sys.argv):
    if a == '--limit' and i + 1 < len(sys.argv):
        LIMIT = int(sys.argv[i + 1])


def ref_of(name):
    return f"G9-REGULARIZACAO IND. {name} (reversao pura per-doc)"


def main():
    o = get_odoo_connection()
    assert o.authenticate(), "FALHA AUTH"
    print(f"UID {o._uid} | MODO: {'CONFIRMAR (ESCREVE)' if CONFIRMAR else 'DRY-RUN'}\n")

    def rr(model, domain, fields, ctx=CTX_BOTH, **kw):
        kw2 = {'fields': fields, 'context': ctx}
        kw2.update(kw)
        return o.execute_kw(model, 'search_read', [domain], kw2)

    # lock guard
    comps = rr('res.company', [('id', 'in', [1, 5])], ['id', 'fiscalyear_lock_date'])
    lock = {c['id']: (c.get('fiscalyear_lock_date') or '') for c in comps}
    if lock.get(1, '') >= D2 or lock.get(5, '') >= D2:
        print(f"[ABORT] periodo FECHADO (lock FB {lock.get(1)} / LF {lock.get(5)}). jan/2025 nao esta aberto.")
        return
    print(f"Lock OK: FB={lock.get(1)} / LF={lock.get(5)} -> jan/2025 ABERTO\n")

    moves = rr('account.move',
               [('journal_id', '=', 847), ('company_id', '=', 5), ('state', '=', 'posted'),
                ('move_type', '=', 'out_invoice'), ('date', '>=', D1), ('date', '<=', D2)],
               ['id', 'name', 'date', 'payment_state'], order='id')
    print(f"NFs j847 jan/2025: {len(moves)}")

    plan = []
    for m in moves:
        l5902 = rr('account.move.line', [('move_id', '=', m['id']), ('l10n_br_operacao_id', 'in', OPS_5902)], ['credit'])
        valor = round(sum(l.get('credit') or 0 for l in l5902), 2)
        if valor <= 0:
            continue
        recv = rr('account.move.line',
                  [('move_id', '=', m['id']), ('account_id.account_type', '=', 'asset_receivable'), ('debit', '>', 0)],
                  ['id', 'account_id', 'partner_id', 'amount_residual', 'reconciled', 'matched_credit_ids'])
        if len(recv) != 1:
            plan.append({**m, 'valor': valor, 'status': f'RECEBIVEL inesperado ({len(recv)}) - PULA'})
            continue
        rv = recv[0]
        dup = rr('account.move', [('ref', '=', ref_of(m['name']))], ['id'])
        rec = {**m, 'valor': valor, 'recv_line': rv['id'], 'acc_cli': rv['account_id'][0],
               'partner': rv['partner_id'][0] if rv['partner_id'] else False,
               'residual': rv['amount_residual'], 'reconciled': rv['reconciled'],
               'n_concil': len(rv['matched_credit_ids'] or []),
               'desconciliar': bool(rv['matched_credit_ids'])}
        rec['status'] = 'JA CORRIGIDA - PULA' if dup else 'ELEGIVEL'
        plan.append(rec)

    eleg = [p for p in plan if p['status'] == 'ELEGIVEL']
    total = round(sum(p['valor'] for p in eleg), 2)
    n_desc = sum(1 for p in eleg if p.get('desconciliar'))
    print(f"  ELEGIVEIS: {len(eleg)} (R$ {total:,.2f}) | a desconciliar: {n_desc} | ja abertas: {len(eleg)-n_desc}")
    for p in [x for x in plan if x['status'] != 'ELEGIVEL']:
        print(f"     PULA {p['name']}: {p['status']}")

    if not CONFIRMAR:
        for p in eleg[:6]:
            print(f"     {p['name']:18} R$ {p['valor']:>10,.2f} residual {p['residual']:>10,.2f} desconciliar={p['desconciliar']} ({p['n_concil']})")
        print(f"     ... ({len(eleg)} total)")
        print("\n[DRY-RUN] nada escrito. --confirmar para efetivar.")
        gerar_excel(plan, False)
        return

    fila = eleg[:LIMIT] if LIMIT else eleg
    print("\n" + "=" * 60 + f"\nEXECUTANDO {len(fila)}{' (LIMIT)' if LIMIT else ''}\n" + "=" * 60)
    for i, p in enumerate(fila, 1):
        ref = ref_of(p['name'])
        # 1) capturar pagamentos (antes de desconciliar) + desconciliar
        payment_lines = []
        if p['desconciliar']:
            rvf = o.execute_kw('account.move.line', 'read', [[p['recv_line']], ['matched_credit_ids']], {'context': CTX_LF})
            pr_ids = rvf[0].get('matched_credit_ids') or []
            if pr_ids:
                prs = o.execute_kw('account.partial.reconcile', 'read', [pr_ids, ['credit_move_id']], {'context': CTX_LF})
                payment_lines = [pr['credit_move_id'][0] for pr in prs if pr.get('credit_move_id')]
            o.execute_kw('account.move.line', 'remove_move_reconcile', [[p['recv_line']]], {'context': CTX_LF})
        # 2) entry LF
        lf_id = o.execute_kw('account.move', 'create', [{
            'move_type': 'entry', 'journal_id': J_DIV_LF, 'company_id': 5, 'date': p['date'], 'ref': ref,
            'line_ids': [
                (0, 0, {'account_id': ACC_PASSIVA_LF, 'partner_id': p['partner'], 'name': ref, 'debit': p['valor'], 'credit': 0.0}),
                (0, 0, {'account_id': p['acc_cli'], 'partner_id': p['partner'], 'name': ref, 'debit': 0.0, 'credit': p['valor']}),
            ]}], {'context': CTX_LF})
        o.execute_kw('account.move', 'action_post', [[lf_id]], {'context': CTX_LF})
        ajuste_credit = o.execute_kw('account.move.line', 'search',
                                     [[('move_id', '=', lf_id), ('account_id', '=', p['acc_cli'])]], {'context': CTX_LF})
        # 3) entry FB
        fb_id = o.execute_kw('account.move', 'create', [{
            'move_type': 'entry', 'journal_id': J_DIV_FB, 'company_id': 1, 'date': p['date'], 'ref': ref,
            'line_ids': [
                (0, 0, {'account_id': ACC_CPV_FB, 'name': ref, 'debit': p['valor'], 'credit': 0.0}),
                (0, 0, {'account_id': ACC_ATIVA_FB, 'name': ref, 'debit': 0.0, 'credit': p['valor']}),
            ]}], {'context': CTX_FB})
        o.execute_kw('account.move', 'action_post', [[fb_id]], {'context': CTX_FB})
        # 4) reconciliar o que conseguir (recebivel + pagamentos + ajuste)
        to_rec = [p['recv_line']] + payment_lines + list(ajuste_credit)
        try:
            o.execute_kw('account.move.line', 'reconcile', [to_rec], {'context': CTX_LF})
            p['reconc'] = 'ok'
        except Exception as e:
            p['reconc'] = f'FALHOU: {str(e)[:50]}'
        chk = o.execute_kw('account.move.line', 'read', [[p['recv_line']], ['amount_residual', 'reconciled']], {'context': CTX_LF})
        p['residual_final'] = chk[0]['amount_residual']
        p['lf_entry'], p['fb_entry'] = lf_id, fb_id
        p['status'] = 'CORRIGIDA' + (' (desconc.)' if p['desconciliar'] else '')
        print(f"  {i}/{len(fila)} {p['name']} LF={lf_id} FB={fb_id} desc={p['desconciliar']} reconc={p['reconc']} residual_final={p['residual_final']:,.2f}")

    feitas = [p for p in fila if p.get('status', '').startswith('CORRIGIDA')]
    print(f"\n[OK] {len(feitas)} NFs (R$ {sum(p['valor'] for p in feitas):,.2f}); {sum(1 for p in feitas if p['desconciliar'])} desconciliadas. Excedente/reconc. final = Contadora.")
    gerar_excel(plan, True)


def gerar_excel(plan, aplicado):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'jan-2025'
    HDR = PatternFill('solid', fgColor='1F4E78'); HF = Font(bold=True, color='FFFFFF')
    cols = ['NF retorno (LF)', 'Data', 'Insumos (R$)', 'A receber (residual)', 'Pagto',
            'Desconciliar?', 'Status', 'Lanç. LF', 'Lanç. FB']
    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=h); cell.fill = HDR; cell.font = HF
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    for i, w in enumerate([20, 12, 14, 16, 10, 12, 22, 11, 11], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    r = 2; tot = 0.0
    for p in sorted(plan, key=lambda z: z['name']):
        ws.cell(row=r, column=1, value=p['name'])
        ws.cell(row=r, column=2, value=p.get('date'))
        ws.cell(row=r, column=3, value=p.get('valor', 0)).number_format = '#,##0.00'
        ws.cell(row=r, column=4, value=p.get('residual', '')).number_format = '#,##0.00'
        ws.cell(row=r, column=5, value=p.get('payment_state', ''))
        ws.cell(row=r, column=6, value='SIM' if p.get('desconciliar') else 'nao')
        ws.cell(row=r, column=7, value=p.get('status', ''))
        ws.cell(row=r, column=8, value=p.get('lf_entry', ''))
        ws.cell(row=r, column=9, value=p.get('fb_entry', ''))
        if p.get('status', '').startswith(('CORRIGIDA', 'ELEGIVEL')):
            tot += p.get('valor', 0)
        r += 1
    ws.cell(row=r + 1, column=2, value='TOTAL').font = Font(bold=True)
    c = ws.cell(row=r + 1, column=3, value=round(tot, 2)); c.number_format = '#,##0.00'; c.font = Font(bold=True)
    ws.freeze_panes = 'A2'
    wb.save(OUT)
    print(f"  Excel: {OUT} ({'APLICADO' if aplicado else 'PLANO'})")


if __name__ == '__main__':
    main()
