#!/usr/bin/env python3
"""G9 SERVIÇO por CFOP (READ-ONLY) — divergência REAL de serviço LF x FB, isolando a linha de serviço.

Compara a LINHA DE SERVIÇO (CFOP 5124/5125 na LF VND) x (CFOP 1124/1125 na FB ENTSI) por NF.
Isola dos insumos retornados (5902/1902), que devem bater. NAO escreve no Odoo.
"""
import sys
from collections import defaultdict
sys.path.insert(0, '/home/rafaelnascimento/projetos/frete_sistema')
from app.odoo.utils.connection import get_odoo_connection
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

CTX = {'allowed_company_ids': [1, 5]}
OUT = '/home/rafaelnascimento/projetos/frete_sistema/docs/industrializacao-fb-lf/G9_Servico_Por_CFOP.xlsx'
ACC_LF, P_FB = 26085, 1
ACC_FB, P_LF = 11038, 35
J_ENTSI = 1001
SERV_LF, SERV_FB = {'5124', '5125'}, {'1124', '1125'}
HDR = PatternFill('solid', fgColor='1F4E78'); HF = Font(bold=True, color='FFFFFF')
TIT = Font(bold=True, size=13, color='1F4E78'); B = Font(bold=True); NUM = '#,##0.00'
BAD = PatternFill('solid', fgColor='FCE4D6'); WARN = PatternFill('solid', fgColor='FFF2CC')


def main():
    o = get_odoo_connection()
    assert o.authenticate(), "FALHA AUTH"

    def sr(m, d, f, **k):
        kw = {'fields': f, 'context': CTX}; kw.update(k)
        return o.execute_kw(m, 'search_read', [d], kw)

    def minfo(ids):
        out = {}
        for i in range(0, len(ids), 300):
            for m in o.execute_kw('account.move', 'read', [ids[i:i + 300]], {'fields': ['l10n_br_chave_nf', 'l10n_br_numero_nf', 'date'], 'context': CTX}):
                out[m['id']] = m
        return out

    def keyset(acc, partner, extra):
        ml = sr('account.move.line', [('account_id', '=', acc), ('partner_id', '=', partner), ('parent_state', '=', 'posted')] + extra, ['move_id'])
        mids = list({l['move_id'][0] for l in ml})
        info = minfo(mids)
        byk = {}
        for mid in mids:
            i = info[mid]
            ch = i.get('l10n_br_chave_nf') or (f"NUM:{i['l10n_br_numero_nf']}" if i.get('l10n_br_numero_nf') else f"M:{mid}")
            byk[ch] = {'mid': mid, 'num': i.get('l10n_br_numero_nf'), 'date': i.get('date')}
        return byk

    lf = keyset(ACC_LF, P_FB, [('debit', '>', 0)])
    fb = keyset(ACC_FB, P_LF, [('journal_id', '=', J_ENTSI), ('credit', '>', 0)])
    casados = set(lf) & set(fb)

    def servico(mids, codes):
        sub, tot = defaultdict(float), defaultdict(float)
        mids = list(mids)
        for i in range(0, len(mids), 200):
            lines = sr('account.move.line', [('move_id', 'in', mids[i:i + 200]), ('display_type', '=', 'product')], ['move_id', 'price_subtotal', 'price_total', 'l10n_br_cfop_id'])
            for l in lines:
                cf = l.get('l10n_br_cfop_id')
                code = (cf[1][:4] if cf else '')
                if code in codes:
                    sub[l['move_id'][0]] += l['price_subtotal'] or 0
                    tot[l['move_id'][0]] += l['price_total'] or 0
        return sub, tot

    sub_lf, tot_lf = servico([lf[c]['mid'] for c in casados], SERV_LF)
    sub_fb, tot_fb = servico([fb[c]['mid'] for c in casados], SERV_FB)

    rows = []
    for ch in casados:
        sl = sub_lf.get(lf[ch]['mid'], 0.0)
        sf = sub_fb.get(fb[ch]['mid'], 0.0)
        tl = tot_lf.get(lf[ch]['mid'], 0.0)
        tf = tot_fb.get(fb[ch]['mid'], 0.0)
        d = sl - sf
        rows.append({'num': lf[ch]['num'], 'date': lf[ch]['date'], 'sl': sl, 'sf': sf, 'd': d,
                     'pct': (d / sl * 100) if sl else 0, 'ret_lf': tl, 'ret_fb': tf, 'dret': tl - tf})
    div = [r for r in rows if abs(r['d']) >= 0.01]
    div.sort(key=lambda r: -abs(r['d']))
    sigA = [r for r in div if abs(r['d']) > 50]       # divergencia material (>R$50)
    soma_serv_lf = sum(r['sl'] for r in rows)
    soma_serv_fb = sum(r['sf'] for r in rows)
    soma_tot_lf = sum(r['ret_lf'] for r in rows)   # ret_lf/ret_fb guardam o TOTAL (com impostos)
    soma_tot_fb = sum(r['ret_fb'] for r in rows)

    # ===== Excel =====
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Resumo'
    ws.column_dimensions['A'].width = 54; ws.column_dimensions['B'].width = 18
    ws['A1'] = 'Serviço de industrialização LF × FB (linha CFOP 5124↔1124)'; ws['A1'].font = TIT
    blk = [
        ('NFs casadas analisadas', len(rows)),
        ('— POR BASE (subtotal, sem PIS/COFINS) —', ''),
        ('Σ base serviço LF', soma_serv_lf),
        ('Σ base serviço FB', soma_serv_fb),
        ('Σ Δ base (= efeito do destaque PIS/COFINS)', soma_serv_lf - soma_serv_fb),
        ('— POR TOTAL (com impostos = o que vai à conta corrente) —', ''),
        ('Σ TOTAL serviço LF', soma_tot_lf),
        ('Σ TOTAL serviço FB', soma_tot_fb),
        ('Σ Δ TOTAL (divergência econômica REAL)', soma_tot_lf - soma_tot_fb),
    ]
    r = 3
    for a, v in blk:
        ws.cell(row=r, column=1, value=a)
        c = ws.cell(row=r, column=2, value=v); c.number_format = NUM if isinstance(v, float) else '0'; c.font = B
        r += 1

    ws2 = wb.create_sheet('Detalhe serviço')
    for col, w in zip('ABCDEFGH', [10, 11, 15, 15, 14, 9, 15, 15]):
        ws2.column_dimensions[col].width = w
    for c, h in enumerate(['Nº NF', 'Data', 'Base serv. LF', 'Base serv. FB', 'Δ base', '% base', 'Total serv. LF', 'Total serv. FB'], 1):
        cell = ws2.cell(row=1, column=c, value=h); cell.fill = HDR; cell.font = HF; cell.alignment = Alignment(horizontal='center', wrap_text=True)
    r = 2
    for row in div:
        for c, v in enumerate([row['num'], row['date'], row['sl'], row['sf'], row['d'], round(row['pct'], 1), row['ret_lf'], row['ret_fb']], 1):
            cell = ws2.cell(row=r, column=c, value=v)
            if c in (3, 4, 5, 7, 8):
                cell.number_format = NUM
        ws2.cell(row=r, column=5).fill = WARN if abs(row['d']) > 50 else PatternFill()
        r += 1
    ws2.freeze_panes = 'A2'
    wb.save(OUT)

    print(f"Excel: {OUT}")
    print(f"  {len(rows)} NFs casadas")
    print(f"  BASE  (sem PIS/COFINS): LF {soma_serv_lf:,.2f} | FB {soma_serv_fb:,.2f} | Δ {soma_serv_lf-soma_serv_fb:,.2f}  (efeito PIS/COFINS)")
    print(f"  TOTAL (com impostos):   LF {soma_tot_lf:,.2f} | FB {soma_tot_fb:,.2f} | Δ {soma_tot_lf-soma_tot_fb:,.2f}  (divergência REAL)")


if __name__ == '__main__':
    main()
