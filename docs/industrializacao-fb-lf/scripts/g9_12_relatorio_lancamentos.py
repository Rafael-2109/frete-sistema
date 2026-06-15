#!/usr/bin/env python3
"""G9 RELATORIO (READ-ONLY) — relacao das NFs regularizadas com o NUMERO do lancamento (nao o ID interno).

Le os entries G9 ja criados (ref 'G9-REGULARIZACAO...') de um mes e gera Excel com:
  NF retorno | Data | Insumos (R$) | Lançamento LF | Lançamento FB | Desconciliada? | Recebível final
mostrando o NUMERO visivel do lancamento (ex DIV/2025/01/0023), nao o id.
Uso: python g9_12_relatorio_lancamentos.py --mes 2025-01
"""
import sys
import re
sys.path.insert(0, '/home/rafaelnascimento/projetos/frete_sistema')
from app.odoo.utils.connection import get_odoo_connection
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

CTX = {'allowed_company_ids': [1, 5]}
OPS_5902 = [2864, 2710]
MES = '2025-01'
for i, a in enumerate(sys.argv):
    if a == '--mes' and i + 1 < len(sys.argv):
        MES = sys.argv[i + 1]
import calendar
Y, M = int(MES[:4]), int(MES[5:7])
D1 = f"{MES}-01"
D2 = f"{MES}-{calendar.monthrange(Y, M)[1]:02d}"
OUT = f'/home/rafaelnascimento/projetos/frete_sistema/docs/industrializacao-fb-lf/G9_Relacao_{MES}.xlsx'
RE_NF = re.compile(r'IND\. (\S+) \(')


def main():
    o = get_odoo_connection()
    assert o.authenticate(), "FALHA AUTH"
    print(f"UID {o._uid} | relatorio {MES}")

    def rr(model, domain, fields, **kw):
        kw2 = {'fields': fields, 'context': CTX}
        kw2.update(kw)
        return o.execute_kw(model, 'search_read', [domain], kw2)

    # entries G9 do mes (LF j894 company5 / FB j893 company1)
    lf = rr('account.move', [('journal_id', '=', 894), ('company_id', '=', 5),
                             ('ref', 'like', 'G9-REGULARIZACAO'), ('date', '>=', D1), ('date', '<=', D2)],
            ['id', 'name', 'date', 'ref'])
    fb = rr('account.move', [('journal_id', '=', 893), ('company_id', '=', 1),
                             ('ref', 'like', 'G9-REGULARIZACAO'), ('date', '>=', D1), ('date', '<=', D2)],
            ['id', 'name', 'date', 'ref'])
    fb_by_nf = {}
    for m in fb:
        mm = RE_NF.search(m['ref'] or '')
        if mm:
            fb_by_nf[mm.group(1)] = m['name']

    rows = []
    for m in lf:
        mm = RE_NF.search(m['ref'] or '')
        if not mm:
            continue
        nf = mm.group(1)
        # valor = debito da linha PASSIVA (26667) do entry LF
        vlines = rr('account.move.line', [('move_id', '=', m['id']), ('account_id', '=', 26667)], ['debit'])
        valor = round(sum(x.get('debit') or 0 for x in vlines), 2)
        # recebivel final da NF + se foi desconciliada (heuristica: tem nosso ajuste => foi tratada)
        nfmv = rr('account.move', [('name', '=', nf), ('company_id', '=', 5), ('journal_id', '=', 847)],
                  ['amount_residual', 'payment_state'])
        rows.append({
            'nf': nf, 'data': m['date'], 'valor': valor,
            'lanc_lf': m['name'], 'lanc_fb': fb_by_nf.get(nf, '(?)'),
            'residual': nfmv[0]['amount_residual'] if nfmv else '',
            'pagto': nfmv[0]['payment_state'] if nfmv else '',
        })
    rows.sort(key=lambda r: r['nf'])
    print(f"  NFs no relatorio: {len(rows)}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Regularizacao {MES}'
    HDR = PatternFill('solid', fgColor='1F4E78')
    HF = Font(bold=True, color='FFFFFF')
    ws['A1'] = f'NFs de retorno de industrialização regularizadas — {MES}'
    ws['A1'].font = Font(bold=True, size=13, color='1F4E78')
    ws.merge_cells('A1:F1')
    cols = ['NF de retorno (LF)', 'Data', 'Insumos regularizados (R$)',
            'Lançamento LF', 'Lançamento FB', 'A receber (final)']
    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.fill = HDR
        cell.font = HF
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    for i, w in enumerate([22, 12, 22, 20, 20, 16], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    r = 4
    tot = 0.0
    for x in rows:
        ws.cell(row=r, column=1, value=x['nf'])
        ws.cell(row=r, column=2, value=x['data'])
        ws.cell(row=r, column=3, value=x['valor']).number_format = '#,##0.00'
        ws.cell(row=r, column=4, value=x['lanc_lf'])
        ws.cell(row=r, column=5, value=x['lanc_fb'])
        cell = ws.cell(row=r, column=6, value=x['residual'])
        if isinstance(x['residual'], (int, float)):
            cell.number_format = '#,##0.00'
        tot += x['valor']
        r += 1
    ws.cell(row=r + 1, column=2, value='TOTAL').font = Font(bold=True)
    c = ws.cell(row=r + 1, column=3, value=round(tot, 2))
    c.number_format = '#,##0.00'
    c.font = Font(bold=True)
    ws.freeze_panes = 'A4'
    wb.save(OUT)
    print(f"  Excel: {OUT} ({len(rows)} NFs, total R$ {tot:,.2f})")


if __name__ == '__main__':
    main()
