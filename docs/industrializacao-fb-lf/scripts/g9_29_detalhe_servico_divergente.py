#!/usr/bin/env python3
"""G9 DETALHE serviço LF x NF x serviço FB (READ-ONLY) — onde estão as divergências (causa B).

Para cada NF casada: serviço líq. LF (a-receber − insumos G9) x a-pagar FB (ENTSI) -> Δ.
Analisa direção (LF cobra mais / FB reconhece mais) e distribuição. NAO escreve no Odoo.
"""
import sys
import re
from collections import defaultdict
sys.path.insert(0, '/home/rafaelnascimento/projetos/frete_sistema')
from app.odoo.utils.connection import get_odoo_connection
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

CTX = {'allowed_company_ids': [1, 5]}
OUT = '/home/rafaelnascimento/projetos/frete_sistema/docs/industrializacao-fb-lf/G9_Detalhe_Servico_Divergente.xlsx'
ACC_LF, P_FB = 26085, 1
ACC_FB, P_LF = 11038, 35
J_ENTSI, J_DIV_LF = 1001, 894
RE_VND = re.compile(r'IND\. (\S+) \(')
HDR = PatternFill('solid', fgColor='1F4E78'); HF = Font(bold=True, color='FFFFFF')
TIT = Font(bold=True, size=13, color='1F4E78'); B = Font(bold=True); NUM = '#,##0.00'
BAD = PatternFill('solid', fgColor='FCE4D6'); WARN = PatternFill('solid', fgColor='FFF2CC')


def main():
    o = get_odoo_connection()
    assert o.authenticate(), "FALHA AUTH"

    def sr(m, d, f, **k):
        kw = {'fields': f, 'context': CTX}; kw.update(k)
        return o.execute_kw(m, 'search_read', [d], kw)

    def minfo(ids):
        out = {}
        for i in range(0, len(ids), 300):
            for m in o.execute_kw('account.move', 'read', [ids[i:i + 300]], {'fields': ['name', 'ref', 'l10n_br_chave_nf', 'l10n_br_numero_nf'], 'context': CTX}):
                out[m['id']] = m
        return out

    def chave(info):
        return info.get('l10n_br_chave_nf') or (f"NUM:{info['l10n_br_numero_nf']}" if info.get('l10n_br_numero_nf') else f"MOVE:{info['name']}")

    # LF a-receber
    lf_l = sr('account.move.line', [('account_id', '=', ACC_LF), ('partner_id', '=', P_FB), ('parent_state', '=', 'posted'), ('debit', '>', 0)], ['move_id', 'debit', 'date'])
    lf_mv = defaultdict(lambda: [0.0, None])
    for l in lf_l:
        lf_mv[l['move_id'][0]][0] += l['debit'] or 0; lf_mv[l['move_id'][0]][1] = l['date']
    lf_i = minfo(list(lf_mv.keys()))
    lf = {}
    for mid, (d, dt) in lf_mv.items():
        lf[chave(lf_i[mid])] = {'name': lf_i[mid]['name'], 'num': lf_i[mid].get('l10n_br_numero_nf'), 'date': dt, 'ar': d}
    # FB a-pagar ENTSI
    fb_l = sr('account.move.line', [('account_id', '=', ACC_FB), ('partner_id', '=', P_LF), ('parent_state', '=', 'posted'), ('journal_id', '=', J_ENTSI), ('credit', '>', 0)], ['move_id', 'credit'])
    fb_mv = defaultdict(float)
    for l in fb_l:
        fb_mv[l['move_id'][0]] += l['credit'] or 0
    fb_i = minfo(list(fb_mv.keys()))
    fb = {}
    for mid, c in fb_mv.items():
        fb[chave(fb_i[mid])] = {'name': fb_i[mid]['name'], 'ap': c}
    # insumos G9 por name VND
    g9 = sr('account.move.line', [('account_id', '=', ACC_LF), ('partner_id', '=', P_FB), ('journal_id', '=', J_DIV_LF), ('parent_state', '=', 'posted'), ('credit', '>', 0)], ['move_id', 'credit'])
    g9_ref = {}
    g9_ids = list({l['move_id'][0] for l in g9})
    for i in range(0, len(g9_ids), 300):
        for m in o.execute_kw('account.move', 'read', [g9_ids[i:i + 300]], {'fields': ['ref'], 'context': CTX}):
            g9_ref[m['id']] = m['ref'] or ''
    insumo = defaultdict(float)
    for l in g9:
        mm = RE_VND.search(g9_ref.get(l['move_id'][0], ''))
        if mm:
            insumo[mm.group(1)] += l['credit'] or 0

    rows = []
    for ch in set(lf) & set(fb):
        L, F = lf[ch], fb[ch]
        ins = insumo.get(L['name'], 0.0)
        serv_lf = L['ar'] - ins
        d = serv_lf - F['ap']
        if abs(d) < 0.01:
            continue
        pct = (d / serv_lf * 100) if serv_lf else 0
        rows.append({'num': L['num'], 'date': L['date'], 'ar': L['ar'], 'ins': ins, 'serv': serv_lf,
                     'ap': F['ap'], 'd': d, 'pct': pct, 'dir': 'LF cobra mais' if d > 0 else 'FB reconhece mais',
                     'lf': L['name'], 'fb': F['name'], 'ch': ch})
    rows.sort(key=lambda r: -abs(r['d']))

    lf_mais = [r for r in rows if r['d'] > 0]
    fb_mais = [r for r in rows if r['d'] < 0]
    faixas = defaultdict(lambda: [0, 0.0])
    for r in rows:
        a = abs(r['pct'])
        k = '<1%' if a < 1 else '1-5%' if a < 5 else '5-20%' if a < 20 else '>20%'
        faixas[k][0] += 1; faixas[k][1] += r['d']

    # ===== Excel =====
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = 'Resumo'
    ws.column_dimensions['A'].width = 48; ws.column_dimensions['B'].width = 16; ws.column_dimensions['C'].width = 16
    ws['A1'] = 'Divergência de serviço LF × FB (causa B)'; ws['A1'].font = TIT
    res = [
        ('', '', ''),
        ('NFs com serviço divergente', len(rows), ''),
        ('Σ Δ (LF − FB)', sum(r['d'] for r in rows), ''),
        ('', '', ''),
        ('LF cobra MAIS que a FB reconhece', len(lf_mais), sum(r['d'] for r in lf_mais)),
        ('FB reconhece MAIS que a LF cobra', len(fb_mais), sum(r['d'] for r in fb_mais)),
        ('', '', ''),
        ('Distribuição por |%| do serviço', 'NFs', 'Σ Δ'),
    ]
    for k in ['<1%', '1-5%', '5-20%', '>20%']:
        if k in faixas:
            res.append((f'   {k}', faixas[k][0], faixas[k][1]))
    r = 3
    for a, b_, c_ in res:
        ws.cell(row=r, column=1, value=a)
        if isinstance(b_, (int, float)):
            ws.cell(row=r, column=2, value=b_).number_format = NUM if isinstance(b_, float) else '0'
        elif b_:
            ws.cell(row=r, column=2, value=b_)
        if isinstance(c_, (int, float)):
            ws.cell(row=r, column=3, value=c_).number_format = NUM
        elif c_:
            ws.cell(row=r, column=3, value=c_)
        r += 1

    ws2 = wb.create_sheet('Detalhe NF a NF')
    for col, w in zip('ABCDEFGHIJK', [10, 11, 14, 13, 14, 14, 14, 9, 22, 22, 22]):
        ws2.column_dimensions[col].width = w
    heads = ['Nº NF', 'Data', 'A receber LF', 'Insumos (G9)', 'Serviço líq. LF', 'Serviço FB', 'Δ (LF−FB)', '% serv.', 'Direção', 'Lanç. LF', 'Lanç. FB']
    for c, h in enumerate(heads, 1):
        cell = ws2.cell(row=1, column=c, value=h); cell.fill = HDR; cell.font = HF; cell.alignment = Alignment(horizontal='center', wrap_text=True)
    r = 2
    for row in rows:
        for c, v in enumerate([row['num'], row['date'], row['ar'], row['ins'], row['serv'], row['ap'], row['d'], round(row['pct'], 1), row['dir'], row['lf'], row['fb']], 1):
            cell = ws2.cell(row=r, column=c, value=v)
            if c in (3, 4, 5, 6, 7):
                cell.number_format = NUM
        ws2.cell(row=r, column=9).fill = WARN if row['d'] > 0 else BAD
        r += 1
    ws2.freeze_panes = 'A2'
    wb.save(OUT)

    print(f"Excel: {OUT}")
    print(f"  {len(rows)} NFs divergentes | Σ Δ = R$ {sum(r['d'] for r in rows):,.2f}")
    print(f"  LF cobra mais: {len(lf_mais)} (R$ {sum(r['d'] for r in lf_mais):,.2f}) | FB reconhece mais: {len(fb_mais)} (R$ {sum(r['d'] for r in fb_mais):,.2f})")
    for k in ['<1%', '1-5%', '5-20%', '>20%']:
        if k in faixas:
            print(f"    {k:6s}: {faixas[k][0]:4d} NFs | Σ Δ R$ {faixas[k][1]:,.2f}")
    print("  TOP 8 maiores divergências:")
    for row in rows[:8]:
        print(f"    NF {row['num']} {row['date']} | servLF {row['serv']:>11,.2f} servFB {row['ap']:>11,.2f} Δ {row['d']:>11,.2f} ({row['pct']:.0f}%)")


if __name__ == '__main__':
    main()
