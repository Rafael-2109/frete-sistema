#!/usr/bin/env python3
"""G9 LISTA ADIANTAMENTOS FB (READ-ONLY) — os 34 pagamentos com sobra (pago a maior).

Para cada: data, pagamento, valor pago, aplicado, sobra, e a(s) fatura(s) que recebeu.
Tambem mede quanto ha de fatura ENTSI EM ABERTO (onde a sobra PODERIA ter sido aplicada).
Gera Excel. NAO escreve no Odoo.
"""
import sys
from collections import defaultdict
sys.path.insert(0, '/home/rafaelnascimento/projetos/frete_sistema')
from app.odoo.utils.connection import get_odoo_connection
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

CTX = {'allowed_company_ids': [1, 5]}
OUT = '/home/rafaelnascimento/projetos/frete_sistema/docs/industrializacao-fb-lf/G9_Adiantamentos_FB.xlsx'
ACC_FB, P_LF = 11038, 35
HDR = PatternFill('solid', fgColor='1F4E78'); HF = Font(bold=True, color='FFFFFF')
TIT = Font(bold=True, size=13, color='1F4E78'); B = Font(bold=True); NUM = '#,##0.00'
WARN = PatternFill('solid', fgColor='FFF2CC')


def main():
    o = get_odoo_connection()
    assert o.authenticate(), "FALHA AUTH"

    base_fb = [('account_id', '=', ACC_FB), ('partner_id', '=', P_LF), ('parent_state', '=', 'posted')]
    adi = o.execute_kw('account.move.line', 'search_read',
                       [base_fb + [('debit', '>', 0), ('reconciled', '=', False), ('amount_residual', '>', 0)]],
                       {'fields': ['date', 'debit', 'amount_residual', 'matched_credit_ids', 'ref', 'journal_id', 'move_id'],
                        'order': 'date', 'context': CTX})

    # resolver faturas reconciliadas (partials -> credit move line -> move name)
    all_p = set()
    for l in adi:
        all_p |= set(l.get('matched_credit_ids') or [])
    parts = o.execute_kw('account.partial.reconcile', 'read', [list(all_p)], {'fields': ['credit_move_id', 'amount'], 'context': CTX}) if all_p else []
    pmap = {p['id']: p for p in parts}
    cm_ids = list({p['credit_move_id'][0] for p in parts if p.get('credit_move_id')})
    cm_lines = o.execute_kw('account.move.line', 'read', [cm_ids], {'fields': ['move_id'], 'context': CTX}) if cm_ids else []
    line2move = {l['id']: l['move_id'][1] for l in cm_lines if l.get('move_id')}

    rows = []
    for l in adi:
        faturas = defaultdict(float)
        for pid in (l.get('matched_credit_ids') or []):
            p = pmap.get(pid)
            if p and p.get('credit_move_id'):
                nm = line2move.get(p['credit_move_id'][0], '?')
                faturas[nm] += p.get('amount', 0) or 0
        fat_str = '; '.join(f'{k} ({v:,.2f})' for k, v in sorted(faturas.items(), key=lambda x: -x[1]))
        rows.append({'date': l['date'], 'pag': l['move_id'][1], 'banco': (l['journal_id'][1] if l.get('journal_id') else ''),
                     'pago': l['debit'], 'aplicado': l['debit'] - l['amount_residual'], 'sobra': l['amount_residual'],
                     'fat': fat_str or '(nenhuma — adiantamento puro)', 'ref': l.get('ref') or ''})

    # faturas ENTSI/outras EM ABERTO na FB (onde a sobra poderia ser aplicada)
    g = o.execute_kw('account.move.line', 'read_group',
                     [base_fb + [('credit', '>', 0), ('reconciled', '=', False), ('amount_residual', '<', 0)], ['amount_residual:sum'], []],
                     {'context': CTX, 'lazy': False})
    apagar_aberto = -(g[0].get('amount_residual', 0) or 0)
    n_aberto = g[0]['__count']
    tot_sobra = sum(r['sobra'] for r in rows)

    # ===== Excel =====
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Adiantamentos FB'
    for col, w in zip('ABCDEFGH', [12, 20, 14, 15, 15, 15, 60, 26]):
        ws.column_dimensions[col].width = w
    ws['A1'] = 'Adiantamentos FB → LF (pagamentos com sobra sobre as faturas)'; ws['A1'].font = TIT
    ws['A2'] = f'{len(rows)} pagamentos | sobra total R$ {tot_sobra:,.2f} | há R$ {apagar_aberto:,.2f} de faturas EM ABERTO ({n_aberto}) onde a sobra poderia ser reconciliada'
    ws['A2'].font = Font(italic=True, size=9, color='808080')
    hr = 4
    for c, h in enumerate(['Data', 'Pagamento', 'Banco', 'Valor pago', 'Aplicado', 'Sobra (adiant.)', 'Fatura(s) que recebeu o pagamento', 'Ref'], 1):
        cell = ws.cell(row=hr, column=c, value=h); cell.fill = HDR; cell.font = HF; cell.alignment = Alignment(horizontal='center', wrap_text=True)
    r = hr + 1
    for row in rows:
        for c, v in enumerate([row['date'], row['pag'], row['banco'], row['pago'], row['aplicado'], row['sobra'], row['fat'], row['ref']], 1):
            cell = ws.cell(row=r, column=c, value=v)
            if c in (4, 5, 6):
                cell.number_format = NUM
        ws.cell(row=r, column=6).fill = WARN
        r += 1
    ws.cell(row=r, column=2, value='TOTAL').font = B
    for c, key in ((4, 'pago'), (5, 'aplicado'), (6, 'sobra')):
        cell = ws.cell(row=r, column=c, value=sum(x[key] for x in rows)); cell.number_format = NUM; cell.font = B; cell.fill = WARN
    ws.freeze_panes = 'A5'
    wb.save(OUT)

    print(f"Excel: {OUT}")
    print(f"  {len(rows)} adiantamentos | sobra total R$ {tot_sobra:,.2f}")
    print(f"  faturas FB EM ABERTO (onde aplicar): R$ {apagar_aberto:,.2f} ({n_aberto} linhas)")
    print(f"  {'Data':12s} {'Pago':>12s} {'Sobra':>12s}  Fatura(s)")
    for row in rows:
        print(f"  {row['date']:12s} {row['pago']:>12,.2f} {row['sobra']:>12,.2f}  {row['fat'][:70]}")


if __name__ == '__main__':
    main()
