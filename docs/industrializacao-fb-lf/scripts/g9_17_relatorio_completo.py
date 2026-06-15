#!/usr/bin/env python3
"""G9 RELATORIO COMPLETO (READ-ONLY) p/ Contadora — 1 Excel multi-aba.

Abas: Resumo | Remessas (por doc) | Retornos regularizados (por doc) | Saldo a pagar FB-LF | A receber aberto (LF)
NAO escreve no Odoo.
"""
import sys
import re
from collections import defaultdict
sys.path.insert(0, '/home/rafaelnascimento/projetos/frete_sistema')
from app.odoo.utils.connection import get_odoo_connection
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

CTX = {'allowed_company_ids': [1, 5]}
OUT = '/home/rafaelnascimento/projetos/frete_sistema/docs/industrializacao-fb-lf/G9_Relatorio_Completo_Contadora.xlsx'
ATIVA_INI, PASSIVA_INI = 61930965.26, -37749509.88
BAIXADO = 32314729.70
RE_NF = re.compile(r'IND\. (\S+) \(')
HDR = PatternFill('solid', fgColor='1F4E78'); HF = Font(bold=True, color='FFFFFF')
TIT = Font(bold=True, size=13, color='1F4E78'); B = Font(bold=True); NUM = '#,##0.00'
BAD = PatternFill('solid', fgColor='FCE4D6'); OKF = PatternFill('solid', fgColor='E2EFDA')


def hdrow(ws, row, headers, widths):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HDR; cell.font = HF; cell.alignment = Alignment(horizontal='center', wrap_text=True)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w


def main():
    o = get_odoo_connection()
    assert o.authenticate(), "FALHA AUTH"

    def rr(m, d, f, **k):
        kw = {'fields': f, 'context': CTX}; kw.update(k)
        return o.execute_kw(m, 'search_read', [d], kw)

    def rg(m, d, f, g=[]):
        return o.execute_kw(m, 'read_group', [d, f, g], {'context': CTX, 'lazy': False})

    def saldo(acc):
        g = rg('account.move.line', [('account_id', '=', acc), ('parent_state', '=', 'posted')], ['debit:sum', 'credit:sum'])
        return (g[0]['debit'] or 0) - (g[0]['credit'] or 0)

    ativa, passiva = saldo(22800), saldo(26667)

    # a pagar FB (FORNECEDORES 11038 partner 35) em aberto
    gfb = rg('account.move.line', [('account_id', '=', 11038), ('partner_id', '=', 35), ('parent_state', '=', 'posted'), ('reconciled', '=', False), ('amount_residual', '!=', 0)], ['amount_residual:sum'])
    apagar_fb = -(gfb[0].get('amount_residual', 0) or 0)
    # a receber LF (CLIENTES 26085 partner 1) em aberto
    glf = rg('account.move.line', [('account_id', '=', 26085), ('partner_id', '=', 1), ('parent_state', '=', 'posted'), ('reconciled', '=', False), ('amount_residual', '>', 0)], ['amount_residual:sum'])
    areceber_lf = glf[0].get('amount_residual', 0) or 0

    wb = openpyxl.Workbook()

    # ===== ABA 1 RESUMO =====
    ws = wb.active; ws.title = 'Resumo'
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 52; ws.column_dimensions['C'].width = 22; ws.column_dimensions['D'].width = 22
    ws['B1'] = 'Industrialização FB↔LF — relatório de regularização (01/2025 a hoje)'; ws['B1'].font = TIT
    blocos = [
        ('', '', ''),
        ('CONTAS DE CONTROLE', 'Antes', 'Depois'),
        ('5101010001 REMESSA IND. (ATIVA) — FB', ATIVA_INI, ativa),
        ('5101020001 REMESSA IND. (PASSIVA) — LF', PASSIVA_INI, passiva),
        ('Total baixado (insumos regularizados)', BAIXADO, ''),
        ('Saldo remanescente ATIVA = exercício 2024 (fora do escopo)', '', ''),
        ('', '', ''),
        ('SALDO A PAGAR FB→LF (2 visões)', '', ''),
        ('A receber da LF — após regularização+FIFO', areceber_lf, 'visão LF'),
        ('A pagar reconhecido pela FB (FORNECEDORES)', apagar_fb, 'visão FB'),
        ('DIVERGÊNCIA (escrituração na FB ainda incompleta — Etapa 5/R2)', areceber_lf - apagar_fb, '⚠'),
        ('', '', ''),
        ('MÉTODO', '', ''),
        ('FASE 1 ajuste: D PASSIVA/C Clientes (LF) + D CPV/C ATIVA (FB)', '', ''),
        ('FASE 2 FIFO: créditos/excedentes compensam a-receber mais antigos', '', ''),
        ('Excedente = compensação conta corrente (nunca devolução)', '', ''),
    ]
    r = 3
    for a, b, c in blocos:
        ws.cell(row=r, column=2, value=a)
        if a in ('CONTAS DE CONTROLE', 'SALDO A PAGAR FB→LF (2 visões)', 'MÉTODO'):
            ws.cell(row=r, column=2).font = B
        for col, v in ((3, b), (4, c)):
            if isinstance(v, (int, float)):
                cell = ws.cell(row=r, column=col, value=v); cell.number_format = NUM; cell.font = B
            elif v:
                ws.cell(row=r, column=col, value=v)
        if 'DIVERGÊNCIA' in a:
            ws.cell(row=r, column=2).fill = BAD; ws.cell(row=r, column=3).fill = BAD
        r += 1

    # ===== ABA 2 REMESSAS por doc =====
    ws = wb.create_sheet('Remessas (por doc)')
    lines = rr('account.move.line', [('account_id', '=', 22800), ('parent_state', '=', 'posted'), ('debit', '>', 0), ('date', '>=', '2025-01-01')], ['move_id', 'debit', 'date'], limit=20000)
    bymove = defaultdict(lambda: [0.0, None])
    for l in lines:
        mid = l['move_id'][0]
        bymove[mid][0] += l['debit']
        bymove[mid][1] = l['date']
    mids = list(bymove.keys())
    names = {}
    for i in range(0, len(mids), 300):
        for m in rr('account.move', [('id', 'in', mids[i:i + 300])], ['name']):
            names[m['id']] = m['name']
    rem = sorted([(names.get(mid, mid), bymove[mid][1], bymove[mid][0]) for mid in mids], key=lambda x: (x[1], str(x[0])))
    # FIFO cobertura: mais antigas baixadas ate BAIXADO
    acc = 0.0
    rem_rows = []
    for nome, data, val in rem:
        if acc + val <= BAIXADO + 0.01:
            st = 'retornada (baixada)'; acc += val
        elif acc < BAIXADO:
            st = 'parcial'; acc += val
        else:
            st = 'EM ABERTO (em poder da LF)'
        rem_rows.append((nome, data, val, st))
    hdrow(ws, 1, ['Remessa (NF 5901)', 'Data', 'Valor insumos (R$)', 'Status (cobertura FIFO estimada)'], [26, 12, 20, 30])
    r = 2
    for nome, data, val, st in rem_rows:
        ws.cell(row=r, column=1, value=str(nome))
        ws.cell(row=r, column=2, value=data)
        ws.cell(row=r, column=3, value=val).number_format = NUM
        cell = ws.cell(row=r, column=4, value=st)
        cell.fill = OKF if 'retornada' in st else (BAD if 'ABERTO' in st else PatternFill())
        r += 1
    ws.cell(row=r + 1, column=2, value='TOTAL').font = B
    cc = ws.cell(row=r + 1, column=3, value=sum(x[2] for x in rem_rows)); cc.number_format = NUM; cc.font = B
    ws.freeze_panes = 'A2'

    # ===== ABA 3 RETORNOS regularizados =====
    ws = wb.create_sheet('Retornos regularizados')
    lf = rr('account.move', [('journal_id', '=', 894), ('company_id', '=', 5), ('ref', 'like', 'G9-REGULARIZACAO')], ['name', 'date', 'ref'], limit=5000)
    fb = rr('account.move', [('journal_id', '=', 893), ('company_id', '=', 1), ('ref', 'like', 'G9-REGULARIZACAO')], ['name', 'ref'], limit=5000)
    valmap = {}
    lf_ids = [m['name'] for m in lf]
    # valor por entry LF (debito PASSIVA)
    lfmoves = {m['name']: m for m in lf}
    fb_by_nf = {}
    for m in fb:
        mm = RE_NF.search(m['ref'] or '')
        if mm:
            fb_by_nf[mm.group(1)] = m['name']
    # valor: somar debito 26667 por move LF
    lfmove_ids = rr('account.move', [('journal_id', '=', 894), ('company_id', '=', 5), ('ref', 'like', 'G9-REGULARIZACAO')], ['id', 'name'], limit=5000)
    id2name = {m['id']: m['name'] for m in lfmove_ids}
    vlines = rr('account.move.line', [('move_id', 'in', list(id2name.keys())), ('account_id', '=', 26667)], ['move_id', 'debit'], limit=20000)
    vbymove = defaultdict(float)
    for l in vlines:
        vbymove[l['move_id'][0]] += l['debit'] or 0
    rows3 = []
    for m in lf:
        mm = RE_NF.search(m['ref'] or '')
        nf = mm.group(1) if mm else '?'
        # achar id do move p/ valor
        rows3.append((nf, m['date'], m['name'], fb_by_nf.get(nf, '?')))
    # valor por nf via id
    name2val = {}
    for mid, nm in id2name.items():
        name2val[nm] = vbymove.get(mid, 0)
    rows3.sort(key=lambda x: x[0])
    hdrow(ws, 1, ['NF retorno (LF)', 'Data', 'Insumos (R$)', 'Lançamento LF', 'Lançamento FB'], [20, 12, 16, 18, 18])
    r = 2
    tot3 = 0.0
    for nf, data, lfn, fbn in rows3:
        v = name2val.get(lfn, 0)
        tot3 += v
        ws.cell(row=r, column=1, value=nf)
        ws.cell(row=r, column=2, value=data)
        ws.cell(row=r, column=3, value=v).number_format = NUM
        ws.cell(row=r, column=4, value=lfn)
        ws.cell(row=r, column=5, value=fbn)
        r += 1
    ws.cell(row=r + 1, column=2, value='TOTAL').font = B
    cc = ws.cell(row=r + 1, column=3, value=tot3); cc.number_format = NUM; cc.font = B
    ws.freeze_panes = 'A2'

    # ===== ABA 4 SALDO A PAGAR (2 visoes) =====
    ws = wb.create_sheet('Saldo a pagar FB-LF')
    ws['A1'] = 'Saldo a pagar FB → LF — duas visões'; ws['A1'].font = TIT; ws.merge_cells('A1:C1')
    hdrow(ws, 3, ['Visão', 'Valor (R$)', 'Observação'], [40, 20, 50])
    dados4 = [
        ('A receber da LF (regularizado + FIFO)', areceber_lf, 'serviço genuíno a receber; conta CLIENTES partner FB'),
        ('A pagar reconhecido pela FB (FORNECEDORES)', apagar_fb, 'o que a FB já registrou a pagar à LF'),
        ('DIVERGÊNCIA', areceber_lf - apagar_fb, 'escrituração dos retornos na FB (Etapa 5/R2) incompleta'),
    ]
    r = 4
    for a, v, obs in dados4:
        ws.cell(row=r, column=1, value=a).font = B if 'DIVERG' in a else Font()
        ws.cell(row=r, column=2, value=v).number_format = NUM
        ws.cell(row=r, column=3, value=obs)
        if 'DIVERG' in a:
            for col in (1, 2, 3):
                ws.cell(row=r, column=col).fill = BAD
        r += 1

    # ===== ABA 5 A RECEBER ABERTO (LF) =====
    ws = wb.create_sheet('A receber aberto (LF)')
    ar = rr('account.move.line', [('account_id', '=', 26085), ('partner_id', '=', 1), ('parent_state', '=', 'posted'), ('reconciled', '=', False), ('amount_residual', '>', 0)], ['move_id', 'date', 'amount_residual'], limit=20000, order='date')
    hdrow(ws, 1, ['Documento', 'Data', 'A receber (R$)'], [24, 12, 18])
    r = 2
    for l in ar:
        ws.cell(row=r, column=1, value=l['move_id'][1] if isinstance(l['move_id'], list) else '')
        ws.cell(row=r, column=2, value=l['date'])
        ws.cell(row=r, column=3, value=l['amount_residual']).number_format = NUM
        r += 1
    ws.cell(row=r + 1, column=2, value='TOTAL').font = B
    cc = ws.cell(row=r + 1, column=3, value=sum(l['amount_residual'] for l in ar)); cc.number_format = NUM; cc.font = B
    ws.freeze_panes = 'A2'

    wb.save(OUT)
    print(f"Excel: {OUT}")
    print(f"  remessas {len(rem_rows)} | retornos {len(rows3)} | a-receber aberto {len(ar)}")
    print(f"  ATIVA {ativa:,.2f} | PASSIVA {passiva:,.2f} | a-receber LF {areceber_lf:,.2f} | a-pagar FB {apagar_fb:,.2f} | gap {areceber_lf-apagar_fb:,.2f}")


if __name__ == '__main__':
    main()
