#!/usr/bin/env python3
"""G9 RELATORIO CONTADORA (READ-ONLY) — gera Excel multi-aba p/ a Contadora decidir a regularizacao.

Le saldos frescos do Odoo + a base de casos (g9_universo_desde_2025.csv).
NAO escreve no Odoo. Saida: G9_Relatorio_Contadora.xlsx
"""
import sys
import csv
from collections import defaultdict
sys.path.insert(0, '/home/rafaelnascimento/projetos/frete_sistema')
from app.odoo.utils.connection import get_odoo_connection
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = '/home/rafaelnascimento/projetos/frete_sistema/docs/industrializacao-fb-lf'
CSV_IN = f'{BASE}/g9_universo_desde_2025.csv'
OUT = f'{BASE}/G9_Relatorio_Contadora.xlsx'
CTX = {'allowed_company_ids': [1, 5]}

NUM = '#,##0.00'
HDR_FILL = PatternFill('solid', fgColor='1F4E78')
HDR_FONT = Font(bold=True, color='FFFFFF', size=11)
SUB_FILL = PatternFill('solid', fgColor='D6E4F0')
BAD_FILL = PatternFill('solid', fgColor='FCE4D6')
OK_FILL = PatternFill('solid', fgColor='E2EFDA')
TITLE = Font(bold=True, size=14, color='1F4E78')
BOLD = Font(bold=True)
THIN = Border(*[Side(style='thin', color='BFBFBF')] * 4)


def main():
    o = get_odoo_connection()
    assert o.authenticate(), "FALHA AUTH"

    def saldo(acc):
        g = o.execute_kw('account.move.line', 'read_group',
                         [[('account_id', '=', acc), ('parent_state', '=', 'posted')],
                          ['debit:sum', 'credit:sum'], []], {'context': CTX})
        d = g[0]['debit'] or 0
        c = g[0]['credit'] or 0
        return d, c, d - c

    ativa_fb = saldo(22800)
    passiva_lf = saldo(26667)
    ativa_lf = saldo(26652)

    # base de casos
    rows = []
    with open(CSV_IN) as f:
        for r in csv.DictReader(f):
            r['valor_insumos_5902'] = float(r['valor_insumos_5902'] or 0)
            rows.append(r)

    wb = openpyxl.Workbook()

    def style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = THIN

    # ============ ABA 1 — RESUMO ============
    ws = wb.active
    ws.title = 'Resumo'
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 52
    ws.column_dimensions['C'].width = 24
    ws['B1'] = 'Regularização contábil — Industrialização por encomenda (FB ↔ LF)'
    ws['B1'].font = TITLE
    ws['B2'] = 'NACOM GOYA (FB, encomendante) × LA FAMIGLIA (LF, industrializadora) — material de terceiros'
    ws['B2'].font = Font(italic=True, color='808080')

    linhas = [
        ('', ''),
        ('O PROBLEMA (desde a implantação do Odoo, 07/2024):', ''),
        ('No regime FB encomenda → LF industrializa → LF devolve o PA, as contas de', ''),
        ('controle deveriam ZERAR a cada ciclo (remessa ↔ retorno). Hoje o retorno', ''),
        ('NÃO baixa nenhuma das duas — os insumos de terceiros inflam o ativo da FB.', ''),
        ('', ''),
        ('SALDOS ACUMULADOS (medido ao vivo, Odoo PROD):', ''),
        ('   5101010001 REMESSA INDUSTRIALIZAÇÃO (ATIVA) — FB', ativa_fb[2]),
        ('   5101020001 REMESSA INDUSTRIALIZAÇÃO (PASSIVA) — LF', passiva_lf[2]),
        ('   5101010001 (ATIVA) movimentada na LF via PERDAS (indevido)', ativa_lf[2]),
        ('', ''),
        ('UNIVERSO A REGULARIZAR (notas de retorno desde 01/2025):', ''),
        ('   Nº de notas de retorno (CFOP 5902)', len(rows)),
        ('   Valor dos insumos sem baixa (desde 01/2025)', sum(r['valor_insumos_5902'] for r in rows)),
        ('', ''),
        ('O QUE PEDIMOS DA CONTADORA (ver aba "Decisões pedidas"):', ''),
        ('   Validar o método de regularização (já testado em 1 nota real) e a', ''),
        ('   contrapartida de resultado (CPV) para o histórico — vide perguntas.', ''),
    ]
    r = 4
    for txt, val in linhas:
        ws.cell(row=r, column=2, value=txt)
        if txt.endswith(':'):
            ws.cell(row=r, column=2).font = BOLD
        if isinstance(val, (int, float)):
            cell = ws.cell(row=r, column=3, value=val)
            if isinstance(val, float):
                cell.number_format = NUM
            cell.font = BOLD
        r += 1

    # ============ ABA 2 — DIAGNOSTICO (correto x em uso) ============
    ws = wb.create_sheet('Diagnóstico (correto x uso)')
    for i, w in enumerate([18, 40, 40], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws['A1'] = 'Contabilização do RETORNO — como ESTÁ x como DEVERIA ser'
    ws['A1'].font = TITLE
    ws.merge_cells('A1:C1')
    hdr = ['Etapa / linha', 'EM USO HOJE (errado)', 'CORRETO']
    diag = [
        ('LF — serviço (CFOP 5124)', 'D CLIENTES / C Receita serviço', '= igual (mantém)'),
        ('LF — insumos (CFOP 5902)', 'C 1150100012 / D CLIENTES (embutido) → PASSIVA 5101020001 NÃO baixa',
         'D 5101020001 (PASSIVA) / C 1150100012 → baixa a obrigação'),
        ('FB — serviço (CFOP 1124)', 'D recebimento / C FORNECEDORES', '= igual (a pagar serviço)'),
        ('FB — insumos (CFOP 1902)', 'op c/ movimento de estoque: re-entra os insumos no estoque + ATIVA 5101010001 NÃO baixa',
         'op simbólica (sem estoque): D custo do PA / C 5101010001 → baixa a ATIVA'),
        ('EFEITO LÍQUIDO', 'Ativo de controle infla (nunca baixa) + risco de double-count no estoque físico',
         'As 2 contas de controle ZERAM a cada ciclo; custo do PA correto'),
    ]
    r = 3
    for c, h in enumerate(hdr, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 3)
    r += 1
    for a, b, c in diag:
        ws.cell(row=r, column=1, value=a).font = BOLD
        ws.cell(row=r, column=2, value=b).fill = BAD_FILL
        ws.cell(row=r, column=3, value=c).fill = OK_FILL
        for col in range(1, 4):
            ws.cell(row=r, column=col).alignment = Alignment(wrap_text=True, vertical='top')
            ws.cell(row=r, column=col).border = THIN
        ws.row_dimensions[r].height = 48
        r += 1

    # ============ ABA 3 — SALDOS por ano ============
    ws = wb.create_sheet('Saldos acumulados')
    ws['A1'] = 'Contas de controle — saldo por ano (Odoo PROD)'
    ws['A1'].font = TITLE
    ws.merge_cells('A1:E1')
    for i, w in enumerate([46, 8, 18, 18, 18], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    hdr = ['Conta', 'Ano', 'Débito', 'Crédito', 'Saldo']
    r = 3
    for c, h in enumerate(hdr, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 5)
    r += 1
    for label, acc in [('5101010001 ATIVA (FB)', 22800),
                       ('5101020001 PASSIVA (LF)', 26667),
                       ('5101010001 ATIVA na LF (via PERDAS, indevido)', 26652)]:
        g = o.execute_kw('account.move.line', 'read_group',
                         [[('account_id', '=', acc), ('parent_state', '=', 'posted')],
                          ['debit:sum', 'credit:sum'], ['date:year']], {'context': CTX, 'lazy': False})
        for row_ in sorted(g, key=lambda x: x.get('date:year') or ''):
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=(row_.get('date:year') or '')[:4])
            for col, key in [(3, 'debit'), (4, 'credit')]:
                cell = ws.cell(row=r, column=col, value=row_.get(key, 0) or 0)
                cell.number_format = NUM
            cell = ws.cell(row=r, column=5, value=(row_.get('debit', 0) or 0) - (row_.get('credit', 0) or 0))
            cell.number_format = NUM
            cell.font = BOLD
            r += 1

    # ============ ABA 4 — UNIVERSO resumo ============
    ws = wb.create_sheet('Universo desde 01-2025')
    ws['A1'] = 'Notas de retorno a regularizar — desde 01/2025'
    ws['A1'].font = TITLE
    ws.merge_cells('A1:D1')
    for i, w in enumerate([34, 12, 22, 30], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    by_year = defaultdict(lambda: [0, 0.0])
    for rr_ in rows:
        y = rr_['data'][:4]
        by_year[y][0] += 1
        by_year[y][1] += rr_['valor_insumos_5902']
    r = 3
    ws.cell(row=r, column=1, value='Por ano')
    ws.cell(row=r, column=1).font = BOLD
    ws.cell(row=r, column=1).fill = SUB_FILL
    r += 1
    for c, h in enumerate(['Ano', 'Nº notas', 'Insumos a regularizar', ''], start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 3)
    r += 1
    tn = tv = 0
    for y in sorted(by_year):
        n, v = by_year[y]
        tn += n
        tv += v
        ws.cell(row=r, column=1, value=y)
        ws.cell(row=r, column=2, value=n)
        ws.cell(row=r, column=3, value=v).number_format = NUM
        r += 1
    ws.cell(row=r, column=1, value='TOTAL').font = BOLD
    ws.cell(row=r, column=2, value=tn).font = BOLD
    cc = ws.cell(row=r, column=3, value=tv)
    cc.number_format = NUM
    cc.font = BOLD
    r += 2

    # por status de periodo
    ws.cell(row=r, column=1, value='Por situação do período contábil (lock dates)')
    ws.cell(row=r, column=1).font = BOLD
    ws.cell(row=r, column=1).fill = SUB_FILL
    r += 1
    ws.cell(row=r, column=1, value='Lock: FB fechada até 2025-04-30 · LF fechada até 2025-12-31')
    ws.cell(row=r, column=1).font = Font(italic=True, color='808080')
    r += 2
    for perna, key in [('LF — baixa da PASSIVA', 'lf_periodo'), ('FB — baixa da ATIVA', 'fb_periodo')]:
        st = defaultdict(lambda: [0, 0.0])
        for rr_ in rows:
            st[rr_[key]][0] += 1
            st[rr_[key]][1] += rr_['valor_insumos_5902']
        ws.cell(row=r, column=1, value=perna).font = BOLD
        r += 1
        for c, h in enumerate(['Situação', 'Nº notas', 'Valor', 'Mecanismo de correção'], start=1):
            ws.cell(row=r, column=c, value=h)
        style_header(ws, r, 4)
        r += 1
        for s in ('ABERTO', 'FECHADO'):
            if s in st:
                n, v = st[s]
                ws.cell(row=r, column=1, value=s)
                ws.cell(row=r, column=2, value=n)
                ws.cell(row=r, column=3, value=v).number_format = NUM
                ws.cell(row=r, column=4, value='lançamento por nota (data original)' if s == 'ABERTO'
                        else 'ajuste agregado em período aberto OU reabrir período')
                r += 1
        r += 1

    # ============ ABA 5 — CASO PILOTO PROVADO ============
    ws = wb.create_sheet('Caso-piloto provado')
    ws['A1'] = 'Caso-piloto testado em produção (reversível) — prova do método'
    ws['A1'].font = TITLE
    ws.merge_cells('A1:D1')
    for i, w in enumerate([26, 30, 30, 18], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    info = [
        ('Nota de retorno (LF)', 'VND/2026/00234 (08/04/2026) — MOLHO SHOYU', '', ''),
        ('Entrada na FB', 'ENTSI/2026/04/0025 (09/04/2026)', '', ''),
        ('Valor dos insumos', 1927.44, '', ''),
        ('', '', '', ''),
        ('Lançamento', 'Débito', 'Crédito', 'Valor'),
        ('LF — DIV/2026/04/0003', '5101020001 PASSIVA', '1120100001 CLIENTES', 1927.44),
        ('FB — DIV/2026/04/0016', '3201000001 CPV (custo prod. vendidos)', '5101010001 ATIVA', 1927.44),
        ('', '', '', ''),
        ('Resultado medido', '', '', ''),
        ('PASSIVA 5101020001 (LF)', 'baixada', '', 1927.44),
        ('ATIVA 5101010001 (FB)', 'baixada', '', 1927.44),
        ('A receber da LF (nota)', 'de 2.798,95 → 871,51', '= a pagar da FB (871,51)', ''),
        ('Reversível?', 'Sim — estorno dos 2 lançamentos desfaz', '', ''),
    ]
    r = 3
    for a, b, c, d in info:
        ws.cell(row=r, column=1, value=a)
        if a and not isinstance(b, (int, float)):
            ws.cell(row=r, column=1).font = BOLD
        if isinstance(b, (int, float)):
            ws.cell(row=r, column=2, value=b).number_format = NUM
        else:
            ws.cell(row=r, column=2, value=b)
        ws.cell(row=r, column=3, value=c)
        if isinstance(d, (int, float)):
            ws.cell(row=r, column=4, value=d).number_format = NUM
        if a == 'Lançamento':
            style_header(ws, r, 4)
        r += 1

    # ============ ABA 6 — DECISOES PEDIDAS ============
    ws = wb.create_sheet('Decisões pedidas')
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 95
    ws['B1'] = 'Decisões / validações pedidas à Contadora'
    ws['B1'].font = TITLE
    perg = [
        '',
        '1) MÉTODO — regularizar nota a nota com lançamento de ajuste que (a) baixa a obrigação na',
        '   LF (D 5101020001 / C Clientes) e (b) baixa o ativo de controle na FB reconhecendo o custo',
        '   dos insumos consumidos. É aceitável? (testado em 1 nota real, reversível — aba anterior)',
        '',
        '2) CONTRAPARTIDA NA FB — como o produto acabado já foi vendido, o custo dos insumos não pode',
        '   voltar ao estoque do PA; propomos lançar em CPV (3201000001 Custo dos Produtos Vendidos).',
        '   Confirma o CPV, ou prefere outra conta (ex.: ajuste de exercícios anteriores)?',
        '',
        '3) PERÍODO FECHADO — FB está fechada até 30/04/2025 e LF até 31/12/2025. Para esses períodos',
        '   podemos: (A) fazer um lançamento de ajuste AGREGADO em período aberto (data atual), ou',
        '   (B) reabrir os períodos. Qual o procedimento aceito (considerando SPED/balanços já entregues)?',
        '',
        '4) IMPACTO FISCAL — reconhecer o custo histórico em resultado altera o lucro (IRPJ/CSLL).',
        '   Há orientação sobre a forma e o período de reconhecimento?',
        '',
        '5) SALDO ANTERIOR A 01/2025 — o acumulado total da ATIVA inclui 2024 (~R$ 21 mi). Entra nesta',
        '   regularização ou é tratado separadamente?',
        '',
        'Obs.: o fluxo correto (2 notas / operação simbólica) já está em implementação para PARAR de',
        'gerar novos casos. Esta regularização trata apenas o passivo histórico.',
    ]
    r = 3
    for txt in perg:
        ws.cell(row=r, column=2, value=txt)
        if txt and txt[0].isdigit():
            ws.cell(row=r, column=2).font = BOLD
        r += 1

    # ============ ABA 7 — BASE (1152 casos) ============
    ws = wb.create_sheet('Base de casos')
    cols = ['data', 'lf_nome', 'valor_insumos_5902', 'lf_periodo', 'fb_nome', 'fb_periodo', 'chave']
    headers = ['Data', 'Nota retorno (LF)', 'Insumos (R$)', 'Período LF', 'Entrada (FB)', 'Período FB', 'Chave NF-e']
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))
    for i, w in enumerate([12, 20, 16, 12, 22, 12, 46], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    r = 2
    for rr_ in rows:
        for c, key in enumerate(cols, start=1):
            v = rr_.get(key, '')
            cell = ws.cell(row=r, column=c, value=v)
            if key == 'valor_insumos_5902':
                cell.number_format = NUM
        r += 1
    ws.freeze_panes = 'A2'

    wb.save(OUT)
    print(f"Excel gerado: {OUT}")
    print(f"  abas: {wb.sheetnames}")
    print(f"  ATIVA FB={ativa_fb[2]:,.2f} | PASSIVA LF={passiva_lf[2]:,.2f} | casos={len(rows)} | total insumos={sum(r['valor_insumos_5902'] for r in rows):,.2f}")


if __name__ == '__main__':
    main()
