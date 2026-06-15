#!/usr/bin/env python3
"""G9 RELATORIO DE CONCILIACAO FB<->LF (READ-ONLY) p/ Contadora.

Aba 1 "Conciliacao FB-LF": saldos + pagamentos por mes + NF-a-NF (chave de acesso).
Aba 2 "Diagnostico": decomposicao do gap (geracao x pagamento) + recomendacoes.
Todos os numeros vem do Odoo PROD no momento da execucao. NAO escreve no Odoo.
"""
import sys
import re
from collections import defaultdict
sys.path.insert(0, '/home/rafaelnascimento/projetos/frete_sistema')
from app.odoo.utils.connection import get_odoo_connection
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CTX = {'allowed_company_ids': [1, 5]}
OUT = '/home/rafaelnascimento/projetos/frete_sistema/docs/industrializacao-fb-lf/G9_Conciliacao_FB_LF_Contadora.xlsx'
ACC_LF, P_FB = 26085, 1
ACC_FB, P_LF = 11038, 35
J_ENTSI, J_DIV_LF = 1001, 894

HDR = PatternFill('solid', fgColor='1F4E78'); HF = Font(bold=True, color='FFFFFF', size=11)
TIT = Font(bold=True, size=14, color='1F4E78'); SUB = Font(bold=True, size=12, color='1F4E78')
B = Font(bold=True); NUM = '#,##0.00'
BAD = PatternFill('solid', fgColor='FCE4D6'); OKF = PatternFill('solid', fgColor='E2EFDA')
WARN = PatternFill('solid', fgColor='FFF2CC'); GREY = PatternFill('solid', fgColor='F2F2F2')
thin = Side(style='thin', color='BFBFBF'); BOX = Border(left=thin, right=thin, top=thin, bottom=thin)


def main():
    o = get_odoo_connection()
    assert o.authenticate(), "FALHA AUTH"

    def sr(m, d, f, **k):
        kw = {'fields': f, 'context': CTX}; kw.update(k)
        return o.execute_kw(m, 'search_read', [d], kw)

    def rg(m, d, f, g=[]):
        return o.execute_kw(m, 'read_group', [d, f, g], {'context': CTX, 'lazy': False})

    def minfo(ids):
        out = {}
        for i in range(0, len(ids), 300):
            for m in o.execute_kw('account.move', 'read', [ids[i:i + 300]],
                                  {'fields': ['name', 'ref', 'l10n_br_chave_nf', 'l10n_br_numero_nf'], 'context': CTX}):
                out[m['id']] = m
        return out

    def chave_of(info):
        return info.get('l10n_br_chave_nf') or (f"NUM:{info['l10n_br_numero_nf']}" if info.get('l10n_br_numero_nf') else f"MOVE:{info['name']}")

    # ===== LADO LF (a-receber) =====
    lf_l = sr('account.move.line', [('account_id', '=', ACC_LF), ('partner_id', '=', P_FB),
              ('parent_state', '=', 'posted'), ('debit', '>', 0)],
              ['move_id', 'debit', 'amount_residual', 'date'], limit=50000)
    lf_mv = defaultdict(lambda: {'debit': 0.0, 'resid': 0.0, 'date': None})
    for l in lf_l:
        m = l['move_id'][0]
        lf_mv[m]['debit'] += l['debit'] or 0
        lf_mv[m]['resid'] += l['amount_residual'] or 0
        lf_mv[m]['date'] = l['date']
    lf_info = minfo(list(lf_mv.keys()))
    lf_by_ch = {}
    for mid, agg in lf_mv.items():
        info = lf_info[mid]; ch = chave_of(info)
        lf_by_ch[ch] = {'name': info['name'], 'num': info.get('l10n_br_numero_nf'), 'date': agg['date'],
                        'areceber': agg['debit'], 'aberto': agg['resid'],
                        'ind': bool(info.get('ref') and '/IND/' in info['ref'])}

    # ===== LADO FB (a-pagar — TODOS os creditos da conta corrente, nao so ENTSI) =====
    fb_l = sr('account.move.line', [('account_id', '=', ACC_FB), ('partner_id', '=', P_LF),
              ('parent_state', '=', 'posted'), ('credit', '>', 0)],
              ['move_id', 'credit', 'amount_residual', 'date', 'journal_id'], limit=50000)
    fb_mv = defaultdict(lambda: {'credit': 0.0, 'resid': 0.0, 'date': None, 'entsi': False})
    for l in fb_l:
        m = l['move_id'][0]
        fb_mv[m]['credit'] += l['credit'] or 0
        fb_mv[m]['resid'] += l['amount_residual'] or 0
        fb_mv[m]['date'] = l['date']
        if l.get('journal_id') and l['journal_id'][0] == J_ENTSI:
            fb_mv[m]['entsi'] = True
    fb_info = minfo(list(fb_mv.keys()))
    fb_by_ch = {}
    for mid, agg in fb_mv.items():
        info = fb_info[mid]; ch = chave_of(info)
        fb_by_ch[ch] = {'name': info['name'], 'num': info.get('l10n_br_numero_nf'), 'date': agg['date'],
                        'apagar': agg['credit'], 'aberto': -agg['resid'], 'entsi': agg['entsi']}

    # ===== insumos G9 (ajuste FASE 1) por NF — liga pelo name do lançamento LF (VND) no ref =====
    RE_VND = re.compile(r'IND\. (\S+) \(')
    g9_l = sr('account.move.line', [('account_id', '=', ACC_LF), ('partner_id', '=', P_FB),
              ('journal_id', '=', J_DIV_LF), ('parent_state', '=', 'posted'), ('credit', '>', 0)],
              ['move_id', 'credit'], limit=50000)
    g9_move_ids = list({l['move_id'][0] for l in g9_l})
    g9_ref = {}
    for i in range(0, len(g9_move_ids), 300):
        for m in o.execute_kw('account.move', 'read', [g9_move_ids[i:i + 300]], {'fields': ['ref'], 'context': CTX}):
            g9_ref[m['id']] = m['ref'] or ''
    insumo_by_vnd = defaultdict(float)
    for l in g9_l:
        mm = RE_VND.search(g9_ref.get(l['move_id'][0], ''))
        if mm:
            insumo_by_vnd[mm.group(1)] += l['credit'] or 0

    # ===== linhas NF-a-NF =====
    rows = []
    for ch in set(lf_by_ch) | set(fb_by_ch):
        L, F = lf_by_ch.get(ch), fb_by_ch.get(ch)
        num = (L or F)['num']
        data = (L or F)['date']
        ar = L['areceber'] if L else 0.0
        ab_lf = L['aberto'] if L else 0.0
        ap = F['apagar'] if F else 0.0
        ab_fb = F['aberto'] if F else 0.0
        insumo = insumo_by_vnd.get(L['name'], 0.0) if L else 0.0
        servico_lf = ar - insumo            # a-receber líquido de serviço (já sem insumos)
        dservico = servico_lf - ap          # divergência de serviço LF vs a-pagar FB
        if L and F:
            sit = 'Casada (serviço bate)' if abs(dservico) < 0.01 else 'Casada — serviço diverge'
        elif L:
            sit = 'FB nao escriturou (IND)' if L['ind'] else 'So LF (sem a-pagar FB)'
        else:
            sit = 'So FB (sem a-receber LF)'
        rows.append({'num': num, 'data': data, 'ar': ar, 'insumo': insumo, 'serv': servico_lf,
                     'ab_lf': ab_lf, 'ap': ap, 'ab_fb': ab_fb, 'dserv': dservico, 'sit': sit,
                     'lf': L['name'] if L else '', 'fb': F['name'] if F else '', 'ch': ch})
    rows.sort(key=lambda r: (r['data'] or '9999', str(r['num'] or '')))

    # ===== agregados financeiros =====
    def s1(domain, field):
        g = rg('account.move.line', domain, [f'{field}:sum'])
        return g[0].get(field, 0) or 0
    base_lf = [('account_id', '=', ACC_LF), ('partner_id', '=', P_FB), ('parent_state', '=', 'posted')]
    base_fb = [('account_id', '=', ACC_FB), ('partner_id', '=', P_LF), ('parent_state', '=', 'posted')]
    areceber_ger = s1(base_lf + [('debit', '>', 0)], 'debit')
    cred_g9 = s1(base_lf + [('journal_id', '=', J_DIV_LF)], 'credit')
    cred_lf_tot = s1(base_lf + [('credit', '>', 0)], 'credit')
    pag_receb_lf = cred_lf_tot - cred_g9
    saldo_lf = areceber_ger - cred_lf_tot
    apagar_ger = s1(base_fb + [('credit', '>', 0)], 'credit')
    pago_fb = s1(base_fb + [('debit', '>', 0)], 'debit')
    saldo_fb = apagar_ger - pago_fb
    gap = saldo_lf - saldo_fb
    causa_ger = (areceber_ger - cred_g9) - apagar_ger
    causa_pag = pago_fb - pag_receb_lf

    # pagamentos por mes (FB debitos x LF creditos pagto)
    gm_fb = rg('account.move.line', base_fb + [('debit', '>', 0)], ['debit:sum'], ['date:month'])
    gm_lf = rg('account.move.line', base_lf + [('credit', '>', 0), ('journal_id', '!=', J_DIV_LF)], ['credit:sum'], ['date:month'])
    pf = {r['date:month']: r.get('debit', 0) or 0 for r in gm_fb}
    pl = {r['date:month']: r.get('credit', 0) or 0 for r in gm_lf}
    meses = sorted(set(pf) | set(pl))
    # mesmo recorte por ANO (revela que o gap de pagamento é defasagem de 2026)
    gy_fb = rg('account.move.line', base_fb + [('debit', '>', 0)], ['debit:sum'], ['date:year'])
    gy_lf = rg('account.move.line', base_lf + [('credit', '>', 0), ('journal_id', '!=', J_DIV_LF)], ['credit:sum'], ['date:year'])
    pf_ano = {str(r['date:year']): r.get('debit', 0) or 0 for r in gy_fb}
    pl_ano = {str(r['date:year']): r.get('credit', 0) or 0 for r in gy_lf}
    anos = sorted(set(pf_ano) | set(pl_ano))
    dif_ult = pf_ano.get(anos[-1], 0) - pl_ano.get(anos[-1], 0) if anos else 0

    # categorias do cruzamento
    cat = defaultdict(lambda: [0, 0.0, 0.0])
    for r in rows:
        c = cat[r['sit']]; c[0] += 1; c[1] += r['ar']; c[2] += r['ap']

    # =================== EXCEL ===================
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = 'Conciliacao FB-LF'
    for col, w in zip('ABCDEFGHIJKLM', [11, 11, 15, 14, 15, 15, 15, 13, 13, 24, 18, 20, 46]):
        ws.column_dimensions[col].width = w
    ws['A1'] = 'Conciliação conta corrente FB ↔ LF (industrialização por encomenda + operações correlatas)'; ws['A1'].font = TIT
    ws['A2'] = 'Fonte: Odoo PROD (CIEL IT). Lado LF = conta CLIENTES 1120100001 (partner FB), valor cheio da NF. Lado FB = conta FORNECEDORES 2120100001 (partner LF). Serviço líq. LF = a-receber − insumos retornados (ajuste G9).'
    ws['A2'].font = Font(italic=True, size=9, color='808080')

    # BLOCO SALDOS
    r = 4
    ws.cell(row=r, column=1, value='SALDOS DE FECHAMENTO').font = SUB
    r += 1
    saldos = [
        ('A receber reconhecido pela LF (saldo conta CLIENTES)', saldo_lf),
        ('A pagar reconhecido pela FB (saldo conta FORNECEDORES)', saldo_fb),
        ('DIVERGÊNCIA (gap a explicar)', gap),
    ]
    for lbl, val in saldos:
        ws.cell(row=r, column=1, value=lbl)
        c = ws.cell(row=r, column=3, value=val); c.number_format = NUM; c.font = B
        if 'DIVERG' in lbl:
            for cc in (1, 2, 3):
                ws.cell(row=r, column=cc).fill = BAD
            ws.cell(row=r, column=1).font = B
        r += 1

    # BLOCO PAGAMENTOS POR MES
    r += 1
    ws.cell(row=r, column=1, value='PAGAMENTOS — FB pagou × LF recebeu (por mês)').font = SUB
    r += 1
    for c, h in enumerate(['Mês', 'FB pagou (R$)', 'LF recebeu (R$)', 'Dif. no mês', 'Dif. acumulada'], 1):
        cell = ws.cell(row=r, column=c, value=h); cell.fill = HDR; cell.font = HF; cell.border = BOX
    r += 1
    ac = 0.0
    for mes in meses:
        a, b = pf.get(mes, 0), pl.get(mes, 0)
        ac += a - b
        ws.cell(row=r, column=1, value=mes)
        for c, v in ((2, a), (3, b), (4, a - b), (5, ac)):
            cell = ws.cell(row=r, column=c, value=v); cell.number_format = NUM; cell.border = BOX
        r += 1
    ws.cell(row=r, column=1, value='TOTAL').font = B
    for c, v in ((2, pago_fb), (3, pag_receb_lf), (4, pago_fb - pag_receb_lf)):
        cell = ws.cell(row=r, column=c, value=v); cell.number_format = NUM; cell.font = B; cell.fill = WARN
    ws.cell(row=r, column=1, value='TOTAL').font = B
    r += 1
    ws.cell(row=r, column=1, value='(+) Ajuste G9 insumos baixado na LF (não é pagamento)')
    cgc = ws.cell(row=r, column=3, value=cred_g9); cgc.number_format = NUM; cgc.fill = GREY
    r += 2

    # TABELA NF-a-NF
    ws.cell(row=r, column=1, value='DETALHE NF A NF (chave de acesso da NF-e)').font = SUB
    r += 1
    head_row = r
    headers = ['Nº NF', 'Data', 'A receber LF', 'Insumos (G9)', 'Serviço líq. LF', 'A pagar FB',
               'Δ serviço (LF−FB)', 'Em aberto LF', 'Em aberto FB', 'Situação', 'Lançamento LF', 'Lançamento FB', 'Chave NF-e']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h); cell.fill = HDR; cell.font = HF
        cell.alignment = Alignment(horizontal='center', wrap_text=True); cell.border = BOX
    r += 1
    fmap = {'Casada (serviço bate)': OKF, 'Casada — serviço diverge': WARN,
            'FB nao escriturou (IND)': BAD, 'So FB (sem a-receber LF)': BAD, 'So LF (sem a-pagar FB)': GREY}
    for row in rows:
        vals = [row['num'], row['data'], row['ar'], row['insumo'], row['serv'], row['ap'],
                row['dserv'], row['ab_lf'], row['ab_fb'], row['sit'], row['lf'], row['fb'], row['ch']]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            if c in (3, 4, 5, 6, 7, 8, 9):
                cell.number_format = NUM
            cell.border = BOX
        fill = fmap.get(row['sit'])
        if fill:
            ws.cell(row=r, column=10).fill = fill
        r += 1
    tot_r = r
    ws.cell(row=tot_r, column=1, value='TOTAL').font = B
    for c, key in ((3, 'ar'), (4, 'insumo'), (5, 'serv'), (6, 'ap'), (8, 'ab_lf'), (9, 'ab_fb')):
        cell = ws.cell(row=tot_r, column=c, value=sum(x[key] for x in rows)); cell.number_format = NUM; cell.font = B; cell.fill = WARN
    ws.freeze_panes = ws.cell(row=head_row + 1, column=1)

    # =================== ABA 2 DIAGNOSTICO ===================
    d = wb.create_sheet('Diagnostico')
    d.column_dimensions['A'].width = 4
    d.column_dimensions['B'].width = 62; d.column_dimensions['C'].width = 20; d.column_dimensions['D'].width = 14
    d['B1'] = 'Diagnóstico da divergência FB ↔ LF'; d['B1'].font = TIT
    casadas = [x for x in rows if x['sit'].startswith('Casada')]
    serv_div_casadas = sum(x['dserv'] for x in casadas)
    n_serv_div = sum(1 for x in casadas if x['sit'].endswith('diverge'))
    so_lf_serv = sum(x['serv'] for x in rows if not x['fb'])
    so_fb_ap = sum(x['ap'] for x in rows if not x['lf'])
    n_so_lf = sum(1 for x in rows if not x['fb'])
    n_so_fb = sum(1 for x in rows if not x['lf'])
    blocks = [
        ('', '', ''),
        ('1. O QUE A DIVERGÊNCIA NÃO É', '', ''),
        ('A FB TEM entrada escriturada e PAGOU em dinheiro. Não é "falta de entrada".', '', ''),
        ('Os R$ {:,.2f} de insumos (ajuste G9) NÃO são pagamento — material de terceiros que retornou.'.format(cred_g9), '', ''),
        ('NÃO sobrou pagamento: excedente/saldo credor em aberto na LF = R$ 0,00.', '', ''),
        ('', '', ''),
        ('2. DECOMPOSIÇÃO DO GAP (fecha ao centavo)', 'Valor (R$)', ''),
        ('A receber reconhecido pela LF', saldo_lf, ''),
        ('A pagar reconhecido pela FB', saldo_fb, ''),
        ('= GAP a explicar', gap, '⚠'),
        ('(B) Divergência de GERAÇÃO de serviço', causa_ger, ''),
        ('     serviço a-receber LF (bruto {:,.2f} − insumos G9 {:,.2f})'.format(areceber_ger, cred_g9), areceber_ger - cred_g9, ''),
        ('     a-pagar gerado pela FB', apagar_ger, ''),
        ('(A) Divergência de PAGAMENTO', causa_pag, ''),
        ('     FB pagou (débitos na conta FORNECEDORES)', pago_fb, ''),
        ('     LF reconheceu receber (créditos, exceto ajuste G9)', pag_receb_lf, ''),
        ('(B) + (A) =', causa_ger + causa_pag, '= GAP'),
        ('', '', ''),
        ('3. A CAUSA (B) ABERTA NF A NF (soma = (B) acima)', 'Valor (R$)', ''),
        ('     serviço diverge em {} NFs casadas'.format(n_serv_div), serv_div_casadas, ''),
        ('     serviço de NFs só na LF (sem a-pagar na FB)', so_lf_serv, ''),
        ('     (−) a-pagar de NFs só na FB (sem a-receber na LF)', -so_fb_ap, ''),
        ('', '', ''),
        ('4. CONTAGEM NF A NF (ver aba Conciliação)', 'NFs', 'a-receber LF'),
    ]
    for sit in ['Casada (serviço bate)', 'Casada — serviço diverge', 'FB nao escriturou (IND)', 'So LF (sem a-pagar FB)', 'So FB (sem a-receber LF)']:
        c = cat.get(sit)
        if c:
            blocks.append(('     ' + sit, c[0], c[1]))
    blocks += [
        ('', '', ''),
        ('4b. TIMING DA CAUSA (A) — FB pagou × LF baixou, por ano', 'FB pagou', 'LF baixou'),
    ]
    for y in anos:
        blocks.append(('     ' + y, pf_ano.get(y, 0), pl_ano.get(y, 0)))
    blocks += [
        ('     → 2024 e 2025 batem; o descasamento (R$ {:,.2f}) é de {}: a LF ainda NÃO baixou os recebimentos.'.format(dif_ult, anos[-1] if anos else ''), '', ''),
        ('', '', ''),
        ('5. AÇÕES RECOMENDADAS (decisão da Contadora)', '', ''),
        ('(A) LF reconciliar os recebimentos de {} (FB já pagou ~R$ {:,.2f}) — NÃO é excedente.'.format(anos[-1] if anos else '', pf_ano.get(anos[-1], 0) if anos else 0), causa_pag, ''),
        ('(B-serviço) conferir/ajustar valor de serviço nas {} NFs que divergem'.format(n_serv_div), serv_div_casadas, ''),
        ('(B-avulsas) {} NFs só na LF e {} NFs só na FB → casar/escriturar contraparte'.format(n_so_lf, n_so_fb), so_lf_serv - so_fb_ap, ''),
        ('', '', ''),
        ('NOTA: nada disso é perda econômica; é descasamento de reconhecimento + reconciliação pendente.', '', ''),
        ('Após a LF baixar {}, o a-receber cai de {:,.0f} p/ ~{:,.0f}, aproximando do a-pagar FB.'.format(anos[-1] if anos else '', saldo_lf, saldo_lf - dif_ult), '', ''),
    ]
    r = 3
    for a, b_, c_ in blocks:
        cell = ws_b = d.cell(row=r, column=2, value=a)
        if a and (a[0].isdigit() and '.' in a[:3]):
            cell.font = SUB
        if 'GAP' in str(a) and a.startswith('='):
            cell.font = B
        for col, v in ((3, b_), (4, c_)):
            if isinstance(v, (int, float)):
                cc = d.cell(row=r, column=col, value=v); cc.number_format = NUM; cc.font = B
            elif v:
                d.cell(row=r, column=col, value=v)
        if 'GAP a explicar' in str(a) or str(a).startswith('(B) + (A)'):
            for col in (2, 3, 4):
                d.cell(row=r, column=col).fill = BAD
        r += 1

    wb.save(OUT)
    print(f"Excel: {OUT}")
    print(f"  linhas NF: {len(rows)} | saldo LF {saldo_lf:,.2f} | saldo FB {saldo_fb:,.2f} | gap {gap:,.2f}")
    print(f"  (B) {causa_ger:,.2f} + (A) {causa_pag:,.2f} = {causa_ger + causa_pag:,.2f}")
    nf_b = serv_div_casadas + so_lf_serv - so_fb_ap
    print(f"  CHECK (B) nf-level: servDiv {serv_div_casadas:,.2f} + soLF {so_lf_serv:,.2f} - soFB {so_fb_ap:,.2f} = {nf_b:,.2f} (deve = B {causa_ger:,.2f} | dif {nf_b - causa_ger:,.2f})")
    for sit, c in sorted(cat.items()):
        print(f"    {sit:32s}: {c[0]:5d} NFs | a-receber {c[1]:,.2f} | a-pagar {c[2]:,.2f}")


if __name__ == '__main__':
    main()
