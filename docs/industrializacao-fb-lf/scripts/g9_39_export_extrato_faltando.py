#!/usr/bin/env python3
"""G9 EXPORTA (READ-ONLY) as transacoes FB->LF 2026 SEM entrada no extrato LF.

Mesma logica do g9_38 (cruza por TRANSACAO/move, multiset por valor): cada pagamento
FB->LF (debito FORNECEDORES, partner LF) deveria ter uma entrada NACOM no extrato da LF.
As transacoes sem par = extrato faltando (ou transito). Exporta detalhe p/ Excel para o
operador localizar e importar o extrato correspondente. NAO escreve nada no Odoo.
"""
import sys
from collections import defaultdict, Counter
sys.path.insert(0, '/home/rafaelnascimento/projetos/frete_sistema')
from app.odoo.utils.connection import get_odoo_connection
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

CTX = {'allowed_company_ids': [1, 5]}
OUT = '/home/rafaelnascimento/projetos/frete_sistema/docs/industrializacao-fb-lf/G9_Extrato_Faltando_LF_2026.xlsx'
ACC_FORN_FB, P_LF = 11038, 35
HDR = PatternFill('solid', fgColor='C00000'); HF = Font(bold=True, color='FFFFFF')
TIT = Font(bold=True, size=13, color='C00000'); B = Font(bold=True); NUM = '#,##0.00'


def main():
    o = get_odoo_connection()
    assert o.authenticate(), "FALHA AUTH"

    def sr(m, d, f, **k):
        kw = {'fields': f, 'context': CTX}; kw.update(k)
        return o.execute_kw(m, 'search_read', [d], kw)

    # A = pagamentos FB -> LF 2026 (debitos FORNECEDORES partner LF), agrupados por move = transacao
    pag = sr('account.move.line', [('account_id', '=', ACC_FORN_FB), ('partner_id', '=', P_LF),
             ('parent_state', '=', 'posted'), ('debit', '>', 0), ('date', '>=', '2026-01-01')],
             ['date', 'debit', 'move_id'])
    trans = defaultdict(lambda: {'val': 0.0, 'date': None, 'nfs': 0})
    for p in pag:
        t = trans[p['move_id'][0]]; t['val'] += p['debit']; t['date'] = p['date']; t['nfs'] += 1

    # B = entradas NACOM no extrato da LF 2026
    jb = [j['id'] for j in sr('account.journal', [('type', '=', 'bank'), ('company_id', '=', 5)], ['id'])]
    ent = sr('account.bank.statement.line', [('journal_id', 'in', jb), ('company_id', '=', 5),
             ('date', '>=', '2026-01-01'), ('amount', '>', 0), ('payment_ref', 'ilike', 'NACOM')], ['amount'])

    # multiset match por valor (centavo) -> faltantes
    Bc = Counter(round(e['amount'], 2) for e in ent)
    falt = []
    for mid, t in trans.items():
        k = round(t['val'], 2)
        if Bc.get(k, 0) > 0:
            Bc[k] -= 1
        else:
            falt.append((mid, t))

    # detalhes dos moves faltantes (name, ref, journal pagador FB, statement line)
    mids = [mid for mid, _ in falt]
    info = {}
    for i in range(0, len(mids), 300):
        for m in o.execute_kw('account.move', 'read', [mids[i:i + 300]],
                              {'fields': ['name', 'ref', 'journal_id', 'statement_line_id'], 'context': CTX}):
            info[m['id']] = m
    # memo real do extrato (statement.line) de cada move
    slids = [info[m]['statement_line_id'][0] for m in info if info[m].get('statement_line_id')]
    memo = {}
    for i in range(0, len(slids), 300):
        for s in o.execute_kw('account.bank.statement.line', 'read', [slids[i:i + 300]],
                              {'fields': ['payment_ref'], 'context': CTX}):
            memo[s['id']] = s['payment_ref'] or ''
    # heuristica de natureza: memo de utilidade/tarifa = NAO eh pagamento FB->LF
    NAOLF = ('AGUA', 'ESGOTO', 'SABESP', 'ENERGIA', 'ELETR', 'TARIFA', 'INTERNET', 'TELEFON')

    def natureza(m):
        u = m.upper()
        if any(t in u for t in NAOLF):
            return 'REVISAR (despesa nao-LF?)'
        if '18.467.441' in u or 'FAMIGLIA' in u:
            return 'PIX FB->LF (destino=CNPJ LF)'
        return 'PIX FB->LF (verificar)'

    rows = []
    for mid, t in falt:
        m = info.get(mid, {})
        sl = m.get('statement_line_id')
        memo_txt = memo.get(sl[0], '') if sl else ''
        rows.append({'date': t['date'], 'val': t['val'], 'nfs': t['nfs'],
                     'name': m.get('name', ''),
                     'ref': str(m.get('ref') or ''),
                     'banco': (m.get('journal_id') or [0, ''])[1],
                     'memo': memo_txt, 'nat': natureza(memo_txt)})
    rows.sort(key=lambda x: (x['date'], -x['val']))

    tot = sum(r['val'] for r in rows)

    suspeitos = [r for r in rows if r['nat'].startswith('REVISAR')]
    tot_susp = sum(r['val'] for r in suspeitos)
    WARN = PatternFill('solid', fgColor='FFC7CE')

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Extrato faltando LF 2026'
    for col, w in zip('ABCDEFGH', [12, 14, 7, 17, 20, 16, 42, 26]):
        ws.column_dimensions[col].width = w
    ws['A1'] = 'G9 — Pagamentos FB->LF 2026 SEM entrada no extrato LF'; ws['A1'].font = TIT
    ws['A2'] = (f'{len(rows)} transacoes = R$ {tot:,.2f} | dos quais {len(suspeitos)} A REVISAR (despesa nao-LF) = '
                f'R$ {tot_susp:,.2f} | gap real FB->LF = R$ {tot - tot_susp:,.2f}')
    ws['A2'].font = Font(italic=True, size=9, color='808080')
    heads = ['Data pagto FB', 'Valor', 'Nº NFs', 'Pagamento (move)', 'Ref bancaria', 'Banco pagador', 'Memo extrato', 'Natureza']
    for c, h in enumerate(heads, 1):
        cell = ws.cell(row=4, column=c, value=h); cell.fill = HDR; cell.font = HF
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    r = 5
    for row in rows:
        vals = [row['date'], row['val'], row['nfs'], row['name'], row['ref'], row['banco'], row['memo'], row['nat']]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            if c == 2:
                cell.number_format = NUM
            if row['nat'].startswith('REVISAR'):
                cell.fill = WARN
        r += 1
    ws.cell(row=r + 1, column=1, value='TOTAL').font = B
    cc = ws.cell(row=r + 1, column=2, value=tot); cc.number_format = NUM; cc.font = B
    ws.freeze_panes = 'A5'
    wb.save(OUT)

    print(f"Excel: {OUT}")
    print(f"  transacoes FB sem entrada no extrato LF: {len(rows)} = R$ {tot:,.2f}")
    print(f"  A REVISAR (despesa nao-LF): {len(suspeitos)} = R$ {tot_susp:,.2f}  |  gap real FB->LF = R$ {tot - tot_susp:,.2f}\n")
    print(f"  {'Data':<11} {'Valor':>12}  {'Pagamento':<17} {'Natureza':<28} Memo")
    for row in rows:
        print(f"  {row['date']:<11} {row['val']:>12,.2f}  {row['name']:<17} {row['nat']:<28} {row['memo'][:45]}")


if __name__ == '__main__':
    main()
