#!/usr/bin/env python3
"""G9 ROLLOUT abril/2026 (reversao pura per-documento) — TODAS as NFs de retorno do mes.

DRY-RUN e o DEFAULT. So efetiva com --confirmar.
Para cada NF de retorno (j847, com 5902) de abr/2026:
  LF: D 5101020001 PASSIVA (26667) / C <conta recebivel da NF>  + concilia com o recebivel -> baixa PASSIVA
  FB: D 3201000001 CPV (22527) / C 5101010001 ATIVA (22800)                                -> baixa ATIVA
Idempotente por REF (pula as ja corrigidas). Gera Excel com a relacao.
"""
import sys
sys.path.insert(0, '/home/rafaelnascimento/projetos/frete_sistema')
from app.odoo.utils.connection import get_odoo_connection
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CTX_BOTH = {'allowed_company_ids': [1, 5]}
CTX_LF = {'allowed_company_ids': [5]}
CTX_FB = {'allowed_company_ids': [1]}
OPS_5902 = [2864, 2710]
ACC_PASSIVA_LF = 26667
ACC_CPV_FB = 22527
ACC_ATIVA_FB = 22800
J_DIV_LF = 894
J_DIV_FB = 893
OUT_XLSX = '/home/rafaelnascimento/projetos/frete_sistema/docs/industrializacao-fb-lf/G9_NFs_corrigidas_abril2026.xlsx'
CONFIRMAR = '--confirmar' in sys.argv


def ref_of(name):
    return f"G9-REGULARIZACAO IND. {name} (reversao pura per-doc)"


def main():
    o = get_odoo_connection()
    assert o.authenticate(), "FALHA AUTH"
    print(f"UID {o._uid} | MODO: {'CONFIRMAR (ESCREVE)' if CONFIRMAR else 'DRY-RUN'}\n")

    def rr(model, domain, fields, ctx=CTX_BOTH, **kw):
        kw2 = {'fields': fields, 'context': ctx}
        kw2.update(kw)
        return o.execute_kw(model, 'search_read', [domain], kw2)

    # 1) NFs retorno abr/2026
    moves = rr('account.move',
               [('journal_id', '=', 847), ('company_id', '=', 5), ('state', '=', 'posted'),
                ('move_type', '=', 'out_invoice'),
                ('date', '>=', '2026-04-01'), ('date', '<=', '2026-04-30')],
               ['id', 'name', 'date', 'l10n_br_chave_nf', 'amount_total'], order='id')
    print(f"NFs j847 abr/2026: {len(moves)}")

    plan = []
    for m in moves:
        mid = m['id']
        # valor insumos 5902
        l5902 = rr('account.move.line', [('move_id', '=', mid), ('l10n_br_operacao_id', 'in', OPS_5902)],
                   ['credit'])
        valor = round(sum(l.get('credit') or 0 for l in l5902), 2)
        if valor <= 0:
            plan.append({**m, 'valor': 0, 'status': 'SEM 5902 (venda pura) - PULA'})
            continue
        # linha de recebivel
        recv = rr('account.move.line',
                  [('move_id', '=', mid), ('account_id.account_type', '=', 'asset_receivable'),
                   ('debit', '>', 0)],
                  ['id', 'account_id', 'partner_id', 'amount_residual', 'debit'])
        if len(recv) != 1:
            plan.append({**m, 'valor': valor, 'status': f'RECEBIVEL inesperado ({len(recv)} linhas) - PULA'})
            continue
        rv = recv[0]
        # idempotencia
        dup = rr('account.move', [('ref', '=', ref_of(m['name']))], ['id'])
        rec = {**m, 'valor': valor, 'recv_line': rv['id'],
               'acc_cli': rv['account_id'][0], 'partner': rv['partner_id'][0] if rv['partner_id'] else False,
               'residual': rv['amount_residual'], 'recv_debit': rv['debit']}
        if dup:
            rec['status'] = f'JA CORRIGIDA (entries {[d["id"] for d in dup]}) - PULA'
        elif rv['amount_residual'] + 0.01 < valor:
            rec['status'] = f'RESIDUAL {rv["amount_residual"]:.2f} < insumos {valor:.2f} - REVISAR'
        else:
            rec['status'] = 'ELEGIVEL'
        plan.append(rec)

    eleg = [p for p in plan if p.get('status') == 'ELEGIVEL']
    pula = [p for p in plan if p.get('status') != 'ELEGIVEL']
    tot_eleg = round(sum(p['valor'] for p in eleg), 2)
    print(f"\n  ELEGIVEIS: {len(eleg)}  (R$ {tot_eleg:,.2f})")
    print(f"  PULA/REVISAR: {len(pula)}")
    for p in pula:
        print(f"     {p['name']:18} R$ {p.get('valor',0):>10,.2f}  {p['status']}")

    if not CONFIRMAR:
        print("\n  --- amostra elegiveis ---")
        for p in eleg[:8]:
            print(f"     {p['name']:18} {p['date']} R$ {p['valor']:>10,.2f} residual {p['residual']:>10,.2f}")
        print(f"     ... ({len(eleg)} no total)")
        print("\n[DRY-RUN] nada escrito. Para efetivar: --confirmar")
        gerar_excel(plan, aplicado=False)
        return

    # ============ EXECUCAO ============
    print("\n" + "=" * 60 + "\nEXECUTANDO\n" + "=" * 60)
    for i, p in enumerate(eleg, 1):
        name, valor, partner, acc_cli, recv_line = p['name'], p['valor'], p['partner'], p['acc_cli'], p['recv_line']
        ref = ref_of(name)
        # LF
        lf_id = o.execute_kw('account.move', 'create', [{
            'move_type': 'entry', 'journal_id': J_DIV_LF, 'company_id': 5, 'date': p['date'], 'ref': ref,
            'line_ids': [
                (0, 0, {'account_id': ACC_PASSIVA_LF, 'partner_id': partner, 'name': ref, 'debit': valor, 'credit': 0.0}),
                (0, 0, {'account_id': acc_cli, 'partner_id': partner, 'name': ref, 'debit': 0.0, 'credit': valor}),
            ]}], {'context': CTX_LF})
        o.execute_kw('account.move', 'action_post', [[lf_id]], {'context': CTX_LF})
        lf_credit = o.execute_kw('account.move.line', 'search',
                                 [[('move_id', '=', lf_id), ('account_id', '=', acc_cli)]], {'context': CTX_LF})
        o.execute_kw('account.move.line', 'reconcile', [[lf_credit[0], recv_line]], {'context': CTX_LF})
        # FB
        fb_id = o.execute_kw('account.move', 'create', [{
            'move_type': 'entry', 'journal_id': J_DIV_FB, 'company_id': 1, 'date': p['date'], 'ref': ref,
            'line_ids': [
                (0, 0, {'account_id': ACC_CPV_FB, 'name': ref, 'debit': valor, 'credit': 0.0}),
                (0, 0, {'account_id': ACC_ATIVA_FB, 'name': ref, 'debit': 0.0, 'credit': valor}),
            ]}], {'context': CTX_FB})
        o.execute_kw('account.move', 'action_post', [[fb_id]], {'context': CTX_FB})
        p['lf_entry'] = lf_id
        p['fb_entry'] = fb_id
        p['status'] = 'CORRIGIDA'
        if i % 10 == 0 or i == len(eleg):
            print(f"  {i}/{len(eleg)} ... {name} LF={lf_id} FB={fb_id}")

    print(f"\n[OK] {len(eleg)} NFs corrigidas (R$ {tot_eleg:,.2f}).")
    gerar_excel(plan, aplicado=True)


def gerar_excel(plan, aplicado):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'NFs abril-2026'
    HDR = PatternFill('solid', fgColor='1F4E78')
    HF = Font(bold=True, color='FFFFFF')
    THIN = Border(*[Side(style='thin', color='BFBFBF')] * 4)
    cols = ['NF retorno (LF)', 'Data', 'Insumos (R$)', 'Status',
            'Lanç. LF', 'Lanç. FB', 'A receber antes', 'Conta recebível']
    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HDR
        cell.font = HF
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN
    widths = [20, 12, 15, 34, 12, 12, 16, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    r = 2
    tot = 0.0
    for p in sorted(plan, key=lambda x: x['name']):
        ws.cell(row=r, column=1, value=p['name'])
        ws.cell(row=r, column=2, value=p['date'])
        ws.cell(row=r, column=3, value=p.get('valor', 0)).number_format = '#,##0.00'
        ws.cell(row=r, column=4, value=p.get('status', ''))
        ws.cell(row=r, column=5, value=p.get('lf_entry', ''))
        ws.cell(row=r, column=6, value=p.get('fb_entry', ''))
        ws.cell(row=r, column=7, value=p.get('residual', '')).number_format = '#,##0.00'
        ws.cell(row=r, column=8, value=p.get('acc_cli', ''))
        if p.get('status') in ('CORRIGIDA', 'ELEGIVEL'):
            tot += p.get('valor', 0)
        r += 1
    ws.cell(row=r + 1, column=2, value='TOTAL').font = Font(bold=True)
    c = ws.cell(row=r + 1, column=3, value=round(tot, 2))
    c.number_format = '#,##0.00'
    c.font = Font(bold=True)
    ws.freeze_panes = 'A2'
    wb.save(OUT_XLSX)
    print(f"  Excel: {OUT_XLSX} ({'APLICADO' if aplicado else 'PLANO'}) — {len(plan)} NFs, total elegivel R$ {tot:,.2f}")


if __name__ == '__main__':
    main()
