#!/usr/bin/env python3
"""G9 PROPOSTA por periodo (READ-ONLY) — gera a proposta de regularizacao de um mes p/ a Contadora.

NAO escreve no Odoo (proposta). Mede, para o mes (default 2025-01):
  - NFs de retorno (j847, com 5902), valor dos insumos;
  - estado do recebivel de cada NF (aberto x conciliado/pago) -> define se a perna LF usa conciliacao;
  - lock dates -> se o periodo esta fechado (precisa data de lancamento em periodo aberto).
Gera Excel-proposta. Uso: python g9_10_proposta_periodo.py [--mes 2025-01]
"""
import sys
import calendar
sys.path.insert(0, '/home/rafaelnascimento/projetos/frete_sistema')
from app.odoo.utils.connection import get_odoo_connection
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CTX = {'allowed_company_ids': [1, 5]}
OPS_5902 = [2864, 2710]

MES = '2025-01'
for i, a in enumerate(sys.argv):
    if a == '--mes' and i + 1 < len(sys.argv):
        MES = sys.argv[i + 1]
Y, M = int(MES[:4]), int(MES[5:7])
D1 = f"{MES}-01"
D2 = f"{MES}-{calendar.monthrange(Y, M)[1]:02d}"
OUT = f'/home/rafaelnascimento/projetos/frete_sistema/docs/industrializacao-fb-lf/G9_Proposta_{MES}.xlsx'


def main():
    o = get_odoo_connection()
    assert o.authenticate(), "FALHA AUTH"
    print(f"UID {o._uid} | PROPOSTA {MES} ({D1}..{D2})\n")

    def rr(model, domain, fields, **kw):
        kw2 = {'fields': fields, 'context': CTX}
        kw2.update(kw)
        return o.execute_kw(model, 'search_read', [domain], kw2)

    # lock dates
    comps = rr('res.company', [('id', 'in', [1, 5])], ['id', 'fiscalyear_lock_date'])
    lock = {c['id']: c.get('fiscalyear_lock_date') for c in comps}
    fb_fechado = (lock.get(1) or '') >= D2
    lf_fechado = (lock.get(5) or '') >= D2
    print(f"Lock: FB={lock.get(1)} (mes {'FECHADO' if fb_fechado else 'ABERTO'}) | LF={lock.get(5)} (mes {'FECHADO' if lf_fechado else 'ABERTO'})\n")

    # NFs do mes
    moves = rr('account.move',
               [('journal_id', '=', 847), ('company_id', '=', 5), ('state', '=', 'posted'),
                ('move_type', '=', 'out_invoice'), ('date', '>=', D1), ('date', '<=', D2)],
               ['id', 'name', 'date', 'l10n_br_chave_nf', 'amount_total', 'payment_state'], order='id')
    print(f"NFs j847 {MES}: {len(moves)}")

    rows = []
    for m in moves:
        l5902 = rr('account.move.line', [('move_id', '=', m['id']), ('l10n_br_operacao_id', 'in', OPS_5902)], ['credit'])
        valor = round(sum(l.get('credit') or 0 for l in l5902), 2)
        if valor <= 0:
            continue
        recv = rr('account.move.line',
                  [('move_id', '=', m['id']), ('account_id.account_type', '=', 'asset_receivable'), ('debit', '>', 0)],
                  ['amount_residual', 'reconciled', 'debit'])
        residual = recv[0]['amount_residual'] if recv else 0
        recon = recv[0]['reconciled'] if recv else False
        if recon or residual + 0.01 < valor:
            perna_lf = 'recebivel pago/insuf. -> baixar PASSIVA contra resultado'
        else:
            perna_lf = 'recebivel aberto -> D PASSIVA / C Clientes (concilia)'
        rows.append({'name': m['name'], 'date': m['date'], 'valor': valor,
                     'residual': residual, 'reconciled': recon, 'payment_state': m.get('payment_state'),
                     'perna_lf': perna_lf})

    total = round(sum(r['valor'] for r in rows), 2)
    abertos = sum(1 for r in rows if 'aberto' in r['perna_lf'])
    pagos = len(rows) - abertos
    print(f"  NFs com retorno: {len(rows)} | total insumos R$ {total:,.2f}")
    print(f"  recebivel ABERTO (concilia): {abertos} | recebivel PAGO/insuf (contra resultado): {pagos}")
    print(f"  payment_states: {dict((s, sum(1 for r in rows if r['payment_state']==s)) for s in set(r['payment_state'] for r in rows))}")

    # Excel proposta
    wb = openpyxl.Workbook()
    HDR = PatternFill('solid', fgColor='1F4E78'); HF = Font(bold=True, color='FFFFFF')
    TITLE = Font(bold=True, size=13, color='1F4E78'); BOLD = Font(bold=True)
    ws = wb.active; ws.title = 'Proposta'
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 56; ws.column_dimensions['C'].width = 24
    ws['B1'] = f'Proposta de regularização — retorno de industrialização — {MES}'
    ws['B1'].font = TITLE
    info = [
        ('', ''),
        (f'Período contábil: FB {"FECHADO" if fb_fechado else "aberto"} · LF {"FECHADO" if lf_fechado else "aberto"} (lock FB {lock.get(1)} / LF {lock.get(5)})', ''),
        ('', ''),
        ('Nº de notas de retorno (CFOP 5902)', len(rows)),
        ('Valor dos insumos a regularizar', total),
        ('  — com recebível em aberto (baixa via D PASSIVA / C Clientes)', abertos),
        ('  — com recebível já pago (baixa da PASSIVA contra resultado)', pagos),
        ('', ''),
        ('Lançamento por nota:', ''),
        ('  FB (todas as 52): D 3201000001 CPV / C 5101010001 (ATIVA) — uniforme', ''),
        (f'  LF c/ recebível ABERTO ({abertos}): D 5101020001 PASSIVA / C Clientes + concilia', ''),
        (f'  LF c/ recebível PAGO/parcial ({pagos}): D 5101020001 PASSIVA / C ??? — A DECIDIR', ''),
        ('', ''),
        ('DECISÃO PEDIDA À CONTADORA:', ''),
        (f'  Período de {MES}: {"ABERTO" if not (fb_fechado or lf_fechado) else "FECHADO"} (lock FB {lock.get(1)} / LF {lock.get(5)}).', ''),
        (f'  Para as {pagos} notas com recebível JÁ PAGO/parcial, a baixa da PASSIVA', ''),
        ('  não pode reduzir o Clientes (já liquidado). Contra qual conta baixar?', ''),
        ('  (ex.: resultado / conta de acerto intercompany FB↔LF / outra)', ''),
    ]
    r = 3
    for txt, val in info:
        ws.cell(row=r, column=2, value=txt)
        if txt.endswith(':'):
            ws.cell(row=r, column=2).font = BOLD
        if isinstance(val, (int, float)):
            cell = ws.cell(row=r, column=3, value=val)
            if isinstance(val, float):
                cell.number_format = '#,##0.00'
            cell.font = BOLD
        r += 1

    ws2 = wb.create_sheet('Relação NFs')
    cols = ['NF retorno (LF)', 'Data', 'Insumos (R$)', 'A receber (residual)', 'Situação pagto', 'Tratamento perna LF']
    for c, h in enumerate(cols, 1):
        cell = ws2.cell(row=1, column=c, value=h); cell.fill = HDR; cell.font = HF
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    for i, w in enumerate([20, 12, 15, 18, 14, 46], 1):
        ws2.column_dimensions[chr(64 + i)].width = w
    r = 2
    for x in sorted(rows, key=lambda z: z['name']):
        ws2.cell(row=r, column=1, value=x['name'])
        ws2.cell(row=r, column=2, value=x['date'])
        ws2.cell(row=r, column=3, value=x['valor']).number_format = '#,##0.00'
        ws2.cell(row=r, column=4, value=x['residual']).number_format = '#,##0.00'
        ws2.cell(row=r, column=5, value=x['payment_state'])
        ws2.cell(row=r, column=6, value=x['perna_lf'])
        r += 1
    ws2.cell(row=r + 1, column=2, value='TOTAL').font = BOLD
    c = ws2.cell(row=r + 1, column=3, value=total); c.number_format = '#,##0.00'; c.font = BOLD
    ws2.freeze_panes = 'A2'
    wb.save(OUT)
    print(f"\nExcel proposta: {OUT}")
    print("\n[FIM — READ-ONLY, nada escrito no Odoo]")


if __name__ == '__main__':
    main()
