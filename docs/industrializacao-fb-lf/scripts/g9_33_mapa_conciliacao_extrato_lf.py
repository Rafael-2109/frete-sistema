#!/usr/bin/env python3
"""G9 MAPA conciliação extrato LF 2026 (READ-ONLY) — entradas FB x faturas a receber.

Identifica as entradas FB (payment_ref NACOM) nao conciliadas, simula FIFO contra as
faturas a receber da FB e mede quantas batem 1:1 (alta confiança) vs agregadas. NAO escreve.
"""
import sys
from collections import defaultdict
sys.path.insert(0, '/home/rafaelnascimento/projetos/frete_sistema')
from app.odoo.utils.connection import get_odoo_connection
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

CTX = {'allowed_company_ids': [1, 5]}
OUT = '/home/rafaelnascimento/projetos/frete_sistema/docs/industrializacao-fb-lf/G9_Mapa_Conciliacao_Extrato_LF.xlsx'
ACC_LF, P_FB = 26085, 1
HDR = PatternFill('solid', fgColor='1F4E78'); HF = Font(bold=True, color='FFFFFF')
TIT = Font(bold=True, size=13, color='1F4E78'); B = Font(bold=True); NUM = '#,##0.00'
OKF = PatternFill('solid', fgColor='E2EFDA'); WARN = PatternFill('solid', fgColor='FFF2CC')


def main():
    o = get_odoo_connection()
    assert o.authenticate(), "FALHA AUTH"

    jb = [j['id'] for j in o.execute_kw('account.journal', 'search_read', [[('type', '=', 'bank'), ('company_id', '=', 5)]], {'fields': ['id'], 'context': CTX})]
    ent = o.execute_kw('account.bank.statement.line', 'search_read',
                       [[('journal_id', 'in', jb), ('company_id', '=', 5), ('date', '>=', '2026-01-01'),
                         ('amount', '>', 0), ('payment_ref', 'ilike', 'NACOM'), ('is_reconciled', '=', False)]],
                       {'fields': ['date', 'amount', 'payment_ref'], 'order': 'date', 'context': CTX})
    # faturas a receber FB (residual>0) FIFO por data
    fat = o.execute_kw('account.move.line', 'search_read',
                       [[('account_id', '=', ACC_LF), ('partner_id', '=', P_FB), ('parent_state', '=', 'posted'),
                         ('debit', '>', 0), ('amount_residual', '>', 0)]],
                       {'fields': ['move_id', 'amount_residual', 'date'], 'order': 'date, id', 'context': CTX})
    pilha = [{'nf': f['move_id'][1], 'r': f['amount_residual'], 'date': f['date']} for f in fat]
    resid_set = sorted(round(f['amount_residual'], 2) for f in fat)
    resid_count = defaultdict(int)
    for v in resid_set:
        resid_count[v] += 1

    rows = []
    n11 = v11 = 0
    coberto = 0.0
    for e in ent:
        amt = e['amount']
        # 1:1? existe fatura com residual == amt (unica)
        um_a_um = resid_count.get(round(amt, 2), 0) == 1
        # FIFO: quais faturas cobre
        falta = amt; usadas = []
        for f in pilha:
            if f['r'] <= 0.005:
                continue
            u = min(falta, f['r']); f['r'] -= u; falta -= u
            usadas.append(f"{f['nf']} ({u:,.2f})")
            if falta <= 0.005:
                break
        aplicado = amt - falta; coberto += aplicado
        if um_a_um:
            n11 += 1; v11 += amt
        rows.append({'date': e['date'], 'amt': amt, 'ref': e['payment_ref'], '11': 'sim' if um_a_um else '',
                     'fifo': '; '.join(usadas[:3]) + (f'  (+{len(usadas)-3})' if len(usadas) > 3 else '')})

    tot = sum(e['amount'] for e in ent)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Entradas FB 2026'
    for col, w in zip('ABCDE', [12, 16, 8, 60, 50]):
        ws.column_dimensions[col].width = w
    ws['A1'] = 'Conciliação extrato LF 2026 — entradas da FB (memo NACOM) × faturas a receber'; ws['A1'].font = TIT
    ws['A2'] = f'{len(ent)} entradas FB = R$ {tot:,.2f} | cobertura FIFO R$ {coberto:,.2f} | batem 1:1 (única fatura = valor): {n11} (R$ {v11:,.2f})'
    ws['A2'].font = Font(italic=True, size=9, color='808080')
    for c, h in enumerate(['Data', 'Valor entrada', '1:1?', 'Faturas FIFO que cobriria', 'Memo'], 1):
        cell = ws.cell(row=4, column=c, value=h); cell.fill = HDR; cell.font = HF; cell.alignment = Alignment(horizontal='center', wrap_text=True)
    r = 5
    for row in rows:
        for c, v in enumerate([row['date'], row['amt'], row['11'], row['fifo'], row['ref']], 1):
            cell = ws.cell(row=r, column=c, value=v)
            if c == 2:
                cell.number_format = NUM
        if row['11']:
            ws.cell(row=r, column=3).fill = OKF
        r += 1
    ws.cell(row=r + 1, column=1, value='TOTAL').font = B
    cc = ws.cell(row=r + 1, column=2, value=tot); cc.number_format = NUM; cc.font = B
    ws.freeze_panes = 'A5'
    wb.save(OUT)

    print(f"Excel: {OUT}")
    print(f"  entradas FB 2026 (NACOM): {len(ent)} = R$ {tot:,.2f}")
    print(f"  batem 1:1 (valor = residual de fatura ÚNICA): {n11} = R$ {v11:,.2f}")
    print(f"  agregadas (cobrem várias faturas / valor não-único): {len(ent)-n11} = R$ {tot-v11:,.2f}")
    print(f"  cobertura FIFO simulada: R$ {coberto:,.2f}")


if __name__ == '__main__':
    main()
