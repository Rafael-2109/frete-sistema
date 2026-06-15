#!/usr/bin/env python3
"""G9 DRY-RUN reconciliacao dos 34 adiantamentos FB (pago a maior) contra as faturas-alvo.

Matching DIRIGIDO: o ref de cada pagamento ja lista as faturas ENTSI/CMPMP que ele
deveria quitar. Casa cada adiantamento contra essas faturas (greedy, controlando residual).
DRY-RUN default (so simula + Excel). --confirmar efetiva o reconcile no Odoo.
NAO efetivar sem autorizacao explicita da Contadora/financeiro.
"""
import sys
import re
from collections import defaultdict
sys.path.insert(0, '/home/rafaelnascimento/projetos/frete_sistema')
from app.odoo.utils.connection import get_odoo_connection
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

CTX = {'allowed_company_ids': [1, 5]}
OUT = '/home/rafaelnascimento/projetos/frete_sistema/docs/industrializacao-fb-lf/G9_DryRun_Reconciliacao_Adiantamentos.xlsx'
ACC_FB, P_LF = 11038, 35
CONFIRMAR = '--confirmar' in sys.argv
RE_FAT = re.compile(r'[A-Z]{3,6}/\d{4}/\d{2}/\d{4}')
HDR = PatternFill('solid', fgColor='1F4E78'); HF = Font(bold=True, color='FFFFFF')
TIT = Font(bold=True, size=13, color='1F4E78'); B = Font(bold=True); NUM = '#,##0.00'
OKF = PatternFill('solid', fgColor='E2EFDA'); WARN = PatternFill('solid', fgColor='FFF2CC'); BAD = PatternFill('solid', fgColor='FCE4D6')


def main():
    o = get_odoo_connection()
    assert o.authenticate(), "FALHA AUTH"
    print(f"UID {o._uid} | {'CONFIRMAR (EFETIVA)' if CONFIRMAR else 'DRY-RUN'}\n")

    base_fb = [('account_id', '=', ACC_FB), ('partner_id', '=', P_LF), ('parent_state', '=', 'posted')]
    adi = o.execute_kw('account.move.line', 'search_read',
                       [base_fb + [('debit', '>', 0), ('reconciled', '=', False), ('amount_residual', '>', 0)]],
                       {'fields': ['date', 'debit', 'amount_residual', 'ref', 'move_id'], 'order': 'date', 'context': CTX})

    # faturas-alvo citadas nos refs
    alvo_names = set()
    for l in adi:
        alvo_names |= set(RE_FAT.findall(l.get('ref') or ''))
    # linhas dessas faturas na conta FORNECEDORES (credito a-pagar) com residual
    fat_moves = o.execute_kw('account.move', 'search_read', [[('name', 'in', list(alvo_names)), ('company_id', '=', 1)]],
                             {'fields': ['name'], 'context': CTX})
    mid2name = {m['id']: m['name'] for m in fat_moves}
    fat_lines = o.execute_kw('account.move.line', 'search_read',
                             [[('move_id', 'in', list(mid2name)), ('account_id', '=', ACC_FB), ('credit', '>', 0), ('parent_state', '=', 'posted')]],
                             {'fields': ['move_id', 'amount_residual', 'reconciled'], 'context': CTX}) if mid2name else []
    fat = {}  # name -> {line_id, resid (a-pagar em aberto, >0)}
    for l in fat_lines:
        nm = mid2name.get(l['move_id'][0])
        fat[nm] = {'line_id': l['id'], 'resid': -(l['amount_residual'] or 0)}

    # simular casamento (controla residual consumido)
    rows, plano = [], []
    tot_rec = tot_sobra = 0.0
    for l in adi:
        valor = l['amount_residual']
        alvos = RE_FAT.findall(l.get('ref') or '')
        aplica = 0.0; det = []; casar_lines = []
        for a in alvos:
            f = fat.get(a)
            if f is None:
                det.append(f'{a}: não encontrada'); continue
            if f['resid'] <= 0.005:
                det.append(f'{a}: já paga'); continue
            usar = min(valor - aplica, f['resid'])
            if usar > 0.005:
                aplica += usar; f['resid'] -= usar
                det.append(f'{a}: R$ {usar:,.2f}'); casar_lines.append(f['line_id'])
            if aplica >= valor - 0.005:
                break
        sobra = valor - aplica
        tot_rec += aplica; tot_sobra += sobra
        status = ('reconciliável total' if sobra < 0.01 else
                  'reconciliável parcial' if aplica > 0.01 else 'fatura-alvo já quitada')
        rows.append({'date': l['date'], 'pag': l['move_id'][1], 'valor': valor, 'aplica': aplica,
                     'sobra': sobra, 'det': '; '.join(det), 'status': status})
        if casar_lines:
            plano.append((l['id'], casar_lines, l['move_id'][1]))

    # ===== Excel =====
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'DryRun Reconciliacao'
    for col, w in zip('ABCDEFG', [12, 22, 16, 16, 16, 22, 70]):
        ws.column_dimensions[col].width = w
    ws['A1'] = 'DRY-RUN — reconciliação dos adiantamentos FB contra as faturas-alvo (do próprio ref do pagamento)'; ws['A1'].font = TIT
    ws['A2'] = f'{len(rows)} adiantamentos | reconciliável R$ {tot_rec:,.2f} | sem alvo disponível R$ {tot_sobra:,.2f} | SIMULAÇÃO — nada efetivado no Odoo'
    ws['A2'].font = Font(italic=True, size=9, color='808080')
    hr = 4
    for c, h in enumerate(['Data', 'Pagamento', 'Valor adiant.', 'Reconciliável', 'Sem alvo', 'Status', 'Faturas-alvo (do ref) e valor casado'], 1):
        cell = ws.cell(row=hr, column=c, value=h); cell.fill = HDR; cell.font = HF; cell.alignment = Alignment(horizontal='center', wrap_text=True)
    r = hr + 1
    for row in rows:
        for c, v in enumerate([row['date'], row['pag'], row['valor'], row['aplica'], row['sobra'], row['status'], row['det']], 1):
            cell = ws.cell(row=r, column=c, value=v)
            if c in (3, 4, 5):
                cell.number_format = NUM
        st = row['status']
        ws.cell(row=r, column=6).fill = OKF if st.endswith('total') else WARN if 'parcial' in st else BAD
        r += 1
    ws.cell(row=r, column=2, value='REGRA: nunca casa valor > residual da fatura (min). Sobra sem alvo NÃO é forçada em fatura alguma.').font = Font(italic=True, size=9, color='C00000')
    r += 1
    ws.cell(row=r, column=2, value='TOTAL').font = B
    for c, key in ((3, 'valor'), (4, 'aplica'), (5, 'sobra')):
        cell = ws.cell(row=r, column=c, value=sum(x[key] for x in rows)); cell.number_format = NUM; cell.font = B; cell.fill = WARN
    ws.freeze_panes = 'A5'
    wb.save(OUT)

    print(f"Excel: {OUT}")
    print(f"  {len(rows)} adiantamentos | reconciliável R$ {tot_rec:,.2f} | sem alvo disponível R$ {tot_sobra:,.2f}")
    by = defaultdict(lambda: [0, 0.0])
    for row in rows:
        by[row['status']][0] += 1; by[row['status']][1] += row['valor']
    for st, c in sorted(by.items()):
        print(f"    {st:24s}: {c[0]:3d} | R$ {c[1]:,.2f}")

    if not CONFIRMAR:
        print("\n[DRY-RUN] nada efetivado. --confirmar efetiva (somente com autorização explícita).")
        return
    print("\n[CONFIRMAR] efetivando reconcile...")
    ok = fail = 0
    for deb_id, cred_ids, pagnome in plano:
        try:
            o.execute_kw('account.move.line', 'reconcile', [[deb_id] + cred_ids], {'context': CTX})
            ok += 1
        except Exception as e:
            fail += 1
            print(f"  FALHA {pagnome}: {str(e)[:90]}")
    print(f"[FIM] reconciliados ok={ok} falhas={fail}")


if __name__ == '__main__':
    main()
