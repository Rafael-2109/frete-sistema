#!/usr/bin/env python3
"""G9 INVESTIGA DUPLICIDADE (READ-ONLY) — os 33 adiantamentos cujas faturas-alvo ja estao pagas.

Para cada adiantamento: soma das faturas-alvo (S) e QUEM ja as quitou (pagamento gemeo).
Se existe um pagamento de valor ~= ao adiantamento que ja quitou as mesmas faturas -> duplicata provavel.
NAO escreve nada no Odoo.
"""
import sys
import re
from collections import defaultdict
sys.path.insert(0, '/home/rafaelnascimento/projetos/frete_sistema')
from app.odoo.utils.connection import get_odoo_connection
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

CTX = {'allowed_company_ids': [1, 5]}
OUT = '/home/rafaelnascimento/projetos/frete_sistema/docs/industrializacao-fb-lf/G9_Duplicidade_Adiantamentos.xlsx'
ACC_FB, P_LF = 11038, 35
RE_FAT = re.compile(r'[A-Z]{3,6}/\d{4}/\d{2}/\d{4}')
HDR = PatternFill('solid', fgColor='1F4E78'); HF = Font(bold=True, color='FFFFFF')
TIT = Font(bold=True, size=13, color='1F4E78'); B = Font(bold=True); NUM = '#,##0.00'
BAD = PatternFill('solid', fgColor='FCE4D6'); WARN = PatternFill('solid', fgColor='FFF2CC'); OKF = PatternFill('solid', fgColor='E2EFDA')


def main():
    o = get_odoo_connection()
    assert o.authenticate(), "FALHA AUTH"

    base_fb = [('account_id', '=', ACC_FB), ('partner_id', '=', P_LF), ('parent_state', '=', 'posted')]
    adi = o.execute_kw('account.move.line', 'search_read',
                       [base_fb + [('debit', '>', 0), ('reconciled', '=', False), ('amount_residual', '>', 0)]],
                       {'fields': ['date', 'debit', 'amount_residual', 'ref', 'move_id'], 'order': 'date', 'context': CTX})

    alvo_names = set()
    for l in adi:
        alvo_names |= set(RE_FAT.findall(l.get('ref') or ''))
    fat_moves = o.execute_kw('account.move', 'search_read', [[('name', 'in', list(alvo_names)), ('company_id', '=', 1)]],
                             {'fields': ['name'], 'context': CTX})
    mid2name = {m['id']: m['name'] for m in fat_moves}
    fat_lines = o.execute_kw('account.move.line', 'search_read',
                             [[('move_id', 'in', list(mid2name)), ('account_id', '=', ACC_FB), ('credit', '>', 0)]],
                             {'fields': ['move_id', 'credit', 'amount_residual', 'matched_debit_ids'], 'context': CTX})
    fat = {}  # name -> {credit, resid, partials}
    for l in fat_lines:
        nm = mid2name[l['move_id'][0]]
        fat[nm] = {'credit': l['credit'], 'resid': -(l['amount_residual'] or 0), 'partials': l.get('matched_debit_ids') or []}

    # resolver os pagamentos que quitaram cada fatura (partial -> debit_move_id)
    all_p = set()
    for f in fat.values():
        all_p |= set(f['partials'])
    parts = o.execute_kw('account.partial.reconcile', 'read', [list(all_p)], {'fields': ['debit_move_id', 'amount'], 'context': CTX}) if all_p else []
    pmap = {p['id']: p for p in parts}
    deb_ids = list({p['debit_move_id'][0] for p in parts if p.get('debit_move_id')})
    deb_lines = o.execute_kw('account.move.line', 'read', [deb_ids], {'fields': ['move_id', 'date', 'debit'], 'context': CTX}) if deb_ids else []
    debinfo = {l['id']: {'pag': l['move_id'][1], 'date': l['date'], 'debit': l['debit']} for l in deb_lines}

    rows = []
    n_dup = 0; v_dup = 0.0
    for l in adi:
        X = l['amount_residual']
        alvos = RE_FAT.findall(l.get('ref') or '')
        S = sum(fat[a]['credit'] for a in alvos if a in fat)
        # quem quitou as faturas-alvo (pagamentos distintos, soma por pagamento)
        pagadores = defaultdict(lambda: {'val': 0.0, 'date': None})
        for a in alvos:
            f = fat.get(a)
            if not f:
                continue
            for pid in f['partials']:
                p = pmap.get(pid)
                if not p or not p.get('debit_move_id'):
                    continue
                di = debinfo.get(p['debit_move_id'][0])
                if di and di['pag'] != l['move_id'][1]:  # exclui o proprio
                    pagadores[di['pag']]['val'] += p.get('amount', 0) or 0
                    pagadores[di['pag']]['date'] = di['date']
        # pagamento gemeo = maior pagador
        gemeo = max(pagadores.items(), key=lambda kv: kv[1]['val'], default=None)
        gem_str = ''
        dup = False
        if gemeo:
            gnome, gd = gemeo
            gem_str = f"{gnome} | {gd['date']} | R$ {gd['val']:,.2f}"
            if abs(gd['val'] - X) <= max(1.0, 0.01 * X):  # gemeo ~= adiantamento
                dup = True
        diag = 'DUPLICATA provável' if dup else ('faturas já quitadas — revisar' if S > 0 else 'sem alvo')
        if dup:
            n_dup += 1; v_dup += X
        rows.append({'date': l['date'], 'pag': l['move_id'][1], 'X': X, 'S': S,
                     'gemeo': gem_str or '(não localizado)', 'diag': diag,
                     'outros': '; '.join(f'{k} R$ {v["val"]:,.0f}' for k, v in sorted(pagadores.items(), key=lambda x: -x[1]['val'])[:3])})

    # ===== Excel =====
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Duplicidade'
    for col, w in zip('ABCDEFG', [12, 22, 15, 15, 40, 24, 50]):
        ws.column_dimensions[col].width = w
    ws['A1'] = 'Investigação de duplicidade — adiantamentos FB cujas faturas-alvo já estão pagas'; ws['A1'].font = TIT
    ws['A2'] = f'{len(rows)} adiantamentos | DUPLICATA provável (gêmeo ~= valor): {n_dup} = R$ {v_dup:,.2f}'
    ws['A2'].font = Font(italic=True, size=9, color='808080')
    hr = 4
    for c, h in enumerate(['Data', 'Adiantamento', 'Valor (X)', 'Soma faturas-alvo (S)', 'Pagamento que JÁ quitou (gêmeo)', 'Diagnóstico', 'Outros pagadores'], 1):
        cell = ws.cell(row=hr, column=c, value=h); cell.fill = HDR; cell.font = HF; cell.alignment = Alignment(horizontal='center', wrap_text=True)
    r = hr + 1
    for row in rows:
        for c, v in enumerate([row['date'], row['pag'], row['X'], row['S'], row['gemeo'], row['diag'], row['outros']], 1):
            cell = ws.cell(row=r, column=c, value=v)
            if c in (3, 4):
                cell.number_format = NUM
        ws.cell(row=r, column=6).fill = BAD if 'DUPLICATA' in row['diag'] else WARN
        r += 1
    ws.freeze_panes = 'A5'
    wb.save(OUT)

    print(f"Excel: {OUT}")
    print(f"  {len(rows)} adiantamentos | DUPLICATA provável: {n_dup} = R$ {v_dup:,.2f}")
    for row in rows:
        print(f"  {row['date']} X={row['X']:>11,.2f} S={row['S']:>11,.2f} [{row['diag']:22s}] gemeo: {row['gemeo'][:55]}")


if __name__ == '__main__':
    main()
