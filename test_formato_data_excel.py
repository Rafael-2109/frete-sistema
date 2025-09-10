#!/usr/bin/env python3
"""
Script para testar o formato de data no Excel
Verifica se a data está sendo gravada como valor datetime real, não como string
"""

import os
import sys
from datetime import datetime, date
from openpyxl import Workbook, load_workbook

# Adicionar o caminho do projeto
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

def testar_formato_data():
    """
    Testa a gravação de datas no Excel
    """
    print("\n" + "="*60)
    print("🧪 TESTE DE FORMATO DE DATA NO EXCEL")
    print("="*60)
    
    # Criar workbook de teste
    wb = Workbook()
    ws = wb.active
    ws.title = "Teste Data"
    
    # Data de teste
    data_teste = date(2025, 9, 12)  # 12/09/2025
    
    # Testar diferentes formas de gravar
    print("\n📝 Gravando datas de diferentes formas:")
    
    # Linha 1: Cabeçalhos
    ws['A1'] = 'Método'
    ws['B1'] = 'Valor'
    ws['C1'] = 'Tipo no Excel'
    
    # Linha 2: Como STRING (ERRADO)
    ws['A2'] = 'String formatada'
    ws['B2'] = data_teste.strftime('%d/%m/%Y')  # String "12/09/2025"
    
    # Linha 3: Como DATE object (CORRETO)
    ws['A3'] = 'Date object Python'
    celula_data = ws['B3']
    celula_data.value = data_teste  # Date object
    celula_data.number_format = 'DD/MM/YYYY'
    
    # Linha 4: Como DATETIME (também CORRETO)
    ws['A4'] = 'Datetime object Python'
    celula_datetime = ws['B4']
    celula_datetime.value = datetime(2025, 9, 12, 10, 30)  # Datetime object
    celula_datetime.number_format = 'DD/MM/YYYY HH:MM'
    
    # Salvar arquivo
    arquivo_teste = '/tmp/teste_formato_data.xlsx'
    wb.save(arquivo_teste)
    print(f"\n✅ Arquivo salvo: {arquivo_teste}")
    
    # Reabrir e verificar tipos
    print("\n🔍 Verificando tipos após salvar e reabrir:")
    wb2 = load_workbook(arquivo_teste)
    ws2 = wb2.active
    
    for row in range(2, 5):
        metodo = ws2[f'A{row}'].value
        valor = ws2[f'B{row}'].value
        tipo = type(valor).__name__
        
        print(f"\n  {metodo}:")
        print(f"    Valor: {valor}")
        print(f"    Tipo Python: {tipo}")
        
        # Verificar se é reconhecido como data
        if isinstance(valor, (date, datetime)):
            print(f"    ✅ Reconhecido como data/datetime pelo Excel")
            print(f"    📅 Pode ser ordenado e filtrado corretamente")
        else:
            print(f"    ❌ Gravado como string - NÃO reconhecido como data")
            print(f"    ⚠️ Não pode ser ordenado corretamente")
    
    print("\n" + "="*60)
    print("📊 CONCLUSÃO:")
    print("="*60)
    print("✅ Usar: celula.value = date_object")
    print("✅ Usar: celula.number_format = 'DD/MM/YYYY'")
    print("❌ NÃO usar: celula.value = data.strftime('%d/%m/%Y')")
    print("="*60 + "\n")

if __name__ == "__main__":
    testar_formato_data()